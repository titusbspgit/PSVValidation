#!/usr/bin/env python3
"""Temporary generator - will be removed"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timezone, timedelta

IST = timezone(timedelta(hours=5, minutes=30))
now = datetime.now(IST)
ts = now.strftime("%Y%m%d_%H%M%S")
filename = f"GPIO_TestPlan_{ts}.xlsx"

wb = openpyxl.Workbook()
# TestPlan sheet
ws1 = wb.active
ws1.title = "TestPlan"

# MetaData sheet
ws2 = wb.create_sheet("MetaData")
ws2.sheet_state = 'veryHidden'

print(filename)
