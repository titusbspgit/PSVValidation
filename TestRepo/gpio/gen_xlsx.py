#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timezone, timedelta
import sys, os, base64

IST = timezone(timedelta(hours=5, minutes=30))
now = datetime.now(IST)
ts = now.strftime("%Y%m%d_%H%M%S")
filename = f"GPIO_TestPlan_{ts}.xlsx"
filepath = f"/tmp/{filename}"

wb = openpyxl.Workbook()

# === TESTPLAN SHEET ===
ws1 = wb.active
ws1.title = "TestPlan"

tp_headers = ["Index","SS / Module","Feature","Test Case Name","Test Description","Speed","Mode","Memory Start Offset","Memory End Offset","Remarks","Test Steps / Procedure","Impacted Registers","Validation / Acceptance Criteria","Code Generation (Required / Not)"]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")

for col_idx, h in enumerate(tp_headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_align

row_data = [
    1,
    "GPIO",
    "Register Write Read",
    "gpio_reg_wr_rd_test",
    "This test verifies the default reset values and write-read integrity of 49 GPIO registers including gp0_gpio_8, gp0_gpio_9, gp0_gpio_10, and additional GPIO registers. In the first phase, each register is read after reset and its value (with bit 0 masked out) is compared against the expected default value, with certain registers skipped based on configuration. In the second phase, six distinct data patterns are written to each writable register and then read back, with the read value compared against the expected value computed using write masks, read masks, and default values for non-writable bits. Some registers classified as VRRW are excluded from the write-read phase. The test passes only if all default value checks and all write-read checks succeed across all registers and all patterns.",
    "NA",
    "NA",
    "",
    "",
    "The test covers 49 GPIO registers defined in addr_array. Headers gpio_def.h and gpio_offset.h provide register address definitions and offset mappings. Some registers at indices 32 and 37-44 are excluded from write-read checks as VRRW registers via skip_array. Registers at indices 37-48 are excluded from default value checks via skip_rst_array. Bit 0 is masked out during default value comparison using 0xfffffffe. The soft_reset_chk function is defined but disabled. A note in the source indicates that the din value may become 1 automatically if not forced, affecting bit-level selection and read value matching.",
    "1. Initialize the test environment and begin the default value verification phase for all 49 GPIO registers.\n2. For each register (gp0_gpio_8, gp0_gpio_9, gp0_gpio_10, and remaining GPIO registers), skip registers flagged for reset-check exclusion or that are not readable.\n3. Read each applicable register after reset and compare the read value (with bit 0 masked out) against the expected default reset value. Record any mismatches.\n4. Begin the write-read verification phase using six test data patterns: all-ones, alternating-bit patterns, mixed patterns, and half-word pattern.\n5. For each test pattern, write the pattern (masked by the per-register write mask) to each writable register that is not flagged for exclusion (VRRW registers are skipped).\n6. Read back each written register (masked by the per-register read mask) and compare against the expected value, which accounts for writable bits receiving the test pattern and non-writable bits retaining their default values.\n7. Record any write-read mismatches.\n8. After completing all six patterns across all registers, verify that no default-value or write-read failures occurred.\n9. Report the test as passed if all checks succeed, or failed if any mismatch was detected.",
    "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
    "All 49 GPIO registers must return their expected default values after reset when read with bit 0 masked out. All writable registers (excluding VRRW-flagged registers) must correctly store and return all six distinct test patterns when read back, with the expected value accounting for write masks, read masks, and default values of non-writable bits. The test passes only if both def_fail_cnt and wr_fail_cnt remain zero after all checks are complete.",
    ""
]

for col_idx, val in enumerate(row_data, 1):
    cell = ws1.cell(row=2, column=col_idx, value=val)
    cell.alignment = wrap_align

col_widths = [8, 15, 22, 28, 60, 8, 8, 20, 20, 50, 60, 35, 55, 25]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws1.freeze_panes = "A2"

# === METADATA SHEET ===
ws2 = wb.create_sheet("MetaData")
ws2.sheet_state = 'veryHidden'

md_headers = ["Index","Test Case Name","Meta Test Description","Meta Test Steps / Procedure","Meta Impacted Registers","Meta Validation / Acceptance Criteria","Meta Headers","Meta Macros","Meta Arrays"]

meta_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

for col_idx, h in enumerate(md_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = meta_fill
    cell.alignment = wrap_align

meta_row = [
    1,
    "gpio_reg_wr_rd_test",
    "This testcase validates the default (reset) values and write-read integrity of 49 GPIO registers. The test includes test_define.c which includes gpio/gpio_def.h and gpio/gpio_offset.h, and defines CNT=49 along with arrays: addr_array[49] containing register address macros (MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, and 46 additional GPIO registers), default_value_array[49] with expected reset values (GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL, etc.), read_mask_array[49] with read masks, write_mask_array[49] with write masks, skip_array[49] for write-read skip flags (indices 32, 37-44 are skipped as VRRW registers), and skip_rst_array[49] for reset-value skip flags (indices 37-48 are skipped). Global counters def_fail_cnt and wr_fail_cnt track failures. SOFT_RST_REG_ADDRESS is defined as 0x00000000 and SOFT_RST_REG_DATA as 0x00000000 but soft_reset_chk() is disabled via #ifdef 0. Phase 1 - chk_rst_val(): Iterates i from 0 to CNT-1, sets addr=addr_array[i], skips if skip_rst_array[i]==1, skips if read_mask_array[i]==0x00000000 (not readable), reads register via data_rd=read_reg(addr), masks with data=(data_rd & 0xfffffffe) to exclude bit 0, compares data against default_value_array[i], increments def_fail_cnt on mismatch. Phase 2 - chk_rd_wr(): Defines chk_val[6]={0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}. Outer loop j from 0 to 5 iterates over patterns. Write sub-loop: for each register i from 0 to CNT-1, skips if skip_array[i]==1, skips if write_mask_array[i]==0x00000000, otherwise calls write_reg(addr, (data_wr & write_mask_array[i])). Read sub-loop: for each register i from 0 to CNT-1, skips if skip_array[i]==1, skips if write_mask_array[i]==0x00000000, skips if read_mask_array[i]==0x00000000, otherwise reads data_rd=(read_reg(addr) & read_mask_array[i]), computes wr_n=(write_mask_array[i] ^ 0xffffffff), computes exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])), compares data_rd against exp_val, increments wr_fail_cnt on mismatch. Final check: if def_fail_cnt > 0 or wr_fail_cnt > 0 then finish(1) else finish(0).",
    "1. test_case() is called as the entry point.\n2. chk_rst_val() is invoked to verify default register values.\n3. Loop i from 0 to CNT-1 (CNT=49):\n a. addr = addr_array[i] (e.g., MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, ...).\n b. If skip_rst_array[i] == 1, skip this register (continue). Indices 37-48 are skipped.\n c. If read_mask_array[i] == 0x00000000, skip this register (not readable).\n d. data_rd = read_reg(addr).\n e. data = (data_rd & 0xfffffffe) — mask out bit 0.\n f. Compare data against default_value_array[i] (e.g., GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL, ...).\n g. If mismatch, increment def_fail_cnt and print failure message with address, expected, and read values.\n h. If match, print pass message under DEBUG_DISPLAY.\n4. chk_rd_wr() is invoked to verify write-read functionality.\n5. Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}.\n6. Outer loop j from 0 to 5 (six test patterns):\n a. data_wr = chk_val[j].\n b. Write sub-loop: Loop i from 0 to CNT-1:\n i. addr = addr_array[i].\n ii. If skip_array[i] == 1, skip (indices 32, 37-44 are VRRW registers).\n iii. If write_mask_array[i] == 0x00000000, skip (not writable).\n iv. write_reg(addr, (data_wr & write_mask_array[i])).\n c. Read sub-loop: Loop i from 0 to CNT-1:\n i. addr = addr_array[i].\n ii. If skip_array[i] == 1, skip.\n iii. If write_mask_array[i] == 0x00000000, skip.\n iv. If read_mask_array[i] == 0x00000000, skip.\n v. data_rd = (read_reg(addr) & read_mask_array[i]).\n vi. wr_n = (write_mask_array[i] ^ 0xffffffff).\n vii. exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])).\n viii. Compare data_rd against exp_val.\n ix. If mismatch, increment wr_fail_cnt and print failure message.\n x. If match, print pass message under DEBUG_DISPLAY.\n7. After both phases complete, evaluate final result.\n8. If (def_fail_cnt > 0 || wr_fail_cnt > 0), call finish(1) — test FAIL.\n9. Else call finish(0) — test PASS.",
    "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10",
    "All 49 GPIO registers must return their expected default values after reset when read with bit 0 masked out. All writable registers (excluding VRRW-flagged registers) must correctly store and return all six distinct test patterns when read back, with the expected value accounting for write masks, read masks, and default values of non-writable bits. The test passes only if both def_fail_cnt and wr_fail_cnt remain zero after all checks are complete.",
    "",
    "",
    ""
]

for col_idx, val in enumerate(meta_row, 1):
    cell = ws2.cell(row=2, column=col_idx, value=val)
    cell.alignment = wrap_align

md_widths = [8, 28, 60, 60, 40, 55, 20, 20, 20]
for i, w in enumerate(md_widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws2.freeze_panes = "A2"

wb.save(filepath)
print(f"FILENAME={filename}")
print(f"FILEPATH={filepath}")
