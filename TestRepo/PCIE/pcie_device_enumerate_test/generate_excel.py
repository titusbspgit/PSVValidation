#!/usr/bin/env python3
"""
Agent 7 - Excel Generator Script
Generates PCIE TestPlan Excel workbook from aggregated JSON data.
This script is temporary and will be removed after execution.
"""
import json
import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# IST timezone
IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime("%Y%m%d_%H%M%S")

IP_NAME = "PCIE"
filename = f"{IP_NAME}_TestPlan_{timestamp}.xlsx"
output_dir = os.path.dirname(os.path.abspath(__file__))
filepath = os.path.join(output_dir, filename)

# Input JSON data
json_data = [
    {
        "Index": "1",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_device_enumerate_test",
        "Feature": "Device Enumeration",
        "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
        "Meta Macros": "NA",
        "Meta Arrays": "NA",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Meta Test Description": "This testcase performs PCIe device enumeration. It begins by writing 0x0 to address 0xE6004100. Depending on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP), it invokes link_training_dm0_x4(4) or link_training_dm1_x4(4) for link training. Cache programming is performed by reading mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, using set_data to modify bit fields [11:14], [3:6], [27:30], [19:22] to 0xf, and writing back. After wait_on(20), the same coherency registers are read-modify-written again. SII0 and SII1 link status is polled at offset 0xC0 until bits match 0xD1 pattern ((data_rd & 0xD1) == 0xD1). Under DM0_RC, the Vendor ID is read from PCIe slave 0 at offset 0x0, the command register at offset 0x4 is written with 0x7, and mem_base_program_dm0_x4() and mem_base_program_dm1_x4() are called. Registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034 are written with 0x1. Cache disable programming is performed by modifying the coherency control registers with 0x0 for bit fields [19:22] and [27:30]. After wait_on(30), PCIe slave 1 BAR registers at offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24 are written with 0xFFFFFFFF, read back, then written with specific values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) and read back. The same sequence is repeated for PCIe slave 0. Finally, address 0xE6004100 is polled until it reads 0x12345678, with wait_on(5) between iterations, then finish(0) is called.",
        "Test Description": "This test performs PCIe device enumeration by training the PCIe link, programming cache coherency control registers for both PCIE0 and PCIE1, polling SII link status until the link is established, reading the Vendor ID from the device identification register, enabling bus master and memory space via the status/command register, programming memory base addresses, and then enumerating BAR registers (BAR0, BAR1, secondary bus/primary bus, IO limit/base, memory limit/base, and prefetchable memory limit/base) on both PCIe slave ports by first writing all-ones to determine BAR sizes, reading back, then programming actual base addresses and verifying by reading back. The test concludes by polling a synchronization register until a completion pattern is received.",
        "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0) - clear synchronization register. 2. Conditional link training: link_training_dm0_x4(4) or link_training_dm1_x4(4) based on DM0_RC/DM1_RC/DM0_EP/DM1_EP defines. 3. Cache programming: read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [11:14]=0xf, [3:6]=0xf, write back. 4. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [27:30]=0xf, [19:22]=0xf, write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. wait_on(20). 7. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF with all four bit fields set to 0xf. 8. Read-modify-write mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF similarly. 9. Duplicate link training and cache programming block (repeated in source). 10. read_sii0_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1. 11. read_sii1_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1. 12. non_secure_prot_nic() called. 13. Under DM0_RC: read_pcie_slv0_reg(0x0) to get Vendor ID. 14. write_pcie_slv0_reg(0x4, 0x7) to enable command register. 15. mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). 16. wait_on(10). 17. write_reg(0xE690000C, 0x1), write_reg(0xE6900010, 0x1), write_reg(0xE6900014, 0x1), write_reg(0xE6900018, 0x1), write_reg(0xE6900030, 0x1), write_reg(0xE6900034, 0x1). 18. Cache disable: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF with bits [19:22]=0x0, [27:30]=0x0. 19. Repeat cache disable for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 20. wait_on(10), then consolidated cache disable write for both PCIE0 and PCIE1. 21. wait_on(30). 22. write_pcie_slv1_reg offsets 0x10-0x24 with 0xFFFFFFFF. 23. read_pcie_slv1_reg offsets 0x10-0x24. 24. write_pcie_slv1_reg offsets 0x10-0x24 with specific base addresses. 25. read_pcie_slv1_reg offsets 0x10-0x24. 26. Repeat steps 22-25 for pcie_slv0. 27. wait_on(10). 28. Poll read_reg(0xE6004100) until value equals 0x12345678, with wait_on(5) between iterations. 29. finish(0).",
        "Test Steps / Procedure": "1. Clear the synchronization register to prepare for the test.\n2. Perform PCIe link training based on the configured mode (Root Complex or Endpoint for DM0/DM1).\n3. Program cache coherency control registers for both PCIE0 and PCIE1 by enabling specific bit fields.\n4. Wait for coherency settings to take effect.\n5. Re-apply cache coherency settings for both PCIE0 and PCIE1.\n6. Poll SII0 link status register until the link is established (expected status pattern detected).\n7. Poll SII1 link status register until the link is established.\n8. Configure non-secure protection settings.\n9. Read the device Vendor ID from the TYPE1_DEV_ID_VEND_ID_REG on PCIe slave 0.\n10. Write to TYPE1_STATUS_COMMAND_REG to enable bus master, memory space, and IO space.\n11. Program memory base addresses for both DM0 and DM1.\n12. Enable configuration registers by writing to six system-level control registers.\n13. Disable cache coherency by clearing specific bit fields in the coherency control registers for both PCIE0 and PCIE1.\n14. Wait for cache disable to take effect.\n15. Write all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG on PCIe slave 1 to determine BAR sizes.\n16. Read back all BAR registers on PCIe slave 1 to capture size information.\n17. Program actual base addresses into the BAR registers on PCIe slave 1.\n18. Read back all BAR registers on PCIe slave 1 to verify programming.\n19. Repeat BAR size detection, programming, and verification for PCIe slave 0.\n20. Poll the synchronization register until the expected completion value is received, confirming test completion.",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
        "Impacted Registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
        "Meta Validation / Acceptance Criteria": "1. SII0 link status polling: read_sii0_reg(0xC0) is polled in a while loop until (data_rd & 0xD1) == 0xD1, confirming link is up. 2. SII1 link status polling: read_sii1_reg(0xC0) is polled until (data_rd & 0xD1) == 0xD1. 3. Vendor ID read: rd_wr_data1 = read_pcie_slv0_reg(0x0) is printed to verify device presence. 4. BAR register read-back: After writing 0xFFFFFFFF to offsets 0x10-0x24 on both slave ports, the values are read back to determine BAR sizes. After programming actual base addresses, values are read back to verify correct programming. 5. Synchronization polling: read_reg(0xE6004100) is polled until data_rd == 0x12345678, confirming the remote side has completed its operations. 6. finish(0) is called to indicate test pass.",
        "Validation / Acceptance Criteria": "1. SII0 and SII1 link status registers must report the expected link-up pattern before proceeding.\n2. The Vendor ID read from TYPE1_DEV_ID_VEND_ID_REG must return a valid device identifier confirming device presence.\n3. BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) must correctly reflect size information after writing all-ones and must retain programmed base addresses after final configuration.\n4. The synchronization register must eventually return the expected completion value, confirming the remote endpoint has completed its operations.\n5. The test must complete successfully with a pass indication.",
        "Remarks": "The test uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select between Root Complex and Endpoint modes for link training. The source contains a duplicated block of link training and cache programming code. SII link status polling uses a busy-wait loop. Cache coherency is enabled before enumeration and disabled afterward. A synchronization mechanism using a polling register is used to coordinate with the remote PCIe endpoint. Multiple wait delays (wait_on) are used throughout the test for timing."
    }
]

# TestPlan columns
tp_columns = [
    "Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
    "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
    "Test Steps / Procedure", "Impacted Registers", "Validation / Acceptance Criteria",
    "Code Generation"
]

# MetaData columns
md_columns = [
    "Index", "SS / Module", "Feature", "Test Case Name",
    "Meta Headers", "Meta Macros", "Meta Arrays",
    "Meta Test Description", "Meta Test Steps / Procedure",
    "Meta Impacted Registers", "Meta Validation / Acceptance Criteria"
]

# Column width categories
long_text_cols = {
    "Test Description", "Remarks", "Test Steps / Procedure",
    "Validation / Acceptance Criteria", "Meta Headers", "Meta Macros",
    "Meta Arrays", "Meta Test Description", "Meta Test Steps / Procedure",
    "Meta Impacted Registers", "Meta Validation / Acceptance Criteria"
}
medium_cols = {"Feature", "Test Case Name", "Impacted Registers", "Code Generation"}

def safe_value(row, key):
    val = row.get(key, None)
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return "NA"
    return val

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
wrap_align = Alignment(wrap_text=True, vertical="top")
header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

def apply_formatting(ws, columns):
    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Freeze first row
    ws.freeze_panes = "A2"

    # Auto-filter
    if len(columns) > 0:
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    # Auto-size columns
    from openpyxl.utils import get_column_letter
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))

        # Apply width limits
        if col_name in long_text_cols:
            width = min(max_len + 2, 80)
        elif col_name in medium_cols:
            width = min(max_len + 2, 40)
        else:
            width = min(max_len + 2, 25)

        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 12)

# Create workbook
wb = Workbook()

# Rename default sheet to TestPlan
ws_tp = wb.active
ws_tp.title = "TestPlan"

# Create MetaData sheet
ws_md = wb.create_sheet("MetaData")

# Populate TestPlan
apply_formatting(ws_tp, tp_columns)
for row_idx, row in enumerate(json_data, 2):
    for col_idx, col_name in enumerate(tp_columns, 1):
        if col_name == "Code Generation":
            val = safe_value(row, "Code Generation")
        else:
            val = safe_value(row, col_name)
        cell = ws_tp.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border

# Populate MetaData
apply_formatting(ws_md, md_columns)
for row_idx, row in enumerate(json_data, 2):
    for col_idx, col_name in enumerate(md_columns, 1):
        val = safe_value(row, col_name)
        cell = ws_md.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border

# Set MetaData to veryHidden
ws_md.sheet_state = "veryHidden"

# Save workbook
wb.save(filepath)
print(f"FILENAME={filename}")
print(f"FILEPATH={filepath}")

# Validation
validation_passed = True
errors = []

# 1. File exists
if not os.path.exists(filepath):
    validation_passed = False
    errors.append("File does not exist")

# 2. File size > 0
if os.path.getsize(filepath) <= 0:
    validation_passed = False
    errors.append("File size is 0")

# 3. Filename ends with .xlsx
if not filename.endswith(".xlsx"):
    validation_passed = False
    errors.append("Filename does not end with .xlsx")

# 4-9. Reopen and validate
try:
    wb2 = load_workbook(filepath)
    if "TestPlan" not in wb2.sheetnames:
        validation_passed = False
        errors.append("TestPlan sheet missing")
    if "MetaData" not in wb2.sheetnames:
        validation_passed = False
        errors.append("MetaData sheet missing")
    if wb2["MetaData"].sheet_state != "veryHidden":
        validation_passed = False
        errors.append(f"MetaData sheet state is {wb2['MetaData'].sheet_state}, expected veryHidden")

    # Check TestPlan columns
    tp_headers = [wb2["TestPlan"].cell(row=1, column=i).value for i in range(1, len(tp_columns)+1)]
    if tp_headers != tp_columns:
        validation_passed = False
        errors.append(f"TestPlan columns mismatch: {tp_headers}")

    # Check MetaData columns
    md_headers = [wb2["MetaData"].cell(row=1, column=i).value for i in range(1, len(md_columns)+1)]
    if md_headers != md_columns:
        validation_passed = False
        errors.append(f"MetaData columns mismatch: {md_headers}")

    # Check row counts (excluding header)
    tp_rows = wb2["TestPlan"].max_row - 1
    md_rows = wb2["MetaData"].max_row - 1
    if tp_rows != len(json_data):
        validation_passed = False
        errors.append(f"TestPlan rows: {tp_rows}, expected: {len(json_data)}")
    if md_rows != len(json_data):
        validation_passed = False
        errors.append(f"MetaData rows: {md_rows}, expected: {len(json_data)}")

    wb2.close()
except Exception as e:
    validation_passed = False
    errors.append(f"Reopen failed: {str(e)}")

if validation_passed:
    print("VALIDATION=PASSED")
else:
    print(f"VALIDATION=FAILED")
    for e in errors:
        print(f"ERROR: {e}")

print(f"ROWS_TESTPLAN={len(json_data)}")
print(f"ROWS_METADATA={len(json_data)}")
print(f"FILESIZE={os.path.getsize(filepath)}")
