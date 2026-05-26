import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

# Constants
REPO_ROOT = Path(__file__).resolve().parents[2] if len(Path(__file__).resolve().parents) >= 2 else Path('.')
DATA_JSON = REPO_ROOT / 'Test_Output' / 'testplan_data.json'
OUTPUT_DIR = REPO_ROOT / 'Test_Output'

# Column schemas in required order
TESTPLAN_COLS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

METADATA_COLS = [
    'Index',
    'Test Case Name',
    'Meta Test Description',
    'Meta Test Steps / Procedure',
    'Meta Impacted Registers',
    'Meta Validation / Acceptance Criteria',
    'Meta Headers',
    'Meta Macros',
    'Meta Arrays'
]

def load_json(path: Path):
    with path.open('r', encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError('json_data must be an array')
    for i, row in enumerate(data):
        if not isinstance(row, dict):
            raise ValueError(f'Row {i} is not an object')
    return data


def ist_now_str():
    ist = timezone(timedelta(hours=5, minutes=30))
    return datetime.now(ist).strftime('%Y%m%d_%H%M%S')


def write_sheet(ws, columns, rows):
    # Header
    bold = Font(bold=True)
    ws.append(columns)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = 'A2'  # freeze first row

    # Rows
    for row in rows:
        ws.append([row.get(col, '') if row.get(col, '') is not None else '' for col in columns])


def main():
    data = load_json(DATA_JSON)

    wb = Workbook()
    ws_plan = wb.active
    ws_plan.title = 'TestPlan'
    ws_meta = wb.create_sheet('MetaData')

    # Prepare row data preserving order
    write_sheet(ws_plan, TESTPLAN_COLS, data)
    write_sheet(ws_meta, METADATA_COLS, data)

    # VeryHidden metadata sheet
    ws_meta.sheet_state = 'veryHidden'

    # Ensure output dir exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Save with IST timestamp
    fname = f'testplan_{ist_now_str()}.xlsx'
    out_path = OUTPUT_DIR / fname
    wb.save(out_path)
    print(f'Wrote Excel to: {out_path}')


if __name__ == '__main__':
    main()
