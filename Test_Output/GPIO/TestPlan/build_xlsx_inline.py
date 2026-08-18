#!/usr/bin/env python3
"""Inline xlsx builder - outputs base64-encoded xlsx to stdout for piping to GitHub API."""
import sys, os, json, base64
from datetime import datetime, timezone, timedelta
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
TS = now_ist.strftime("%Y%m%d_%H%M%S")
FN = f"GPIO_TestPlan_{TS}.xlsx"
FP = f"Test_Output/GPIO/TestPlan/{FN}"

H = ["Index","SS / Module","Feature","Test Case Name","Test Description","Speed","Mode","Memory Start Offset","Memory End Offset","Remarks","Test Steps / Procedure","Impacted Registers","Validation / Acceptance Criteria","Code Generation (Required / Not)","Meta Test Description","Meta Test Steps / Procedure","Meta Impacted Registers","Meta Validation / Acceptance Criteria","Meta Headers","Meta Macros","Meta Arrays"]

R1 = [1,"GPIO","Register Write Read Verification","gpio_reg_wr_rd_test","Verify register write and read operations for GPIO per-pin registers gp0_gpio_8 through gp0_gpio_30. The test reads default reset values, writes multiple patterns (0x00, 0xFF, 0x55, 0xAA, 0xA5, 0x5A), reads back and verifies using field-specific read/write masks considering RO, WO, RW, and RW2 field types.","NA","Polling","0x0","0x58","","1. Read default reset value from each GPIO register (gp0_gpio_8 to gp0_gpio_30) and verify against expected reset value 0x00100000 (io_ctrl=1 at bit 20).\n2. Write pattern 0x00000000 to each register.\n3. Read back and verify: RW/RW2 fields updated, RO fields unchanged, WO fields read as 0.\n4. Repeat steps 2-3 for patterns 0x000E0000, 0x000A0000, 0x00140000, 0x00060000, 0x00180000 (masked for writable fields).\n5. Trigger soft reset via SOFT_RST_REG_ADDRESS.\n6. Re-read all registers and verify reset values restored.","gp0_gpio_8; gp0_gpio_9; gp0_gpio_10; gp0_gpio_11; gp0_gpio_12; gp0_gpio_13; gp0_gpio_14; gp0_gpio_15; gp0_gpio_16; gp0_gpio_17; gp0_gpio_18; gp0_gpio_19; gp0_gpio_20; gp0_gpio_21; gp0_gpio_22; gp0_gpio_23; gp0_gpio_24; gp0_gpio_25; gp0_gpio_26; gp0_gpio_27; gp0_gpio_28; gp0_gpio_29; gp0_gpio_30","PASS: All registers return expected reset value 0x00100000 on initial read. All write patterns are correctly reflected in RW/RW2 fields on read-back. RO fields (data_in, intr_raw_sts) remain unchanged after write. WO field (intr_clr) reads as 0. After soft reset, all registers return to reset value. FAIL: Any mismatch between expected and actual values.","Not Required","Verify register write and read operations for GPIO per-pin registers gp0_gpio_8 through gp0_gpio_30. The test reads default reset values, writes multiple patterns (0x00, 0xFF, 0x55, 0xAA, 0xA5, 0x5A), reads back and verifies using field-specific read/write masks considering RO, WO, RW, and RW2 field types.","1. Read default reset value from each GPIO register (gp0_gpio_8 to gp0_gpio_30) and verify against expected reset value 0x00100000 (io_ctrl=1 at bit 20).\n2. Write pattern 0x00000000 to each register.\n3. Read back and verify: RW/RW2 fields updated, RO fields unchanged, WO fields read as 0.\n4. Repeat steps 2-3 for patterns 0x000E0000, 0x000A0000, 0x00140000, 0x00060000, 0x00180000 (masked for writable fields).\n5. Trigger soft reset via SOFT_RST_REG_ADDRESS.\n6. Re-read all registers and verify reset values restored.","MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27; MIZAR_GPIO_GP0_GPIO_28; MIZAR_GPIO_GP0_GPIO_29; MIZAR_GPIO_GP0_GPIO_30; SOFT_RST_REG_ADDRESS","PASS: All registers return expected reset value 0x00100000 on initial read. All write patterns are correctly reflected in RW/RW2 fields on read-back. RO fields (data_in, intr_raw_sts) remain unchanged after write. WO field (intr_clr) reads as 0. After soft reset, all registers return to reset value. FAIL: Any mismatch between expected and actual values.","","",""]

R2 = [2,"GPIO","Level Select Interrupt Enable","test_gpio_level_sel_intr_en","Verify GPIO level-select interrupt functionality for all 32 GPIO pins (pin 8 through pin 39). The test configures each per-pin GPIO register in input mode with level-triggered interrupt enabled, enables the group interrupt via the group interrupt enable register, enables LSS sysreg interrupt routing via the sysreg interrupt enable register, drives an external stimulus, and verifies interrupt assertion through the per-pin interrupt raw status bit, the group interrupt status register, and the LSS sysreg raw status register. The test covers both active HIGH level (level_sel=1) and active LOW level (level_sel=0) interrupt generation. After each interrupt, the test clears the per-pin interrupt, verifies the clear succeeded, disables the group interrupt, verifies the group status is cleared, clears the sysreg raw status, and verifies the sysreg clear succeeded. The test uses GIC IRQ 87 for GPIO0 interrupt routing.","NA","ISR","0x0","0x88","Requires GIC interrupt controller to be initialized and IRQ 87 routed for GPIO0. External stimulus hardware must be connected at the designated SRAM location to drive GPIO pin levels. The test depends on conditional compilation with GPIO0 defined. LSS sysreg interrupt enable must be configured before GPIO interrupts can propagate to the GIC.","1. Enable GIC IRQ for GPIO0 interrupt routing.\n2. Write to the LSS sysreg interrupt enable register (intr_en1) to enable GPIO0 interrupt at bit 1.\n3. For each GPIO pin (pin 8 to pin 39), configure the per-pin register (gp0_gpio_8 through gp0_gpio_39) in input mode with level-select interrupt set to active HIGH (io_ctrl=1, level_sel=1).\n4. Enable the group interrupt for the target pin by writing the corresponding bit to the group interrupt enable register (gp0_intr1_intr_en1).\n5. Drive external stimulus HIGH by writing to the external stimulus address to trigger the level interrupt.\n6. Wait for the interrupt service routine to execute.\n7. In the ISR, read the per-pin register and verify the interrupt raw status bit (bit 1) is asserted.\n8. Read the group interrupt status register (gp0_intr1_intr_sts1) and verify the corresponding pin bit is set.\n9. Clear the per-pin interrupt by writing to the interrupt clear field (bit 16) in the per-pin register.\n10. Read back the per-pin register and verify the interrupt has been cleared successfully (expected value 0x100001).\n11. Disable the group interrupt by writing 0x00000000 to the group interrupt enable register.\n12. Read the group interrupt status register and verify it reads 0x0.\n13. Write to the LSS sysreg raw status clear register (raw_stcr1) to clear the GPIO0 interrupt status.\n14. Read back the sysreg raw status register and verify the GPIO0 interrupt bit is cleared.\n15. Clear the GIC IRQ.\n16. Repeat steps 3-15 for all 32 GPIO pins with active HIGH level.\n17. Repeat the entire sequence (steps 3-16) with level-select set to active LOW (io_ctrl=1, level_sel=0) and drive external stimulus LOW for each pin.\n18. Verify the test completes with zero errors.","intr_en1; gp0_gpio_8; gp0_intr1_intr_en1; gp0_intr1_intr_sts1; raw_stcr1","PASS: For each of the 32 GPIO pins in active HIGH mode (level_sel=1): the per-pin register interrupt raw status bit (bit 1) is asserted when the external stimulus is HIGH. The group interrupt status register (gp0_intr1_intr_sts1) shows the corresponding pin bit set. After clearing the interrupt via the per-pin register interrupt clear field, the per-pin register reads 0x100001. After disabling the group interrupt enable, the group interrupt status register reads 0x0. After clearing the sysreg raw status register (raw_stcr1), the GPIO0 interrupt bit is cleared. For each of the 32 GPIO pins in active LOW mode (level_sel=0): the same validation sequence passes when the external stimulus is driven LOW for the target pin. The test completes with zero accumulated errors. FAIL: Any per-pin interrupt raw status not asserted, any group interrupt status mismatch, any interrupt clear failure (per-pin register not equal to 0x100001), any group interrupt status not clearing to 0x0, any sysreg raw status not clearing, or any interrupt not occurring.","Not Required","Verify GPIO level-select interrupt functionality for all 32 GPIO pins (GPIO_8 to GPIO_39). The test enables GIC IRQ 87 (GPIO0) via GIC_EnableIRQ(87). It writes LSS_SYSREG_INTR_EN1_GPIO0_INTR to MIZAR_LSS_SYSREG_INTR_EN1 to enable sysreg interrupt routing. In the first loop (active HIGH, i=0..31): writes 0x00180000 to MIZAR_GPIO_GP0_GPIO_8+(i*4) setting io_ctrl=1 (bit 20) and level_sel=1 (bit 19), writes (1<<i) to MIZAR_GPIO_GP0_INTR1_INTR_EN1 (offset 0x84) to enable group interrupt for pin i, writes 0xffffffff to 0xA0243ffc to drive external stimulus HIGH, sets int_pend=1 and polls while(int_pend==1) with wait_on(10). In Default_IRQHandler: sets int_pend=0, writes 0xffffffff to 0xA0243ffc, reads MIZAR_GPIO_GP0_GPIO_8+(i*4) into rdata, checks (rdata and 0x2) != 0x0 for intr_raw_sts assertion, reads MIZAR_GPIO_GP0_INTR1_INTR_STS1 (offset 0x88) into rdata_grp, checks (rdata_grp and (1<<i)) != 0 for group interrupt, writes 0x00110000 to MIZAR_GPIO_GP0_GPIO_8+(i*4) to clear interrupt, waits wait_on(20), reads back and checks rdata==0x100001, writes 0x00000000 to MIZAR_GPIO_GP0_INTR1_INTR_EN1 to disable group interrupt, reads MIZAR_GPIO_GP0_INTR1_INTR_STS1 and checks rdata_grp==0x0, writes LSS_SYSREG_RAW_STCR1_GPIO0_INTR to MIZAR_LSS_SYSREG_RAW_STCR1, reads back and checks cleared, then calls GIC_ClearIRQ(87). Second loop (active LOW) similar with 0x00100000. finish(test_err).","1. GIC_EnableIRQ(87) to enable GPIO0 GIC interrupt.\n2. write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR).\n3. Active HIGH loop (i=0 to 31): write_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4), 0x00180000); write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 1<<i); write_reg(0xA0243ffc, 0xffffffff).\n4. ISR: read_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4)), check (rdata & 0x2)!=0.\n5. read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1), check (rdata_grp & (1<<i))!=0.\n6. write_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4), 0x00110000) to clear.\n7. Verify read_reg == 0x100001.\n8. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x0) disable.\n9. Verify read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1)==0x0.\n10. write_reg/read_reg MIZAR_LSS_SYSREG_RAW_STCR1, verify cleared.\n11. GIC_ClearIRQ(87).\n12. Active LOW loop similar with 0x00100000.\n13. finish(test_err).","MIZAR_LSS_SYSREG_INTR_EN1; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1; 0xA0243ffc; MIZAR_GPIO_GP0_INTR1_INTR_STS1; MIZAR_LSS_SYSREG_RAW_STCR1","PASS conditions: (1) (rdata & 0x2) != 0x0 after read_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4)). (2) (rdata_grp & (1<<i)) != 0 after read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1). (3) read_reg == 0x100001 after clear. (4) read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) == 0x0 after disable. (5) sysreg raw_stcr1 cleared. FAIL if test_err > 0.","#include<lss_sysreg.h>; #include<stdio.h>; #include<test_define.c>; #include<test_common.h>; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>","#define CNT 49","const unsigned long int addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,...}; const unsigned int default_value_array[49]={...}; const unsigned int read_mask_array[49]={...}; const unsigned int write_mask_array[49]={...}; const int skip_array[49]={0,...,0};"]

CW = [8,14,30,32,60,8,10,20,18,40,70,60,65,28,60,70,60,65,50,30,50]

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"
    hf = Font(name="Calibri",size=11,bold=True,color="FFFFFF")
    hfill = PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
    ha = Alignment(horizontal="center",vertical="center",wrap_text=True)
    ca = Alignment(vertical="top",wrap_text=True)
    tb = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    for ci,h in enumerate(H,1):
        c = ws.cell(row=1,column=ci,value=h)
        c.font=hf; c.fill=hfill; c.alignment=ha; c.border=tb
    for ri,rd in enumerate([R1,R2],2):
        for ci,v in enumerate(rd,1):
            c = ws.cell(row=ri,column=ci,value=v)
            c.alignment=ca; c.border=tb; c.font=Font(name="Calibri",size=10)
    for ci,w in enumerate(CW,1):
        ws.column_dimensions[ws.cell(row=1,column=ci).column_letter].width=w
    ws.freeze_panes="A2"
    wm = wb.create_sheet(title="MetaData")
    meta_json = json.dumps([{H[i]:R1[i] for i in range(21)},{H[i]:R2[i] for i in range(21)}],indent=2)
    wm.cell(row=1,column=1,value=meta_json)
    wm.sheet_state="veryHidden"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def main():
    xlsx_bytes = build()
    # Save locally
    with open(FN,"wb") as f:
        f.write(xlsx_bytes)
    print(f"FILE_PATH={FP}",file=sys.stderr)
    print(f"FILENAME={FN}",file=sys.stderr)
    print(f"SIZE={len(xlsx_bytes)}",file=sys.stderr)
    # Output base64 to stdout for GitHub API
    sys.stdout.write(base64.b64encode(xlsx_bytes).decode())

if __name__=="__main__":
    main()
