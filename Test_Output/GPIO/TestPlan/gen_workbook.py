#!/usr/bin/env python3
"""
GPIO TestPlan XLSX Generator
Generates GPIO_TestPlan_20260820_123218.xlsx
Run: python3 gen_workbook.py
"""
import json, sys
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── JSON Data ──
json_data = [
  {
    "index": 1,
    "ss_module": "GPIO",
    "test_case_name": "gpio_reg_wr_rd_test",
    "feature": "Register Read/Write Validation",
    "test_description": "This test validates the GPIO GP0 register block by performing two checks: (1) Reset value verification - reads each GPIO register and verifies the data matches the expected default reset values after masking read-only bits. (2) Write/Read verification - writes six different test patterns (all-ones, alternating bits, mixed patterns) to each writable GPIO register, reads back the values, and verifies correctness by accounting for read masks, write masks, and default values of non-writable bit fields. The test covers registers gp0_gpio_8, gp0_gpio_9, and gp0_gpio_10. The test reports PASS if all default value checks and write/read checks succeed, otherwise FAIL.",
    "test_steps": "1. Initialize the test environment and load the GPIO register configuration arrays including register addresses, default values, read masks, write masks, and skip control arrays.\n2. Perform reset value verification for each GPIO register (gp0_gpio_8, gp0_gpio_9, gp0_gpio_10): Read each register, apply the appropriate read mask, and compare the result against the expected default reset value. Skip registers marked in the skip-reset array or those that are not readable.\n3. Perform write/read verification using six test data patterns (all-ones, alternating-bit patterns, and mixed patterns) for each GPIO register:\n   a. Write each test pattern to the register, masked by the write mask to protect read-only and write-only fields.\n   b. Read back the register value, masked by the read mask.\n   c. Compute the expected value considering writable bits from the written pattern and non-writable bits retaining their default values.\n   d. Compare the read-back value against the expected value.\n4. Skip any registers flagged in the skip array or those with zero write mask (not writable) or zero read mask (not readable).\n5. Evaluate the final test result: PASS if all default value checks and all write/read pattern checks succeed with zero mismatches; FAIL otherwise.",
    "impacted_registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
    "validation_acceptance_criteria": "1. All GPIO registers must return their expected default reset values when read after reset.\n2. For each of the six test patterns, the write/read-back value must match the expected value computed using read masks, write masks, and default values for non-writable fields.\n3. The test must complete with zero default-value mismatches and zero write/read mismatches to be considered PASS.\n4. Registers marked as skip or having zero read/write masks must be correctly excluded from the respective checks.",
    "speed": "NA",
    "mode": "NA",
    "remarks": "The test uses six distinct data patterns to exercise all writable bit positions across the GPIO registers. Bit 0 (data_in field) is masked out during default value comparison since it reflects the live pin state. The skip arrays allow selective exclusion of specific registers from reset and write/read checks. The soft reset check function is disabled in the source code.",
    "meta_test_description": "Phase 1 chk_rst_val and Phase 2 chk_rd_wr for GPIO GP0 registers...",
    "meta_test_steps": "1. Include headers...\n2. Define arrays...",
    "meta_impacted_registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10"
  },
  {
    "index": 2,
    "ss_module": "GPIO",
    "test_case_name": "test_gpio_level_sel_intr_en",
    "feature": "GPIO Level Select Interrupt Enable",
    "test_description": "This test validates the GPIO level-select interrupt enable functionality for 32 GPIO pins (pins 8 through 39). The test operates in two phases: Phase 1 tests active-high level interrupts and Phase 2 tests active-low level interrupts. The test passes only if all 64 interrupt cycles complete without any errors.",
    "test_steps": "1. Enable the GIC interrupt line...\n2. Enable the system-level GPIO interrupt...",
    "impacted_registers": "intr_en1; gp0_gpio_8; raw_stcr1",
    "validation_acceptance_criteria": "1. For each of the 32 GPIO pins in active-high level mode...\n2. The test must complete all 64 interrupt cycles with zero errors.",
    "speed": "NA",
    "mode": "NA",
    "remarks": "The test exercises both active-high and active-low level-select interrupt modes for all 32 GPIO pins.",
    "meta_test_description": "GPIO level-select interrupt test...",
    "meta_test_steps": "1. Include headers...\n2. Enable GIC IRQ...",
    "meta_impacted_registers": "MIZAR_LSS_SYSREG_INTR_EN1; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1; 0xA0243ffc; MIZAR_GPIO_GP0_INTR1_INTR_STS1; MIZAR_LSS_SYSREG_RAW_STCR1"
  }
]

# ── Create Workbook ──
wb = Workbook()
ws_tp = wb.active
ws_tp.title = "TestPlan"
ws_md = wb.create_sheet("MetaData")

# ── Styles ──
hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")

# ── TestPlan Headers ──
tp_headers = ["Index","SS / Module","Feature","Test Case Name","Test Description",
              "Speed","Mode","Memory Start Offset","Memory End Offset","Remarks",
              "Test Steps / Procedure","Impacted Registers",
              "Validation / Acceptance Criteria","Code Generation"]

for c, h in enumerate(tp_headers, 1):
    cell = ws_tp.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = wrap

# ── TestPlan Data ──
for r, tc in enumerate(json_data, 2):
    ws_tp.cell(row=r, column=1, value=tc["index"]).alignment = wrap
    ws_tp.cell(row=r, column=2, value=tc["ss_module"]).alignment = wrap
    ws_tp.cell(row=r, column=3, value=tc["feature"]).alignment = wrap
    ws_tp.cell(row=r, column=4, value=tc["test_case_name"]).alignment = wrap
    ws_tp.cell(row=r, column=5, value=tc["test_description"]).alignment = wrap
    ws_tp.cell(row=r, column=6, value=tc["speed"]).alignment = wrap
    ws_tp.cell(row=r, column=7, value=tc["mode"]).alignment = wrap
    ws_tp.cell(row=r, column=8, value="").alignment = wrap
    ws_tp.cell(row=r, column=9, value="").alignment = wrap
    ws_tp.cell(row=r, column=10, value=tc["remarks"]).alignment = wrap
    ws_tp.cell(row=r, column=11, value=tc["test_steps"]).alignment = wrap
    ws_tp.cell(row=r, column=12, value=tc["impacted_registers"]).alignment = wrap
    ws_tp.cell(row=r, column=13, value=tc["validation_acceptance_criteria"]).alignment = wrap
    ws_tp.cell(row=r, column=14, value="").alignment = wrap

# ── MetaData Headers ──
md_headers = ["Index","Test Case Name","Meta Test Description",
              "Meta Test Steps / Procedure","Meta Impacted Registers",
              "Meta Validation / Acceptance Criteria",
              "Meta Headers","Meta Macros","Meta Arrays"]

for c, h in enumerate(md_headers, 1):
    cell = ws_md.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = wrap

# ── MetaData Data ──
for r, tc in enumerate(json_data, 2):
    ws_md.cell(row=r, column=1, value=tc["index"]).alignment = wrap
    ws_md.cell(row=r, column=2, value=tc["test_case_name"]).alignment = wrap
    ws_md.cell(row=r, column=3, value=tc.get("meta_test_description","")).alignment = wrap
    ws_md.cell(row=r, column=4, value=tc.get("meta_test_steps","")).alignment = wrap
    ws_md.cell(row=r, column=5, value=tc.get("meta_impacted_registers","")).alignment = wrap
    ws_md.cell(row=r, column=6, value=tc.get("validation_acceptance_criteria","")).alignment = wrap
    ws_md.cell(row=r, column=7, value="").alignment = wrap
    ws_md.cell(row=r, column=8, value="").alignment = wrap
    ws_md.cell(row=r, column=9, value="").alignment = wrap

# ── Freeze & Column Widths ──
ws_tp.freeze_panes = "A2"
ws_md.freeze_panes = "A2"

for ws in [ws_tp, ws_md]:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

# ── Hide MetaData ──
ws_md.sheet_state = "veryHidden"

# ── Save ──
fname = "GPIO_TestPlan_20260820_123218.xlsx"
wb.save(fname)
print(f"SUCCESS: {fname} ({__import__('os').path.getsize(fname)} bytes)")
