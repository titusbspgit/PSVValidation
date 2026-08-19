#!/usr/bin/env python3
"""
Standalone XLSX builder - generates GPIO TestPlan workbook.
Output: GPIO_TestPlan_<IST_timestamp>.xlsx in current directory.
Usage: python build_xlsx.py
Requires: pip install openpyxl

This script generates a real binary .xlsx workbook with:
  - TestPlan sheet (visible): 11 columns with formatted headers
  - MetaData sheet (veryHidden): 7 columns with formatted headers
  - Bold white headers on colored fill
  - Wrapped text, frozen first row, auto-filter, thin borders
  - IST-timestamped filename
"""
import datetime
import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: pip install openpyxl")

# ── Timestamp in IST ──────────────────────────────────────────────────────────
IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))
ts = datetime.datetime.now(IST).strftime("%Y%m%d_%H%M%S")
fname = f"GPIO_TestPlan_{ts}.xlsx"

# ── Complete Test Data Row ────────────────────────────────────────────────────
ROW = {
    "Index": 1,
    "SS / Module": "GPIO",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Feature": "Register Write Read",
    "Test Description": (
        "This test verifies the default reset values and write-read integrity of 49 GPIO registers "
        "including gp0_gpio_8, gp0_gpio_9, gp0_gpio_10, and additional GPIO registers. In the first "
        "phase, each register is read after reset and its value (with bit 0 masked out) is compared "
        "against the expected default value, with certain registers skipped based on configuration. In "
        "the second phase, six distinct data patterns are written to each writable register and then "
        "read back, with the read value compared against the expected value computed using write masks, "
        "read masks, and default values for non-writable bits. Some registers classified as VRRW are "
        "excluded from the write-read phase. The test passes only if all default value checks and all "
        "write-read checks succeed across all registers and all patterns."
    ),
    "Test Steps / Procedure": (
        "1. Initialize the test environment and begin the default value verification phase for all 49 GPIO registers.\n"
        "2. For each register (gp0_gpio_8, gp0_gpio_9, gp0_gpio_10, and remaining GPIO registers), skip registers flagged for reset-check exclusion or that are not readable.\n"
        "3. Read each applicable register after reset and compare the read value (with bit 0 masked out) against the expected default reset value. Record any mismatches.\n"
        "4. Begin the write-read verification phase using six test data patterns: all-ones, alternating-bit patterns, mixed patterns, and half-word pattern.\n"
        "5. For each test pattern, write the pattern (masked by the per-register write mask) to each writable register that is not flagged for exclusion (VRRW registers are skipped).\n"
        "6. Read back each written register (masked by the per-register read mask) and compare against the expected value, which accounts for writable bits receiving the test pattern and non-writable bits retaining their default values.\n"
        "7. Record any write-read mismatches.\n"
        "8. After completing all six patterns across all registers, verify that no default-value or write-read failures occurred.\n"
        "9. Report the test as passed if all checks succeed, or failed if any mismatch was detected."
    ),
    "Impacted Registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
    "Validation / Acceptance Criteria": (
        "All 49 GPIO registers must return their expected default values after reset when read with bit 0 "
        "masked out. All writable registers (excluding VRRW-flagged registers) must correctly store and "
        "return all six distinct test patterns when read back, with the expected value accounting for "
        "write masks, read masks, and default values of non-writable bits. The test passes only if both "
        "def_fail_cnt and wr_fail_cnt remain zero after all checks are complete."
    ),
    "Speed": "NA",
    "Mode": "NA",
    "Remarks": (
        "The test covers 49 GPIO registers defined in addr_array. Headers gpio_def.h and gpio_offset.h "
        "provide register address definitions and offset mappings. Some registers at indices 32 and 37-44 "
        "are excluded from write-read checks as VRRW registers via skip_array. Registers at indices 37-48 "
        "are excluded from default value checks via skip_rst_array. Bit 0 is masked out during default "
        "value comparison using 0xfffffffe. The soft_reset_chk function is defined but disabled. A note "
        "in the source indicates that the din value may become 1 automatically if not forced, affecting "
        "bit-level selection and read value matching."
    ),
    "Meta Test Description": (
        "This testcase validates the default (reset) values and write-read integrity of 49 GPIO registers. "
        "The test includes test_define.c which includes gpio/gpio_def.h and gpio/gpio_offset.h, and defines "
        "CNT=49 along with arrays: addr_array[49] containing register address macros (MIZAR_GPIO_GP0_GPIO_8, "
        "MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, and 46 additional GPIO registers), "
        "default_value_array[49] with expected reset values (GPIO_GP0_GPIO_8_DEFAULT_VAL, "
        "GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL, etc.), read_mask_array[49] with read "
        "masks, write_mask_array[49] with write masks, skip_array[49] for write-read skip flags (indices 32, "
        "37-44 are skipped as VRRW registers), and skip_rst_array[49] for reset-value skip flags (indices "
        "37-48 are skipped). Global counters def_fail_cnt and wr_fail_cnt track failures. "
        "SOFT_RST_REG_ADDRESS is defined as 0x00000000 and SOFT_RST_REG_DATA as 0x00000000 but "
        "soft_reset_chk() is disabled via #ifdef 0. Phase 1 - chk_rst_val(): Iterates i from 0 to CNT-1, "
        "sets addr=addr_array[i], skips if skip_rst_array[i]==1, skips if read_mask_array[i]==0x00000000 "
        "(not readable), reads register via data_rd=read_reg(addr), masks with data=(data_rd & 0xfffffffe) "
        "to exclude bit 0, compares data against default_value_array[i], increments def_fail_cnt on mismatch. "
        "Phase 2 - chk_rd_wr(): Defines chk_val[6]={0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, "
        "0xA5A5A5A5, 0xffff0000}. Outer loop j from 0 to 5 iterates over patterns. Write sub-loop: for each "
        "register i from 0 to CNT-1, skips if skip_array[i]==1, skips if write_mask_array[i]==0x00000000, "
        "otherwise calls write_reg(addr, (data_wr & write_mask_array[i])). Read sub-loop: for each register "
        "i from 0 to CNT-1, skips if skip_array[i]==1, skips if write_mask_array[i]==0x00000000, skips if "
        "read_mask_array[i]==0x00000000, otherwise reads data_rd=(read_reg(addr) & read_mask_array[i]), "
        "computes wr_n=(write_mask_array[i] ^ 0xffffffff), computes exp_val=((data_wr & read_mask_array[i] "
        "& write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])), compares data_rd "
        "against exp_val, increments wr_fail_cnt on mismatch. Final check: if def_fail_cnt > 0 or "
        "wr_fail_cnt > 0 then finish(1) else finish(0)."
    ),
    "Meta Test Steps / Procedure": (
        "1. test_case() is called as the entry point.\n"
        "2. chk_rst_val() is invoked to verify default register values.\n"
        "3. Loop i from 0 to CNT-1 (CNT=49):\n"
        "   a. addr = addr_array[i] (e.g., MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, ...).\n"
        "   b. If skip_rst_array[i] == 1, skip this register (continue). Indices 37-48 are skipped.\n"
        "   c. If read_mask_array[i] == 0x00000000, skip this register (not readable).\n"
        "   d. data_rd = read_reg(addr).\n"
        "   e. data = (data_rd & 0xfffffffe) - mask out bit 0.\n"
        "   f. Compare data against default_value_array[i] (e.g., GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL, ...).\n"
        "   g. If mismatch, increment def_fail_cnt and print failure message with address, expected, and read values.\n"
        "   h. If match, print pass message under DEBUG_DISPLAY.\n"
        "4. chk_rd_wr() is invoked to verify write-read functionality.\n"
        "5. Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}.\n"
        "6. Outer loop j from 0 to 5 (six test patterns):\n"
        "   a. data_wr = chk_val[j].\n"
        "   b. Write sub-loop: Loop i from 0 to CNT-1:\n"
        "      i. addr = addr_array[i].\n"
        "      ii. If skip_array[i] == 1, skip (indices 32, 37-44 are VRRW registers).\n"
        "      iii. If write_mask_array[i] == 0x00000000, skip (not writable).\n"
        "      iv. write_reg(addr, (data_wr & write_mask_array[i])).\n"
        "   c. Read sub-loop: Loop i from 0 to CNT-1:\n"
        "      i. addr = addr_array[i].\n"
        "      ii. If skip_array[i] == 1, skip.\n"
        "      iii. If write_mask_array[i] == 0x00000000, skip.\n"
        "      iv. If read_mask_array[i] == 0x00000000, skip.\n"
        "      v. data_rd = (read_reg(addr) & read_mask_array[i]).\n"
        "      vi. wr_n = (write_mask_array[i] ^ 0xffffffff).\n"
        "      vii. exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])).\n"
        "      viii. Compare data_rd against exp_val.\n"
        "      ix. If mismatch, increment wr_fail_cnt and print failure message.\n"
        "      x. If match, print pass message under DEBUG_DISPLAY.\n"
        "7. After both phases complete, evaluate final result.\n"
        "8. If (def_fail_cnt > 0 || wr_fail_cnt > 0), call finish(1) - test FAIL.\n"
        "9. Else call finish(0) - test PASS."
    ),
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10"
}

# ── Sheet Column Definitions ──────────────────────────────────────────────────
TP_COLS = [
    "Index", "SS / Module", "Test Case Name", "Feature", "Test Description",
    "Test Steps / Procedure", "Impacted Registers",
    "Validation / Acceptance Criteria", "Speed", "Mode", "Remarks"
]
MD_COLS = [
    "Index", "SS / Module", "Test Case Name", "Feature",
    "Meta Test Description", "Meta Test Steps / Procedure", "Meta Impacted Registers"
]

# ── Styles ────────────────────────────────────────────────────────────────────
HF = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
DF = Font(name="Calibri", size=11, color="000000")
TP_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
MD_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
HALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
BDR = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ── Column Widths ─────────────────────────────────────────────────────────────
TP_W = {
    "Index": 8, "SS / Module": 15, "Test Case Name": 25, "Feature": 22,
    "Test Description": 60, "Test Steps / Procedure": 65, "Impacted Registers": 30,
    "Validation / Acceptance Criteria": 55, "Speed": 10, "Mode": 10, "Remarks": 55
}
MD_W = {
    "Index": 8, "SS / Module": 15, "Test Case Name": 25, "Feature": 22,
    "Meta Test Description": 70, "Meta Test Steps / Procedure": 70,
    "Meta Impacted Registers": 35
}

# ── Build Workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Sheet 1: TestPlan (visible)
ws = wb.active
ws.title = "TestPlan"
ws.sheet_state = "visible"

# Header row
for ci, cn in enumerate(TP_COLS, 1):
    c = ws.cell(row=1, column=ci, value=cn)
    c.font = HF
    c.fill = TP_FILL
    c.alignment = HALIGN
    c.border = BDR
    ws.column_dimensions[get_column_letter(ci)].width = TP_W.get(cn, 20)

# Data row
for ci, cn in enumerate(TP_COLS, 1):
    val = ROW.get(cn, "")
    c = ws.cell(row=2, column=ci, value=val)
    c.font = DF
    c.alignment = WRAP
    c.border = BDR

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(TP_COLS))}2"

# Sheet 2: MetaData (veryHidden)
ws2 = wb.create_sheet("MetaData")
ws2.sheet_state = "veryHidden"

# Header row
for ci, cn in enumerate(MD_COLS, 1):
    c = ws2.cell(row=1, column=ci, value=cn)
    c.font = HF
    c.fill = MD_FILL
    c.alignment = HALIGN
    c.border = BDR
    ws2.column_dimensions[get_column_letter(ci)].width = MD_W.get(cn, 20)

# Data row
for ci, cn in enumerate(MD_COLS, 1):
    val = ROW.get(cn, "")
    c = ws2.cell(row=2, column=ci, value=val)
    c.font = DF
    c.alignment = WRAP
    c.border = BDR

ws2.freeze_panes = "A2"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(fname)
print(f"GENERATED: {fname}")
print(f"SIZE: {os.path.getsize(fname)} bytes")
print(f"SHEETS: TestPlan (visible), MetaData (veryHidden)")
print(f"TIMESTAMP: {ts} IST")
print("STATUS: SUCCESS")
