#!/usr/bin/env python3
"""Generate GPIO TestPlan XLSX - temporary generator script."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, os

wb = openpyxl.Workbook()

# ── Sheet 1: TestPlan (visible) ──
ws1 = wb.active
ws1.title = "TestPlan"

tp_headers = [
    "Index", "SS / Module", "Test Case Name", "Feature",
    "Test Description", "Test Steps / Procedure",
    "Impacted Registers", "Validation / Acceptance Criteria",
    "Speed", "Mode", "Remarks"
]

tp_data = [
    1,
    "GPIO",
    "gpio_reg_wr_rd_test",
    "Register Write Read",
    "This test verifies the default reset values and write-read integrity of 49 GPIO registers including gp0_gpio_8, gp0_gpio_9, gp0_gpio_10, and others. In the first phase, each register is read after reset and its value is compared against the expected default value, with certain registers skipped based on configuration. In the second phase, six distinct data patterns are written to each writable register and read back, with the read value compared against the expected value computed using the write mask, read mask, and default value. The test passes only if all default value checks and all write-read checks succeed across all registers and all patterns.",
    "1. Initialize the test and begin the default value verification phase for all 49 GPIO registers.\n2. For each register (gp0_gpio_8, gp0_gpio_9, gp0_gpio_10, and others), skip registers flagged for reset-check exclusion or that are not readable.\n3. Read each applicable register and compare the read value (with bit 0 masked) against the expected default reset value. Record any mismatches.\n4. Begin the write-read verification phase using six test data patterns (all-ones, alternating bit patterns, and mixed patterns).\n5. For each test pattern, write the pattern (masked by the write mask) to each writable register that is not flagged for exclusion.\n6. Read back each written register (masked by the read mask) and compare against the expected value, which accounts for writable bits, read-only bits, and default values of non-writable bits.\n7. Record any write-read mismatches.\n8. After completing all patterns and all registers, verify that no default-value or write-read failures occurred.\n9. Report the test as passed if all checks succeed, or failed if any mismatch was detected.",
    "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
    "All 49 GPIO registers must return their expected default values after reset. All writable registers must correctly store and return six distinct test patterns when read back, accounting for write masks, read masks, and default values of non-writable bits. The test passes only if zero mismatches are detected across both the default value check and the write-read check phases.",
    "NA",
    "NA",
    "The test covers 49 GPIO registers. Some registers are excluded from the write-read phase (VRRW registers) and some from the reset-value phase via skip arrays. Bit 0 is masked out during default value comparison. The test uses headers gpio_def.h and gpio_offset.h for register address definitions and offset mappings."
]

header_font = Font(name='Calibri', bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col_idx, header in enumerate(tp_headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for col_idx, value in enumerate(tp_data, 1):
    cell = ws1.cell(row=2, column=col_idx, value=value)
    cell.alignment = cell_align
    cell.border = thin_border

col_widths_tp = [8, 15, 28, 22, 60, 65, 35, 60, 10, 10, 55]
for i, w in enumerate(col_widths_tp, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A2"

# ── Sheet 2: MetaData (VeryHidden) ──
ws2 = wb.create_sheet(title="MetaData")

meta_headers = [
    "Index", "SS / Module", "Test Case Name", "Feature",
    "Meta Test Description", "Meta Test Steps / Procedure",
    "Meta Impacted Registers"
]

meta_data = [
    1,
    "GPIO",
    "gpio_reg_wr_rd_test",
    "Register Write Read",
    "This testcase validates the default (reset) values and write-read functionality of 49 GPIO registers. It includes two phases: (1) chk_rst_val() iterates over addr_array[0..CNT-1] containing register address macros (MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, etc.), skips entries flagged in skip_rst_array, skips registers with read_mask_array[i]==0x00000000, reads each register via read_reg(addr), masks the read data with 0xfffffffe, and compares against default_value_array[i]. Mismatches increment def_fail_cnt. (2) chk_rd_wr() writes six test patterns (0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000) into each register masked with write_mask_array[i], then reads back each register masked with read_mask_array[i], computes the expected value as ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])) where wr_n = (write_mask_array[i] ^ 0xffffffff), and compares. Mismatches increment wr_fail_cnt. The test calls finish(0) on success or finish(1) if any failure counter is non-zero. Arrays are defined in test_define.c which includes gpio/gpio_def.h and gpio/gpio_offset.h. Some registers are skipped for write-read (skip_array) and some for reset-value check (skip_rst_array), notably VRRW registers.",
    "1. test_case() entry point is called.\n2. chk_rst_val() is invoked to check default register values.\n3. Loop i from 0 to CNT-1 (49 registers):\n a. addr = addr_array[i] (e.g., MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, ...).\n b. If skip_rst_array[i] == 1, skip this register (continue).\n c. If read_mask_array[i] == 0x00000000, skip (register not readable).\n d. data_rd = read_reg(addr).\n e. data = (data_rd & 0xfffffffe) — mask out bit 0.\n f. Compare data against default_value_array[i].\n g. If mismatch, increment def_fail_cnt and print failure.\n4. chk_rd_wr() is invoked to check write-read functionality.\n5. Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}.\n6. Outer loop j from 0 to 5 (six patterns):\n a. data_wr = chk_val[j].\n b. Write phase: Loop i from 0 to CNT-1:\n i. addr = addr_array[i].\n ii. If skip_array[i] == 1, skip.\n iii. If write_mask_array[i] == 0x00000000, skip (not writable).\n iv. write_reg(addr, (data_wr & write_mask_array[i])).\n c. Read phase: Loop i from 0 to CNT-1:\n i. addr = addr_array[i].\n ii. If skip_array[i] == 1, skip.\n iii. If write_mask_array[i] == 0x00000000, skip.\n iv. If read_mask_array[i] == 0x00000000, skip.\n v. data_rd = (read_reg(addr) & read_mask_array[i]).\n vi. wr_n = (write_mask_array[i] ^ 0xffffffff).\n vii. exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])).\n viii. Compare data_rd against exp_val.\n ix. If mismatch, increment wr_fail_cnt and print failure.\n7. After both phases, check if def_fail_cnt > 0 or wr_fail_cnt > 0.\n8. If any failures, call finish(1) (test fail).\n9. If no failures, call finish(0) (test pass).",
    "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10"
]

meta_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

for col_idx, header in enumerate(meta_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = meta_fill
    cell.alignment = header_align
    cell.border = thin_border

for col_idx, value in enumerate(meta_data, 1):
    cell = ws2.cell(row=2, column=col_idx, value=value)
    cell.alignment = cell_align
    cell.border = thin_border

col_widths_meta = [8, 15, 28, 22, 70, 70, 45]
for i, w in enumerate(col_widths_meta, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A2"
ws2.sheet_state = 'veryHidden'

wb.save("/tmp/GPIO_TestPlan_20250620_154500.xlsx")
print("DONE")
