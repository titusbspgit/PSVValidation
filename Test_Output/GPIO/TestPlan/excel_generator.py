#!/usr/bin/env python3
"""
TestPlan Excel Generator - Agent 7
Generates GPIO_TestPlan_YYYYMMDD_HHMMSS.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timezone, timedelta
import os, json, base64

# IST timezone
IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime("%Y%m%d_%H%M%S")
filename = f"GPIO_TestPlan_{timestamp}.xlsx"

# Input data
data = [
    {
        "index": 1,
        "testcase_name": "gpio_reg_wr_rd_test",
        "ss_module": "GPIO",
        "feature": "Register Read Write Validation",
        "test_description": "This test validates the GPIO register block by performing reset value verification and write-read verification across all 49 GPIO registers. First, it reads each register after reset and verifies the value matches the expected default. Second, it writes six distinct data patterns (all-ones, alternating bits, nibble patterns, and half-word pattern) to each writable register, reads back the value, and verifies correctness using the register's read and write masks. Registers that are not readable or not writable are automatically skipped. VRRW-type registers are excluded from write-read checks as configured. The test reports pass if all comparisons succeed, or fail if any mismatch is detected.",
        "speed": "NA",
        "mode": "NA",
        "remarks": "The soft reset check function is disabled (compiled out). VRRW-type registers (interrupt raw status/clear, IO control groups, data-out groups, data-in groups) are skipped during write-read verification via the skip array. The reset value check masks out bit 0 using 0xFFFFFFFE to handle data_in pin level variability. The register block uses AHB bus interface with 32-bit address width and hreset_n as the active-low asynchronous reset signal.",
        "test_steps": "1. Initialize the test environment and load the GPIO register configuration arrays containing 49 register entries with their addresses, default values, read masks, write masks, and skip conditions.\n2. Perform reset value verification: For each of the 49 GPIO registers (gp0_gpio_8 through the full register set), read the register value after reset.\n3. Skip registers marked in the reset-skip configuration or registers that are not readable (read mask is zero).\n4. Compare each read value (masked with 0xFFFFFFFE to exclude bit 0 variability) against the expected default value from the register specification.\n5. Record any reset value mismatches as failures.\n6. Perform write-read verification: For each of six test patterns (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555, 0xF5F5F5F5, 0xA5A5A5A5, 0xFFFF0000), write the pattern (masked with the write mask) to each writable register.\n7. Skip registers marked in the write-read skip configuration, non-writable registers (write mask is zero), and VRRW-type registers.\n8. Read back each written register and apply the read mask.\n9. Compute the expected read-back value accounting for read-only bits retaining their default values and writable bits reflecting the written pattern.\n10. Compare the actual read-back value against the computed expected value and record any mismatches.\n11. Report overall test result: PASS if no mismatches detected in either reset value check or write-read check; FAIL otherwise.",
        "impacted_registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
        "validation_criteria": "1. All GPIO registers must return their expected default values after reset (with bit 0 masked out). 2. For each of the six write patterns, all writable register fields must correctly reflect the written data when read back through the read mask. 3. Read-only fields must retain their default values regardless of write attempts. 4. Non-readable and non-writable registers must be correctly skipped without errors. 5. VRRW-type registers must be excluded from write-read checks as configured. 6. The test must report PASS (finish 0) when all checks succeed and FAIL (finish 1) when any mismatch is detected.",
        "meta_test_description": "This testcase validates the GPIO register block (gp0) by performing two main checks...",
        "meta_test_steps": "1. Include headers: stdio.h, stdlib.h, test_common.h...",
        "meta_impacted_registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10",
    }
]

# Create workbook
wb = openpyxl.Workbook()
ws_tp = wb.active
ws_tp.title = "TestPlan"
ws_md = wb.create_sheet("MetaData")

# Headers
tp_headers = ["Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
              "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
              "Test Steps / Procedure", "Impacted Registers", "Validation / Acceptance Criteria", "Code Generation"]

md_headers = ["Index", "Test Case Name", "Meta Test Description", "Meta Test Steps / Procedure",
              "Meta Impacted Registers", "Meta Validation / Acceptance Criteria",
              "Meta Headers", "Meta Macros", "Meta Arrays"]

# Formatting
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
wrap_alignment = Alignment(wrap_text=True, vertical="top")

# Write TestPlan headers
for col, header in enumerate(tp_headers, 1):
    cell = ws_tp.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_alignment

# Write MetaData headers
for col, header in enumerate(md_headers, 1):
    cell = ws_md.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_alignment

# Populate data
for row_idx, item in enumerate(data, 2):
    # TestPlan row
    ws_tp.cell(row=row_idx, column=1, value=item.get("index", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=2, value=item.get("ss_module", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=3, value=item.get("feature", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=4, value=item.get("testcase_name", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=5, value=item.get("test_description", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=6, value=item.get("speed", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=7, value=item.get("mode", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=8, value=item.get("memory_start_offset", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=9, value=item.get("memory_end_offset", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=10, value=item.get("remarks", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=11, value=item.get("test_steps", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=12, value=item.get("impacted_registers", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=13, value=item.get("validation_criteria", "")).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=14, value=item.get("code_generation", "")).alignment = wrap_alignment

    # MetaData row
    ws_md.cell(row=row_idx, column=1, value=item.get("index", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=2, value=item.get("testcase_name", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=3, value=item.get("meta_test_description", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=4, value=item.get("meta_test_steps", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=5, value=item.get("meta_impacted_registers", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=6, value=item.get("meta_validation_criteria", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=7, value=item.get("meta_headers", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=8, value=item.get("meta_macros", "")).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=9, value=item.get("meta_arrays", "")).alignment = wrap_alignment

# Freeze first row
ws_tp.freeze_panes = "A2"
ws_md.freeze_panes = "A2"

# Auto-size columns
for ws in [ws_tp, ws_md]:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 12)

# MetaData veryHidden
ws_md.sheet_state = "veryHidden"

# Save
output_path = filename
wb.save(output_path)
print(f"GENERATED: {output_path}")
print(f"SIZE: {os.path.getsize(output_path)}")
