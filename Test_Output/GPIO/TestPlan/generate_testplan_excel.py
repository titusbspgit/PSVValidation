#!/usr/bin/env python3
"""
GPIO Test Plan Excel Generator
================================
Generates a real .xlsx workbook with:
  - Sheet 1: "TestPlan" (visible) — 14 columns, 3 data rows
  - Sheet 2: "MetaData" (VeryHidden) — 9 columns, 3 data rows

Formatting:
  - Header row: Bold, Blue background (#4472C4), White font
  - All cells: Wrap text enabled
  - Column widths: Auto-adjusted for readability
  - First row frozen in both sheets
  - MetaData sheet state: veryHidden

Output:
  - File: testplan_<YYYYMMDD_HHMMSS>.xlsx  (timestamp in IST)
  - Saved in the same directory as this script

Usage:
  python generate_testplan_excel.py
"""

import json
import os
import sys
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ============================================================
# EMBEDDED JSON DATA — 3 test case entries (exact from workflow)
# ============================================================
JSON_DATA = [
    {
        "Index": "1",
        "SS / Module": "GPIO",
        "Feature": "GPIO Register Read/Write and Reset Verification",
        "Test Case Name": "gpio_reg_wr_rd_test",
        "Test Description": "Validates reset default values and masked read/write behavior of GPIO GP0 pad registers (pads 8 through 27). For each register, readable bits are checked against reset defaults; then multiple data patterns are written to writable bits and read back to confirm expected values while non-writable bits retain their defaults.",
        "Meta Test Description": "The test exercises a subset of GPIO GP0 pad registers (indices 8..27) using arrays of register addresses, default values, read masks, and write masks. It performs two phases: (1) default value check per register with a readback masked by 0xfffffffe before comparing against the expected default; (2) write/read verification using six predefined 32-bit patterns. For each register, writes apply only to writable bits using the per-register write mask, and the readback is masked by the per-register read mask. The expected value is constructed by combining the written pattern on readable+writable bits with the retained default on readable but non-writable bits. Fail counters track any mismatches across both phases; test result is pass only if both default and write/read checks report zero failures.",
        "Speed": "NA",
        "Mode": "Polling",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Remarks": "Input fields may read as logic 1 when left un-driven; forcing inputs to 0 can affect internal selection behavior and cause mismatches during default-value reads. Ensure the configured register count aligns with the array lengths (20) to avoid out-of-bounds access during iteration.",
        "Test Steps / Procedure": "1. Initialize the test environment and counters for default and write/read failures. 2. For each GPIO GP0 pad register from gp0_gpio_8 to gp0_gpio_27: a) If marked non-readable, skip default verification; otherwise read the register and compare readable bits against the documented reset defaults. 3. For each of six data patterns, iterate the same register set: a) If a register is marked non-writable, skip the write; otherwise write the pattern to the writable bits only. b) Read back the register and verify that readable+writable bits match the written pattern and that readable but non-writable bits retain their reset defaults. 4. Accumulate per-register failures across both phases. 5. Declare the test PASS if no failures are recorded; otherwise declare FAIL.",
        "Meta Test Steps / Procedure": "test_case(): 1) Call chk_rst_val(); 2) Call chk_rd_wr(); 3) If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0). chk_rst_val(): For i=0..(CNT-1): addr = addr_array[i]; if (skip_rst_array[i] == 1) continue; if (read_mask_array[i] == 0x00000000) continue; data_rd = read_reg(addr); data = (data_rd & 0xfffffffe); if (data == default_value_array[i]) PASS else { def_fail_cnt++; log mismatch }. chk_rd_wr(): Define chk_val[6] = {0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}; For each pattern j: data_wr = chk_val[j]; Write phase: For i=0..(CNT-1): addr = addr_array[i]; if (skip_array[i] == 1) continue; if (write_mask_array[i] == 0x00000000) continue; write_reg(addr, (data_wr & write_mask_array[i])); Read/verify phase: For i=0..(CNT-1): addr = addr_array[i]; if (skip_array[i] == 1) continue; if (write_mask_array[i] == 0x00000000) continue; if (read_mask_array[i] == 0x00000000) continue; data_rd = (read_reg(addr) & read_mask_array[i]); wr_n = (write_mask_array[i] ^ 0xffffffff); exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if (data_rd == exp_val) PASS else { wr_fail_cnt++; log mismatch }. soft_reset_chk(): Defined but compiled out (#ifdef 0); would write/restore a soft-reset register at 0x00000000 with waits if enabled.",
        "Impacted Registers": "gp0_gpio_8, gp0_gpio_9, gp0_gpio_10, gp0_gpio_11, gp0_gpio_12, gp0_gpio_13, gp0_gpio_14, gp0_gpio_15, gp0_gpio_16, gp0_gpio_17, gp0_gpio_18, gp0_gpio_19, gp0_gpio_20, gp0_gpio_21, gp0_gpio_22, gp0_gpio_23, gp0_gpio_24, gp0_gpio_25, gp0_gpio_26, gp0_gpio_27",
        "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, MIZAR_GPIO_GP0_GPIO_11, MIZAR_GPIO_GP0_GPIO_12, MIZAR_GPIO_GP0_GPIO_13, MIZAR_GPIO_GP0_GPIO_14, MIZAR_GPIO_GP0_GPIO_15, MIZAR_GPIO_GP0_GPIO_16, MIZAR_GPIO_GP0_GPIO_17, MIZAR_GPIO_GP0_GPIO_18, MIZAR_GPIO_GP0_GPIO_19, MIZAR_GPIO_GP0_GPIO_20, MIZAR_GPIO_GP0_GPIO_21, MIZAR_GPIO_GP0_GPIO_22, MIZAR_GPIO_GP0_GPIO_23, MIZAR_GPIO_GP0_GPIO_24, MIZAR_GPIO_GP0_GPIO_25, MIZAR_GPIO_GP0_GPIO_26, MIZAR_GPIO_GP0_GPIO_27, GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL, GPIO_GP0_GPIO_11_DEFAULT_VAL, GPIO_GP0_GPIO_12_DEFAULT_VAL, GPIO_GP0_GPIO_13_DEFAULT_VAL, GPIO_GP0_GPIO_14_DEFAULT_VAL, GPIO_GP0_GPIO_15_DEFAULT_VAL, GPIO_GP0_GPIO_16_DEFAULT_VAL, GPIO_GP0_GPIO_17_DEFAULT_VAL, GPIO_GP0_GPIO_18_DEFAULT_VAL, GPIO_GP0_GPIO_19_DEFAULT_VAL, GPIO_GP0_GPIO_20_DEFAULT_VAL, GPIO_GP0_GPIO_21_DEFAULT_VAL, GPIO_GP0_GPIO_22_DEFAULT_VAL, GPIO_GP0_GPIO_23_DEFAULT_VAL, GPIO_GP0_GPIO_24_DEFAULT_VAL, GPIO_GP0_GPIO_25_DEFAULT_VAL, GPIO_GP0_GPIO_26_DEFAULT_VAL, GPIO_GP0_GPIO_27_DEFAULT_VAL, GPIO_GP0_GPIO_8_READ_MASK, GPIO_GP0_GPIO_9_READ_MASK, GPIO_GP0_GPIO_10_READ_MASK, GPIO_GP0_GPIO_11_READ_MASK, GPIO_GP0_GPIO_12_READ_MASK, GPIO_GP0_GPIO_13_READ_MASK, GPIO_GP0_GPIO_14_READ_MASK, GPIO_GP0_GPIO_15_READ_MASK, GPIO_GP0_GPIO_16_READ_MASK, GPIO_GP0_GPIO_17_READ_MASK, GPIO_GP0_GPIO_18_READ_MASK, GPIO_GP0_GPIO_19_READ_MASK, GPIO_GP0_GPIO_20_READ_MASK, GPIO_GP0_GPIO_21_READ_MASK, GPIO_GP0_GPIO_22_READ_MASK, GPIO_GP0_GPIO_23_READ_MASK, GPIO_GP0_GPIO_24_READ_MASK, GPIO_GP0_GPIO_25_READ_MASK, GPIO_GP0_GPIO_26_READ_MASK, GPIO_GP0_GPIO_27_READ_MASK, GPIO_GP0_GPIO_8_WRITE_MASK, GPIO_GP0_GPIO_9_WRITE_MASK, GPIO_GP0_GPIO_10_WRITE_MASK, GPIO_GP0_GPIO_11_WRITE_MASK, GPIO_GP0_GPIO_12_WRITE_MASK, GPIO_GP0_GPIO_13_WRITE_MASK, GPIO_GP0_GPIO_14_WRITE_MASK, GPIO_GP0_GPIO_15_WRITE_MASK, GPIO_GP0_GPIO_16_WRITE_MASK, GPIO_GP0_GPIO_17_WRITE_MASK, GPIO_GP0_GPIO_18_WRITE_MASK, GPIO_GP0_GPIO_19_WRITE_MASK, GPIO_GP0_GPIO_20_WRITE_MASK, GPIO_GP0_GPIO_21_WRITE_MASK, GPIO_GP0_GPIO_22_WRITE_MASK, GPIO_GP0_GPIO_23_WRITE_MASK, GPIO_GP0_GPIO_24_WRITE_MASK, GPIO_GP0_GPIO_25_WRITE_MASK, GPIO_GP0_GPIO_26_WRITE_MASK, GPIO_GP0_GPIO_27_WRITE_MASK",
        "Validation / Acceptance Criteria": "Pass if all selected GPIO GP0 registers (pads 8-27) meet both: (a) default readback of readable bits equals the documented reset defaults, and (b) after writing each test pattern, readable+writable bits equal the written pattern and readable but non-writable bits retain their reset defaults. Any mismatch results in test failure.",
        "Meta Validation / Acceptance Criteria": "Default phase: data = (read_reg(addr) & 0xfffffffe); PASS if (data == default_value_array[i]); otherwise def_fail_cnt++. Write/read phase: data_rd = (read_reg(addr) & read_mask_array[i]); wr_n = (write_mask_array[i] ^ 0xffffffff); exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); PASS if (data_rd == exp_val); otherwise wr_fail_cnt++. Final: if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) else finish(0).",
        "Code Generation (Required / Not)": "Not",
        "Meta Headers": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<gpio/gpio_def.h>\n#include<gpio/gpio_offset.h>",
        "Meta Macros": "#define SOFT_RST_REG_ADDRESS\t0x00000000\n#define SOFT_RST_REG_DATA\t0x00000000\n#define CNT 49",
        "Meta Arrays": "const unsigned long int addr_array[20]={MIZAR_GPIO_GP0_GPIO_8,...,MIZAR_GPIO_GP0_GPIO_27}; const unsigned int default_value_array[20]={GPIO_GP0_GPIO_8_DEFAULT_VAL,...,GPIO_GP0_GPIO_27_DEFAULT_VAL}; const unsigned int read_mask_array[20]={GPIO_GP0_GPIO_8_READ_MASK,...,GPIO_GP0_GPIO_27_READ_MASK}; const unsigned int write_mask_array[20]={GPIO_GP0_GPIO_8_WRITE_MASK,...,GPIO_GP0_GPIO_27_WRITE_MASK}; const unsigned int skip_array[20]={0,...,0}; const unsigned int skip_rst_array[20]={0,...,0}; unsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000};"
    },
    {
        "Test_Plan_ID": "NA",
        "IP_NAME": "GPIO",
        "Test_Name": "test_gpio_negedge_intr_en",
        "Test_Folder": "TestRepo/gpio/test_gpio_negedge_intr_en",
        "Test_Type": "Interrupt",
        "Test_Category": "Interrupt Verification",
        "Test_Description": "Validates negative-edge GPIO pad interrupts across 32 pads (pads 8-39). The test enables per-pad interrupts, configures pad control for negative-edge detection, then generates a falling edge on each pad in sequence and verifies interrupt status, clearing, and GIC handling.",
        "Source_File(s)": "program.c",
        "Define_File(s)": "test_define.c",
        "Registers_Under_Test": "Per-pad GPIO registers (base at pad 8, 32 consecutive pads); GPIO interrupt raw status/clear (set/clear) register; GPIO interrupt enable register; GPIO group interrupt status register; System interrupt enable and raw status/clear registers; IO control/group register at 0xA0243ffc",
        "Register_Count": "32",
        "Key_Functions": "test_case(), Default_IRQHandler(), write_reg(), read_reg(), wait_on(), finish(), GIC_EnableIRQ(), GIC_ClearIRQ()",
        "Includes": "<stdio.h>, <lss_sysreg.h>, <test_common.h>, <gpio/gpio_def.h>, <gpio/gpio_offset.h>",
        "Interrupt_Used": "Yes",
        "IRQ_Numbers": "87 (GPIO0), 88 (GPIO1)",
        "GIC_Enabled": "Yes",
        "Pass_Criteria": "For each pad: interrupt is observed without timeout; per-pad/group interrupt status reflects only the targeted pad; status is fully cleared after service; no unexpected input/state bits remain set; final error count is zero.",
        "Fail_Criteria": "Timeout waiting for the negative-edge interrupt; incorrect input/state bit after edge; group status bit not set for targeted pad; status not cleared to zero after service; any error increments the test error counter.",
        "Test_Patterns": "Drive all pads high via IO control, then for each pad drive a single pad low to create a falling edge; repeat across 32 pads.",
        "GPIO_Pads_Tested": "32 (pads 8-39)",
        "Interrupt_Type": "Negative edge",
        "Interrupt_Clear_Method": "Write to the GPIO raw status/clear register for the specific pad; clear the corresponding system raw status/clear register; clear the GIC pending interrupt.",
        "Sysreg_Integration": "System interrupt enable selects the GPIO instance and routes it to the GIC (IRQ 87 for GPIO0 or 88 for GPIO1). System raw status/clear is used to acknowledge at the system level.",
        "IO_Control_Groups": "IO control/group register at 0xA0243ffc is used to set all pads high and then force a single pad low; per-pad control fields (e.g., enable and edge configuration bits) are programmed in the pad registers.",
        "Notes": "GPIO instance selection (GPIO0 vs GPIO1) is compile-time controlled; ensure IO control can actively drive pad levels to create a valid falling edge; ensure no external contention; delays are required for hardware settling between level changes.",
        "GitHub_Source_URL": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en"
    },
    {
        "Test_Plan_ID": "NA",
        "IP_NAME": "GPIO",
        "Test_Name": "test_gpio_pedge_all_pads_en",
        "Test_Folder": "TestRepo/gpio/test_gpio_pedge_all_pads_en",
        "Test_Type": "Interrupt",
        "Test_Category": "Interrupt Verification",
        "Test_Description": "Validates positive-edge GPIO pad interrupts across 32 pads (pads 8-39). The test enables per-pad interrupts, configures pad control for rising-edge detection, then generates a rising edge on each pad in sequence and verifies interrupt status, clearing, and GIC handling including group aggregation.",
        "Source_File(s)": "program.c",
        "Define_File(s)": "test_define.c",
        "Registers_Under_Test": "Per-pad GPIO registers for 32 pads (starting from pad 8); GPIO interrupt enable register; GPIO per-pad/group interrupt status register(s); System interrupt enable and raw status/clear registers; IO control/group registers to drive pad levels",
        "Register_Count": "32",
        "Key_Functions": "test_case(), Default_IRQHandler(), write_reg(), read_reg(), wait_on(), finish(), GIC_EnableIRQ(), GIC_ClearIRQ()",
        "Includes": "<stdio.h>, <lss_sysreg.h>, <test_common.h>, <gpio/gpio_def.h>, <gpio/gpio_offset.h>",
        "Interrupt_Used": "Yes",
        "IRQ_Numbers": "87 (GPIO0), 88 (GPIO1)",
        "GIC_Enabled": "Yes",
        "Pass_Criteria": "For each pad, a rising edge triggers an interrupt without timeout; the correct per-pad and group interrupt status bit sets; service clears all related status to zero; no unexpected bits remain set; final error count is zero.",
        "Fail_Criteria": "Timeout waiting for the positive-edge interrupt; unexpected input/state after the edge; group status not reflecting the targeted pad; status not fully cleared after service; any such event increments the error counter.",
        "Test_Patterns": "Drive all pads low via IO control; for each iteration, drive one pad high to create a rising edge; repeat across all 32 pads.",
        "GPIO_Pads_Tested": "32 (pads 8-39)",
        "Interrupt_Type": "Positive edge",
        "Interrupt_Clear_Method": "Clear the GPIO per-pad raw status/clear register for the serviced pad; acknowledge at the system level using the system raw status/clear register; clear any pending GIC interrupt.",
        "Sysreg_Integration": "The selected GPIO instance is enabled in the system interrupt controller and routed to GIC IRQ 87 (GPIO0) or 88 (GPIO1). System raw status/clear is used for top-level acknowledge.",
        "IO_Control_Groups": "IO control/group registers are used to set all pads low and then force a single pad high per iteration; per-pad control fields are configured for positive-edge detection.",
        "Notes": "Ensure IO control actively drives pad levels to create clean rising edges; avoid external contention on the pads; allow settling delays between transitions; build selects GPIO instance (GPIO0 vs GPIO1) consistently with IRQ routing.",
        "GitHub_Source_URL": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en"
    }
]

# ============================================================
# COLUMN DEFINITIONS
# ============================================================
TESTPLAN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_COLUMNS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

# ============================================================
# FIELD MAPPING — normalizes heterogeneous JSON keys
# ============================================================
# For each target column, list possible source keys in priority order
TESTPLAN_FIELD_MAP = {
    "Index":                          ["Index", "Test_Plan_ID"],
    "SS / Module":                    ["SS / Module", "IP_NAME"],
    "Feature":                        ["Feature", "Test_Category"],
    "Test Case Name":                 ["Test Case Name", "Test_Name"],
    "Test Description":               ["Test Description", "Test_Description"],
    "Speed":                          ["Speed"],
    "Mode":                           ["Mode", "Interrupt_Type"],
    "Memory Start Offset":            ["Memory Start Offset"],
    "Memory End Offset":              ["Memory End Offset"],
    "Remarks":                        ["Remarks", "Notes"],
    "Test Steps / Procedure":         ["Test Steps / Procedure", "Test_Patterns"],
    "Impacted Registers":             ["Impacted Registers", "Registers_Under_Test"],
    "Validation / Acceptance Criteria":["Validation / Acceptance Criteria", "Pass_Criteria"],
    "Code Generation (Required / Not)":["Code Generation (Required / Not)"],
}

METADATA_FIELD_MAP = {
    "Index":                              ["Index", "Test_Plan_ID"],
    "Test Case Name":                     ["Test Case Name", "Test_Name"],
    "Meta Test Description":              ["Meta Test Description", "Test_Description"],
    "Meta Test Steps / Procedure":        ["Meta Test Steps / Procedure", "Test_Patterns"],
    "Meta Impacted Registers":            ["Meta Impacted Registers", "Registers_Under_Test"],
    "Meta Validation / Acceptance Criteria":["Meta Validation / Acceptance Criteria", "Pass_Criteria"],
    "Meta Headers":                       ["Meta Headers", "Includes"],
    "Meta Macros":                        ["Meta Macros", "IRQ_Numbers"],
    "Meta Arrays":                        ["Meta Arrays"],
}


def resolve_field(entry, candidates, default="NA"):
    """Return the value of the first matching key found in entry."""
    for key in candidates:
        if key in entry and entry[key] is not None and str(entry[key]).strip():
            return str(entry[key])
    return default


def build_rows(data, columns, field_map):
    """Build a list-of-lists (rows) from JSON data using the field map."""
    rows = []
    for idx, entry in enumerate(data, start=1):
        row = []
        for col in columns:
            candidates = field_map.get(col, [col])
            val = resolve_field(entry, candidates)
            # Override Index with sequential number if source is NA or missing
            if col == "Index" and (val == "NA" or not val.strip()):
                val = str(idx)
            row.append(val)
        rows.append(row)
    return rows


def auto_col_width(ws, min_width=12, max_width=60):
    """Set column widths based on content length."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                cell_len = max(len(line) for line in str(cell.value).split("\n")) if cell.value else 0
            except Exception:
                cell_len = 0
            if cell_len > max_len:
                max_len = cell_len
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def create_workbook(output_dir):
    """Create the Excel workbook and save it."""

    # --- Timestamp in IST ---
    IST = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(IST)
    ts = now_ist.strftime("%Y%m%d_%H%M%S")
    filename = f"testplan_{ts}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # --- Build row data ---
    tp_rows = build_rows(JSON_DATA, TESTPLAN_COLUMNS, TESTPLAN_FIELD_MAP)
    md_rows = build_rows(JSON_DATA, METADATA_COLUMNS, METADATA_FIELD_MAP)

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # ===================== Sheet 1: TestPlan =====================
    ws_tp = wb.active
    ws_tp.title = "TestPlan"

    # Header row
    for c, col_name in enumerate(TESTPLAN_COLUMNS, start=1):
        cell = ws_tp.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align

    # Data rows
    for r, row_data in enumerate(tp_rows, start=2):
        for c, val in enumerate(row_data, start=1):
            cell = ws_tp.cell(row=r, column=c, value=val)
            cell.alignment = wrap_align

    ws_tp.freeze_panes = "A2"
    auto_col_width(ws_tp)

    # ===================== Sheet 2: MetaData =====================
    ws_md = wb.create_sheet(title="MetaData")

    # Header row
    for c, col_name in enumerate(METADATA_COLUMNS, start=1):
        cell = ws_md.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align

    # Data rows
    for r, row_data in enumerate(md_rows, start=2):
        for c, val in enumerate(row_data, start=1):
            cell = ws_md.cell(row=r, column=c, value=val)
            cell.alignment = wrap_align

    ws_md.freeze_panes = "A2"
    auto_col_width(ws_md)

    # --- Set MetaData sheet to VeryHidden ---
    ws_md.sheet_state = "veryHidden"

    # --- Save ---
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    print(f"SUCCESS: Workbook saved to {filepath}")
    print(f"  - TestPlan sheet: {len(tp_rows)} data rows, {len(TESTPLAN_COLUMNS)} columns")
    print(f"  - MetaData sheet: {len(md_rows)} data rows, {len(METADATA_COLUMNS)} columns (veryHidden)")
    print(f"  - Filename: {filename}")
    return filepath, filename


if __name__ == "__main__":
    # Output directory = same directory as this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filepath, filename = create_workbook(script_dir)
    # Write filename to env for GitHub Actions
    gh_output = os.environ.get("GITHUB_OUTPUT", "")
    if gh_output:
        with open(gh_output, "a") as f:
            f.write(f"excel_filename={filename}\n")
            f.write(f"excel_filepath={filepath}\n")
