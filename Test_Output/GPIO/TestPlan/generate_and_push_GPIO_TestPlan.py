#!/usr/bin/env python3
"""
GPIO TestPlan Generator + GitHub Pusher
========================================
Generates GPIO_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx (IST) and pushes to GitHub.

Usage:
    pip install openpyxl requests
    GITHUB_TOKEN=ghp_xxx python generate_and_push_GPIO_TestPlan.py

Output:
    - TestPlan sheet: 28 columns x 2 data rows (visible)
    - MetaData sheet: 11 columns x 2 data rows (veryHidden)
    - Styling: Dark blue #003366 headers, white bold Calibri 11pt, thin borders
    - Pushed to: Test_Output/GPIO/TestPlan/GPIO_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx
"""
import os, sys, base64, json, requests
from datetime import datetime, timezone, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# IST TIMESTAMP
# ============================================================
IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
ts_file = now_ist.strftime("%Y%m%d_%H%M%S")
ts_display = now_ist.strftime("%Y-%m-%d %H:%M:%S IST")
FILENAME = f"GPIO_TestPlan_{ts_file}.xlsx"

# ============================================================
# COLUMN HEADERS
# ============================================================
TP_HEADERS = [
    "Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
    "Meta Test Description", "Speed", "Mode", "Memory Start Offset",
    "Memory End Offset", "Remarks", "Test Steps / Procedure",
    "Meta Test Steps / Procedure", "Impacted Registers",
    "Meta Impacted Registers", "Validation / Acceptance Criteria",
    "Meta Validation / Acceptance Criteria", "Code Generation (Required / Not)",
    "Meta Headers", "Meta Macros", "Meta Arrays", "Register Mapping",
    "Resolved Base Address", "Source Folder", "Source Files", "IP Name",
    "Base Define", "Generation Timestamp (IST)"
]

MD_HEADERS = [
    "Index", "Test Case Name", "Meta Test Description",
    "Meta Test Steps / Procedure", "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria", "Meta Headers", "Meta Macros",
    "Meta Arrays", "Register Mapping", "Resolved Base Address"
]

# ============================================================
# FINAL AGGREGATED DATA — 2 TESTCASES (IMMUTABLE)
# ============================================================
FINAL_JSON = [
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "Register Read/Write",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "This test validates the GPIO register block by performing two checks: (1) Default value verification \u2014 reads each GPIO register after reset and verifies the value matches the expected default, and (2) Write-Read verification \u2014 writes six distinct data patterns (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555, 0xF5F5F5F5, 0xA5A5A5A5, 0xFFFF0000) to each writable GPIO register and reads back to confirm the written value is retained correctly. The registers under test are gp0_gpio_8, gp0_gpio_9, and gp0_gpio_10. Certain registers are skipped based on skip arrays and mask configurations. The test reports PASS if all default value checks and all write-read checks succeed, and FAIL if any mismatch is detected.",
    "Meta Test Description": "The testcase gpio_reg_wr_rd_test performs two phases of register validation on the GPIO IP block. Phase 1 (chk_rst_val): Iterates over addr_array[0..CNT-1] where CNT=49. For each index, checks skip_rst_array[i]; if 1, skips. Checks read_mask_array[i]; if 0x00000000, skips (not readable). Otherwise calls read_reg(addr_array[i]), masks result with 0xfffffffe, and compares to default_value_array[i]. On mismatch, increments def_fail_cnt. Phase 2 (chk_rd_wr): Iterates over 6 patterns in chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}. For each pattern: (a) Write pass \u2014 iterates addr_array, skips if skip_array[i]==1 or write_mask_array[i]==0x00000000, otherwise calls write_reg(addr, data_wr & write_mask_array[i]). (b) Read-verify pass \u2014 iterates addr_array, skips if skip_array[i]==1 or write_mask_array[i]==0x00000000 or read_mask_array[i]==0x00000000, otherwise reads data_rd = read_reg(addr) & read_mask_array[i], computes wr_n = write_mask_array[i] ^ 0xffffffff, computes exp_val = (data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]), compares data_rd to exp_val, increments wr_fail_cnt on mismatch. Finally test_case() checks if def_fail_cnt > 0 || wr_fail_cnt > 0: calls finish(1) for FAIL, finish(0) for PASS. The addr_array contains MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10. A soft_reset_chk() function exists but is disabled inside #ifdef 0 (dead code).",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Certain registers are skipped during default value check and write-read check based on skip array and reset-skip array configurations. Registers with a read mask of 0x00000000 are skipped for read operations, and registers with a write mask of 0x00000000 are skipped for write operations. The default value comparison masks the read data with 0xFFFFFFFE (bit 0 excluded). The din bit may auto-set to 1 if no value is forced, which can affect default value matching for level-select related bits.",
    "Test Steps / Procedure": "1. Initialize the test environment and load the GPIO register address table, default value table, read mask table, and write mask table.\n2. Perform default value verification: For each GPIO register (gp0_gpio_8, gp0_gpio_9, gp0_gpio_10), read the register value after reset and compare it against the expected default value. Skip registers flagged in the reset-skip configuration or those that are not readable.\n3. Record any default value mismatches.\n4. Perform write-read verification with pattern 0xFFFFFFFF: Write the pattern to each writable GPIO register applying the write mask, then read back each register applying the read mask and verify the read value matches the expected value.\n5. Perform write-read verification with pattern 0xAAAAAAAA: Repeat the write-read cycle and verify.\n6. Perform write-read verification with pattern 0x55555555: Repeat the write-read cycle and verify.\n7. Perform write-read verification with pattern 0xF5F5F5F5: Repeat the write-read cycle and verify.\n8. Perform write-read verification with pattern 0xA5A5A5A5: Repeat the write-read cycle and verify.\n9. Perform write-read verification with pattern 0xFFFF0000: Repeat the write-read cycle and verify.\n10. Evaluate overall result: If all default value checks and all write-read checks pass with zero mismatches, report PASS. Otherwise report FAIL.",
    "Meta Test Steps / Procedure": "1. Entry: test_case() is called.\n2. Call chk_rst_val():\n   a. Loop i = 0 to CNT-1 (CNT=49).\n   b. addr = addr_array[i] (contains MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10).\n   c. If skip_rst_array[i] == 1, skip (continue).\n   d. If read_mask_array[i] == 0x00000000, skip (not readable).\n   e. data_rd = read_reg(addr).\n   f. data = (data_rd & 0xfffffffe).\n   g. If data == default_value_array[i], PASS for this register.\n   h. Else, def_fail_cnt++ and print failure with addr, expected, and read values.\n3. Call chk_rd_wr():\n   a. Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}.\n   b. Outer loop j = 0 to 5: data_wr = chk_val[j].\n   c. Write pass: Loop i = 0 to CNT-1.\n      - addr = addr_array[i].\n      - If skip_array[i] == 1, skip.\n      - If write_mask_array[i] == 0x00000000, skip (not writable).\n      - Else write_reg(addr, data_wr & write_mask_array[i]).\n   d. Read-verify pass: Loop i = 0 to CNT-1.\n      - addr = addr_array[i].\n      - If skip_array[i] == 1, skip.\n      - If write_mask_array[i] == 0x00000000, skip.\n      - If read_mask_array[i] == 0x00000000, skip.\n      - Else: data_rd = read_reg(addr) & read_mask_array[i].\n      - wr_n = write_mask_array[i] ^ 0xffffffff.\n      - exp_val = (data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]).\n      - If data_rd == exp_val, PASS.\n      - Else wr_fail_cnt++ and print failure.\n4. Final check in test_case(): If def_fail_cnt > 0 || wr_fail_cnt > 0, call finish(1) [FAIL]. Else call finish(0) [PASS].\n5. soft_reset_chk() exists but is disabled (#ifdef 0 \u2014 dead code). Contains read_reg(SOFT_RST_REG_ADDRESS), write_reg(SOFT_RST_REG_ADDRESS, SOFT_RST_REG_DATA), wait_on(1000) \u2014 all inactive.",
    "Impacted Registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10",
    "Validation / Acceptance Criteria": "PASS: All GPIO registers (gp0_gpio_8, gp0_gpio_9, gp0_gpio_10) return their expected default values after reset, and all six write-read pattern cycles (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555, 0xF5F5F5F5, 0xA5A5A5A5, 0xFFFF0000) produce matching read-back values when compared against the expected value computed from write mask, read mask, and default value. FAIL: Any default value mismatch or any write-read mismatch is detected in any register for any pattern.",
    "Meta Validation / Acceptance Criteria": "Default value check PASS condition: (read_reg(addr_array[i]) & 0xfffffffe) == default_value_array[i] for all non-skipped, readable registers. Default value check FAIL condition: data != default_value_array[i] \u2192 def_fail_cnt++. Write-Read check PASS condition: (read_reg(addr) & read_mask_array[i]) == ((data_wr & read_mask_array[i] & write_mask_array[i]) | ((write_mask_array[i] ^ 0xffffffff) & read_mask_array[i] & default_value_array[i])) for each of 6 patterns {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}. Write-Read check FAIL condition: data_rd != exp_val \u2192 wr_fail_cnt++. Overall PASS: def_fail_cnt == 0 && wr_fail_cnt == 0 \u2192 finish(0). Overall FAIL: def_fail_cnt > 0 || wr_fail_cnt > 0 \u2192 finish(1).",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>; #include <stdlib.h>; #include \"test_common.h\"; #include \"test_define.c\"; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>",
    "Meta Macros": "#define SOFT_RST_REG_ADDRESS 0x00000000; #define SOFT_RST_REG_DATA 0x00000000; #define CNT 49",
    "Meta Arrays": "const unsigned long int addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,MIZAR_GPIO_GP0_GPIO_10,}; const unsigned int default_value_array[49]={GPIO_GP0_GPIO_8_DEFAULT_VAL,GPIO_GP0_GPIO_9_DEFAULT_VAL,GPIO_GP0_GPIO_10_DEFAULT_VAL,}; const unsigned int read_mask_array[49]={GPIO_GP0_GPIO_8_READ_MASK,GPIO_GP0_GPIO_9_READ_MASK,GPIO_GP0_GPIO_10_READ_MASK,}; const unsigned int write_mask_array[49]={GPIO_GP0_GPIO_8_WRITE_MASK,GPIO_GP0_GPIO_9_WRITE_MASK,GPIO_GP0_GPIO_10_WRITE_MASK,}; const unsigned int skip_array[49]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,1,1,1,1,1,1,1,1,0,0,0,0,}; const unsigned int skip_rst_array[49]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,1,1,1,1,1,1,1,1,1,1,1,}; unsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}",
    "Register Mapping": "MIZAR_GPIO_GP0_GPIO_8=gp0_gpio_8; MIZAR_GPIO_GP0_GPIO_9=gp0_gpio_9; MIZAR_GPIO_GP0_GPIO_10=gp0_gpio_10",
    "Resolved Base Address": "0xA001A000",
    "Source Folder": "TestRepo/gpio/gpio_reg_wr_rd_test",
    "Source Files": "program.c; test_define.c",
    "IP Name": "GPIO",
    "Base Define": "#define GPIO0 1",
    "Generation Timestamp (IST)": ""
  },
  {
    "Index": "2",
    "SS / Module": "GPIO",
    "Feature": "Level Select Interrupt Enable",
    "Test Case Name": "test_gpio_level_sel_intr_en",
    "Test Description": "This test validates the GPIO level-select interrupt enable functionality for 32 GPIO pads (GPIO 8 through GPIO 39). It performs two phases: (1) Active-high level interrupt \u2014 configures each GPIO pin for input mode with active-high level-select interrupt detection (bits 19 and 20 set), enables the group interrupt, triggers the interrupt, and verifies that the raw interrupt status bit is set in the individual GPIO register and the group interrupt status register (gp0_intr1_intr_sts1). It then clears the interrupt and verifies the clear operation. (2) Active-low level interrupt \u2014 configures each GPIO pin for input mode with active-low level-select interrupt detection (bit 20 set, bit 19 clear), enables the group interrupt, triggers the interrupt with the inverted pad value, and performs the same verification and clear sequence. The test also verifies system-level interrupt status clearing. The test reports PASS if all 32 pads pass both active-high and active-low interrupt checks with zero errors, and FAIL otherwise.",
    "Meta Test Description": "The testcase test_gpio_level_sel_intr_en validates level-triggered GPIO interrupt functionality on the GPIO IP block for 32 pads. It uses ISR-driven methodology with Default_IRQHandler(). Phase 1 (Active-high level): In test_case(), GIC_EnableIRQ(87) is called (GPIO0 active). write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR) enables the system-level interrupt. Loop i=0 to 31: (a) write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00180000) configures input mode with active-high level interrupt (bits 19,20 = 1,1). (b) wait_on(50). (c) wr_val = 1<<i; write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val) enables group interrupt for pad i. (d) wait_on(10). (e) write_reg(0xA0243ffc, 0xffffffff) writes SRAM trigger. (f) int_pend=1; while(int_pend==1) polls waiting for ISR. Phase 2 (Active-low level): Loop i=0 to 31: (a) write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00100000) configures input mode with active-low level interrupt (bit 20=1, bit 19=0). (b) wait_on(50). (c) wr_val = 1<<i; write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val). (d) wait_on(10). (e) write_reg(0xA0243ffc, ~(wr_val)) writes inverted trigger. (f) int_pend=1; while(int_pend==1) polls waiting for ISR. In Default_IRQHandler(): (1) int_pend=0. (2) write_reg(0xA0243ffc, 0xffffffff). (3) rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)). (4) Check (rdata & 0x2) != 0x0 for raw interrupt status. (5) rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); check (rdata_grp & (1<<i)) != 0 for group interrupt. (6) write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00110000) to clear interrupt (bit 16 set). (7) wait_on(20). (8) rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)); check rdata == 0x100001 for successful clear. (9) write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000) disables group interrupt. (10) rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); check rdata_grp == 0x0 for group interrupt cleared. (11) write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR) clears system-level status. (12) rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); check (rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0. (13) GIC_ClearIRQ(87). On any failure, test_err is incremented. Final: finish(test_err) \u2014 PASS if test_err==0, FAIL otherwise.",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "This test requires external stimulus or pin-level forcing to trigger the level-based interrupt on each GPIO input pin. The test iterates over 32 GPIO pads (GPIO 8 through GPIO 39) using address arithmetic based on the gp0_gpio_8 base register with 4-byte stride. The system-level interrupt enable and status registers belong to an external LSS_SYSREG module and are not part of the GPIO register specification. The hardcoded address 0xA0243ffc is used as an SRAM trigger location. GIC IRQ number 87 is used for GPIO0 interrupt routing.",
    "Test Steps / Procedure": "1. Enable the GIC interrupt for the GPIO block.\n2. Enable the system-level interrupt for the GPIO block in the system interrupt controller.\n3. For each of the 32 GPIO pads (pad 0 through pad 31), perform the active-high level interrupt test:\n   a. Configure the GPIO pin register (gp0_gpio_8 + offset) for input mode with active-high level-select interrupt detection.\n   b. Enable the group interrupt for the current pad in the interrupt enable register (gp0_intr1_intr_en1).\n   c. Write the SRAM trigger value to initiate the interrupt.\n   d. Wait for the interrupt service routine to execute.\n4. For each of the 32 GPIO pads, perform the active-low level interrupt test:\n   a. Configure the GPIO pin register for input mode with active-low level-select interrupt detection.\n   b. Enable the group interrupt for the current pad.\n   c. Write the inverted trigger value to initiate the interrupt.\n   d. Wait for the interrupt service routine to execute.\n5. In the interrupt handler, verify the raw interrupt status bit is set in the individual GPIO register.\n6. Verify the group interrupt status bit is set in the interrupt status register (gp0_intr1_intr_sts1) for the active pad.\n7. Clear the interrupt by writing the clear bit to the GPIO register and verify the register reads the expected cleared value.\n8. Disable the group interrupt and verify the group interrupt status register reads zero.\n9. Clear the system-level interrupt status and verify it is cleared.\n10. Clear the GIC interrupt.\n11. Evaluate overall result: If all pads pass both active-high and active-low interrupt checks with zero errors, report PASS. Otherwise report FAIL.",
    "Meta Test Steps / Procedure": "1. Entry: test_case() is called.\n2. GIC_EnableIRQ(87) \u2014 enable GIC interrupt for GPIO0.\n3. write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR) \u2014 enable system-level GPIO0 interrupt.\n4. Active-high level interrupt loop: for(i = 0; i < 32; i++):\n   a. wr_val = 1 << i.\n   b. write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4), 0x00180000) \u2014 configure input mode, level interrupt, active-high (bits 19,20 = 1,1).\n   c. wait_on(50).\n   d. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val) \u2014 enable group interrupt for pad i.\n   e. wait_on(10).\n   f. write_reg(0xA0243ffc, 0xffffffff) \u2014 write SRAM trigger.\n   g. int_pend = 1; while(int_pend == 1) { printf(\"Waiting for interrupt\"); wait_on(10); } \u2014 poll until ISR clears int_pend.\n5. wait_on(100).\n6. Active-low level interrupt loop: for(i = 0; i < 32; i++):\n   a. wr_val = 1 << i.\n   b. write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4), 0x00100000) \u2014 configure input mode, level interrupt, active-low (bit 20=1, bit 19=0).\n   c. wait_on(50).\n   d. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val) \u2014 enable group interrupt for pad i.\n   e. wait_on(10).\n   f. write_reg(0xA0243ffc, ~(wr_val)) \u2014 write inverted SRAM trigger.\n   g. int_pend = 1; while(int_pend == 1) { printf(\"Waiting for interrupt\"); wait_on(10); }.\n7. finish(test_err) \u2014 PASS if test_err == 0, FAIL if test_err > 0.\n8. Default_IRQHandler() execution:\n   a. wr_val = 1 << i; int_pend = 0.\n   b. write_reg(0xA0243ffc, 0xffffffff).\n   c. rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4)).\n   d. Check (rdata & 0x2) != 0x0 \u2014 raw interrupt status bit set.\n   e. If raw status set:\n      i. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1).\n      ii. Check (rdata_grp & (1 << i)) != 0 \u2014 group interrupt for pad i. If fail: test_err++.\n      iii. write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4), 0x00110000) \u2014 clear interrupt (bit 16 set).\n      iv. wait_on(20).\n      v. rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4)).\n      vi. Check rdata == 0x100001 \u2014 interrupt cleared successfully. If fail: test_err++.\n      vii. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000) \u2014 disable group interrupt.\n      viii. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1).\n      ix. Check rdata_grp == 0x0 \u2014 group interrupt cleared. If fail: test_err++.\n      x. write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR) \u2014 clear system-level status.\n      xi. rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1).\n      xii. Check (rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0 \u2014 system status cleared. If fail: test_err++.\n   f. Else (raw status not set): printf(\"Interrupt Not occured\"); test_err++.\n   g. GIC_ClearIRQ(87).",
    "Impacted Registers": "gp0_gpio_8; gp0_intr1_intr_en1; gp0_intr1_intr_sts1",
    "Meta Impacted Registers": "MIZAR_LSS_SYSREG_INTR_EN1; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1; 0xA0243ffc; MIZAR_GPIO_GP0_INTR1_INTR_STS1; MIZAR_LSS_SYSREG_RAW_STCR1",
    "Validation / Acceptance Criteria": "PASS: For all 32 GPIO pads in both active-high and active-low level interrupt modes: (1) The raw interrupt status bit (bit 1) is set in the individual GPIO register when the interrupt is triggered. (2) The corresponding group interrupt bit is set in the interrupt status register (gp0_intr1_intr_sts1). (3) After clearing the interrupt, the GPIO register reads the expected cleared value (0x100001). (4) After disabling the group interrupt, the interrupt status register (gp0_intr1_intr_sts1) reads zero. (5) The system-level interrupt status is successfully cleared. FAIL: Any of the above checks fail for any pad, or the interrupt does not occur at all, incrementing the error counter. The test reports FAIL if the total error count is greater than zero.",
    "Meta Validation / Acceptance Criteria": "Raw interrupt status check: (read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)) & 0x2) != 0x0 \u2014 PASS if set, FAIL (test_err++) if not set. Group interrupt status check: (read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) & (1<<i)) != 0 \u2014 PASS if set, FAIL (test_err++) if zero. Interrupt clear check: read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)) == 0x100001 \u2014 PASS if match, FAIL (test_err++) if mismatch. Group interrupt clear check: read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) == 0x0 \u2014 PASS if zero, FAIL (test_err++) if non-zero. System-level status clear check: (read_reg(MIZAR_LSS_SYSREG_RAW_STCR1) & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0 \u2014 PASS if cleared, FAIL (test_err++) if still set. Overall: finish(test_err) \u2014 PASS if test_err == 0, FAIL if test_err > 0.",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include<lss_sysreg.h>; #include<stdio.h>; #include<test_define.c>; #include<test_common.h>; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>",
    "Meta Macros": "#define CNT 49",
    "Meta Arrays": "const unsigned long int addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,MIZAR_GPIO_GP0_GPIO_10,}; const unsigned int default_value_array[49]={GPIO_GP0_GPIO_8_DEFAULT_VAL,GPIO_GP0_GPIO_9_DEFAULT_VAL,GPIO_GP0_GPIO_10_DEFAULT_VAL,}; const unsigned int read_mask_array[49]={GPIO_GP0_GPIO_8_READ_MASK,GPIO_GP0_GPIO_9_READ_MASK,GPIO_GP0_GPIO_10_READ_MASK,}; const unsigned int write_mask_array[49]={GPIO_GP0_GPIO_8_WRITE_MASK,GPIO_GP0_GPIO_9_WRITE_MASK,GPIO_GP0_GPIO_10_WRITE_MASK,}; const int skip_array[49]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,}",
    "Register Mapping": "MIZAR_LSS_SYSREG_INTR_EN1=NA; MIZAR_GPIO_GP0_GPIO_8=gp0_gpio_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1=gp0_intr1_intr_en1; 0xA0243ffc=NA; MIZAR_GPIO_GP0_INTR1_INTR_STS1=gp0_intr1_intr_sts1; MIZAR_LSS_SYSREG_RAW_STCR1=NA",
    "Resolved Base Address": "0xA001A000",
    "Source Folder": "TestRepo/gpio/test_gpio_level_sel_intr_en",
    "Source Files": "program.c; test_define.c",
    "IP Name": "GPIO",
    "Base Define": "#define GPIO0 1",
    "Generation Timestamp (IST)": ""
  }
]

# Stamp the generation timestamp
for row in FINAL_JSON:
    row["Generation Timestamp (IST)"] = ts_display

# ============================================================
# STEP 1: GENERATE XLSX
# ============================================================
print("=" * 70)
print("STEP 1: Generating Excel file...")
print("=" * 70)

wb = Workbook()
DARK_BLUE   = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
WHITE_BOLD  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT   = Font(name="Calibri", size=10)
WRAP        = Alignment(horizontal="left", vertical="top", wrap_text=True)
HDR_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# -- TestPlan sheet (visible) --
ws = wb.active
ws.title = "TestPlan"
for c, h in enumerate(TP_HEADERS, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = WHITE_BOLD
    cell.fill = DARK_BLUE
    cell.alignment = HDR_ALIGN
    cell.border = THIN_BORDER

for r_idx, row_data in enumerate(FINAL_JSON, 2):
    for c_idx, h in enumerate(TP_HEADERS, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=row_data.get(h, ""))
        cell.font = DATA_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

ws.freeze_panes = "A2"

# Auto-width (capped at 59)
for c, h in enumerate(TP_HEADERS, 1):
    max_len = len(h)
    for row_data in FINAL_JSON:
        val = str(row_data.get(h, ""))
        first_line = val.split("\n")[0] if val else ""
        max_len = max(max_len, min(len(first_line), 55))
    col_letter = ws.cell(row=1, column=c).column_letter
    ws.column_dimensions[col_letter].width = max_len + 4

# -- MetaData sheet (veryHidden) --
ms = wb.create_sheet("MetaData")
for c, h in enumerate(MD_HEADERS, 1):
    cell = ms.cell(row=1, column=c, value=h)
    cell.font = WHITE_BOLD
    cell.fill = DARK_BLUE
    cell.alignment = HDR_ALIGN
    cell.border = THIN_BORDER

for r_idx, row_data in enumerate(FINAL_JSON, 2):
    for c_idx, h in enumerate(MD_HEADERS, 1):
        cell = ms.cell(row=r_idx, column=c_idx, value=row_data.get(h, ""))
        cell.font = DATA_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

ms.freeze_panes = "A2"
ms.sheet_state = "veryHidden"

# Save to buffer
buf = BytesIO()
wb.save(buf)
xlsx_bytes = buf.getvalue()
buf.close()

# Also save locally
with open(FILENAME, "wb") as f:
    f.write(xlsx_bytes)

print(f"  Filename  : {FILENAME}")
print(f"  Size      : {len(xlsx_bytes)} bytes")
print(f"  Timestamp : {ts_display}")
print(f"  TestPlan  : {len(TP_HEADERS)} cols x {len(FINAL_JSON)} data rows (visible)")
print(f"  MetaData  : {len(MD_HEADERS)} cols x {len(FINAL_JSON)} data rows (veryHidden)")
print(f"  STEP 1 COMPLETE\n")

# ============================================================
# STEP 2: PUSH TO GITHUB
# ============================================================
print("=" * 70)
print("STEP 2: Pushing to GitHub...")
print("=" * 70)

TOKEN = os.environ.get("GITHUB_TOKEN", "")
if not TOKEN:
    print("ERROR: Set GITHUB_TOKEN environment variable.")
    print("Usage: GITHUB_TOKEN=ghp_xxx python generate_and_push_GPIO_TestPlan.py")
    sys.exit(1)

REPO   = "titusbspgit/PSVValidation"
BRANCH = "main"
PATH   = f"Test_Output/GPIO/TestPlan/{FILENAME}"
API    = f"https://api.github.com/repos/{REPO}/contents/{PATH}"
GH_HEADERS = {
    "Authorization": f"token {TOKEN}",
    "Accept": "application/vnd.github.v3+json"
}

# Check if file already exists
existing_sha = None
r = requests.get(API, headers=GH_HEADERS, params={"ref": BRANCH})
if r.status_code == 200:
    existing_sha = r.json().get("sha")

payload = {
    "message": f"Add {FILENAME} - GPIO TestPlan with {len(FINAL_JSON)} testcases (auto-generated)",
    "content": base64.b64encode(xlsx_bytes).decode("ascii"),
    "branch": BRANCH
}
if existing_sha:
    payload["sha"] = existing_sha

r = requests.put(API, headers=GH_HEADERS, json=payload)
if r.status_code not in (200, 201):
    print(f"  PUSH FAILED: HTTP {r.status_code}")
    print(f"  Response: {r.text[:500]}")
    sys.exit(1)

result = r.json()
commit_sha   = result["commit"]["sha"]
html_url     = result["content"]["html_url"]
file_sha     = result["content"]["sha"]
file_size    = result["content"]["size"]

print(f"  Path       : {PATH}")
print(f"  Commit SHA : {commit_sha}")
print(f"  File SHA   : {file_sha}")
print(f"  Size       : {file_size} bytes")
print(f"  HTML URL   : {html_url}")
print(f"  STEP 2 COMPLETE\n")

# ============================================================
# STEP 3: VERIFY
# ============================================================
print("=" * 70)
print("STEP 3: Verifying push...")
print("=" * 70)

v = requests.get(API, headers=GH_HEADERS, params={"ref": BRANCH})
if v.status_code == 200:
    vj = v.json()
    print(f"  File exists  : {vj['name']}")
    print(f"  Size         : {vj['size']} bytes")
    print(f"  SHA          : {vj['sha']}")
    print(f"  Download URL : {vj.get('download_url', 'N/A')}")
    assert vj["name"] == FILENAME, f"Name mismatch: {vj['name']} != {FILENAME}"
    assert vj["sha"] == file_sha, f"SHA mismatch: {vj['sha']} != {file_sha}"
    print(f"  Validation   : ALL CHECKS PASSED")
    print(f"  STEP 3 COMPLETE - VERIFIED\n")
else:
    print(f"  VERIFICATION FAILED: HTTP {v.status_code}")
    sys.exit(1)

# ============================================================
# STEP 4: FINAL SUMMARY
# ============================================================
print("=" * 70)
print("  WORKFLOW EXECUTION COMPLETE")
print("=" * 70)
print(f"  Output File  : {FILENAME}")
print(f"  GitHub Path  : {PATH}")
print(f"  Repository   : {REPO}")
print(f"  Branch       : {BRANCH}")
print(f"  Commit SHA   : {commit_sha}")
print(f"  File SHA     : {file_sha}")
print(f"  File Size    : {file_size} bytes")
print(f"  HTML URL     : {html_url}")
print(f"  Timestamp    : {ts_display}")
print(f"  Testcases    : {len(FINAL_JSON)}")
print(f"    1. gpio_reg_wr_rd_test         [Polling] [Register Read/Write]")
print(f"    2. test_gpio_level_sel_intr_en  [ISR]     [Level Select Interrupt Enable]")
print(f"  Sheets:")
print(f"    TestPlan : {len(TP_HEADERS)} columns x {len(FINAL_JSON)} rows (visible)")
print(f"    MetaData : {len(MD_HEADERS)} columns x {len(FINAL_JSON)} rows (veryHidden)")
print(f"  Styling:")
print(f"    Headers  : Dark Blue #003366, White Bold Calibri 11pt, Centered")
print(f"    Data     : Calibri 10pt, Left/Top, Wrap Text")
print(f"    Borders  : Thin on all cells")
print(f"    Freeze   : Row 1 frozen on both sheets")
print("=" * 70)

# Output structured status
status_output = {
    "status": "SUCCESS",
    "execution_mode": "Fallback Automation",
    "output_file_path": PATH,
    "github_url": html_url,
    "commit_sha": commit_sha,
    "file_sha": file_sha,
    "file_size_bytes": file_size,
    "timestamp_ist": ts_display,
    "testcase_count": len(FINAL_JSON),
    "testplan_columns": len(TP_HEADERS),
    "metadata_columns": len(MD_HEADERS)
}
print("\n" + json.dumps(status_output, indent=2))
