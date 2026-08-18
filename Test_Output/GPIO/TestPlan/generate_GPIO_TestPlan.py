#!/usr/bin/env python3
"""GPIO TestPlan Excel Generator
Generates GPIO_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx with:
  - TestPlan sheet (visible): 28 columns, 2 data rows
  - MetaData sheet (veryHidden): 11 columns, 2 data rows
Pushes to GitHub via PyGithub if GITHUB_TOKEN env var is set.
"""
import json
import os
import sys
import base64
from datetime import datetime, timedelta, timezone
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

# ── Constants ──
OWNER = "titusbspgit"
REPO = "PSVValidation"
BRANCH = "main"
OUTPUT_DIRECTORY = "Test_Output/GPIO/TestPlan"
IP_NAME = "GPIO"

# ── 28 TestPlan columns ──
TESTPLAN_COLS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Meta Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Meta Test Steps / Procedure",
    "Impacted Registers",
    "Meta Impacted Registers",
    "Validation / Acceptance Criteria",
    "Meta Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
    "Register Mapping",
    "Resolved Base Address",
    "Source Folder",
    "Source Files",
    "IP Name",
    "Base Define",
    "Generation Timestamp (IST)"
]

# ── 11 MetaData columns ──
METADATA_COLS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
    "Register Mapping",
    "Resolved Base Address"
]

# ── Row Data (2 testcases) ──
ROW_DATA = [
    {
        "Index": "1",
        "SS / Module": "GPIO",
        "Feature": "NA",
        "Test Case Name": "gpio_reg_wr_rd_test",
        "Test Description": "Validates reset values and read/write behavior of gp0_gpio_8 through gp0_gpio_30. For each register, the test skips entries flagged as non-testable, verifies readable reset values with bit 0 ignored, writes a series of data patterns masked by the write mask, and confirms read-back values match the expected combination of written bits and preserved default bits for non-writable fields. The test reports pass only if all applicable checks succeed.",
        "Meta Test Description": "The testcase executes test_case(), which performs two phases: (1) default/reset value verification via chk_rst_val(), and (2) write/read verification via chk_rd_wr(). In chk_rst_val(): for i=0..CNT-1, addr = addr_array[i]. If skip_rst_array[i] == 1, the entry is skipped. If read_mask_array[i] == 0x00000000, the entry is treated as not readable and skipped. Otherwise, data_rd = read_reg(addr), data = (data_rd & 0xfffffffe), and data is compared to default_value_array[i]. In chk_rd_wr(): the code iterates over six write patterns {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}. For each pattern and each register, write_reg(addr, (data_wr & write_mask_array[i])) is performed, then read-back verified against exp_val. After both phases, finish(1) if failures, else finish(0).",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Remarks": "Requires platform support for read_reg, write_reg, and finish functions. Registers marked in the skip lists are intentionally not exercised. Reset-value comparison ignores bit 0 of the read value. The optional soft reset sequence is disabled in this build.",
        "Test Steps / Procedure": "1. For each register gp0_gpio_8 through gp0_gpio_30: if not skipped and readable, read and verify reset value with bit 0 ignored.\n2. For each of six data patterns and each register: if not skipped, writable, and readable, write pattern masked by write mask, read back and verify.\n3. Declare PASS only if all checks pass; otherwise FAIL.",
        "Meta Test Steps / Procedure": "Entry: test_case()\n1) Call chk_rst_val(): Loop i=0..CNT-1, read_reg(addr), mask with 0xfffffffe, compare to default_value_array[i].\n2) Call chk_rd_wr(): 6 patterns, write masked, read back, compare to exp_val.\n3) If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0).",
        "Impacted Registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10; gp0_gpio_11; gp0_gpio_12; gp0_gpio_13; gp0_gpio_14; gp0_gpio_15; gp0_gpio_16; gp0_gpio_17; gp0_gpio_18; gp0_gpio_19; gp0_gpio_20; gp0_gpio_21; gp0_gpio_22; gp0_gpio_23; gp0_gpio_24; gp0_gpio_25; gp0_gpio_26; gp0_gpio_27; gp0_gpio_28; gp0_gpio_29; gp0_gpio_30",
        "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27; MIZAR_GPIO_GP0_GPIO_28; MIZAR_GPIO_GP0_GPIO_29; MIZAR_GPIO_GP0_GPIO_30; SOFT_RST_REG_ADDRESS",
        "Validation / Acceptance Criteria": "PASS: For every non-skipped, readable register in gp0_gpio_8..gp0_gpio_30, the reset readback (with bit 0 ignored) equals the expected default value; and for every non-skipped, writable and readable register across all six data patterns, the masked readback matches the expected combination. FAIL: Any deviation.",
        "Meta Validation / Acceptance Criteria": "Default check: (read_reg(addr_array[i]) & 0xfffffffe) == default_value_array[i]. Write/read check: exp_val computed and compared. Overall PASS if (def_fail_cnt == 0 && wr_fail_cnt == 0).",
        "Code Generation (Required / Not)": "Not",
        "Meta Headers": "#include <stdio.h>; #include <stdlib.h>; #include \"test_common.h\"; #include \"test_define.c\"; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>",
        "Meta Macros": "#define SOFT_RST_REG_ADDRESS 0x00000000; #define SOFT_RST_REG_DATA 0x00000000; #define CNT 49",
        "Meta Arrays": "addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,...,MIZAR_GPIO_GP0_GPIO_30}; default_value_array[49]; read_mask_array[49]; write_mask_array[49]; skip_array[49]; skip_rst_array[49]",
        "Register Mapping": "MIZAR_GPIO_GP0_GPIO_8 -> gp0_gpio_8 (offset 0x0, base 0xA001A000, matched); MIZAR_GPIO_GP0_GPIO_9 -> gp0_gpio_9 (offset 0x4, base 0xA001A000, matched); MIZAR_GPIO_GP0_GPIO_10 -> gp0_gpio_10 (offset 0x8, base 0xA001A000, matched)",
        "Resolved Base Address": "0xA001A000 (MIZAR_GPIO_BASE with #define GPIO0 1)",
        "Source Folder": "TestRepo/gpio/gpio_reg_wr_rd_test",
        "Source Files": "program.c; test_define.c",
        "IP Name": "GPIO",
        "Base Define": "#define GPIO0 1",
        "Generation Timestamp (IST)": "PLACEHOLDER_IST"
    },
    {
        "Index": "2",
        "SS / Module": "GPIO",
        "Feature": "NA",
        "Test Case Name": "test_gpio_level_sel_intr_en",
        "Test Description": "Validates level-based interrupt generation and clearing for GP0 GPIO pins 8 through 39. The test configures each pin for input mode with level-triggered interrupts, enables the group interrupt via gp0_intr1_intr_en1, stimulates the corresponding pin state via a memory-backed pattern, and waits for the interrupt handler to run. It then verifies that the per-pin raw status is set, the group status reflects the active pin, and that both per-pin and group status clear correctly when commanded.",
        "Meta Test Description": "The testcase initializes error tracking and conditionally enables the appropriate platform interrupt line. It programs the system-level interrupt enable for the selected GPIO instance. For i = 0..31 (GPIO8..GPIO39), it configures each per-pin register at (gp0_gpio_8 + i*4) with 0x00180000 (input mode with level interrupt), enables the corresponding group interrupt bit by writing (1<<i) to gp0_intr1_intr_en1, writes 0xFFFFFFFF to 0xA0243ffc to stimulate activity, and waits for ISR. A second pass uses 0x00100000 for opposite level selection. The ISR verifies raw status, group status, clears per-pin and group interrupts, and clears system-level status.",
        "Speed": "NA",
        "Mode": "ISR",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Remarks": "Requires platform interrupt routing and an installed interrupt handler. Group interrupt enable and status registers must be accessible. The test relies on a memory-backed stimulus location (0xA0243ffc) to drive pin-level conditions.",
        "Test Steps / Procedure": "1. Configure system to route GPIO interrupt and enable it.\n2. For each GPIO pin 8-39, program per-pin register for input mode with level-high interrupt.\n3. Enable corresponding bit in gp0_intr1_intr_en1.\n4. Apply stimulus and wait for interrupt service.\n5. In ISR, verify per-pin raw status, group status, clear per-pin and group interrupts, clear system status.\n6. Repeat with level-low selection.\n7. Declare PASS if all checks succeed; otherwise FAIL.",
        "Meta Test Steps / Procedure": "Entry: test_case()\n1) Init: test_err = 0; GIC_EnableIRQ.\n2) System-level interrupt enable via MIZAR_LSS_SYSREG_INTR_EN1.\n3) Level-high loop (i=0..31): write_reg per-pin, enable group, stimulus, wait ISR.\n4) Level-low loop (i=0..31): similar with 0x00100000.\n5) ISR: verify raw status, group status, clear, verify cleared.\n6) finish(test_err).",
        "Impacted Registers": "gp0_gpio_8; gp0_intr1_intr_en1; gp0_intr1_intr_sts1",
        "Meta Impacted Registers": "MIZAR_LSS_SYSREG_INTR_EN1; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1; 0xA0243ffc; MIZAR_GPIO_GP0_INTR1_INTR_STS1; MIZAR_LSS_SYSREG_RAW_STCR1",
        "Validation / Acceptance Criteria": "PASS: For each GPIO pin 8-39 in both level-high and level-low configurations, an interrupt is generated, group status reflects the active pin, per-pin and group status clear correctly, and system-level interrupt status is cleared. FAIL: Any missing interrupt, incorrect group status, failed clear, or uncleared system status.",
        "Meta Validation / Acceptance Criteria": "ISR: (rdata & 0x2) != 0x0 for raw status. (rdata_grp & (1<<i)) != 0 for group. After clear: per-pin reads 0x100001. After disable: gp0_intr1_intr_sts1 reads 0x0. System RAWSTCR1 bit cleared. Overall PASS when test_err == 0.",
        "Code Generation (Required / Not)": "Not",
        "Meta Headers": "#include<lss_sysreg.h>; #include<stdio.h>; #include<test_define.c>; #include<test_common.h>; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>",
        "Meta Macros": "#define CNT 49",
        "Meta Arrays": "addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,...}; default_value_array[49]; read_mask_array[49]; write_mask_array[49]; skip_array[49]",
        "Register Mapping": "MIZAR_GPIO_GP0_GPIO_8 -> gp0_gpio_8 (offset 0x0, base 0xA001A000, matched); MIZAR_GPIO_GP0_INTR1_INTR_EN1 -> gp0_intr1_intr_en1 (matched); MIZAR_GPIO_GP0_INTR1_INTR_STS1 -> gp0_intr1_intr_sts1 (matched); MIZAR_LSS_SYSREG_INTR_EN1 -> system register (external); MIZAR_LSS_SYSREG_RAW_STCR1 -> system register (external)",
        "Resolved Base Address": "0xA001A000 (MIZAR_GPIO_BASE with #define GPIO0 1)",
        "Source Folder": "TestRepo/gpio/test_gpio_level_sel_intr_en",
        "Source Files": "program.c; test_define.c",
        "IP Name": "GPIO",
        "Base Define": "#define GPIO0 1",
        "Generation Timestamp (IST)": "PLACEHOLDER_IST"
    }
]


def build_workbook(ist_str):
    """Build the .xlsx workbook in memory."""
    wb = Workbook()

    # ── Styles ──
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── TestPlan Sheet ──
    ws = wb.active
    ws.title = "TestPlan"

    # Header row
    for col_idx, col_name in enumerate(TESTPLAN_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Data rows
    for row_idx, row_obj in enumerate(ROW_DATA, start=2):
        for col_idx, col_name in enumerate(TESTPLAN_COLS, start=1):
            val = row_obj.get(col_name, "")
            if val == "PLACEHOLDER_IST":
                val = ist_str
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    for col_idx in range(1, len(TESTPLAN_COLS) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(TESTPLAN_COLS[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 14), 60)

    ws.freeze_panes = "A2"

    # ── MetaData Sheet (veryHidden) ──
    meta = wb.create_sheet("MetaData")
    meta.sheet_state = "veryHidden"

    for col_idx, col_name in enumerate(METADATA_COLS, start=1):
        cell = meta.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    for row_idx, row_obj in enumerate(ROW_DATA, start=2):
        for col_idx, col_name in enumerate(METADATA_COLS, start=1):
            val = row_obj.get(col_name, "")
            cell = meta.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    for col_idx in range(1, len(METADATA_COLS) + 1):
        meta.column_dimensions[get_column_letter(col_idx)].width = 30

    return wb


def main():
    # IST timestamp
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    ist_str = now_ist.strftime("%Y-%m-%d %H:%M:%S IST")
    date_part = now_ist.strftime("%Y%m%d")
    time_part = now_ist.strftime("%H%M%S")
    filename = f"{IP_NAME}_TestPlan_{date_part}_{time_part}.xlsx"

    wb = build_workbook(ist_str)

    # Save to BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx_bytes = buf.read()

    # Try pushing to GitHub
    token = os.environ.get("GITHUB_TOKEN", "")
    if token:
        try:
            from github import Github
            g = Github(token)
            repo = g.get_repo(f"{OWNER}/{REPO}")
            path = f"{OUTPUT_DIRECTORY}/{filename}"
            content_b64 = base64.b64encode(xlsx_bytes).decode("ascii")
            msg = f"Add {IP_NAME} TestPlan (IST {ist_str})"
            repo.create_file(path, msg, xlsx_bytes, branch=BRANCH)
            print(f"Pushed to GitHub: {path}")
        except Exception as e:
            print(f"GitHub push failed: {e}")
            # Fallback: save locally
            os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)
            out_path = os.path.join(OUTPUT_DIRECTORY, filename)
            with open(out_path, "wb") as f:
                f.write(xlsx_bytes)
            print(f"Saved locally: {out_path}")
    else:
        os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)
        out_path = os.path.join(OUTPUT_DIRECTORY, filename)
        with open(out_path, "wb") as f:
            f.write(xlsx_bytes)
        print(f"Saved locally: {out_path}")


if __name__ == "__main__":
    main()
