#!/usr/bin/env python3
"""
GPIO TestPlan Excel Generator & GitHub Pusher
=============================================
This script generates a REAL .xlsx binary Excel file with:
  - TestPlan sheet (visible) - 21 columns, 1 header + 2 data rows
  - MetaData sheet (veryHidden) - full JSON in cell A1
Then pushes the .xlsx directly to GitHub.

Usage:
  pip install openpyxl PyGithub
  export GITHUB_TOKEN="your_pat_here"
  python generate_and_push_xlsx.py
"""

import os
import sys
import json
import base64
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from github import Github
except ImportError:
    print("ERROR: PyGithub not installed. Run: pip install PyGithub")
    sys.exit(1)

from io import BytesIO

# ============================================================
# Configuration
# ============================================================
REPO_OWNER = "titusbspgit"
REPO_NAME = "PSVValidation"
BRANCH = "main"
OUTPUT_DIR = "Test_Output/GPIO/TestPlan/"
IP_NAME = "GPIO"

# IST timezone (GMT+05:30)
IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
TIMESTAMP = now_ist.strftime("%Y%m%d_%H%M%S")
FILENAME = f"{IP_NAME}_TestPlan_{TIMESTAMP}.xlsx"
FILE_PATH = f"{OUTPUT_DIR}{FILENAME}"

# ============================================================
# 21 Column Headers
# ============================================================
HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

# ============================================================
# Row 1: gpio_reg_wr_rd_test
# ============================================================
ROW1 = [
    1,
    "GPIO",
    "Register Write Read Verification",
    "gpio_reg_wr_rd_test",
    "Verify register write and read operations for GPIO per-pin registers gp0_gpio_8 through gp0_gpio_30. The test reads default reset values, writes multiple patterns (0x00, 0xFF, 0x55, 0xAA, 0xA5, 0x5A), reads back and verifies using field-specific read/write masks considering RO, WO, RW, and RW2 field types.",
    "NA",
    "Polling",
    "0x0",
    "0x58",
    "",
    "1. Read default reset value from each GPIO register (gp0_gpio_8 to gp0_gpio_30) and verify against expected reset value 0x00100000 (io_ctrl=1 at bit 20).\n2. Write pattern 0x00000000 to each register.\n3. Read back and verify: RW/RW2 fields updated, RO fields unchanged, WO fields read as 0.\n4. Repeat steps 2-3 for patterns 0x000E0000, 0x000A0000, 0x00140000, 0x00060000, 0x00180000 (masked for writable fields).\n5. Trigger soft reset via SOFT_RST_REG_ADDRESS.\n6. Re-read all registers and verify reset values restored.",
    "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10; gp0_gpio_11; gp0_gpio_12; gp0_gpio_13; gp0_gpio_14; gp0_gpio_15; gp0_gpio_16; gp0_gpio_17; gp0_gpio_18; gp0_gpio_19; gp0_gpio_20; gp0_gpio_21; gp0_gpio_22; gp0_gpio_23; gp0_gpio_24; gp0_gpio_25; gp0_gpio_26; gp0_gpio_27; gp0_gpio_28; gp0_gpio_29; gp0_gpio_30",
    "PASS: All registers return expected reset value 0x00100000 on initial read. All write patterns are correctly reflected in RW/RW2 fields on read-back. RO fields (data_in, intr_raw_sts) remain unchanged after write. WO field (intr_clr) reads as 0. After soft reset, all registers return to reset value. FAIL: Any mismatch between expected and actual values.",
    "Not Required",
    "Verify register write and read operations for GPIO per-pin registers gp0_gpio_8 through gp0_gpio_30. The test reads default reset values, writes multiple patterns (0x00, 0xFF, 0x55, 0xAA, 0xA5, 0x5A), reads back and verifies using field-specific read/write masks considering RO, WO, RW, and RW2 field types.",
    "1. Read default reset value from each GPIO register (gp0_gpio_8 to gp0_gpio_30) and verify against expected reset value 0x00100000 (io_ctrl=1 at bit 20).\n2. Write pattern 0x00000000 to each register.\n3. Read back and verify: RW/RW2 fields updated, RO fields unchanged, WO fields read as 0.\n4. Repeat steps 2-3 for patterns 0x000E0000, 0x000A0000, 0x00140000, 0x00060000, 0x00180000 (masked for writable fields).\n5. Trigger soft reset via SOFT_RST_REG_ADDRESS.\n6. Re-read all registers and verify reset values restored.",
    "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27; MIZAR_GPIO_GP0_GPIO_28; MIZAR_GPIO_GP0_GPIO_29; MIZAR_GPIO_GP0_GPIO_30; SOFT_RST_REG_ADDRESS",
    "PASS: All registers return expected reset value 0x00100000 on initial read. All write patterns are correctly reflected in RW/RW2 fields on read-back. RO fields (data_in, intr_raw_sts) remain unchanged after write. WO field (intr_clr) reads as 0. After soft reset, all registers return to reset value. FAIL: Any mismatch between expected and actual values.",
    "",
    "",
    "",
]

# ============================================================
# Row 2: test_gpio_level_sel_intr_en
# ============================================================
ROW2 = [
    2,
    "GPIO",
    "Level Select Interrupt Enable",
    "test_gpio_level_sel_intr_en",
    "Verify GPIO level-select interrupt functionality for all 32 GPIO pins (pin 8 through pin 39). The test configures each per-pin GPIO register in input mode with level-triggered interrupt enabled, enables the group interrupt via the group interrupt enable register, enables LSS sysreg interrupt routing via the sysreg interrupt enable register, drives an external stimulus, and verifies interrupt assertion through the per-pin interrupt raw status bit, the group interrupt status register, and the LSS sysreg raw status register. The test covers both active HIGH level (level_sel=1) and active LOW level (level_sel=0) interrupt generation. After each interrupt, the test clears the per-pin interrupt, verifies the clear succeeded, disables the group interrupt, verifies the group status is cleared, clears the sysreg raw status, and verifies the sysreg clear succeeded. The test uses GIC IRQ 87 for GPIO0 interrupt routing.",
    "NA",
    "ISR",
    "0x0",
    "0x88",
    "Requires GIC interrupt controller to be initialized and IRQ 87 routed for GPIO0. External stimulus hardware must be connected at the designated SRAM location to drive GPIO pin levels. The test depends on conditional compilation with GPIO0 defined. LSS sysreg interrupt enable must be configured before GPIO interrupts can propagate to the GIC.",
    "1. Enable GIC IRQ for GPIO0 interrupt routing.\n2. Write to the LSS sysreg interrupt enable register (intr_en1) to enable GPIO0 interrupt at bit 1.\n3. For each GPIO pin (pin 8 to pin 39), configure the per-pin register (gp0_gpio_8 through gp0_gpio_39) in input mode with level-select interrupt set to active HIGH (io_ctrl=1, level_sel=1).\n4. Enable the group interrupt for the target pin by writing the corresponding bit to the group interrupt enable register (gp0_intr1_intr_en1).\n5. Drive external stimulus HIGH by writing to the external stimulus address to trigger the level interrupt.\n6. Wait for the interrupt service routine to execute.\n7. In the ISR, read the per-pin register and verify the interrupt raw status bit (bit 1) is asserted.\n8. Read the group interrupt status register (gp0_intr1_intr_sts1) and verify the corresponding pin bit is set.\n9. Clear the per-pin interrupt by writing to the interrupt clear field (bit 16) in the per-pin register.\n10. Read back the per-pin register and verify the interrupt has been cleared successfully (expected value 0x100001).\n11. Disable the group interrupt by writing 0x00000000 to the group interrupt enable register.\n12. Read the group interrupt status register and verify it reads 0x0.\n13. Write to the LSS sysreg raw status clear register (raw_stcr1) to clear the GPIO0 interrupt status.\n14. Read back the sysreg raw status register and verify the GPIO0 interrupt bit is cleared.\n15. Clear the GIC IRQ.\n16. Repeat steps 3-15 for all 32 GPIO pins with active HIGH level.\n17. Repeat the entire sequence (steps 3-16) with level-select set to active LOW (io_ctrl=1, level_sel=0) and drive external stimulus LOW for each pin.\n18. Verify the test completes with zero errors.",
    "intr_en1; gp0_gpio_8; gp0_intr1_intr_en1; gp0_intr1_intr_sts1; raw_stcr1",
    "PASS: For each of the 32 GPIO pins in active HIGH mode (level_sel=1): the per-pin register interrupt raw status bit (bit 1) is asserted when the external stimulus is HIGH. The group interrupt status register (gp0_intr1_intr_sts1) shows the corresponding pin bit set. After clearing the interrupt via the per-pin register interrupt clear field, the per-pin register reads 0x100001. After disabling the group interrupt enable, the group interrupt status register reads 0x0. After clearing the sysreg raw status register (raw_stcr1), the GPIO0 interrupt bit is cleared. For each of the 32 GPIO pins in active LOW mode (level_sel=0): the same validation sequence passes when the external stimulus is driven LOW for the target pin. The test completes with zero accumulated errors. FAIL: Any per-pin interrupt raw status not asserted, any group interrupt status mismatch, any interrupt clear failure (per-pin register not equal to 0x100001), any group interrupt status not clearing to 0x0, any sysreg raw status not clearing, or any interrupt not occurring.",
    "Not Required",
    "Verify GPIO level-select interrupt functionality for all 32 GPIO pins (GPIO_8 to GPIO_39). The test enables GIC IRQ 87 (GPIO0) via GIC_EnableIRQ(87). It writes LSS_SYSREG_INTR_EN1_GPIO0_INTR to MIZAR_LSS_SYSREG_INTR_EN1 to enable sysreg interrupt routing. In the first loop (active HIGH, i=0..31): writes 0x00180000 to MIZAR_GPIO_GP0_GPIO_8+(i*4) setting io_ctrl=1 (bit 20) and level_sel=1 (bit 19), writes (1<<i) to MIZAR_GPIO_GP0_INTR1_INTR_EN1 (offset 0x84) to enable group interrupt for pin i, writes 0xffffffff to 0xA0243ffc to drive external stimulus HIGH, sets int_pend=1 and polls while(int_pend==1) with wait_on(10). In Default_IRQHandler: sets int_pend=0, writes 0xffffffff to 0xA0243ffc, reads MIZAR_GPIO_GP0_GPIO_8+(i*4) into rdata, checks (rdata & 0x2) != 0x0 for intr_raw_sts assertion, reads MIZAR_GPIO_GP0_INTR1_INTR_STS1 (offset 0x88) into rdata_grp, checks (rdata_grp & (1<<i)) != 0 for group interrupt, writes 0x00110000 to MIZAR_GPIO_GP0_GPIO_8+(i*4) to clear interrupt (intr_clr bit 16=1, io_ctrl bit 20=1), waits wait_on(20), reads back MIZAR_GPIO_GP0_GPIO_8+(i*4) and checks rdata==0x100001, writes 0x00000000 to MIZAR_GPIO_GP0_INTR1_INTR_EN1 to disable group interrupt, reads MIZAR_GPIO_GP0_INTR1_INTR_STS1 and checks rdata_grp==0x0, writes LSS_SYSREG_RAW_STCR1_GPIO0_INTR to MIZAR_LSS_SYSREG_RAW_STCR1, reads back MIZAR_LSS_SYSREG_RAW_STCR1 and checks (rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR)==0, then calls GIC_ClearIRQ(87). In the second loop (active LOW, i=0..31): writes 0x00100000 to MIZAR_GPIO_GP0_GPIO_8+(i*4) setting io_ctrl=1 (bit 20) and level_sel=0 (bit 19), writes (1<<i) to MIZAR_GPIO_GP0_INTR1_INTR_EN1, writes ~(wr_val) to 0xA0243ffc to drive stimulus LOW for target pin, sets int_pend=1 and polls. ISR performs same validation sequence. finish(test_err).",
    '1. GIC_EnableIRQ(87) to enable GPIO0 GIC interrupt.\n2. write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR) to enable sysreg GPIO0 interrupt routing.\n3. Active HIGH loop (i=0 to 31): wr_val = 1<<i; write_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4), 0x00180000) sets io_ctrl=1 (bit 20) and level_sel=1 (bit 19); wait_on(50); write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val) enables group interrupt for pin i; wait_on(10); write_reg(0xA0243ffc, 0xffffffff) drives external stimulus HIGH; int_pend=1; while(int_pend==1) { printf("Waiting for interrupt"); wait_on(10); }.\n4. Default_IRQHandler entry: int_pend=0; write_reg(0xA0243ffc, 0xffffffff); rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4)).\n5. Check (rdata & 0x2) != 0x0 for intr_raw_sts assertion. If false: printf("Interrupt Not occured"); test_err++.\n6. rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); check (rdata_grp & (1<<i)) != 0. If false: printf("ERROR: Group Interrupt not occured"); test_err++.\n7. write_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4), 0x00110000) to clear interrupt (intr_clr bit 16=1, io_ctrl bit 20=1); wait_on(20).\n8. rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4)); check rdata==0x100001. If not: printf("ERROR : Interrupt clear failed"); test_err++.\n9. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000) to disable group interrupt.\n10. rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); check rdata_grp==0x0. If not: printf("ERROR : Group Interrupt clear failed"); test_err++.\n11. write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); rdata=read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); check (rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR)==0. If not: printf("sysreg status not cleared"); test_err++.\n12. GIC_ClearIRQ(87).\n13. Active LOW loop (i=0 to 31): wr_val=1<<i; write_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4), 0x00100000) sets io_ctrl=1 (bit 20) and level_sel=0 (bit 19); wait_on(50); write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val); wait_on(10); write_reg(0xA0243ffc, ~(wr_val)) drives stimulus LOW for target pin; int_pend=1; while(int_pend==1) { printf("Waiting for interrupt"); wait_on(10); }.\n14. ISR performs same validation sequence as steps 4-12.\n15. finish(test_err).',
    "MIZAR_LSS_SYSREG_INTR_EN1; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1; 0xA0243ffc; MIZAR_GPIO_GP0_INTR1_INTR_STS1; MIZAR_LSS_SYSREG_RAW_STCR1",
    "PASS conditions: (1) In Default_IRQHandler, (rdata & 0x2) != 0x0 after read_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4)). (2) (rdata_grp & (1<<i)) != 0 after read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1). (3) After write_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4), 0x00110000) and wait_on(20), read_reg(MIZAR_GPIO_GP0_GPIO_8+(i*4)) == 0x100001. (4) After write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000), read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) == 0x0. (5) After write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR), (read_reg(MIZAR_LSS_SYSREG_RAW_STCR1) & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0. FAIL if test_err > 0.",
    "#include<lss_sysreg.h>; #include<stdio.h>; #include<test_define.c>; #include<test_common.h>; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>",
    "#define CNT 49",
    "const unsigned long int addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,...}; const unsigned int default_value_array[49]={...}; const unsigned int read_mask_array[49]={...}; const unsigned int write_mask_array[49]={...}; const int skip_array[49]={0,0,...,0};",
]

# ============================================================
# Full JSON for MetaData sheet
# ============================================================
METADATA_JSON = json.dumps([
    {
        "Index": 1,
        "SS / Module": "GPIO",
        "Feature": "Register Write Read Verification",
        "Test Case Name": "gpio_reg_wr_rd_test",
        "Test Description": ROW1[4],
        "Speed": "NA",
        "Mode": "Polling",
        "Memory Start Offset": "0x0",
        "Memory End Offset": "0x58",
        "Remarks": "",
        "Test Steps / Procedure": ROW1[10],
        "Impacted Registers": ROW1[11],
        "Validation / Acceptance Criteria": ROW1[12],
        "Code Generation (Required / Not)": "Not Required",
        "Meta Test Description": ROW1[14],
        "Meta Test Steps / Procedure": ROW1[15],
        "Meta Impacted Registers": ROW1[16],
        "Meta Validation / Acceptance Criteria": ROW1[17],
        "Meta Headers": "",
        "Meta Macros": "",
        "Meta Arrays": "",
    },
    {
        "Index": 2,
        "SS / Module": "GPIO",
        "Feature": "Level Select Interrupt Enable",
        "Test Case Name": "test_gpio_level_sel_intr_en",
        "Test Description": ROW2[4],
        "Speed": "NA",
        "Mode": "ISR",
        "Memory Start Offset": "0x0",
        "Memory End Offset": "0x88",
        "Remarks": ROW2[9],
        "Test Steps / Procedure": ROW2[10],
        "Impacted Registers": ROW2[11],
        "Validation / Acceptance Criteria": ROW2[12],
        "Code Generation (Required / Not)": "Not Required",
        "Meta Test Description": ROW2[14],
        "Meta Test Steps / Procedure": ROW2[15],
        "Meta Impacted Registers": ROW2[16],
        "Meta Validation / Acceptance Criteria": ROW2[17],
        "Meta Headers": ROW2[18],
        "Meta Macros": ROW2[19],
        "Meta Arrays": ROW2[20],
    },
], indent=2)

# ============================================================
# Column widths
# ============================================================
COL_WIDTHS = [8, 14, 30, 32, 60, 8, 10, 20, 18, 40, 70, 60, 65, 28, 60, 70, 60, 65, 50, 30, 50]


def build_workbook():
    """Build the xlsx workbook in memory and return bytes."""
    wb = Workbook()

    # ---- TestPlan sheet ----
    ws = wb.active
    ws.title = "TestPlan"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate([ROW1, ROW2], 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # ---- MetaData sheet (veryHidden) ----
    ws_meta = wb.create_sheet(title="MetaData")
    ws_meta.cell(row=1, column=1, value=METADATA_JSON)
    ws_meta.sheet_state = "veryHidden"

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def push_to_github(xlsx_bytes):
    """Push the xlsx binary to GitHub."""
    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        print("ERROR: GITHUB_TOKEN environment variable not set.")
        sys.exit(1)

    g = Github(token)
    repo = g.get_repo(f"{REPO_OWNER}/{REPO_NAME}")

    content_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    commit_msg = f"Add {FILENAME} - GPIO TestPlan Excel (auto-generated)"

    # Check if file exists
    try:
        existing = repo.get_contents(FILE_PATH, ref=BRANCH)
        result = repo.update_file(
            path=FILE_PATH,
            message=commit_msg,
            content=xlsx_bytes,
            sha=existing.sha,
            branch=BRANCH,
        )
    except Exception:
        result = repo.create_file(
            path=FILE_PATH,
            message=commit_msg,
            content=xlsx_bytes,
            branch=BRANCH,
        )

    commit_sha = result["commit"].sha
    github_url = f"https://github.com/{REPO_OWNER}/{REPO_NAME}/blob/{BRANCH}/{FILE_PATH}"

    return commit_sha, github_url


def main():
    print(f"Generating {FILENAME}...")
    xlsx_bytes = build_workbook()
    print(f"  Workbook size: {len(xlsx_bytes):,} bytes")

    # Save locally
    with open(FILENAME, "wb") as f:
        f.write(xlsx_bytes)
    print(f"  Saved locally: {FILENAME}")

    # Push to GitHub
    print(f"Pushing to GitHub: {FILE_PATH}")
    commit_sha, github_url = push_to_github(xlsx_bytes)

    print("\n" + "=" * 60)
    print("SUCCESS")
    print("=" * 60)
    print(json.dumps({
        "status": "SUCCESS",
        "execution_mode": "Direct Excel Generation",
        "output_file_path": FILE_PATH,
        "github_url": github_url,
        "commit_sha": commit_sha,
        "filename": FILENAME,
        "file_size_bytes": len(xlsx_bytes),
        "timestamp_ist": now_ist.isoformat(),
    }, indent=2))


if __name__ == "__main__":
    main()
