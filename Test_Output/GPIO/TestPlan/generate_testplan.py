#!/usr/bin/env python3
"""
GPIO TestPlan XLSX Generator
Generates a real Excel workbook with TestPlan and MetaData sheets.
Run: python3 generate_testplan.py
"""
import json
import os
import sys
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

# ══════════════════════════════════════════════════════════════
# FULL JSON DATA - 2 testcases
# ══════════════════════════════════════════════════════════════
json_data = [
  {
    "index": 1,
    "ss_module": "GPIO",
    "test_case_name": "gpio_reg_wr_rd_test",
    "feature": "Register Read/Write Validation",
    "meta_test_description": "This testcase validates the default (reset) values and read/write functionality of GPIO GP0 registers (MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10). The test is structured in two phases:\n\nPhase 1 - chk_rst_val(): Iterates over addr_array[] containing register addresses. For each register, it checks skip_rst_array[i] to determine if the register should be skipped. If read_mask_array[i] is 0x00000000, the register is not readable and is skipped. Otherwise, it performs read_reg(addr), masks the read data with 0xfffffffe (clearing bit 0), and compares against default_value_array[i]. Mismatches increment def_fail_cnt.\n\nPhase 2 - chk_rd_wr(): Uses six test patterns: 0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000. For each pattern, it iterates over addr_array[]. Registers are skipped if skip_array[i]==1 or write_mask_array[i]==0x00000000. Writes are performed as write_reg(addr, data_wr & write_mask_array[i]). Read-back is performed as read_reg(addr) & read_mask_array[i]. Expected value is computed as: (data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]) where wr_n = write_mask_array[i] ^ 0xffffffff. Mismatches increment wr_fail_cnt.\n\nFinal verdict: finish(1) if any fail count > 0, else finish(0).\n\nsoft_reset_chk() is disabled (#ifdef 0). SOFT_RST_REG_ADDRESS is excluded per instructions.\n\nHeaders included: gpio/gpio_def.h, gpio/gpio_offset.h, test_common.h.\nCNT = 49. Only 3 register addresses are populated in the visible addr_array entries.",
    "test_description": "This test validates the GPIO GP0 register block by performing two checks: (1) Reset value verification - reads each GPIO register and verifies the data matches the expected default reset values after masking read-only bits. (2) Write/Read verification - writes six different test patterns (all-ones, alternating bits, mixed patterns) to each writable GPIO register, reads back the values, and verifies correctness by accounting for read masks, write masks, and default values of non-writable bit fields. The test covers registers gp0_gpio_8, gp0_gpio_9, and gp0_gpio_10. The test reports PASS if all default value checks and write/read checks succeed, otherwise FAIL.",
    "meta_test_steps": "1. Include headers: gpio/gpio_def.h, gpio/gpio_offset.h, test_common.h, test_define.c.\n2. Define arrays: addr_array[49] = {MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, ...}, default_value_array[49], read_mask_array[49], write_mask_array[49], skip_array[49], skip_rst_array[49].\n3. Initialize fail counters: def_fail_cnt = 0, wr_fail_cnt = 0.\n4. Call chk_rst_val():\n   4a. Loop i = 0 to CNT-1.\n   4b. addr = addr_array[i].\n   4c. If skip_rst_array[i] == 1, skip (continue).\n   4d. If read_mask_array[i] == 0x00000000, skip (not readable).\n   4e. data_rd = read_reg(addr).\n   4f. data = data_rd & 0xfffffffe (mask out bit 0).\n   4g. Compare data with default_value_array[i].\n   4h. If mismatch, increment def_fail_cnt and print failure.\n5. Call chk_rd_wr():\n   5a. Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}.\n   5b. Loop j = 0 to 5 (for each test pattern).\n   5c. data_wr = chk_val[j].\n   5d. Write phase: Loop i = 0 to CNT-1.\n       - addr = addr_array[i].\n       - If skip_array[i] == 1, skip.\n       - If write_mask_array[i] == 0x00000000, skip (not writable).\n       - write_reg(addr, data_wr & write_mask_array[i]).\n   5e. Read phase: Loop i = 0 to CNT-1.\n       - addr = addr_array[i].\n       - If skip_array[i] == 1, skip.\n       - If write_mask_array[i] == 0x00000000, skip.\n       - If read_mask_array[i] == 0x00000000, skip.\n       - data_rd = read_reg(addr) & read_mask_array[i].\n       - wr_n = write_mask_array[i] ^ 0xffffffff.\n       - exp_val = (data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]).\n       - Compare data_rd with exp_val.\n       - If mismatch, increment wr_fail_cnt and print failure.\n6. Check final result: if def_fail_cnt > 0 || wr_fail_cnt > 0, call finish(1) (FAIL); else call finish(0) (PASS).\n7. soft_reset_chk() is disabled (#ifdef 0 block) and not executed.",
    "test_steps": "1. Initialize the test environment and load the GPIO register configuration arrays including register addresses, default values, read masks, write masks, and skip control arrays.\n2. Perform reset value verification for each GPIO register (gp0_gpio_8, gp0_gpio_9, gp0_gpio_10): Read each register, apply the appropriate read mask, and compare the result against the expected default reset value. Skip registers marked in the skip-reset array or those that are not readable.\n3. Perform write/read verification using six test data patterns (all-ones, alternating-bit patterns, and mixed patterns) for each GPIO register:\n   a. Write each test pattern to the register, masked by the write mask to protect read-only and write-only fields.\n   b. Read back the register value, masked by the read mask.\n   c. Compute the expected value considering writable bits from the written pattern and non-writable bits retaining their default values.\n   d. Compare the read-back value against the expected value.\n4. Skip any registers flagged in the skip array or those with zero write mask (not writable) or zero read mask (not readable).\n5. Evaluate the final test result: PASS if all default value checks and all write/read pattern checks succeed with zero mismatches; FAIL otherwise.",
    "meta_impacted_registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10",
    "impacted_registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
    "validation_acceptance_criteria": "1. All GPIO registers must return their expected default reset values when read after reset.\n2. For each of the six test patterns, the write/read-back value must match the expected value computed using read masks, write masks, and default values for non-writable fields.\n3. The test must complete with zero default-value mismatches and zero write/read mismatches to be considered PASS.\n4. Registers marked as skip or having zero read/write masks must be correctly excluded from the respective checks.",
    "speed": "NA",
    "mode": "NA",
    "remarks": "The test uses six distinct data patterns to exercise all writable bit positions across the GPIO registers. Bit 0 (data_in field) is masked out during default value comparison since it reflects the live pin state. The skip arrays allow selective exclusion of specific registers from reset and write/read checks. The soft reset check function is disabled in the source code.",
    "register_details": [
      {"macro": "MIZAR_GPIO_GP0_GPIO_8", "register_name": "gp0_gpio_8", "base_address": "0xA001A000", "offset": "0x0", "resolved_address": "0xA001A000", "operation": "read_modify_write", "resolution_status": "resolved", "mapping_status": "matched", "register_width": 32, "fields": [{"field_name": "data_in", "bit_position": "0", "access_type": "RO", "reset_value": "0", "description": "Determines the level gpio pin"}, {"field_name": "intr_raw_sts", "bit_position": "1", "access_type": "RO", "reset_value": "0", "description": "RAWST bit: Interrupt raw status"}, {"field_name": "intr_clr", "bit_position": "16", "access_type": "WO", "reset_value": "0", "description": "Writing 1 clears the interrupt raw status"}, {"field_name": "pedge_intr_en", "bit_position": "17", "access_type": "RW", "reset_value": "0", "description": "Pos-edge interrupt enable"}, {"field_name": "nedge_intr_en", "bit_position": "18", "access_type": "RW", "reset_value": "0", "description": "Neg-edge interrupt enable"}, {"field_name": "level_sel", "bit_position": "19", "access_type": "RW", "reset_value": "0", "description": "GPIO Pin signal level select"}, {"field_name": "io_ctrl", "bit_position": "20", "access_type": "RW2", "reset_value": "1", "description": "GPIO mode select"}, {"field_name": "dout", "bit_position": "21", "access_type": "RW2", "reset_value": "0", "description": "Output data"}]},
      {"macro": "MIZAR_GPIO_GP0_GPIO_9", "register_name": "gp0_gpio_9", "base_address": "0xA001A000", "offset": "0x4", "resolved_address": "0xA001A004", "operation": "read_modify_write", "resolution_status": "resolved", "mapping_status": "matched", "register_width": 32, "fields": [{"field_name": "data_in", "bit_position": "0", "access_type": "RO", "reset_value": "0", "description": "Determines the level gpio pin"}, {"field_name": "intr_raw_sts", "bit_position": "1", "access_type": "RO", "reset_value": "0", "description": "RAWST bit"}, {"field_name": "intr_clr", "bit_position": "16", "access_type": "WO", "reset_value": "0", "description": "Interrupt clear"}, {"field_name": "pedge_intr_en", "bit_position": "17", "access_type": "RW", "reset_value": "0", "description": "Pos-edge interrupt enable"}, {"field_name": "nedge_intr_en", "bit_position": "18", "access_type": "RW", "reset_value": "0", "description": "Neg-edge interrupt enable"}, {"field_name": "level_sel", "bit_position": "19", "access_type": "RW", "reset_value": "0", "description": "Level select"}, {"field_name": "io_ctrl", "bit_position": "20", "access_type": "RW2", "reset_value": "1", "description": "IO control"}, {"field_name": "dout", "bit_position": "21", "access_type": "RW2", "reset_value": "0", "description": "Output data"}]},
      {"macro": "MIZAR_GPIO_GP0_GPIO_10", "register_name": "gp0_gpio_10", "base_address": "0xA001A000", "offset": "0x8", "resolved_address": "0xA001A008", "operation": "read_modify_write", "resolution_status": "resolved", "mapping_status": "matched", "register_width": 32, "fields": [{"field_name": "data_in", "bit_position": "0", "access_type": "RO", "reset_value": "0", "description": "Determines the level gpio pin"}, {"field_name": "intr_raw_sts", "bit_position": "1", "access_type": "RO", "reset_value": "0", "description": "RAWST bit"}, {"field_name": "intr_clr", "bit_position": "16", "access_type": "WO", "reset_value": "0", "description": "Interrupt clear"}, {"field_name": "pedge_intr_en", "bit_position": "17", "access_type": "RW", "reset_value": "0", "description": "Pos-edge interrupt enable"}, {"field_name": "nedge_intr_en", "bit_position": "18", "access_type": "RW", "reset_value": "0", "description": "Neg-edge interrupt enable"}, {"field_name": "level_sel", "bit_position": "19", "access_type": "RW", "reset_value": "0", "description": "Level select"}, {"field_name": "io_ctrl", "bit_position": "20", "access_type": "RW2", "reset_value": "1", "description": "IO control"}, {"field_name": "dout", "bit_position": "21", "access_type": "RW2", "reset_value": "0", "description": "Output data"}]}
    ]
  },
  {
    "index": 2,
    "ss_module": "GPIO",
    "test_case_name": "test_gpio_level_sel_intr_en",
    "feature": "GPIO Level Select Interrupt Enable",
    "meta_test_description": "This testcase validates the GPIO level-select interrupt functionality for GPIO pins 8 through 39 (32 GPIOs). The test is structured in two phases using MIZAR_GPIO_GP0_GPIO_8 as the base address with computed offsets (i * 4) for each GPIO pin.\n\nPhase 1 - Active High Level Interrupt: For each GPIO (i = 0 to 31):\n  - Enables GIC IRQ (87 for GPIO0, 88 for GPIO1).\n  - Writes MIZAR_LSS_SYSREG_INTR_EN1 with LSS_SYSREG_INTR_EN1_GPIO0_INTR (or GPIO1_INTR) to enable sysreg interrupt.\n  - Writes 0x00180000 to MIZAR_GPIO_GP0_GPIO_8 + (i * 4), setting io_ctrl (bit 20) = 1 (input mode), level_sel (bit 19) = 1 (active high level interrupt), nedge_intr_en (bit 18) = 1.\n  - Waits 50 cycles via wait_on(50).\n  - Writes wr_val = (1 << i) to MIZAR_GPIO_GP0_INTR1_INTR_EN1 to enable the group interrupt for the specific GPIO.\n  - Waits 10 cycles, writes 0xffffffff to 0xA0243ffc (SRAM trigger location).\n  - Sets int_pend = 1 and polls in a while loop with wait_on(10) until int_pend becomes 0 (set by IRQ handler).\n\nPhase 2 - Active Low Level Interrupt: For each GPIO (i = 0 to 31):\n  - Writes 0x00100000 to MIZAR_GPIO_GP0_GPIO_8 + (i * 4), setting io_ctrl (bit 20) = 1 (input mode), level_sel (bit 19) = 0 (active low level interrupt).\n  - Waits 50 cycles.\n  - Writes wr_val = (1 << i) to MIZAR_GPIO_GP0_INTR1_INTR_EN1.\n  - Waits 10 cycles, writes ~(wr_val) to 0xA0243ffc.\n  - Sets int_pend = 1 and polls until interrupt is serviced.\n\nIRQ Handler (Default_IRQHandler):\n  - Sets int_pend = 0.\n  - Writes 0xffffffff to 0xA0243ffc.\n  - Reads MIZAR_GPIO_GP0_GPIO_8 + (i * 4) into rdata.\n  - Checks (rdata & 0x2) != 0x0 to verify intr_raw_sts (bit 1) is set.\n  - If raw status set: reads MIZAR_GPIO_GP0_INTR1_INTR_STS1 and verifies (rdata_grp & (1 << i)) != 0 for group interrupt status.\n  - Writes 0x00110000 to MIZAR_GPIO_GP0_GPIO_8 + (i * 4) to clear interrupt (intr_clr bit 16 = 1, io_ctrl bit 20 = 1).\n  - Waits 20 cycles, reads back MIZAR_GPIO_GP0_GPIO_8 + (i * 4) and verifies rdata == 0x100001.\n  - Writes 0x00000000 to MIZAR_GPIO_GP0_INTR1_INTR_EN1 to disable group interrupt.\n  - Reads MIZAR_GPIO_GP0_INTR1_INTR_STS1 and verifies rdata_grp == 0x0 (group interrupt cleared).\n  - Writes LSS_SYSREG_RAW_STCR1_GPIO0_INTR (or GPIO1_INTR) to MIZAR_LSS_SYSREG_RAW_STCR1 to clear sysreg status.\n  - Reads back MIZAR_LSS_SYSREG_RAW_STCR1 and verifies the GPIO interrupt bit is cleared.\n  - Clears GIC IRQ.\n  - If any check fails, increments test_err.\n\nFinal verdict: finish(test_err). PASS if test_err == 0, FAIL otherwise.\n\nHeaders included: lss_sysreg.h, stdio.h, test_define.c, test_common.h, gpio/gpio_def.h, gpio/gpio_offset.h.\nCNT = 49. addr_array defined but not iterated in program.c; computed addressing used instead.",
    "test_description": "This test validates the GPIO level-select interrupt enable functionality for 32 GPIO pins (pins 8 through 39). The test operates in two phases: Phase 1 tests active-high level interrupts by configuring each GPIO pin in input mode with level_sel set to high, enabling the corresponding group interrupt, triggering the interrupt via an SRAM write, and verifying the interrupt is received and processed correctly. Phase 2 tests active-low level interrupts by configuring each GPIO pin in input mode with level_sel set to low, enabling the group interrupt, triggering with the inverted pattern, and verifying interrupt reception. In the interrupt handler, the test verifies the raw interrupt status in gp0_gpio_8 (bit 1), confirms the group interrupt status, clears the interrupt by writing to the interrupt clear bit, verifies the clear was successful, disables the group interrupt, confirms the group status is cleared, and clears the system-level interrupt status in raw_stcr1. The test also enables and clears the system-level interrupt enable via intr_en1. The test passes only if all 64 interrupt cycles (32 active-high + 32 active-low) complete without any errors.",
    "meta_test_steps": "1. Include headers: lss_sysreg.h, stdio.h, test_define.c, test_common.h. test_define.c includes gpio/gpio_def.h, gpio/gpio_offset.h.\n2. Declare variables: gpio_number, test_err, i (global); rdata, wr_val (local); extern int_pend.\n3. Initialize test_err = 0.\n4. Enable GIC IRQ: GIC_EnableIRQ(87) under #ifdef GPIO0, GIC_EnableIRQ(88) under #ifdef GPIO1.\n5. Enable sysreg interrupt: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR) under #ifdef GPIO0, or LSS_SYSREG_INTR_EN1_GPIO1_INTR under #ifdef GPIO1.\n6. Phase 1 - Active High Level Interrupt Loop (i = 0 to 31):\n   6a. wr_val = 1 << i.\n   6b. write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4), 0x00180000).\n   6c. wait_on(50).\n   6d. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val).\n   6e. wait_on(10).\n   6f. write_reg(0xA0243ffc, 0xffffffff).\n   6g. int_pend = 1.\n   6h. Poll: while(int_pend == 1) { wait_on(10); }.\n7. wait_on(100) between phases.\n8. Phase 2 - Active Low Level Interrupt Loop (i = 0 to 31):\n   8a-8h. Similar to Phase 1 with level_sel=0 and inverted trigger pattern.\n9. IRQ Handler verifies raw status, group status, clears interrupt, verifies clear, disables group, clears sysreg status.\n10. finish(test_err).",
    "test_steps": "1. Enable the GIC interrupt line corresponding to the GPIO instance under test.\n2. Enable the system-level GPIO interrupt by writing to the intr_en1 register to set the appropriate GPIO interrupt enable bit.\n3. For each of the 32 GPIO pins (active-high level interrupt phase):\n   a. Configure the GPIO pin in input mode with level_sel set to 1 (active high) and negative-edge interrupt enabled by writing to the corresponding gp0_gpio_8 register (computed offset for each pin).\n   b. Enable the group interrupt for the specific GPIO pin by writing to the GPIO group interrupt enable register.\n   c. Trigger the interrupt by writing to the SRAM trigger location.\n   d. Wait for the interrupt to be serviced by polling the interrupt pending flag.\n4. For each of the 32 GPIO pins (active-low level interrupt phase):\n   a. Configure the GPIO pin in input mode with level_sel set to 0 (active low) by writing to the corresponding gp0_gpio_8 register.\n   b. Enable the group interrupt for the specific GPIO pin.\n   c. Trigger the interrupt by writing the inverted pattern to the SRAM trigger location.\n   d. Wait for the interrupt to be serviced.\n5. In the interrupt handler, for each interrupt received:\n   a. Read the corresponding gp0_gpio_8 register and verify the raw interrupt status bit (bit 1) is asserted.\n   b. Read the GPIO group interrupt status register and verify the correct GPIO bit is set.\n   c. Clear the interrupt by writing to the interrupt clear bit (bit 16) of the corresponding gp0_gpio_8 register.\n   d. Read back the gp0_gpio_8 register and verify the interrupt has been cleared successfully.\n   e. Disable the group interrupt and verify the group interrupt status register reads zero.\n   f. Clear the system-level interrupt status in raw_stcr1 and verify it is cleared by reading back.\n   g. Clear the GIC interrupt.\n6. Evaluate the final test result: PASS if all 64 interrupt cycles complete with zero errors; FAIL otherwise.",
    "meta_impacted_registers": "MIZAR_LSS_SYSREG_INTR_EN1; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1; 0xA0243ffc; MIZAR_GPIO_GP0_INTR1_INTR_STS1; MIZAR_LSS_SYSREG_RAW_STCR1",
    "impacted_registers": "intr_en1; gp0_gpio_8; raw_stcr1",
    "validation_acceptance_criteria": "1. For each of the 32 GPIO pins in active-high level mode, the raw interrupt status bit (bit 1) in the corresponding gp0_gpio_8 register must be asserted when the interrupt is triggered.\n2. For each of the 32 GPIO pins in active-low level mode, the raw interrupt status bit (bit 1) in the corresponding gp0_gpio_8 register must be asserted when the interrupt is triggered.\n3. The group interrupt status register must reflect the correct GPIO bit being set for each triggered interrupt.\n4. After writing to the interrupt clear bit (bit 16), the gp0_gpio_8 register must read back the expected cleared value (0x100001).\n5. After disabling the group interrupt, the group interrupt status register must read zero.\n6. After clearing the system-level interrupt status in raw_stcr1, the GPIO interrupt bit must read as zero.\n7. The test must complete all 64 interrupt cycles (32 active-high + 32 active-low) with zero errors to be considered PASS.",
    "speed": "NA",
    "mode": "NA",
    "remarks": "The test exercises both active-high and active-low level-select interrupt modes for all 32 GPIO pins in the group. GPIO pins are configured in input mode throughout the test. The interrupt flow covers the full path from GPIO pin-level raw status through group interrupt status to system-level sysreg interrupt status. An SRAM trigger location is used to stimulate the interrupt. The addr_array defined in test_define.c is not iterated in program.c; instead, computed addressing from the base register with offset (i * 4) is used. Three registers could not be mapped to specification documents: the GPIO group interrupt enable register (offset 0x84), the GPIO group interrupt status register (offset 0x88), and the SRAM trigger address (0xA0243ffc).",
    "register_details": [
      {"macro": "MIZAR_LSS_SYSREG_INTR_EN1", "register_name": "intr_en1", "base_address": "0xA0000000", "offset": "0x28", "resolved_address": "0xA0000028", "operation": "write", "resolution_status": "resolved", "mapping_status": "matched", "register_width": 32, "fields": [{"field_name": "gpio0_intr", "bit_position": "1", "access_type": "INTREN1", "reset_value": "0", "description": "GPIO0 interrupt enable"}, {"field_name": "gpio1_intr", "bit_position": "2", "access_type": "INTREN1", "reset_value": "0", "description": "GPIO1 interrupt enable"}]},
      {"macro": "MIZAR_GPIO_GP0_GPIO_8", "register_name": "gp0_gpio_8", "base_address": "0xA001A000", "offset": "0x0", "resolved_address": "0xA001A000", "operation": "read_modify_write", "resolution_status": "resolved", "mapping_status": "matched", "register_width": 32, "fields": [{"field_name": "data_in", "bit_position": "0", "access_type": "RO", "reset_value": "0", "description": "GPIO pin level"}, {"field_name": "intr_raw_sts", "bit_position": "1", "access_type": "RO", "reset_value": "0", "description": "Interrupt raw status"}, {"field_name": "intr_clr", "bit_position": "16", "access_type": "WO", "reset_value": "0", "description": "Interrupt clear"}, {"field_name": "level_sel", "bit_position": "19", "access_type": "RW", "reset_value": "0", "description": "Level select"}, {"field_name": "io_ctrl", "bit_position": "20", "access_type": "RW2", "reset_value": "1", "description": "IO control"}]},
      {"macro": "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "register_name": "NA", "base_address": "0xA001A000", "offset": "0x84", "resolved_address": "0xA001A084", "operation": "write", "resolution_status": "resolved", "mapping_status": "unresolved", "register_width": 32, "fields": []},
      {"macro": "0xA0243ffc", "register_name": "NA", "base_address": "NA", "offset": "0xA0243ffc", "resolved_address": "0xA0243ffc", "operation": "write", "resolution_status": "direct_hex", "mapping_status": "unresolved", "register_width": 32, "fields": []},
      {"macro": "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "register_name": "NA", "base_address": "0xA001A000", "offset": "0x88", "resolved_address": "0xA001A088", "operation": "read", "resolution_status": "resolved", "mapping_status": "unresolved", "register_width": 32, "fields": []},
      {"macro": "MIZAR_LSS_SYSREG_RAW_STCR1", "register_name": "raw_stcr1", "base_address": "0xA0000000", "offset": "0x24", "resolved_address": "0xA0000024", "operation": "read_modify_write", "resolution_status": "resolved", "mapping_status": "matched", "register_width": 32, "fields": [{"field_name": "gpio0_intr", "bit_position": "1", "access_type": "RAWSTCR1", "reset_value": "0", "description": "GPIO0 interrupt raw status/clear"}, {"field_name": "gpio1_intr", "bit_position": "2", "access_type": "RAWSTCR1", "reset_value": "0", "description": "GPIO1 interrupt raw status/clear"}]}
    ]
  }
]

# ══════════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ══════════════════════════════════════════════════════════════
wb = Workbook()
ws_tp = wb.active
ws_tp.title = "TestPlan"
ws_md = wb.create_sheet("MetaData")

# Styles
hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")

# ── TestPlan Sheet ──
tp_headers = ["Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
              "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
              "Test Steps / Procedure", "Impacted Registers",
              "Validation / Acceptance Criteria", "Code Generation"]

for c, h in enumerate(tp_headers, 1):
    cell = ws_tp.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = wrap_align

for r, tc in enumerate(json_data, 2):
    ws_tp.cell(row=r, column=1, value=tc["index"]).alignment = wrap_align
    ws_tp.cell(row=r, column=2, value=tc["ss_module"]).alignment = wrap_align
    ws_tp.cell(row=r, column=3, value=tc["feature"]).alignment = wrap_align
    ws_tp.cell(row=r, column=4, value=tc["test_case_name"]).alignment = wrap_align
    ws_tp.cell(row=r, column=5, value=tc["test_description"]).alignment = wrap_align
    ws_tp.cell(row=r, column=6, value=tc["speed"]).alignment = wrap_align
    ws_tp.cell(row=r, column=7, value=tc["mode"]).alignment = wrap_align
    ws_tp.cell(row=r, column=8, value="").alignment = wrap_align
    ws_tp.cell(row=r, column=9, value="").alignment = wrap_align
    ws_tp.cell(row=r, column=10, value=tc["remarks"]).alignment = wrap_align
    ws_tp.cell(row=r, column=11, value=tc["test_steps"]).alignment = wrap_align
    ws_tp.cell(row=r, column=12, value=tc["impacted_registers"]).alignment = wrap_align
    ws_tp.cell(row=r, column=13, value=tc["validation_acceptance_criteria"]).alignment = wrap_align
    ws_tp.cell(row=r, column=14, value="").alignment = wrap_align

# ── MetaData Sheet ──
md_headers = ["Index", "Test Case Name", "Meta Test Description",
              "Meta Test Steps / Procedure", "Meta Impacted Registers",
              "Meta Validation / Acceptance Criteria",
              "Meta Headers", "Meta Macros", "Meta Arrays"]

for c, h in enumerate(md_headers, 1):
    cell = ws_md.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = wrap_align

for r, tc in enumerate(json_data, 2):
    # Build Meta Headers from meta_test_description
    meta_headers = ""
    meta_desc = tc.get("meta_test_description", "")
    if "Headers included:" in meta_desc:
        hdr_line = meta_desc.split("Headers included:")[-1].split("\n")[0].strip()
        meta_headers = hdr_line.rstrip(".")
    
    # Build Meta Macros from register_details
    macros_list = []
    for rd in tc.get("register_details", []):
        m = rd["macro"]
        rn = rd.get("register_name", "NA")
        ba = rd.get("base_address", "NA")
        off = rd.get("offset", "NA")
        ra = rd.get("resolved_address", "NA")
        op = rd.get("operation", "NA")
        rs = rd.get("resolution_status", "NA")
        ms = rd.get("mapping_status", "NA")
        macros_list.append(f"{m} -> {rn} | base:{ba} offset:{off} addr:{ra} | op:{op} | res:{rs} map:{ms}")
    meta_macros = "\n".join(macros_list)
    
    # Build Meta Arrays from register_details fields
    arrays_list = []
    for rd in tc.get("register_details", []):
        if rd.get("fields"):
            fields_str = "; ".join([f"{f['field_name']}[{f['bit_position']}]({f['access_type']})=rst:{f['reset_value']}" for f in rd["fields"]])
            arrays_list.append(f"{rd['register_name']}({rd['register_width']}b): {fields_str}")
    meta_arrays = "\n".join(arrays_list)
    
    ws_md.cell(row=r, column=1, value=tc["index"]).alignment = wrap_align
    ws_md.cell(row=r, column=2, value=tc["test_case_name"]).alignment = wrap_align
    ws_md.cell(row=r, column=3, value=tc.get("meta_test_description", "")).alignment = wrap_align
    ws_md.cell(row=r, column=4, value=tc.get("meta_test_steps", "")).alignment = wrap_align
    ws_md.cell(row=r, column=5, value=tc.get("meta_impacted_registers", "")).alignment = wrap_align
    ws_md.cell(row=r, column=6, value=tc.get("validation_acceptance_criteria", "")).alignment = wrap_align
    ws_md.cell(row=r, column=7, value=meta_headers).alignment = wrap_align
    ws_md.cell(row=r, column=8, value=meta_macros).alignment = wrap_align
    ws_md.cell(row=r, column=9, value=meta_arrays).alignment = wrap_align

# ── Freeze Panes ──
ws_tp.freeze_panes = "A2"
ws_md.freeze_panes = "A2"

# ── Auto-size Columns ──
for ws in [ws_tp, ws_md]:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_len = max(max_len, len(line))
        adjusted = min(max_len + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 12)

# ── Hide MetaData ──
ws_md.sheet_state = "veryHidden"

# ── Generate IST Timestamp ──
ist = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(ist)
timestamp = now_ist.strftime("%Y%m%d_%H%M%S")
filename = f"GPIO_TestPlan_{timestamp}.xlsx"

# ── Save ──
script_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(script_dir, filename)
wb.save(output_path)
file_size = os.path.getsize(output_path)

# ── Verify ──
from openpyxl import load_workbook
wb_check = load_workbook(output_path)
assert "TestPlan" in wb_check.sheetnames
assert "MetaData" in wb_check.sheetnames
assert wb_check["TestPlan"].max_row == 3  # header + 2 data rows
assert wb_check["MetaData"].max_row == 3
wb_check.close()

print(f"SUCCESS")
print(f"Filename: {filename}")
print(f"Path: {output_path}")
print(f"Size: {file_size} bytes")
print(f"TestPlan rows: 2")
print(f"MetaData rows: 2")
print(f"Validation: PASSED")
