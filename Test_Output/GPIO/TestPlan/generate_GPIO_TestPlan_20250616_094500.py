#!/usr/bin/env python3
"""
GPIO TestPlan XLSX Generator & GitHub Pusher
=============================================
Run this script ONCE to:
  1. Generate GPIO_TestPlan_20250616_094500.xlsx locally
  2. Push it to GitHub at Test_Output/GPIO/TestPlan/

Requirements: pip install openpyxl requests
Usage: python generate_GPIO_TestPlan_20250616_094500.py

Set environment variable GITHUB_TOKEN before running:
  export GITHUB_TOKEN=ghp_xxxxxxxxxxxx
"""
import os
import sys
import base64
import json
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("WARNING: requests not installed. XLSX will be generated locally only.")
    requests = None

# ============================================================
# CONFIGURATION
# ============================================================
OUTPUT_FILENAME = "GPIO_TestPlan_20250616_094500.xlsx"
GITHUB_OWNER = "titusbspgit"
GITHUB_REPO = "PSVValidation"
GITHUB_BRANCH = "main"
GITHUB_PATH = f"Test_Output/GPIO/TestPlan/{OUTPUT_FILENAME}"

# ============================================================
# TESTPLAN DATA (21 fields, 1 test case) - EXACTLY as aggregated
# ============================================================
HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Meta Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Meta Test Steps / Procedure",
    "Impacted Registers",
    "Meta Impacted Registers",
    "Validation / Acceptance Criteria",
    "Meta Validation / Acceptance Criteria",
    "Code Generation",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays"
]

DATA_ROW = [
    1,
    "GPIO",
    "Register Read Write Verification",
    "gpio_reg_wr_rd_test",
    "This test verifies the register read and write functionality for GPIO GP0 per-pin configuration registers. The test reads default reset values from registers gp0_gpio_8 (offset 0x0), gp0_gpio_9 (offset 0x4), and gp0_gpio_10 (offset 0x8), compares them against expected default values using read masks, then writes multiple data patterns (0xFFFFFFFF, 0x55555555, 0xAAAAAAAA, 0xA5A5A5A5, 0xF5F5F5F5, 0xFFFF0000) to each register using write masks, reads back, and verifies correctness. A soft reset verification flow is present but conditionally disabled.",
    "Test reads registers MIZAR_GPIO_GP0_GPIO_8 (base 0xA001A000 + offset 0x0), MIZAR_GPIO_GP0_GPIO_9 (base 0xA001A000 + offset 0x4), MIZAR_GPIO_GP0_GPIO_10 (base 0xA001A000 + offset 0x8) via read_reg(addr_array[i]). Default values compared using GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL with corresponding READ_MASK macros. Write patterns applied via write_reg(addr_array[i], (wr_data[j] & WRITE_MASK[i]) | (DEFAULT_VAL[i] & ~WRITE_MASK[i])). Readback verified as (read_val & READ_MASK[i]) == ((wr_data[j] & WRITE_MASK[i]) | (DEFAULT_VAL[i] & ~WRITE_MASK[i])) & READ_MASK[i]. SOFT_RST_REG_ADDRESS (0x00000000) used in disabled #ifdef 0 block for soft_reset_chk().",
    "NA",
    "Polling",
    "0x0",
    "0x8",
    "SOFT_RST_REG_ADDRESS is inside a disabled #ifdef 0 block and will not execute. RO fields (data_in bit[0], intr_raw_sts bit[1]) cannot be written and retain reset values on write attempts. WO field (intr_clr bit[16]) reads back as 0. RW2 fields (io_ctrl bit[20], dout bit[21]) behave as standard read-write for this test. Register width is 32 bits, AHB bus, ahb_clk domain, hreset_n active-low async reset.",
    "1. Initialize the GPIO subsystem and configure AHB bus access.\n2. Read the default value of register gp0_gpio_8 at offset 0x0.\n3. Read the default value of register gp0_gpio_9 at offset 0x4.\n4. Read the default value of register gp0_gpio_10 at offset 0x8.\n5. Compare each read value (masked with read mask 0x003F0003) against the expected default value 0x00100000 (io_ctrl=1 at bit 20).\n6. For each register, write pattern 0xFFFFFFFF applying write mask 0x003F0000, read back, and verify.\n7. Repeat step 6 with pattern 0x55555555.\n8. Repeat step 6 with pattern 0xAAAAAAAA.\n9. Repeat step 6 with patterns 0xA5A5A5A5, 0xF5F5F5F5, and 0xFFFF0000.\n10. Report PASS if all read-back values match expected values after masking; report FAIL otherwise.",
    "1. addr_array[49] initialized with {MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10}.\n2. chk_rst_val(): for i=0 to CNT-1: rd_data = read_reg(addr_array[i]); compare (rd_data & READ_MASK[i]) vs (DEFAULT_VAL[i] & READ_MASK[i]).\n3. chk_rd_wr(): for each wr_data in {0xFFFFFFFF, 0x55555555, 0xAAAAAAAA, 0xA5A5A5A5, 0xF5F5F5F5, 0xFFFF0000}: for i=0 to CNT-1: write_reg(addr_array[i], (wr_data & WRITE_MASK[i]) | (DEFAULT_VAL[i] & ~WRITE_MASK[i])); rd_data = read_reg(addr_array[i]); verify (rd_data & READ_MASK[i]) == ((wr_data & WRITE_MASK[i]) | (DEFAULT_VAL[i] & ~WRITE_MASK[i])) & READ_MASK[i].\n4. soft_reset_chk() inside #ifdef 0: write_reg(SOFT_RST_REG_ADDRESS, SOFT_RST_REG_DATA); read_reg(SOFT_RST_REG_ADDRESS) \u2014 DISABLED.\n5. main() calls chk_rst_val() then chk_rd_wr(); prints PASS/FAIL.",
    "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10",
    "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; SOFT_RST_REG_ADDRESS",
    "PASS: All register default value reads match expected value 0x00100000 after applying read mask 0x003F0003. All write-read cycles for 6 data patterns return expected masked values for all 3 registers. FAIL: Any single read-back mismatch after masking causes test failure.",
    "PASS: For all i in [0, CNT): (read_reg(addr_array[i]) & READ_MASK[i]) == (DEFAULT_VAL[i] & READ_MASK[i]) AND for all wr_data patterns: (read_reg(addr_array[i]) & READ_MASK[i]) == ((wr_data & WRITE_MASK[i]) | (DEFAULT_VAL[i] & ~WRITE_MASK[i])) & READ_MASK[i]. FAIL: Any mismatch in above comparisons. Write Mask: 0x003F0000. Read Mask: 0x003F0003. Default: 0x00100000.",
    "Not",
    '#include <stdio.h>; #include <stdlib.h>; #include <string.h>; #include "common.h"; #include "test_define.c"; #include "gpio_headers.h"',
    "#define GPIO0 1; #define SOFT_RST_REG_ADDRESS 0x00000000; #define SOFT_RST_REG_DATA 0x00000400; #define CNT 3",
    "addr_array[49]={MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10}; default_val_array[49]={GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL}; rd_mask_array[49]={GPIO_GP0_GPIO_8_READ_MASK, GPIO_GP0_GPIO_9_READ_MASK, GPIO_GP0_GPIO_10_READ_MASK}; wr_mask_array[49]={GPIO_GP0_GPIO_8_WRITE_MASK, GPIO_GP0_GPIO_9_WRITE_MASK, GPIO_GP0_GPIO_10_WRITE_MASK}; wr_data[6]={0xFFFFFFFF, 0x55555555, 0xAAAAAAAA, 0xA5A5A5A5, 0xF5F5F5F5, 0xFFFF0000}; rd_data_array[49]; exp_data_array[49]"
]

METADATA = [
    ["IP_Name", "GPIO"],
    ["Token", "#define GPIO0 1"],
    ["Base_Address", "0xA001A000"],
    ["Register_Block", "gp0"],
    ["Bus_Type", "AHB"],
    ["Bus_Width", "32"],
    ["Clock", "ahb_clk"],
    ["Reset_Signal", "hreset_n"],
    ["Reset_Type", "Active-low asynchronous"],
    ["Source_Module", "gpio"],
    ["Generation_Timestamp_IST", "2025-06-16T09:45:00+05:30"],
    ["Generator", "Ag_Excel_Generator Agent"],
    ["Workflow", "PSV TestPlan Generation Pipeline"],
    ["Total_Test_Cases", "1"],
    ["Resolved_Registers", "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10"],
    ["Unresolved_Macros", "SOFT_RST_REG_ADDRESS"],
    ["Write_Patterns", "0xFFFFFFFF; 0x55555555; 0xAAAAAAAA; 0xA5A5A5A5; 0xF5F5F5F5; 0xFFFF0000"],
    ["Read_Mask", "0x003F0003"],
    ["Write_Mask", "0x003F0000"],
    ["Default_Value", "0x00100000"],
    ["Spec_Source_Register", "gp0_autoreg.xlsx"],
    ["Spec_Source_System", "LSS_REGS (1).pdf"],
    ["GitHub_Source_Path", "TestRepo/gpio/gpio_reg_wr_rd_test/"],
]


def create_workbook():
    """Create the XLSX workbook with TestPlan + MetaData sheets."""
    wb = Workbook()

    # ===== SHEET 1: TestPlan (visible) =====
    ws = wb.active
    ws.title = "TestPlan"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    data_font = Font(name='Calibri', size=10)

    # Row 1: Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Row 2: Data
    for col_idx, value in enumerate(DATA_ROW, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

    ws.freeze_panes = 'A2'

    col_widths = {
        1: 8, 2: 15, 3: 30, 4: 28, 5: 60, 6: 60, 7: 10, 8: 12,
        9: 20, 10: 20, 11: 50, 12: 60, 13: 60, 14: 40, 15: 50,
        16: 55, 17: 60, 18: 16, 19: 50, 20: 50, 21: 70
    }
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 200

    # ===== SHEET 2: MetaData (very hidden) =====
    ws_meta = wb.create_sheet(title="MetaData")
    ws_meta.sheet_state = 'veryHidden'

    meta_hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    meta_hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')

    for col_idx, hdr in enumerate(["Property", "Value"], 1):
        cell = ws_meta.cell(row=1, column=col_idx, value=hdr)
        cell.font = meta_hdr_font
        cell.fill = meta_hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, row_data in enumerate(METADATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_meta.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    ws_meta.column_dimensions['A'].width = 30
    ws_meta.column_dimensions['B'].width = 80
    ws_meta.freeze_panes = 'A2'

    return wb


def save_locally(wb):
    """Save workbook to local file and return bytes."""
    # Save to BytesIO for GitHub push
    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    # Also save to disk
    local_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILENAME)
    with open(local_path, 'wb') as f:
        f.write(xlsx_bytes)

    print(f"[LOCAL] Saved: {local_path}")
    print(f"[LOCAL] Size: {len(xlsx_bytes)} bytes")
    return xlsx_bytes, local_path


def push_to_github(xlsx_bytes):
    """Push the XLSX binary to GitHub via REST API."""
    if requests is None:
        print("[GITHUB] Skipped - requests library not available")
        return None

    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        print("[GITHUB] Skipped - GITHUB_TOKEN not set")
        print("[GITHUB] Set it with: export GITHUB_TOKEN=ghp_xxxxxxxxxxxx")
        return None

    api_url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_PATH}"
    headers_api = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json"
    }

    # Check if file already exists (to get SHA for update)
    existing_sha = None
    resp = requests.get(api_url, headers=headers_api, params={"ref": GITHUB_BRANCH})
    if resp.status_code == 200:
        existing_sha = resp.json().get("sha")

    payload = {
        "message": f"Add {OUTPUT_FILENAME} - GPIO TestPlan (1 test case, 21 columns)\n\nGenerated by Ag_Excel_Generator Agent\nIST: 2025-06-16 09:45:00\nSheets: TestPlan (visible), MetaData (veryHidden)\nTest: gpio_reg_wr_rd_test\nRegisters: gp0_gpio_8, gp0_gpio_9, gp0_gpio_10",
        "content": base64.b64encode(xlsx_bytes).decode("ascii"),
        "branch": GITHUB_BRANCH
    }
    if existing_sha:
        payload["sha"] = existing_sha

    resp = requests.put(api_url, headers=headers_api, json=payload)

    if resp.status_code in (200, 201):
        result = resp.json()
        content_sha = result.get("content", {}).get("sha", "N/A")
        commit_sha = result.get("commit", {}).get("sha", "N/A")
        html_url = result.get("content", {}).get("html_url", "N/A")
        print(f"[GITHUB] Push SUCCESS")
        print(f"[GITHUB] File SHA: {content_sha}")
        print(f"[GITHUB] Commit SHA: {commit_sha}")
        print(f"[GITHUB] URL: {html_url}")
        return {
            "content_sha": content_sha,
            "commit_sha": commit_sha,
            "html_url": html_url
        }
    else:
        print(f"[GITHUB] Push FAILED: {resp.status_code}")
        print(f"[GITHUB] Response: {resp.text}")
        return None


def main():
    print("=" * 70)
    print("GPIO TestPlan XLSX Generator & GitHub Pusher")
    print("=" * 70)
    print(f"Output: {OUTPUT_FILENAME}")
    print(f"Target: {GITHUB_OWNER}/{GITHUB_REPO}/{GITHUB_PATH}")
    print(f"Branch: {GITHUB_BRANCH}")
    print()

    # Step 1: Create workbook
    print("[STEP 1] Creating XLSX workbook...")
    wb = create_workbook()
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  TestPlan headers: {len(HEADERS)}")
    print(f"  TestPlan data rows: 1")
    print(f"  MetaData rows: {len(METADATA)}")
    print()

    # Step 2: Save locally
    print("[STEP 2] Saving locally...")
    xlsx_bytes, local_path = save_locally(wb)
    print()

    # Step 3: Push to GitHub
    print("[STEP 3] Pushing to GitHub...")
    github_result = push_to_github(xlsx_bytes)
    print()

    # Step 4: Summary
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    status_json = {
        "status": "SUCCESS" if github_result else "SUCCESS (local only)",
        "execution_mode": "Fallback Automation",
        "output_file_path": GITHUB_PATH,
        "local_file_path": local_path,
        "github_url": github_result["html_url"] if github_result else "NA",
        "file_size_bytes": len(xlsx_bytes),
        "sheets": ["TestPlan (visible)", "MetaData (veryHidden)"],
        "headers_count": len(HEADERS),
        "data_rows": 1,
        "metadata_rows": len(METADATA)
    }
    print(json.dumps(status_json, indent=2))
    print("=" * 70)


if __name__ == '__main__':
    main()
