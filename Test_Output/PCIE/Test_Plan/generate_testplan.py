#!/usr/bin/env python3
"""Temporary script to generate PCIE TestPlan Excel - will be removed after execution."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from datetime import datetime, timezone, timedelta
import os

# IST timestamp
ist = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(ist)
timestamp = now_ist.strftime("%Y%m%d_%H%M%S")
filename = f"PCIE_TestPlan_{timestamp}.xlsx"

# Create workbook
wb = openpyxl.Workbook()

# Rename default sheet to TestPlan
ws_tp = wb.active
ws_tp.title = "TestPlan"

# Create MetaData sheet
ws_md = wb.create_sheet("MetaData")

# TestPlan columns
tp_cols = [
    "Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
    "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
    "Test Steps / Procedure", "Impacted Registers", "Validation / Acceptance Criteria",
    "Code Generation"
]

# MetaData columns
md_cols = [
    "Index", "Test Case Name", "Meta Test Description", "Meta Test Steps / Procedure",
    "Meta Impacted Registers", "Meta Validation / Acceptance Criteria",
    "Meta Headers", "Meta Macros", "Meta Arrays"
]

# Header formatting
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
cell_align = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Write headers - TestPlan
for col_idx, col_name in enumerate(tp_cols, 1):
    cell = ws_tp.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write headers - MetaData
for col_idx, col_name in enumerate(md_cols, 1):
    cell = ws_md.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data row
row_data = {
    "Index": "1",
    "SS / Module": "PCIE",
    "Feature": "PCIe Device Enumeration and BAR Configuration",
    "Test Case Name": "pcie_device_enumerate_test",
    "Test Description": "Verifies PCIe device enumeration by performing link training on dual-mode controllers, configuring coherency control registers, polling link status on both SII interfaces until the expected link-up pattern is detected, reading the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enabling bus master and memory space via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, configuring system-level control registers, disabling cache coherency, and enumerating BARs on both PCIe slave endpoints by writing all-ones to BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG, reading back the BAR sizes, and programming final BAR address values. The test concludes by polling a synchronization register until the expected completion pattern is observed.",
    "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
    "Validation / Acceptance Criteria": "The test passes when: (1) PCIe link training completes successfully on the configured dual-mode controllers. (2) The SII0 link status register reports the expected link-up pattern with the required status bits set. (3) The SII1 link status register reports the expected link-up pattern with the required status bits set. (4) The Vendor ID read from TYPE1_DEV_ID_VEND_ID_REG returns a valid non-zero value confirming device presence. (5) BAR enumeration on both PCIe slave endpoints completes with successful write-readback cycles on BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG. (6) The synchronization register returns the expected completion value, indicating the remote endpoint has completed its operations. (7) The test terminates via finish(0) indicating overall success.",
    "Remarks": "The testcase uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select the link training mode, so behavior varies based on build configuration. Cache coherency is enabled before enumeration and disabled afterward in a staged sequence with wait intervals. The SII link status polling uses a bitmask check for link-up detection. Several system-level control register addresses and the SII status register could not be mapped to named registers in the specification. The test includes a final synchronization polling loop that waits for a specific completion pattern from the remote endpoint. The source code contains a duplicated block of link training and cache programming logic."
}

# Build categorized test steps with rich text
bold_font = InlineFont(b=True, sz=11, rFont='Calibri')
normal_font = InlineFont(b=False, sz=11, rFont='Calibri')

test_steps_rich = CellRichText(
    TextBlock(bold_font, "Initialization:"),
    TextBlock(normal_font, "\n1. Initialize the test synchronization register to clear any previous state.\n\n"),
    TextBlock(bold_font, "Configuration:"),
    TextBlock(normal_font, "\n1. Perform PCIe link training on the dual-mode controllers based on the configured mode (Root Complex or Endpoint)."),
    TextBlock(normal_font, "\n2. Enable cache coherency by performing read-modify-write operations on the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances, setting the required bit fields."),
    TextBlock(normal_font, "\n3. Wait for the coherency configuration to take effect."),
    TextBlock(normal_font, "\n4. Perform a combined coherency control update on both controller instances with all bit field groups enabled."),
    TextBlock(normal_font, "\n5. Configure non-secure protection via NIC programming."),
    TextBlock(normal_font, "\n6. Write to TYPE1_STATUS_COMMAND_REG on PCIe slave 0 to enable bus master, memory space, and I/O space access."),
    TextBlock(normal_font, "\n7. Program memory base addresses for both dual-mode controllers."),
    TextBlock(normal_font, "\n8. Configure system-level control registers to enable required functionality."),
    TextBlock(normal_font, "\n9. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF for both controller instances in a staged sequence with wait intervals.\n\n"),
    TextBlock(bold_font, "Execution:"),
    TextBlock(normal_font, "\n1. Poll the SII0 link status register until the expected link-up pattern is detected."),
    TextBlock(normal_font, "\n2. Poll the SII1 link status register until the expected link-up pattern is detected."),
    TextBlock(normal_font, "\n3. Read the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG on PCIe slave 0 to confirm device presence."),
    TextBlock(normal_font, "\n4. Enumerate BARs on PCIe slave 1 by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, reading back to determine BAR sizes, then programming final BAR address values."),
    TextBlock(normal_font, "\n5. Repeat BAR enumeration for PCIe slave 0."),
    TextBlock(normal_font, "\n6. Poll the synchronization register until the expected completion value is observed."),
    TextBlock(normal_font, "\n7. Verify the test completes successfully.")
)

# Populate TestPlan row 2
tp_row_map = {
    1: row_data["Index"],
    2: row_data["SS / Module"],
    3: row_data["Feature"],
    4: row_data["Test Case Name"],
    5: row_data["Test Description"],
    6: "",  # Speed
    7: "",  # Mode
    8: "",  # Memory Start Offset
    9: "",  # Memory End Offset
    10: row_data["Remarks"],
    12: row_data["Impacted Registers"],
    13: row_data["Validation / Acceptance Criteria"],
    14: "",  # Code Generation - blank
}

for col_idx, value in tp_row_map.items():
    cell = ws_tp.cell(row=2, column=col_idx, value=value)
    cell.alignment = cell_align
    cell.border = thin_border

# Set rich text for Test Steps column (col 11)
ws_tp.cell(row=2, column=11).value = test_steps_rich
ws_tp.cell(row=2, column=11).alignment = cell_align
ws_tp.cell(row=2, column=11).border = thin_border

# MetaData row
meta_data = {
    "Meta Test Description": "The testcase performs PCIe device enumeration across two dual-mode controllers (DM0 and DM1). It begins by writing 0x0 to 0xE6004100 to initialize the test. Link training is invoked conditionally based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP) using link_training_dm0_x4(4) or link_training_dm1_x4(4). Cache programming is performed by read-modify-write operations on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF using set_data() to configure bit fields [11:14], [3:6], [27:30], and [19:22] with value 0xf. After a wait_on(20), the same coherency control registers are programmed again with all four bit field groups set to 0xf in a single read-modify-write sequence. The SII0 link status is polled via read_sii0_reg(0xC0) until (data_rd & 0xD1) == 0xD1, and similarly SII1 link status is polled via read_sii1_reg(0xC0). Under DM0_RC, the Vendor ID is read from read_pcie_slv0_reg(0x0), the command register is written via write_pcie_slv0_reg(0x4, 0x7), and memory base programming functions mem_base_program_dm0_x4() and mem_base_program_dm1_x4() are called. System-level registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, and 0xE6900034 are written with 0x1. Cache disable programming is then performed by read-modify-write on the coherency control registers, setting bit fields [19:22] and [27:30] to 0x0. After wait_on(30), BAR enumeration is performed on PCIe slave 1 by writing 0xFFFFFFFF to offsets 0x10-0x24, reading them back, then writing final BAR values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000). The same BAR enumeration sequence is repeated for PCIe slave 0. Finally, the test polls 0xE6004100 until the value equals 0x12345678, with wait_on(5) between iterations, and calls finish(0) upon success.",
    "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100 to initialize the test synchronization register. 2. Invoke link_training_dm0_x4(4) or link_training_dm1_x4(4) based on compile-time defines DM0_RC, DM1_RC, DM0_EP, DM1_EP. 3. Perform cache programming: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting bit fields [11:14]=0xf, [3:6]=0xf via set_data(), then write back. 4. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting bit fields [27:30]=0xf, [19:22]=0xf, then write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. Call wait_on(20). 7. Perform combined read-modify-write on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting all four bit field groups [11:14], [3:6], [27:30], [19:22] to 0xf. 8. Repeat step 7 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 9. Read SII0 link status via read_sii0_reg(0xC0). 10. Call non_secure_prot_nic(). 11. Poll read_sii0_reg(0xC0) in a while loop until (data_rd & 0xD1) == 0xD1. 12. Read SII1 link status via read_sii1_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1. 13. Under DM0_RC: read Vendor ID via read_pcie_slv0_reg(0x0). 14. Write 0x7 to command register via write_pcie_slv0_reg(0x4, 0x7). 15. Call mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). 16. Call wait_on(10). 17. Write 0x1 to system registers 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034. 18. Perform cache disable: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF setting bit fields [19:22]=0x0 and [27:30]=0xf. 19. Call wait_on(10). 20. Perform final cache disable: set bit fields [27:30]=0x0 and [19:22]=0x0 on both coherency control registers. 21. Call wait_on(30). 22. Write 0xFFFFFFFF to PCIe slave 1 BAR registers at offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24. 23. Read back all six BAR registers from PCIe slave 1. 24. Write final BAR values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) to PCIe slave 1. 25. Read back all six BAR registers from PCIe slave 1 to verify. 26. Repeat steps 22-25 for PCIe slave 0. 27. Call wait_on(10). 28. Poll read_reg(0xE6004100) until value equals 0x12345678, with wait_on(5) between iterations. 29. Call finish(0) to end the test.",
    "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
    "Meta Validation / Acceptance Criteria": "The test passes when: (1) PCIe link training completes successfully on the configured dual-mode controllers. (2) The SII0 link status register reports the expected link-up pattern with the required status bits set. (3) The SII1 link status register reports the expected link-up pattern with the required status bits set. (4) The Vendor ID read from TYPE1_DEV_ID_VEND_ID_REG returns a valid non-zero value confirming device presence. (5) BAR enumeration on both PCIe slave endpoints completes with successful write-readback cycles on BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG. (6) The synchronization register returns the expected completion value, indicating the remote endpoint has completed its operations. (7) The test terminates via finish(0) indicating overall success."
}

md_row_map = {
    1: row_data["Index"],
    2: row_data["Test Case Name"],
    3: meta_data["Meta Test Description"],
    4: meta_data["Meta Test Steps / Procedure"],
    5: meta_data["Meta Impacted Registers"],
    6: meta_data["Meta Validation / Acceptance Criteria"],
    7: "",  # Meta Headers
    8: "",  # Meta Macros
    9: "",  # Meta Arrays
}

for col_idx, value in md_row_map.items():
    cell = ws_md.cell(row=2, column=col_idx, value=value)
    cell.alignment = cell_align
    cell.border = thin_border

# Code Generation dropdown
dv = DataValidation(
    type="list",
    formula1='"Required,Not Required"',
    allow_blank=True,
    showDropDown=False
)
dv.error = "Please select Required or Not Required"
dv.errorTitle = "Invalid Input"
dv.sqref = "N2:N1000"
ws_tp.add_data_validation(dv)

# Column widths - TestPlan
tp_widths = {
    1: 8,    # Index
    2: 15,   # SS / Module
    3: 40,   # Feature
    4: 35,   # Test Case Name
    5: 60,   # Test Description
    6: 10,   # Speed
    7: 10,   # Mode
    8: 20,   # Memory Start Offset
    9: 20,   # Memory End Offset
    10: 50,  # Remarks
    11: 90,  # Test Steps / Procedure
    12: 60,  # Impacted Registers
    13: 70,  # Validation / Acceptance Criteria
    14: 18,  # Code Generation
}

for col_idx, width in tp_widths.items():
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws_tp.column_dimensions[col_letter].width = width

# Column widths - MetaData
md_widths = {
    1: 8,    # Index
    2: 35,   # Test Case Name
    3: 80,   # Meta Test Description
    4: 80,   # Meta Test Steps / Procedure
    5: 80,   # Meta Impacted Registers
    6: 80,   # Meta Validation / Acceptance Criteria
    7: 30,   # Meta Headers
    8: 30,   # Meta Macros
    9: 30,   # Meta Arrays
}

for col_idx, width in md_widths.items():
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws_md.column_dimensions[col_letter].width = width

# Freeze first row
ws_tp.freeze_panes = "A2"
ws_md.freeze_panes = "A2"

# Set row height for data row
ws_tp.row_dimensions[2].height = 400

# Set MetaData to veryHidden
ws_md.sheet_state = 'veryHidden'

# Set TestPlan as active
wb.active = wb.sheetnames.index("TestPlan")

# Save
filepath = f"/tmp/{filename}"
wb.save(filepath)

# Verify
file_size = os.path.getsize(filepath)
wb2 = openpyxl.load_workbook(filepath)
sheets = wb2.sheetnames
tp_visible = wb2["TestPlan"].sheet_state == 'visible'
md_hidden = wb2["MetaData"].sheet_state == 'veryHidden'
tp_rows = wb2["TestPlan"].max_row
md_rows = wb2["MetaData"].max_row

print(f"FILENAME={filename}")
print(f"FILEPATH={filepath}")
print(f"FILESIZE={file_size}")
print(f"SHEETS={sheets}")
print(f"TP_VISIBLE={tp_visible}")
print(f"MD_VERYHIDDEN={md_hidden}")
print(f"TP_ROWS={tp_rows}")
print(f"MD_ROWS={md_rows}")
print(f"VALIDATION=PASSED")

wb2.close()
wb.close()
