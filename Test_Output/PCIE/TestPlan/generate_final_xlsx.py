#!/usr/bin/env python3
"""PCIE TestPlan Excel Generator - Agent 7
Generates PCIE_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx with TestPlan and MetaData sheets.
"""
import json
import os
import sys
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# IST timezone
IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime("%Y%m%d_%H%M%S")
filename = f"PCIE_TestPlan_{timestamp}.xlsx"

# JSON data
json_data = [
  {
    "Index": "1",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_device_enumerate_test",
    "Feature": "Device Enumeration",
    "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
    "Meta Macros": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DEBUG_DISPLAY",
    "Meta Arrays": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs PCIe device enumeration. It begins by writing 0x0 to address 0xE6004100, then conditionally invokes link training for DM0 or DM1 in RC or EP mode using link_training_dm0_x4(4) or link_training_dm1_x4(4). Cache programming is performed by reading mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, modifying specific bit fields (bits 3-6, 11-14, 19-22, 27-30) using set_data(), and writing back. The test polls read_sii0_reg(0xC0) and read_sii1_reg(0xC0) waiting for link status bits (mask 0xD1) to indicate link-up. Under DM0_RC, it reads the Vendor ID from read_pcie_slv0_reg(0x0), writes 0x7 to write_pcie_slv0_reg(0x4), and calls mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). System registers at 0xE690000C through 0xE6900034 are written with 0x1. Cache is then disabled by clearing bits 19-22 and 27-30 in the coherency control registers. BAR registers (offsets 0x10-0x24) on both slv0 and slv1 are written with 0xFFFFFFFF, read back, then programmed with specific base addresses (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) and read back again. Finally, the test polls 0xE6004100 waiting for the value 0x12345678 before calling finish(0).",
    "Test Description": "This test performs PCIe device enumeration by initializing link training, programming cache coherency control registers for both PCIE0 and PCIE1 controllers, polling link status registers until the link is established, reading the Vendor ID, enabling bus master and memory space access, programming memory base addresses, configuring system-level registers, disabling cache coherency, enumerating BAR registers on both slave ports by writing all-ones and reading back to determine BAR sizes, then programming BAR registers with specific base addresses, and finally polling a synchronization register until enumeration completion is signaled.",
    "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100 to initialize synchronization register. 2. Conditionally call link_training_dm0_x4(4) or link_training_dm1_x4(4) based on DM0_RC, DM1_RC, DM0_EP, DM1_EP defines. 3. CACHE PROGRAMMING: Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [11:14]=0xf and [3:6]=0xf, write back. 4. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF again, set_data bits [27:30]=0xf and [19:22]=0xf, write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. wait_on(20). 7. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [11:14]=0xf, [3:6]=0xf, [27:30]=0xf, [19:22]=0xf, write back. 8. Read mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [11:14]=0xf, [3:6]=0xf, [27:30]=0xf, [19:22]=0xf, write back. 9. Repeat link training and cache programming block (duplicate code block). 10. Read read_sii0_reg(0xC0) into data_rd. Call non_secure_prot_nic(). 11. Poll read_sii0_reg(0xC0) in while loop until (data_rd & 0xD1) == 0xD1. 12. Read read_sii1_reg(0xC0) into data_rd. Poll in while loop until (data_rd & 0xD1) == 0xD1. 13. Under DM0_RC: read_pcie_slv0_reg(0x0) to get Vendor ID. write_pcie_slv0_reg(0x4, 0x7). Call mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). wait_on(10). 14. Write 0x1 to 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034. 15. DISABLE_CACHE PROGRAMMING: Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [11:14]=0xf, [3:6]=0xf, write back. Read again, set_data bits [27:30]=0xf, [19:22]=0x0, write back. Repeat for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 16. wait_on(10). Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [11:14]=0xf, [3:6]=0xf, [27:30]=0x0, [19:22]=0x0, write back. Repeat for PCIE1. 17. wait_on(30). 18. Write 0xFFFFFFFF to write_pcie_slv1_reg offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24. Read back each. 19. Write specific base addresses to write_pcie_slv1_reg offsets 0x10=0x0, 0x14=0x4, 0x18=0x20000000, 0x1c=0x40000000, 0x20=0x60000000, 0x24=0x80000000. Read back each. 20. Write 0xFFFFFFFF to write_pcie_slv0_reg offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24. Read back each. 21. Write specific base addresses to write_pcie_slv0_reg offsets 0x10=0x0, 0x14=0x4, 0x18=0x20000000, 0x1c=0x40000000, 0x20=0x60000000, 0x24=0x80000000. Read back each. 22. wait_on(10). 23. Poll read_reg(0xE6004100) in while loop with wait_on(5) until value equals 0x12345678. 24. Call finish(0).",
    "Test Steps / Procedure": "1. Initialize the synchronization register by clearing it. 2. Perform PCIe link training for the configured dual-mode controller (DM0 or DM1) in either Root Complex or Endpoint mode with x4 lane width. 3. Program the COHERENCY_CONTROL_3_OFF register for both PCIE0 and PCIE1 controllers to enable cache coherency by setting specific bit fields. 4. Wait for the configuration to take effect. 5. Re-apply cache coherency settings to both PCIE0 and PCIE1 COHERENCY_CONTROL_3_OFF registers with all relevant bit fields enabled. 6. Poll the gic register on SII0 interface until link status bits indicate link-up (expected pattern confirmed). 7. Configure non-secure protection via NIC programming. 8. Poll the gic register on SII1 interface until link status bits indicate link-up. 9. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG to retrieve the Vendor ID of the connected device. 10. Write to TYPE1_STATUS_COMMAND_REG to enable bus master, memory space, and I/O space access. 11. Program memory base addresses for both DM0 and DM1 controllers. 12. Enable system-level configuration registers by writing enable values. 13. Disable cache coherency by clearing the upper bit fields in COHERENCY_CONTROL_3_OFF for both PCIE0 and PCIE1 controllers. 14. Wait for cache disable to take effect, then clear remaining coherency fields. 15. Enumerate BAR registers (BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on slave port 1 by writing all-ones and reading back to determine BAR sizes. 16. Program BAR registers on slave port 1 with specific base addresses. 17. Repeat BAR enumeration and programming on slave port 0. 18. Poll the synchronization register until the expected completion value is received, confirming successful enumeration.",
    "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
    "Impacted Registers": "COHERENCY_CONTROL_3_OFF; gic; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
    "Meta Validation / Acceptance Criteria": "1. Poll read_sii0_reg(0xC0): while loop checks (data_rd & 0xD1) != 0xD1, exits when link status bits [7,6,4,0] are all set. 2. Poll read_sii1_reg(0xC0): same condition (data_rd & 0xD1) == 0xD1. 3. Read read_pcie_slv0_reg(0x0) to verify Vendor ID is valid (printed via printf). 4. BAR sizing: write 0xFFFFFFFF to offsets 0x10-0x24 on slv1 and slv0, read back to determine implemented BAR bits. 5. BAR programming: write specific base addresses to offsets 0x10-0x24 on slv1 and slv0, read back to verify programmed values. 6. Final polling: read_reg(0xE6004100) must equal 0x12345678 to indicate successful enumeration completion. Test calls finish(0) on success.",
    "Validation / Acceptance Criteria": "1. The gic register on both SII0 and SII1 interfaces must indicate link-up by having the expected link status bit pattern set. 2. The TYPE1_DEV_ID_VEND_ID_REG must return a valid Vendor ID confirming device presence. 3. BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports must respond correctly to all-ones writes for BAR sizing and must retain programmed base address values on readback. 4. The synchronization register must return the expected completion value (indicating successful enumeration) before the test finishes. 5. The test passes by calling finish(0) after all polling conditions are met.",
    "Remarks": "The test uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select the link training mode; the BAR enumeration and Vendor ID read paths are active under DM0_RC. The source contains a duplicated code block for link training and cache programming. Multiple wait_on() calls are used for timing synchronization between configuration phases. The test polls two separate SII interfaces (SII0 and SII1) for link status. Several system-level registers at absolute addresses could not be mapped to canonical register names. The PCIE1 coherency control macro could not be resolved to a register specification."
  },
  {
    "Index": "2",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_dma_write_test",
    "Feature": "DMA Write",
    "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
    "Meta Macros": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DEBUG_DISPLAY",
    "Meta Arrays": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs PCIe DMA write and read-back operations across all four DMA channels (0-3). It begins by writing 0x0 to address 0xE6004100 to initialize the synchronization register. Link training is performed conditionally for DM0 or DM1 in RC or EP mode via link_training_dm0_x4(4) or link_training_dm1_x4(4). Under DM0_RC, the test polls read_sii0_reg(0xC0) with mask 0xD1 until link-up is confirmed. Under DM1_RC, it polls read_sii1_reg(0xC0) similarly. The Vendor ID is read from read_pcie_slv0_reg(0x0) (DM0_RC) or read_pcie_slv1_reg(0x0) (DM1_RC), and 0x7 is written to offset 0x4 to enable bus master and memory space. BAR programming and memory base programming are performed via bar_program_dm0_x4()/mem_base_program_dm0_x4() or bar_program_dm1_x4()/mem_base_program_dm1_x4(). Non-secure protection is configured via non_secure_prot_nic(). The test polls read_reg(0xE6004100) until value equals 0x12345678. Source memory at src_addr0 (0xE6000000) is preloaded with 128 words of 0xC0DEBEED and 128 words of 0xF00DDEAF. GIC is set up via GIC_Set() and GIC_EnableAllIRQ(). DMA write interrupt mask is cleared by writing 0x0 to mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF (or PCIE1 equivalents for DM1_RC). For each of channels 0-3, DMA write channel is programmed via program_dma_wch0/1/2/3() with source, destination, and length (0x40), then the DMA write doorbell register mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF is written with the channel number. The test waits in a while(int_pend) loop for the interrupt handler to clear int_pend. After all write channels complete, DMA read channels 0-3 are programmed via program_dma_rch0/1/2/3() and triggered via mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF with the channel number, again waiting for interrupt completion. The Default_IRQHandler reads mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF, masks with 0x0000000F, clears interrupts by writing the status to mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF, and calls GIC_ClearIRQ(0x20) for DM0 or GIC_ClearIRQ(0x23) for DM1. The test calls finish(0) upon completion.",
    "Test Description": "This test validates PCIe DMA write and read-back functionality across all four DMA channels. After link training and link-up confirmation, the test reads the device Vendor ID, enables bus master and memory space access, programs BARs and memory base addresses, and preloads source memory with known data patterns. DMA write interrupt masks are cleared for both write and read engines. Each of the four DMA write channels is sequentially programmed with source address, destination address, and transfer length, then triggered via the DMA_WRITE_DOORBELL_OFF register. The test waits for an interrupt-driven completion signal after each channel transfer. After all four write channels complete, the test performs DMA read-back on all four channels using the DMA_READ_DOORBELL_OFF register, again waiting for interrupt-driven completion per channel. The interrupt handler reads the DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF registers, clears the interrupts via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF, and acknowledges the GIC interrupt. The test finishes successfully after all DMA write and read-back operations complete.",
    "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100 to initialize synchronization register. 2. Conditionally call link_training_dm0_x4(4) or link_training_dm1_x4(4) based on DM0_RC, DM1_RC, DM0_EP, DM1_EP defines. 3. Under DM0_RC: Read read_sii0_reg(0xC0) into data_rd. Poll in while loop until (data_rd & 0xD1) == 0xD1. 4. Under DM0_RC: Read read_pcie_slv0_reg(0x0) to get Vendor ID. Write write_pcie_slv0_reg(0x4, 0x7). Call bar_program_dm0_x4(). wait_on(10). Call mem_base_program_dm0_x4(). 5. Under DM1_RC: Read read_sii1_reg(0xC0) into data_rd. Poll in while loop until (data_rd & 0xD1) == 0xD1. 6. Under DM1_RC: Read read_pcie_slv1_reg(0x0) to get Vendor ID. Write write_pcie_slv1_reg(0x4, 0x7). Call bar_program_dm1_x4(). wait_on(10). Call mem_base_program_dm1_x4(). 7. Call non_secure_prot_nic(). 8. Poll read_reg(0xE6004100) in while loop with wait_on(5) until value equals 0x12345678. 9. Set len=0x40, src_addr0=0xE6000000. Set wr_addr0/rd_addr0 and other channel addresses based on DM0_RC (0xA7xxxxxx) or DM1_RC (0xC7xxxxxx). Set dst_addr0=0xE6001000, dst_addr1/2/3=0xE6020000. 10. Preload source memory: for i=0..127, write_reg(src_addr0 + 4*i, 0xC0DEBEED). For i=0..127, write_reg((src_addr0+400) + 4*i, 0xF00DDEAF). 11. Set int_pend=1. Call GIC_Set(). Call GIC_EnableAllIRQ(). 12. Under DM0_RC: Write 0x0 to mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF. Write 0x0 to mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF. 13-22. Program and trigger DMA write/read channels 0-3. 23-24. IRQ handler clears interrupts.",
    "Test Steps / Procedure": "1. Initialize the synchronization register by clearing it. 2. Perform PCIe link training for the configured dual-mode controller in Root Complex or Endpoint mode with x4 lane width. 3. Poll the gic register on the appropriate SII interface until link status bits confirm link-up. 4. Read the TYPE1_DEV_ID_VEND_ID_REG to retrieve the Vendor ID of the connected device. 5. Write to TYPE1_STATUS_COMMAND_REG to enable bus master, memory space, and I/O space access. 6. Program BARs and memory base addresses for the active controller. 7. Configure non-secure protection via NIC programming. 8. Poll the synchronization register until the expected handshake value is received. 9. Preload source memory with known data patterns for DMA transfer verification. 10. Initialize the GIC and enable all IRQs for interrupt-driven DMA completion. 11. Clear DMA interrupt masks by writing to DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF registers. 12. For each DMA write channel (0 through 3): program the channel with source address, destination address, and transfer length, then trigger the transfer by writing the channel number to DMA_WRITE_DOORBELL_OFF and wait for interrupt-driven completion. 13. For each DMA read channel (0 through 3): program the channel with remote source address, local destination address, and transfer length, then trigger the transfer by writing the channel number to DMA_READ_DOORBELL_OFF and wait for interrupt-driven completion. 14. Verify that the interrupt handler reads DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF, clears interrupts via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF, and acknowledges the GIC interrupt. 15. Confirm all DMA transfers complete successfully and call finish.",
    "Meta Impacted Registers": "0xE6004100; 0xC0; 0x0; 0x4; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_CLEAR_OFF",
    "Impacted Registers": "gic; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; DMA_WRITE_INT_MASK_OFF; DMA_READ_INT_MASK_OFF; DMA_WRITE_DOORBELL_OFF; DMA_READ_DOORBELL_OFF; DMA_WRITE_INT_STATUS_OFF; DMA_READ_INT_STATUS_OFF; DMA_WRITE_INT_CLEAR_OFF; DMA_READ_INT_CLEAR_OFF",
    "Meta Validation / Acceptance Criteria": "1. Poll link status. 2. Verify Vendor ID. 3. Poll sync register. 4-5. DMA write/read channel completion via interrupt. 6. IRQ handler validates status registers. 7. finish(0) after all 8 DMA transfers.",
    "Validation / Acceptance Criteria": "1. The gic register on the appropriate SII interface must indicate link-up by having the expected link status bit pattern set. 2. The TYPE1_DEV_ID_VEND_ID_REG must return a valid Vendor ID confirming device presence. 3. The synchronization register must return the expected handshake value before DMA operations begin. 4. Each DMA write channel transfer must complete successfully, confirmed by an interrupt where the DMA_WRITE_INT_STATUS_OFF register shows a non-zero channel status in the lower 4 bits. 5. Each DMA read channel transfer must complete successfully, confirmed by an interrupt where the DMA_READ_INT_STATUS_OFF register shows a non-zero channel status in the lower 4 bits. 6. All DMA interrupts must be properly cleared via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF registers, and the GIC interrupt must be acknowledged. 7. The test passes by calling finish(0) after all four write and four read DMA channel transfers complete without errors.",
    "Remarks": "The test uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select the link training and DMA programming paths. DM0_RC uses PCIE0 DMA registers and GIC IRQ 0x20, while DM1_RC uses PCIE1 DMA registers and GIC IRQ 0x23. Source memory is preloaded with two distinct data patterns (0xC0DEBEED and 0xF00DDEAF) at different offsets for transfer verification. The transfer length is 0x40 bytes per channel. The test relies on interrupt-driven completion using the int_pend flag toggled by Default_IRQHandler. Multiple wait_on() calls are used for timing synchronization between channel operations. The synchronization register at the absolute address could not be mapped to a canonical register name. All PCIE1 DMA macros could not be resolved to offset values but were matched to the same canonical register names as their PCIE0 equivalents."
  },
  {
    "Index": "3",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_mem_wr_rd_test",
    "Feature": "Memory Write Read",
    "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
    "Meta Macros": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DM0; DM1; DEBUG_DISPLAY",
    "Meta Arrays": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs PCIe memory write and read-back verification through the PCIe slave interfaces.",
    "Test Description": "This test validates PCIe memory write and read-back operations through the PCIe slave interfaces. After link training and link-up confirmation by polling the gic register, the test reads the device Vendor ID from the TYPE1_DEV_ID_VEND_ID_REG, enables bus master and memory space access via the TYPE1_STATUS_COMMAND_REG, programs BARs and memory base addresses, and configures non-secure protection. Cache coherency is first enabled then disabled through the COHERENCY_CONTROL_3_OFF register for both PCIE0 and PCIE1 controllers. The test then performs memory write-read verification at multiple memory offsets through the PCIe slave port, writing known data patterns and reading them back to verify data integrity. The test supports Root Complex and Endpoint modes for both DM0 and DM1 controllers. Finally, the test polls a synchronization register until the expected completion value is received, confirming successful memory access.",
    "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100. 2. Link training. 3-8. Cache programming. 9-14. Link-up polling and BAR/mem programming. 15-18. Cache disable. 19-22. Memory write-read operations. 23-24. Poll sync and finish.",
    "Test Steps / Procedure": "1. Initialize the synchronization register by clearing it. 2. Perform PCIe link training for the configured dual-mode controller (DM0 or DM1) in either Root Complex or Endpoint mode with x4 lane width. 3. Program the COHERENCY_CONTROL_3_OFF register for both PCIE0 and PCIE1 controllers to enable cache coherency by setting specific bit fields. 4. Wait for the configuration to take effect. 5. Re-apply cache coherency settings to both PCIE0 and PCIE1 COHERENCY_CONTROL_3_OFF registers with all relevant bit fields enabled. 6. Poll the gic register on the appropriate SII interface until link status bits indicate link-up. 7. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG to retrieve the Vendor ID of the connected device. 8. Write to TYPE1_STATUS_COMMAND_REG to enable bus master, memory space, and I/O space access. 9. Program BARs and memory base addresses for the active controller and mode. 10. Configure non-secure protection via NIC programming. 11. Update the synchronization register with a handshake value. 12. Disable cache coherency by progressively clearing the upper bit fields in COHERENCY_CONTROL_3_OFF for both PCIE0 and PCIE1 controllers. 13. Wait for cache disable to take effect. 14. Perform memory write-read verification at multiple memory offsets through the PCIe slave port, writing known data patterns and reading them back. 15. Poll the synchronization register until the expected completion value is received, confirming successful memory access. 16. Finish the test upon successful completion.",
    "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4",
    "Impacted Registers": "COHERENCY_CONTROL_3_OFF; gic; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG",
    "Meta Validation / Acceptance Criteria": "1. Poll link status. 2. Verify Vendor ID. 3. Memory write-read verification. 4. Final polling. 5. finish(0).",
    "Validation / Acceptance Criteria": "1. The gic register on the appropriate SII interface must indicate link-up by having the expected link status bit pattern set. 2. The TYPE1_DEV_ID_VEND_ID_REG must return a valid Vendor ID confirming device presence. 3. All memory write-read operations through the PCIe slave port must complete successfully, with read-back data matching the written data patterns at each tested memory offset. 4. The synchronization register must return the expected completion value before the test finishes. 5. The test passes by calling finish(0) after all memory write-read verifications and polling conditions are met.",
    "Remarks": "The test uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select the link training mode and memory write-read paths. Additional conditional defines DM0 and DM1 control the link-up polling path. Under DM0_EP, a long wait_on(30000) is used before proceeding. Different data patterns are used for RC mode (controller-specific patterns) versus EP mode (uniform pattern). The pcie_slv0_mem_wr_rd and pcie_slv1_mem_wr_rd helper functions perform the actual memory write and read-back verification. Cache coherency is first enabled then explicitly disabled before the memory operations. The synchronization register is written with an intermediate handshake value before cache disable, and the final polling waits for a different completion value. The synchronization register at the absolute address and the PCIE1 coherency control macro could not be mapped to canonical register names."
  },
  {
    "Index": "4",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_reg_wr_rd_test",
    "Feature": "Register Write Read",
    "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; <pcie.h>",
    "Meta Macros": "NA",
    "Meta Arrays": "rc0_ctl_addr[5]; rc1_ctl_addr[5]; ctl_default[5]; sii0_addr[3]; sii1_addr[3]; sii_default[3]; sii0_write_mask[3]; sii1_write_mask[3]; phy0_addr[3]; phy1_addr[3]; phy0_default[3]; phy1_default[3]; phy0_write_mask[3]; phy1_write_mask[3]; chk_val[6]; chk_val_phy[3]",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs register reset-value verification and register write-read verification for PCIe controller registers across three register groups.",
    "Test Description": "This test validates register reset default values and register write-read integrity for PCIe controller registers across three register groups: DBI DSP controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) for both PCIE0 and PCIE1 instances, SII registers (transmit header and PHY control registers) for both PCIE0 and PCIE1 instances, and PHY registers for both PHY0 and PHY1 instances. The test first reads all registers and verifies they contain their expected reset default values. Then it performs write-read verification by writing multiple data patterns to each register, reading back, and comparing the read data against the expected written value, accounting for register-specific write masks. The test passes only if all reset-value checks and all write-read checks succeed with zero errors.",
    "Meta Test Steps / Procedure": "1. Declare global arrays. 2. test_case() calls chk_rst_val() then chk_rd_wr(). 3-9. Reset value checks. 10-21. Write-read checks. 22. finish(err2 || err1).",
    "Test Steps / Procedure": "1. Read all five DBI DSP controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on PCIE0 and verify each contains its expected reset default value. 2. Read the same five DBI DSP controller registers on PCIE1 and verify reset default values. 3. Read all three SII registers (transmit header and PHY control registers) on PCIE0 and verify reset default values. 4. Read all three SII registers on PCIE1 and verify reset default values. 5. Release PHY reset by writing the reset control value to the PHY reset control register for both PCIE0 and PCIE1. 6. Read all three PHY registers on PHY0 and PHY1, extract the relevant 16-bit field based on address alignment, and verify reset default values. 7. For each of three data patterns (all-ones, alternating-bits, inverted alternating-bits): write the pattern to all five DBI DSP controller registers on both PCIE0 and PCIE1, write the masked pattern to all three SII registers on both PCIE0 and PCIE1, re-apply PHY reset, and write the PHY-specific pattern to all three PHY registers on both PHY0 and PHY1. 8. Read back all DBI DSP controller registers on both PCIE0 and PCIE1 and verify the read data matches the written pattern. 9. Read back all SII registers on both PCIE0 and PCIE1 and verify the read data matches the masked written pattern. 10. Read back all PHY registers on both PHY0 and PHY1, extract the relevant 16-bit field, apply the write mask, and verify the read data matches the expected masked PHY pattern. 11. Confirm the test passes with zero accumulated errors across all reset-value and write-read checks.",
    "Meta Impacted Registers": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8",
    "Impacted Registers": "MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
    "Meta Validation / Acceptance Criteria": "1. Reset value check. 2. Write-read check. 3. Final pass/fail: finish(err2 || err1).",
    "Validation / Acceptance Criteria": "1. All DBI DSP controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on both PCIE0 and PCIE1 must read their expected reset default values. 2. All SII registers on both PCIE0 and PCIE1 must read their expected reset default values. 3. All PHY registers on both PHY0 and PHY1 must read their expected reset default values after PHY reset is released. 4. For each write-read data pattern, all DBI DSP controller registers on both PCIE0 and PCIE1 must read back the exact written value. 5. For each write-read data pattern, all SII registers on both PCIE0 and PCIE1 must read back the written value masked with the register-specific write mask. 6. For each write-read data pattern, all PHY registers on both PHY0 and PHY1 must read back the written PHY value within the valid write mask bits. 7. The test passes with zero accumulated errors across all checks, confirmed by finish being called with a zero argument.",
    "Remarks": "The test covers three distinct register groups across two PCIe controller instances (PCIE0 and PCIE1): DBI DSP controller registers, SII registers, and PHY registers. SII registers have per-register write masks (two fully writable, one partially writable with mask 0xF000F). PHY registers use 16-bit access with alignment-based extraction (lower 16 bits if address is 4-byte aligned, upper 16 bits otherwise) and a 13-bit write mask (0x1FFF). The write-read loop iterates over only the first three of six defined check values. PHY reset is re-applied before each PHY register access cycle. The SII macros and PHY reset control macros could not be resolved to canonical register names. The PHY registers at absolute addresses could not be mapped to canonical register names."
  }
]

# TestPlan columns
tp_cols = ["Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
           "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
           "Test Steps / Procedure", "Impacted Registers", "Validation / Acceptance Criteria",
           "Code Generation"]

# MetaData columns
md_cols = ["Index", "Test Case Name", "Meta Test Description", "Meta Test Steps / Procedure",
           "Meta Impacted Registers", "Meta Validation / Acceptance Criteria",
           "Meta Headers", "Meta Macros", "Meta Arrays"]

# Create workbook
wb = Workbook()

# TestPlan sheet
ws_tp = wb.active
ws_tp.title = "TestPlan"

# MetaData sheet
ws_md = wb.create_sheet("MetaData")

# Formatting
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")

def write_sheet(ws, columns, data):
    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align
    # Auto-size columns
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
        adjusted_width = min(max_len + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    # Freeze first row
    ws.freeze_panes = "A2"

write_sheet(ws_tp, tp_cols, json_data)
write_sheet(ws_md, md_cols, json_data)

# Set MetaData sheet to veryHidden
ws_md.sheet_state = "veryHidden"

# Save
output_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(output_dir, filename)
wb.save(output_path)

# Validate
assert os.path.exists(output_path), f"File not found: {output_path}"
assert os.path.getsize(output_path) > 0, "File is empty"
wb2 = load_workbook(output_path)
assert "TestPlan" in wb2.sheetnames, "TestPlan sheet missing"
assert "MetaData" in wb2.sheetnames, "MetaData sheet missing"
assert wb2["MetaData"].sheet_state == "veryHidden", "MetaData not veryHidden"

print(f"SUCCESS: Generated {filename}")
print(f"Path: {output_path}")
print(f"Size: {os.path.getsize(output_path)} bytes")
print(f"TestPlan rows: {ws_tp.max_row - 1}")
print(f"MetaData rows: {ws_md.max_row - 1}")
print(f"Validation: PASSED")
