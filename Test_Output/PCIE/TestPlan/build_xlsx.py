#!/usr/bin/env python3
"""Minimal inline XLSX builder - outputs base64 of the workbook to stdout."""
import sys, io, base64, os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    R=True
except: R=False

IST=timezone(timedelta(hours=5,minutes=30))
TS=datetime.now(IST).strftime("%Y%m%d_%H%M%S")
FN=f"PCIE_TestPlan_{TS}.xlsx"

S={"pcie_device_enumerate_test":{"Initialization":["Clear the synchronization register to prepare for the test."],"Configuration":["Initiate PCIe link training for the configured dual-mode controller at Gen4 speed.","Enable cache coherency by performing read-modify-write on COHERENCY_CONTROL_3_OFF for both PCIe controller instances, setting the required bit fields.","Write to TYPE1_STATUS_COMMAND_REG to enable bus master, memory space, and I/O space access.","Program memory base addresses for both dual-mode controllers.","Configure system-level control registers to enable required functionality.","Disable cache coherency by performing read-modify-write on COHERENCY_CONTROL_3_OFF for both PCIe controller instances, clearing the required bit fields."],"Execution":["Wait for the configuration to take effect.","Poll the link status register on both SII interfaces until the expected link-up condition is detected.","Read the TYPE1_DEV_ID_VEND_ID_REG to retrieve the Vendor ID of the enumerated device.","Enumerate BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports by writing all-ones, reading back to determine BAR size, then programming actual base addresses.","Read back all BAR registers to verify the programmed values.","Poll the synchronization register until the expected completion handshake value is received.","End the test with a pass indication."]},"pcie_reg_wr_rd_test":{"Initialization":["Initialize error counters and register address arrays for both PCIe controller instances, SII interfaces, and PHY registers."],"Configuration":["Write to the PHY reset control registers on both PCIe controller instances to release PHY from reset.","Write the PHY reset control registers again on both controller instances."],"Execution":["Read all DBI control registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on both PCIe controller instances and verify they match their expected reset default values.","Read all SII registers (transmit header and PHY control registers) on both SII interfaces and verify they match their expected reset default values.","Read PHY lane registers on both controller instances with 16-bit extraction logic and verify they match their expected reset default values.","Begin the write-read-compare test loop using multiple data patterns.","For each data pattern, write the pattern to all DBI control registers on both PCIe controller instances.","Write the masked pattern to all SII registers on both SII interfaces using the appropriate write masks.","Write the PHY-specific pattern to all PHY lane registers on both controller instances using the appropriate write masks.","Read back all DBI control registers on both controller instances and verify the read values match the written pattern.","Read back all SII registers on both SII interfaces and verify the read values match the masked written pattern.","Read back all PHY lane registers on both controller instances with 16-bit extraction and mask logic, and verify the read values match the expected masked PHY pattern.","Report pass if all reset default checks and write-read-compare checks pass across all registers and all patterns; report fail if any mismatch is detected."]}}

def ps(n):
    d=S.get(n,{});l=[]
    for c in["Initialization","Configuration","Execution","Interrupt"]:
        if c in d and d[c]:
            if l:l.append("")
            l.append(f"{c}:")
            for i,s in enumerate(d[c],1):l.append(f"{i}. {s}")
    return"\n".join(l)

def rs(n):
    if not R:return ps(n)
    d=S.get(n,{});bf=InlineFont(b=True,sz=11);nf=InlineFont(b=False,sz=11);p=[];f=True
    for c in["Initialization","Configuration","Execution","Interrupt"]:
        if c in d and d[c]:
            if not f:p.append(TextBlock(nf,"\n\n"))
            p.append(TextBlock(bf,f"{c}:\n"))
            for i,s in enumerate(d[c],1):
                p.append(TextBlock(nf,f"{i}. {s}"))
                if i<len(d[c]):p.append(TextBlock(nf,"\n"))
            f=False
    return CellRichText(*p) if p else""

D=[{"Index":"1","SS / Module":"PCIE","Test Case Name":"pcie_device_enumerate_test","Feature":"PCIe Device Enumeration","Test Description":"Verifies PCIe device enumeration by performing link training, configuring cache coherency control registers (COHERENCY_CONTROL_3_OFF) for both PCIe controllers, polling link status until the expected link-up condition is detected, reading the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enabling bus master and memory/IO space access via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, and enumerating Base Address Registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both PCIe slave ports by writing all-ones to determine BAR size and then programming actual base addresses. The test concludes by polling a synchronization register for a completion handshake value.","Impacted Registers":"COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG","Validation / Acceptance Criteria":"The test passes when: (1) PCIe link training completes successfully for the configured controller mode. (2) The link status polling on both SII interfaces returns the expected link-up condition. (3) The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID. (4) BAR enumeration on both slave ports completes successfully with all BAR registers accepting the programmed base addresses. (5) The synchronization register returns the expected completion handshake value. The test ends with finish(0) indicating success.","Remarks":"The testcase uses conditional compilation to support multiple PCIe modes (DM0_RC, DM1_RC, DM0_EP, DM1_EP). Link training speed is set to Gen4. Cache coherency programming is performed in multiple phases with waits between them. BAR enumeration follows the standard PCIe enumeration procedure. Several system-level control registers could not be mapped to known PCIe DBI register names. The COHERENCY_CONTROL_3_OFF register is accessed on both PCIe controller instances (PCIE0 and PCIE1).","Meta Test Description":"This testcase performs PCIe device enumeration across two dual-mode controllers (DM0 and DM1). It begins by writing 0x0 to 0xE6004100, then initiates link training via conditional compilation. Cache programming is performed by read-modify-write of mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. The test polls link status, reads Vendor ID, enables bus master, programs memory bases, performs BAR enumeration, and polls for completion.","Meta Test Steps / Procedure":"1. Initialize global variables. 2. Write 0x0 to 0xE6004100. 3. Perform link training. 4-6. Cache programming for both controllers. 7. Wait. 8-9. Program coherency registers again. 10-11. Poll link status. 12-14. Read Vendor ID, enable bus master, program memory bases. 15-16. Wait and write system registers. 17-21. Disable cache. 22-27. BAR enumeration. 28-30. Poll and finish.","Meta Impacted Registers":"0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24"},{"Index":"2","SS / Module":"PCIE","Test Case Name":"pcie_reg_wr_rd_test","Feature":"PCIe Register Read Write Verification","Test Description":"Verifies the reset default values and read-write integrity of PCIe DBI, SII, and PHY registers across both PCIe controller instances. The test first reads all target registers and compares them against their expected reset default values. It then writes the PHY reset control registers to release PHY from reset. After verifying reset defaults, the test performs a write-read-compare cycle using multiple data patterns on MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF, SII transmit header registers, SII PHY control registers, and PHY lane registers. Write masks are applied before writing. Read-back values are compared against expected values with appropriate masking. The test reports pass if all comparisons match.","Impacted Registers":"MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF","Validation / Acceptance Criteria":"The test passes when: (1) All DBI control registers on both PCIe controller instances read back their expected reset default values of zero. (2) All SII registers read back expected defaults. (3) All PHY lane registers read back expected defaults after 16-bit extraction. (4) For each write-read-compare data pattern, all registers read back the expected values with appropriate masking. The test fails if any mismatch is detected.","Remarks":"The testcase covers both PCIe controller instances (PCIE0 and PCIE1) with identical register sets. Three data patterns are used for write-read-compare testing. Write masks are applied to SII and PHY registers. PHY register reads use 16-bit extraction logic. The PHY reset control registers are written before PHY register access.","Meta Test Description":"This testcase verifies the reset default values and read-write functionality of PCIe registers across both PCIe controller instances (PCIE0 and PCIE1), SII interfaces, and PHY registers.","Meta Test Steps / Procedure":"1. Declare global variables. 2-11. Initialize register address arrays, default values, and write masks. 12. Call chk_rst_val(). 13-20. Read and verify reset defaults. 21. Call chk_rd_wr(). 22-37. Write-read-compare loop with 3 patterns. 38. Call finish(err2 || err1).","Meta Impacted Registers":"mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8"}]

wb=Workbook();ws=wb.active;ws.title="TestPlan";wm=wb.create_sheet("MetaData")
TC=["Index","SS / Module","Feature","Test Case Name","Test Description","Speed","Mode","Memory Start Offset","Memory End Offset","Remarks","Test Steps / Procedure","Impacted Registers","Validation / Acceptance Criteria","Code Generation"]
MC=["Index","Test Case Name","Meta Test Description","Meta Test Steps / Procedure","Meta Impacted Registers","Meta Validation / Acceptance Criteria","Meta Headers","Meta Macros","Meta Arrays"]
hf=Font(name="Calibri",size=11,bold=True,color="FFFFFF");hl=PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
ha=Alignment(horizontal="center",vertical="top",wrap_text=True);ca=Alignment(vertical="top",wrap_text=True)
bd=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
for i,c in enumerate(TC,1):
    x=ws.cell(row=1,column=i,value=c);x.font=hf;x.fill=hl;x.alignment=ha;x.border=bd
for i,c in enumerate(MC,1):
    x=wm.cell(row=1,column=i,value=c);x.font=hf;x.fill=hl;x.alignment=ha;x.border=bd
for ri,tc in enumerate(D,2):
    n=tc["Test Case Name"]
    v=[tc.get("Index",""),tc.get("SS / Module",""),tc.get("Feature",""),n,tc.get("Test Description",""),"","","","",tc.get("Remarks",""),None,tc.get("Impacted Registers",""),tc.get("Validation / Acceptance Criteria",""),""]
    for ci,vl in enumerate(v,1):
        if ci==11:continue
        x=ws.cell(row=ri,column=ci,value=vl);x.alignment=ca;x.border=bd
    sc=ws.cell(row=ri,column=11)
    try:sc.value=rs(n)
    except:sc.value=ps(n)
    sc.alignment=ca;sc.border=bd
    mv=[tc.get("Index",""),n,tc.get("Meta Test Description",""),tc.get("Meta Test Steps / Procedure",""),tc.get("Meta Impacted Registers",""),tc.get("Validation / Acceptance Criteria",""),"","",""]
    for ci,vl in enumerate(mv,1):
        x=wm.cell(row=ri,column=ci,value=vl);x.alignment=ca;x.border=bd
dv=DataValidation(type="list",formula1='"Required,Not Required"',allow_blank=True,showDropDown=False)
dv.sqref=f"{get_column_letter(14)}2:{get_column_letter(14)}1000";ws.add_data_validation(dv)
tw={"Index":8,"SS / Module":15,"Feature":30,"Test Case Name":35,"Test Description":60,"Speed":10,"Mode":10,"Memory Start Offset":20,"Memory End Offset":20,"Remarks":50,"Test Steps / Procedure":90,"Impacted Registers":60,"Validation / Acceptance Criteria":70,"Code Generation":18}
for i,c in enumerate(TC,1):ws.column_dimensions[get_column_letter(i)].width=tw.get(c,20)
mw={"Index":8,"Test Case Name":35,"Meta Test Description":80,"Meta Test Steps / Procedure":80,"Meta Impacted Registers":80,"Meta Validation / Acceptance Criteria":70,"Meta Headers":30,"Meta Macros":30,"Meta Arrays":30}
for i,c in enumerate(MC,1):wm.column_dimensions[get_column_letter(i)].width=mw.get(c,20)
for r in range(2,len(D)+2):ws.row_dimensions[r].height=250;wm.row_dimensions[r].height=200
ws.freeze_panes="A2";wm.freeze_panes="A2";wm.sheet_state="veryHidden";wb.active=0

buf=io.BytesIO();wb.save(buf);buf.seek(0);raw=buf.read()
out=os.path.join(os.getcwd(),FN);
with open(out,"wb") as f:f.write(raw)
sz=len(raw)
wb2=load_workbook(out)
ok="TestPlan" in wb2.sheetnames and "MetaData" in wb2.sheetnames and wb2["MetaData"].sheet_state=="veryHidden" and sz>5000
tp=wb2["TestPlan"].max_row-1;md=wb2["MetaData"].max_row-1;wb2.close()
print(f"FILENAME={FN}");print(f"SIZE={sz}");print(f"TP={tp}");print(f"MD={md}");print(f"OK={ok}")
