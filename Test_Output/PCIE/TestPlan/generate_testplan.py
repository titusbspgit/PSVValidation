#!/usr/bin/env python3
"""
Auto-generated TestPlan Excel Generator for PCIE IP.
Execute this script to generate the PCIE_TestPlan Excel workbook.
Requirements: pip install openpyxl
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
import os

def generate_testplan():
    IST = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(IST)
    filename = f"PCIE_TestPlan_{now_ist.strftime('%Y%m%d')}_{now_ist.strftime('%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()

    # ---- TestPlan Sheet ----
    ws_tp = wb.active
    ws_tp.title = "TestPlan"

    tp_columns = [
        "Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
        "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
        "Test Steps / Procedure", "Impacted Registers",
        "Validation / Acceptance Criteria", "Code Generation"
    ]

    tp_data = [
        {
            "Index": "1",
            "SS / Module": "PCIE",
            "Feature": "PCIe Device Enumeration and BAR Configuration",
            "Test Case Name": "pcie_device_enumerate_test",
            "Test Description": "Verifies PCIe device enumeration and BAR configuration across two dual-mode controllers. The test initiates link training in x4 mode, programs cache coherency settings via the COHERENCY_CONTROL_3_OFF register for both PCIe controllers, polls SII link status registers until the link is established, reads the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enables bus master and memory/IO space access via TYPE1_STATUS_COMMAND_REG, programs memory base addresses, configures system-level control registers, disables cache coherency, enumerates all six BAR registers (BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports by writing all-ones to determine BAR sizes and then programming final base addresses, reads back BAR values for verification, and polls a completion status register until the expected completion value is observed.",
            "Speed": "",
            "Mode": "",
            "Memory Start Offset": "",
            "Memory End Offset": "",
            "Remarks": "The testcase uses compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select the link training mode, so behavior varies based on build configuration. Cache coherency programming is performed in multiple phases with waits between them, followed by a disable phase. The SII link status polling uses a bitmask condition and may loop indefinitely if the link does not come up. Several system-level control register addresses and the SII register space could not be mapped to named registers in the provided specification. The completion status register is polled in an unbounded loop with periodic waits. The BAR enumeration covers offsets that map to Type 1 header configuration space registers on both PCIe slave ports.",
            "Test Steps / Procedure": "1. Initialize the PCIe test environment and clear the completion status register. 2. Initiate PCIe link training in x4 mode for the configured dual-mode controller. 3. Program cache coherency control settings in the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances by enabling the relevant bit fields. 4. Wait for the cache programming to take effect. 5. Perform a consolidated cache coherency programming pass for both controllers. 6. Poll the SII0 link status register until the link-up condition is detected (specific status bits are set). 7. Configure non-secure protection via NIC settings. 8. Poll the SII1 link status register until the link-up condition is detected. 9. Read the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG on slave port 0. 10. Enable bus master, memory space, and I/O space access by writing to TYPE1_STATUS_COMMAND_REG on slave port 0. 11. Program memory base addresses for both dual-mode controllers. 12. Enable system-level control registers for PCIe operation. 13. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF for both controllers. 14. Wait for cache disable to take effect. 15. Enumerate BAR registers on slave port 1: write all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG to determine BAR sizes, then program final base addresses and read back for verification. 16. Repeat BAR enumeration on slave port 0. 17. Poll the completion status register until the expected completion value is observed. 18. Report test completion.",
            "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
            "Validation / Acceptance Criteria": "The test passes when: (1) PCIe link training completes successfully for the configured dual-mode controller in x4 mode. (2) SII0 and SII1 link status registers report the expected link-up condition with the required status bits set. (3) The Vendor ID is successfully read from TYPE1_DEV_ID_VEND_ID_REG. (4) TYPE1_STATUS_COMMAND_REG is successfully written to enable bus master, memory space, and I/O space. (5) All six BAR registers on both slave ports accept the all-ones write pattern and return valid BAR size information on read-back. (6) Final BAR base address values are successfully programmed and verified by read-back on both slave ports. (7) The completion status register returns the expected completion value during polling. (8) The test completes by calling finish with a success code.",
            "Code Generation": ""
        }
    ]

    # ---- MetaData Sheet ----
    ws_md = wb.create_sheet("MetaData")

    md_columns = [
        "Index", "Test Case Name", "Meta Test Description",
        "Meta Test Steps / Procedure", "Meta Impacted Registers",
        "Meta Validation / Acceptance Criteria", "Meta Headers",
        "Meta Macros", "Meta Arrays"
    ]

    md_data = [
        {
            "Index": "1",
            "Test Case Name": "pcie_device_enumerate_test",
            "Meta Test Description": "This testcase performs PCIe device enumeration across two dual-mode controllers (DM0 and DM1). It begins by writing 0x0 to 0xE6004100 and initiating link training in x4 mode based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). It then performs cache programming by executing read-modify-write sequences on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF using set_data() to configure bit fields [11:14], [3:6], [27:30], and [19:22] with value 0xF. After a wait_on(20), it repeats the cache programming with all four bit fields set in a single read-modify-write pass. It then polls read_sii0_reg(0xC0) until bits matching mask 0xD1 equal 0xD1, calls non_secure_prot_nic(), and polls read_sii1_reg(0xC0) with the same 0xD1 condition. Under DM0_RC, it reads the Vendor ID via read_pcie_slv0_reg(0x0), writes 0x7 to write_pcie_slv0_reg(0x4) to enable bus master, memory space, and I/O space, then calls mem_base_program_dm0_x4() and mem_base_program_dm1_x4() followed by wait_on(10). It writes 0x1 to six system-level registers (0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034). It then performs cache disable programming by repeating read-modify-write on both coherency control registers with bit fields [19:22] and [27:30] set to 0x0. After wait_on(30), it enumerates BAR registers on both PCIe slave ports (slv0 and slv1) by writing 0xFFFFFFFF to offsets 0x10 through 0x24, reading them back to determine BAR sizes, then programming final BAR values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) and reading them back for verification. Finally, it polls 0xE6004100 waiting for value 0x12345678 with wait_on(5) between iterations, and calls finish(0) upon success.",
            "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100. 2. Invoke link_training_dm0_x4(4) or link_training_dm1_x4(4) based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). 3. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, use set_data() to set bits [11:14]=0xF and [3:6]=0xF, write back (cache programming). 4. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF again, set bits [27:30]=0xF and [19:22]=0xF, write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. Call wait_on(20). 7. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set all four bit fields [11:14], [3:6], [27:30], [19:22] to 0xF in one pass, write back. 8. Read mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, set all four bit fields to 0xF, write back. 9. Read read_sii0_reg(0xC0), poll in while loop until (data_rd & 0xD1) == 0xD1. 10. Call non_secure_prot_nic(). 11. Read read_sii1_reg(0xC0), poll in while loop until (data_rd & 0xD1) == 0xD1. 12. Under DM0_RC: read Vendor ID via read_pcie_slv0_reg(0x0), print it. 13. Write 0x7 to write_pcie_slv0_reg(0x4) to enable bus master, memory space, and I/O space. 14. Call mem_base_program_dm0_x4() and mem_base_program_dm1_x4(), then wait_on(10). 15. Write 0x1 to system registers 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034. 16. Perform cache disable programming: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF with bits [19:22]=0x0 and [27:30]=0xF, then [27:30]=0x0 and [19:22]=0x0. 17. Repeat cache disable for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 18. Call wait_on(10), then consolidated cache disable pass for both controllers, then wait_on(30). 19. Write 0xFFFFFFFF to write_pcie_slv1_reg offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24. 20. Read back all six BAR offsets via read_pcie_slv1_reg to determine BAR sizes. 21. Write final BAR values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) to slv1 offsets 0x10-0x24. 22. Read back all six BAR offsets from slv1 for verification. 23. Repeat steps 19-22 for pcie_slv0. 24. Call wait_on(10). 25. Poll read_reg(0xE6004100) in while loop until value equals 0x12345678, with wait_on(5) between iterations. 26. Call finish(0).",
            "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
            "Meta Validation / Acceptance Criteria": "",
            "Meta Headers": "",
            "Meta Macros": "",
            "Meta Arrays": ""
        }
    ]

    # Formatting
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    def populate_sheet(ws, columns, data):
        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment

        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = wrap_alignment

        # Auto-size columns
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(col_name)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), 80))
            adjusted_width = min(max_length + 4, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze first row
        ws.freeze_panes = "A2"

    populate_sheet(ws_tp, tp_columns, tp_data)
    populate_sheet(ws_md, md_columns, md_data)

    # Set MetaData sheet to veryHidden
    ws_md.sheet_state = "veryHidden"

    # Save
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(script_dir, filename)
    wb.save(filepath)
    print(f"Generated: {filepath}")
    print(f"Filename: {filename}")

    # Validate
    wb2 = openpyxl.load_workbook(filepath)
    assert "TestPlan" in wb2.sheetnames
    assert "MetaData" in wb2.sheetnames
    assert os.path.getsize(filepath) > 0
    print("Validation: PASSED")
    return filename

if __name__ == "__main__":
    generate_testplan()
