#!/usr/bin/env python3
"""PCIE TestPlan XLSX Generator
Generates PCIE_TestPlan_20250627_120000.xlsx from the companion CSV file.
Requires: pip install openpyxl
Usage: python PCIE_TestPlan_20250627_120000_generator.py
"""
import csv
import os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

def generate_xlsx():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_file = os.path.join(script_dir, "PCIE_TestPlan_20250627_120000.csv")
    xlsx_file = os.path.join(script_dir, "PCIE_TestPlan_20250627_120000.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "PCIE_TestPlan"

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    cell.alignment = cell_alignment

    # Column widths
    col_widths = [8, 12, 30, 20, 80, 80, 80, 80, 60, 50, 60, 12, 40, 50, 60, 35, 45, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_file)
    print(f"XLSX generated: {xlsx_file}")
    print(f"Total testcases: 4")
    print(f"Testcases: pcie_device_enumerate_test, pcie_dma_write_test, pcie_mem_wr_rd_test, pcie_reg_wr_rd_test")

if __name__ == '__main__':
    generate_xlsx()
