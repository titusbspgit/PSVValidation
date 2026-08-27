#!/usr/bin/env python3
"""Minimal XLSX generator - run with: python3 _run_gen.py"""
import openpyxl, os, sys, base64, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timezone, timedelta

IST = timezone(timedelta(hours=5, minutes=30))
ts = datetime.now(IST).strftime('%Y%m%d_%H%M%S')
fn = f'PCIE_TestPlan_{ts}.xlsx'

wb = openpyxl.Workbook()
ws = wb.active; ws.title = 'TestPlan'
wm = wb.create_sheet('MetaData')

hf=Font(bold=True,color='FFFFFF',size=11)
hfl=PatternFill(start_color='4472C4',end_color='4472C4',fill_type='solid')
ha=Alignment(horizontal='center',vertical='top',wrap_text=True)
ca=Alignment(vertical='top',wrap_text=True)
tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

for i,n in enumerate(['Index','SS / Module','Feature','Test Case Name','Test Description','Speed','Mode','Memory Start Offset','Memory End Offset','Remarks','Test Steps / Procedure','Impacted Registers','Validation / Acceptance Criteria','Code Generation'],1):
    c=ws.cell(row=1,column=i,value=n);c.font=hf;c.fill=hfl;c.alignment=ha;c.border=tb
for i,n in enumerate(['Index','Test Case Name','Meta Test Description','Meta Test Steps / Procedure','Meta Impacted Registers','Meta Validation / Acceptance Criteria','Meta Headers','Meta Macros','Meta Arrays'],1):
    c=wm.cell(row=1,column=i,value=n);c.font=hf;c.fill=hfl;c.alignment=ha;c.border=tb

ws.freeze_panes='A2'; wm.freeze_panes='A2'
dv=DataValidation(type='list',formula1='"Required,Not Required"',allow_blank=True,showDropDown=False)
dv.sqref='N2:N1000'; ws.add_data_validation(dv)
wm.sheet_state='veryHidden'; wb.active=0

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)
b64 = base64.b64encode(buf.read()).decode('ascii')
print(b64[:200])
print(f'...total_b64_len={len(b64)}')
print(f'FILENAME={fn}')
