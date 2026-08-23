#!/usr/bin/env python3
"""PCIE TestPlan Excel Generator - PCIE_TestPlan_20250627_182530
Generates a real .xlsx workbook from the JSON data using openpyxl.
Run: python PCIE_TestPlan_20250627_182530_generator.py
Requires: pip install openpyxl
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_FILE = os.path.join(SCRIPT_DIR, 'PCIE_TestPlan_20250627_182530_data.json')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'PCIE_TestPlan_20250627_182530.xlsx')

TP_COLS = ['Index','SS / Module','Feature','Test Case Name','Test Description',
           'Speed','Mode','Memory Start Offset','Memory End Offset','Remarks',
           'Test Steps / Procedure','Impacted Registers',
           'Validation / Acceptance Criteria','Code Generation']

MD_COLS = ['Index','Test Case Name','Meta Test Description',
           'Meta Test Steps / Procedure','Meta Impacted Registers',
           'Meta Validation / Acceptance Criteria','Meta Headers',
           'Meta Macros','Meta Arrays']

def build():
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    wb = Workbook()
    # --- TestPlan sheet ---
    ws_tp = wb.active
    ws_tp.title = 'TestPlan'
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    wrap = Alignment(wrap_text=True, vertical='top')
    for ci, col_name in enumerate(TP_COLS, 1):
        c = ws_tp.cell(row=1, column=ci, value=col_name)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = wrap
    for ri, row in enumerate(data, 2):
        for ci, col_name in enumerate(TP_COLS, 1):
            val = row.get(col_name, '')
            c = ws_tp.cell(row=ri, column=ci, value=val)
            c.alignment = wrap
    ws_tp.freeze_panes = 'A2'
    # --- MetaData sheet ---
    ws_md = wb.create_sheet('MetaData')
    for ci, col_name in enumerate(MD_COLS, 1):
        c = ws_md.cell(row=1, column=ci, value=col_name)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = wrap
    for ri, row in enumerate(data, 2):
        for ci, col_name in enumerate(MD_COLS, 1):
            val = row.get(col_name, '')
            c = ws_md.cell(row=ri, column=ci, value=val)
            c.alignment = wrap
    ws_md.freeze_panes = 'A2'
    ws_md.sheet_state = 'veryHidden'
    # --- Auto-size columns ---
    for ws in [ws_tp, ws_md]:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    wb.save(OUTPUT_FILE)
    print(f'Workbook saved: {OUTPUT_FILE}')
    print(f'File size: {os.path.getsize(OUTPUT_FILE)} bytes')
    # Validate
    from openpyxl import load_workbook
    vwb = load_workbook(OUTPUT_FILE)
    assert 'TestPlan' in vwb.sheetnames, 'TestPlan sheet missing'
    assert 'MetaData' in vwb.sheetnames, 'MetaData sheet missing'
    assert vwb['TestPlan'].max_row == len(data) + 1
    assert vwb['MetaData'].max_row == len(data) + 1
    assert vwb['MetaData'].sheet_state == 'veryHidden'
    print('Validation PASSED')

if __name__ == '__main__':
    build()
