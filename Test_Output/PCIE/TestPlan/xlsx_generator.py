#!/usr/bin/env python3
"""Generate PCIE TestPlan XLSX. Run: python xlsx_generator.py"""
import os, sys, io, base64
from datetime import datetime, timezone, timedelta
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    RICH = True
except ImportError:
    RICH = False

IST = timezone(timedelta(hours=5, minutes=30))
ts = datetime.now(IST).strftime("%Y%m%d_%H%M%S")
FN = f"PCIE_TestPlan_{ts}.xlsx"

STEPS = {
  "pcie_device_enumerate_test": {
    "Initialization": ["Clear the synchronization register to prepare for the test."],
    "Configuration": [
      "Initiate PCIe link training for the configured dual-mode controller at Gen4 speed.",
      "Enable cache coherency by performing read-modify-write on COHERENCY_CONTROL_3_OFF for both PCIe controller instances, setting the required bit fields.",
      "Write to TYPE1_STATUS_COMMAND_REG to enable bus master, memory space, and I/O space access.",
      "Program memory base addresses for both dual-mode controllers.",
      "Configure system-level control registers to enable required functionality.",
      "Disable cache coherency by performing read-modify-write on COHERENCY_CONTROL_3_OFF for both PCIe controller instances, clearing the required bit fields."
    ],
    "Execution": [
      "Wait for the configuration to take effect.",
      "Poll the link status register on both SII interfaces until the expected link-up condition is detected.",
      "Read the TYPE1_DEV_ID_VEND_ID_REG to retrieve the Vendor ID of the enumerated device.",
      "Enumerate BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports by writing all-ones, reading back to determine BAR size, then programming actual base addresses.",
      "Read back all BAR registers to verify the programmed values.",
      "Poll the synchronization register until the expected completion handshake value is received.",
      "End the test with a pass indication."
    ]
  },
  "pcie_reg_wr_rd_test": {
    "Initialization": ["Initialize error counters and register address arrays for both PCIe controller instances, SII interfaces, and PHY registers."],
    "Configuration": [
      "Write to the PHY reset control registers on both PCIe controller instances to release PHY from reset.",
      "Write the PHY reset control registers again on both controller instances."
    ],
    "Execution": [
      "Read all DBI control registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on both PCIe controller instances and verify they match their expected reset default values.",
      "Read all SII registers (transmit header and PHY control registers) on both SII interfaces and verify they match their expected reset default values.",
      "Read PHY lane registers on both controller instances with 16-bit extraction logic and verify they match their expected reset default values.",
      "Begin the write-read-compare test loop using multiple data patterns.",
      "For each data pattern, write the pattern to all DBI control registers on both PCIe controller instances.",
      "Write the masked pattern to all SII registers on both SII interfaces using the appropriate write masks.",
      "Write the PHY-specific pattern to all PHY lane registers on both controller instances using the appropriate write masks.",
      "Read back all DBI control registers on both controller instances and verify the read values match the written pattern.",
      "Read back all SII registers on both SII interfaces and verify the read values match the masked written pattern.",
      "Read back all PHY lane registers on both controller instances with 16-bit extraction and mask logic, and verify the read values match the expected masked PHY pattern.",
      "Report pass if all reset default checks and write-read-compare checks pass across all registers and all patterns; report fail if any mismatch is detected."
    ]
  }
}

def plain(n):
    sd=STEPS.get(n,{}); lines=[]
    for c in ["Initialization","Configuration","Execution","Interrupt"]:
        if c in sd and sd[c]:
            if lines: lines.append("")
            lines.append(f"{c}:")
            for i,s in enumerate(sd[c],1): lines.append(f"{i}. {s}")
    return "\n".join(lines)

def rich(n):
    if not RICH: return plain(n)
    sd=STEPS.get(n,{}); bf=InlineFont(b=True,sz=11); nf=InlineFont(b=False,sz=11); parts=[]; first=True
    for c in ["Initialization","Configuration","Execution","Interrupt"]:
        if c in sd and sd[c]:
            if not first: parts.append(TextBlock(nf,"\n\n"))
            parts.append(TextBlock(bf,f"{c}:\n"))
            for i,s in enumerate(sd[c],1):
                parts.append(TextBlock(nf,f"{i}. {s}"))
                if i<len(sd[c]): parts.append(TextBlock(nf,"\n"))
            first=False
    return CellRichText(*parts) if parts else ""

DATA=[
  {"Index":"1","SS / Module":"PCIE","Test Case Name":"pcie_device_enumerate_test","Feature":"PCIe Device Enumeration",
   "Test Description":"Verifies PCIe device enumeration by performing link training, configuring cache coherency control registers (COHERENCY_CONTROL_3_OFF) for both PCIe controllers, polling link status until the expected link-up condition is detected, reading the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enabling bus master and memory/IO space access via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, and enumerating Base Address Registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both PCIe slave ports by writing all-ones to determine BAR size and then programming actual base addresses. The test concludes by polling a synchronization register for a completion handshake value.",
   "Impacted Registers":"COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
   "Validation / Acceptance Criteria":"The test passes when: (1) PCIe link training completes successfully for the configured controller mode. (2) The link status polling on both SII interfaces returns the expected link-up condition. (3) The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID. (4) BAR enumeration on both slave ports completes successfully with all BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) accepting the programmed base addresses. (5) The synchronization register returns the expected completion handshake value during the final polling loop. The test ends with finish(0) indicating success.",
   "Remarks":"The testcase uses conditional compilation to support multiple PCIe modes (DM0_RC, DM1_RC, DM0_EP, DM1_EP). Link training speed is set to Gen4. Cache coherency programming is performed in multiple phases with waits between them. The link status polling uses a bitmask check for the expected link-up pattern. BAR enumeration follows the standard PCIe enumeration procedure of writing all-ones and reading back to determine BAR size. Several system-level control registers could not be mapped to known PCIe DBI register names. One link status polling register could not be resolved to a known register name. The COHERENCY_CONTROL_3_OFF register is accessed on both PCIe controller instances (PCIE0 and PCIE1). The macro for the PCIE1 instance could not be fully resolved during macro resolution.",
   "Meta Test Description":"This testcase performs PCIe device enumeration across two dual-mode controllers (DM0 and DM1). It begins by writing 0x0 to 0xE6004100, then initiates link training via conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP modes using link_training_dm0_x4 or link_training_dm1_x4 with speed 4). Cache programming is performed by read-modify-write of mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF using set_data to set bit fields [11:14], [3:6], [27:30], and [19:22] to 0xF. After wait_on(20), the same coherency registers are programmed again. The test then polls read_sii0_reg(0xC0) until (data_rd & 0xD1) == 0xD1, and similarly polls read_sii1_reg(0xC0). Under DM0_RC mode, it reads the Vendor ID from read_pcie_slv0_reg(0x0), writes 0x7 to write_pcie_slv0_reg(0x4) to enable bus master, memory space, and I/O space, then calls mem_base_program_dm0_x4 and mem_base_program_dm1_x4. System-level registers 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, and 0xE6900034 are written with 0x1. Cache is then disabled by read-modify-write of the coherency control registers setting fields [19:22] and [27:30] to 0x0. After wait_on(30), BAR enumeration is performed on both slave ports: BAR registers at offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24 are written with 0xFFFFFFFF, read back to determine BAR size, then programmed with actual base addresses (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000). This is done for both pcie_slv1 and pcie_slv0. Finally, the test polls 0xE6004100 until the value equals 0x12345678, then calls finish(0).",
   "Meta Test Steps / Procedure":"1. Initialize global variables data_rd, data_wr, rd_wr_data1, err1, err2. 2. Write 0x0 to 0xE6004100 to clear the synchronization register. 3. Perform link training via link_training_dm0_x4(4) or link_training_dm1_x4(4) based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). 4. Cache programming: Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, apply set_data on bit fields [11:14] and [3:6] with value 0xF, write back. 5. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF again, apply set_data on bit fields [27:30] and [19:22] with value 0xF, write back. 6. Repeat steps 4-5 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 7. Call wait_on(20). 8. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set all four bit field groups [11:14], [3:6], [27:30], [19:22] to 0xF, write back. 9. Repeat step 8 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 10. Read read_sii0_reg(0xC0) and poll in a while loop until (data_rd & 0xD1) == 0xD1. 11. Read read_sii1_reg(0xC0) and poll in a while loop until (data_rd & 0xD1) == 0xD1. 12. Under DM0_RC: Read Vendor ID via read_pcie_slv0_reg(0x0). 13. Write 0x7 to write_pcie_slv0_reg(0x4) to enable bus master, memory space, and I/O space. 14. Call mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). 15. Call wait_on(10). 16. Write 0x1 to system registers 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034. 17. Disable cache: Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting fields [19:22] to 0x0. 18. Repeat step 17 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 19. Call wait_on(10). 20. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting fields [27:30] and [19:22] to 0x0. 21. Repeat step 20 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 22. Call wait_on(30). 23. Write 0xFFFFFFFF to pcie_slv1 BAR registers at offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24. 24. Read back pcie_slv1 BAR registers at offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24 to determine BAR sizes. 25. Write actual base addresses to pcie_slv1 BAR registers: 0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000. 26. Read back pcie_slv1 BAR registers to verify programmed values. 27. Repeat steps 23-26 for pcie_slv0 BAR registers. 28. Call wait_on(10). 29. Poll read_reg(0xE6004100) until data_rd equals 0x12345678, with wait_on(5) between iterations. 30. Call finish(0) to end the test.",
   "Meta Impacted Registers":"0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24"},
  {"Index":"2","SS / Module":"PCIE","Test Case Name":"pcie_reg_wr_rd_test","Feature":"PCIe Register Read Write Verification",
   "Test Description":"Verifies the reset default values and read-write integrity of PCIe DBI, SII, and PHY registers across both PCIe controller instances. The test first reads all target registers and compares them against their expected reset default values. It then writes the PHY reset control registers to release PHY from reset. After verifying reset defaults, the test performs a write-read-compare cycle using multiple data patterns on MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF, SII transmit header registers, SII PHY control registers, and PHY lane registers. Write masks are applied to SII and PHY registers before writing. Read-back values are compared against expected values with appropriate masking. The test reports pass if all comparisons match and fail if any mismatch is detected.",
   "Impacted Registers":"MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
   "Validation / Acceptance Criteria":"The test passes when: (1) All DBI control registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on both PCIe controller instances read back their expected reset default values of zero. (2) All SII transmit header and PHY control registers on both SII interfaces read back their expected reset default values of zero. (3) All PHY lane registers on both controller instances read back their expected reset default values of zero after 16-bit extraction. (4) For each of the three write-read-compare data patterns, all DBI control registers on both controller instances read back the exact written value. (5) All SII registers on both interfaces read back the written value masked with the appropriate write mask. (6) All PHY lane registers on both controller instances read back the written PHY pattern masked appropriately after 16-bit extraction. The test fails and reports an error message if any read-back value does not match the expected value. The final result is determined by finish(err2 || err1), where a non-zero error count indicates failure.",
   "Remarks":"The testcase covers both PCIe controller instances (PCIE0 and PCIE1) with identical register sets. Three data patterns are used for write-read-compare testing on DBI and SII registers, and three separate PHY-specific patterns are used for PHY registers. Write masks are applied to SII registers and PHY registers to account for read-only or reserved bit fields. PHY register reads use 16-bit extraction logic based on address alignment. The PHY reset control registers are written before PHY register access to ensure the PHY is out of reset. Several SII registers, PHY reset control registers, and PHY lane registers could not be mapped to canonical register names in the specification document. The DBI registers are tested on both controller instances and map to the same canonical register names.",
   "Meta Test Description":"This testcase verifies the reset default values and read-write functionality of PCIe registers across both PCIe controller instances (PCIE0 and PCIE1), SII interfaces, and PHY registers. The test defines six register address arrays: rc0_ctl_addr[5] containing {mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF}, rc1_ctl_addr[5] containing {mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE1_DBI_DSP_UTILITY_OFF}, sii0_addr[3] containing {mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2, mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3, mizar_PCIE0_SII_PHY_CONTROL_23}, sii1_addr[3] containing {mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2, mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3, mizar_PCIE1_SII_PHY_CONTROL_23}, phy0_addr[3] containing {0xE68860B8, 0xE68862B8, 0xE68864B8}, and phy1_addr[3] containing {0xE68A60B8, 0xE68A62B8, 0xE68A64B8}.",
   "Meta Test Steps / Procedure":"1. Declare global variables data_rd, data_wr, data1_rd, err1=0, err2=0. 2. Initialize rc0_ctl_addr[5] with {mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF}. 3. Initialize rc1_ctl_addr[5] with {mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE1_DBI_DSP_UTILITY_OFF}. 4. Initialize ctl_default[5] = {0x0, 0x0, 0x0, 0x0, 0x0}. 5. Initialize sii0_addr[3] with {mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2, mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3, mizar_PCIE0_SII_PHY_CONTROL_23}. 6. Initialize sii1_addr[3] with {mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2, mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3, mizar_PCIE1_SII_PHY_CONTROL_23}. 7. Initialize sii_default[3] = {0x0, 0x0, 0x0}. 8. Initialize sii0_write_mask[3] = {0xFFFFFFFF, 0xFFFFFFFF, 0xF000F} and sii1_write_mask[3] = {0xFFFFFFFF, 0xFFFFFFFF, 0xF000F}. 9. Initialize phy0_addr[3] = {0xE68860B8, 0xE68862B8, 0xE68864B8} and phy1_addr[3] = {0xE68A60B8, 0xE68A62B8, 0xE68A64B8}. 10. Initialize phy0_default[3] = {0x0, 0x0, 0x0} and phy1_default[3] = {0x0, 0x0, 0x0}. 11. Initialize phy0_write_mask[3] = {0x1FFF, 0x1FFF, 0x1FFF} and phy1_write_mask[3] = {0x1FFF, 0x1FFF, 0x1FFF}. 12. Call chk_rst_val(): 13. Loop i=0 to 4: read_reg(rc0_ctl_addr[i]), compare data_rd against ctl_default[i], increment err1 on mismatch. 14. Loop i=0 to 4: read_reg(rc1_ctl_addr[i]), compare data_rd against ctl_default[i], increment err2 on mismatch. 15. Loop i=0 to 2: read_reg(sii0_addr[i]), compare data_rd against sii_default[i], increment err2 on mismatch. 16. Loop i=0 to 2: read_reg(sii1_addr[i]), compare data_rd against sii_default[i], increment err2 on mismatch. 17. write_reg(mizar_PCIE0_SII_PHY_RST_CONTROL, 0x01203000) to release PCIE0 PHY reset. 18. write_reg(mizar_PCIE1_SII_PHY_RST_CONTROL, 0x01203000) to release PCIE1 PHY reset. 19. Loop i=0 to 2: read_reg(phy0_addr[i]), apply 16-bit extraction, compare against phy0_default[i], increment err2 on mismatch. 20. Loop i=0 to 2: read_reg(phy1_addr[i]), apply same 16-bit extraction, compare against phy1_default[i], increment err2 on mismatch. 21. Call chk_rd_wr(): 22-37. Write-read-compare loop with 3 patterns. 38. Call finish(err2 || err1).",
   "Meta Impacted Registers":"mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8"}
]

wb=Workbook(); ws=wb.active; ws.title="TestPlan"; wm=wb.create_sheet("MetaData")
TC=["Index","SS / Module","Feature","Test Case Name","Test Description","Speed","Mode","Memory Start Offset","Memory End Offset","Remarks","Test Steps / Procedure","Impacted Registers","Validation / Acceptance Criteria","Code Generation"]
MC=["Index","Test Case Name","Meta Test Description","Meta Test Steps / Procedure","Meta Impacted Registers","Meta Validation / Acceptance Criteria","Meta Headers","Meta Macros","Meta Arrays"]
hf_=Font(name="Calibri",size=11,bold=True,color="FFFFFF"); hfl=PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
ha_=Alignment(horizontal="center",vertical="top",wrap_text=True); ca_=Alignment(vertical="top",wrap_text=True)
bd=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
for i,c in enumerate(TC,1):
    cl=ws.cell(row=1,column=i,value=c); cl.font=hf_; cl.fill=hfl; cl.alignment=ha_; cl.border=bd
for i,c in enumerate(MC,1):
    cl=wm.cell(row=1,column=i,value=c); cl.font=hf_; cl.fill=hfl; cl.alignment=ha_; cl.border=bd
for ri,tc in enumerate(DATA,2):
    n=tc["Test Case Name"]
    v=[tc.get("Index",""),tc.get("SS / Module",""),tc.get("Feature",""),n,tc.get("Test Description",""),"","","","",tc.get("Remarks",""),None,tc.get("Impacted Registers",""),tc.get("Validation / Acceptance Criteria",""),""]
    for ci,val in enumerate(v,1):
        if ci==11: continue
        cl=ws.cell(row=ri,column=ci,value=val); cl.alignment=ca_; cl.border=bd
    sc=ws.cell(row=ri,column=11)
    try: sc.value=rich(n)
    except: sc.value=plain(n)
    sc.alignment=ca_; sc.border=bd
    mv=[tc.get("Index",""),n,tc.get("Meta Test Description",""),tc.get("Meta Test Steps / Procedure",""),tc.get("Meta Impacted Registers",""),tc.get("Validation / Acceptance Criteria",""),"","",""]
    for ci,val in enumerate(mv,1):
        cl=wm.cell(row=ri,column=ci,value=val); cl.alignment=ca_; cl.border=bd
dv=DataValidation(type="list",formula1='"Required,Not Required"',allow_blank=True,showDropDown=False)
dv.sqref=f"{get_column_letter(14)}2:{get_column_letter(14)}1000"; ws.add_data_validation(dv)
tw={"Index":8,"SS / Module":15,"Feature":30,"Test Case Name":35,"Test Description":60,"Speed":10,"Mode":10,"Memory Start Offset":20,"Memory End Offset":20,"Remarks":50,"Test Steps / Procedure":90,"Impacted Registers":60,"Validation / Acceptance Criteria":70,"Code Generation":18}
for i,c in enumerate(TC,1): ws.column_dimensions[get_column_letter(i)].width=tw.get(c,20)
mww={"Index":8,"Test Case Name":35,"Meta Test Description":80,"Meta Test Steps / Procedure":80,"Meta Impacted Registers":80,"Meta Validation / Acceptance Criteria":70,"Meta Headers":30,"Meta Macros":30,"Meta Arrays":30}
for i,c in enumerate(MC,1): wm.column_dimensions[get_column_letter(i)].width=mww.get(c,20)
for r in range(2,len(DATA)+2): ws.row_dimensions[r].height=250; wm.row_dimensions[r].height=200
ws.freeze_panes="A2"; wm.freeze_panes="A2"; wm.sheet_state="veryHidden"; wb.active=0
out=os.path.join(os.getcwd(),FN); wb.save(out)
print(f"FILE:{FN}"); print(f"SIZE:{os.path.getsize(out)}"); print(f"PATH:{out}")
wb2=load_workbook(out)
assert "TestPlan" in wb2.sheetnames and "MetaData" in wb2.sheetnames
assert wb2["MetaData"].sheet_state=="veryHidden"
print(f"TP_ROWS:{wb2['TestPlan'].max_row-1}"); print(f"MD_ROWS:{wb2['MetaData'].max_row-1}")
print("VALIDATION:PASSED"); wb2.close()

# Also write base64 to stdout for programmatic consumption
buf=io.BytesIO(); wb.save(buf); buf.seek(0)
b64=base64.b64encode(buf.read()).decode("ascii")
print(f"BASE64_START"); print(b64); print(f"BASE64_END")
