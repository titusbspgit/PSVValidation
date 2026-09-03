#!/usr/bin/env python3
"""Generate PCIE TestPlan Excel workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timezone, timedelta
import os, sys, base64

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime("%Y%m%d_%H%M%S")
filename = f"PCIE_TestPlan_{timestamp}.xlsx"

json_data = [
    {
        "Index": "1",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_device_enumerate_test",
        "Feature": "Device Enumeration",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Test Description": "This test validates PCIe device enumeration...",
        "Test Steps / Procedure": "1. Clear the synchronization register...",
        "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; ...",
        "Validation / Acceptance Criteria": "1. The SII0 link status register must reach...",
        "Remarks": "The source code contains a duplicated block...",
        "Meta Headers": "...",
        "Meta Macros": "NA",
        "Meta Arrays": "NA",
        "Meta Test Description": "...",
        "Meta Test Steps / Procedure": "...",
        "Meta Impacted Registers": "...",
        "Meta Validation / Acceptance Criteria": "..."
    }
]

wb = openpyxl.Workbook()
ws_tp = wb.active
ws_tp.title = "TestPlan"
ws_md = wb.create_sheet("MetaData")

# Headers
tp_cols = ["Index","SS / Module","Feature","Test Case Name","Test Description","Speed","Mode","Memory Start Offset","Memory End Offset","Remarks","Test Steps / Procedure","Impacted Registers","Validation / Acceptance Criteria","Code Generation"]
md_cols = ["Index","Test Case Name","Meta Test Description","Meta Test Steps / Procedure","Meta Impacted Registers","Meta Validation / Acceptance Criteria","Meta Headers","Meta Macros","Meta Arrays"]

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")

for i, col in enumerate(tp_cols, 1):
    cell = ws_tp.cell(row=1, column=i, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

for i, col in enumerate(md_cols, 1):
    cell = ws_md.cell(row=1, column=i, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

ws_tp.freeze_panes = "A2"
ws_md.freeze_panes = "A2"
ws_md.sheet_state = "veryHidden"

wb.save(filename)
print(f"FILENAME={filename}")
# Output base64 for upload
with open(filename, "rb") as f:
    b64 = base64.b64encode(f.read()).decode()
print(f"B64LEN={len(b64)}")
print(f"FILESIZE={os.path.getsize(filename)}")
