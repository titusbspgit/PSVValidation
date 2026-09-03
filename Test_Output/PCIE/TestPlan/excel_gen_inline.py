#!/usr/bin/env python3
"""
Inline Excel Generator for PCIE TestPlan.
Generates XLSX, converts to base64, writes to stdout.
"""
import sys
import os
import base64
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not available")
    sys.exit(1)

from datetime import datetime, timezone, timedelta

def main():
    IST = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(IST)
    timestamp = now_ist.strftime("%Y%m%d_%H%M%S")
    filename = f"PCIE_TestPlan_{timestamp}.xlsx"

    tp_columns = [
        "Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
        "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
        "Test Steps / Procedure", "Impacted Registers", "Validation / Acceptance Criteria",
        "Code Generation"
    ]
    md_columns = [
        "Index", "Test Case Name", "Meta Test Description", "Meta Test Steps / Procedure",
        "Meta Impacted Registers", "Meta Validation / Acceptance Criteria",
        "Meta Headers", "Meta Macros", "Meta Arrays"
    ]

    data = [{
        "Index": "1",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_device_enumerate_test",
        "Feature": "Device Enumeration",
        "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
        "Meta Macros": "NA",
        "Meta Arrays": "NA",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Meta Test Description": "This testcase performs PCIe device enumeration...",
        "Test Description": "This test performs PCIe device enumeration...",
        "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0)...",
        "Test Steps / Procedure": "1. Clear the synchronization register...",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF...",
        "Impacted Registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG...",
        "Meta Validation / Acceptance Criteria": "1. SII0 link status polling...",
        "Validation / Acceptance Criteria": "1. SII0 and SII1 link status registers...",
        "Remarks": "The test uses conditional compilation..."
    }]

    wb = openpyxl.Workbook()
    ws_tp = wb.active
    ws_tp.title = "TestPlan"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(vertical="top", wrap_text=True)

    for ci, cn in enumerate(tp_columns, 1):
        c = ws_tp.cell(row=1, column=ci, value=cn)
        c.font = hf; c.fill = hfill; c.alignment = ha

    for ri, rec in enumerate(data, 2):
        for ci, cn in enumerate(tp_columns, 1):
            v = rec.get(cn, "")
            c = ws_tp.cell(row=ri, column=ci, value=v if v else "")
            c.alignment = ca

    ws_tp.freeze_panes = "A2"
    for ci, cn in enumerate(tp_columns, 1):
        ws_tp.column_dimensions[get_column_letter(ci)].width = min(60, max(12, len(cn) + 4))

    ws_md = wb.create_sheet(title="MetaData")
    for ci, cn in enumerate(md_columns, 1):
        c = ws_md.cell(row=1, column=ci, value=cn)
        c.font = hf; c.fill = hfill; c.alignment = ha

    for ri, rec in enumerate(data, 2):
        for ci, cn in enumerate(md_columns, 1):
            v = rec.get(cn, "")
            c = ws_md.cell(row=ri, column=ci, value=v if v else "")
            c.alignment = ca

    ws_md.freeze_panes = "A2"
    for ci, cn in enumerate(md_columns, 1):
        ws_md.column_dimensions[get_column_letter(ci)].width = min(60, max(12, len(cn) + 4))

    ws_md.sheet_state = "veryHidden"

    buf = io.BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    
    print(f"FILENAME={filename}")
    print(f"FILESIZE={buf.tell()}")
    print(f"ROWS_TP=1")
    print(f"ROWS_MD=1")
    print(f"B64START")
    print(b64)
    print(f"B64END")

if __name__ == "__main__":
    main()
