#!/usr/bin/env python3
"""
PCIE TestPlan XLSX Generator - Agent 7
Generates PCIE_TestPlan_20260823_235734.xlsx with Main and MetaData sheets.
Run: python3 generate_PCIE_TestPlan.py
Requires: pip install openpyxl
"""
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──
OUTPUT_FILENAME = "PCIE_TestPlan_20260823_235734.xlsx"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, OUTPUT_FILENAME)

# ── Styles ──
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ── Testcase Data (4 testcases) ──
testcases = [
    {
        "Index": "1",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_device_enumerate_test",
        "Feature": "PCIe Device Enumeration",
        "Test Description": "Verifies PCIe device enumeration on dual PCIe controllers. The test initializes the system, performs link training for x4 lane configuration, programs cache coherency control registers, and polls link status registers until both controllers report link-up. Under Root Complex mode, the test reads the TYPE1_DEV_ID_VEND_ID_REG to obtain the Vendor ID, writes to TYPE1_STATUS_COMMAND_REG to enable IO space, memory space, and bus master access, and programs memory base addresses. System-level configuration registers are set. Cache coherency is subsequently disabled. BAR sizing is performed on both PCIe slave interfaces by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, reading back to determine BAR sizes, then programming actual base addresses and verifying. The test polls a completion status register until the expected completion value is observed.",
        "Test Steps / Procedure": "1. Initialize the system by writing to the system control register to clear its state.\n2. Perform PCIe link training for x4 lane width on the applicable controller mode (Root Complex or Endpoint).\n3. Program the COHERENCY_CONTROL_3_OFF register on both PCIe controllers to enable cache coherency by setting the appropriate bit fields.\n4. Wait for the coherency configuration to take effect.\n5. Poll the link status register on PCIe controller 0 until link-up is confirmed.\n6. Poll the link status register on PCIe controller 1 until link-up is confirmed.\n7. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG to obtain the device Vendor ID.\n8. Write to TYPE1_STATUS_COMMAND_REG to enable IO space, memory space, and bus master access.\n9. Program memory base addresses for both controllers.\n10. Configure system-level protection and control registers.\n11. Disable cache coherency by clearing the relevant bit fields in the COHERENCY_CONTROL_3_OFF register on both controllers.\n12. Perform BAR sizing on both PCIe slave interfaces by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, then reading back to determine addressable size.\n13. Program actual base address values into the BAR and bridge window registers on both slave interfaces and read back to verify.\n14. Poll the completion status register until the expected completion value is observed.\n15. End the test upon successful completion.",
        "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
        "Validation / Acceptance Criteria": "The test passes when: (1) Both PCIe controllers achieve link-up, confirmed by polling the link status register until the expected link-up pattern is observed. (2) The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID upon read. (3) BAR sizing returns valid size information after writing all-ones and reading back from BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG on both slave interfaces. (4) Programmed BAR values are read back successfully. (5) The completion status register returns the expected completion value, at which point finish(0) is called indicating test success.",
        "Remarks": "The testcase operates on dual PCIe controllers (controller 0 and controller 1) with both slave interfaces (slv0 and slv1). Link training mode is selected via compile-time defines for Root Complex or Endpoint configurations. The source contains a duplicated code block for link training and cache programming. Polling is used for link-up detection with a specific bitmask pattern and for final test completion. Multiple wait delays are inserted between configuration phases. Several system-level registers used for protection and control configuration could not be mapped to named registers in the specification. The link status register accessed via a separate bus interface accessor could not be mapped to a named register in the specification. One coherency control macro for the second PCIe controller could not be resolved to a register offset.",
        "Meta Test Description": "The testcase performs PCIe device enumeration across two PCIe controllers (DM0 and DM1). It begins by writing 0x0 to 0xE6004100 to initialize the system. Link training is conditionally invoked based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP) using link_training_dm0_x4(4) or link_training_dm1_x4(4) for x4 lane width. Cache coherency programming is performed by reading mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, modifying specific bit fields (bits 3-6, 11-14, 19-22, 27-30) using set_data(), and writing back the modified values. A wait_on(20) delay is inserted between coherency configuration phases. The test then polls read_sii0_reg(0xC0) in a while loop until (data_rd & 0xD1) == 0xD1, indicating link-up status for controller 0. Similarly, read_sii1_reg(0xC0) is polled for controller 1 link-up. Under DM0_RC configuration, the Vendor ID is read from read_pcie_slv0_reg(0x0), the command register is written via write_pcie_slv0_reg(0x4, 0x7) to enable IO, Memory, and Bus Master, and memory base programming functions are called. System-level registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, and 0xE6900034 are written with 0x1. Cache coherency is then disabled. BAR sizing is performed on both slave interfaces. Finally, 0xE6004100 is polled until it reads 0x12345678. The test concludes with finish(0).",
        "Meta Test Steps / Procedure": "1. Initialize global variables data_rd, data_wr, rd_wr_data1, err1=0, err2=0.\n2. Write 0x0 to 0xE6004100 to initialize system state.\n3. Conditionally invoke link_training_dm0_x4(4) or link_training_dm1_x4(4).\n4. CACHE PROGRAMMING: Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, apply set_data() to modify bits 11-14 with 0xf and bits 3-6 with 0xf, write back.\n5. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF again, modify bits 27-30 and 19-22 with 0xf, write back.\n6. Repeat steps 4-5 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF.\n7. Call wait_on(20).\n8-9. Repeat cache programming for both controllers.\n10. Repeat link training and cache programming sequence.\n11-15. Poll link status registers until link-up.\n16-18. Read Vendor ID, write command register, program memory base.\n19-20. Write system registers.\n21-24. Disable cache programming.\n25-30. BAR sizing on slv0 and slv1.\n31-32. Poll 0xE6004100 until 0x12345678.\n33. Call finish(0).",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24"
    },
    {
        "Index": "2",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_dma_write_test",
        "Feature": "PCIe DMA Write and Read Transfer",
        "Test Description": "Verifies PCIe DMA write and read data transfers across four DMA channels on a PCIe controller. The test initializes the system, performs link training for x4 lane configuration, and polls the link status register until link-up is confirmed. The Vendor ID is read from TYPE1_DEV_ID_VEND_ID_REG, and the TYPE1_STATUS_COMMAND_REG is written to enable IO space, memory space, and bus master access. BAR and memory base programming are performed. Source memory is preloaded with known data patterns. DMA write and read interrupt masks are unmasked by writing to DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF. For each of four DMA channels, the write channel is programmed and the DMA_WRITE_DOORBELL_OFF register is written with the channel number to trigger the transfer. The test waits for an interrupt to signal completion. After all write channels complete, the same sequence is performed for four DMA read channels using DMA_READ_DOORBELL_OFF. The interrupt handler reads DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF, masks the status, and writes the status to DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF to clear the interrupts.",
        "Test Steps / Procedure": "1. Initialize the system by writing to the system control register to clear its state.\n2. Perform PCIe link training for x4 lane width on the applicable controller mode.\n3. Poll the link status register until link-up is confirmed for the active controller.\n4. Read the TYPE1_DEV_ID_VEND_ID_REG to obtain the device Vendor ID.\n5. Write to TYPE1_STATUS_COMMAND_REG to enable IO space, memory space, and bus master access.\n6. Perform BAR programming and memory base programming for the active controller.\n7. Poll the completion status register until the expected synchronization value is observed.\n8. Preload source memory regions with known data patterns for DMA transfers.\n9. Enable interrupts via GIC setup.\n10. Unmask DMA write and read interrupts by writing to DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF.\n11. For each of four DMA write channels (0-3), program the channel and trigger the transfer by writing the channel number to DMA_WRITE_DOORBELL_OFF and wait for the interrupt-driven completion signal.\n12. For each of four DMA read channels (0-3), program the channel and trigger the transfer by writing the channel number to DMA_READ_DOORBELL_OFF and wait for the interrupt-driven completion signal.\n13. In the interrupt handler, read DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF, then write the status to DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF to clear the interrupts.\n14. End the test upon successful completion of all DMA channel transfers.",
        "Impacted Registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; DMA_WRITE_INT_MASK_OFF; DMA_READ_INT_MASK_OFF; DMA_WRITE_DOORBELL_OFF; DMA_READ_DOORBELL_OFF; DMA_WRITE_INT_STATUS_OFF; DMA_READ_INT_STATUS_OFF; DMA_WRITE_INT_CLEAR_OFF; DMA_READ_INT_CLEAR_OFF",
        "Validation / Acceptance Criteria": "The test passes when: (1) The PCIe link achieves link-up. (2) The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID. (3) All four DMA write channel transfers complete successfully via interrupt. (4) All four DMA read channel transfers complete successfully via interrupt. (5) The interrupt handler correctly reads and clears DMA_WRITE_INT_STATUS_OFF, DMA_READ_INT_STATUS_OFF, DMA_WRITE_INT_CLEAR_OFF, and DMA_READ_INT_CLEAR_OFF. (6) The test completes with finish(0).",
        "Remarks": "The testcase operates on one of two PCIe controllers selected via compile-time defines. Four DMA write channels and four DMA read channels are exercised sequentially, each triggered by writing the channel number to the doorbell register. Interrupt-driven completion is used. Source memory is preloaded with two distinct data patterns. The link status register and several system-level registers could not be mapped to named registers in the specification. All DMA registers for the second PCIe controller could not be resolved.",
        "Meta Test Description": "The testcase performs PCIe DMA write and read transfers across four DMA channels on either PCIe controller 0 (DM0) or controller 1 (DM1). It begins by writing 0x0 to 0xE6004100 to initialize the system. Link training is conditionally invoked. Under DM0_RC, read_sii0_reg(0xC0) is polled until (data_rd & 0xD1) == 0xD1. The Vendor ID is read via read_pcie_slv0_reg(0x0), the command register is written via write_pcie_slv0_reg(0x4, 0x7). Source memory is preloaded with 0xC0DEBEED and 0xF00DDEAF patterns. DMA write and read interrupt masks are unmasked. For each of channels 0-3, the write/read channel is programmed and the doorbell register is written to trigger the transfer. The interrupt handler reads status, masks lower 4 bits, and clears interrupts. The test concludes with finish(0).",
        "Meta Test Steps / Procedure": "1. Initialize global variables.\n2. Write 0x0 to 0xE6004100.\n3. Conditionally invoke link training.\n4. Poll read_sii0_reg(0xC0) until link-up.\n5. Read Vendor ID via read_pcie_slv0_reg(0x0).\n6. Write 0x7 to write_pcie_slv0_reg(0x4).\n7. Call bar_program and mem_base_program.\n8. Poll read_reg(0xE6004100) until 0x12345678.\n9. Preload source memory with 0xC0DEBEED and 0xF00DDEAF.\n10. Enable interrupts via GIC.\n11. Write 0x0 to mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF.\n12. For channels 0-3: program write channel, write channel number to mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF, wait for interrupt.\n13. For channels 0-3: program read channel, write channel number to mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF, wait for interrupt.\n14. IRQ Handler: Read mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF, clear via mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF.\n15. Call finish(0).",
        "Meta Impacted Registers": "0xE6004100; 0xC0; 0x0; 0x4; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_CLEAR_OFF"
    },
    {
        "Index": "3",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_mem_wr_rd_test",
        "Feature": "PCIe Memory Write and Read",
        "Test Description": "Verifies PCIe memory write and read operations through the PCIe slave interfaces on dual PCIe controllers. The test initializes the system, performs link training for x4 lane configuration, programs cache coherency control registers on both controllers, and polls the link status register until link-up is confirmed. Under Root Complex mode, the test reads the TYPE1_DEV_ID_VEND_ID_REG to obtain the Vendor ID, writes to TYPE1_STATUS_COMMAND_REG to enable IO space, memory space, and bus master access, and performs BAR and memory base programming. Under Endpoint mode, EP-specific BAR programming is performed. Cache coherency is subsequently disabled by clearing the relevant bit fields in the COHERENCY_CONTROL_3_OFF register on both controllers. Memory write-read verification is performed through the PCIe slave interface at multiple memory addresses with known data patterns. The test polls a completion status register until the expected completion value is observed.",
        "Test Steps / Procedure": "1. Initialize the system by writing to the system control register to clear its state.\n2. Perform PCIe link training for x4 lane width on the applicable controller mode (Root Complex or Endpoint).\n3. Program the COHERENCY_CONTROL_3_OFF register on both PCIe controllers to enable cache coherency.\n4. Wait for the coherency configuration to take effect.\n5. Poll the link status register until link-up is confirmed for the active controller.\n6. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG to obtain the device Vendor ID.\n7. Write to TYPE1_STATUS_COMMAND_REG to enable IO space, memory space, and bus master access.\n8. Perform BAR programming and memory base programming for the active controller.\n9. In Endpoint mode, perform EP-specific BAR programming and memory base programming.\n10. Configure system-level protection registers.\n11. Write a synchronization value to the system control register.\n12. Disable cache coherency by clearing the relevant bit fields in the COHERENCY_CONTROL_3_OFF register on both controllers.\n13. Perform memory write-read verification through the PCIe slave interface at multiple memory addresses with known data patterns.\n14. Poll the completion status register until the expected completion value is observed.\n15. End the test upon successful completion.",
        "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG",
        "Validation / Acceptance Criteria": "The test passes when: (1) The PCIe link achieves link-up. (2) The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID. (3) Memory write-read operations through the PCIe slave interface complete successfully for all tested memory addresses and data patterns. (4) The completion status register returns the expected completion value, at which point finish(0) is called indicating test success.",
        "Remarks": "The testcase supports both Root Complex and Endpoint modes selected via compile-time defines. In Root Complex mode, three memory addresses are tested with distinct data patterns per controller. In Endpoint mode, five BAR1 offset addresses are tested with the same data pattern. Cache coherency is enabled before link-up and disabled before memory operations. A long wait delay is used in Endpoint mode for link stabilization. The link status register accessed via a separate bus interface accessor could not be mapped to a named register in the specification. One coherency control macro for the second PCIe controller could not be resolved to a register offset.",
        "Meta Test Description": "The testcase performs PCIe memory write and read operations through the PCIe slave interfaces. It begins by writing 0x0 to 0xE6004100 to initialize the system. Link training is conditionally invoked. Cache coherency programming is performed by reading mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, modifying specific bit fields using set_data(), and writing back. Link status is polled via read_sii0_reg(0xC0) or read_sii1_reg(0xC0). Under RC mode, Vendor ID is read, command register is written. BAR and memory base programming are performed. Cache coherency is disabled. Memory write-read verification is performed via pcie_slv0_mem_wr_rd() or pcie_slv1_mem_wr_rd() at multiple addresses. 0xE6004100 is polled until 0x12345678. The test concludes with finish(0).",
        "Meta Test Steps / Procedure": "1. Initialize global variables.\n2. Write 0x0 to 0xE6004100.\n3. Conditionally invoke link training.\n4. CACHE PROGRAMMING on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF.\n5. Call wait_on(20).\n6. Repeat cache programming.\n7. Read read_sii0_reg(0xC0), poll until link-up.\n8. Under DM0_RC: Read read_pcie_slv0_reg(0x0), write write_pcie_slv0_reg(0x4, 0x7), call bar and mem_base programs.\n9. Call non_secure_prot_nic().\n10. Write 0x11111111 to 0xE6004100.\n11. DISABLE_CACHE PROGRAMMING.\n12. Call wait_on(30).\n13-15. Perform memory write-read operations.\n16. Poll read_reg(0xE6004100) until 0x12345678.\n17. Call finish(0).",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4"
    },
    {
        "Index": "4",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_reg_wr_rd_test",
        "Feature": "PCIe Register Write-Read Verification",
        "Test Description": "Verifies PCIe register reset values and write-read integrity across multiple register groups on both PCIe controllers. The test first checks that all DBI controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF), SII interface registers, and PHY registers read their expected default values after reset. It then performs write-read verification by writing multiple test patterns and reading back to confirm data integrity. Write masks are applied for SII and PHY registers to account for read-only bit fields. The test reports pass or fail based on accumulated error counts from all comparison operations.",
        "Test Steps / Procedure": "1. Read all DBI controller registers on both PCIe controllers and verify they contain their expected default reset values.\n2. Read all SII interface registers on both controllers and verify they contain their expected default reset values.\n3. Write to the PHY reset control register on both controllers to enable PHY access.\n4. Read all PHY registers on both controllers with appropriate alignment-based masking and verify they contain their expected default reset values.\n5. Write a set of test patterns to all DBI controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on both controllers.\n6. Write test patterns with appropriate write masks to all SII interface registers on both controllers.\n7. Write PHY-specific test patterns with appropriate write masks to all PHY registers on both controllers.\n8. Read back all DBI controller registers and verify the written values match.\n9. Read back all SII interface registers and verify the written values match accounting for write masks.\n10. Read back all PHY registers with alignment-based masking and verify the written values match accounting for write masks.\n11. Repeat steps 5-10 for multiple test patterns.\n12. Report pass if no mismatches are detected across all reset-value and write-read checks; otherwise report fail.",
        "Impacted Registers": "MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
        "Validation / Acceptance Criteria": "The test passes when: (1) All DBI controller registers on both PCIe controllers read their expected default reset values of zero. (2) All SII interface registers on both controllers read their expected default reset values of zero. (3) All PHY registers on both controllers read their expected default reset values of zero after alignment-based masking. (4) For each test pattern written to the DBI controller registers, the read-back values match the written values exactly. (5) For each test pattern written to the SII interface registers, the read-back values match the written values after applying the appropriate write masks. (6) For each PHY-specific test pattern written to the PHY registers, the read-back values match the written values after applying alignment-based masking and write masks. (7) The test completes with finish(0) when both err1 and err2 are zero.",
        "Remarks": "The testcase does not perform link training or link-up polling. It directly accesses DBI controller registers, SII interface registers, and PHY registers without requiring PCIe link establishment. Write masks are applied for SII registers and PHY registers to account for read-only bit fields. PHY register reads use alignment-based masking where odd-aligned addresses are shifted right by 16 bits and even-aligned addresses are masked with the lower 16 bits. Three test pattern iterations are performed out of six defined patterns. All SII interface registers and PHY reset control registers could not be mapped to named registers in the specification. All PHY registers accessed via hardcoded addresses could not be mapped to named registers in the specification. All register macros for the second PCIe controller could not be resolved to register offsets.",
        "Meta Test Description": "The testcase performs register reset-value verification and write-read testing across multiple PCIe register groups on both PCIe controllers. Global arrays define the register addresses: rc0_ctl_addr[5] contains {mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF}. rc1_ctl_addr[5] contains the corresponding PCIE1 macros. sii0_addr[3] and sii1_addr[3] contain SII register macros. phy0_addr[3] and phy1_addr[3] contain PHY register addresses. chk_rst_val() reads all registers and compares against expected defaults. chk_rd_wr() writes multiple test patterns and reads back to verify. Write masks are applied for SII and PHY registers. The test concludes with finish(err2 || err1).",
        "Meta Test Steps / Procedure": "1. Initialize global variables.\n2. Define register address arrays rc0_ctl_addr[5], rc1_ctl_addr[5], sii0_addr[3], sii1_addr[3], phy0_addr[3], phy1_addr[3].\n3. Define default value arrays and write mask arrays.\n4. Call chk_rst_val(): Read all registers and compare against defaults.\n5. Write 0x01203000 to mizar_PCIE0_SII_PHY_RST_CONTROL and mizar_PCIE1_SII_PHY_RST_CONTROL.\n6. Read PHY registers with alignment-based masking and compare against defaults.\n7. Call chk_rd_wr(): For each test pattern, write to all register groups, read back and compare.\n8. Call finish(err2 || err1).",
        "Meta Impacted Registers": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8"
    }
]

# ── Main Sheet Columns ──
main_cols = [
    "Index", "SS / Module", "Test Case Name", "Feature",
    "Test Description", "Test Steps / Procedure", "Impacted Registers",
    "Validation / Acceptance Criteria", "Remarks"
]

# ── MetaData Sheet Columns ──
meta_cols = [
    "Index", "SS / Module", "Test Case Name", "Feature",
    "Meta Test Description", "Meta Test Steps / Procedure", "Meta Impacted Registers"
]

def build_workbook():
    wb = openpyxl.Workbook()

    # ── Main Sheet ──
    ws_main = wb.active
    ws_main.title = "Main"
    for c, col_name in enumerate(main_cols, 1):
        cell = ws_main.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for r, tc in enumerate(testcases, 2):
        for c, col_name in enumerate(main_cols, 1):
            cell = ws_main.cell(row=r, column=c, value=tc.get(col_name, "NA"))
            cell.alignment = cell_align
            cell.border = thin_border
    main_widths = [8, 15, 35, 30, 60, 60, 40, 60, 50]
    for i, w in enumerate(main_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = w
    ws_main.auto_filter.ref = f"A1:{get_column_letter(len(main_cols))}1"
    ws_main.freeze_panes = "A2"

    # ── MetaData Sheet ──
    ws_meta = wb.create_sheet("MetaData")
    for c, col_name in enumerate(meta_cols, 1):
        cell = ws_meta.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for r, tc in enumerate(testcases, 2):
        for c, col_name in enumerate(meta_cols, 1):
            cell = ws_meta.cell(row=r, column=c, value=tc.get(col_name, "NA"))
            cell.alignment = cell_align
            cell.border = thin_border
    meta_widths = [8, 15, 35, 30, 80, 80, 60]
    for i, w in enumerate(meta_widths, 1):
        ws_meta.column_dimensions[get_column_letter(i)].width = w
    ws_meta.auto_filter.ref = f"A1:{get_column_letter(len(meta_cols))}1"
    ws_meta.freeze_panes = "A2"

    return wb

if __name__ == "__main__":
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    size = os.path.getsize(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")
    print(f"File size: {size} bytes")
    print(f"Sheets: Main ({len(main_cols)} columns, {len(testcases)} rows), MetaData ({len(meta_cols)} columns, {len(testcases)} rows)")
    print("Done.")
