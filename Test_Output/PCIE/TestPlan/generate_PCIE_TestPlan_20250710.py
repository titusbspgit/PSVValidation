#!/usr/bin/env python3
"""Generate PCIE_TestPlan_20250710_033000.xlsx with openpyxl."""
import json, os, sys
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print('ERROR: openpyxl not installed. Run: pip install openpyxl')
    sys.exit(1)

# ── DATA ──────────────────────────────────────────────────────────────────
json_data = [
  {
    "Index": "1",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_device_enumerate_test",
    "Feature": "Device Enumeration",
    "Meta Headers": '<stdlib.h>; <stdio.h>; <test_common.h>; "pcie.h"',
    "Meta Macros": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DEBUG_DISPLAY",
    "Meta Arrays": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs PCIe device enumeration. It begins by writing 0x0 to register at 0xE6004100 and then initiates link training for dual-mode controllers (DM0/DM1) in x4 configuration depending on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). Cache coherency programming is performed by reading mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, using set_data to modify bit fields [11:14], [3:6], [27:30], [19:22] with value 0xf, and writing back. After wait_on(20), the same coherency registers are programmed again. The test then polls read_sii0_reg(0xC0) in a while loop until (data_rd & 0xD1) == 0xD1, and similarly polls read_sii1_reg(0xC0). Under DM0_RC, it reads the Vendor ID from read_pcie_slv0_reg(0x0), writes 0x7 to write_pcie_slv0_reg(0x4), and calls mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). System-level registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034 are written with 0x1. Cache disable programming is then performed by modifying the same coherency control registers with 0x0 in bit fields [27:30] and [19:22]. After wait_on(30), BAR registers (offsets 0x10-0x24) for both PCIe slave port 0 and slave port 1 are written with 0xFFFFFFFF, read back, then written with specific address values and read back again. Finally, the test polls read_reg(0xE6004100) in a while loop with wait_on(5) until the value equals 0x12345678, then calls finish(0).",
    "Test Description": "This test performs PCIe device enumeration by initializing link training for dual-mode controllers in x4 configuration, programming cache coherency settings in the COHERENCY_CONTROL_3_OFF register for both PCIe port 0 and port 1, polling the gic register on SII0 and SII1 interfaces until link status is confirmed, reading the Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enabling memory and bus master via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, configuring system-level control registers, disabling cache coherency, enumerating BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports by writing all-ones to determine BAR sizes and then programming specific base addresses, and finally polling a synchronization register until a completion handshake value is received.",
    "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0) to initialize control register. 2. Call link_training_dm0_x4(4) or link_training_dm1_x4(4) based on DM0_RC/DM1_RC/DM0_EP/DM1_EP compile defines. 3. CACHE PROGRAMMING: Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, use set_data to modify bits [11:14]=0xf, [3:6]=0xf, write back. 4. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF again, set_data bits [27:30]=0xf, [19:22]=0xf, write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. wait_on(20). 7. Repeat cache programming for both PCIE0 and PCIE1 coherency registers with all four bit fields set to 0xf. 8-24. [Full steps as documented].",
    "Test Steps / Procedure": "1. Initialize the system control register to clear any previous state. 2. Perform PCIe link training for dual-mode controllers in x4 lane configuration based on the configured role. 3. Program cache coherency settings in the COHERENCY_CONTROL_3_OFF register for both PCIe port 0 and port 1. 4. Wait for coherency settings to take effect. 5. Re-apply cache coherency programming for both ports. 6. Poll the gic register on the SII0 interface until the link status indicates ready. 7. Configure non-secure protection via NIC. 8. Poll the gic register on the SII1 interface until the link status indicates ready. 9. Read the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG on slave port 0. 10. Write to TYPE1_STATUS_COMMAND_REG on slave port 0 to enable memory space access and bus mastering. 11. Program memory base addresses for both dual-mode controllers. 12. Enable system-level control registers. 13. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF. 14. Wait for cache disable to take effect. 15. Enumerate BAR registers on slave port 1. 16. Enumerate BAR registers on slave port 0. 17. Poll the synchronization register until the expected completion handshake value is received. 18. Complete the test.",
    "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
    "Impacted Registers": "COHERENCY_CONTROL_3_OFF; gic; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
    "Meta Validation / Acceptance Criteria": "1. Poll read_sii0_reg(0xC0): while loop checks (data_rd & 0xD1) != 0xD1. 2. Poll read_sii1_reg(0xC0): same condition. 3. read_pcie_slv0_reg(0x0) reads Vendor ID. 4. BAR enumeration: write 0xFFFFFFFF to offsets 0x10-0x24 on both slave ports. 5. Final polling: read_reg(0xE6004100) polled until data_rd == 0x12345678. 6. Test completes by calling finish(0).",
    "Validation / Acceptance Criteria": "1. The gic register on SII0 must return a value where the masked bits match the expected link-ready pattern. 2. The gic register on SII1 must similarly confirm link readiness. 3. The TYPE1_DEV_ID_VEND_ID_REG must return a valid Vendor ID. 4. BAR registers on both slave ports must respond correctly to all-ones writes for BAR sizing and must retain programmed base address values on read-back. 5. The synchronization register must eventually return the expected completion handshake value. 6. The test passes when finish is called with a success code of 0.",
    "Remarks": "Test uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select link training mode; multiple wait_on delays are used between cache programming phases; polling loops on SII0 and SII1 interfaces have no explicit timeout; the PCIE1 coherency control register macro could not be resolved from available headers."
  },
  {
    "Index": "2",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_dma_write_test",
    "Feature": "DMA Write and Read",
    "Meta Headers": '<stdlib.h>; <stdio.h>; <test_common.h>; "pcie.h"',
    "Meta Macros": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DEBUG_DISPLAY",
    "Meta Arrays": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs PCIe DMA write and read-back operations across all four DMA channels (Channel 0-3) for dual-mode controllers.",
    "Test Description": "This test validates PCIe DMA write and read-back functionality across all four DMA channels for dual-mode controllers in x4 configuration. It initializes the PCIe link, confirms link readiness by polling the gic register, reads the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enables memory space access and bus mastering via TYPE1_STATUS_COMMAND_REG, programs BAR and memory base addresses, and waits for a synchronization handshake. Source memory is preloaded with known data patterns. The GIC interrupt controller is configured. DMA write and read interrupt masks are cleared via DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF. For each of the four DMA channels, a write transfer is programmed and triggered via DMA_WRITE_DOORBELL_OFF. The interrupt handler reads DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF to determine which channels completed, then clears the interrupts via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF.",
    "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0). 2. Call link_training. 3-20. [Full DMA programming steps].",
    "Test Steps / Procedure": "1. Initialize the system control register. 2. Perform PCIe link training. 3. Poll the gic register on the appropriate SII interface. 4. Read the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG. 5. Write to TYPE1_STATUS_COMMAND_REG. 6. Program BAR registers and memory base addresses. 7. Configure non-secure protection via NIC. 8. Poll the synchronization register. 9. Preload source memory with known data patterns. 10. Configure the GIC interrupt controller. 11. Unmask DMA write and read interrupts. 12. For each of the four DMA write channels: program and trigger via DMA_WRITE_DOORBELL_OFF. 13. For each of the four DMA read channels: program and trigger via DMA_READ_DOORBELL_OFF. 14. In the interrupt handler, read and clear DMA status. 15. Complete the test.",
    "Meta Impacted Registers": "0xE6004100; 0xC0; 0x0; 0x4; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_CLEAR_OFF",
    "Impacted Registers": "gic; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; DMA_WRITE_INT_MASK_OFF; DMA_READ_INT_MASK_OFF; DMA_WRITE_DOORBELL_OFF; DMA_READ_DOORBELL_OFF; DMA_WRITE_INT_STATUS_OFF; DMA_READ_INT_STATUS_OFF; DMA_WRITE_INT_CLEAR_OFF; DMA_READ_INT_CLEAR_OFF",
    "Meta Validation / Acceptance Criteria": "1. Poll read_sii0_reg(0xC0) under DM0_RC. 2. Poll read_sii1_reg(0xC0) under DM1_RC. 3. Vendor ID read. 4. Synchronization poll. 5-10. DMA completion and interrupt validation.",
    "Validation / Acceptance Criteria": "1. The gic register on the SII interface must confirm link readiness. 2. TYPE1_DEV_ID_VEND_ID_REG must return a valid Vendor ID. 3. The synchronization register must return the expected handshake value. 4. Each DMA write channel transfer must complete successfully. 5. Each DMA read channel transfer must complete successfully. 6. All eight DMA transfers must complete without timeout. 7. The test passes when finish is called with a success code of 0.",
    "Remarks": "Test uses conditional compilation; DMA transfer length is fixed at 0x40 bytes; source data is preloaded with two known patterns; interrupt-driven completion uses a global int_pend flag; PCIE1 DMA register macros could not be resolved; GIC IRQ numbers differ between DM0 (0x20) and DM1 (0x23)."
  },
  {
    "Index": "3",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_mem_wr_rd_test",
    "Feature": "Memory Write and Read",
    "Meta Headers": '<stdlib.h>; <stdio.h>; <test_common.h>; "pcie.h"',
    "Meta Macros": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DM0; DM1; DEBUG_DISPLAY",
    "Meta Arrays": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs PCIe memory write and read-back verification across the PCIe slave memory space.",
    "Test Description": "This test validates PCIe memory write and read-back operations through the PCIe slave memory space for dual-mode controllers in x4 configuration. It initializes the PCIe link, programs cache coherency settings in the COHERENCY_CONTROL_3_OFF register for both PCIe port 0 and port 1, polls the gic register on the SII interface until link readiness is confirmed, reads the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enables memory space access and bus mastering via TYPE1_STATUS_COMMAND_REG, programs BAR registers and memory base addresses, configures non-secure protection, signals synchronization, then disables cache coherency. The test then performs memory write-read verification by writing known data patterns to multiple PCIe memory addresses and reading them back.",
    "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0). 2-27. [Full memory write-read steps].",
    "Test Steps / Procedure": "1. Initialize the system control register. 2. Perform PCIe link training. 3. Program cache coherency settings in COHERENCY_CONTROL_3_OFF. 4. Wait for coherency settings. 5. Re-apply cache coherency programming. 6. Poll the gic register. 7. Read the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG. 8. Write to TYPE1_STATUS_COMMAND_REG. 9. Program BAR registers and memory base addresses. 10. Configure non-secure protection via NIC. 11. Write a synchronization signal. 12. Disable cache coherency. 13. Wait for cache disable. 14. Perform memory write-read verification. 15. Poll the synchronization register. 16. Complete the test.",
    "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4",
    "Impacted Registers": "COHERENCY_CONTROL_3_OFF; gic; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG",
    "Meta Validation / Acceptance Criteria": "1. Poll read_sii0_reg(0xC0). 2. Poll read_sii1_reg(0xC0). 3. Vendor ID read. 4. pcie_slv0_mem_wr_rd / pcie_slv1_mem_wr_rd functions. 5. Final polling. 6. Error counters. 7. finish(0).",
    "Validation / Acceptance Criteria": "1. The gic register on the SII interface must confirm link readiness. 2. TYPE1_DEV_ID_VEND_ID_REG must return a valid Vendor ID. 3. Each PCIe memory write-read operation must successfully write the known data pattern and read back the same value. 4. The synchronization register must eventually return the expected completion handshake value. 5. The test passes when finish is called with a success code of 0.",
    "Remarks": "Test uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP, DM0, DM1); cache coherency is enabled then disabled in a phased sequence; under Endpoint mode a long wait_on(30000) is used; global error counters err1 and err2 are declared but not explicitly checked; the PCIE1 coherency control register macro could not be resolved."
  },
  {
    "Index": "4",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_reg_wr_rd_test",
    "Feature": "Register Write and Read Verification",
    "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; <pcie.h>",
    "Meta Macros": "NA",
    "Meta Arrays": "unsigned int rc0_ctl_addr[5] = {mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF}; unsigned int rc1_ctl_addr[5] = {...}; unsigned int phy0_addr[3] = {0xE68860B8, 0xE68862B8, 0xE68864B8}; unsigned int phy1_addr[3] = {0xE68A60B8, 0xE68A62B8, 0xE68A64B8}; int chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0x00000000, 0xA5A5A5A5, 0xffff0000}; int chk_val_phy[3] = {0x7baf, 0x1, 0x003b}",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Meta Test Description": "This testcase performs register reset value verification and write-read verification for PCIe DBI DSP controller registers, SII registers, and PHY registers across both PCIE0 and PCIE1 ports.",
    "Test Description": "This test verifies the reset default values and write-read functionality of PCIe registers across three register domains: DBI DSP controller registers, SII interface registers, and PHY registers for both PCIE port 0 and PCIE port 1. In the reset value check phase, the test reads MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, and UTILITY_OFF registers on both ports and verifies they contain the expected default values. In the write-read check phase, the test writes multiple data patterns to all registers, then reads them back and verifies the written values are retained correctly, accounting for register write masks on SII and PHY registers.",
    "Meta Test Steps / Procedure": "1. Call chk_rst_val(). 2-27. [Full register check steps].",
    "Test Steps / Procedure": "1. Read all five DBI DSP controller registers on PCIe port 0 and verify reset default values. 2. Read the same five registers on PCIe port 1 and verify. 3. Read three SII interface registers on port 0 and verify. 4. Read three SII interface registers on port 1 and verify. 5. Write to the PHY reset control registers on both ports. 6. Read three PHY registers on port 0 with 16-bit extraction and verify. 7. Read three PHY registers on port 1 and verify. 8. Begin write-read verification: for each of three data patterns, write to all registers. 9. Write the pattern to all three SII interface registers. 10. Write PHY-specific patterns. 11. Read back all DBI DSP controller registers and verify. 12. Read back all SII interface registers and verify. 13. Read back all PHY registers and verify. 14. Repeat for all three data patterns. 15. Verify the test completes with no accumulated errors.",
    "Meta Impacted Registers": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8",
    "Impacted Registers": "MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
    "Meta Validation / Acceptance Criteria": "1. Reset value check: all registers must read default 0x0. 2-13. Write-read check with patterns {0xffffffff, 0xaaaaaaaa, 0x55555555}.",
    "Validation / Acceptance Criteria": "1. All five DBI DSP controller registers on both ports must read back their expected reset default values. 2. All three SII interface registers on both ports must read back their expected reset default values. 3. All three PHY registers on both ports must read back their expected reset default values. 4. During write-read verification, all registers must retain the written data pattern. 5. All SII registers must retain the written data pattern masked by the register write mask. 6. All PHY registers must retain the written PHY data pattern masked by the PHY write mask. 7. The test passes only when both error counters are zero.",
    "Remarks": "Test covers three register domains: DBI DSP controller registers (5 per port), SII interface registers (3 per port), and PHY registers (3 per port) across both PCIE0 and PCIE1; SII registers use per-register write masks; PHY registers use 16-bit extraction logic and a 13-bit write mask; PCIE1 DBI DSP register macros, all SII register macros, and PHY reset control macros could not be resolved; PHY register hex addresses could not be mapped to named registers; two separate error counters track failures independently."
  }
]

# ── SHEET DEFINITIONS ─────────────────────────────────────────────────────
TP_COLS = ['Index', 'SS / Module', 'Feature', 'Test Case Name',
           'Test Description', 'Speed', 'Mode', 'Memory Start Offset',
           'Memory End Offset', 'Remarks', 'Test Steps / Procedure',
           'Impacted Registers', 'Validation / Acceptance Criteria',
           'Code Generation']

MD_COLS = ['Index', 'Test Case Name', 'Meta Test Description',
           'Meta Test Steps / Procedure', 'Meta Impacted Registers',
           'Meta Validation / Acceptance Criteria', 'Meta Headers',
           'Meta Macros', 'Meta Arrays']

# ── WORKBOOK CREATION ─────────────────────────────────────────────────────
wb = Workbook()
ws_tp = wb.active
ws_tp.title = 'TestPlan'
ws_md = wb.create_sheet('MetaData')

hdr_font = Font(bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
wrap = Alignment(wrap_text=True, vertical='top')

def write_sheet(ws, columns, data):
    # Header
    for c, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = wrap
    # Data rows
    for r, row_data in enumerate(data, 2):
        for c, col_name in enumerate(columns, 1):
            val = row_data.get(col_name, '')
            cell = ws.cell(row=r, column=c, value=val if val else '')
            cell.alignment = wrap
    # Auto-size
    for c, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for r in range(2, len(data) + 2):
            val = str(ws.cell(row=r, column=c).value or '')
            max_len = max(max_len, min(len(val), 80))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 60)
    ws.freeze_panes = 'A2'

write_sheet(ws_tp, TP_COLS, json_data)
write_sheet(ws_md, MD_COLS, json_data)

# Very Hidden
ws_md.sheet_state = 'veryHidden'

# Save
fname = 'PCIE_TestPlan_20250710_033000.xlsx'
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
wb.save(out_path)
print(f'SUCCESS: {out_path} ({os.path.getsize(out_path)} bytes)')
print(f'Sheets: {wb.sheetnames}')
