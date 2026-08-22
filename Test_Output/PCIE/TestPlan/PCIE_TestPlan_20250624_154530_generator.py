#!/usr/bin/env python3
"""PCIE TestPlan Excel Generator
Converts PCIE_TestPlan_20250624_154530.csv to PCIE_TestPlan_20250624_154530.xlsx
Run: python3 PCIE_TestPlan_20250624_154530_generator.py
"""
import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(script_dir, 'PCIE_TestPlan_20250624_154530.csv')
    xlsx_path = os.path.join(script_dir, 'PCIE_TestPlan_20250624_154530.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'TestPlan'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    wrap = Alignment(wrap_text=True, vertical='top')

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = wrap
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

    ws.freeze_panes = 'A2'

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = min(40, max(12, len(str(ws.cell(1, col).value or '')) + 4))

    wb.save(xlsx_path)
    print(f'Created: {xlsx_path}')
    print(f'Rows: {ws.max_row - 1} data rows')

if __name__ == '__main__':
    main()
