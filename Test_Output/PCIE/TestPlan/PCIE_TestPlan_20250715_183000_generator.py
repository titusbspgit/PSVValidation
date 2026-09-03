#!/usr/bin/env python3
import json
import os
import sys
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

json_data = [
    {
        "Index": "1",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_device_enumerate_test",
        "Feature": "Device Enumeration",
        "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
        "Meta Macros": "NA",
        "Meta Arrays": "NA",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Meta Test Description": "This testcase performs PCIe device enumeration. It begins by writing 0x0 to 0xE6004100, then initiates link training via link_training_dm0_x4 or link_training_dm1_x4 depending on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). Cache coherency programming is performed by read-modify-write sequences on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, setting bit fields [11:14], [3:6], [27:30], and [19:22] to 0xf using set_data(). After wait_on(20), the same coherency registers are programmed again with all four bit fields set to 0xf. The test then polls read_sii0_reg(0xC0) in a while loop until (data_rd & 0xD1) == 0xD1, and similarly polls read_sii1_reg(0xC0). Under DM0_RC, the Vendor ID is read from read_pcie_slv0_reg(0x0), the command register is written via write_pcie_slv0_reg(0x4, 0x7), and memory base programming is invoked via mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). System registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034 are written with 0x1. Cache disable programming follows by read-modify-write on both coherency control registers, setting fields [19:22] and [27:30] to 0x0. BAR registers (offsets 0x10-0x24) on both PCIe slave ports (slv0 and slv1) are written with 0xFFFFFFFF, read back, then written with specific address values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) and read back again. Finally, the test polls 0xE6004100 in a while loop until data_rd equals 0x12345678, with wait_on(5) between iterations, then calls finish(0).",
        "Test Description": "This test validates PCIe device enumeration. It initializes the PCIe link by performing link training for x4 configuration. Cache coherency control is programmed on both PCIE0 and PCIE1 instances by modifying specific bit fields in the COHERENCY_CONTROL_3_OFF register. The test polls a link status register on both SII0 and SII1 interfaces until the link reaches the expected ready state. Under Root Complex mode, the Vendor ID is read from the TYPE1_DEV_ID_VEND_ID_REG, and the TYPE1_STATUS_COMMAND_REG is configured to enable memory and bus master access. Memory base address programming is then performed for both dual-mode ports. System-level configuration registers are written to enable specific functions. The cache coherency settings are then disabled. BAR sizing is performed on both PCIe slave ports by writing all-ones to BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG, reading back the size masks, then programming the actual base addresses. The test concludes by polling a synchronization register until a specific completion pattern is received, confirming successful enumeration.",
        "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0) \u2014 clear synchronization register. 2. Conditional link training: link_training_dm0_x4(4) or link_training_dm1_x4(4) based on DM0_RC/DM1_RC/DM0_EP/DM1_EP defines. 3. CACHE PROGRAMMING: read_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF), set_data bits [11:14]=0xf, [3:6]=0xf, write_reg back. 4. Read same register, set_data bits [27:30]=0xf, [19:22]=0xf, write_reg back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. wait_on(20). 7. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data all four bit fields [11:14], [3:6], [27:30], [19:22] to 0xf, write back. 8. Read mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data all four bit fields to 0xf, write back. 9. Repeat link training and cache programming block (duplicate in source). 10. read_sii0_reg(0xC0) \u2014 poll in while loop until (data_rd & 0xD1) == 0xD1. 11. non_secure_prot_nic() call. 12. read_sii1_reg(0xC0) \u2014 poll in while loop until (data_rd & 0xD1) == 0xD1. 13. Under DM0_RC: read_pcie_slv0_reg(0x0) to get Vendor ID, write_pcie_slv0_reg(0x4, 0x7), call mem_base_program_dm0_x4() and mem_base_program_dm1_x4(), wait_on(10). 14. write_reg(0xE690000C, 0x1), write_reg(0xE6900010, 0x1), write_reg(0xE6900014, 0x1), write_reg(0xE6900018, 0x1), write_reg(0xE6900030, 0x1), write_reg(0xE6900034, 0x1). 15. DISABLE_CACHE PROGRAMMING: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF with bits [19:22]=0x0, [27:30]=0xf then 0x0. 16. Same disable cache for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 17. wait_on(10). 18. Repeat disable cache with all four fields set to 0x0 for both PCIE0 and PCIE1. 19. wait_on(30). 20. write_pcie_slv1_reg offsets 0x10-0x24 with 0xFFFFFFFF, read back each. 21. write_pcie_slv1_reg offsets 0x10-0x24 with specific addresses (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000), read back each. 22. write_pcie_slv0_reg offsets 0x10-0x24 with 0xFFFFFFFF, read back each. 23. write_pcie_slv0_reg offsets 0x10-0x24 with specific addresses, read back each. 24. wait_on(10). 25. Poll read_reg(0xE6004100) in while loop until data_rd == 0x12345678, with wait_on(5) between iterations. 26. finish(0).",
        "Test Steps / Procedure": "1. Clear the synchronization register to prepare for the enumeration sequence. 2. Initiate PCIe link training in x4 lane configuration for the appropriate dual-mode port based on the build configuration (Root Complex or Endpoint). 3. Program cache coherency control on both PCIE0 and PCIE1 COHERENCY_CONTROL_3_OFF registers by enabling specific bit fields for cache coherency. 4. Wait for the coherency settings to take effect. 5. Re-apply coherency control settings with all relevant bit fields enabled on both PCIE0 and PCIE1 instances. 6. Poll the SII0 link status register until the link ready condition is met (expected bit pattern confirmed). 7. Configure non-secure protection via NIC programming. 8. Poll the SII1 link status register until the link ready condition is met. 9. In Root Complex mode, read the Vendor ID from TYPE1_DEV_ID_VEND_ID_REG to verify device presence. 10. Write to TYPE1_STATUS_COMMAND_REG to enable memory space access and bus master capability. 11. Execute memory base address programming for both dual-mode PCIe ports. 12. Write to system-level configuration registers to enable required functions. 13. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF on both PCIE0 and PCIE1 instances. 14. Perform BAR sizing on PCIe slave port 1 by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, then reading back the size masks. 15. Program actual base addresses into the BAR registers on PCIe slave port 1 and verify by reading back. 16. Repeat BAR sizing and base address programming on PCIe slave port 0. 17. Poll the synchronization register until the expected completion pattern is received, confirming successful device enumeration. 18. Complete the test.",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
        "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
        "Meta Validation / Acceptance Criteria": "1. Poll read_sii0_reg(0xC0): while loop checks (data_rd & 0xD1) != 0xD1, exits when link status bits [7,6,4,0] are all set (0xD1 mask). 2. Poll read_sii1_reg(0xC0): same condition (data_rd & 0xD1) == 0xD1. 3. Vendor ID read from read_pcie_slv0_reg(0x0) is printed for verification. 4. BAR registers (0x10-0x24) on slv1 and slv0 are written with 0xFFFFFFFF and read back to determine BAR size. Then written with specific base addresses and read back to confirm programming. 5. Final poll: read_reg(0xE6004100) must equal 0x12345678 to exit the while loop and pass. 6. finish(0) is called to indicate test completion with pass status.",
        "Validation / Acceptance Criteria": "1. The SII0 link status register must reach the expected ready state with the required link-up bits set before the test proceeds. 2. The SII1 link status register must similarly reach the expected ready state. 3. The Vendor ID read from TYPE1_DEV_ID_VEND_ID_REG must return a valid device identifier confirming device presence on the PCIe bus. 4. BAR registers (BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports must accept the all-ones write and return valid size masks on read-back, and must accept the programmed base addresses and return them correctly on read-back. 5. The synchronization register must return the expected completion pattern to confirm that the remote endpoint has completed its enumeration sequence. 6. The test must complete successfully via the finish call with a pass indication.",
        "Remarks": "The source code contains a duplicated block of link training and cache programming. Conditional compilation via DM0_RC, DM1_RC, DM0_EP, DM1_EP selects the link training path. The mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF macro is unresolved in the register specification. Several system-level registers used for configuration could not be mapped to the PCIe DBI register specification. The SII0 and SII1 link status register at offset used for polling could not be mapped to the specification. The test uses wait_on() delays between critical configuration steps. The non_secure_prot_nic() call configures NIC-level security settings as a dependency."
    },
    {
        "Index": "2",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_dma_write_test",
        "Feature": "DMA Write and Read",
        "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
        "Meta Macros": "NA",
        "Meta Arrays": "NA",
        "Speed": "NA",
        "Mode": "Interrupt Mode",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Meta Test Description": "This testcase validates PCIe DMA write and read operations across all four DMA channels (Channel 0-3) on both PCIE0 (DM0_RC) and PCIE1 (DM1_RC) instances.",
        "Test Description": "This test validates PCIe DMA write and read operations across all four DMA channels on both PCIE0 and PCIE1 controller instances. The test initializes the PCIe link via link training in x4 configuration, polls the link status until the link is up, reads the Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, and configures the TYPE1_STATUS_COMMAND_REG to enable memory and bus master access. BAR and memory base programming are performed. Source memory is preloaded with known data patterns. The GIC interrupt controller is initialized and all IRQs are enabled. DMA write and read interrupt masks are cleared by writing to DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF. For each of the four DMA write channels, the channel is programmed with source, destination, and length parameters, and the DMA_WRITE_DOORBELL_OFF register is written with the channel number to trigger the transfer. The test waits for the DMA completion interrupt before proceeding to the next channel. After all four write channels complete, the same sequence is repeated for four DMA read channels using DMA_READ_DOORBELL_OFF. The interrupt handler reads DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF to determine which channel completed, then clears the interrupts via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF. The test completes successfully after all eight DMA transfers finish.",
        "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0). 2. Link training. 3-9. Link-up polling, Vendor ID read, BAR programming. 10-21. DMA channel programming and doorbell writes for all 4 write and 4 read channels with interrupt-driven completion.",
        "Test Steps / Procedure": "1. Clear the synchronization register to prepare for the test. 2. Initiate PCIe link training in x4 lane configuration for the appropriate dual-mode port. 3. Poll the link status register until the link-up condition is confirmed. 4. Read the Vendor ID from TYPE1_DEV_ID_VEND_ID_REG to verify device presence on the bus. 5. Write to TYPE1_STATUS_COMMAND_REG to enable memory space access and bus master capability. 6. Perform BAR programming and memory base address programming for the active PCIe port. 7. Configure non-secure protection via NIC programming. 8. Poll the synchronization register until the expected completion pattern is received from the remote endpoint. 9. Preload source memory with known data patterns for DMA transfer verification. 10. Initialize the GIC interrupt controller and enable all IRQs. 11. Unmask DMA write and read interrupts by writing to DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF. 12. For each DMA write channel (0 through 3): program the channel with source address, destination address, and transfer length, then trigger the transfer by writing the channel number to DMA_WRITE_DOORBELL_OFF and wait for the completion interrupt. 13. For each DMA read channel (0 through 3): program the channel with remote source address, local destination address, and transfer length, then trigger the transfer by writing the channel number to DMA_READ_DOORBELL_OFF and wait for the completion interrupt. 14. In the interrupt handler, read DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF to identify the completed channel, then clear the interrupt by writing to DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF. 15. After all eight DMA transfers complete, finalize the test.",
        "Meta Impacted Registers": "0xE6004100; 0xC0; 0x0; 0x4; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_CLEAR_OFF",
        "Impacted Registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; DMA_WRITE_INT_MASK_OFF; DMA_READ_INT_MASK_OFF; DMA_WRITE_DOORBELL_OFF; DMA_READ_DOORBELL_OFF; DMA_WRITE_INT_STATUS_OFF; DMA_READ_INT_STATUS_OFF; DMA_WRITE_INT_CLEAR_OFF; DMA_READ_INT_CLEAR_OFF",
        "Meta Validation / Acceptance Criteria": "1. Link-up polling confirms link ready. 2-10. DMA interrupt-driven completion for all 8 channels.",
        "Validation / Acceptance Criteria": "1. The link status register must reach the expected ready state with link-up bits set before proceeding. 2. The Vendor ID read from TYPE1_DEV_ID_VEND_ID_REG must return a valid device identifier. 3. The synchronization register must return the expected completion pattern confirming endpoint readiness. 4. Each DMA write channel transfer must complete successfully, confirmed by a DMA write completion interrupt being received and the DMA_WRITE_INT_STATUS_OFF register showing a non-zero channel completion status. 5. Each DMA read channel transfer must complete successfully, confirmed by a DMA read completion interrupt being received and the DMA_READ_INT_STATUS_OFF register showing a non-zero channel completion status. 6. All DMA interrupts must be properly cleared via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF after each channel completion. 7. All four write channels and all four read channels (eight total DMA transfers) must complete without hanging in the interrupt wait loop. 8. The test must complete successfully via the finish call with a pass indication.",
        "Remarks": "The test uses interrupt-driven DMA completion via the GIC interrupt controller and a Default_IRQHandler. Conditional compilation via DM0_RC, DM1_RC, DM0_EP, DM1_EP selects the link training and DMA programming path. All PCIE1 DMA register macros are unresolved in the register specification."
    },
    {
        "Index": "3",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_mem_wr_rd_test",
        "Feature": "Memory Write and Read",
        "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
        "Meta Macros": "NA",
        "Meta Arrays": "NA",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Meta Test Description": "This testcase validates PCIe memory space write and read operations through the PCIe slave ports.",
        "Test Description": "This test validates PCIe memory space write and read operations through the PCIe slave ports. It initializes the PCIe link by performing link training in x4 configuration. Cache coherency control is programmed on both PCIE0 and PCIE1 instances by modifying specific bit fields in the COHERENCY_CONTROL_3_OFF register. The test polls a link status register until the link reaches the expected ready state. Under Root Complex mode, the Vendor ID is read from TYPE1_DEV_ID_VEND_ID_REG, and TYPE1_STATUS_COMMAND_REG is configured to enable memory and bus master access. BAR programming and memory base address programming are performed. A readiness signal is written to the synchronization register. Cache coherency settings are then disabled in a staged sequence. PCIe memory write-read operations are performed at multiple memory addresses through the appropriate slave port, writing known data patterns and reading them back to verify data integrity across the PCIe link. The test concludes by polling the synchronization register until the expected completion pattern is received from the remote endpoint, confirming successful memory access.",
        "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0). 2-30. Link training, cache programming, link polling, BAR programming, memory write-read operations, synchronization polling.",
        "Test Steps / Procedure": "1. Clear the synchronization register to prepare for the test. 2. Initiate PCIe link training in x4 lane configuration for the appropriate dual-mode port based on the build configuration (Root Complex or Endpoint). 3. Program cache coherency control on both PCIE0 and PCIE1 COHERENCY_CONTROL_3_OFF registers by enabling specific bit fields for cache coherency. 4. Wait for the coherency settings to take effect. 5. Re-apply coherency control settings with all relevant bit fields enabled on both PCIE0 and PCIE1 instances. 6. Poll the link status register until the link ready condition is met (expected bit pattern confirmed). 7. In Root Complex mode, read the Vendor ID from TYPE1_DEV_ID_VEND_ID_REG to verify device presence. 8. Write to TYPE1_STATUS_COMMAND_REG to enable memory space access and bus master capability. 9. Perform BAR programming and memory base address programming for the active PCIe port. 10. In Endpoint mode, perform endpoint-specific BAR programming and memory base programming. 11. Configure non-secure protection via NIC programming. 12. Write a readiness signal to the synchronization register to notify the remote endpoint. 13. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF on both PCIE0 and PCIE1 instances in a staged sequence. 14. Wait for the cache disable settings to stabilize. 15. Perform PCIe memory write-read operations at multiple memory addresses through the appropriate slave port, writing known data patterns and reading them back. 16. Poll the synchronization register until the expected completion pattern is received from the remote endpoint, confirming successful memory access. 17. Complete the test.",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4",
        "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG",
        "Meta Validation / Acceptance Criteria": "1. Link polling confirms link ready. 2-6. Memory write-read data integrity verification.",
        "Validation / Acceptance Criteria": "1. The link status register must reach the expected ready state with the required link-up bits set before the test proceeds. 2. The Vendor ID read from TYPE1_DEV_ID_VEND_ID_REG must return a valid device identifier confirming device presence on the PCIe bus. 3. Each PCIe memory write-read operation must successfully write the known data pattern to the target memory address and read back the same value, verifying data integrity across the PCIe link. 4. The synchronization register must return the expected completion pattern to confirm that the remote endpoint has completed its operations. 5. The test must complete successfully via the finish call with a pass indication.",
        "Remarks": "Conditional compilation via DM0_RC, DM1_RC, DM0_EP, DM1_EP selects the link training path and memory write-read target addresses. The mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF macro is unresolved in the register specification."
    },
    {
        "Index": "4",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_reg_wr_rd_test",
        "Feature": "Register Write and Read",
        "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; <pcie.h>",
        "Meta Macros": "NA",
        "Meta Arrays": "rc0_ctl_addr[5]; rc1_ctl_addr[5]; ctl_default[5]; sii0_addr[3]; sii1_addr[3]; sii_default[3]; sii0_write_mask[3]; sii1_write_mask[3]; phy0_addr[3]; phy1_addr[3]; phy0_default[3]; phy1_default[3]; phy0_write_mask[3]; phy1_write_mask[3]; chk_val[6]; chk_val_phy[3]",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Meta Test Description": "This testcase validates register reset default values and register write-read functionality across multiple PCIe register groups.",
        "Test Description": "This test validates register reset default values and register write-read functionality across multiple PCIe register groups on both PCIE0 and PCIE1 controller instances. The test operates in two phases. In the first phase, all target registers are read and their values are compared against expected reset defaults to verify correct power-on initialization. The register groups tested include DBI DSP controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF), SII interface registers, and PHY registers. The PHY reset control registers are written to bring the PHY out of reset before reading PHY register defaults. In the second phase, multiple data patterns are written to each register and read back to verify write-read integrity. The data patterns include all-ones, alternating bit patterns, and zero. SII registers are written with write-mask-applied values, and PHY registers use PHY-specific test patterns with 13-bit write masks. Read-back values are compared against expected values accounting for write masks. The test passes only if all reset default checks and all write-read checks succeed with zero errors.",
        "Meta Test Steps / Procedure": "1. chk_rst_val() reads all registers and compares against defaults. 2-24. chk_rd_wr() writes patterns and reads back for verification.",
        "Test Steps / Procedure": "1. Read all DBI DSP controller registers on PCIE0 (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) and verify each returns the expected reset default value of zero. 2. Read all DBI DSP controller registers on PCIE1 and verify each returns the expected reset default value of zero. 3. Read all SII interface registers on PCIE0 and verify each returns the expected reset default value of zero. 4. Read all SII interface registers on PCIE1 and verify each returns the expected reset default value of zero. 5. Write to the PHY reset control registers on both PCIE0 and PCIE1 to bring the PHY out of reset. 6. Read all PHY registers on PCIE0 and PCIE1 with 16-bit extraction logic and verify each returns the expected reset default value of zero. 7. Begin the write-read verification phase using three data patterns: all-ones, alternating-bit, and alternating-bit-inverted. 8. For each data pattern, write the pattern to all DBI DSP controller registers on both PCIE0 and PCIE1, then read back and verify the written value matches. 9. For each data pattern, write the pattern with the appropriate write mask applied to all SII interface registers on both PCIE0 and PCIE1, then read back and verify the masked value matches. 10. For each data pattern, write the PHY-specific test pattern with the 13-bit write mask applied to all PHY registers on both PCIE0 and PCIE1, then read back with 16-bit extraction and verify the masked value matches. 11. Verify that the cumulative error count is zero, confirming all reset default checks and write-read checks passed. 12. Complete the test with a pass or fail indication based on the error count.",
        "Meta Impacted Registers": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8",
        "Impacted Registers": "MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
        "Meta Validation / Acceptance Criteria": "1. Reset default check verifies all registers read 0x0. 2-3. Write-read check verifies patterns.",
        "Validation / Acceptance Criteria": "1. All DBI DSP controller registers on both PCIE0 and PCIE1 (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) must read back the expected reset default value of zero. 2. All SII interface registers on both PCIE0 and PCIE1 must read back the expected reset default value of zero. 3. All PHY registers on both PCIE0 and PCIE1 must read back the expected reset default value of zero after 16-bit extraction. 4. For each of the three write-read test patterns, all DBI DSP controller registers must read back the exact written value. 5. For each test pattern, all SII interface registers must read back the written value masked by the applicable write mask. 6. For each PHY-specific test pattern, all PHY registers must read back the written value masked by the 13-bit write mask after 16-bit extraction. 7. The cumulative error count must be zero for the test to pass. Any mismatch in reset default or write-read verification results in test failure.",
        "Remarks": "The test exercises three distinct register groups: DBI DSP controller registers, SII interface registers, and PHY registers, each with different write mask and extraction behaviors. PHY registers use 16-bit extraction based on address alignment (upper or lower 16 bits). SII registers have per-register write masks applied during write-read verification. All PCIE1 DBI DSP and SII register macros are unresolved in the register specification."
    }
]

def generate_xlsx():
    IST = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(IST)
    timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
    filename = f'PCIE_TestPlan_{timestamp}.xlsx'
    output_dir = os.environ.get('OUTPUT_DIR', '.')
    filepath = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()

    # TestPlan sheet
    ws_tp = wb.active
    ws_tp.title = 'TestPlan'
    tp_cols = ['Index', 'SS / Module', 'Feature', 'Test Case Name', 'Test Description',
               'Speed', 'Mode', 'Memory Start Offset', 'Memory End Offset', 'Remarks',
               'Test Steps / Procedure', 'Impacted Registers', 'Validation / Acceptance Criteria',
               'Code Generation']

    # MetaData sheet
    ws_md = wb.create_sheet('MetaData')
    md_cols = ['Index', 'Test Case Name', 'Meta Test Description', 'Meta Test Steps / Procedure',
               'Meta Impacted Registers', 'Meta Validation / Acceptance Criteria',
               'Meta Headers', 'Meta Macros', 'Meta Arrays']

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    wrap_align = Alignment(wrap_text=True, vertical='top')

    # Write TestPlan headers
    for col_idx, col_name in enumerate(tp_cols, 1):
        cell = ws_tp.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align

    # Write TestPlan data
    for row_idx, row_data in enumerate(json_data, 2):
        for col_idx, col_name in enumerate(tp_cols, 1):
            val = row_data.get(col_name, '')
            cell = ws_tp.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap_align

    # Write MetaData headers
    for col_idx, col_name in enumerate(md_cols, 1):
        cell = ws_md.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align

    # Write MetaData data
    for row_idx, row_data in enumerate(json_data, 2):
        for col_idx, col_name in enumerate(md_cols, 1):
            val = row_data.get(col_name, '')
            cell = ws_md.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap_align

    # Auto-size columns
    for ws in [ws_tp, ws_md]:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            max_len = max(max_len, len(line))
            width = min(max(max_len + 2, 12), 60)
            ws.column_dimensions[col_letter].width = width

    # Freeze first row
    ws_tp.freeze_panes = 'A2'
    ws_md.freeze_panes = 'A2'

    # Set MetaData sheet to veryHidden
    ws_md.sheet_state = 'veryHidden'

    # Save
    wb.save(filepath)
    print(f'GENERATED:{filepath}')
    print(f'FILENAME:{filename}')

    # Validate
    vwb = load_workbook(filepath)
    sheets = vwb.sheetnames
    assert 'TestPlan' in sheets, 'TestPlan sheet missing'
    assert 'MetaData' in sheets, 'MetaData sheet missing'
    tp_rows = vwb['TestPlan'].max_row - 1
    md_rows = vwb['MetaData'].max_row - 1
    fsize = os.path.getsize(filepath)
    assert fsize > 0, 'File is empty'
    print(f'VALIDATION:PASSED')
    print(f'ROWS_TP:{tp_rows}')
    print(f'ROWS_MD:{md_rows}')
    print(f'SIZE:{fsize}')
    return filepath, filename

if __name__ == '__main__':
    generate_xlsx()
