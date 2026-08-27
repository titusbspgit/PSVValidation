#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
import os, sys, base64

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
filename = f'PCIE_TestPlan_{timestamp}.xlsx'
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

wb = openpyxl.Workbook()
ws_tp = wb.active
ws_tp.title = 'TestPlan'
ws_md = wb.create_sheet('MetaData')

tp_cols = ['Index','SS / Module','Feature','Test Case Name','Test Description','Speed','Mode','Memory Start Offset','Memory End Offset','Remarks','Test Steps / Procedure','Impacted Registers','Validation / Acceptance Criteria','Code Generation']
md_cols = ['Index','Test Case Name','Meta Test Description','Meta Test Steps / Procedure','Meta Impacted Registers','Meta Validation / Acceptance Criteria','Meta Headers','Meta Macros','Meta Arrays']

hf = Font(bold=True, color='FFFFFF', size=11)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ha = Alignment(horizontal='center', vertical='top', wrap_text=True)
ca = Alignment(vertical='top', wrap_text=True)
tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for ci, cn in enumerate(tp_cols, 1):
    c = ws_tp.cell(row=1, column=ci, value=cn)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
for ci, cn in enumerate(md_cols, 1):
    c = ws_md.cell(row=1, column=ci, value=cn)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

ts1 = "Initialization:\n1. Clear the synchronization control register to prepare for the test.\n2. Initiate PCIe link training in x4 mode for the applicable dual-mode controller configuration (Root Complex or Endpoint).\n\nConfiguration:\n1. Enable cache coherency by programming the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances, setting the master read and write cache mode and value fields.\n2. Perform a combined cache coherency programming pass for both controller instances.\n3. Enable bus master, memory space, and I/O space access by writing to TYPE1_STATUS_COMMAND_REG.\n4. Write enable values to system-level control registers.\n5. Disable cache coherency by clearing the cache value fields in COHERENCY_CONTROL_3_OFF for both controller instances.\n\nExecution:\n1. Wait for the cache programming to take effect.\n2. Poll the SII0 link status register until the link-up condition is confirmed.\n3. Poll the SII1 link status register until the link-up condition is confirmed.\n4. Read the device vendor ID from TYPE1_DEV_ID_VEND_ID_REG on the first slave port and verify it is valid.\n5. Execute memory base address programming for both dual-mode controllers.\n6. Wait for the disable to take effect and perform a combined disable pass.\n7. Determine BAR sizes on the second slave port by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, then reading back.\n8. Program actual base addresses into the BAR registers on the second slave port and verify by reading back.\n9. Repeat BAR size determination and base address programming on the first slave port.\n10. Poll the synchronization register until the expected completion value is received.\n11. Report test completion."

ts2 = "Initialization:\n1. Initialize the test environment and define target register groups for both PCIe controller instances: DBI DSP control registers, SII registers, and PHY registers.\n\nConfiguration:\n1. Write to the PHY reset control registers on both controller instances to bring the PHY out of reset.\n\nExecution:\n1. Read all DBI DSP control registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on both controller instances and verify they contain the expected reset default values of zero.\n2. Read all SII transmit header and PHY control registers on both controller instances and verify they contain the expected reset default values of zero.\n3. Read all PHY lane registers on both controller instances with 16-bit extraction based on address alignment and verify they contain the expected reset default values.\n4. Begin the write-read verification phase using multiple test patterns.\n5. For each test pattern, write the pattern to all DBI DSP control registers on both controller instances.\n6. Write the masked test pattern to all SII registers on both controller instances, applying the appropriate write masks.\n7. Reassert PHY reset control on both controller instances before writing PHY registers.\n8. Write PHY-specific test patterns to all PHY lane registers on both controller instances, applying PHY write masks.\n9. Read back all DBI DSP control registers on both controller instances and verify the read data matches the written pattern.\n10. Read back all SII registers on both controller instances and verify the read data matches the masked written pattern.\n11. Read back all PHY lane registers on both controller instances with 16-bit extraction and verify the read data matches the masked PHY pattern.\n12. Repeat steps 5 through 11 for all remaining test patterns.\n13. Report test pass if all register comparisons succeed across all groups and patterns; report test fail if any mismatch is detected."

td1 = 'Verifies PCIe device enumeration by performing link training on dual-mode controllers, programming cache coherency control via the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances, polling the SII link status registers until link-up is confirmed, reading the device vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enabling bus master and memory space via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, configuring BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports by writing all-ones to determine BAR size and then programming actual base addresses, disabling cache coherency, writing system-level control registers, and polling a synchronization register until the expected completion value is received.'
td2 = 'Verifies the reset default values and write-read integrity of PCIe DBI DSP registers, SII registers, and PHY registers across both PCIe controller instances. The test first reads all target registers and compares against expected default values of zero. It then brings the PHY out of reset by writing to the PHY reset control registers. PHY registers are read with 16-bit extraction based on address alignment and verified against defaults. In the write-read phase, multiple test patterns are written to MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, and UTILITY_OFF on both controller instances, along with SII transmit header and PHY control registers, and PHY lane registers. Each register is read back and compared against the written value, applying write masks where applicable. The PHY reset control is reasserted before each PHY write-read cycle. The test reports pass if all comparisons succeed across all register groups and all test patterns, and fail if any mismatch is detected.'

rm1 = 'The testcase uses compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select the link training and enumeration path, meaning the active code path depends on the build configuration. Polling is used on SII link status registers and a synchronization register with wait delays between iterations. Cache coherency is enabled at the start and disabled after memory base programming. Several system-level control registers and the SII link status register could not be mapped to named registers in the specification. The PCIE1 instance of the coherency control register could not be resolved through the available header files. The BAR register offsets are used on both slave port 0 and slave port 1 through different accessor functions.'
rm2 = 'The testcase covers both PCIE0 and PCIE1 controller instances with identical register groups and test patterns. PHY reset control registers are written before each PHY register access to ensure the PHY is in the correct state for register read and write operations. PHY registers use 16-bit access with extraction logic based on address alignment. Write masks are applied to SII registers and PHY registers to account for read-only or reserved bit fields. Only three of the six defined DBI DSP test patterns are used in the write-read loop. Several registers including the PCIE1 DBI DSP registers, all SII registers, PHY reset control registers, and PHY lane registers could not be mapped to named registers in the available specification. The test uses two separate error counters for different register groups and combines them for the final pass or fail determination.'

ir1 = 'COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG'
ir2 = 'MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF'

vc1 = 'The test passes when all of the following conditions are met: (1) PCIe link training completes successfully for the configured dual-mode controllers. (2) The SII0 and SII1 link status polling loops exit successfully, confirming link-up on both PCIe instances. (3) The vendor ID read from TYPE1_DEV_ID_VEND_ID_REG returns a valid non-zero value. (4) The TYPE1_STATUS_COMMAND_REG is successfully written to enable bus master and memory space access. (5) BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) on both slave ports return expected size masks when written with all-ones and retain programmed base addresses when read back. (6) The synchronization register polling loop exits when the expected completion value is received. (7) The test completes by calling finish(0) indicating a pass.'
vc2 = 'The test passes when all of the following conditions are met: (1) All DBI DSP control registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) on both controller instances read back the expected reset default value of zero during the reset value check phase. (2) All SII transmit header and PHY control registers on both controller instances read back the expected reset default value of zero. (3) All PHY lane registers on both controller instances read back the expected reset default value of zero after PHY reset is deasserted, with correct 16-bit extraction applied. (4) During the write-read phase, all DBI DSP control registers on both controller instances read back the exact test pattern that was written for each of the three test pattern iterations. (5) All SII registers on both controller instances read back the correctly masked test pattern for each iteration. (6) All PHY lane registers on both controller instances read back the correctly masked PHY test pattern with 16-bit extraction for each iteration. (7) The test completes with both error counters at zero, resulting in a pass indication.'

tp_data = [
    ['1','PCIE','PCIe Device Enumeration and Link Training','pcie_device_enumerate_test',td1,'','','','',rm1,ts1,ir1,vc1,''],
    ['2','PCIE','PCIe Register Write-Read Verification','pcie_reg_wr_rd_test',td2,'','','','',rm2,ts2,ir2,vc2,''],
]
for ri, rd in enumerate(tp_data, 2):
    for ci, val in enumerate(rd, 1):
        c = ws_tp.cell(row=ri, column=ci, value=val)
        c.alignment = ca; c.border = tb

mtd1 = 'The testcase performs PCIe device enumeration across two dual-mode controllers (DM0 and DM1). It begins by writing 0x0 to 0xE6004100 to clear a control register. Link training is initiated via link_training_dm0_x4(4) or link_training_dm1_x4(4) depending on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). Cache coherency programming is performed by read-modify-write operations on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, setting bit fields CFG_MSTR_ARCACHE_MODE (bits 3:6), CFG_MSTR_AWCACHE_MODE (bits 11:14), CFG_MSTR_ARCACHE_VALUE (bits 19:22), and CFG_MSTR_AWCACHE_VALUE (bits 27:30) to 0xF using set_data(). A wait_on(20) delay is inserted between cache programming phases. The SII0 link status register at offset 0xC0 is polled via read_sii0_reg(0xC0) until (data_rd & 0xD1) == 0xD1, confirming link-up. Similarly, SII1 link status is polled via read_sii1_reg(0xC0). Under DM0_RC, the vendor ID is read from read_pcie_slv0_reg(0x0), the command register is written via write_pcie_slv0_reg(0x4,0x7), and memory base programming is performed via mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). System-level registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, and 0xE6900034 are written with 0x1. Cache coherency is then disabled by writing 0x0 to the ARCACHE_VALUE and AWCACHE_VALUE fields of both PCIE0 and PCIE1 COHERENCY_CONTROL_3_OFF registers. After a wait_on(30), BAR registers (offsets 0x10 through 0x24) on both PCIe slave ports (slv0 and slv1) are written with 0xFFFFFFFF, read back, then programmed with specific base addresses (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) and read back again. Finally, the test polls 0xE6004100 via read_reg() until the value equals 0x12345678, then calls finish(0) to indicate test completion.'
mts1 = '1. Write 0x0 to 0xE6004100 to clear the synchronization/control register. 2. Invoke link_training_dm0_x4(4) or link_training_dm1_x4(4) based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to initiate PCIe link training in x4 mode. 3. Perform cache coherency enable programming: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting bits [3:6] (CFG_MSTR_ARCACHE_MODE) and bits [11:14] (CFG_MSTR_AWCACHE_MODE) to 0xF using set_data(). 4. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting bits [19:22] (CFG_MSTR_ARCACHE_VALUE) and bits [27:30] (CFG_MSTR_AWCACHE_VALUE) to 0xF. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. Call wait_on(20) for delay. 7. Perform combined cache programming: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting all four cache fields (bits 3:6, 11:14, 19:22, 27:30) to 0xF in a single sequence. 8. Repeat step 7 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 9. Read SII0 link status via read_sii0_reg(0xC0). 10. Call non_secure_prot_nic() for NIC protection setup. 11. Poll read_sii0_reg(0xC0) in a while loop until (data_rd & 0xD1) == 0xD1 confirming PCIE0 link-up. 12. Read SII1 link status via read_sii1_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1 confirming PCIE1 link-up. 13. Under DM0_RC: read vendor ID via read_pcie_slv0_reg(0x0) and print. 14. Write 0x7 to command register via write_pcie_slv0_reg(0x4, 0x7) to enable bus master, memory space, and I/O space. 15. Call mem_base_program_dm0_x4() and mem_base_program_dm1_x4() for memory base programming. 16. Call wait_on(10). 17. Write 0x1 to system registers 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034. 18. Perform cache coherency disable: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, setting CFG_MSTR_ARCACHE_VALUE (bits 19:22) and CFG_MSTR_AWCACHE_VALUE (bits 27:30) to 0x0. 19. Call wait_on(10) then perform combined disable for both controllers. 20. Call wait_on(30). 21. Write 0xFFFFFFFF to BAR registers on slv1 (offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24) to determine BAR sizes. 22. Read back all slv1 BAR registers. 23. Program slv1 BAR registers with actual base addresses (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000). 24. Read back all slv1 BAR registers to verify. 25. Repeat steps 21-24 for slv0 BAR registers. 26. Call wait_on(10). 27. Poll read_reg(0xE6004100) until value equals 0x12345678 with wait_on(5) between iterations. 28. Call finish(0) to indicate test pass.'
mir1 = '0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24'
mvc1 = vc1

mtd2 = 'The testcase verifies reset default values and write-read functionality of PCIe registers across three register groups on both PCIE0 and PCIE1 controller instances. It includes headers test_common.h and pcie.h. Global arrays define the register addresses: rc0_ctl_addr[5] holds {mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF}, rc1_ctl_addr[5] holds the corresponding PCIE1 macros {mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE1_DBI_DSP_UTILITY_OFF}, sii0_addr[3] holds {mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2, mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3, mizar_PCIE0_SII_PHY_CONTROL_23}, sii1_addr[3] holds the corresponding PCIE1 SII macros, phy0_addr[3] holds {0xE68860B8, 0xE68862B8, 0xE68864B8}, and phy1_addr[3] holds {0xE68A60B8, 0xE68A62B8, 0xE68A64B8}. Default value arrays ctl_default[5], sii_default[3], phy0_default[3], and phy1_default[3] are all initialized to 0x0. Write mask arrays sii0_write_mask[3] and sii1_write_mask[3] are {0xFFFFFFFF, 0xFFFFFFFF, 0xF000F}, and phy0_write_mask[3] and phy1_write_mask[3] are {0x1FFF, 0x1FFF, 0x1FFF}. The test_case() function calls chk_rst_val() followed by chk_rd_wr(), then calls finish(err2 || err1).'
mts2 = '1. Include headers test_common.h and pcie.h. 2. Initialize global arrays. 3. Initialize default arrays. 4. Initialize write mask arrays. 5. Initialize error counters err1=0, err2=0. 6. Enter test_case(), call chk_rst_val(). 7. In chk_rst_val(): loop i=0 to 4, read_reg(rc0_ctl_addr[i]), compare data_rd against ctl_default[i], increment err1 on mismatch. 8. Loop i=0 to 4, read_reg(rc1_ctl_addr[i]), compare data_rd against ctl_default[i], increment err2 on mismatch. 9. Loop i=0 to 2, read_reg(sii0_addr[i]), compare data_rd against sii_default[i], increment err2 on mismatch. 10. Loop i=0 to 2, read_reg(sii1_addr[i]), compare data_rd against sii_default[i], increment err2 on mismatch. 11. Write mizar_PCIE0_SII_PHY_RST_CONTROL with 0x01203000. 12. Write mizar_PCIE1_SII_PHY_RST_CONTROL with 0x01203000. 13-32. Write-read verification with test patterns.'
mir2 = 'mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8'
mvc2 = vc2

md_data = [
    ['1','pcie_device_enumerate_test',mtd1,mts1,mir1,mvc1,'','',''],
    ['2','pcie_reg_wr_rd_test',mtd2,mts2,mir2,mvc2,'','',''],
]
for ri, rd in enumerate(md_data, 2):
    for ci, val in enumerate(rd, 1):
        c = ws_md.cell(row=ri, column=ci, value=val)
        c.alignment = ca; c.border = tb

tp_w = {'A':8,'B':15,'C':40,'D':35,'E':60,'F':10,'G':10,'H':20,'I':20,'J':50,'K':90,'L':60,'M':70,'N':18}
for col_l, w in tp_w.items():
    ws_tp.column_dimensions[col_l].width = w
md_w = {'A':8,'B':35,'C':80,'D':80,'E':80,'F':80,'G':40,'H':40,'I':40}
for col_l, w in md_w.items():
    ws_md.column_dimensions[col_l].width = w

ws_tp.freeze_panes = 'A2'
ws_md.freeze_panes = 'A2'
for ri in range(2, 4):
    ws_tp.row_dimensions[ri].height = 400
    ws_md.row_dimensions[ri].height = 300

dv = DataValidation(type='list', formula1='"Required,Not Required"', allow_blank=True, showDropDown=False)
dv.error = 'Please select Required or Not Required'
dv.errorTitle = 'Invalid Input'
dv.prompt = 'Select Code Generation status'
dv.promptTitle = 'Code Generation'
dv.sqref = 'N2:N1000'
ws_tp.add_data_validation(dv)

ws_md.sheet_state = 'veryHidden'
wb.active = 0
wb.save(output_path)

if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
    wb2 = openpyxl.load_workbook(output_path)
    assert 'TestPlan' in wb2.sheetnames
    assert 'MetaData' in wb2.sheetnames
    with open(output_path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode('ascii')
    print(f'FILENAME={filename}')
    print(f'SIZE={os.path.getsize(output_path)}')
    print(f'B64LEN={len(b64)}')
    # Write b64 to a text file for retrieval
    b64_path = output_path + '.b64'
    with open(b64_path, 'w') as f:
        f.write(b64)
    print(f'B64PATH={b64_path}')
else:
    print('FAILURE')
    sys.exit(1)
