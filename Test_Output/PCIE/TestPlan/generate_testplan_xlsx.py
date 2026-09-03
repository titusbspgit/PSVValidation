#!/usr/bin/env python3
"""Generate PCIE_TestPlan XLSX from JSON data.
Usage: python generate_testplan_xlsx.py
Requires: pip install openpyxl
"""
import json
import os
import sys
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(SCRIPT_DIR, 'PCIE_TestPlan_20250716_100530_data.json')

if not os.path.exists(DATA_FILE):
    json_files = [f for f in os.listdir(SCRIPT_DIR) if f.endswith('_data.json')]
    if json_files:
        DATA_FILE = os.path.join(SCRIPT_DIR, sorted(json_files)[-1])
    else:
        print('ERROR: No data JSON file found')
        sys.exit(1)

with open(DATA_FILE, 'r') as f:
    testcases = json.load(f)

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
output_file = os.path.join(SCRIPT_DIR, f'PCIE_TestPlan_{timestamp}.xlsx')

# === STYLES ===
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
cell_align = Alignment(vertical='top', wrap_text=True)
meta_header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')

# === TESTPLAN SHEET COLUMNS ===
TP_COLUMNS = [
    'Index', 'SS / Module', 'Feature', 'Test Case Name',
    'Test Description', 'Speed', 'Mode',
    'Memory Start Offset', 'Memory End Offset',
    'Remarks', 'Test Steps / Procedure',
    'Impacted Registers', 'Validation / Acceptance Criteria',
    'Code Generation'
]

# === METADATA SHEET COLUMNS ===
MD_COLUMNS = [
    'Index', 'Test Case Name',
    'Meta Test Description', 'Meta Test Steps / Procedure',
    'Meta Impacted Registers', 'Meta Validation / Acceptance Criteria',
    'Meta Headers', 'Meta Macros', 'Meta Arrays'
]

# === CREATE WORKBOOK ===
wb = Workbook()

# --- TestPlan Sheet ---
ws_tp = wb.active
ws_tp.title = 'TestPlan'

for col_idx, col_name in enumerate(TP_COLUMNS, 1):
    cell = ws_tp.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row_idx, tc in enumerate(testcases, 2):
    for col_idx, col_name in enumerate(TP_COLUMNS, 1):
        value = tc.get(col_name, '')
        cell = ws_tp.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_align
        cell.border = thin_border

for col_idx, col_name in enumerate(TP_COLUMNS, 1):
    col_letter = ws_tp.cell(row=1, column=col_idx).column_letter
    max_len = len(col_name)
    for row in range(2, len(testcases) + 2):
        val = str(ws_tp.cell(row=row, column=col_idx).value or '')
        lines = val.split('\n')
        for line in lines:
            if len(line) > max_len:
                max_len = len(line)
    ws_tp.column_dimensions[col_letter].width = min(max_len + 4, 60)

ws_tp.freeze_panes = 'A2'
ws_tp.auto_filter.ref = ws_tp.dimensions

# --- MetaData Sheet ---
ws_md = wb.create_sheet('MetaData')

for col_idx, col_name in enumerate(MD_COLUMNS, 1):
    cell = ws_md.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = meta_header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row_idx, tc in enumerate(testcases, 2):
    for col_idx, col_name in enumerate(MD_COLUMNS, 1):
        value = tc.get(col_name, '')
        cell = ws_md.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_align
        cell.border = thin_border

for col_idx, col_name in enumerate(MD_COLUMNS, 1):
    col_letter = ws_md.cell(row=1, column=col_idx).column_letter
    ws_md.column_dimensions[col_letter].width = 50

ws_md.freeze_panes = 'A2'
ws_md.sheet_state = 'veryHidden'

# === SAVE ===
wb.save(output_file)

# === VALIDATE ===
file_size = os.path.getsize(output_file)
wb_check = load_workbook(output_file)
sheets = wb_check.sheetnames
tp_rows = wb_check['TestPlan'].max_row - 1
md_rows = wb_check['MetaData'].max_row - 1
wb_check.close()

print(f'SUCCESS: Generated {output_file}')
print(f'  File size: {file_size} bytes')
print(f'  Sheets: {sheets}')
print(f'  TestPlan rows: {tp_rows}')
print(f'  MetaData rows: {md_rows}')
print(f'  Timestamp (IST): {now_ist.strftime("%Y-%m-%d %H:%M:%S")}')
print(f'  Validation: PASSED')
