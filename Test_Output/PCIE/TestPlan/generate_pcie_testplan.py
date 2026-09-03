#!/usr/bin/env python3
"""Generate PCIE TestPlan Excel workbook from JSON data.
Run: python generate_pcie_testplan.py
Requires: openpyxl (pip install openpyxl)
"""
import json, os, sys
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
TIMESTAMP = now_ist.strftime("%Y%m%d_%H%M%S")
FILENAME = f"PCIE_TestPlan_{TIMESTAMP}.xlsx"

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "PCIE_TestPlan_20250703_190000_data.json")

TP_COLS = [
    "Index", "SS / Module", "Feature", "Test Case Name",
    "Test Description", "Speed", "Mode", "Memory Start Offset",
    "Memory End Offset", "Remarks", "Test Steps / Procedure",
    "Impacted Registers", "Validation / Acceptance Criteria", "Code Generation"
]

MD_COLS = [
    "Index", "Test Case Name", "Meta Test Description",
    "Meta Test Steps / Procedure", "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria", "Meta Headers",
    "Meta Macros", "Meta Arrays"
]

HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

def auto_width(ws, max_w=60):
    for col_cells in ws.columns:
        mx = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                longest = max(len(l) for l in lines) if lines else 0
                mx = max(mx, longest)
        ws.column_dimensions[col_letter].width = min(max(mx + 2, 12), max_w)

def build():
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"Loaded {len(data)} test cases from {DATA_FILE}")

    wb = Workbook()
    # --- TestPlan sheet ---
    ws_tp = wb.active
    ws_tp.title = "TestPlan"
    ws_tp.append(TP_COLS)
    for r in data:
        row = [r.get(c, "") for c in TP_COLS]
        ws_tp.append(row)
    for cell in ws_tp[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = WRAP
    for row in ws_tp.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    ws_tp.freeze_panes = "A2"
    auto_width(ws_tp)

    # --- MetaData sheet ---
    ws_md = wb.create_sheet("MetaData")
    ws_md.append(MD_COLS)
    for r in data:
        row = [r.get(c, "") for c in MD_COLS]
        ws_md.append(row)
    for cell in ws_md[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = WRAP
    for row in ws_md.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    ws_md.freeze_panes = "A2"
    auto_width(ws_md)
    ws_md.sheet_state = "veryHidden"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), FILENAME)
    wb.save(out)
    print(f"Saved: {out}  ({os.path.getsize(out)} bytes)")

    # Validate
    vwb = load_workbook(out)
    assert "TestPlan" in vwb.sheetnames, "Missing TestPlan sheet"
    assert "MetaData" in vwb.sheetnames, "Missing MetaData sheet"
    tp_rows = vwb["TestPlan"].max_row - 1
    md_rows = vwb["MetaData"].max_row - 1
    print(f"Validation PASSED  TestPlan rows={tp_rows}  MetaData rows={md_rows}")
    return out, tp_rows, md_rows

if __name__ == "__main__":
    build()
