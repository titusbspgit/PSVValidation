#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json, os

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

data = json.loads(open('/tmp/testcases.json').read())

# ── Main Sheet ──
ws_main = wb.active
ws_main.title = "Main"
main_cols = ["Index","SS / Module","Test Case Name","Feature","Test Description",
             "Test Steps / Procedure","Impacted Registers","Validation / Acceptance Criteria","Remarks"]

for c, col_name in enumerate(main_cols, 1):
    cell = ws_main.cell(row=1, column=c, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for r, tc in enumerate(data, 2):
    for c, col_name in enumerate(main_cols, 1):
        cell = ws_main.cell(row=r, column=c, value=tc.get(col_name, "NA"))
        cell.alignment = cell_align
        cell.border = thin_border

main_widths = [8, 15, 35, 30, 60, 60, 40, 60, 50]
for i, w in enumerate(main_widths, 1):
    ws_main.column_dimensions[get_column_letter(i)].width = w

ws_main.auto_filter.ref = f"A1:{get_column_letter(len(main_cols))}1"
ws_main.freeze_panes = "A2"

# ── MetaData Sheet ──
ws_meta = wb.create_sheet("MetaData")
meta_cols = ["Index","SS / Module","Test Case Name","Feature",
             "Meta Test Description","Meta Test Steps / Procedure","Meta Impacted Registers"]

for c, col_name in enumerate(meta_cols, 1):
    cell = ws_meta.cell(row=1, column=c, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for r, tc in enumerate(data, 2):
    for c, col_name in enumerate(meta_cols, 1):
        cell = ws_meta.cell(row=r, column=c, value=tc.get(col_name, "NA"))
        cell.alignment = cell_align
        cell.border = thin_border

meta_widths = [8, 15, 35, 30, 80, 80, 60]
for i, w in enumerate(meta_widths, 1):
    ws_meta.column_dimensions[get_column_letter(i)].width = w

ws_meta.auto_filter.ref = f"A1:{get_column_letter(len(meta_cols))}1"
ws_meta.freeze_panes = "A2"

out = "/tmp/PCIE_TestPlan.xlsx"
wb.save(out)
print(f"Saved: {out} ({os.path.getsize(out)} bytes)")
