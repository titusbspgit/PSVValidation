#!/usr/bin/env python3
"""
Agent 7 - Excel Workbook Generator for PCIE TestPlan
Generates PCIE_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx using openpyxl
"""
import json
import os
import sys
from datetime import datetime, timezone, timedelta

# openpyxl imports
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# INPUT DATA - Final Aggregated Agent 6 JSON
# ============================================================
json_data = [
    {
        "Index": "1",
        "SS / Module": "PCIE",
        "Test Case Name": "pcie_device_enumerate_test",
        "Feature": "Device Enumeration",
        "Meta Headers": "<stdlib.h>; <stdio.h>; <test_common.h>; \"pcie.h\"",
        "Meta Macros": "NA",
        "Meta Arrays": "NA",
        "Speed": "NA",
        "Mode": "NA",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Meta Test Description": "This testcase performs PCIe device enumeration. It begins by writing 0x0 to address 0xE6004100. Depending on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP), it invokes link_training_dm0_x4(4) or link_training_dm1_x4(4) for link training. Cache programming is performed by reading mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, using set_data to modify bit fields [11:14], [3:6], [27:30], [19:22] to 0xf, and writing back. After wait_on(20), the same coherency registers are read-modify-written again. SII0 and SII1 link status is polled at offset 0xC0 until bits match 0xD1 pattern ((data_rd & 0xD1) == 0xD1). Under DM0_RC, the Vendor ID is read from PCIe slave 0 at offset 0x0, the command register at offset 0x4 is written with 0x7, and mem_base_program_dm0_x4() and mem_base_program_dm1_x4() are called. Registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034 are written with 0x1. Cache disable programming is performed by modifying the coherency control registers with 0x0 for bit fields [19:22] and [27:30]. After wait_on(30), PCIe slave 1 BAR registers at offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24 are written with 0xFFFFFFFF, read back, then written with specific values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) and read back. The same sequence is repeated for PCIe slave 0. Finally, address 0xE6004100 is polled until it reads 0x12345678, with wait_on(5) between iterations, then finish(0) is called.",
        "Test Description": "This test performs PCIe device enumeration by training the PCIe link, programming cache coherency control registers for both PCIE0 and PCIE1, polling SII link status until the link is established, reading the Vendor ID from the device identification register, enabling bus master and memory space via the status/command register, programming memory base addresses, and then enumerating BAR registers (BAR0, BAR1, secondary bus/primary bus, IO limit/base, memory limit/base, and prefetchable memory limit/base) on both PCIe slave ports by first writing all-ones to determine BAR sizes, reading back, then programming actual base addresses and verifying by reading back. The test concludes by polling a synchronization register until a completion pattern is received.",
        "Meta Test Steps / Procedure": "1. write_reg(0xE6004100, 0x0) - clear synchronization register. 2. Conditional link training: link_training_dm0_x4(4) or link_training_dm1_x4(4) based on DM0_RC/DM1_RC/DM0_EP/DM1_EP defines. 3. Cache programming: read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [11:14]=0xf, [3:6]=0xf, write back. 4. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set_data bits [27:30]=0xf, [19:22]=0xf, write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. wait_on(20). 7. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF with all four bit fields set to 0xf. 8. Read-modify-write mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF similarly. 9. Duplicate link training and cache programming block (repeated in source). 10. read_sii0_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1. 11. read_sii1_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1. 12. non_secure_prot_nic() called. 13. Under DM0_RC: read_pcie_slv0_reg(0x0) to get Vendor ID. 14. write_pcie_slv0_reg(0x4, 0x7) to enable command register. 15. mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). 16. wait_on(10). 17. write_reg(0xE690000C, 0x1), write_reg(0xE6900010, 0x1), write_reg(0xE6900014, 0x1), write_reg(0xE6900018, 0x1), write_reg(0xE6900030, 0x1), write_reg(0xE6900034, 0x1). 18. Cache disable: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF with bits [19:22]=0x0, [27:30]=0x0. 19. Repeat cache disable for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 20. wait_on(10), then consolidated cache disable write for both PCIE0 and PCIE1. 21. wait_on(30). 22. write_pcie_slv1_reg offsets 0x10-0x24 with 0xFFFFFFFF. 23. read_pcie_slv1_reg offsets 0x10-0x24. 24. write_pcie_slv1_reg offsets 0x10-0x24 with specific base addresses. 25. read_pcie_slv1_reg offsets 0x10-0x24. 26. Repeat steps 22-25 for pcie_slv0. 27. wait_on(10). 28. Poll read_reg(0xE6004100) until value equals 0x12345678, with wait_on(5) between iterations. 29. finish(0).",
        "Test Steps / Procedure": "1. Clear the synchronization register to prepare for the test.\n2. Perform PCIe link training based on the configured mode (Root Complex or Endpoint for DM0/DM1).\n3. Program cache coherency control registers for both PCIE0 and PCIE1 by enabling specific bit fields.\n4. Wait for coherency settings to take effect.\n5. Re-apply cache coherency settings for both PCIE0 and PCIE1.\n6. Poll SII0 link status register until the link is established (expected status pattern detected).\n7. Poll SII1 link status register until the link is established.\n8. Configure non-secure protection settings.\n9. Read the device Vendor ID from the TYPE1_DEV_ID_VEND_ID_REG on PCIe slave 0.\n10. Write to TYPE1_STATUS_COMMAND_REG to enable bus master, memory space, and IO space.\n11. Program memory base addresses for both DM0 and DM1.\n12. Enable configuration registers by writing to six system-level control registers.\n13. Disable cache coherency by clearing specific bit fields in the coherency control registers for both PCIE0 and PCIE1.\n14. Wait for cache disable to take effect.\n15. Write all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG on PCIe slave 1 to determine BAR sizes.\n16. Read back all BAR registers on PCIe slave 1 to capture size information.\n17. Program actual base addresses into the BAR registers on PCIe slave 1.\n18. Read back all BAR registers on PCIe slave 1 to verify programming.\n19. Repeat BAR size detection, programming, and verification for PCIe slave 0.\n20. Poll the synchronization register until the expected completion value is received, confirming test completion.",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
        "Impacted Registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
        "Meta Validation / Acceptance Criteria": "1. SII0 link status polling: read_sii0_reg(0xC0) is polled in a while loop until (data_rd & 0xD1) == 0xD1, confirming link is up. 2. SII1 link status polling: read_sii1_reg(0xC0) is polled until (data_rd & 0xD1) == 0xD1. 3. Vendor ID read: rd_wr_data1 = read_pcie_slv0_reg(0x0) is printed to verify device presence. 4. BAR register read-back: After writing 0xFFFFFFFF to offsets 0x10-0x24 on both slave ports, the values are read back to determine BAR sizes. After programming actual base addresses, values are read back to verify correct programming. 5. Synchronization polling: read_reg(0xE6004100) is polled until data_rd == 0x12345678, confirming the remote side has completed its operations. 6. finish(0) is called to indicate test pass.",
        "Validation / Acceptance Criteria": "1. SII0 and SII1 link status registers must report the expected link-up pattern before proceeding.\n2. The Vendor ID read from TYPE1_DEV_ID_VEND_ID_REG must return a valid device identifier confirming device presence.\n3. BAR registers (BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, PREF_MEM_LIMIT_PREF_MEM_BASE_REG) must correctly reflect size information after writing all-ones and must retain programmed base addresses after final configuration.\n4. The synchronization register must eventually return the expected completion value, confirming the remote endpoint has completed its operations.\n5. The test must complete successfully with a pass indication.",
        "Remarks": "The test uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select between Root Complex and Endpoint modes for link training. The source contains a duplicated block of link training and cache programming code. SII link status polling uses a busy-wait loop. Cache coherency is enabled before enumeration and disabled afterward. A synchronization mechanism using a polling register is used to coordinate with the remote PCIe endpoint. Multiple wait delays (wait_on) are used throughout the test for timing."
    }
]

# ============================================================
# CONFIGURATION
# ============================================================
IP_NAME = "PCIE"
IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime("%Y%m%d_%H%M%S")
filename = f"{IP_NAME}_TestPlan_{timestamp}.xlsx"
output_dir = "/tmp"
filepath = os.path.join(output_dir, filename)

# ============================================================
# COLUMN DEFINITIONS
# ============================================================
testplan_columns = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation"
]

metadata_columns = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria"
]

# Long text columns for width limits
long_text_cols_tp = {
    "Test Description", "Remarks", "Test Steps / Procedure",
    "Validation / Acceptance Criteria"
}
long_text_cols_md = {
    "Meta Headers", "Meta Macros", "Meta Arrays",
    "Meta Test Description", "Meta Test Steps / Procedure",
    "Meta Impacted Registers", "Meta Validation / Acceptance Criteria"
}

# ============================================================
# STYLES
# ============================================================
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ============================================================
# HELPER: safe value
# ============================================================
def safe_val(row, key):
    val = row.get(key, None)
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return "NA"
    return val

# ============================================================
# CREATE WORKBOOK
# ============================================================
wb = Workbook()

# --- TestPlan Sheet ---
ws_tp = wb.active
ws_tp.title = "TestPlan"

# Write header
for col_idx, col_name in enumerate(testplan_columns, 1):
    cell = ws_tp.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data rows
for row_idx, row_data in enumerate(json_data, 2):
    for col_idx, col_name in enumerate(testplan_columns, 1):
        if col_name == "Code Generation":
            val = safe_val(row_data, "Code Generation")
        else:
            val = safe_val(row_data, col_name)
        cell = ws_tp.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = cell_alignment
        cell.border = thin_border

# --- MetaData Sheet ---
ws_md = wb.create_sheet(title="MetaData")

# Write header
for col_idx, col_name in enumerate(metadata_columns, 1):
    cell = ws_md.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data rows
for row_idx, row_data in enumerate(json_data, 2):
    for col_idx, col_name in enumerate(metadata_columns, 1):
        val = safe_val(row_data, col_name)
        cell = ws_md.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = cell_alignment
        cell.border = thin_border

# ============================================================
# SET METADATA SHEET TO veryHidden
# ============================================================
ws_md.sheet_state = "veryHidden"

# ============================================================
# FREEZE FIRST ROW
# ============================================================
ws_tp.freeze_panes = "A2"
ws_md.freeze_panes = "A2"

# ============================================================
# AUTO FILTERS
# ============================================================
ws_tp.auto_filter.ref = f"A1:{get_column_letter(len(testplan_columns))}1"
ws_md.auto_filter.ref = f"A1:{get_column_letter(len(metadata_columns))}1"

# ============================================================
# AUTO-SIZE COLUMNS WITH MAX WIDTH LIMITS
# ============================================================
def auto_size_columns(ws, columns, long_text_set):
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 200))
        if col_name in long_text_set:
            width = min(max_len + 2, 80)
        elif max_len > 30:
            width = min(max_len + 2, 40)
        else:
            width = min(max_len + 2, 25)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

auto_size_columns(ws_tp, testplan_columns, long_text_cols_tp)
auto_size_columns(ws_md, metadata_columns, long_text_cols_md)

# ============================================================
# SAVE WORKBOOK
# ============================================================
wb.save(filepath)
wb.close()

# ============================================================
# POST-SAVE VALIDATION
# ============================================================
validation_results = []

# 1. File exists
validation_results.append(("File exists", os.path.exists(filepath)))

# 2. File size > 0
file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
validation_results.append(("File size > 0", file_size > 0))

# 3. Filename ends with .xlsx
validation_results.append(("Filename ends with .xlsx", filename.endswith(".xlsx")))

# 4-11. Reopen and validate
try:
    wb_check = load_workbook(filepath)
    validation_results.append(("Workbook reopens", True))
    
    # 5. Contains TestPlan sheet
    validation_results.append(("Contains TestPlan", "TestPlan" in wb_check.sheetnames))
    
    # 6. Contains MetaData sheet
    validation_results.append(("Contains MetaData", "MetaData" in wb_check.sheetnames))
    
    # 7. MetaData is veryHidden
    md_sheet = wb_check["MetaData"]
    validation_results.append(("MetaData veryHidden", md_sheet.sheet_state == "veryHidden"))
    
    # 8. TestPlan columns
    tp_sheet = wb_check["TestPlan"]
    tp_headers = [tp_sheet.cell(row=1, column=c).value for c in range(1, len(testplan_columns)+1)]
    validation_results.append(("TestPlan columns match", tp_headers == testplan_columns))
    
    # 9. MetaData columns
    md_headers = [md_sheet.cell(row=1, column=c).value for c in range(1, len(metadata_columns)+1)]
    validation_results.append(("MetaData columns match", md_headers == metadata_columns))
    
    # 10. TestPlan row count
    tp_data_rows = tp_sheet.max_row - 1
    validation_results.append(("TestPlan row count", tp_data_rows == len(json_data)))
    
    # 11. MetaData row count
    md_data_rows = md_sheet.max_row - 1
    validation_results.append(("MetaData row count", md_data_rows == len(json_data)))
    
    wb_check.close()
except Exception as e:
    validation_results.append(("Workbook reopens", False))

all_passed = all(v[1] for v in validation_results)
validation_status = "PASSED" if all_passed else "FAILED"

# ============================================================
# OUTPUT RESULTS
# ============================================================
print(f"FILENAME={filename}")
print(f"FILEPATH={filepath}")
print(f"FILESIZE={file_size}")
print(f"VALIDATION={validation_status}")
print(f"ROWS_TESTPLAN={len(json_data)}")
print(f"ROWS_METADATA={len(json_data)}")
print(f"TIMESTAMP={timestamp}")
for name, result in validation_results:
    print(f"  CHECK: {name} = {result}")
