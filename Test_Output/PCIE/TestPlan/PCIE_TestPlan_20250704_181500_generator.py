#!/usr/bin/env python3
"""
Auto-generated TestPlan Excel Generator
Generates: PCIE_TestPlan_20250704_181500.xlsx
Triggered: workflow_dispatch or push
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate():
    # Input data
    data = [
        {
            "Index": "1",
            "SS / Module": "PCIE",
            "Feature": "Device Enumeration",
            "Test Case Name": "pcie_device_enumerate_test",
            "Test Description": "Verifies PCIe device enumeration by performing link training, configuring cache coherency control registers for both PCIe controllers, polling link status registers until the expected link-up condition is detected, reading the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enabling bus master and memory/IO space via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, performing BAR sizing and assignment on BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG for both PCIe slave ports, configuring bus number and IO/memory limit registers (SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG), disabling cache coherency, and polling a synchronization register until the expected completion value is observed.",
            "Speed": "",
            "Mode": "",
            "Memory Start Offset": "",
            "Memory End Offset": "",
            "Remarks": "The testcase uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select the PCIe controller mode and link training path. It includes multiple wait operations for timing synchronization. Polling is performed on link status registers with a specific bitmask condition and on a synchronization register for test completion. Cache coherency control is enabled and then disabled during the test flow. Several register accesses remain unmapped to canonical register names. BAR sizing and programming is performed on both PCIe slave port 0 and slave port 1.",
            "Test Steps / Procedure": "1. Initialize the PCIe subsystem by writing to the synchronization register. 2. Perform link training for the configured PCIe controller mode (Root Complex or Endpoint). 3. Enable cache coherency control for both PCIe controller 0 and controller 1 by configuring the coherency control registers with appropriate bit field settings. 4. Wait for coherency settings to take effect. 5. Re-program coherency control registers with full bit field enablement for both controllers. 6. Poll the link status register on PCIe interface 0 until the expected link-up condition is detected. 7. Configure non-secure protection settings. 8. Poll the link status register on PCIe interface 1 until the expected link-up condition is detected. 9. Read the Vendor ID from TYPE1_DEV_ID_VEND_ID_REG on PCIe slave port 0. 10. Enable bus master, memory space, and I/O space access by writing to TYPE1_STATUS_COMMAND_REG. 11. Program memory base addresses for both PCIe controllers. 12. Write enable values to system configuration registers. 13. Disable cache coherency by clearing the relevant bit fields in the coherency control registers for both controllers. 14. Wait for cache disable to take effect. 15. Perform BAR sizing on PCIe slave port 1 by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, then reading back the supported size. 16. Assign actual BAR address values to PCIe slave port 1 registers and verify by reading back. 17. Repeat BAR sizing and assignment for PCIe slave port 0. 18. Poll the synchronization register until the expected completion value is observed. 19. Report test completion.",
            "Impacted Registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
            "Validation / Acceptance Criteria": "The test passes when: the link status registers on both PCIe interfaces report the expected link-up condition (masked status equals the expected pattern). The Vendor ID is successfully read from TYPE1_DEV_ID_VEND_ID_REG. BAR sizing returns valid supported-size values when all-ones are written to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG. Assigned BAR values are read back correctly. The synchronization register reaches the expected completion value. The test calls finish(0) indicating success.",
            "Code Generation": "",
            "Meta Test Description": "The testcase performs PCIe device enumeration by first writing to 0xE6004100 to initialize, then conditionally invoking link training for DM0 or DM1 in RC or EP mode via compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). It programs cache coherency control by performing read-modify-write operations on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF using set_data to configure specific bit fields (bits 3-6, 11-14, 19-22, 27-30). After a wait_on(20), it re-programs the same coherency registers. It then polls read_sii0_reg(0xC0) until the value masked with 0xD1 equals 0xD1, and similarly polls read_sii1_reg(0xC0). Under DM0_RC, it reads the Vendor ID from PCIe slave 0 at offset 0x0, writes 0x7 to offset 0x4 to enable bus master, memory space, and I/O space, then calls mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). It writes 0x1 to addresses 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, and 0xE6900034. It then disables cache by re-programming coherency control registers with 0x0 for bits 19-22 and 27-30. After wait_on(30), it performs BAR sizing on PCIe slave 1 by writing 0xFFFFFFFF to offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24, reading them back, then writing actual BAR values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000) and reading back again. The same BAR sizing and programming is repeated for PCIe slave 0. Finally, it polls 0xE6004100 until the value equals 0x12345678, with wait_on(5) between iterations, and calls finish(0) upon success.",
            "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100 to initialize. 2. Conditionally call link_training_dm0_x4(4) or link_training_dm1_x4(4) based on DM0_RC, DM1_RC, DM0_EP, DM1_EP defines. 3. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF: set bits 11-14 to 0xf, bits 3-6 to 0xf, then write back. 4. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF: set bits 27-30 to 0xf, bits 19-22 to 0xf, then write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. wait_on(20). 7. Read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting all four bit groups (3-6, 11-14, 19-22, 27-30) to 0xf and write back. 8. Read-modify-write mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF similarly. 9. Read read_sii0_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1. 10. Call non_secure_prot_nic(). 11. Read read_sii1_reg(0xC0) and poll until (data_rd & 0xD1) == 0xD1. 12. Under DM0_RC: read Vendor ID via read_pcie_slv0_reg(0x0). 13. Write 0x7 to write_pcie_slv0_reg(0x4) to enable bus master, memory space, and I/O space. 14. Call mem_base_program_dm0_x4() and mem_base_program_dm1_x4(). 15. wait_on(10). 16. Write 0x1 to 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034. 17. Disable cache: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF setting bits 19-22 and 27-30 to 0x0. 18. wait_on(10). 19. Final disable cache: set bits 27-30 and 19-22 to 0x0 for both PCIE0 and PCIE1 coherency registers. 20. wait_on(30). 21. BAR sizing on PCIe slave 1: write 0xFFFFFFFF to offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24, then read back all. 22. Write actual BAR values to PCIe slave 1: 0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000 at offsets 0x10-0x24, then read back. 23. Repeat BAR sizing and programming for PCIe slave 0. 24. wait_on(10). 25. Poll read_reg(0xE6004100) until value equals 0x12345678, with wait_on(5) between iterations. 26. Call finish(0).",
            "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
            "Meta Validation / Acceptance Criteria": "The test passes when: the link status registers on both PCIe interfaces report the expected link-up condition (masked status equals the expected pattern). The Vendor ID is successfully read from TYPE1_DEV_ID_VEND_ID_REG. BAR sizing returns valid supported-size values when all-ones are written to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG. Assigned BAR values are read back correctly. The synchronization register reaches the expected completion value. The test calls finish(0) indicating success.",
            "Meta Headers": "",
            "Meta Macros": "",
            "Meta Arrays": ""
        }
    ]

    # TestPlan columns
    tp_columns = [
        "Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
        "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
        "Test Steps / Procedure", "Impacted Registers", "Validation / Acceptance Criteria",
        "Code Generation"
    ]

    # MetaData columns
    md_columns = [
        "Index", "Test Case Name", "Meta Test Description", "Meta Test Steps / Procedure",
        "Meta Impacted Registers", "Meta Validation / Acceptance Criteria",
        "Meta Headers", "Meta Macros", "Meta Arrays"
    ]

    # Create workbook
    wb = Workbook()

    # TestPlan sheet
    ws_tp = wb.active
    ws_tp.title = "TestPlan"
    for col_idx, col_name in enumerate(tp_columns, 1):
        ws_tp.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(tp_columns, 1):
            ws_tp.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # MetaData sheet
    ws_md = wb.create_sheet("MetaData")
    for col_idx, col_name in enumerate(md_columns, 1):
        ws_md.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(md_columns, 1):
            ws_md.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # Formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    MAX_COL_WIDTH = 60

    for ws in [ws_tp, ws_md]:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                cell.alignment = wrap_alignment
                if cell.value:
                    for line in str(cell.value).split('\n'):
                        max_length = max(max_length, len(line))
            adjusted_width = min(max(max_length + 2, 12), MAX_COL_WIDTH)
            ws.column_dimensions[col_letter].width = adjusted_width

    # Set MetaData sheet to veryHidden
    ws_md.sheet_state = "veryHidden"

    # Save
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "PCIE_TestPlan_20250704_181500.xlsx")
    wb.save(output_path)
    wb.close()
    print(f"Generated: {output_path}")
    return output_path

if __name__ == "__main__":
    generate()
