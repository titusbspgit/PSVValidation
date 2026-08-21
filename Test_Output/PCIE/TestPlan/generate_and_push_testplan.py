#!/usr/bin/env python3
"""
PCIE TestPlan XLSX Generator
Generates PCIE_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx with openpyxl
and pushes to GitHub via API.

Usage: python3 generate_and_push_testplan.py
"""
import json, os, sys, base64, urllib.request, urllib.error
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
ts = now_ist.strftime("%Y%m%d_%H%M%S")
FILENAME = f"PCIE_TestPlan_{ts}.xlsx"

OWNER = "titusbspgit"
REPO = "PSVValidation"
BRANCH = "main"
OUTPUT_DIR = "Test_Output/PCIE/TestPlan"
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")

json_data = [
  {
    "index": 1,
    "SS_Module": "PCIE",
    "Test_Case_Name": "pcie_reg_wr_rd_test",
    "Feature": "Register Read/Write Validation",
    "Test_Description": "This test validates the reset default values and read-write accessibility of PCIe controller registers across both PCIE0 and PCIE1 instances. It covers DBI DSP registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF), SII registers (Transmit Header2, Transmit Header3, PHY Control 23), PHY reset control registers, and PHY lane registers. The test first reads all registers and compares against expected default values of 0x0. Then it performs write-read-back verification using multiple data patterns (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555) with appropriate write masks applied for SII and PHY registers.",
    "Meta_Test_Description": "This testcase validates reset default values and read-write functionality of PCIe registers for both PCIE0 and PCIE1 controller instances. It includes: (1) DBI DSP controller registers stored in rc0_ctl_addr[] (mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF) and rc1_ctl_addr[] (PCIE1 equivalents), all with default 0x0. (2) SII registers in sii0_addr[] (mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2, mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3, mizar_PCIE0_SII_PHY_CONTROL_23) and sii1_addr[] (PCIE1 equivalents), with write masks sii0_write_mask[] = {0xFFFFFFFF, 0xFFFFFFFF, 0xF000F} and sii1_write_mask[] = {0xFFFFFFFF, 0xFFFFFFFF, 0xF000F}. (3) PHY reset control via write_reg(mizar_PCIE0_SII_PHY_RST_CONTROL, 0x01203000) and write_reg(mizar_PCIE1_SII_PHY_RST_CONTROL, 0x01203000). (4) PHY lane registers in phy0_addr[] = {0xE68860B8, 0xE68862B8, 0xE68864B8} and phy1_addr[] = {0xE68A60B8, 0xE68A62B8, 0xE68A64B8} with write masks phy0_write_mask[] = phy1_write_mask[] = {0x1FFF, 0x1FFF, 0x1FFF}. PHY registers use 16-bit access with alignment-based shifting: if address % 4 != 0, data is shifted right by 16 bits; otherwise masked with 0x0000FFFF. The chk_rst_val() function reads all registers and compares against defaults. The chk_rd_wr() function iterates 3 times with chk_val[] = {0xFFFFFFFF, 0xAAAAAAAA, 0x55555555} for controller/SII registers and chk_val_phy[] = {0x7BAF, 0x1, 0x003B} for PHY registers. Error counters err1 and err2 track failures. finish(err2 || err1) determines pass/fail.",
    "Test_Steps_Procedure": "1. Read all five PCIE0 DBI DSP registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) and verify each returns the expected default value of 0x0.\n2. Read all five PCIE1 DBI DSP registers (same register set as PCIE0 on the second instance) and verify each returns the expected default value of 0x0.\n3. Read all three PCIE0 SII registers (Transmit Header2, Transmit Header3, PHY Control 23) and verify each returns the expected default value of 0x0.\n4. Read all three PCIE1 SII registers (Transmit Header2, Transmit Header3, PHY Control 23) and verify each returns the expected default value of 0x0.\n5. Write the PHY reset control value 0x01203000 to both PCIE0 and PCIE1 PHY Reset Control registers to bring PHY out of reset.\n6. Read all three PCIE0 PHY lane registers using 16-bit aligned access and verify each returns the expected default value of 0x0.\n7. Read all three PCIE1 PHY lane registers using 16-bit aligned access and verify each returns the expected default value of 0x0.\n8. For each of three data patterns (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555), write the pattern to all five PCIE0 DBI DSP registers and all five PCIE1 DBI DSP registers.\n9. For each data pattern, write the pattern masked with the appropriate write mask to all three PCIE0 SII registers and all three PCIE1 SII registers.\n10. Write the PHY reset control value to both PCIE0 and PCIE1 PHY Reset Control registers before each PHY write iteration.\n11. For each of three PHY-specific data patterns (0x7BAF, 0x1, 0x003B), write the pattern masked with 0x1FFF to all three PCIE0 PHY lane registers and all three PCIE1 PHY lane registers.\n12. Read back all five PCIE0 DBI DSP registers and verify the read data matches the written pattern.\n13. Read back all five PCIE1 DBI DSP registers and verify the read data matches the written pattern.\n14. Read back all three PCIE0 SII registers and verify the read data matches the written pattern masked with the write mask.\n15. Read back all three PCIE1 SII registers and verify the read data matches the written pattern masked with the write mask.\n16. Read back all three PCIE0 PHY lane registers using 16-bit aligned access and verify the read data masked with 0x1FFF matches the expected PHY pattern masked with 0x1FFF.\n17. Read back all three PCIE1 PHY lane registers using 16-bit aligned access and verify the read data masked with 0x1FFF matches the expected PHY pattern masked with 0x1FFF.\n18. Verify that no errors were accumulated across all reset-value checks and read-write-back checks. Report pass if both error counters are zero, fail otherwise.",
    "Meta_Test_Steps_Procedure": "1. Global variables: data_rd, data_wr, data1_rd, err1=0, err2=0. Arrays: rc0_ctl_addr[5]={mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF}, rc1_ctl_addr[5]={mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE1_DBI_DSP_UTILITY_OFF}, ctl_default[5]={0x0,0x0,0x0,0x0,0x0}, sii0_addr[3]={mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2, mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3, mizar_PCIE0_SII_PHY_CONTROL_23}, sii1_addr[3]={mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2, mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3, mizar_PCIE1_SII_PHY_CONTROL_23}, sii_default[3]={0x0,0x0,0x0}, sii0_write_mask[3]={0xFFFFFFFF,0xFFFFFFFF,0xF000F}, sii1_write_mask[3]={0xFFFFFFFF,0xFFFFFFFF,0xF000F}, phy0_addr[3]={0xE68860B8,0xE68862B8,0xE68864B8}, phy1_addr[3]={0xE68A60B8,0xE68A62B8,0xE68A64B8}, phy0_default[3]={0x0,0x0,0x0}, phy1_default[3]={0x0,0x0,0x0}, phy0_write_mask[3]={0x1FFF,0x1FFF,0x1FFF}, phy1_write_mask[3]={0x1FFF,0x1FFF,0x1FFF}.\n2. test_case() entry: calls chk_rst_val(), then chk_rd_wr(), then finish(err2 || err1).\n3. chk_rst_val(): Loop i=0..4: data_rd = read_reg(rc0_ctl_addr[i]); if(data_rd != ctl_default[i]) err1++, print FAILED. Loop i=0..4: data_rd = read_reg(rc1_ctl_addr[i]); if(data_rd != ctl_default[i]) err2++, print FAILED. Loop i=0..2: data_rd = read_reg(sii0_addr[i]); if(data_rd != sii_default[i]) err2++, print FAILED. Loop i=0..2: data_rd = read_reg(sii1_addr[i]); if(data_rd != sii_default[i]) err2++, print FAILED. write_reg(mizar_PCIE0_SII_PHY_RST_CONTROL, 0x01203000); write_reg(mizar_PCIE1_SII_PHY_RST_CONTROL, 0x01203000). Loop i=0..2: data_rd = read_reg(phy0_addr[i]); if(phy0_addr[i]%4) data_rd = data_rd >> 16; else data_rd = data_rd & 0x0000FFFF; if(data_rd != phy0_default[i]) err2++, print FAILED. Loop i=0..2: data_rd = read_reg(phy1_addr[i]); if(phy1_addr[i]%4) data_rd = data_rd >> 16; else data_rd = data_rd & 0x0000FFFF; if(data_rd != phy1_default[i]) err2++, print FAILED.\n4. chk_rd_wr(): chk_val[6]={0xFFFFFFFF,0xAAAAAAAA,0x55555555,0x00000000,0xA5A5A5A5,0xFFFF0000}; chk_val_phy[3]={0x7BAF,0x1,0x003B}. Loop j=0..2: Write phase: Loop i=0..4: write_reg(rc0_ctl_addr[i], chk_val[j]). Loop i=0..4: write_reg(rc1_ctl_addr[i], chk_val[j]). Loop i=0..2: write_reg(sii0_addr[i], chk_val[j] & sii0_write_mask[i]). Loop i=0..2: write_reg(sii1_addr[i], chk_val[j] & sii1_write_mask[i]). write_reg(mizar_PCIE0_SII_PHY_RST_CONTROL, 0x01203000); write_reg(mizar_PCIE1_SII_PHY_RST_CONTROL, 0x01203000). Loop i=0..2: write_reg(phy0_addr[i], chk_val_phy[j] & phy0_write_mask[i]). Loop i=0..2: write_reg(phy1_addr[i], chk_val_phy[j] & phy1_write_mask[i]). Read-back phase: Loop i=0..4: data_rd = read_reg(rc0_ctl_addr[i]); if(data_rd != chk_val[j]) err1++, print FAILED. Loop i=0..4: data_rd = read_reg(rc1_ctl_addr[i]); if(data_rd != chk_val[j]) err1++, print FAILED. Loop i=0..2: data_rd = read_reg(sii0_addr[i]); if(data_rd != (chk_val[j] & sii0_write_mask[i])) err1++, print FAILED. Loop i=0..2: data_rd = read_reg(sii1_addr[i]); if(data_rd != (chk_val[j] & sii1_write_mask[i])) err1++, print FAILED. Loop i=0..2: data_rd = read_reg(phy0_addr[i]); if(phy0_addr[i]%4) data_rd >>= 16; else data_rd &= 0x0000FFFF; if((data_rd & phy0_write_mask[i]) != (chk_val_phy[j] & 0x00001FFF)) err1++, print FAILED. Loop i=0..2: data_rd = read_reg(phy1_addr[i]); if(phy1_addr[i]%4) data_rd >>= 16; else data_rd &= 0x0000FFFF; if((data_rd & phy1_write_mask[i]) != (chk_val_phy[j] & 0x00001FFF)) err1++, print FAILED.\n5. finish(err2 || err1): pass if both err1 and err2 are 0, fail otherwise.",
    "Validation_Acceptance_Criteria": "1. All DBI DSP registers on both PCIE0 and PCIE1 instances must read back their expected default value of 0x00000000 after reset.\n2. All SII registers on both PCIE0 and PCIE1 instances must read back their expected default value of 0x00000000 after reset.\n3. All PHY lane registers on both PCIE0 and PCIE1 instances must read back their expected default value of 0x0 after PHY reset de-assertion.\n4. For each write pattern (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555), all DBI DSP registers must read back the exact written value.\n5. For each write pattern, all SII registers must read back the written value masked with the register-specific write mask (full 32-bit for Transmit Header registers, 0xF000F for PHY Control 23).\n6. For each PHY-specific write pattern (0x7BAF, 0x1, 0x003B), all PHY lane registers must read back the written value within the 13-bit write mask (0x1FFF) using 16-bit aligned access.\n7. Both error counters must be zero for the test to pass.",
    "Impacted_Registers": "MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
    "Meta_Impacted_Registers": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8",
    "Speed": "NA",
    "Mode": "NA",
    "Memory_Offset": "0x58; 0x60; 0x720; 0x8F4; 0xC80; NA; NA; NA; NA; NA; NA; NA; NA; NA; NA; NA; NA; NA; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8",
    "Remarks": "Test covers two PCIe controller instances (PCIE0 and PCIE1) across three register domains: DBI DSP controller registers, SII wrapper registers, and PHY lane registers. SII registers use instance-specific write masks; PHY registers use 16-bit aligned access with a 13-bit write mask (0x1FFF). PHY reset control registers are written with value 0x01203000 to de-assert PHY reset before PHY register access. Three data patterns are used for write-read-back verification of controller and SII registers, and three separate PHY-specific patterns are used for PHY lane registers. PCIE1 DBI DSP registers, all SII registers, PHY reset control registers, and PHY lane hex-addressed registers could not be mapped to specification register names due to missing header or specification data.",
    "Headers_Include": "#include <stdlib.h>; #include <stdio.h>; #include <test_common.h>; #include <pcie.h>",
    "Register_Operations": [
      {"token": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0x58", "resolution_status": "partially_resolved", "register_name": "MSI_CAP_OFF_08H_REG", "mapping_status": "matched"},
      {"token": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0x60", "resolution_status": "partially_resolved", "register_name": "MSI_CAP_OFF_10H_REG", "mapping_status": "matched"},
      {"token": "mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0x720", "resolution_status": "partially_resolved", "register_name": "FILTER_MASK_2_OFF", "mapping_status": "matched"},
      {"token": "mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0x8F4", "resolution_status": "partially_resolved", "register_name": "AXI_MSTR_MSG_ADDR_HIGH_OFF", "mapping_status": "matched"},
      {"token": "mizar_PCIE0_DBI_DSP_UTILITY_OFF", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0xC80", "resolution_status": "partially_resolved", "register_name": "UTILITY_OFF", "mapping_status": "matched"},
      {"token": "mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_DBI_DSP_UTILITY_OFF", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE0_SII_PHY_CONTROL_23", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_SII_PHY_CONTROL_23", "token_type": "macro", "operation": "read_modify_write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE0_SII_PHY_RST_CONTROL", "token_type": "macro", "operation": "write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "mizar_PCIE1_SII_PHY_RST_CONTROL", "token_type": "macro", "operation": "write", "base_value": "NA", "offset_value": "NA", "resolution_status": "unresolved", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "0xE68860B8", "token_type": "hex", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0xE68860B8", "resolution_status": "direct_hex", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "0xE68862B8", "token_type": "hex", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0xE68862B8", "resolution_status": "direct_hex", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "0xE68864B8", "token_type": "hex", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0xE68864B8", "resolution_status": "direct_hex", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "0xE68A60B8", "token_type": "hex", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0xE68A60B8", "resolution_status": "direct_hex", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "0xE68A62B8", "token_type": "hex", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0xE68A62B8", "resolution_status": "direct_hex", "register_name": "NA", "mapping_status": "unresolved"},
      {"token": "0xE68A64B8", "token_type": "hex", "operation": "read_modify_write", "base_value": "NA", "offset_value": "0xE68A64B8", "resolution_status": "direct_hex", "register_name": "NA", "mapping_status": "unresolved"}
    ]
  }
]

def build_workbook():
    wb = Workbook()
    ws_tp = wb.active
    ws_tp.title = "TestPlan"
    ws_md = wb.create_sheet("MetaData")

    tp_headers = ["Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
                  "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
                  "Test Steps / Procedure", "Impacted Registers",
                  "Validation / Acceptance Criteria", "Code Generation"]
    md_headers = ["Index", "Test Case Name", "Meta Test Description",
                  "Meta Test Steps / Procedure", "Meta Impacted Registers",
                  "Meta Validation / Acceptance Criteria", "Meta Headers",
                  "Meta Macros", "Meta Arrays"]

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    for ci, h in enumerate(tp_headers, 1):
        c = ws_tp.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = wrap

    for ci, h in enumerate(md_headers, 1):
        c = ws_md.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = wrap

    for ri, row in enumerate(json_data, 2):
        reg_ops = row.get("Register_Operations", [])
        macros_str = "; ".join([op["token"] for op in reg_ops])
        arrays_str = json.dumps(reg_ops, indent=None)

        tp_vals = [
            row["index"], row["SS_Module"], row["Feature"], row["Test_Case_Name"],
            row["Test_Description"], row["Speed"], row["Mode"], row["Memory_Offset"],
            "", row["Remarks"], row["Test_Steps_Procedure"], row["Impacted_Registers"],
            row["Validation_Acceptance_Criteria"], ""
        ]
        for ci, val in enumerate(tp_vals, 1):
            c = ws_tp.cell(row=ri, column=ci, value=val)
            c.alignment = wrap

        md_vals = [
            row["index"], row["Test_Case_Name"], row["Meta_Test_Description"],
            row["Meta_Test_Steps_Procedure"], row["Meta_Impacted_Registers"],
            row["Validation_Acceptance_Criteria"], row["Headers_Include"],
            macros_str, arrays_str
        ]
        for ci, val in enumerate(md_vals, 1):
            c = ws_md.cell(row=ri, column=ci, value=val)
            c.alignment = wrap

    ws_tp.freeze_panes = "A2"
    ws_md.freeze_panes = "A2"
    ws_md.sheet_state = "veryHidden"

    for ws in [ws_tp, ws_md]:
        for col_cells in ws.columns:
            mx = 12
            for cell in col_cells:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    longest = max(len(l) for l in lines) if lines else 0
                    mx = max(mx, min(longest + 2, 80))
            ws.column_dimensions[col_cells[0].column_letter].width = mx

    return wb

def main():
    wb = build_workbook()
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), FILENAME)
    wb.save(out_path)
    print(f"Workbook saved: {out_path}")
    print(f"Filename: {FILENAME}")

    # Validate
    wb2 = load_workbook(out_path)
    assert "TestPlan" in wb2.sheetnames
    assert "MetaData" in wb2.sheetnames
    fsize = os.path.getsize(out_path)
    assert fsize > 0
    print(f"Validation PASSED (size={fsize} bytes, sheets={wb2.sheetnames})")

    # Push to GitHub if token available
    if GITHUB_TOKEN:
        with open(out_path, "rb") as f:
            content_b64 = base64.b64encode(f.read()).decode("utf-8")
        api_url = f"https://api.github.com/repos/{OWNER}/{REPO}/contents/{OUTPUT_DIR}/{FILENAME}"
        payload = json.dumps({
            "message": "Added generated TestPlan Excel",
            "content": content_b64,
            "branch": BRANCH
        }).encode("utf-8")
        req = urllib.request.Request(api_url, data=payload, method="PUT")
        req.add_header("Authorization", f"token {GITHUB_TOKEN}")
        req.add_header("Content-Type", "application/json")
        req.add_header("Accept", "application/vnd.github.v3+json")
        try:
            resp = urllib.request.urlopen(req)
            result = json.loads(resp.read().decode("utf-8"))
            print(f"Pushed to GitHub: {result['content']['html_url']}")
            print(f"Commit SHA: {result['commit']['sha']}")
        except urllib.error.HTTPError as e:
            print(f"GitHub push failed: {e.code} {e.read().decode()}")
    else:
        print("No GITHUB_TOKEN set. Skipping push.")

if __name__ == "__main__":
    main()
