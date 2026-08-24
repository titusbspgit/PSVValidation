#!/usr/bin/env python3
import json, base64, sys
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

data = json.loads(sys.stdin.read())

tp_cols = ["Index","SS / Module","Feature","Test Case Name","Test Description",
           "Speed","Mode","Memory Start Offset","Memory End Offset","Remarks",
           "Test Steps / Procedure","Impacted Registers",
           "Validation / Acceptance Criteria","Code Generation"]

md_cols = ["Index","Test Case Name","Meta Test Description",
           "Meta Test Steps / Procedure","Meta Impacted Registers",
           "Meta Validation / Acceptance Criteria",
           "Meta Headers","Meta Macros","Meta Arrays"]

wb = Workbook()
ws_tp = wb.active
ws_tp.title = "TestPlan"
ws_md = wb.create_sheet("MetaData")

hf = Font(bold=True, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")

for ws, cols in [(ws_tp, tp_cols), (ws_md, md_cols)]:
    for ci, cn in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.font = hf; c.fill = hfill; c.alignment = wrap
    ws.freeze_panes = "A2"

for ri, row in enumerate(data, 2):
    for ci, col in enumerate(tp_cols, 1):
        c = ws_tp.cell(row=ri, column=ci, value=row.get(col, ""))
        c.alignment = wrap
    for ci, col in enumerate(md_cols, 1):
        c = ws_md.cell(row=ri, column=ci, value=row.get(col, ""))
        c.alignment = wrap

for ws in [ws_tp, ws_md]:
    for ci in range(1, ws.max_column + 1):
        ml = 0
        for row in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row:
                if cell.value:
                    ml = max(ml, min(len(str(cell.value)), 80))
        ws.column_dimensions[get_column_letter(ci)].width = max(12, min(ml + 2, 82))

ws_md.sheet_state = "veryHidden"

buf = BytesIO()
wb.save(buf)
b = buf.getvalue()
print(base64.b64encode(b).decode())
