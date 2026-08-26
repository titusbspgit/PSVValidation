#!/usr/bin/env python3
"""PCIE TestPlan XLSX Generator - Agent 7"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
import json, os, sys

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
TS = now_ist.strftime("%Y%m%d_%H%M%S")
FILENAME = f"PCIE_TestPlan_{TS}.xlsx"
OUTPATH = f"/tmp/{FILENAME}"

# ── JSON DATA ──
json_data = [
  {
    "Index": "1",
    "SS / Module": "PCIE",
    "Test Case Name": "pcie_device_enumerate_test",
    "Feature": "PCIe Device Enumeration",
    "Test Description": "Verifies PCIe device enumeration by performing link training, configuring cache coherency via the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances, polling the SII link status registers until link-up is confirmed, reading the TYPE1_DEV_ID_VEND_ID_REG to retrieve the endpoint Vendor ID, enabling IO, Memory, and Bus Master access via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, performing BAR sizing and assignment on BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG for both PCIe slave ports, and polling a system handshake register for the expected completion value.",
    "Remarks": "The testcase uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select between Root Complex and Endpoint modes and between dual-mode controller instances. Polling with wait_on() delays is used for link status confirmation and completion handshake. Several system-level registers used for configuration could not be mapped to named registers in the specification. The SII link status register at the polled offset could not be mapped to a named register. One coherency control macro for the second PCIe instance could not be resolved from the available headers.",
    "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
    "Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. PCIe link training completes successfully for the configured controller mode. 2. The SII0 link status register reports link-up with the expected bit pattern confirming data link layer active and link-up status for PCIE0. 3. The SII1 link status register reports link-up with the expected bit pattern for PCIE1. 4. The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID from the endpoint device. 5. TYPE1_STATUS_COMMAND_REG is successfully written to enable IO, Memory, and Bus Master access. 6. BAR sizing returns valid size information for all BARs on both PCIe slave ports. 7. BAR programming with actual address values completes and read-back confirms correct values. 8. The system handshake register returns the expected completion value, confirming the remote endpoint has completed its enumeration sequence. 9. The test calls finish(0) indicating successful completion.",
    "Test Steps / Procedure": "1. Initialize the system by writing to the system control register. 2. Perform PCIe link training for the configured controller mode (Root Complex or Endpoint). 3. Enable cache coherency by programming the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances with appropriate bit field values. 4. Wait for the coherency settings to take effect. 5. Poll the SII0 link status register until PCIE0 link-up is confirmed. 6. Poll the SII1 link status register until PCIE1 link-up is confirmed. 7. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG from the endpoint to retrieve the Vendor ID. 8. Enable IO space, Memory space, and Bus Master access by writing to TYPE1_STATUS_COMMAND_REG. 9. Program memory base addresses for both PCIe controller instances. 10. Write to system-level configuration registers to enable required system settings. 11. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF for both instances. 12. Perform BAR sizing on PCIe slave port 1 by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, then reading back to determine sizes. 13. Program actual BAR values for PCIe slave port 1 and verify by reading back. 14. Repeat BAR sizing and programming for PCIe slave port 0. 15. Poll the system handshake register until the expected completion value is received. 16. Confirm test completion."
  }
]

print(f"Generating: {OUTPATH}")
print("openpyxl version:", openpyxl.__version__)
wb = openpyxl.Workbook()
print("Workbook created successfully")
ws = wb.active
ws.title = "TestPlan"
print(f"Saved to: {OUTPATH}")
wb.save(OUTPATH)
print(f"File size: {os.path.getsize(OUTPATH)} bytes")
