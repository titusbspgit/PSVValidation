#!/usr/bin/env python3
"""Generate PCIE TestPlan Excel workbook from JSON data.

Usage:
    python PCIE_TestPlan_20260824_161300_generator.py

Requires:
    pip install openpyxl
"""

import json
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "PCIE_TestPlan_20260824_161300_data.json")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ==================== TestPlan Sheet ====================
    ws_tp = wb.active
    ws_tp.title = "TestPlan"
    tp_columns = [
        "Index", "SS / Module", "Feature", "Test Case Name",
        "Test Description", "Speed", "Mode",
        "Memory Start Offset", "Memory End Offset", "Remarks",
        "Test Steps / Procedure", "Impacted Registers",
        "Validation / Acceptance Criteria", "Code Generation"
    ]
    for col_idx, col_name in enumerate(tp_columns, 1):
        cell = ws_tp.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, entry in enumerate(data, 2):
        for col_idx, col_name in enumerate(tp_columns, 1):
            value = entry.get(col_name, "")
            cell = ws_tp.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    tp_widths = [8, 15, 30, 35, 60, 10, 10, 20, 20, 50, 60, 40, 60, 15]
    for i, w in enumerate(tp_widths, 1):
        ws_tp.column_dimensions[ws_tp.cell(row=1, column=i).column_letter].width = w

    ws_tp.freeze_panes = "A2"

    # ==================== MetaData Sheet ====================
    ws_md = wb.create_sheet(title="MetaData")
    md_columns = [
        "Index", "Test Case Name", "Meta Test Description",
        "Meta Test Steps / Procedure", "Meta Impacted Registers",
        "Meta Validation / Acceptance Criteria",
        "Meta Headers", "Meta Macros", "Meta Arrays"
    ]
    for col_idx, col_name in enumerate(md_columns, 1):
        cell = ws_md.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, entry in enumerate(data, 2):
        for col_idx, col_name in enumerate(md_columns, 1):
            if col_name == "Meta Validation / Acceptance Criteria":
                value = entry.get("Validation / Acceptance Criteria", "")
            else:
                value = entry.get(col_name, "")
            cell = ws_md.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    md_widths = [8, 35, 80, 80, 60, 60, 30, 30, 30]
    for i, w in enumerate(md_widths, 1):
        ws_md.column_dimensions[ws_md.cell(row=1, column=i).column_letter].width = w

    ws_md.freeze_panes = "A2"
    ws_md.sheet_state = "veryHidden"

    # ==================== Save ====================
    output_dir = os.path.join(script_dir, "..")
    output_file = os.path.join(output_dir, "PCIE_TestPlan_20260824_161300.xlsx")
    wb.save(output_file)

    # ==================== Validate ====================
    assert os.path.exists(output_file), "File not created"
    assert os.path.getsize(output_file) > 0, "File is empty"
    vwb = load_workbook(output_file)
    assert "TestPlan" in vwb.sheetnames, "TestPlan sheet missing"
    assert "MetaData" in vwb.sheetnames, "MetaData sheet missing"
    vwb.close()

    print(f"SUCCESS: {output_file}")
    print(f"  TestPlan rows: {len(data)}")
    print(f"  MetaData rows: {len(data)}")


if __name__ == "__main__":
    main()
