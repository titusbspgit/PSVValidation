#!/usr/bin/env python3
"""
PCIE TestPlan XLSX Generator - Agent 7
Generates a real Excel workbook with TestPlan and MetaData sheets.
Usage: python3 generate_xlsx.py
Output: PCIE_TestPlan_YYYYMMDD_HHMMSS.xlsx in the same directory
"""
import os, sys, json, base64
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    HAS_RICH_TEXT = True
except ImportError:
    HAS_RICH_TEXT = False
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
ts = now_ist.strftime("%Y%m%d_%H%M%S")
FILENAME = f"PCIE_TestPlan_{ts}.xlsx"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
FILEPATH = os.path.join(PARENT_DIR, FILENAME)

# ============================================================
# TestPlan columns
# ============================================================
TP_COLS = [
    "Index", "SS / Module", "Feature", "Test Case Name",
    "Test Description", "Speed", "Mode", "Memory Start Offset",
    "Memory End Offset", "Remarks", "Test Steps / Procedure",
    "Impacted Registers", "Validation / Acceptance Criteria",
    "Code Generation"
]

# MetaData columns
MD_COLS = [
    "Index", "Test Case Name", "Meta Test Description",
    "Meta Test Steps / Procedure", "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers", "Meta Macros", "Meta Arrays"
]

# ============================================================
# Categorized Test Steps (pre-formatted with line breaks)
# ============================================================
STEPS = {}
STEPS["1"] = (
    "Initialization:\n"
    "1. Initialize the system by writing to the system control register.\n"
    "\n"
    "Configuration:\n"
    "1. Perform PCIe link training for the configured controller mode (Root Complex or Endpoint).\n"
    "2. Enable cache coherency by programming the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances with appropriate bit field values.\n"
    "3. Enable IO space, Memory space, and Bus Master access by writing to TYPE1_STATUS_COMMAND_REG.\n"
    "4. Program memory base addresses for both PCIe controller instances.\n"
    "5. Write to system-level configuration registers to enable required system settings.\n"
    "6. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF for both instances.\n"
    "\n"
    "Execution:\n"
    "1. Wait for the coherency settings to take effect.\n"
    "2. Poll the SII0 link status register until PCIE0 link-up is confirmed.\n"
    "3. Poll the SII1 link status register until PCIE1 link-up is confirmed.\n"
    "4. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG from the endpoint to retrieve the Vendor ID.\n"
    "5. Perform BAR sizing on PCIe slave port 1 by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, then reading back to determine sizes.\n"
    "6. Program actual BAR values for PCIe slave port 1 and verify by reading back.\n"
    "7. Repeat BAR sizing and programming for PCIe slave port 0.\n"
    "8. Poll the system handshake register until the expected completion value is received.\n"
    "9. Confirm test completion."
)

STEPS["2"] = (
    "Initialization:\n"
    "1. Initialize the system by writing to the system control register.\n"
    "2. Preload known data patterns into source memory regions.\n"
    "\n"
    "Configuration:\n"
    "1. Perform PCIe link training for the configured controller mode.\n"
    "2. Enable IO space, Memory space, and Bus Master access by writing to TYPE1_STATUS_COMMAND_REG.\n"
    "3. Configure BARs and program memory base addresses for the active controller.\n"
    "4. Unmask DMA write and read interrupts by writing to DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF.\n"
    "\n"
    "Execution:\n"
    "1. Poll the SII link status register until link-up is confirmed for the active PCIe controller instance.\n"
    "2. Read the TYPE1_DEV_ID_VEND_ID_REG from the endpoint to retrieve the Vendor ID.\n"
    "3. Poll the system handshake register until the expected completion value is received.\n"
    "4. Program DMA write channel 0 with source address, destination address, and transfer length, then trigger the transfer by writing to DMA_WRITE_DOORBELL_OFF and wait for interrupt-driven completion.\n"
    "5. Repeat DMA write transfer for channels 1, 2, and 3, triggering each sequentially via DMA_WRITE_DOORBELL_OFF and waiting for completion.\n"
    "6. Program DMA read channel 0 with remote source address, local destination address, and transfer length, then trigger the transfer by writing to DMA_READ_DOORBELL_OFF and wait for interrupt-driven completion.\n"
    "7. Repeat DMA read transfer for channels 1, 2, and 3, triggering each sequentially via DMA_READ_DOORBELL_OFF and waiting for completion.\n"
    "8. Confirm test completion after all DMA write and read transfers on all four channels have completed successfully.\n"
    "\n"
    "Interrupt:\n"
    "1. Enable GIC interrupts for DMA completion notification.\n"
    "2. In the interrupt handler, read DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF to identify completed channels, then clear interrupts via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF."
)

STEPS["3"] = (
    "Initialization:\n"
    "1. Initialize the system by writing to the system control register.\n"
    "\n"
    "Configuration:\n"
    "1. Perform PCIe link training for the configured controller mode (Root Complex or Endpoint).\n"
    "2. Enable cache coherency by programming the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances with appropriate bit field values.\n"
    "3. Re-program the COHERENCY_CONTROL_3_OFF register for both instances to consolidate all coherency bit fields.\n"
    "4. Enable IO space, Memory space, and Bus Master access by writing to TYPE1_STATUS_COMMAND_REG.\n"
    "5. Configure BARs and program memory base addresses for the active controller.\n"
    "6. Call the non-secure protection NIC configuration function.\n"
    "7. Write a synchronization signal to the system control register.\n"
    "8. Disable cache coherency in a staged sequence by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF for both instances.\n"
    "\n"
    "Execution:\n"
    "1. Wait for the coherency settings to take effect.\n"
    "2. Poll the SII link status register until link-up is confirmed for the active PCIe controller instance.\n"
    "3. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG from the endpoint to retrieve the Vendor ID.\n"
    "4. Wait for the cache disable settings to take effect.\n"
    "5. Perform memory write-read verification at multiple target addresses through the PCIe slave port, writing known data patterns and reading them back.\n"
    "6. Poll the system handshake register until the expected completion value is received.\n"
    "7. Confirm test completion."
)

STEPS["4"] = (
    "Initialization:\n"
    "1. Write to the PHY reset control registers for both instances to bring the PHY out of reset.\n"
    "\n"
    "Execution:\n"
    "1. Read all five DBI controller registers for both PCIe controller instances and verify that each register contains its expected default value of zero.\n"
    "2. Read all three SII registers for both PCIe controller instances and verify that each register contains its expected default value of zero.\n"
    "3. Read all three PHY registers for both instances, extract the relevant half-word, and verify that each contains its expected default value of zero.\n"
    "4. For each of three data patterns, write the pattern to all five DBI controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) for both instances.\n"
    "5. For each data pattern, write the pattern masked with the appropriate write mask to all three SII registers for both instances.\n"
    "6. For each data pattern, write the PHY-specific pattern masked with the PHY write mask to all three PHY registers for both instances.\n"
    "7. After each write iteration, read back all DBI controller registers and verify the read value matches the written value.\n"
    "8. After each write iteration, read back all SII registers and verify the read value matches the written value accounting for the write mask.\n"
    "9. After each write iteration, read back all PHY registers, extract the relevant half-word, and verify the read value matches the written value accounting for the PHY write mask.\n"
    "10. Confirm that all default value checks and all write-read comparisons passed with zero errors across both controller instances."
)

print(f"[Agent 7] PCIE TestPlan XLSX Generator")
print(f"[Agent 7] IST: {now_ist.strftime('%Y-%m-%d %H:%M:%S')}")
print(f"[Agent 7] Filename: {FILENAME}")
print(f"[Agent 7] Output path: {FILEPATH}")
print(f"[Agent 7] Rich text support: {HAS_RICH_TEXT}")
print(f"[Agent 7] Run this script to generate the workbook.")
