#!/usr/bin/env python3
"""PCIE TestPlan Excel Generator
Run: python3 PCIE_TestPlan_20250705_141500_generate.py
Produces: PCIE_TestPlan_20250705_141500.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── IST timestamp ──
FILENAME = "PCIE_TestPlan_20250705_141500.xlsx"

# ── Column definitions ──
TP_COLS = [
    "Index", "SS / Module", "Feature", "Test Case Name", "Test Description",
    "Speed", "Mode", "Memory Start Offset", "Memory End Offset", "Remarks",
    "Test Steps / Procedure", "Impacted Registers",
    "Validation / Acceptance Criteria", "Code Generation"
]
MD_COLS = [
    "Index", "Test Case Name", "Meta Test Description",
    "Meta Test Steps / Procedure", "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria", "Meta Headers",
    "Meta Macros", "Meta Arrays"
]

# ── Categorized Test Steps ──
STEPS_ROW1 = (
    "Initialization:\n"
    "1. Initialize the system by writing to the system control register.\n"
    "2. Perform PCIe link training for the configured controller mode (Root Complex or Endpoint).\n\n"
    "Configuration:\n"
    "1. Enable cache coherency by programming the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances with appropriate bit field values.\n"
    "2. Wait for the coherency settings to take effect.\n"
    "3. Enable IO space, Memory space, and Bus Master access by writing to TYPE1_STATUS_COMMAND_REG.\n"
    "4. Program memory base addresses for both PCIe controller instances.\n"
    "5. Write to system-level configuration registers to enable required system settings.\n"
    "6. Disable cache coherency by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF for both instances.\n\n"
    "Execution:\n"
    "1. Poll the SII0 link status register until PCIE0 link-up is confirmed.\n"
    "2. Poll the SII1 link status register until PCIE1 link-up is confirmed.\n"
    "3. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG from the endpoint to retrieve the Vendor ID.\n"
    "4. Perform BAR sizing on PCIe slave port 1 by writing all-ones to BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG, then reading back to determine sizes.\n"
    "5. Program actual BAR values for PCIe slave port 1 and verify by reading back.\n"
    "6. Repeat BAR sizing and programming for PCIe slave port 0.\n"
    "7. Poll the system handshake register until the expected completion value is received.\n"
    "8. Confirm test completion."
)

STEPS_ROW2 = (
    "Initialization:\n"
    "1. Initialize the system by writing to the system control register.\n"
    "2. Perform PCIe link training for the configured controller mode.\n"
    "3. Preload known data patterns into source memory regions.\n\n"
    "Configuration:\n"
    "1. Enable IO space, Memory space, and Bus Master access by writing to TYPE1_STATUS_COMMAND_REG.\n"
    "2. Configure BARs and program memory base addresses for the active controller.\n"
    "3. Unmask DMA write and read interrupts by writing to DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF.\n\n"
    "Execution:\n"
    "1. Poll the SII link status register until link-up is confirmed for the active PCIe controller instance.\n"
    "2. Read the TYPE1_DEV_ID_VEND_ID_REG from the endpoint to retrieve the Vendor ID.\n"
    "3. Poll the system handshake register until the expected completion value is received.\n"
    "4. Program DMA write channel 0 with source address, destination address, and transfer length, then trigger the transfer by writing to DMA_WRITE_DOORBELL_OFF and wait for interrupt-driven completion.\n"
    "5. Repeat DMA write transfer for channels 1, 2, and 3, triggering each sequentially via DMA_WRITE_DOORBELL_OFF and waiting for completion.\n"
    "6. Program DMA read channel 0 with remote source address, local destination address, and transfer length, then trigger the transfer by writing to DMA_READ_DOORBELL_OFF and wait for interrupt-driven completion.\n"
    "7. Repeat DMA read transfer for channels 1, 2, and 3, triggering each sequentially via DMA_READ_DOORBELL_OFF and waiting for completion.\n"
    "8. Confirm test completion after all DMA write and read transfers on all four channels have completed successfully.\n\n"
    "Interrupt:\n"
    "1. Enable GIC interrupts for DMA completion notification.\n"
    "2. In the interrupt handler, read DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF to identify completed channels, then clear interrupts via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF."
)

STEPS_ROW3 = (
    "Initialization:\n"
    "1. Initialize the system by writing to the system control register.\n"
    "2. Perform PCIe link training for the configured controller mode (Root Complex or Endpoint).\n\n"
    "Configuration:\n"
    "1. Enable cache coherency by programming the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances with appropriate bit field values.\n"
    "2. Wait for the coherency settings to take effect.\n"
    "3. Re-program the COHERENCY_CONTROL_3_OFF register for both instances to consolidate all coherency bit fields.\n"
    "4. Enable IO space, Memory space, and Bus Master access by writing to TYPE1_STATUS_COMMAND_REG.\n"
    "5. Configure BARs and program memory base addresses for the active controller.\n"
    "6. Call the non-secure protection NIC configuration function.\n"
    "7. Write a synchronization signal to the system control register.\n"
    "8. Disable cache coherency in a staged sequence by clearing the relevant bit fields in COHERENCY_CONTROL_3_OFF for both instances.\n"
    "9. Wait for the cache disable settings to take effect.\n\n"
    "Execution:\n"
    "1. Poll the SII link status register until link-up is confirmed for the active PCIe controller instance.\n"
    "2. In Root Complex mode, read the TYPE1_DEV_ID_VEND_ID_REG from the endpoint to retrieve the Vendor ID.\n"
    "3. Perform memory write-read verification at multiple target addresses through the PCIe slave port, writing known data patterns and reading them back.\n"
    "4. Poll the system handshake register until the expected completion value is received.\n"
    "5. Confirm test completion."
)

STEPS_ROW4 = (
    "Initialization:\n"
    "1. Read all five DBI controller registers for both PCIe controller instances and verify that each register contains its expected default value of zero.\n"
    "2. Read all three SII registers for both PCIe controller instances and verify that each register contains its expected default value of zero.\n"
    "3. Write to the PHY reset control registers for both instances to bring the PHY out of reset.\n"
    "4. Read all three PHY registers for both instances, extract the relevant half-word, and verify that each contains its expected default value of zero.\n\n"
    "Configuration:\n"
    "1. For each of three data patterns, write the pattern to all five DBI controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) for both instances.\n"
    "2. For each data pattern, write the pattern masked with the appropriate write mask to all three SII registers for both instances.\n"
    "3. For each data pattern, write the PHY-specific pattern masked with the PHY write mask to all three PHY registers for both instances.\n\n"
    "Execution:\n"
    "1. After each write iteration, read back all DBI controller registers and verify the read value matches the written value.\n"
    "2. After each write iteration, read back all SII registers and verify the read value matches the written value accounting for the write mask.\n"
    "3. After each write iteration, read back all PHY registers, extract the relevant half-word, and verify the read value matches the written value accounting for the PHY write mask.\n"
    "4. Confirm that all default value checks and all write-read comparisons passed with zero errors across both controller instances."
)

# ── Row data ──
ROWS = [
    {
        "Index": 1,
        "SS / Module": "PCIE",
        "Feature": "PCIe Device Enumeration",
        "Test Case Name": "pcie_device_enumerate_test",
        "Test Description": "Verifies PCIe device enumeration by performing link training, configuring cache coherency via the COHERENCY_CONTROL_3_OFF register for both PCIe controller instances, polling the SII link status registers until link-up is confirmed, reading the TYPE1_DEV_ID_VEND_ID_REG to retrieve the endpoint Vendor ID, enabling IO, Memory, and Bus Master access via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, performing BAR sizing and assignment on BAR0_REG, BAR1_REG, SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG, SEC_STAT_IO_LIMIT_IO_BASE_REG, MEM_LIMIT_MEM_BASE_REG, and PREF_MEM_LIMIT_PREF_MEM_BASE_REG for both PCIe slave ports, and polling a system handshake register for the expected completion value.",
        "Speed": "",
        "Mode": "",
        "Memory Start Offset": "",
        "Memory End Offset": "",
        "Remarks": "The testcase uses conditional compilation (DM0_RC, DM1_RC, DM0_EP, DM1_EP) to select between Root Complex and Endpoint modes and between dual-mode controller instances. Polling with wait_on() delays is used for link status confirmation and completion handshake. Several system-level registers used for configuration could not be mapped to named registers in the specification. The SII link status register at the polled offset could not be mapped to a named register. One coherency control macro for the second PCIe instance could not be resolved from the available headers.",
        "Test Steps / Procedure": STEPS_ROW1,
        "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
        "Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. PCIe link training completes successfully for the configured controller mode. 2. The SII0 link status register reports link-up with the expected bit pattern confirming data link layer active and link-up status for PCIE0. 3. The SII1 link status register reports link-up with the expected bit pattern for PCIE1. 4. The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID from the endpoint device. 5. TYPE1_STATUS_COMMAND_REG is successfully written to enable IO, Memory, and Bus Master access. 6. BAR sizing returns valid size information for all BARs on both PCIe slave ports. 7. BAR programming with actual address values completes and read-back confirms correct values. 8. The system handshake register returns the expected completion value, confirming the remote endpoint has completed its enumeration sequence. 9. The test calls finish(0) indicating successful completion.",
        "Code Generation": "",
        "Meta Test Description": "This testcase performs PCIe device enumeration across two PCIe controller instances (DM0 and DM1). It begins by writing 0x0 to 0xE6004100 to initialize the system. Link training is invoked conditionally based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP) using link_training_dm0_x4(4) or link_training_dm1_x4(4). Cache coherency programming is performed by read-modify-write operations on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, setting bit fields [11:14], [3:6], [27:30], and [19:22] to 0xF using set_data(). After a wait_on(20), the same coherency registers are programmed again. The SII0 link status register at offset 0xC0 is polled via read_sii0_reg() until bits matching mask 0xD1 are all set, confirming link-up for PCIE0. Similarly, SII1 link status at 0xC0 is polled via read_sii1_reg() for PCIE1 link-up. Under DM0_RC, the Vendor ID is read from the endpoint via read_pcie_slv0_reg(0x0), the command register at offset 0x4 is written with 0x7 via write_pcie_slv0_reg() to enable IO, Memory, and Bus Master, and memory base programming functions mem_base_program_dm0_x4() and mem_base_program_dm1_x4() are called. System-level configuration registers at 0xE690000C through 0xE6900034 are written with 0x1. Cache coherency is then disabled by writing 0x0 to bit fields [19:22] and [27:30] of both coherency control registers. After wait_on(30), BAR sizing is performed on PCIe slave port 1 by writing 0xFFFFFFFF to offsets 0x10-0x24 via write_pcie_slv1_reg(), reading back via read_pcie_slv1_reg(), then programming actual BAR values. The same BAR sizing and programming sequence is repeated for PCIe slave port 0 via write_pcie_slv0_reg() and read_pcie_slv0_reg(). Finally, the test polls 0xE6004100 via read_reg() in a while loop waiting for the value 0x12345678 as the completion handshake, with wait_on(5) between iterations, and calls finish(0) upon success.",
        "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100 to initialize the system. 2. Invoke link_training_dm0_x4(4) or link_training_dm1_x4(4) based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). 3. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, apply set_data() to set bit fields [11:14] and [3:6] to 0xF, write back. 4. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF again, set bit fields [27:30] and [19:22] to 0xF, write back. 5. Repeat steps 3-4 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 6. Call wait_on(20). 7. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set all four bit field groups [11:14], [3:6], [27:30], [19:22] to 0xF, write back. 8. Repeat step 7 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 9. Read read_sii0_reg(0xC0) and poll in a while loop until (data_rd & 0xD1) == 0xD1 to confirm PCIE0 link-up. 10. Read read_sii1_reg(0xC0) and poll in a while loop until (data_rd & 0xD1) == 0xD1 to confirm PCIE1 link-up. 11. Under DM0_RC: read Vendor ID via read_pcie_slv0_reg(0x0), write 0x7 to write_pcie_slv0_reg(0x4) to enable IO/Memory/Bus Master, call mem_base_program_dm0_x4() and mem_base_program_dm1_x4(), wait_on(10). 12. Write 0x1 to system registers 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, 0xE6900034. 13. Disable cache: read-modify-write mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF setting [19:22] and [27:30] to 0x0. 14. Repeat step 13 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 15. Call wait_on(10), then consolidate cache disable for both controllers. 16. Call wait_on(30). 17. Write 0xFFFFFFFF to PCIe slave port 1 offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24 for BAR sizing. 18. Read back PCIe slave port 1 offsets 0x10-0x24 to determine BAR sizes. 19. Write actual BAR values to PCIe slave port 1 offsets 0x10-0x24. 20. Read back PCIe slave port 1 offsets 0x10-0x24 to verify. 21. Repeat steps 17-20 for PCIe slave port 0. 22. Call wait_on(10). 23. Poll read_reg(0xE6004100) in a while loop with wait_on(5) until value equals 0x12345678. 24. Call finish(0) to indicate test pass.",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
        "Meta Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. PCIe link training completes successfully for the configured controller mode. 2. The SII0 link status register reports link-up with the expected bit pattern confirming data link layer active and link-up status for PCIE0. 3. The SII1 link status register reports link-up with the expected bit pattern for PCIE1. 4. The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID from the endpoint device. 5. TYPE1_STATUS_COMMAND_REG is successfully written to enable IO, Memory, and Bus Master access. 6. BAR sizing returns valid size information for all BARs on both PCIe slave ports. 7. BAR programming with actual address values completes and read-back confirms correct values. 8. The system handshake register returns the expected completion value, confirming the remote endpoint has completed its enumeration sequence. 9. The test calls finish(0) indicating successful completion."
    },
    {
        "Index": 2,
        "SS / Module": "PCIE",
        "Feature": "PCIe DMA Write and Read",
        "Test Case Name": "pcie_dma_write_test",
        "Test Description": "Verifies PCIe DMA write and read data transfer operations across four DMA channels for both PCIe controller instances. The test performs link training, confirms link-up by polling the SII link status register, reads the TYPE1_DEV_ID_VEND_ID_REG to retrieve the endpoint Vendor ID, enables IO, Memory, and Bus Master access via TYPE1_STATUS_COMMAND_REG, configures BARs and memory base addresses, preloads source data into memory, unmasks DMA interrupts via DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF, sequentially programs and triggers DMA write transfers on channels 0 through 3 using DMA_WRITE_DOORBELL_OFF, waits for each transfer to complete via interrupt-driven handshake, then programs and triggers DMA read-back transfers on channels 0 through 3 using DMA_READ_DOORBELL_OFF. The interrupt handler reads DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF to determine which channels completed, then clears the interrupts via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF. The test confirms successful completion of all DMA transfers across all channels.",
        "Speed": "",
        "Mode": "",
        "Memory Start Offset": "",
        "Memory End Offset": "",
        "Remarks": "The testcase uses conditional compilation to select between two PCIe controller instances and between Root Complex and Endpoint modes. DMA transfers are performed sequentially across four channels, with each channel waiting for interrupt-driven completion before the next channel is triggered. The interrupt handler uses a shared flag to synchronize between the main test flow and the ISR. Source data is preloaded with two distinct patterns for verification purposes. Several register macros for the second PCIe controller instance could not be resolved from the available headers. The SII link status register and the system control register could not be mapped to named registers in the specification.",
        "Test Steps / Procedure": STEPS_ROW2,
        "Impacted Registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; DMA_WRITE_INT_MASK_OFF; DMA_READ_INT_MASK_OFF; DMA_WRITE_DOORBELL_OFF; DMA_READ_DOORBELL_OFF; DMA_WRITE_INT_STATUS_OFF; DMA_READ_INT_STATUS_OFF; DMA_WRITE_INT_CLEAR_OFF; DMA_READ_INT_CLEAR_OFF",
        "Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. PCIe link training completes successfully for the configured controller mode. 2. The SII link status register reports link-up with the expected bit pattern confirming data link layer active and link-up status. 3. The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID from the endpoint device. 4. TYPE1_STATUS_COMMAND_REG is successfully written to enable IO, Memory, and Bus Master access. 5. The system handshake register returns the expected completion value. 6. All four DMA write channel transfers complete successfully, each confirmed by an interrupt where DMA_WRITE_INT_STATUS_OFF reports the corresponding channel completion and the interrupt is cleared via DMA_WRITE_INT_CLEAR_OFF. 7. All four DMA read channel transfers complete successfully, each confirmed by an interrupt where DMA_READ_INT_STATUS_OFF reports the corresponding channel completion and the interrupt is cleared via DMA_READ_INT_CLEAR_OFF. 8. The test calls finish(0) indicating successful completion of all DMA operations.",
        "Code Generation": "",
        "Meta Test Description": "This testcase validates PCIe DMA write and read operations across four DMA channels on two PCIe controller instances (DM0 and DM1). The test begins by writing 0x0 to 0xE6004100 to initialize the system. Link training is invoked conditionally. Under DM0_RC, the SII0 link status register at offset 0xC0 is polled via read_sii0_reg() until (data_rd & 0xD1) == 0xD1. The Vendor ID is read via read_pcie_slv0_reg(0x0), and the command register at offset 0x4 is written with 0x7. DMA interrupts are unmasked via mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF. DMA write channels 0-3 are triggered via mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF. DMA read channels 0-3 are triggered via mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF. The interrupt handler reads mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF, clears via mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF and mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF.",
        "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100 to initialize the system. 2. Invoke link training based on compile-time defines. 3. Under DM0_RC: poll read_sii0_reg(0xC0) until (data_rd & 0xD1) == 0xD1. 4. Read Vendor ID via read_pcie_slv0_reg(0x0), write 0x7 to write_pcie_slv0_reg(0x4). 5. Call bar_program_dm0_x4(), mem_base_program_dm0_x4(). 6. Poll read_reg(0xE6004100) until value equals 0x12345678. 7. Preload source data: write 0xC0DEBEED and 0xF00DDEAF patterns. 8. Enable GIC interrupts. 9. Write 0x0 to mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF and DMA_READ_INT_MASK_OFF. 10. Program and trigger DMA write channels 0-3 via DMA_WRITE_DOORBELL_OFF. 11. Program and trigger DMA read channels 0-3 via DMA_READ_DOORBELL_OFF. 12. IRQ handler reads DMA_WRITE_INT_STATUS_OFF and DMA_READ_INT_STATUS_OFF, clears via DMA_WRITE_INT_CLEAR_OFF and DMA_READ_INT_CLEAR_OFF. 13. Call finish(0).",
        "Meta Impacted Registers": "0xE6004100; 0xC0; 0x0; 0x4; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_CLEAR_OFF",
        "Meta Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. PCIe link training completes successfully. 2. SII link status confirms link-up. 3. TYPE1_DEV_ID_VEND_ID_REG returns valid Vendor ID. 4. TYPE1_STATUS_COMMAND_REG enables IO, Memory, Bus Master. 5. System handshake returns expected value. 6. All four DMA write channels complete with interrupt confirmation. 7. All four DMA read channels complete with interrupt confirmation. 8. finish(0) called."
    },
    {
        "Index": 3,
        "SS / Module": "PCIE",
        "Feature": "PCIe Memory Write and Read",
        "Test Case Name": "pcie_mem_wr_rd_test",
        "Test Description": "Verifies PCIe memory write and read data integrity across both PCIe controller instances in Root Complex and Endpoint modes. The test performs link training, configures cache coherency via the COHERENCY_CONTROL_3_OFF register for both controller instances, polls the SII link status registers until link-up is confirmed, reads the TYPE1_DEV_ID_VEND_ID_REG to retrieve the endpoint Vendor ID, enables IO, Memory, and Bus Master access via TYPE1_STATUS_COMMAND_REG, configures BARs and memory base addresses, writes a synchronization signal to a system control register, disables cache coherency in a staged sequence, then performs memory write-read verification at multiple target addresses through the PCIe slave ports. The test confirms data integrity by writing known patterns and reading them back, and polls a system handshake register for the expected completion value.",
        "Speed": "",
        "Mode": "",
        "Memory Start Offset": "",
        "Memory End Offset": "",
        "Remarks": "The testcase uses conditional compilation to select between Root Complex and Endpoint modes and between dual-mode controller instances. In Root Complex mode, three memory addresses are tested with distinct data patterns. In Endpoint mode, five BAR1 addresses are tested with a uniform data pattern. Cache coherency is enabled before link-up and disabled in a staged two-step sequence after BAR and memory base configuration. The SII link status register and the system control register could not be mapped to named registers in the specification. One coherency control macro for the second PCIe instance could not be resolved from the available headers. Error counters are declared but not explicitly checked in the main test flow, suggesting the memory write-read utility function handles comparison internally.",
        "Test Steps / Procedure": STEPS_ROW3,
        "Impacted Registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG",
        "Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. PCIe link training completes successfully for the configured controller mode. 2. The SII link status register reports link-up with the expected bit pattern confirming data link layer active and link-up status. 3. The TYPE1_DEV_ID_VEND_ID_REG returns a valid Vendor ID from the endpoint device. 4. TYPE1_STATUS_COMMAND_REG is successfully written to enable IO, Memory, and Bus Master access. 5. All memory write-read operations through the PCIe slave ports complete successfully with data integrity confirmed by the write-read verification function. 6. The system handshake register returns the expected completion value, confirming the remote endpoint has completed its sequence. 7. The test calls finish(0) indicating successful completion.",
        "Code Generation": "",
        "Meta Test Description": "This testcase validates PCIe memory write and read operations across both PCIe controller instances (DM0 and DM1) in Root Complex and Endpoint modes. Cache coherency is programmed via mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. SII link status at 0xC0 is polled. Vendor ID read via read_pcie_slv0_reg(0x0). Command register at 0x4 written with 0x7. Memory write-read tests performed via pcie_slv0_mem_wr_rd() and pcie_slv1_mem_wr_rd(). System handshake at 0xE6004100 polled for 0x12345678.",
        "Meta Test Steps / Procedure": "1. Write 0x0 to 0xE6004100. 2. Invoke link training. 3. Program COHERENCY_CONTROL_3_OFF for both instances. 4. Call wait_on(20). 5. Re-program COHERENCY_CONTROL_3_OFF. 6. Poll SII link status at 0xC0. 7. Read Vendor ID via read_pcie_slv0_reg(0x0). 8. Write 0x7 to offset 0x4. 9. Configure BARs and memory base. 10. Call non_secure_prot_nic(). 11. Write 0x11111111 to 0xE6004100. 12. Disable cache coherency in staged sequence. 13. Call wait_on(30). 14. Perform memory write-read tests. 15. Poll 0xE6004100 for 0x12345678. 16. Call finish(0).",
        "Meta Impacted Registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4",
        "Meta Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. PCIe link training completes successfully. 2. SII link status confirms link-up. 3. TYPE1_DEV_ID_VEND_ID_REG returns valid Vendor ID. 4. TYPE1_STATUS_COMMAND_REG enables IO, Memory, Bus Master. 5. All memory write-read operations pass. 6. System handshake returns expected value. 7. finish(0) called."
    },
    {
        "Index": 4,
        "SS / Module": "PCIE",
        "Feature": "PCIe Register Write/Read Verification",
        "Test Case Name": "pcie_reg_wr_rd_test",
        "Test Description": "Verifies the reset default values and write-read data integrity of PCIe DBI controller registers, SII registers, and PHY registers across two PCIe controller instances. The test first reads all target registers and compares them against their expected default values of zero. It then writes the PHY reset control registers to bring the PHY out of reset before checking PHY register defaults. In the write-read phase, the test iterates through multiple data patterns and writes each pattern to all DBI controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF), SII registers, and PHY registers, applying appropriate write masks. After each write, the test reads back every register and compares the read value against the expected written value, accounting for write masks. The test reports pass if all default value checks and all write-read comparisons succeed with zero errors across both controller instances.",
        "Speed": "",
        "Mode": "",
        "Memory Start Offset": "",
        "Memory End Offset": "",
        "Remarks": "The testcase covers three register groups: DBI controller registers, SII registers, and PHY registers across two PCIe controller instances. Write masks are applied to SII and PHY registers to account for read-only or reserved bit fields. PHY registers use half-word access with bit extraction based on address alignment. The PHY reset control registers are written before PHY register access to ensure the PHY is out of reset. Only three of the six defined data patterns are used in the write-read loop. Several SII register macros, PHY reset control macros, and PHY register addresses for both controller instances could not be mapped to named registers in the specification. The second PCIe controller instance register macros could not be resolved from the available headers.",
        "Test Steps / Procedure": STEPS_ROW4,
        "Impacted Registers": "MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
        "Validation / Acceptance Criteria": "The test passes when all of the following conditions are met: 1. All five DBI controller registers for both PCIe instances read back their expected default value of zero during the reset value check phase. 2. All three SII registers for both instances read back their expected default value of zero. 3. All three PHY registers for both instances read back their expected default value of zero after PHY reset is released. 4. For each of the three write-read data patterns, all five DBI controller registers (MSI_CAP_OFF_08H_REG, MSI_CAP_OFF_10H_REG, FILTER_MASK_2_OFF, AXI_MSTR_MSG_ADDR_HIGH_OFF, UTILITY_OFF) for both instances read back the exact value that was written. 5. For each data pattern, all SII registers for both instances read back the written value masked with the applicable write mask. 6. For each PHY-specific data pattern, all PHY registers for both instances read back the written value masked with the PHY write mask after half-word extraction. 7. Both error counters remain zero, and the test completes with a pass indication.",
        "Code Generation": "",
        "Meta Test Description": "This testcase validates reset default values and write-read integrity of PCIe registers across multiple register groups. Register arrays: rc0_ctl_addr[5] = {mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG, mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG, mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF, mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF, mizar_PCIE0_DBI_DSP_UTILITY_OFF}, sii0_addr[3], phy0_addr[3] = {0xE68860B8, 0xE68862B8, 0xE68864B8}. chk_rst_val() reads all registers and compares against defaults. chk_rd_wr() writes patterns {0xffffffff, 0xaaaaaaaa, 0x55555555} and reads back with mask comparison.",
        "Meta Test Steps / Procedure": "1. Read rc0_ctl_addr[i] for i=0..4, compare against ctl_default[i]. 2. Read rc1_ctl_addr[i] for i=0..4, compare against ctl_default[i]. 3. Read sii0_addr[i] and sii1_addr[i] for i=0..2, compare against sii_default[i]. 4. Write mizar_PCIE0_SII_PHY_RST_CONTROL = 0x01203000 and mizar_PCIE1_SII_PHY_RST_CONTROL = 0x01203000. 5. Read phy0_addr[i] and phy1_addr[i] for i=0..2 with half-word extraction. 6. For j=0..2: write chk_val[j] to all register arrays. 7. Read back all registers and compare with expected values accounting for write masks. 8. Call finish(err2 || err1).",
        "Meta Impacted Registers": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8",
        "Meta Validation / Acceptance Criteria": "The test passes when: 1. All DBI registers read default zero. 2. All SII registers read default zero. 3. All PHY registers read default zero after reset release. 4. All write-read patterns match for DBI registers. 5. All write-read patterns match for SII registers with write masks. 6. All write-read patterns match for PHY registers with PHY write masks. 7. Both error counters are zero."
    }
]


def build_workbook():
    wb = Workbook()
    # ── TestPlan sheet ──
    ws_tp = wb.active
    ws_tp.title = "TestPlan"
    # ── MetaData sheet ──
    ws_md = wb.create_sheet("MetaData")
    ws_md.sheet_state = "veryHidden"

    # ── Styles ──
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ── Write TestPlan headers ──
    for ci, col_name in enumerate(TP_COLS, 1):
        c = ws_tp.cell(row=1, column=ci, value=col_name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    # ── Write MetaData headers ──
    for ci, col_name in enumerate(MD_COLS, 1):
        c = ws_md.cell(row=1, column=ci, value=col_name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    # ── Populate TestPlan rows ──
    for ri, row in enumerate(ROWS, 2):
        for ci, col_name in enumerate(TP_COLS, 1):
            val = row.get(col_name, "")
            c = ws_tp.cell(row=ri, column=ci, value=val)
            c.alignment = cell_align
            c.border = thin_border

    # ── Populate MetaData rows ──
    for ri, row in enumerate(ROWS, 2):
        md_vals = [
            row.get("Index", ""),
            row.get("Test Case Name", ""),
            row.get("Meta Test Description", ""),
            row.get("Meta Test Steps / Procedure", ""),
            row.get("Meta Impacted Registers", ""),
            row.get("Meta Validation / Acceptance Criteria", ""),
            "",  # Meta Headers
            "",  # Meta Macros
            ""   # Meta Arrays
        ]
        for ci, val in enumerate(md_vals, 1):
            c = ws_md.cell(row=ri, column=ci, value=val)
            c.alignment = cell_align
            c.border = thin_border

    # ── Column widths ──
    tp_widths = {
        "A": 7, "B": 14, "C": 30, "D": 32, "E": 60,
        "F": 10, "G": 10, "H": 20, "I": 18, "J": 50,
        "K": 90, "L": 60, "M": 70, "N": 16
    }
    for col_letter, width in tp_widths.items():
        ws_tp.column_dimensions[col_letter].width = width

    md_widths = {
        "A": 7, "B": 32, "C": 80, "D": 80, "E": 80,
        "F": 70, "G": 30, "H": 30, "I": 30
    }
    for col_letter, width in md_widths.items():
        ws_md.column_dimensions[col_letter].width = width

    # ── Row heights for TestPlan ──
    for ri in range(2, len(ROWS) + 2):
        ws_tp.row_dimensions[ri].height = 300

    # ── Freeze panes ──
    ws_tp.freeze_panes = "A2"
    ws_md.freeze_panes = "A2"

    # ── Code Generation dropdown ──
    dv = DataValidation(
        type="list",
        formula1='"Required,Not Required"',
        allow_blank=True,
        showDropDown=False
    )
    dv.error = "Please select Required or Not Required"
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select Code Generation status"
    dv.promptTitle = "Code Generation"
    code_gen_col = get_column_letter(len(TP_COLS))  # N
    dv.sqref = f"{code_gen_col}2:{code_gen_col}1000"
    ws_tp.add_data_validation(dv)

    # ── Save ──
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), FILENAME)
    wb.save(out_path)
    print(f"Workbook saved: {out_path}")
    print(f"File size: {os.path.getsize(out_path)} bytes")

    # ── Verify ──
    from openpyxl import load_workbook
    wb2 = load_workbook(out_path)
    assert "TestPlan" in wb2.sheetnames
    assert "MetaData" in wb2.sheetnames
    assert wb2["MetaData"].sheet_state == "veryHidden"
    assert wb2["TestPlan"].max_row == len(ROWS) + 1
    assert wb2["MetaData"].max_row == len(ROWS) + 1
    print("Validation PASSED")
    return out_path


if __name__ == "__main__":
    build_workbook()
