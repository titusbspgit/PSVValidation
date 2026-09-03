#!/usr/bin/env python3
"""Generate PCIE_TestPlan XLSX from JSON data.
Usage: python generate_xlsx.py
Requires: pip install openpyxl
"""
import json
import os
from datetime import datetime, timezone, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_FILES = [f for f in os.listdir(SCRIPT_DIR) if f.endswith('.xlsx.json')]
if not JSON_FILES:
    print("ERROR: No .xlsx.json file found in", SCRIPT_DIR)
    exit(1)

DATA_FILE = os.path.join(SCRIPT_DIR, sorted(JSON_FILES)[-1])
META_FILES = [f for f in os.listdir(SCRIPT_DIR) if f.endswith('_metadata.json')]
META_FILE = os.path.join(SCRIPT_DIR, sorted(META_FILES)[-1]) if META_FILES else None

with open(DATA_FILE, 'r') as f:
    testcases = json.load(f)

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
output_file = os.path.join(SCRIPT_DIR, f'PCIE_TestPlan_{timestamp}.xlsx')

COLUMNS = [
    'Index', 'SS / Module', 'Test Case Name', 'Feature',
    'Meta Headers', 'Meta Macros', 'Meta Arrays',
    'Speed', 'Mode', 'Memory Start Offset', 'Memory End Offset',
    'Meta Test Description', 'Test Description',
    'Meta Test Steps / Procedure', 'Test Steps / Procedure',
    'Meta Impacted Registers', 'Impacted Registers',
    'Meta Validation / Acceptance Criteria', 'Validation / Acceptance Criteria',
    'Remarks'
]

wb = Workbook()
ws = wb.active
ws.title = 'TestPlan'

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
cell_align = Alignment(vertical='top', wrap_text=True)

for col_idx, col_name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row_idx, tc in enumerate(testcases, 2):
    for col_idx, col_name in enumerate(COLUMNS, 1):
        value = tc.get(col_name, 'NA')
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_align
        cell.border = thin_border

for col_idx in range(1, len(COLUMNS) + 1):
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 30

ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = 'A2'

ws_meta = wb.create_sheet('Metadata')
meta_header_font = Font(bold=True, color='FFFFFF', size=11)
meta_header_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')

meta_data = [
    ['Property', 'Value'],
    ['IP_NAME', 'PCIE'],
    ['Total Testcases', str(len(testcases))],
    ['Generation Timestamp (IST)', now_ist.strftime('%Y-%m-%d %H:%M:%S IST')],
    ['Source Repository', 'titusbspgit/PSVValidation'],
    ['Source Branch', 'main'],
    ['Source Subdirectory', 'TestRepo/PCIE'],
    ['Output Directory', 'Test_Output/PCIE/TestPlan/'],
    ['Output Filename', os.path.basename(output_file)],
    ['Register Spec File', 'DWC_pcie_dbi_cpcie_dsp.txt'],
]

if META_FILE and os.path.exists(META_FILE):
    with open(META_FILE, 'r') as f:
        meta_json = json.load(f)
    for tc in meta_json.get('testcases', []):
        meta_data.append([f"Testcase {tc['index']}", f"{tc['folder_name']} - {tc['feature']}"])

for row_idx, row_data in enumerate(meta_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_meta.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.font = meta_header_font
            cell.fill = meta_header_fill
            cell.alignment = header_align
        else:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col_idx == 1:
                cell.font = Font(bold=True)

ws_meta.column_dimensions['A'].width = 30
ws_meta.column_dimensions['B'].width = 60

wb.save(output_file)
print(f'SUCCESS: Generated {output_file}')
print(f'  - TestPlan sheet: {len(testcases)} rows x {len(COLUMNS)} columns')
print(f'  - Metadata sheet: {len(meta_data)} rows')
print(f'  - Timestamp (IST): {now_ist.strftime("%Y-%m-%d %H:%M:%S")}')
