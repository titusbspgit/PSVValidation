import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Dict

import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

INPUT_PATH = Path('automation/testplan_input.json')

TESTPLAN_HEADERS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

METADATA_HEADERS = [
    'Index',
    'Test Case Name',
    'Meta Test Description',
    'Meta Test Steps / Procedure',
    'Meta Impacted Registers',
    'Meta Validation / Acceptance Criteria',
    'Meta Headers',
    'Meta Macros',
    'Meta Arrays'
]

TESTPLAN_COL_WIDTHS = [8, 14, 26, 24, 64, 10, 10, 18, 18, 26, 64, 64, 52, 24]
METADATA_COL_WIDTHS = [8, 24, 64, 64, 64, 64, 40, 40, 64]

BLUE_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
ALIGN_WRAP_TOP = Alignment(wrap_text=True, vertical='top')


def load_input() -> Dict:
    if not INPUT_PATH.exists():
        print(f"ERROR: Input file not found: {INPUT_PATH}")
        sys.exit(2)
    try:
        with open(INPUT_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except Exception as e:
        print(f"ERROR: Failed to read input JSON: {e}")
        sys.exit(2)


def validate_json_array(arr):
    if not isinstance(arr, list):
        print('ERROR: json_data must be a JSON array')
        sys.exit(3)


def set_header_style(ws, headers: List[str], col_widths: List[int]):
    ws.append(headers)
    for idx, cell in enumerate(ws[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        # Set column widths
        if idx-1 < len(col_widths):
            ws.column_dimensions[chr(64 + idx)].width = col_widths[idx-1]
    ws.freeze_panes = 'A2'
    # Apply auto filter on header row
    last_col_letter = chr(64 + len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"


def append_testplan_row(ws, obj: Dict):
    row = [
        obj.get('Index', ''),
        obj.get('SS / Module', ''),
        obj.get('Feature', ''),
        obj.get('Test Case Name', ''),
        obj.get('Test Description', ''),
        obj.get('Speed', ''),
        obj.get('Mode', ''),
        obj.get('Memory Start Offset', ''),
        obj.get('Memory End Offset', ''),
        obj.get('Remarks', ''),
        obj.get('Test Steps / Procedure', ''),
        obj.get('Impacted Registers', ''),
        obj.get('Validation / Acceptance Criteria', ''),
        obj.get('Code Generation (Required / Not)', ''),
    ]
    ws.append(row)


def append_metadata_row(ws, obj: Dict):
    row = [
        obj.get('Index', ''),
        obj.get('Test Case Name', ''),
        obj.get('Meta Test Description', ''),
        obj.get('Meta Test Steps / Procedure', ''),
        obj.get('Meta Impacted Registers', ''),
        obj.get('Meta Validation / Acceptance Criteria', ''),
        obj.get('Meta Headers', ''),
        obj.get('Meta Macros', ''),
        obj.get('Meta Arrays', ''),
    ]
    ws.append(row)


def style_data_rows(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = ALIGN_WRAP_TOP


def build_workbook(json_data: List[Dict]) -> Workbook:
    wb = Workbook()
    # Create TestPlan sheet as active
    ws_tp = wb.active
    ws_tp.title = 'TestPlan'
    set_header_style(ws_tp, TESTPLAN_HEADERS, TESTPLAN_COL_WIDTHS)

    # Create MetaData sheet and set to VeryHidden later
    ws_md = wb.create_sheet('MetaData')
    set_header_style(ws_md, METADATA_HEADERS, METADATA_COL_WIDTHS)

    # Append rows preserving order
    for obj in json_data:
        append_testplan_row(ws_tp, obj)
        append_metadata_row(ws_md, obj)

    # Style data rows
    style_data_rows(ws_tp)
    style_data_rows(ws_md)

    # Set MetaData sheet to VeryHidden
    ws_md.sheet_state = 'veryHidden'

    return wb


def save_and_validate(wb: Workbook, output_directory: str, ip_name: str) -> str:
    tz = pytz.timezone('Asia/Kolkata')
    ts = datetime.now(tz).strftime('%Y%m%d_%H%M%S')
    filename = f"{ip_name}_TestPlan_{ts}.xlsx"
    out_dir = Path(output_directory)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename

    # Save
    wb.save(out_path)

    # Reopen and validate
    wb2 = load_workbook(out_path)
    if 'TestPlan' not in wb2.sheetnames or 'MetaData' not in wb2.sheetnames:
        raise RuntimeError('Validation failed: required sheets missing')
    ws_tp = wb2['TestPlan']
    ws_md = wb2['MetaData']

    # Validate headers
    tp_headers = [c.value for c in ws_tp[1]]
    md_headers = [c.value for c in ws_md[1]]
    if tp_headers != TESTPLAN_HEADERS:
        raise RuntimeError('Validation failed: TestPlan headers mismatch')
    if md_headers != METADATA_HEADERS:
        raise RuntimeError('Validation failed: MetaData headers mismatch')

    # Validate MetaData is VeryHidden
    if ws_md.sheet_state != 'veryHidden':
        raise RuntimeError('Validation failed: MetaData not VeryHidden')

    return str(out_path)


def main():
    data = load_input()
    json_data = data.get('json_data', [])
    output_directory = data.get('output_directory', 'Test_Output/GPIO/TestPlan/')
    ip_name = data.get('IP_NAME', 'IP')

    validate_json_array(json_data)

    wb = build_workbook(json_data)
    out_path = save_and_validate(wb, output_directory, ip_name)

    print(f"SUCCESS: Generated workbook at {out_path}")


if __name__ == '__main__':
    main()
