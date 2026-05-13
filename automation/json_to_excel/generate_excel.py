import json, os, re, sys, subprocess
from datetime import datetime, timedelta, timezone
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Inputs
JSON_INPUT_PATH = os.environ.get('JSON_INPUT_PATH', 'automation/json_to_excel/input.json')
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', 'Test_Output/GPIO/TestPlan')
IP_NAME = os.environ.get('IP_NAME', 'GPIO')
COMMIT_MESSAGE = os.environ.get('COMMIT_MESSAGE', 'Final formatted Excel generated from JSON input')

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_COLS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

WRAP_COLS = {
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
}

BLUE_FILL = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')
THIN = Side(border_style='thin', color='000000')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT_TOP_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=False)
CENTER_TOP = Alignment(horizontal='center', vertical='top', wrap_text=False)


def fail(msg):
    print(f"ERROR: {msg}")
    sys.exit(1)


def load_json_array(path):
    if not os.path.exists(path):
        fail(f"JSON input not found at {path}")
    with open(path, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
        except Exception as e:
            fail(f"Invalid JSON: {e}")
    if not isinstance(data, list) or len(data) == 0:
        fail("JSON must be a non-empty array of objects")
    for i, rec in enumerate(data, 1):
        if not isinstance(rec, dict):
            fail(f"JSON element at index {i-1} is not an object")
    return data


def union_keys_preserve_first_seen(rows):
    keys = []
    seen = set()
    for rec in rows:
        for k in rec.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    return keys


def normalize_steps(text):
    if text is None:
        return ''
    s = str(text).strip()
    if not s:
        return ''
    # Split by line breaks first to retain logical items if present
    lines = [ln.strip() for ln in re.split(r'[\r\n]+', s) if ln.strip()]
    items = []
    if len(lines) == 1:
        # Try to split inline enumerations like "1) ... 2) ..." or "1. ... 2. ..."
        parts = re.split(r'(?:(?<=\))\s*(?=\d+\))|(?:(?<=\.)\s*(?=\d+\.))', s)
        # Fallback if no split happened
        if len(parts) == 1:
            # Try splitting on semicolons if used
            parts = [p.strip() for p in re.split(r'\s*;\s*', s) if p.strip()]
        # Clean potential numeric prefixes
        for p in parts:
            p2 = re.sub(r'^\s*\d+[\)\.-]?\s*', '', p).strip()
            if p2:
                items.append(p2)
    else:
        for ln in lines:
            p2 = re.sub(r'^\s*\d+[\)\.-]?\s*', '', ln).strip()
            if p2:
                items.append(p2)
    if not items:
        items = [s]
    # Renumber strictly 1., 2., 3., ... in a single wrapped cell using newlines
    return "\n".join([f"{i+1}. {it}" for i, it in enumerate(items)])


def write_sheet_data(ws, headers, rows):
    # Headers
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    # Rows
    for r, rec in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            val = rec.get(h, '')
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = 'A2'


def apply_header_style(ws, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL


def apply_borders(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def autofit_columns(ws):
    max_col = ws.max_column
    widths = [0]*(max_col+1)
    for c in range(1, max_col+1):
        max_len = 0
        for r in range(1, ws.max_row+1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            # consider wrapped lines
            for line in s.split('\n'):
                if len(line) > max_len:
                    max_len = len(line)
        widths[c] = min(max(10, max_len + 2), 80)
    for c in range(1, max_col+1):
        ws.column_dimensions[get_column_letter(c)].width = widths[c]


def autofit_row_heights(ws, wrap_cols_idx):
    # Approximate: base height 15pt per line
    base = 15
    for r in range(2, ws.max_row+1):
        max_lines = 1
        for c in wrap_cols_idx:
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            lines = str(v).count('\n') + 1
            if lines > max_lines:
                max_lines = lines
        ws.row_dimensions[r].height = base * max_lines


def apply_wrap_and_alignment(ws, headers):
    hdr_to_idx = {h: i+1 for i, h in enumerate(headers)}
    wrap_idx = [hdr_to_idx[h] for h in headers if h in WRAP_COLS]
    # Enable wrap on specific columns
    for r in range(2, ws.max_row+1):
        for h, c in hdr_to_idx.items():
            cell = ws.cell(row=r, column=c)
            if h in WRAP_COLS:
                cell.alignment = LEFT_TOP_WRAP
            elif h == 'Index':
                cell.alignment = CENTER_TOP
            else:
                # Default text left, top
                cell.alignment = LEFT_TOP
    # Header style
    apply_header_style(ws, ws.max_column)
    # Autofit
    autofit_columns(ws)
    autofit_row_heights(ws, wrap_idx)


def apply_data_validation(ws, headers):
    if 'Code Generation (Required / Not)' not in headers:
        return
    col_idx = headers.index('Code Generation (Required / Not)') + 1
    max_row = ws.max_row
    dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{max_row}")


def main():
    rows = load_json_array(JSON_INPUT_PATH)
    all_keys = union_keys_preserve_first_seen(rows)

    # Phase 1: create base workbook with Data sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    write_sheet_data(ws, all_keys, rows)

    # Create META sheet and set very hidden
    ws_meta = wb.create_sheet('Meta_data_sheet')
    meta_headers_present = [c for c in META_COLS if c in all_keys]
    # Write meta headers
    for c, h in enumerate(meta_headers_present, 1):
        ws_meta.cell(row=1, column=c, value=h)
    for r, rec in enumerate(rows, 2):
        for c, h in enumerate(meta_headers_present, 1):
            ws_meta.cell(row=r, column=c, value=rec.get(h, ''))
    ws_meta.sheet_state = 'veryHidden'

    # STEP 7: Normalize main sheet on the same worksheet
    ws.title = 'TestPlan'
    # Build normalized rows for main cols, removing META
    normalized_headers = [h for h in MAIN_COLS]
    normalized_rows = []
    for rec in rows:
        new_rec = {}
        for h in normalized_headers:
            val = rec.get(h, '')
            if h in ('Test Steps / Procedure', 'Validation / Acceptance Criteria'):
                val = normalize_steps(val)
            new_rec[h] = val
        normalized_rows.append(new_rec)

    # Clear existing content by overwriting and trimming
    # Write headers
    for c, h in enumerate(normalized_headers, 1):
        ws.cell(row=1, column=c, value=h)
    # Write rows
    for r, rec in enumerate(normalized_rows, 2):
        for c, h in enumerate(normalized_headers, 1):
            ws.cell(row=r, column=c, value=rec.get(h, ''))
    # Trim extra columns/rows if any
    if ws.max_column > len(normalized_headers):
        ws.delete_cols(len(normalized_headers)+1, ws.max_column - len(normalized_headers))
    if ws.max_row > len(normalized_rows) + 1:
        ws.delete_rows(len(normalized_rows) + 2, ws.max_row - (len(normalized_rows) + 1))

    # Strict formatting
    apply_wrap_and_alignment(ws, normalized_headers)
    apply_borders(ws)

    # Data validation on single column
    apply_data_validation(ws, normalized_headers)

    # Safety check: ensure no sheet named 'Data'
    if any(sh.title == 'Data' for sh in wb.worksheets):
        # Attempt deletion
        for sh in wb.worksheets:
            if sh.title == 'Data':
                wb.remove(sh)
        if any(sh.title == 'Data' for sh in wb.worksheets):
            fail("Safety check failed: 'Data' sheet still exists")

    # IST timestamp
    try:
        from zoneinfo import ZoneInfo
        ist_now = datetime.now(ZoneInfo('Asia/Kolkata'))
    except Exception:
        ist_now = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    date_str = ist_now.strftime('%Y%m%d')
    time_str = ist_now.strftime('%H%M%S')
    filename = f"{IP_NAME}_TestPlan_{date_str}_{time_str}.xlsx"

    # Ensure output dir exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, filename)

    # Save and validate
    wb.save(out_path)
    try:
        _ = load_workbook(out_path, read_only=True)
    except Exception as e:
        fail(f"XLSX validation failed: {e}")

    # Commit changes
    subprocess.run(['git', 'config', 'user.name', 'Ag-Emb-Mpsoc-Stage1 Agent'], check=True)
    subprocess.run(['git', 'config', 'user.email', 'actions@github.com'], check=True)
    subprocess.run(['git', 'add', out_path], check=True)
    # Avoid empty commit
    diff = subprocess.run(['git', 'diff', '--cached', '--quiet'])
    if diff.returncode != 0:
        subprocess.run(['git', 'commit', '-m', COMMIT_MESSAGE], check=True)
        subprocess.run(['git', 'push', 'origin', 'HEAD:main'], check=True)

    print(out_path)

if __name__ == '__main__':
    main()
