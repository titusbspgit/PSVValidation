import json, os, re, zipfile, hashlib
from copy import deepcopy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLS = [
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
]

ALLOWED_VALIDATION = ["Required", "Blank", "Not Required"]

INPUT_JSON_PATH = os.path.join("automation", "testplan_input.json")
OUTPUT_DIR = os.environ.get("OUTPUT_FILE_PATH", "Test_Output/GPIO/TestPlan")
OUTPUT_NAME = os.environ.get("OUTPUT_FILE_NAME", "GPIO_TestPlan.xlsx")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, OUTPUT_NAME)
REPORT_PATH = os.path.join(OUTPUT_DIR, OUTPUT_NAME.replace('.xlsx', '.report.json'))

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill("solid", fgColor="4472C4")  # Blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=False)
LEFT_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT_TOP = Alignment(horizontal="right", vertical="top")


def read_json():
    with open(INPUT_JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("JSON input must be a non-empty array")
    return data


def build_schema(data):
    seen = []
    for obj in data:
        if isinstance(obj, dict):
            for k in obj.keys():
                if k not in seen:
                    seen.append(k)
        else:
            raise ValueError("All items in JSON array must be objects")
    return seen


def normalize_rows(data, schema):
    norm = []
    for obj in data:
        row = {}
        for k in schema:
            row[k] = obj.get(k, "") if obj.get(k, "") is not None else ""
        for k in obj.keys():
            if k not in schema:
                schema.append(k)
                row[k] = obj.get(k, "") if obj.get(k, "") is not None else ""
        norm.append(row)
    return norm, schema


def create_workbook():
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Data")
    return wb, ws


def write_base_sheet(ws, schema, rows):
    for col_idx, key in enumerate(schema, start=1):
        c = ws.cell(row=1, column=col_idx, value=key)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(schema, start=1):
            val = row.get(key, "")
            ws.cell(row=r_idx, column=c_idx, value=val).border = THIN_BORDER
    ws.freeze_panes = "A2"


def estimate_and_apply_dimensions(ws, schema, rows):
    max_len = {k: len(str(k)) for k in schema}
    for row in rows:
        for k, v in row.items():
            l = max([len(s) for s in str(v).splitlines()]) if v is not None else 0
            if l > max_len.get(k, 0):
                max_len[k] = l
    for c_idx, k in enumerate(schema, start=1):
        width = min(max_len[k] + 2, 120)
        ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = width


def create_meta_sheet(wb, schema, rows):
    ws = wb.create_sheet("Meta_data_sheet")
    for col_idx, key in enumerate(META_COLS, start=1):
        c = ws.cell(row=1, column=col_idx, value=key)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(META_COLS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(key, "")).border = THIN_BORDER
    ws.sheet_state = "veryHidden"


def renumber_multiline(value):
    if value is None:
        return ""
    s = str(value)
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    if not lines:
        return s
    out = []
    pat = re.compile(r"^([\-\*\u2022\u25CF\u25A0\u25E6]+\s*)|(\d+[\.)]\s*)")
    for i, ln in enumerate(lines, start=1):
        ln = pat.sub("", ln).strip()
        out.append(f"{i}. {ln}")
    return "\n".join(out)


def rebuild_as_testplan(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    data_rows = [dict(zip(headers, r)) for r in rows[1:]]

    ws.title = "TestPlan"

    final_headers = MAIN_ORDER

    ws.delete_rows(1, ws.max_row)

    for c_idx, h in enumerate(final_headers, start=1):
        c = ws.cell(row=1, column=c_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER

    for r_idx, row in enumerate(data_rows, start=2):
        for c_idx, h in enumerate(final_headers, start=1):
            val = row.get(h, "")
            if h in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                val = renumber_multiline(val)
            ws.cell(row=r_idx, column=c_idx, value=val).border = THIN_BORDER

    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for h in WRAP_COLS:
        if h in header_to_col:
            col = header_to_col[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).alignment = LEFT_TOP_WRAP

    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).alignment = CENTER

    if "Index" in header_to_col:
        idx_col = header_to_col["Index"]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, idx_col).alignment = CENTER
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value != "Index":
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, c).alignment is None or not ws.cell(r, c).alignment.wrap_text:
                    ws.cell(r, c).alignment = LEFT_TOP

    schema = final_headers
    data = []
    for r in range(2, ws.max_row + 1):
        data.append({h: ws.cell(r, header_to_col[h]).value for h in schema})
    estimate_and_apply_dimensions(ws, schema, data)

    base_height = 15
    for r in range(2, ws.max_row + 1):
        max_lines = 1
        for h in WRAP_COLS:
            if h in header_to_col:
                val = ws.cell(r, header_to_col[h]).value
                if val is None:
                    continue
                lines = str(val).splitlines()
                if len(lines) > max_lines:
                    max_lines = len(lines)
        ws.row_dimensions[r].height = base_height * max_lines

    ws.freeze_panes = "A2"

    if "Code Generation (Required / Not)" in header_to_col:
        col = header_to_col["Code Generation (Required / Not)"]
        col_letter = ws.cell(1, col).column_letter
        dv = DataValidation(type="list", formula1='"' + ",".join(ALLOWED_VALIDATION) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        if ws.max_row >= 2:
            dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")


def validate_xlsx(path):
    if not os.path.exists(path):
        return False
    try:
        with zipfile.ZipFile(path, 'r') as z:
            names = set(z.namelist())
            return ('[Content_Types].xml' in names) and any(n.startswith('xl/workbook') for n in names)
    except Exception:
        return False


def build_report():
    report = {"file": OUTPUT_PATH}
    # bytes, sha256
    with open(OUTPUT_PATH, 'rb') as f:
        data = f.read()
    report["size_bytes"] = len(data)
    report["sha256"] = hashlib.sha256(data).hexdigest()
    # zip entries
    with zipfile.ZipFile(OUTPUT_PATH, 'r') as z:
        names = set(z.namelist())
    report["zip_has_[Content_Types].xml"] = ('[Content_Types].xml' in names)
    report["zip_has_xl_workbook"] = any(n.startswith('xl/workbook') for n in names)
    report["zip_entries_count"] = len(names)
    # workbook inspection
    wb = load_workbook(OUTPUT_PATH)
    report["sheets"] = []
    for ws in wb.worksheets:
        report["sheets"].append({"title": ws.title, "state": ws.sheet_state})
    report["has_Data_sheet"] = ('Data' in [s.title for s in wb.worksheets])
    report["expected_sheets_ok"] = (set([s["title"] for s in report["sheets"]]) == set(["TestPlan", "Meta_data_sheet"]))
    # formatting checks on TestPlan
    fmt = {"header_bold_blue_centered": True, "wrapped_columns": {}, "thin_borders_all_populated": True, "numbering_checks": {}}
    ws = wb["TestPlan"]
    max_col = ws.max_column
    # header checks
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        bold_ok = bool(cell.font and cell.font.bold)
        fill_ok = (cell.fill and getattr(cell.fill.fgColor, 'rgb', None) is not None and cell.fill.fgColor.rgb.endswith('4472C4')) or (cell.fill and getattr(cell.fill, 'fgColor', None) is not None and getattr(cell.fill.fgColor, 'rgb', '') .endswith('4472C4')) or (cell.fill and getattr(cell.fill, 'fgColor', None) in ("4472C4",))
        align_ok = (cell.alignment and cell.alignment.horizontal == 'center' and cell.alignment.vertical == 'center')
        if not (bold_ok and fill_ok and align_ok):
            fmt["header_bold_blue_centered"] = False
            break
    # wrap checks
    headers = {ws.cell(1, c).value: c for c in range(1, max_col + 1)}
    for h in WRAP_COLS:
        ok = True
        if h in headers:
            col = headers[h]
            for r in range(2, ws.max_row + 1):
                if not (ws.cell(r, col).alignment and ws.cell(r, col).alignment.wrap_text):
                    ok = False
                    break
        else:
            ok = False
        fmt["wrapped_columns"][h] = ok
    # thin borders (sample few rows to limit cost)
    for r in range(1, min(ws.max_row, 10) + 1):
        for c in range(1, max_col + 1):
            b = ws.cell(r, c).border
            sides = [b.left.style, b.right.style, b.top.style, b.bottom.style]
            if not all(s in ("thin", None) for s in sides):
                fmt["thin_borders_all_populated"] = False
    # numbering checks
    for h in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
        if h in headers and ws.max_row >= 2:
            val = str(ws.cell(2, headers[h]).value or "")
            fmt["numbering_checks"][h] = bool(re.search(r"(^|\n)1\. ", val))
        else:
            fmt["numbering_checks"][h] = False
    report["formatting"] = fmt
    # data validation checks
    dvs = []
    for dv in ws.data_validations.dataValidation:
        dvs.append({
            "type": dv.type,
            "formula1": dv.formula1,
            "ranges": list(dv.cells),
            "allow_blank": dv.allowBlank,
        })
    report["data_validations"] = dvs
    # compute exclusive validation on Code Generation column
    cg_col = headers.get("Code Generation (Required / Not)")
    only_cg = True
    allowed_list_ok = False
    header_row_excluded = True
    if dvs:
        for dv in dvs:
            if dv["type"] != "list":
                only_cg = False
            if '"' + ','.join(ALLOWED_VALIDATION) + '"' == (dv["formula1"] or ""):
                allowed_list_ok = True
            # ensure ranges map to cg_col and start at row 2
            for r in dv["ranges"]:
                # r like A2:A10
                m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", r)
                if m:
                    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                    col_idx = ws[ c1 + '1' ].column
                    col_idx2 = ws[ c2 + '1' ].column
                    if col_idx != cg_col or col_idx2 != cg_col:
                        only_cg = False
                    if r1 == 1:
                        header_row_excluded = False
                else:
                    only_cg = False
    report["validation_summary"] = {
        "only_code_generation_column": only_cg and (cg_col is not None),
        "allowed_values_exact": allowed_list_ok,
        "header_row_excluded": header_row_excluded,
    }

    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2)


def main():
    data = read_json()
    schema = build_schema(data)
    rows, schema = normalize_rows(data, schema)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb, ws = create_workbook()
    write_base_sheet(ws, schema, rows)
    estimate_and_apply_dimensions(ws, schema, rows)
    create_meta_sheet(wb, schema, rows)
    rebuild_as_testplan(ws)

    if 'Data' in [s.title for s in wb.worksheets]:
        s = wb['Data']
        wb.remove(s)
    titles = [s.title for s in wb.worksheets]
    if set(titles) - set(['TestPlan', 'Meta_data_sheet']):
        raise RuntimeError('Unexpected worksheets present: ' + ','.join(titles))

    wb.save(OUTPUT_PATH)

    if not validate_xlsx(OUTPUT_PATH):
        raise RuntimeError('Generated file is not a valid XLSX package')

    build_report()

    print(f"Saved: {OUTPUT_PATH}")
    print(f"Report: {REPORT_PATH}")

if __name__ == "__main__":
    main()
