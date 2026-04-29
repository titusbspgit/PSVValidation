#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os
import sys
import re
import zipfile
from datetime import datetime

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# -----------------------------
# Embedded JSON input (raw)
# -----------------------------
JSON_INPUT = r'''{ "TC1": { "Index": "1", "SS / Module": "GPIO", "Feature": "Supports AHB 32-bit for the programming of CSR", "Test Case Name": "gpio_reg_wr_rd_test", "Test Description": "Verifies reset defaults and masked read/write behavior across a set of GPIO control and status registers, ensuring reads match documented defaults and writes are honored only for writable bits.", "Speed": "NA", "Mode": "NA", "Memory Start Offset": "NA", "Memory End Offset": "NA", "Remarks": "VRRW-type registers are skipped for write/read testing; group input registers are excluded from default-value checks. A known behavior is noted: input data may default high when not driven, affecting comparisons if external drive is not controlled.", "Test Steps / Procedure": "Entry: test_case\n1. Perform default value checks across the following registers: GP0_GPIO_8..GP0_GPIO_39, GPIO_INTR_RAW_STCLR1, GP0_INTR1_INTR_EN1, GP0_INTR1_INTR_STS1, GP0_INTR2_INTR_EN1, GP0_INTR2_INTR_STS1, GPIO_IO_CTRL_GROUP1..GPIO_IO_CTRL_GROUP4, GPIO_DOUT_GROUP1..GPIO_DOUT_GROUP4, GPIO_DIN_GROUP1..GPIO_DIN_GROUP4.\n - For each register in the list, if it is in the default-skip set, skip the read. If its read mask equals zero, skip the read. Otherwise, READ the register, clear the least-significant bit from the read value, and compare strictly against the documented default value for that register.\n2. Conduct masked write-read verification for six data patterns across the same register set (excluding those in the write-skip set and those with zero write mask):\n - For each pattern, for each register that is writable, WRITE the pattern masked by the register’s write-enable mask.\n - Then, for each register eligible for readback (nonzero read mask and nonzero write mask and not skipped), READ the register, apply the register’s read mask to the value, and compute the expected value as a composition of the written bits on writable positions and preserved default bits on non-writable positions. Compare readback versus expected; count mismatches.\n3. Completion: If any default mismatches or write/read mismatches were detected, declare failure; otherwise, declare pass.", "Impacted Registers": [ "GP0_GPIO_8, GP0_GPIO_9, GP0_GPIO_10, GP0_GPIO_11, GP0_GPIO_12, GP0_GPIO_13, GP0_GPIO_14, GP0_GPIO_15, GP0_GPIO_16, GP0_GPIO_17, GP0_GPIO_18, GP0_GPIO_19, GP0_GPIO_20, GP0_GPIO_21, GP0_GPIO_22, GP0_GPIO_23, GP0_GPIO_24, GP0_GPIO_25, GP0_GPIO_26, GP0_GPIO_27, GP0_GPIO_28, GP0_GPIO_29, GP0_GPIO_30, GP0_GPIO_31, GP0_GPIO_32, GP0_GPIO_33"], "Validation / Acceptance Criteria": "Pass if: (1) For every register included in default checks, the read value (after clearing bit 0) equals the documented default; (2) For each of the six test patterns and each register included in write/read checks, the readback (after the register’s read mask is applied) equals the expected value computed from the written pattern on writable bits and preserved default values on non-writable bits; and (3) the test ends with a zero failure count. Any mismatch in either phase constitutes failure.", "Code Generation (Required / Not)": "", "Hidden_Test_Case_Name": "gpio_reg_wr_rd_test", "Hidden_Test_Description": "This test checks default (reset) values for a list of GPIO CSRs and verifies masked write/read functionality for each using provided read/write masks.", "Hidden_Remarks": "SKIPPING VRRW registers (write/read checks) and skipping default reads for certain group registers via skip_rst_array. Note: when reading default values, DIN becomes 1 automatically if no external force is applied; forcing zero changes bit-level selection and may cause mismatches with expected values.", "Hidden_Test_Steps_Procedure": "Entry point: test_case()\n1) Call chk_rst_val():\n - for (i = 0; i < CNT; i++):\n a) addr = addr_array[i] where addr_array = { MIZAR_GPIO_GP0_GPIO_8..MIZAR_GPIO_GP0_GPIO_39, MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_GPIO_GP0_INTR2_INTR_EN1, MIZAR_GPIO_GP0_INTR2_INTR_STS1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP1..MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, MIZAR_GPIO_GPIO_DOUT_GROUP1..MIZAR_GPIO_GPIO_DOUT_GROUP4, MIZAR_GPIO_GPIO_DIN_GROUP1..MIZAR_GPIO_GPIO_DIN_GROUP4 }.\n b) If (skip_rst_array[i] == 1) continue.\n c) If (read_mask_array[i] == 0x00000000) continue.\n d) READ: data_rd = read_reg(addr) [Operation: READ on macro addr].\n e) data = (data_rd & 0xFFFFFFFE).\n f) If (data == default_value_array[i]) pass; else def_fail_cnt++ and print failure.\n2) Call chk_rd_wr():\n - chk_val[6] = { 0xFFFFFFFF, 0xAAAAAAAA, 0x55555555, 0xF5F5F5F5, 0xA5A5A5A5, 0xFFFF0000 }.\n - for (j = 0; j < 6; j++):\n a) data_wr = chk_val[j].\n b) Write phase: for (i = 0; i < CNT; i++):\n i) addr = addr_array[i].\n ii) If (skip_array[i] == 1) continue.\n iii) If (write_mask_array[i] == 0x00000000) continue.\n iv) WRITE: write_reg(addr, (data_wr & write_mask_array[i])) [Operation: WRITE on macro addr with write mask applied].\n c) Read/compare phase: for (i = 0; i < CNT; i++):\n i) addr = addr_array[i].\n ii) If (skip_array[i] == 1) continue.\n iii) If (write_mask_array[i] == 0x00000000) continue.\n iv) If (read_mask_array[i] == 0x00000000) continue.\n v) READ: data_rd = (read_reg(addr) & read_mask_array[i]) [Operation: READ on macro addr with read mask applied].\n vi) wr_n = (write_mask_array[i] ^ 0xFFFFFFFF).\n vii) exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])).\n viii) If (data_rd == exp_val) pass; else wr_fail_cnt++ and print failure.\n3) If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0).\nTiming: No active waits in executed paths.\nPer-access macro references appear via addr_array as listed above.", "Hidden_Impacted_Registers": [ "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_GPIO_9", "MIZAR_GPIO_GP0_GPIO_10", "MIZAR_GPIO_GP0_GPIO_11", "MIZAR_GPIO_GP0_GPIO_12", "MIZAR_GPIO_GP0_GPIO_13", "MIZAR_GPIO_GP0_GPIO_14", "MIZAR_GPIO_GP0_GPIO_15", "MIZAR_GPIO_GP0_GPIO_16", "MIZAR_GPIO_GP0_GPIO_17", "MIZAR_GPIO_GP0_GPIO_18", "MIZAR_GPIO_GP0_GPIO_19", "MIZAR_GPIO_GP0_GPIO_20", "MIZAR_GPIO_GP0_GPIO_21", "MIZAR_GPIO_GP0_GPIO_22", "MIZAR_GPIO_GP0_GPIO_23", "MIZAR_GPIO_GP0_GPIO_24", "MIZAR_GPIO_GP0_GPIO_25", "MIZAR_GPIO_GP0_GPIO_26", "MIZAR_GPIO_GP0_GPIO_27", "MIZAR_GPIO_GP0_GPIO_28", "MIZAR_GPIO_GP0_GPIO_29", "MIZAR_GPIO_GP0_GPIO_30", "MIZAR_GPIO_GP0_GPIO_31", "MIZAR_GPIO_GP0_GPIO_32", "MIZAR_GPIO_GP0_GPIO_33", "MIZAR_GPIO_GP0_GPIO_34", "MIZAR_GPIO_GP0_GPIO_35", "MIZAR_GPIO_GP0_GPIO_36", "MIZAR_GPIO_GP0_GPIO_37", "MIZAR_GPIO_GP0_GPIO_38", "MIZAR_GPIO_GP0_GPIO_39", "MIZAR_GPIO_GPIO_INTR_RAW_STCLR1", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_GPIO_GP0_INTR2_INTR_EN1", "MIZAR_GPIO_GP0_INTR2_INTR_STS1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GPIO_DOUT_GROUP1", "MIZAR_GPIO_GPIO_DOUT_GROUP2", "MIZAR_GPIO_GPIO_DOUT_GROUP3", "MIZAR_GPIO_GPIO_DOUT_GROUP4", "MIZAR_GPIO_GPIO_DIN_GROUP1", "MIZAR_GPIO_GPIO_DIN_GROUP2", "MIZAR_GPIO_GPIO_DIN_GROUP3", "MIZAR_GPIO_GPIO_DIN_GROUP4" ], "Hidden_Validation_Acceptance_Criteria": "Default check: (read_reg(addr) & 0xFFFFFFFE) == default_value_array[i] for each address not skipped and with nonzero read mask. Write/read check: For each pattern, for each address not skipped and with nonzero masks, (read_reg(addr) & read_mask) == ((data_wr & read_mask & write_mask) | ((~write_mask) & read_mask & default_value)). Test passes if def_fail_cnt == 0 and wr_fail_cnt == 0; otherwise fails." }, "TC2": { "Index": "2", "SS / Module": "GPIO", "Feature": "Interrupts can be generated based on positive edge or negative edge or level high or level low detection at GPIO input", "Test Case Name": "test_gpio_negedge_intr_en", "Test Description": "Configures input mode and negative-edge interrupts per pin, then for each pin creates a falling edge and verifies that the pin and group interrupt statuses set and clear correctly without timeouts.", "Speed": "NA", "Mode": "Interrupt", "Memory Start Offset": "0xA0243ffc", "Memory End Offset": "0xA0243ffc", "Remarks": "The wait flag is armed before generating the edge to avoid race conditions. A bounded wait with a timeout of 5000 iterations (with 10-unit waits per iteration) prevents infinite hangs. All pads are driven high initially to define a known state.", "Test Steps / Procedure": "Entry: test_case\n1. Enable platform interrupt source for the selected GPIO instance at the system register level, and enable the corresponding CPU interrupt line (87 or 88).\n2. Drive the pad driver to all-high at 0xA0243ffc to establish a known starting level.\n3. Configure per-pin control for pins 8 through 39: set input mode, enable negative-edge detection, and clear any latched raw status.\n - For each pin, perform a WRITE to the corresponding per-pin control register (GP0_GPIO_8..GP0_GPIO_39), then wait 10 time units.\n4. For each bit position 0..31 (corresponding to pins 8..39):\n a. Pre-clear the group raw status via GPIO_INTR_RAW_STCLR1 with a write-one-to-clear for the current bit.\n b. Enable only the current bit in the group enable register GP0_INTR1_INTR_EN1 and wait 10 time units.\n c. Arm the waiter flag and generate a falling edge for only the current bit: write all-ones to 0xA0243ffc, wait 30 time units, then write an all-ones pattern with the current bit held low to 0xA0243ffc.\n d. Wait for the interrupt with a timeout of 5000 iterations, waiting 10 time units per iteration. If the timeout expires, record an error for the corresponding pin.\n5. On interrupt, the handler runs:\n a. Deassert the waiter flag and restore the pad driver to all-high at 0xA0243ffc.\n b. READ the per-pin control register for the active pin (GP0_GPIO_8..GP0_GPIO_39) and verify the input state reflects a low level following the falling edge.\n c. If the per-pin raw status indicates an event, READ the group status (GP0_INTR1_INTR_STS1) and confirm the bit corresponding to the active pin is set.\n d. Clear per-pin raw status by writing the clear bit while keeping input mode asserted to the same per-pin control register; also clear the group raw bit via GPIO_INTR_RAW_STCLR1 for that pin.\n e. Verify the group status (GP0_INTR1_INTR_STS1) reads back as zero after clears.\n f. Clear the system-level raw status (RAW_STCR1) for the selected GPIO instance and clear the CPU interrupt (87 or 88).\n g. If the per-pin raw status was not set, record an error.\n6. Completion: Report the accumulated error count as the test result.", "Impacted Registers": [ "INTR_EN1", "GP0_GPIO_8", "GP0_GPIO_9", "GP0_GPIO_10", "GP0_GPIO_11", "GP0_GPIO_12", "GP0_GPIO_13", "GP0_GPIO_14", "GP0_GPIO_15", "GP0_GPIO_16", "GP0_GPIO_17", "GP0_GPIO_18", "GP0_GPIO_19", "GP0_GPIO_20", "GP0_GPIO_21", "GP0_GPIO_22", "GP0_GPIO_23", "GP0_GPIO_24", "GP0_GPIO_25", "GP0_GPIO_26", "GP0_GPIO_27", "GP0_GPIO_28", "GP0_GPIO_29", "GP0_GPIO_30", "GP0_GPIO_31", "GP0_GPIO_32", "GP0_GPIO_33", "GP0_GPIO_34", "GP0_GPIO_35", "GP0_GPIO_36", "GP0_GPIO_37", "GP0_GPIO_38", "GP0_GPIO_39", "GPIO_INTR_RAW_STCLR1", "GP0_INTR1_INTR_EN1", "GP0_INTR1_INTR_STS1", "RAW_STCR1" ], "Validation / Acceptance Criteria": "Per pin: No timeout occurs while waiting for the interrupt; per-pin input state indicates low immediately after the falling edge; per-pin raw status indicates an event; the group status bit corresponding to the pin is set and then reads as zero after per-pin and group clears; the system raw status is cleared after service. The overall test passes if the accumulated error count is zero.", "Code Generation (Required / Not)": "", "Hidden_Test_Case_Name": "test_gpio_negedge_intr_en", "Hidden_Test_Description": "Negative-edge interrupt enable and verification across GPIO pins 8..39 with per-pin and group status checks and clears.", "Hidden_Remarks": "Arm the wait flag (int_pend=1) before generating the falling edge to avoid a race. Uses a bounded wait with timeout=5000 and wait_on(10) per poll. All pads are driven high initially (0xA0243ffc = 0xFFFFFFFF).", "Hidden_Test_Steps_Procedure": "Entry point: test_case()\n1) Initialize:\n - test_err = 0.\n - Conditionally enable GIC IRQ (87 or 88).\n - WRITE: MIZAR_LSS_SYSREG_INTR_EN1 = LSS_SYSREG_INTR_EN1_GPIO0_INTR or LSS_SYSREG_INTR_EN1_GPIO1_INTR (enable sysreg interrupt output).\n - WRITE: 0xA0243ffc = 0xFFFFFFFF (drive all pads high).\n2) Configure GPIOs 8..39 for input + negedge + clear raw:\n - for (i = 0; i < 32; i++): addr1 = MIZAR_GPIO_GP0_GPIO_8 + (i4);\n WRITE: write_reg(addr1, (1<<20) | (1<<18) | (1<<16)); // doe=1, neie=1, iclr=1\n wait_on(10).\n3) For each bit (i = 0..31):\n - wr_val = 1u << i.\n - WRITE: MIZAR_GPIO_GPIO_INTR_RAW_STCLR1 = wr_val (pre-clear group raw, W1C).\n - WRITE: MIZAR_GPIO_GP0_INTR1_INTR_EN1 = wr_val (enable only this bit).\n - wait_on(10).\n - int_pend = 1.\n - Generate falling edge on bit i:\n WRITE: 0xA0243ffc = 0xFFFFFFFF; wait_on(30);\n WRITE: 0xA0243ffc = ~wr_val; // bit i goes low\n - Poll for ISR: timeout=5000; while (int_pend && timeout--) wait_on(10);\n - If (timeout == 0) { print timeout error; test_err++; }\n4) ISR (Default_IRQHandler):\n - local_wr = 1u << i; int_pend = 0.\n - WRITE: 0xA0243ffc = 0xFFFFFFFF (restore high).\n - READ: rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i4)).\n - If ((rdata & 0x1) != 0) test_err++ (DIN should be 0 after negedge).\n - If ((rdata & 0x2) != 0x0) { // per-pin raw set\n READ: rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1);\n If ((rdata_grp & local_wr) == 0) test_err++;\n WRITE: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i4), (1<<20) | (1<<16)); // doe=1, iclr=1\n WRITE: MIZAR_GPIO_GPIO_INTR_RAW_STCLR1 = local_wr; // clear group raw\n READ: rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1);\n If (rdata_grp != 0x0) test_err++;\n#ifdef GPIO0\n WRITE: MIZAR_LSS_SYSREG_RAW_STCR1 = LSS_SYSREG_RAW_STCR1_GPIO0_INTR; GIC_ClearIRQ(87);\n#endif\n#ifdef GPIO1\n WRITE: MIZAR_LSS_SYSREG_RAW_STCR1 = LSS_SYSREG_RAW_STCR1_GPIO1_INTR; GIC_ClearIRQ(88);\n#endif\n } else { test_err++; }\n5) finish(test_err).\nTiming: wait_on(10) after config and enables; wait_on(30) before forcing low; ISR polling loop uses timeout=5000 with wait_on(10) per iteration.", "Hidden_Impacted_Registers": [ "MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_GPIO_9", "MIZAR_GPIO_GP0_GPIO_10", "MIZAR_GPIO_GP0_GPIO_11", "MIZAR_GPIO_GP0_GPIO_12", "MIZAR_GPIO_GP0_GPIO_13", "MIZAR_GPIO_GP0_GPIO_14", "MIZAR_GPIO_GP0_GPIO_15", "MIZAR_GPIO_GP0_GPIO_16", "MIZAR_GPIO_GP0_GPIO_17", "MIZAR_GPIO_GP0_GPIO_18", "MIZAR_GPIO_GP0_GPIO_19", "MIZAR_GPIO_GP0_GPIO_20", "MIZAR_GPIO_GP0_GPIO_21", "MIZAR_GPIO_GP0_GPIO_22", "MIZAR_GPIO_GP0_GPIO_23", "MIZAR_GPIO_GP0_GPIO_24", "MIZAR_GPIO_GP0_GPIO_25", "MIZAR_GPIO_GP0_GPIO_26", "MIZAR_GPIO_GP0_GPIO_27", "MIZAR_GPIO_GP0_GPIO_28", "MIZAR_GPIO_GP0_GPIO_29", "MIZAR_GPIO_GP0_GPIO_30", "MIZAR_GPIO_GP0_GPIO_31", "MIZAR_GPIO_GP0_GPIO_32", "MIZAR_GPIO_GP0_GPIO_33", "MIZAR_GPIO_GP0_GPIO_34", "MIZAR_GPIO_GP0_GPIO_35", "MIZAR_GPIO_GP0_GPIO_36", "MIZAR_GPIO_GP0_GPIO_37", "MIZAR_GPIO_GP0_GPIO_38", "MIZAR_GPIO_GP0_GPIO_39", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_RAW_STCR1" ], "Hidden_Validation_Acceptance_Criteria": "For each i in 0..31: no timeout in wait loop; in ISR: (rdata & 0x1) == 0 for DIN after negedge; per-pin raw bit indicated (rdata & 0x2) != 0; group status has bit set for local_wr; after clearing per-pin (iclr=1) and group W1C, readback of GP0_INTR1_INTR_STS1 == 0; system RAW_STCR1 bit is cleared after write. Test passes if test_err == 0." }, "TC3": { "Index": "3", "SS / Module": "GPIO", "Feature": "Interrupts can be generated based on positive edge or negative edge or level high or level low detection at GPIO input", "Test Case Name": "test_gpio_pedge_all_pads_en", "Test Description": "Enables positive-edge detection for all pins, configures input mode via group control, enables group interrupt output, and for each pin generates a rising edge while verifying group status set/clear and system-level status clear without timeouts.", "Speed": "NA", "Mode": "Interrupt", "Memory Start Offset": "0xA0243ffc", "Memory End Offset": "0xA0243ffc", "Remarks": "Waiter flag is armed before producing the rising edge to avoid races. Timeout of 2000 iterations with a 10-unit wait per iteration bounds the wait. Group interrupt is masked during ISR service and re-enabled afterward.", "Test Steps / Procedure": "Entry: test_case\n1. Enable the platform interrupt source for the selected GPIO instance at the system register level and enable the appropriate CPU interrupt line (87 or 88).\n2. Enable positive-edge detection for pins 8 through 39 by writing to each per-pin control register (GP0_GPIO_8..GP0_GPIO_39).\n3. After a 10-unit wait, configure pins 8..39 for input mode using the group IO control registers (GPIO_IO_CTRL_GROUP1..GPIO_IO_CTRL_GROUP4), then wait another 10 time units.\n4. Enable all bits in the group enable register GP0_INTR1_INTR_EN1.\n5. For each pin index 0..31:\n a. Drive the pad driver low at 0xA0243ffc and wait 10 units.\n b. Arm the waiter flag and drive the pad driver high at 0xA0243ffc to create a single rising edge.\n c. Poll for the interrupt with a timeout of 2000 iterations, waiting 10 units per iteration; on timeout, record an error and break from the loop.\n d. Optionally drive low again and wait 10 units to prepare for the next iteration.\n6. On interrupt, the handler runs:\n a. Read the group status register (GP0_INTR1_INTR_STS1) and mask the group by writing 0 to the group enable register.\n b. If any bit is set in the group status, treat as success for occurrence; otherwise record an error.\n c. Clear per-pin raw status by writing the clear bit for each per-pin control (GP0_GPIO_8..GP0_GPIO_39), wait 2 units, and verify the group status is zero.\n d. Clear the system raw status in RAW_STCR1 for the selected GPIO instance, then read back to confirm the bit is cleared; on failure, record an error.\n e. Re-enable the group output by writing all ones to GP0_INTR1_INTR_EN1, and clear the CPU interrupt (87 or 88).\n7. Completion: Report the accumulated error count as the result.", "Impacted Registers": [ "INTR_EN1", "GP0_GPIO_8", "GP0_GPIO_9", "GP0_GPIO_10", "GP0_GPIO_11", "GP0_GPIO_12", "GP0_GPIO_13", "GP0_GPIO_14", "GP0_GPIO_15", "GP0_GPIO_16", "GP0_GPIO_17", "GP0_GPIO_18", "GP0_GPIO_19", "GP0_GPIO_20", "GP0_GPIO_21", "GP0_GPIO_22", "GP0_GPIO_23", "GP0_GPIO_24", "GP0_GPIO_25", "GP0_GPIO_26", "GP0_GPIO_27", "GP0_GPIO_28", "GP0_GPIO_29", "GP0_GPIO_30", "GP0_GPIO_31", "GP0_GPIO_32", "GP0_GPIO_33", "GP0_GPIO_34", "GP0_GPIO_35", "GP0_GPIO_36", "GP0_GPIO_37", "GP0_GPIO_38", "GP0_GPIO_39", "GPIO_IO_CTRL_GROUP1", "GPIO_IO_CTRL_GROUP2", "GPIO_IO_CTRL_GROUP3", "GPIO_IO_CTRL_GROUP4", "GP0_INTR1_INTR_EN1", "GP0_INTR1_INTR_STS1", "RAW_STCR1" ], "Validation / Acceptance Criteria": "For each pin tested: no timeout occurs while waiting for the interrupt; group interrupt status indicates occurrence; after per-pin clears the group status reads zero; after clearing the system raw status, a readback shows the status cleared; after handler completion, group output is re-enabled for the next iteration. The test passes if the accumulated error count is zero.", "Code Generation (Required / Not)": "", "Hidden_Test_Case_Name": "test_gpio_pedge_all_pads_en", "Hidden_Test_Description": "Positive-edge interrupt enable for all pads with group IO configuration, per-pin raw clear, group status verification, and system-level raw clear checks.", "Hidden_Remarks": "int_pend is set before generating the rising edge. Timeout loop uses 2000 iterations with wait_on(10) to bound waiting. Group interrupt is masked during handler and re-enabled before exit.", "Hidden_Test_Steps_Procedure": "Entry point: test_case()\n1) Setup:\n - Conditionally enable GIC IRQ (87 or 88).\n - WRITE: MIZAR_LSS_SYSREG_INTR_EN1 = LSS_SYSREG_INTR_EN1_GPIO0_INTR or LSS_SYSREG_INTR_EN1_GPIO1_INTR (enable sysreg interrupt output).\n2) Configure posedge detection per pin:\n - for (i = 0; i < 32; i++): WRITE: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i4), 0x00020000); // peie=1\n - wait_on(10).\n3) Configure input mode via group IO control:\n - WRITE: MIZAR_GPIO_GPIO_IO_CTRL_GROUP1 = 0x000000FF;\n - WRITE: MIZAR_GPIO_GPIO_IO_CTRL_GROUP2 = 0x000000FF;\n - WRITE: MIZAR_GPIO_GPIO_IO_CTRL_GROUP3 = 0x000000FF;\n - WRITE: MIZAR_GPIO_GPIO_IO_CTRL_GROUP4 = 0x000000FF;\n - wait_on(10).\n4) Enable group interrupt output:\n - WRITE: MIZAR_GPIO_GP0_INTR1_INTR_EN1 = 0xFFFFFFFF.\n5) For each pin i in 0..31:\n - WRITE: 0xA0243ffc = 0x00000000; wait_on(10);\n - int_pend = 1; WRITE: 0xA0243ffc = 0xFFFFFFFF (posedge).\n - timeout = 2000; while (int_pend == 1 && --timeout > 0) wait_on(10);\n - If (timeout == 0) { print timeout error; test_err++; break; }\n - WRITE: 0xA0243ffc = 0x00000000; wait_on(10).\n6) ISR (Default_IRQHandler):\n - wr_val = 1 << i; int_pend = 0.\n - READ: rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1);\n - WRITE: MIZAR_GPIO_GP0_INTR1_INTR_EN1 = 0x00000000; // mask during service\n - If ((rdata_grp & 0xFFFFFFFF) != 0) success else { print error; test_err++; }\n - for (j = 0; j < 32; j++): WRITE: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (j*4), 0x00010000); // iclr=1\n - wait_on(2).\n - READ: rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); If (rdata_grp != 0) { print error; test_err++; }\n#ifdef GPIO0\n - WRITE: MIZAR_LSS_SYSREG_RAW_STCR1 = LSS_SYSREG_RAW_STCR1_GPIO0_INTR; READ back MIZAR_LSS_SYSREG_RAW_STCR1 and ensure bit cleared; else test_err++.\n#endif\n#ifdef GPIO1\n - WRITE: MIZAR_LSS_SYSREG_RAW_STCR1 = LSS_SYSREG_RAW_STCR1_GPIO1_INTR; READ back MIZAR_LSS_SYSREG_RAW_STCR1 and ensure bit cleared; else test_err++.\n#endif\n - WRITE: MIZAR_GPIO_GP0_INTR1_INTR_EN1 = 0xFFFFFFFF; GIC_ClearIRQ(87/88 as applicable).\n7) finish(test_err).\nTiming: wait_on(10) after configurations and around edge generation; timeout loop uses 2000 iterations with wait_on(10); ISR uses wait_on(2) after per-pin clears.", "Hidden_Impacted_Registers": [ "MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_GPIO_9", "MIZAR_GPIO_GP0_GPIO_10", "MIZAR_GPIO_GP0_GPIO_11", "MIZAR_GPIO_GP0_GPIO_12", "MIZAR_GPIO_GP0_GPIO_13", "MIZAR_GPIO_GP0_GPIO_14", "MIZAR_GPIO_GP0_GPIO_15", "MIZAR_GPIO_GP0_GPIO_16", "MIZAR_GPIO_GP0_GPIO_17", "MIZAR_GPIO_GP0_GPIO_18", "MIZAR_GPIO_GP0_GPIO_19", "MIZAR_GPIO_GP0_GPIO_20", "MIZAR_GPIO_GP0_GPIO_21", "MIZAR_GPIO_GP0_GPIO_22", "MIZAR_GPIO_GP0_GPIO_23", "MIZAR_GPIO_GP0_GPIO_24", "MIZAR_GPIO_GP0_GPIO_25", "MIZAR_GPIO_GP0_GPIO_26", "MIZAR_GPIO_GP0_GPIO_27", "MIZAR_GPIO_GP0_GPIO_28", "MIZAR_GPIO_GP0_GPIO_29", "MIZAR_GPIO_GP0_GPIO_30", "MIZAR_GPIO_GP0_GPIO_31", "MIZAR_GPIO_GP0_GPIO_32", "MIZAR_GPIO_GP0_GPIO_33", "MIZAR_GPIO_GP0_GPIO_34", "MIZAR_GPIO_GP0_GPIO_35", "MIZAR_GPIO_GP0_GPIO_36", "MIZAR_GPIO_GP0_GPIO_37", "MIZAR_GPIO_GP0_GPIO_38", "MIZAR_GPIO_GP0_GPIO_39", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_RAW_STCR1" ], "Hidden_Validation_Acceptance_Criteria": "For each tested pin: no timeout in the polling loop; group status (GP0_INTR1_INTR_STS1) indicates interrupt occurrence; after per-pin clears, group status reads 0; after writing RAW_STCR1 for the selected instance, readback indicates the bit is cleared. Test passes if test_err == 0." } }'''

IP_NAME = os.environ.get('IP_NAME', 'GPIO')
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', 'Test_Output/GPIO/TestPlan')
BRANCH = os.environ.get('BRANCH', 'main')
# Optional: Fixed timestamp passed by workflow; if not provided, use current IST
FIXED_TIMESTAMP = os.environ.get('FIXED_TIMESTAMP', '')

META_COLUMNS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_COLUMNS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)',
]

WRAP_COLUMNS = set([
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria',
])

BLUE_FILL = PatternFill(fill_type='solid', fgColor='4472C4')
WHITE_FONT = Font(color='FFFFFF', bold=True)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
TEXT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
INDEX_ALIGN = Alignment(horizontal='center', vertical='top', wrap_text=True)
BORDER_THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def parse_json(json_text):
    try:
        data = json.loads(json_text)
    except Exception as e:
        print(f"ERROR: Invalid JSON input: {e}")
        sys.exit(2)
    # Accept dict (ordered) or list
    if isinstance(data, dict):
        records = [v for k, v in data.items()]
    elif isinstance(data, list):
        records = data
    else:
        print("ERROR: Top-level JSON must be an object or an array")
        sys.exit(2)
    if not records:
        print("ERROR: JSON array is empty")
        sys.exit(2)
    # Ensure each record is a dict
    for i, r in enumerate(records, 1):
        if not isinstance(r, dict):
            print(f"ERROR: Record {i} is not an object")
            sys.exit(2)
    return records


def build_header_order(records):
    header = []
    seen = set()
    for r in records:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                header.append(k)
    return header


def approximate_width(value):
    if value is None:
        return 0
    if not isinstance(value, str):
        value = str(value)
    # account for newlines: take longest line
    parts = value.split('\n')
    return max((len(p) for p in parts), default=0)


def set_column_widths(ws, max_row, max_col, cap=80):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            l = approximate_width(v)
            if l > max_len:
                max_len = l
        width = min(max_len + 2, cap)
        ws.column_dimensions[letter].width = width


def strip_bullets(line: str) -> str:
    return re.sub(r'^(?:\s*(?:\d+\.|[-*•])\s*)', '', line).strip()


def renumber_multiline(text: str) -> str:
    if text is None:
        return ''
    if not isinstance(text, str):
        text = str(text)
    # normalize newlines
    lines = [ln.strip() for ln in text.replace('\r\n', '\n').replace('\r', '\n').split('\n')]
    lines = [ln for ln in lines if ln.strip() != '']
    numbered = []
    for i, ln in enumerate(lines, 1):
        ln2 = strip_bullets(ln)
        numbered.append(f"{i}. {ln2}")
    return '\n'.join(numbered)


def ensure_only_allowed_sheets(wb):
    names = [ws.title for ws in wb.worksheets]
    if 'TestPlan' not in names or 'Meta_data_sheet' not in names:
        return False
    # No sheet named 'Data' allowed
    if 'Data' in names:
        # try to delete
        try:
            ws = wb['Data']
            wb.remove(ws)
        except Exception:
            return False
    # Ensure only two sheets exist
    names = [ws.title for ws in wb.worksheets]
    if set(names) != {'TestPlan', 'Meta_data_sheet'}:
        # Allow exactly these two
        return False
    # Meta sheet must be veryHidden
    try:
        if wb['Meta_data_sheet'].sheet_state != 'veryHidden':
            return False
    except Exception:
        return False
    return True


def validate_xlsx(path: str) -> bool:
    # Valid Zip and basic OOXML structure
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            namelist = set(zf.namelist())
            required = {
                '[Content_Types].xml',
                '_rels/.rels',
                'xl/workbook.xml',
            }
            if not required.issubset(namelist):
                print('ERROR: XLSX missing required OOXML parts')
                return False
    except Exception as e:
        print(f'ERROR: XLSX zip validation failed: {e}')
        return False
    # Load with openpyxl
    try:
        _ = load_workbook(path, read_only=True)
    except Exception as e:
        print(f'ERROR: openpyxl load failed: {e}')
        return False
    return True


def main():
    records = parse_json(JSON_INPUT)
    header_order = build_header_order(records)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'  # authoritative staging sheet

    # Write headers
    for j, h in enumerate(header_order, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
    ws.freeze_panes = 'A2'

    # Write data rows (preserve values). Missing -> blank
    for i, rec in enumerate(records, start=2):
        for j, h in enumerate(header_order, 1):
            v = rec.get(h, '')
            ws.cell(row=i, column=j, value=v)

    set_column_widths(ws, ws.max_row, ws.max_column)

    # Create META sheet
    meta = wb.create_sheet('Meta_data_sheet')
    # header
    for j, h in enumerate(META_COLUMNS, 1):
        cell = meta.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)
    # rows
    for i, rec in enumerate(records, start=2):
        for j, h in enumerate(META_COLUMNS, 1):
            v = rec.get(h, '')
            if isinstance(v, (list, dict)):
                try:
                    v = json.dumps(v, ensure_ascii=False, separators=(', ', ': '))
                except Exception:
                    v = str(v)
            meta.cell(row=i, column=j, value=v)
    # Very hidden
    meta.sheet_state = 'veryHidden'

    # Normalize MAIN sheet directly on staging sheet
    ws = wb['Data']
    ws.title = 'TestPlan'  # rename existing, do NOT create new visible sheet

    # Remove META columns
    current_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    meta_indices = [idx + 1 for idx, name in enumerate(current_headers) if name in META_COLUMNS]
    for col_idx in sorted(meta_indices, reverse=True):
        ws.delete_cols(col_idx, 1)

    # Rebuild header/data to exactly MAIN_COLUMNS order, insert missing columns as blanks
    # Read current rows into list of dicts
    current_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    row_dicts = []
    for r in range(2, ws.max_row + 1):
        d = {}
        for c, h in enumerate(current_headers, 1):
            d[h] = ws.cell(row=r, column=c).value
        row_dicts.append(d)

    # Overwrite header to MAIN_COLUMNS and data accordingly
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c, value=None)
    for j, h in enumerate(MAIN_COLUMNS, 1):
        ws.cell(row=1, column=j, value=h)
    # clear extra columns beyond MAIN
    if ws.max_column > len(MAIN_COLUMNS):
        ws.delete_cols(len(MAIN_COLUMNS) + 1, ws.max_column - len(MAIN_COLUMNS))

    # write rows in MAIN order
    for r_index, d in enumerate(row_dicts, start=2):
        for j, h in enumerate(MAIN_COLUMNS, 1):
            v = d.get(h, '')
            # MAIN sheet: if list/array, join with ', '
            if isinstance(v, list):
                v = ', '.join(str(x) for x in v)
            ws.cell(row=r_index, column=j, value=v)

    # Mandatory numbering in two columns
    def apply_numbering(col_name: str):
        try:
            j = MAIN_COLUMNS.index(col_name) + 1
        except ValueError:
            return
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=j).value
            ws.cell(row=r, column=j, value=renumber_multiline(v))

    apply_numbering('Test Steps / Procedure')
    apply_numbering('Validation / Acceptance Criteria')

    # Strict formatting
    max_row, max_col = ws.max_row, ws.max_column

    # Header formatting
    for j in range(1, max_col + 1):
        cell = ws.cell(row=1, column=j)
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = HEADER_ALIGN

    # Data rows formatting
    for r in range(2, max_row + 1):
        for j in range(1, max_col + 1):
            cell = ws.cell(row=r, column=j)
            # Wrap for specific columns
            if ws.cell(row=1, column=j).value in WRAP_COLUMNS:
                cell.alignment = TEXT_ALIGN
            else:
                # Index centered, others left/top
                if ws.cell(row=1, column=j).value == 'Index':
                    cell.alignment = INDEX_ALIGN
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

    # Borders on all populated cells
    for r in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            ws.cell(row=r, column=j).border = BORDER_THIN

    # Freeze top row remains
    ws.freeze_panes = 'A2'

    # Autofit columns and approximate row heights
    set_column_widths(ws, ws.max_row, ws.max_column)

    # Data validation ONLY for Code Generation (Required / Not)
    try:
        code_col_idx = MAIN_COLUMNS.index('Code Generation (Required / Not)') + 1
        col_letter = get_column_letter(code_col_idx)
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
    except Exception as e:
        print(f"WARNING: Data validation setup failed: {e}")

    # Enforce final sheet visibility and existence
    if 'Data' in [s.title for s in wb.worksheets]:
        try:
            wb.remove(wb['Data'])
        except Exception:
            pass

    # Re-check allowed sheets
    if not ensure_only_allowed_sheets(wb):
        print('ERROR: Final workbook sheet visibility/state validation failed')
        sys.exit(3)

    # Determine output file name
    if FIXED_TIMESTAMP:
        ts = FIXED_TIMESTAMP
    else:
        if ZoneInfo is not None:
            tz = ZoneInfo('Asia/Kolkata')
            now = datetime.now(tz)
        else:
            # Fallback without zoneinfo: assume local time is IST (best effort)
            now = datetime.now()
        ts = now.strftime('%Y%m%d_%H%M%S')

    file_name = f"{IP_NAME}_TestPlan_{ts}.xlsx"
    out_dir = OUTPUT_DIR.rstrip('/')
    out_path = os.path.join(out_dir, file_name)
    os.makedirs(out_dir, exist_ok=True)

    wb.save(out_path)

    if not validate_xlsx(out_path):
        print('ERROR: XLSX validation failed after save')
        sys.exit(4)

    # Expose path to GitHub Actions
    print(f"Generated: {out_path}")
    gh_out = os.environ.get('GITHUB_OUTPUT')
    if gh_out:
        with open(gh_out, 'a', encoding='utf-8') as f:
            f.write(f"xlsx_path={out_path}\n")


if __name__ == '__main__':
    main()

# retrigger workflow
