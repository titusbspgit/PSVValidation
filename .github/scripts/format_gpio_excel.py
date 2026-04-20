import os
import re
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

IP_NAME = "GPIO"
INPUT_DIR = os.path.join("Test_Output", IP_NAME, "TestPlan")
OUTPUT_DIR = INPUT_DIR

META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}


def find_latest_xlsx(path: str) -> str:
    candidates = []
    for name in os.listdir(path):
        if name.lower().endswith(".xlsx"):
            candidates.append(name)
    if not candidates:
        raise FileNotFoundError("No .xlsx found in " + path)
    # Prefer files matching naming pattern that include timestamp; sort lexicographically
    candidates.sort()
    return os.path.join(path, candidates[-1])


def get_headers(ws):
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value is not None else "")
    return headers


def copy_meta_sheet(main_ws, wb):
    headers = get_headers(main_ws)
    header_to_col = {str(h): idx + 1 for idx, h in enumerate(headers)}
    meta_ws = wb.create_sheet("Meta_data_sheet")
    # Write headers for META columns
    for j, h in enumerate(META_COLS, start=1):
        meta_ws.cell(row=1, column=j, value=h)
    max_row = main_ws.max_row
    for i in range(2, max_row + 1):
        for j, h in enumerate(META_COLS, start=1):
            col_idx = header_to_col.get(h)
            val = main_ws.cell(row=i, column=col_idx).value if col_idx else None
            meta_ws.cell(row=i, column=j, value=val)
    # Very hidden
    meta_ws.sheet_state = 'veryHidden'
    return meta_ws


def build_normalized_main(main_ws, wb):
    headers = get_headers(main_ws)
    header_to_col = {str(h): idx + 1 for idx, h in enumerate(headers)}

    # Create a temporary sheet with target schema
    tmp_ws = wb.create_sheet("_TMP_")
    for j, h in enumerate(MAIN_ORDER, start=1):
        tmp_ws.cell(row=1, column=j, value=h)

    max_row = main_ws.max_row
    for i in range(2, max_row + 1):
        for j, h in enumerate(MAIN_ORDER, start=1):
            col_idx = header_to_col.get(h)
            val = main_ws.cell(row=i, column=col_idx).value if col_idx else None
            tmp_ws.cell(row=i, column=j, value=val)

    # Delete original main sheet and rename
    wb.remove(main_ws)
    tmp_ws.title = "TestPlan"
    return tmp_ws


def apply_formatting(ws):
    # Header formatting
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    data_top = Alignment(vertical="top")
    left = Alignment(horizontal="left", vertical="top", wrap_text=False)
    center = Alignment(horizontal="center", vertical="top", wrap_text=False)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Determine column indices for wrap columns
    headers = get_headers(ws)
    wrap_idx = {headers.index(h) + 1 for h in WRAP_COLS if h in headers}

    max_row = ws.max_row
    max_col = ws.max_column

    # Freeze top row and add filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Apply header formatting
    for col in range(1, max_col + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = header_align
        c.border = border

    # Apply data formatting and wrapping
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if c in wrap_idx:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            else:
                # Index column is the first column
                if c == 1:
                    cell.alignment = center
                else:
                    cell.alignment = left
            cell.border = border

    # Autofit columns (approximate based on content length)
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                ln = 0
            else:
                s = str(val)
                # Consider line breaks for wrapped columns
                ln = max((len(line) for line in s.splitlines()), default=0)
            if ln > max_len:
                max_len = ln
        # Add padding
        ws.column_dimensions[column_letter].width = min(max_len + 2, 80)


def main():
    src_path = find_latest_xlsx(INPUT_DIR)
    wb = load_workbook(src_path)

    # Identify main sheet: prefer 'TestPlan' else first visible
    main_ws = None
    if 'TestPlan' in wb.sheetnames:
        main_ws = wb['TestPlan']
    else:
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state == 'visible':
                main_ws = ws
                break
    if main_ws is None:
        raise RuntimeError("No visible worksheet found")

    # Create Meta sheet and copy meta columns
    copy_meta_sheet(main_ws, wb)

    # Build normalized main sheet with required columns only
    main_ws = build_normalized_main(main_ws, wb)

    # Ensure only one visible sheet named 'TestPlan' plus very hidden Meta_data_sheet
    for name in list(wb.sheetnames):
        if name not in ("TestPlan", "Meta_data_sheet"):
            ws = wb[name]
            wb.remove(ws)

    # Apply formatting to TestPlan only
    apply_formatting(main_ws)

    # Timezone IST
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    ts = now_ist.strftime("%Y%m%d_%H%M%S")
    out_name = f"{IP_NAME}_TestPlan_{ts}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_name)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
