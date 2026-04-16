import json, sys
from openpyxl import load_workbook

p_json = sys.argv[1] if len(sys.argv) > 1 else 'Test_Output/GPIO/input.json'
p_xlsx = sys.argv[2] if len(sys.argv) > 2 else 'Test_Output/GPIO/TestPlanGPIO.xlsx'

with open(p_json, 'r', encoding='utf-8') as f:
    data = json.load(f)
wb = load_workbook(p_xlsx, data_only=True)
ws = wb.active
# Read header
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))[:len(data[0].keys())]]
if hdr != list(data[0].keys()):
    raise SystemExit(f'ERROR: Excel headers differ.\nExcel: {hdr}\nJSON: {list(data[0].keys())}')
# Read rows
rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    rows.append(list(row[:len(hdr)]))
json_rows = [[r[k] for k in hdr] for r in data]
if len(rows) != len(json_rows):
    raise SystemExit(f'ERROR: Row count mismatch. Excel={len(rows)} JSON={len(json_rows)}')
for i, (a, b) in enumerate(zip(rows, json_rows), 1):
    if a != b:
        raise SystemExit(f'ERROR: Row {i} mismatch. Excel={a} JSON={b}')
print('OK: Excel values match JSON exactly.')
