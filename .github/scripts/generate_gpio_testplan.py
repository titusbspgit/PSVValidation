import os
import json
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Inputs via environment
IP_NAME = os.getenv("IP_NAME", "GPIO")
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "Test_Output/GPIO/TestPlan/")

# Aggregated JSON (verbatim from Stage 6)
aggregated_json = [
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "NA",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Read and record values from resolved GPIO GP0 registers: gp0_gpio_8, gp0_gpio_14, gp0_gpio_18, gp0_gpio_22, and gp0_gpio_26.",
    "Meta Test Description": "The testcase sequentially performs read operations on a list of GPIO GP0 register macros. Based on the provided UI-spec mapping: MIZAR_GPIO_GP0_GPIO_8 -> gp0_gpio_8 (matched), MIZAR_GPIO_GP0_GPIO_9 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_10 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_11 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_12 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_13 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_14 -> gp0_gpio_14 (matched), MIZAR_GPIO_GP0_GPIO_15 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_16 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_17 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_18 -> gp0_gpio_18 (matched), MIZAR_GPIO_GP0_GPIO_19 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_20 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_21 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_22 -> gp0_gpio_22 (matched), MIZAR_GPIO_GP0_GPIO_23 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_24 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_25 -> NA (unresolved), MIZAR_GPIO_GP0_GPIO_26 -> gp0_gpio_26 (matched), MIZAR_GPIO_GP0_GPIO_27 -> NA (unresolved), SOFT_RST_REG_ADDRESS -> NA (unresolved). Only matched items are considered resolved to concrete register names; all others remain unresolved as provided. No expected values, comparisons, interrupts, delays, or write operations are specified in the inputs.",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "NA",
    "Test Steps / Procedure": "1. Identify the resolved GPIO GP0 registers from the mapping: gp0_gpio_8, gp0_gpio_14, gp0_gpio_18, gp0_gpio_22, gp0_gpio_26.\n2. Read each resolved register once and record the captured values.\n3. Note any items marked unresolved in the mapping as out of scope for this run.",
    "Meta Test Steps / Procedure": "1) Read macro MIZAR_GPIO_GP0_GPIO_8 -> mapped to gp0_gpio_8; capture the read value.\n2) Read macro MIZAR_GPIO_GP0_GPIO_9 -> unresolved (NA); record as not mapped.\n3) Read macro MIZAR_GPIO_GP0_GPIO_10 -> unresolved (NA); record as not mapped.\n4) Read macro MIZAR_GPIO_GP0_GPIO_11 -> unresolved (NA); record as not mapped.\n5) Read macro MIZAR_GPIO_GP0_GPIO_12 -> unresolved (NA); record as not mapped.\n6) Read macro MIZAR_GPIO_GP0_GPIO_13 -> unresolved (NA); record as not mapped.\n7) Read macro MIZAR_GPIO_GP0_GPIO_14 -> mapped to gp0_gpio_14; capture the read value.\n8) Read macro MIZAR_GPIO_GP0_GPIO_15 -> unresolved (NA); record as not mapped.\n9) Read macro MIZAR_GPIO_GP0_GPIO_16 -> unresolved (NA); record as not mapped.\n10) Read macro MIZAR_GPIO_GP0_GPIO_17 -> unresolved (NA); record as not mapped.\n11) Read macro MIZAR_GPIO_GP0_GPIO_18 -> mapped to gp0_gpio_18; capture the read value.\n12) Read macro MIZAR_GPIO_GP0_GPIO_19 -> unresolved (NA); record as not mapped.\n13) Read macro MIZAR_GPIO_GP0_GPIO_20 -> unresolved (NA); record as not mapped.\n14) Read macro MIZAR_GPIO_GP0_GPIO_21 -> unresolved (NA); record as not mapped.\n15) Read macro MIZAR_GPIO_GP0_GPIO_22 -> mapped to gp0_gpio_22; capture the read value.\n16) Read macro MIZAR_GPIO_GP0_GPIO_23 -> unresolved (NA); record as not mapped.\n17) Read macro MIZAR_GPIO_GP0_GPIO_24 -> unresolved (NA); record as not mapped.\n18) Read macro MIZAR_GPIO_GP0_GPIO_25 -> unresolved (NA); record as not mapped.\n19) Read macro MIZAR_GPIO_GP0_GPIO_26 -> mapped to gp0_gpio_26; capture the read value.\n20) Read macro MIZAR_GPIO_GP0_GPIO_27 -> unresolved (NA); record as not mapped.\n21) Read macro SOFT_RST_REG_ADDRESS -> unresolved (NA); record as not mapped.\nNo explicit compares, masks, loops, waits, or assertions are provided in the inputs; steps reflect read operations and logging only.",
    "Impacted Registers": "gp0_gpio_8; gp0_gpio_14; gp0_gpio_18; gp0_gpio_22; gp0_gpio_26",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27; SOFT_RST_REG_ADDRESS",
    "Validation / Acceptance Criteria": "Reads of gp0_gpio_8, gp0_gpio_14, gp0_gpio_18, gp0_gpio_22, and gp0_gpio_26 complete successfully and values are captured; no expected value checks are defined.",
    "Meta Validation / Acceptance Criteria": "NA",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "NA",
    "Meta Macros": "#define GPIO0 1",
    "Meta Arrays": "NA"
  }
]

TESTPLAN_COLUMNS = [
    "Index","SS / Module","Feature","Test Case Name","Test Description",
    "Speed","Mode","Memory Start Offset","Memory End Offset","Remarks",
    "Test Steps / Procedure","Impacted Registers","Validation / Acceptance Criteria",
    "Code Generation (Required / Not)"
]

METADATA_COLUMNS = [
    "Index","Meta Test Description","Meta Test Steps / Procedure",
    "Meta Impacted Registers","Meta Validation / Acceptance Criteria",
    "Meta Headers","Meta Macros","Meta Arrays"
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFB8CCE4")  # light blue
HEADER_FONT = Font(bold=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def set_column_widths(ws, widths):
    for idx, header in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths[header]


def build_workbook(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Write headers
    for col_idx, header in enumerate(TESTPLAN_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN

    # Rows
    row_idx = 2
    for item in data:
        for col_idx, header in enumerate(TESTPLAN_COLUMNS, start=1):
            val = item.get(header, "")
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = WRAP_ALIGN
        row_idx += 1

    # Freeze top row
    ws.freeze_panes = "A2"

    # Column widths
    widths = {
        "Index": 8,
        "SS / Module": 16,
        "Feature": 12,
        "Test Case Name": 24,
        "Test Description": 60,
        "Speed": 10,
        "Mode": 10,
        "Memory Start Offset": 20,
        "Memory End Offset": 20,
        "Remarks": 16,
        "Test Steps / Procedure": 70,
        "Impacted Registers": 40,
        "Validation / Acceptance Criteria": 50,
        "Code Generation (Required / Not)": 22,
    }
    set_column_widths(ws, widths)

    # MetaData sheet
    ws_meta = wb.create_sheet("MetaData")
    for col_idx, header in enumerate(METADATA_COLUMNS, start=1):
        cell = ws_meta.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN

    row_idx = 2
    for item in data:
        for col_idx, header in enumerate(METADATA_COLUMNS, start=1):
            val = item.get(header, "")
            c = ws_meta.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = WRAP_ALIGN
        row_idx += 1

    ws_meta.freeze_panes = "A2"
    meta_widths = {
        "Index": 8,
        "Meta Test Description": 80,
        "Meta Test Steps / Procedure": 80,
        "Meta Impacted Registers": 70,
        "Meta Validation / Acceptance Criteria": 40,
        "Meta Headers": 20,
        "Meta Macros": 40,
        "Meta Arrays": 20,
    }
    set_column_widths(ws_meta, meta_widths)

    # Hide MetaData veryHidden
    ws_meta.sheet_state = "veryHidden"

    return wb


def main():
    # Compute IST timestamp
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    ts = now_ist.strftime("%Y%m%d_%H%M%S")

    filename = f"{IP_NAME}_TestPlan_{ts}.xlsx"
    out_dir = OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, filename)

    wb = build_workbook(aggregated_json)
    wb.save(out_path)

    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
