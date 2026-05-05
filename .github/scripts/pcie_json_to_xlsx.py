#!/usr/bin/env python3
import argparse
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timezone, timedelta
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MACRO_MAP = OrderedDict({
    "mizar_PCIE1_DBI_USP_CAP_ID_NXT_PTR_REG": "PCIE_CAP_ID_PCIE_NEXT_CAP_PTR_PCIE_CAP_REG",
    "mizar_PCIE1_DBI_USP_DEVICE_CONTROL_DEVICE_STATUS": "DBI_USP_DEVICE_CONTROL_DEVICE_STATUS",
    "mizar_PCIE1_DBI_USP_PL_DEBUG1_OFF": "DBI_USP_PL_DEBUG1_OFF",
    "mizar_PCIE1_SII_PHY_RST_CONTROL": "SII_PHY_RST_CONTROL",
    "mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF": "DBI_DSP_COHERENCY_CONTROL_3_OFF",
    "mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF": "DBI_DSP_COHERENCY_CONTROL_3_OFF",
})

WRAP_COLS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}


def parse_args():
    ap = argparse.ArgumentParser(description="Generate formatted XLSX from JSON")
    ap.add_argument("--input", required=True, help="Path to JSON input file")
    ap.add_argument("--output-dir", required=True, help="Directory to place XLSX")
    ap.add_argument("--ip-name", required=True, help="IP name for filename prefix")
    return ap.parse_args()


def load_json_records(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f, object_pairs_hook=OrderedDict)
    if isinstance(data, dict):
        # Convert TC1, TC2, ... object to array of records
        records = []
        for k in sorted(data.keys(), key=lambda x: (len(x), x)):
            records.append(OrderedDict(data[k]))
    elif isinstance(data, list):
        records = data
    else:
        raise ValueError("JSON root must be object or array")
    if not records:
        raise ValueError("JSON has no records")
    return records


def build_schema(records):
    schema = []
    seen = set()
    for rec in records:
        for key in rec.keys():
            if key not in seen:
                seen.add(key)
                schema.append(key)
    return schema


def autosize_columns(ws):
    maxlen = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            s = "" if val is None else str(val)
            maxlen[idx] = max(maxlen.get(idx, 0), len(s))
    for idx, m in maxlen.items():
        # rough factor to approximate Excel width; cap widths
        width = min(max(10, m + 2), 80)
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_borders(ws):
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def number_lines(text: str) -> str:
    if text is None:
        return ""
    lines = str(text).splitlines()
    out = []
    n = 1
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        # strip existing leading bullets like 1), 1., -, * etc.
        s = re.sub(r"^([0-9]+)[\)\.:\-\s]*", "", s)
        s = re.sub(r"^[\-\*\u2022\u00B7]+\s*", "", s)
        out.append(f"{n}. {s}")
        n += 1
    return "\n".join(out) if out else str(text)


def apply_numbering(ws, col_names):
    header = [c.value for c in ws[1]]
    idxs = [header.index(n) + 1 for n in col_names if n in header]
    for ridx in range(2, ws.max_row + 1):
        for cidx in idxs:
            cell = ws.cell(row=ridx, column=cidx)
            cell.value = number_lines(cell.value)


def macro_replace_in_sheet(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                s = cell.value
                for k, v in MACRO_MAP.items():
                    if k in s:
                        s = s.replace(k, v)
                cell.value = s


def validate_xlsx(path):
    if not zipfile.is_zipfile(path):
        return False
    ok = False
    with zipfile.ZipFile(path, 'r') as z:
        members = set(z.namelist())
        ok = ("[Content_Types].xml" in members and "xl/workbook.xml" in members)
    if not ok:
        return False
    try:
        _ = load_workbook(path)
    except Exception:
        return False
    return True


def main():
    args = parse_args()
    os.makedirs(args.output_dir, exist_ok=True)

    records = load_json_records(args.input)
    schema = build_schema(records)

    # Create workbook with Data sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers
    for c, key in enumerate(schema, start=1):
        ws.cell(row=1, column=c, value=key)

    # Write rows
    for r, rec in enumerate(records, start=2):
        for c, key in enumerate(schema, start=1):
            ws.cell(row=r, column=c, value=rec.get(key, ""))

    # Base formatting
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    autosize_columns(ws)

    # META sheet
    meta = wb.create_sheet("Meta_data_sheet")
    for c, key in enumerate(META_COLS, start=1):
        meta.cell(row=1, column=c, value=key)
    for r, rec in enumerate(records, start=2):
        for c, key in enumerate(META_COLS, start=1):
            meta.cell(row=r, column=c, value=rec.get(key, ""))
    meta.sheet_state = "veryHidden"

    # Rename Data -> TestPlan and transform in-place
    ws.title = "TestPlan"
    # Rebuild content with MAIN_ORDER on the same sheet
    data_rows = []
    for rec in records:
        row_vals = [rec.get(col, "") for col in MAIN_ORDER]
        data_rows.append(row_vals)
    # Clear sheet
    ws.delete_rows(1, ws.max_row)
    # Write new header and rows
    for c, key in enumerate(MAIN_ORDER, start=1):
        ws.cell(row=1, column=c, value=key)
    for r, row_vals in enumerate(data_rows, start=2):
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=r, column=c, value=val)

    # Macro replacement only on visible TestPlan sheet
    macro_replace_in_sheet(ws)

    # Strict formatting for TestPlan
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
    # Wrap text on specified columns
    header = [c.value for c in ws[1]]
    wrap_idxs = {header.index(n) + 1 for n in WRAP_COLS if n in header}
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if c in wrap_idxs:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            else:
                cell.alignment = Alignment(vertical="top", horizontal=("center" if header[c-1] == "Index" else "left"))
    autosize_columns(ws)
    # Approximate row height based on wrapped content lines
    base_height = 15
    for r in range(2, ws.max_row + 1):
        lines = 1
        for c in wrap_idxs:
            val = ws.cell(row=r, column=c).value
            if val is not None:
                lines = max(lines, str(val).count("\n") + 1)
        ws.row_dimensions[r].height = base_height * lines

    add_borders(ws)

    # Numbering inside cells for specified columns
    apply_numbering(ws, [
        "Test Steps / Procedure",
        "Validation / Acceptance Criteria",
    ])

    # Data validation for 'Code Generation (Required / Not)'
    if "Code Generation (Required / Not)" in header:
        code_idx = header.index("Code Generation (Required / Not)") + 1
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
        dv.error = "Select one of: Required, Blank, Not Required"
        dv.errorTitle = "Invalid selection"
        ws.add_data_validation(dv)
        rng = f"{get_column_letter(code_idx)}2:{get_column_letter(code_idx)}{ws.max_row}"
        dv.add(rng)

    # Enforce final sheet visibility
    if "Data" in wb.sheetnames and wb["Data"] != ws:
        wb.remove(wb["Data"])  # should not exist; safety

    # Compute IST timestamp
    ist = timezone(timedelta(hours=5, minutes=30))
    now = datetime.now(ist)
    ts = now.strftime("%Y%m%d_%H%M%S")
    fname = f"{args.ip_name}_TestPlan_{ts}.xlsx"
    out_path = os.path.join(args.output_dir, fname)

    wb.save(out_path)

    # Validate
    ok = validate_xlsx(out_path)
    if not ok:
        print("XLSX validation failed", file=sys.stderr)
        sys.exit(2)

    print(f"FINAL_XLSX_PATH: {out_path}")
    print(f"ROWS: {len(records)}")
    # Columns = length of MAIN_ORDER (visible sheet)
    print(f"COLUMNS: {len(MAIN_ORDER)}")


if __name__ == "__main__":
    main()
