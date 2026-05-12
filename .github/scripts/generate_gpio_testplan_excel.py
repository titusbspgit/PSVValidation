import os, json, re, io, zipfile
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Config
IP_NAME = os.getenv('IP_NAME', 'GPIO')
OUTPUT_DIR = os.getenv('OUTPUT_DIR', 'Test_Output/GPIO/TestPlan')
JSON_PAYLOAD = os.getenv('JSON_PAYLOAD', '').strip()

if not JSON_PAYLOAD:
    raise SystemExit('ERROR: JSON_PAYLOAD is empty')

try:
    data = json.loads(JSON_PAYLOAD)
except Exception as e:
    raise SystemExit(f'ERROR: Invalid JSON input: {e}')

if not isinstance(data, list) or len(data) == 0:
    raise SystemExit('ERROR: JSON input must be a non-empty array of objects')

# Build ordered union of keys preserving first-seen order
keys_order = []
seen = set()
for row in data:
    if not isinstance(row, dict):
        raise SystemExit('ERROR: Each array element must be an object')
    for k in row.keys():
        if k not in seen:
            seen.add(k)
            keys_order.append(k)

# Create workbook and Data sheet
wb = Workbook()
ws = wb.active
ws.title = 'Data'

# Header style
header_font = Font(bold=True)
header_fill = PatternFill('solid', fgColor='8DB4E2')  # blue-ish
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align_text = Alignment(horizontal='left', vertical='top', wrap_text=False)
cell_align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
cell_align_num = Alignment(horizontal='center', vertical='top', wrap_text=False)
thin = Side(style='thin', color='000000')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

# Write header
for c, key in enumerate(keys_order, 1):
    cell = ws.cell(row=1, column=c, value=key)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Helper: normalize numbered lists inside cell (replace bullets or 1) with 1., 2., ...)
def normalize_numbering(text: str) -> str:
    if text is None:
        return ''
    # Split on newlines
    lines = str(text).splitlines()
    items = []
    for line in lines:
        s = line.strip()
        if not s:
            continue
        # Remove leading markers like 1), 1., -, •, *, etc.
        s = re.sub(r'^([0-9]+)[\.)]\s*', '', s)
        s = re.sub(r'^[\-\*•\u2022]+\s*', '', s)
        items.append(s)
    if not items:
        return ''
    return "\n".join(f"{i+1}. {items[i]}" for i in range(len(items)))

# Column indices for special formatting later
col_index_by_name = {}
for idx, k in enumerate(keys_order, 1):
    col_index_by_name[k] = idx

# Write data rows exactly, but pre-normalize specific columns for numbering in main later (will be re-applied)
for r, row in enumerate(data, start=2):
    for c, key in enumerate(keys_order, 1):
        val = row.get(key, '')
        ws.cell(row=r, column=c, value=val)

# Freeze header
ws.freeze_panes = 'A2'

# Create Meta_data_sheet and copy META columns
META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]
meta_ws = wb.create_sheet('Meta_data_sheet')
for c, key in enumerate(META_COLS, 1):
    meta_ws.cell(row=1, column=c, value=key).font = header_font
    for r, row in enumerate(data, start=2):
        meta_ws.cell(row=r, column=c, value=row.get(key, ''))
# Very hidden
meta_ws.sheet_state = 'veryHidden'

# Rename Data to TestPlan and remove META columns in-place, reorder remaining columns
MAIN_ORDER = [
    'Index', 'SS / Module', 'Feature', 'Test Case Name', 'Test Description',
    'Speed', 'Mode', 'Memory Start Offset', 'Memory End Offset', 'Remarks',
    'Test Steps / Procedure', 'Impacted Registers', 'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

# Build a mapping from current header to column index
header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
# Determine columns to keep in the specified order
keep_indices = [header_row.index(h) + 1 for h in MAIN_ORDER if h in header_row]
# Create an ordered snapshot for the TestPlan sheet
rows_snapshot = []
for r in range(1, ws.max_row+1):
    rows_snapshot.append([ws.cell(row=r, column=i).value for i in keep_indices])

# Overwrite current sheet with kept columns only
ws.title = 'TestPlan'
# Clear all cells by recreating sheet content in-place
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        ws.cell(row=r, column=c).value = None

# Write headers in MAIN_ORDER and then data
for c, h in enumerate(MAIN_ORDER, 1):
    ws.cell(row=1, column=c, value=h)
for r, row_vals in enumerate(rows_snapshot[1:], start=2):
    for c, v in enumerate(row_vals, 1):
        ws.cell(row=r, column=c, value=v)

# Apply formatting on TestPlan
max_row = ws.max_row
max_col = ws.max_column
# Header styling
for c in range(1, max_col+1):
    cell = ws.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Wrap text for selected columns
WRAP_COLS = set(['Test Description', 'Remarks', 'Test Steps / Procedure', 'Validation / Acceptance Criteria'])
for c in range(1, max_col+1):
    header = ws.cell(row=1, column=c).value
    for r in range(2, max_row+1):
        cell = ws.cell(row=r, column=c)
        if header in WRAP_COLS:
            # Normalize numbering for specific two columns
            if header in ['Test Steps / Procedure', 'Validation / Acceptance Criteria']:
                cell.value = normalize_numbering(cell.value)
            cell.alignment = cell_align_wrap
        else:
            # Center numeric-like for Index, leave others left-aligned
            if header == 'Index':
                cell.alignment = cell_align_num
            else:
                cell.alignment = cell_align_text
        cell.border = border_all

# Thin borders for header as well
for c in range(1, max_col+1):
    ws.cell(row=1, column=c).border = border_all

# Auto-fit column widths (approximate based on content length)
def compute_width(col_idx):
    max_len = 0
    for r in range(1, max_row+1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            l = 0
        else:
            s = str(v)
            l = max(len(line) for line in s.splitlines()) if s else 0
        if l > max_len:
            max_len = l
    return min(120, max(12, int(max_len * 1.2) + 2))

for c in range(1, max_col+1):
    ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = compute_width(c)

# Data validation only for 'Code Generation (Required / Not)'
try:
    codegen_col = MAIN_ORDER.index('Code Generation (Required / Not)') + 1
    dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(f"{ws.cell(row=2, column=codegen_col).coordinate}:{ws.cell(row=max_row, column=codegen_col).coordinate}")
except ValueError:
    pass  # Column not present; skip

# Safety check: only TestPlan (visible) and Meta_data_sheet (veryHidden)
# Ensure no sheet named 'Data'
for sheet in list(wb.sheetnames):
    if sheet == 'Data':
        del wb[sheet]

# Validate XLSX (zip structure) by saving to a bytes buffer first
buf = io.BytesIO()
wb.save(buf)
buf.seek(0)
with zipfile.ZipFile(buf, 'r') as zf:
    # Basic OOXML entries
    assert '[Content_Types].xml' in zf.namelist()
    assert 'xl/workbook.xml' in zf.namelist()

# Determine IST timestamp and final filename
ist = ZoneInfo('Asia/Kolkata')
now_ist = datetime.now(ist)
ts_date = now_ist.strftime('%Y-%m-%d')
ts_time = now_ist.strftime('%H:%M:%S')
final_stamp = now_ist.strftime('%Y%m%d_%H%M%S')
final_name = f"{IP_NAME}_TestPlan_{final_stamp}.xlsx"

# Ensure output dir exists
os.makedirs(OUTPUT_DIR, exist_ok=True)
final_path = os.path.join(OUTPUT_DIR, final_name)

# Save final workbook to repository working directory
with open(final_path, 'wb') as f:
    f.write(buf.getvalue())

# Commit the Excel back to the repo
os.system('git config user.name "github-actions[bot]"')
os.system('git config user.email "41898282+github-actions[bot]@users.noreply.github.com"')
os.system(f'git add "{final_path}"')
commit_msg = f"Add {IP_NAME} TestPlan Excel autogenerated on IST {ts_date} {ts_time}"
os.system(f'git commit -m "{commit_msg}" || echo "Nothing to commit"')
os.system('git push')

print(f"FINAL_EXCEL_PATH={final_path}")
print(f"FINAL_FILENAME={final_name}")
