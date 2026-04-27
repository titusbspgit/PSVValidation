import json, os, sys
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

BLUE_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
TOPLEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOPCENTER = Alignment(horizontal="center", vertical="top", wrap_text=False)


def json_dumps_if_needed(v):
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def union_keys_preserve_order(objs):
    seen = []
    seen_set = set()
    for obj in objs:
        for k in obj.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def autofit_columns(ws):
    # approximate by max length per column
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = cell.value
                if val is None:
                    l = 0
                else:
                    l = len(str(val))
                if l > max_len:
                    max_len = l
            except Exception:
                pass
        width = min(100, max(10, max_len + 2))
        ws.column_dimensions[col_letter].width = width


def apply_table_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def create_data_sheet(wb, testcases):
    ws = wb.active
    ws.title = "Data"
    # Build union keys preserving first-seen order
    keys = union_keys_preserve_order(testcases)
    # Write header
    ws.append(keys)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = CENTER
    # Rows
    for obj in testcases:
        row = [json_dumps_if_needed(obj.get(k, "")) for k in keys]
        ws.append(row)
    ws.freeze_panes = "A2"
    autofit_columns(ws)
    return ws, keys


def create_meta_sheet(wb, testcases):
    ws = wb.create_sheet("Meta_data_sheet")
    ws.append(META_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = CENTER
    for obj in testcases:
        row = [json_dumps_if_needed(obj.get(k, "")) for k in META_COLUMNS]
        ws.append(row)
    # Very hidden
    ws.sheet_state = "veryHidden"
    return ws


def normalize_main_sheet(wb, data_ws, data_keys):
    # Remove META columns and keep only MAIN in specified order
    data = []
    header_idx = {k: i for i, k in enumerate(data_keys)}
    for r in data_ws.iter_rows(min_row=2, max_row=data_ws.max_row, values_only=True):
        row_obj = {k: r[header_idx[k]] if k in header_idx else "" for k in data_keys}
        data.append(row_obj)

    # Rename to TestPlan
    data_ws.title = "TestPlan"

    # Clear sheet and write MAIN columns only, in order
    for row in data_ws[1:data_ws.max_row]:
        for cell in row:
            cell.value = None
    data_ws.delete_rows(1, data_ws.max_row)

    data_ws.append(MAIN_COLUMNS)
    for cell in data_ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL

    for obj in data:
        row = [json_dumps_if_needed(obj.get(k, "")) for k in MAIN_COLUMNS]
        data_ws.append(row)

    # Wrap text in selected columns
    wrap_cols = {
        "Test Description",
        "Remarks",
        "Test Steps / Procedure",
        "Validation / Acceptance Criteria",
    }
    header_map = {cell.value: idx for idx, cell in enumerate(data_ws[1], start=1)}
    for r in range(2, data_ws.max_row + 1):
        for name in wrap_cols:
            c = header_map.get(name)
            if c:
                data_ws.cell(row=r, column=c).alignment = TOPLEFT_WRAP

    # Align other columns
    for r in range(2, data_ws.max_row + 1):
        for c in range(1, data_ws.max_column + 1):
            if c not in [header_map.get(n) for n in wrap_cols if header_map.get(n)]:
                # Index numeric-ish center, others left
                hdr = data_ws.cell(row=1, column=c).value
                if hdr == "Index":
                    data_ws.cell(row=r, column=c).alignment = TOPCENTER
                else:
                    # Left/top by default
                    if not data_ws.cell(row=r, column=c).alignment:
                        data_ws.cell(row=r, column=c).alignment = TOPLEFT_WRAP

    # Data validation for Code Generation (Required / Not)
    code_col = header_map.get("Code Generation (Required / Not)")
    if code_col:
        col_letter = data_ws.cell(row=1, column=code_col).column_letter
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
        data_ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{data_ws.max_row}")

    # Borders and sizing
    apply_table_borders(data_ws)
    autofit_columns(data_ws)

    return data_ws


def create_overview_and_companions(wb, metadata, testcases):
    # Overview_Summary
    ov = wb.create_sheet("Overview_Summary")
    rows = [
        ("IP Name", metadata.get("ip_name", "")),
        ("Repository", f"{metadata.get('repo_owner','')}/{metadata.get('repo_name','')}"),
        ("Branch", metadata.get("branch", "")),
        ("Subdirectory Root", metadata.get("subdirectory_root", "")),
        ("Generated At (IST)", metadata.get("generated_at_ist", "")),
        ("Timezone", metadata.get("timezone", "")),
    ]
    ov.append(["Field", "Value"]) 
    for k, v in rows:
        ov.append([k, json_dumps_if_needed(v)])
    # Source links
    ov.append(["Source Links", ""]) 
    for link in metadata.get("source_links", []):
        ov.append(["", link])
    autofit_columns(ov)

    # Assumptions
    asm = wb.create_sheet("Assumptions")
    asm.append(["Assumption"]) 
    for a in metadata.get("assumptions", []):
        asm.append([a])
    autofit_columns(asm)

    # Environment_Dependencies
    envs = wb.create_sheet("Environment_Dependencies")
    envs.append(["Dependency"]) 
    for e in metadata.get("environment_dependencies", []):
        envs.append([e])
    autofit_columns(envs)

    # Traceability
    tr = wb.create_sheet("Traceability")
    tr.append(["Index", "Test Case Name", "Feature", "Objective"]) 
    for tc in testcases:
        tr.append([
            tc.get("Index", ""),
            tc.get("Test Case Name", ""),
            tc.get("Feature", ""),
            tc.get("Test Description", ""),
        ])
    autofit_columns(tr)

    # Execution_Matrix
    em = wb.create_sheet("Execution_Matrix")
    matrix = metadata.get("execution_matrix", [])
    if matrix:
        keys = union_keys_preserve_order(matrix)
        em.append(keys)
        for row in matrix:
            em.append([json_dumps_if_needed(row.get(k, "")) for k in keys])
    autofit_columns(em)

    # Risks
    rk = wb.create_sheet("Risks")
    rk.append(["Risk"]) 
    for r in metadata.get("risks", []):
        rk.append([r])
    autofit_columns(rk)


def main():
    input_json_path = os.getenv("INPUT_JSON_PATH", ".github/testplan_artifacts/I2C_TestPlan.json")
    output_dir = os.getenv("OUTPUT_DIR", "Test_Output/I2C/TestPlan")
    ip_name = os.getenv("IP_NAME", "I2C")

    if not os.path.exists(input_json_path):
        print(f"ERROR: JSON artifact not found at {input_json_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Validate JSON structure
    if not isinstance(data, dict) or "testcases" not in data or not isinstance(data["testcases"], list):
        print("ERROR: Invalid JSON structure: expected object with 'testcases' array", file=sys.stderr)
        sys.exit(1)

    metadata = data.get("metadata", {})
    testcases = data["testcases"]

    wb = Workbook()
    data_ws, data_keys = create_data_sheet(wb, testcases)
    create_meta_sheet(wb, testcases)
    normalize_main_sheet(wb, data_ws, data_keys)
    create_overview_and_companions(wb, metadata, testcases)

    # Timestamp in IST for filename
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    stamp = now_ist.strftime("%Y%m%d_%H%M%S")
    os.makedirs(output_dir, exist_ok=True)
    out_name = f"{ip_name}_TestPlan_{stamp}.xlsx"
    out_path = os.path.join(output_dir, out_name)
    wb.save(out_path)
    print(out_path)

if __name__ == "__main__":
    main()
