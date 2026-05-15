#!/usr/bin/env python3
import argparse
import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

META_ORDER_SPEC = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
    "Hidden_Header_Includes",
    "Hidden_Macro_Define",           # spec name (may not exist)
    "Hidden_Macro_Defines",          # present in provided data
    "Hidden_Skip_Array_Definition",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

BLUE_FILL = PatternFill("solid", fgColor="FF4F81BD")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=False)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)

WRAP_COLS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

VALIDATION_COL = "Code Generation (Required / Not)"
VALID_LIST = "Required, Blank, Not Required"


def parse_args():
    p = argparse.ArgumentParser(description="Generate formatted TestPlan Excel from JSON")
    p.add_argument("--json-file", required=True, help="Path to input JSON array file")
    p.add_argument("--output-dir", required=True, help="Repo-relative output directory for Excel")
    p.add_argument("--ip-name", required=True, help="IP name for filename prefix")
    p.add_argument("--tz", default="Asia/Kolkata", help="IANA timezone for timestamp (default: Asia/Kolkata)")
    return p.parse_args()


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("JSON must be a non-empty array of objects")
    for i, row in enumerate(data):
        if not isinstance(row, dict):
            raise ValueError(f"JSON element at index {i} is not an object")
    return data


def union_keys_preserve_order(rows):
    seen = []
    s = set()
    for obj in rows:
        for k in obj.keys():
            if k not in s:
                s.add(k)
                seen.append(k)
    return seen


def best_fit_col_width(values):
    max_len = 0
    for v in values:
        if v is None:
            lns = [""]
        else:
            s = str(v)
            lns = s.split("\n")
        for ln in lns:
            if len(ln) > max_len:
                max_len = len(ln)
    return min(max(10, int(max_len * 1.1) + 2), 120)


def renumber_multiline(value):
    if value is None:
        return ""
    lines = [ln.strip() for ln in str(value).split("\n")]
    lines = [ln for ln in lines if ln]
    if not lines:
        return ""
    return "\n".join(f"{i+1}. {ln.lstrip('0123456789). .')}" for i, ln in enumerate(lines))


def build_base_workbook(rows):
    keys = union_keys_preserve_order(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header
    for j, k in enumerate(keys, start=1):
        cell = ws.cell(row=1, column=j, value=k)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER

    # Rows
    for i, obj in enumerate(rows, start=2):
        for j, k in enumerate(keys, start=1):
            v = obj.get(k, None)
            ws.cell(row=i, column=j, value=v)

    ws.freeze_panes = "A2"

    # Auto-fit cols
    for j, k in enumerate(keys, start=1):
        col_vals = [k] + [r.get(k, None) for r in rows]
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = best_fit_col_width(col_vals)

    # Borders
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in r:
            c.border = THIN_BORDER

    return wb, keys


def create_meta_sheet(wb, rows, all_keys):
    ws = wb.create_sheet("Meta_data_sheet")
    hidden_in_data = [k for k in all_keys if k.startswith("Hidden_")]
    meta_keys = [k for k in META_ORDER_SPEC if k in hidden_in_data]
    for k in hidden_in_data:
        if k not in meta_keys:
            meta_keys.append(k)

    # Header
    for j, k in enumerate(meta_keys, start=1):
        cell = ws.cell(row=1, column=j, value=k)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER

    # Data
    for i, obj in enumerate(rows, start=2):
        for j, k in enumerate(meta_keys, start=1):
            ws.cell(row=i, column=j, value=obj.get(k, None))

    # Styling
    for j in range(1, len(meta_keys) + 1):
        col_letter = ws.cell(row=1, column=j).column_letter
        ws.column_dimensions[col_letter].width = best_fit_col_width([ws.cell(row=r, column=j).value for r in range(1, ws.max_row + 1)])
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in r:
            c.border = THIN_BORDER

    ws.sheet_state = 'veryHidden'

    return meta_keys


def normalize_main_sheet(wb, rows):
    ws = wb["Data"]
    ws.title = "TestPlan"

    normalized_rows = []
    for obj in rows:
        row = {}
        for k in MAIN_ORDER:
            row[k] = obj.get(k, None)
        normalized_rows.append(row)

    # Clear & rewrite
    ws.delete_rows(1, ws.max_row)

    for j, k in enumerate(MAIN_ORDER, start=1):
        cell = ws.cell(row=1, column=j, value=k)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER

    for i, obj in enumerate(normalized_rows, start=2):
        for j, k in enumerate(MAIN_ORDER, start=1):
            v = obj.get(k, None)
            if k in {"Test Steps / Procedure", "Validation / Acceptance Criteria"}:
                v = renumber_multiline(v)
                cell = ws.cell(row=i, column=j, value=v)
                cell.alignment = LEFT_WRAP
            elif k in WRAP_COLS:
                cell = ws.cell(row=i, column=j, value=v)
                cell.alignment = LEFT_WRAP
            else:
                cell = ws.cell(row=i, column=j, value=v)
                if k == "Index":
                    cell.alignment = RIGHT
                else:
                    cell.alignment = LEFT

    # Freeze header row again (since sheet was rewritten)
    ws.freeze_panes = "A2"

    # Column widths and borders
    for j, k in enumerate(MAIN_ORDER, start=1):
        col_vals = [k] + [ws.cell(row=i, column=j).value for i in range(2, ws.max_row + 1)]
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = best_fit_col_width(col_vals)

    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(MAIN_ORDER)):
        for c in r:
            c.border = THIN_BORDER

    # Data validation on single column only
    if VALIDATION_COL in MAIN_ORDER:
        col_idx = MAIN_ORDER.index(VALIDATION_COL) + 1
        dv = DataValidation(type="list", formula1=f'"{VALID_LIST}"', allow_blank=True, showErrorMessage=True)
        dv.error = "Select a value from the list"
        dv.errorTitle = "Invalid Option"
        ws.add_data_validation(dv)
        if ws.max_row >= 2:
            start_row = 2
            end_row = ws.max_row
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

    # Approximate row heights after wrapping
    for i in range(2, ws.max_row + 1):
        max_lines = 1
        for col_name in WRAP_COLS:
            if col_name in MAIN_ORDER:
                j = MAIN_ORDER.index(col_name) + 1
                v = ws.cell(row=i, column=j).value
                if v is None:
                    continue
                lines = str(v).split("\n")
                if len(lines) > max_lines:
                    max_lines = len(lines)
        ws.row_dimensions[i].height = min(15 * max_lines, 300)

    return wb


def validate_xlsx(path):
    with ZipFile(path, 'r') as z:
        names = set(z.namelist())
        if "[Content_Types].xml" not in names:
            return False
        if "xl/workbook.xml" not in names:
            return False
    return True


def main():
    args = parse_args()
    rows = load_json(args.json_file)

    wb, keys = build_base_workbook(rows)
    create_meta_sheet(wb, rows, keys)
    normalize_main_sheet(wb, rows)

    # Safety: ensure no sheet named 'Data' remains
    if 'Data' in wb.sheetnames:
        del wb['Data']

    # Filename with IST timestamp
    tz = ZoneInfo(args.tz)
    now_ist = datetime.now(tz)
    fname = f"{args.ip_name}_TestPlan_{now_ist.strftime('%Y%m%d')}_{now_ist.strftime('%H%M%S')}.xlsx"
    out_dir = args.output_dir
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, fname)
    wb.save(out_path)

    if not validate_xlsx(out_path):
        raise SystemExit("XLSX validation failed")

    print(f"OUTPUT_FILE={out_path}")

if __name__ == "__main__":
    main()
