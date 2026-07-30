#!/usr/bin/env python3
import json
import os
import sys
import argparse
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception as e:
    print(f"ERROR: openpyxl not available: {e}")
    sys.exit(1)

TESTPLAN_COLS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_COLS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

def read_json(json_file: str = None, json_env: str = None):
    data = None
    if json_file:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    elif json_env:
        data = json.loads(json_env)
    else:
        raise ValueError("No JSON input provided")
    if not isinstance(data, list):
        raise ValueError("json_data must be an array of objects")
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Item at index {i} is not an object")
    return data


def style_header(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def auto_width(ws):
    col_widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            s = "" if val is None else str(val)
            length = max([len(line) for line in s.splitlines()]) if s else 0
            if idx not in col_widths:
                col_widths[idx] = 0
            if length > col_widths[idx]:
                col_widths[idx] = length
    for idx, width in col_widths.items():
        adjusted = min(max(width + 2, 12), 100)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = adjusted


def write_sheet(ws, cols, rows):
    ws.append(cols)
    style_header(ws)
    for item in rows:
        row_vals = []
        for key in cols:
            v = item.get(key, "")
            row_vals.append(v)
        ws.append(row_vals)
    # Wrap text for all cells and align top-left
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cols)):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    auto_width(ws)


def generate_workbook(data, output_dir: str, ip_name: str):
    wb = Workbook()
    # Default sheet -> rename to TestPlan
    ws_plan = wb.active
    ws_plan.title = "TestPlan"
    write_sheet(ws_plan, TESTPLAN_COLS, data)

    # MetaData sheet
    ws_meta = wb.create_sheet("MetaData")
    write_sheet(ws_meta, METADATA_COLS, data)

    # Very hide MetaData
    ws_meta.sheet_state = 'veryHidden'

    # IST timestamp
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    ts = now_ist.strftime("%Y%m%d_%H%M%S")

    # Filename rule
    base_name = f"{ip_name}_TestPlan_{ts}.xlsx" if ip_name else f"testplan_{ts}.xlsx"

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / base_name

    wb.save(str(out_path))
    return str(out_path)


def main():
    parser = argparse.ArgumentParser(description="Generate TestPlan Excel from JSON")
    parser.add_argument('--json-file', default=os.environ.get('TESTPLAN_JSON_FILE'))
    parser.add_argument('--output-dir', default=os.environ.get('OUTPUT_DIRECTORY', 'Test_Output'))
    parser.add_argument('--ip-name', default=os.environ.get('IP_NAME', ''))
    args = parser.parse_args()

    # Prefer file, else env TESTPLAN_JSON
    json_env = os.environ.get('TESTPLAN_JSON')
    data = read_json(json_file=args.json_file, json_env=json_env)

    # Normalize rows so all required columns exist to avoid data loss
    def normalize(item):
        norm = {}
        for k in set(TESTPLAN_COLS + METADATA_COLS):
            norm[k] = item.get(k, "")
        return norm
    norm_rows = [normalize(d) for d in data]

    out = generate_workbook(norm_rows, args.output_dir, args.ip_name)
    print(out)

if __name__ == '__main__':
    main()
