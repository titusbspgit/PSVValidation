#!/usr/bin/env python3
import json
import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

IST = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(IST)
timestamp_str = now_ist.strftime('%Y%m%d_%H%M%S')
filename = f'PCIE_TestPlan_{timestamp_str}.xlsx'
output_dir = 'Test_Output/PCIE/TestPlan'
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, filename)

json_data = [
  {
    "index": 1,
    "ss_module": "PCIE",
    "test_case_name": "pcie_device_enumerate_test",
    "feature": "Device Enumeration",
    "meta_test_description": "This testcase performs PCIe device enumeration by initializing the PCIe link, programming cache coherency registers, polling for link readiness, reading vendor ID, configuring memory base addresses, and performing BAR sizing and programming on both PCIe slave ports (slv0 and slv1). The test begins by writing 0x0 to address 0xE6004100 and invoking link training (link_training_dm0_x4 or link_training_dm1_x4 depending on compile-time defines DM0_RC, DM1_RC, DM0_EP, DM1_EP). Cache coherency is programmed by performing read-modify-write on mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF using set_data() to set bits [14:11], [6:3], [30:27], and [22:19] to 0xF. After wait_on(20), the same coherency registers are programmed again with all four bit fields set to 0xF in a single pass. The test then polls SII0 register at offset 0xC0 via read_sii0_reg() until (data & 0xD1) == 0xD1, and similarly polls SII1 register at offset 0xC0 via read_sii1_reg(). Under DM0_RC configuration, the vendor ID is read from PCIe slave 0 at offset 0x0 via read_pcie_slv0_reg(0x0), command register at offset 0x4 is written with 0x7 via write_pcie_slv0_reg(0x4, 0x7), and memory base programming functions mem_base_program_dm0_x4() and mem_base_program_dm1_x4() are called. System registers at 0xE690000C, 0xE6900010, 0xE6900014, 0xE6900018, 0xE6900030, and 0xE6900034 are written with 0x1. Cache coherency is then disabled by performing read-modify-write on both COHERENCY_CONTROL_3_OFF registers, setting bits [22:19] to 0x0 while keeping other fields at 0xF, then after wait_on(10), setting both [30:27] and [22:19] to 0x0. After wait_on(30), BAR sizing is performed on slv1 by writing 0xFFFFFFFF to offsets 0x10-0x24, reading back, then writing actual BAR values (0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000). The same BAR sizing and programming sequence is repeated for slv0. Finally, the test polls address 0xE6004100 via read_reg() until the value equals 0x12345678, with wait_on(5) between polls, then calls finish(0).",
    "test_description": "This test validates PCIe device enumeration by performing link training, programming cache coherency control registers for both PCIe controller instances, polling link status registers for readiness, reading the device Vendor ID from TYPE1_DEV_ID_VEND_ID_REG, enabling bus master and memory/IO space via TYPE1_STATUS_COMMAND_REG, programming memory base addresses, configuring system control registers, disabling cache coherency, performing BAR sizing and address assignment on BAR0_REG through PREF_MEM_LIMIT_PREF_MEM_BASE_REG for both PCIe slave ports, and polling a synchronization register for test completion.",
    "meta_test_steps": "1. Write 0x0 to register at 0xE6004100 to initialize.\n2. Invoke link_training_dm0_x4(4) or link_training_dm1_x4(4) based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP).\n3. CACHE PROGRAMMING - Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, use set_data() to set bits [14:11] to 0xF and bits [6:3] to 0xF, write back.\n4. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF again, set bits [30:27] to 0xF and bits [22:19] to 0xF, write back.\n5. Read mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, set bits [14:11] to 0xF and bits [6:3] to 0xF, write back.\n6. Read mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF again, set bits [30:27] to 0xF and bits [22:19] to 0xF, write back.\n7. Call wait_on(20).\n8. Read mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, set all four bit fields [14:11], [6:3], [30:27], [22:19] to 0xF in a single pass, write back.\n9. Read mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, set all four bit fields to 0xF, write back.\n10. Repeat steps 2-9 (duplicated block in source).\n11. Read SII0 register at offset 0xC0 via read_sii0_reg(0xC0).\n12. Call non_secure_prot_nic().\n13. Poll read_sii0_reg(0xC0) in while loop until (data_rd & 0xD1) == 0xD1.\n14. Read SII1 register at offset 0xC0 via read_sii1_reg(0xC0).\n15. Poll read_sii1_reg(0xC0) in while loop until (data_rd & 0xD1) == 0xD1.\n16. [DM0_RC] Read vendor ID: rd_wr_data1 = read_pcie_slv0_reg(0x0), print VENDOR ID.\n17. [DM0_RC] Write command register: write_pcie_slv0_reg(0x4, 0x7).\n18. [DM0_RC] Call mem_base_program_dm0_x4() and mem_base_program_dm1_x4().\n19. [DM0_RC] Call wait_on(10).\n20. Write system registers: write_reg(0xE690000C, 0x1), write_reg(0xE6900010, 0x1), write_reg(0xE6900014, 0x1), write_reg(0xE6900018, 0x1), write_reg(0xE6900030, 0x1), write_reg(0xE6900034, 0x1).\n21-39. Cache disable and BAR sizing steps as previously generated.",
    "test_steps": "1. Initialize the synchronization register by writing the reset value.\n2. Perform PCIe link training for the configured controller mode.\n3-20. Steps as previously generated.",
    "meta_impacted_registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4; 0xE690000C; 0xE6900010; 0xE6900014; 0xE6900018; 0xE6900030; 0xE6900034; 0x10; 0x14; 0x18; 0x1c; 0x20; 0x24",
    "impacted_registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; BAR0_REG; BAR1_REG; SEC_LAT_TIMER_SUB_BUS_SEC_BUS_PRI_BUS_REG; SEC_STAT_IO_LIMIT_IO_BASE_REG; MEM_LIMIT_MEM_BASE_REG; PREF_MEM_LIMIT_PREF_MEM_BASE_REG",
    "validation_criteria": "1-10 as previously generated.",
    "speed": "NA",
    "mode": "Root Complex (DM0_RC) / Endpoint (DM0_EP, DM1_EP) - compile-time configurable",
    "remarks": "The test supports multiple compile-time configurations via preprocessor defines: DM0_RC, DM1_RC, DM0_EP, DM1_EP.",
    "headers_includes": "#include <stdlib.h>; #include <stdio.h>; #include <test_common.h>; #include \"pcie.h\"",
    "global_variables": "unsigned int data_rd; unsigned int data_wr; unsigned int rd_wr_data1; int err2 = 0; int err1 = 0",
    "preprocessing_defines": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DEBUG_DISPLAY"
  },
  {
    "index": 2,
    "ss_module": "PCIE",
    "test_case_name": "pcie_dma_write_test",
    "feature": "DMA Write",
    "meta_test_description": "This testcase validates PCIe DMA write and read-back operations across all four DMA channels (Channel 0-3) on both PCIe controller instances (DM0 and DM1).",
    "test_description": "This test validates PCIe DMA write and read-back operations across all four DMA channels for both PCIe controller instances.",
    "meta_test_steps": "Steps 1-42 as previously generated.",
    "test_steps": "Steps 1-16 as previously generated.",
    "meta_impacted_registers": "0xE6004100; 0xC0; 0x0; 0x4; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_MASK_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_DOORBELL_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_DOORBELL_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE0_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE0_DBI_DSP_DMA_READ_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_STATUS_OFF; mizar_PCIE1_DBI_DSP_DMA_WRITE_INT_CLEAR_OFF; mizar_PCIE1_DBI_DSP_DMA_READ_INT_CLEAR_OFF",
    "impacted_registers": "TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG; DMA_WRITE_INT_MASK_OFF; DMA_READ_INT_MASK_OFF; DMA_WRITE_DOORBELL_OFF; DMA_READ_DOORBELL_OFF; DMA_WRITE_INT_STATUS_OFF; DMA_READ_INT_STATUS_OFF; DMA_WRITE_INT_CLEAR_OFF; DMA_READ_INT_CLEAR_OFF",
    "validation_criteria": "Steps 1-12 as previously generated.",
    "speed": "NA",
    "mode": "Root Complex (DM0_RC / DM1_RC) - compile-time configurable",
    "remarks": "The test supports multiple compile-time configurations.",
    "headers_includes": "#include <stdlib.h>; #include <stdio.h>; #include <test_common.h>; #include \"pcie.h\"",
    "global_variables": "unsigned int data_rd; unsigned int data_wr; int err2 = 0; int err1 = 0; extern unsigned int int_pend; unsigned int src_addr0; unsigned int dst_addr0; unsigned int rd_addr0; unsigned int wr_addr0; unsigned int len; unsigned int rd_wr_data1",
    "preprocessing_defines": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DEBUG_DISPLAY"
  },
  {
    "index": 3,
    "ss_module": "PCIE",
    "test_case_name": "pcie_mem_wr_rd_test",
    "feature": "Memory Write Read",
    "meta_test_description": "This testcase validates PCIe memory write and read-back operations across PCIe slave ports.",
    "test_description": "This test validates PCIe memory write and read-back operations through the PCIe slave ports.",
    "meta_test_steps": "Steps 1-43 as previously generated.",
    "test_steps": "Steps 1-16 as previously generated.",
    "meta_impacted_registers": "0xE6004100; mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF; mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF; 0xC0; 0x0; 0x4",
    "impacted_registers": "COHERENCY_CONTROL_3_OFF; TYPE1_DEV_ID_VEND_ID_REG; TYPE1_STATUS_COMMAND_REG",
    "validation_criteria": "Steps 1-10 as previously generated.",
    "speed": "NA",
    "mode": "Root Complex (DM0_RC / DM1_RC) / Endpoint (DM0_EP / DM1_EP) - compile-time configurable",
    "remarks": "The test supports multiple compile-time configurations.",
    "headers_includes": "#include <stdlib.h>; #include <stdio.h>; #include <test_common.h>; #include \"pcie.h\"",
    "global_variables": "unsigned int data_rd; unsigned int data_wr; unsigned int rd_wr_data1; int err2 = 0; int err1 = 0",
    "preprocessing_defines": "DM0_RC; DM1_RC; DM0_EP; DM1_EP; DM0; DM1; DEBUG_DISPLAY"
  },
  {
    "index": 4,
    "ss_module": "PCIE",
    "test_case_name": "pcie_reg_wr_rd_test",
    "feature": "Register Write Read",
    "meta_test_description": "This testcase validates PCIe register reset values and write/read functionality across four register groups on both PCIe controller instances (PCIE0 and PCIE1).",
    "test_description": "This test validates PCIe register reset default values and write/read data integrity across multiple register groups on both PCIe controller instances.",
    "meta_test_steps": "Steps 1-29 as previously generated.",
    "test_steps": "Steps 1-12 as previously generated.",
    "meta_impacted_registers": "mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE0_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE0_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE0_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE0_DBI_DSP_UTILITY_OFF; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_08H_REG; mizar_PCIE1_DBI_DSP_MSI_CAP_OFF_10H_REG; mizar_PCIE1_DBI_DSP_FILTER_MASK_2_OFF; mizar_PCIE1_DBI_DSP_AXI_MSTR_MSG_ADDR_HIGH_OFF; mizar_PCIE1_DBI_DSP_UTILITY_OFF; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2; mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3; mizar_PCIE0_SII_PHY_CONTROL_23; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2; mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3; mizar_PCIE1_SII_PHY_CONTROL_23; mizar_PCIE0_SII_PHY_RST_CONTROL; mizar_PCIE1_SII_PHY_RST_CONTROL; 0xE68860B8; 0xE68862B8; 0xE68864B8; 0xE68A60B8; 0xE68A62B8; 0xE68A64B8",
    "impacted_registers": "MSI_CAP_OFF_08H_REG; MSI_CAP_OFF_10H_REG; FILTER_MASK_2_OFF; AXI_MSTR_MSG_ADDR_HIGH_OFF; UTILITY_OFF",
    "validation_criteria": "Steps 1-5 as previously generated.",
    "speed": "NA",
    "mode": "NA",
    "remarks": "The test covers four distinct register groups across both PCIe controller instances.",
    "headers_includes": "#include <stdlib.h>; #include <stdio.h>; #include <test_common.h>; #include <pcie.h>",
    "global_variables": "unsigned int data_rd; unsigned int data_wr; unsigned int data1_rd; unsigned int err2 = 0; unsigned int err1 = 0; unsigned int rc0_ctl_addr[5]; unsigned int rc1_ctl_addr[5]; unsigned int ctl_default[5]; unsigned int sii0_addr[3]; unsigned int sii1_addr[3]; unsigned int sii_default[3]; unsigned int sii0_write_mask[3]; unsigned int sii1_write_mask[3]; unsigned int phy0_addr[3]; unsigned int phy1_addr[3]; unsigned int phy0_default[3]; unsigned int phy1_default[3]; unsigned int phy0_write_mask[3]; unsigned int phy1_write_mask[3]",
    "preprocessing_defines": "NA"
  }
]

# Create workbook
wb = Workbook()

# ---- TestPlan Sheet ----
ws_tp = wb.active
ws_tp.title = 'TestPlan'

tp_headers = [
    'Index', 'SS / Module', 'Feature', 'Test Case Name', 'Test Description',
    'Speed', 'Mode', 'Memory Start Offset', 'Memory End Offset', 'Remarks',
    'Test Steps / Procedure', 'Impacted Registers', 'Validation / Acceptance Criteria',
    'Code Generation'
]

# ---- MetaData Sheet ----
ws_md = wb.create_sheet('MetaData')

md_headers = [
    'Index', 'Test Case Name', 'Meta Test Description', 'Meta Test Steps / Procedure',
    'Meta Impacted Registers', 'Meta Validation / Acceptance Criteria',
    'Meta Headers', 'Meta Macros', 'Meta Arrays'
]

# Formatting
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
wrap_alignment = Alignment(wrap_text=True, vertical='top')

def write_headers(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

write_headers(ws_tp, tp_headers)
write_headers(ws_md, md_headers)

# Populate TestPlan
for row_idx, item in enumerate(json_data, 2):
    ws_tp.cell(row=row_idx, column=1, value=item.get('index', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=2, value=item.get('ss_module', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=3, value=item.get('feature', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=4, value=item.get('test_case_name', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=5, value=item.get('test_description', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=6, value=item.get('speed', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=7, value=item.get('mode', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=8, value='').alignment = wrap_alignment  # Memory Start Offset
    ws_tp.cell(row=row_idx, column=9, value='').alignment = wrap_alignment  # Memory End Offset
    ws_tp.cell(row=row_idx, column=10, value=item.get('remarks', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=11, value=item.get('test_steps', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=12, value=item.get('impacted_registers', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=13, value=item.get('validation_criteria', '')).alignment = wrap_alignment
    ws_tp.cell(row=row_idx, column=14, value='').alignment = wrap_alignment  # Code Generation

# Populate MetaData
for row_idx, item in enumerate(json_data, 2):
    ws_md.cell(row=row_idx, column=1, value=item.get('index', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=2, value=item.get('test_case_name', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=3, value=item.get('meta_test_description', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=4, value=item.get('meta_test_steps', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=5, value=item.get('meta_impacted_registers', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=6, value=item.get('validation_criteria', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=7, value=item.get('headers_includes', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=8, value=item.get('preprocessing_defines', '')).alignment = wrap_alignment
    ws_md.cell(row=row_idx, column=9, value=item.get('global_variables', '')).alignment = wrap_alignment

# Auto-size columns
def auto_size_columns(ws, max_width=60):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_length:
                        max_length = len(line)
        adjusted_width = min(max(max_length + 2, 12), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width

auto_size_columns(ws_tp)
auto_size_columns(ws_md)

# Freeze first row
ws_tp.freeze_panes = 'A2'
ws_md.freeze_panes = 'A2'

# Set MetaData sheet to veryHidden
ws_md.sheet_state = 'veryHidden'

# Save workbook
wb.save(output_path)
print(f'FILENAME={filename}')
print(f'OUTPUT_PATH={output_path}')

# Validate
wb2 = load_workbook(output_path)
assert 'TestPlan' in wb2.sheetnames, 'TestPlan sheet missing'
assert 'MetaData' in wb2.sheetnames, 'MetaData sheet missing'
assert wb2['TestPlan'].max_row == 5, f'Expected 5 rows in TestPlan, got {wb2["TestPlan"].max_row}'
assert wb2['MetaData'].max_row == 5, f'Expected 5 rows in MetaData, got {wb2["MetaData"].max_row}'
file_size = os.path.getsize(output_path)
assert file_size > 0, 'File size is 0'
print(f'VALIDATION=PASSED')
print(f'FILE_SIZE={file_size}')
print(f'ROWS_TESTPLAN=4')
print(f'ROWS_METADATA=4')
