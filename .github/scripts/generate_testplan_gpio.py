import os
import json
import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def main():
    src = os.environ.get('FINAL_JSON', '[]')
    data = json.loads(src)
    if not isinstance(data, list) or not data:
        raise SystemExit('No data provided in FINAL_JSON')
    entry = data[0]

    # Preserve key order by specifying explicit columns (flattened)
    columns = [
        'stage',
        'repo',
        'branch',
        'output_directory',
        'file_name',
        'ist_timestamp_token',
        'commit.sha',
        'commit.message',
        'commit.file_path',
        'commit.github_url',
        'commit.raw_url',
        'verification.ref',
        'verification.object_type',
        'verification.object_sha',
    ]

    def get_nested(d, path):
        cur = d
        for p in path.split('.'):
            if isinstance(cur, dict) and p in cur:
                cur = cur[p]
            else:
                return ''
        return '' if cur is None else str(cur)

    headers = columns
    row = [get_nested(entry, c) if '.' in c else ('' if entry.get(c) is None else str(entry.get(c))) for c in columns]

    # Create workbook
    wb = Workbook()

    # TestPlan sheet
    ws = wb.active
    ws.title = 'TestPlan'

    # Headers formatting
    header_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    header_font = Font(bold=True)
    wrap_align = Alignment(wrap_text=True, vertical='top')

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align

    for j, val in enumerate(row, 1):
        cell = ws.cell(row=2, column=j, value=val)
        cell.alignment = wrap_align

    ws.freeze_panes = 'A2'

    # Reasonable column widths
    for j in range(1, len(headers) + 1):
        hdr = headers[j - 1]
        width = max(14, min(60, int(len(hdr) * 1.2)))
        ws.column_dimensions[get_column_letter(j)].width = width

    # MetaData sheet (very hidden)
    md = wb.create_sheet('MetaData')
    md.sheet_state = 'veryHidden'

    # Compute IST timestamp at runtime
    ist_now = datetime.datetime.now(ZoneInfo('Asia/Kolkata'))
    ts_token = ist_now.strftime('%Y%m%d_%H%M%S')

    # Meta rows
    meta_rows = [
        ('source_json', src),
        ('repo', entry.get('repo', '')),
        ('branch', entry.get('branch', '')),
        ('output_directory', entry.get('output_directory', '')),
        ('ip_name', os.environ.get('IP_NAME', '')),
        ('ist_timestamp_token', ts_token),
        ('generator', 'Ag_Excel_Generator Agent'),
    ]

    md.cell(row=1, column=1, value='key').font = header_font
    md.cell(row=1, column=2, value='value').font = header_font

    for i, (k, v) in enumerate(meta_rows, start=2):
        md.cell(row=i, column=1, value=k)
        md.cell(row=i, column=2, value=v)

    # Output path
    output_dir = os.environ.get('OUTPUT_DIR', 'Test_Output/GPIO/TestPlan/').strip('/\\')
    fname = f"{os.environ.get('IP_NAME', 'IP')}_TestPlan_{ts_token}.xlsx"
    out_path = os.path.join(output_dir, fname)
    os.makedirs(output_dir, exist_ok=True)

    # Save workbook (real .xlsx)
    wb.save(out_path)

    print(out_path)


if __name__ == '__main__':
    main()
