#!/usr/bin/env python3
import json
import os
import argparse
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception as e:
    raise SystemExit(f"Missing dependency openpyxl: {e}. Install with 'pip install openpyxl'.")

TESTPLAN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_COLUMNS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

BLUE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def parse_args():
    p = argparse.ArgumentParser(description="Generate TestPlan Excel from JSON (with MetaData veryHidden)")
    p.add_argument("--json", required=True, help="Path to input JSON array file")
    p.add_argument("--owner", required=True)
    p.add_argument("--repo", required=True)
    p.add_argument("--branch", required=True)
    p.add_argument("--ip", required=True, help="IP_NAME, e.g., USB")
    p.add_argument("--source_subdir", required=True, help="Source subdirectory, e.g., TestRepo/usb")
    p.add_argument("--folder_count", required=True, type=int)
    p.add_argument("--output_dir", required=True, help="Output directory for Excel (relative to repo root)")
    p.add_argument("--timestamp", required=True, help="IST timestamp in format YYYYMMDD_HHMMSS for filename")
    return p.parse_args()


def ensure_list_of_objects(data: Any) -> List[Dict[str, Any]]:
    if not isinstance(data, list):
        raise ValueError("json_data must be a JSON array")
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Element at index {i} is not an object")
    return data  # type: ignore


def ist_now_strings(ts_override: str = None):
    ist = timezone(timedelta(hours=5, minutes=30))
    if ts_override:
        # trust format YYYYMMDD_HHMMSS
        dt = datetime.strptime(ts_override, "%YMMDD_%H%M%S") if False else None
        # We cannot parse reliably because format may be wrong if passed wrong.
        # Instead, simply prepare human string from override by slicing
        y = ts_override[0:4]
        m = ts_override[4:6]
        d = ts_override[6:8]
        H = ts_override[9:11]
        M = ts_override[11:13]
        S = ts_override[13:15]
        human = f"{y}-{m}-{d} {H}:{M}:{S}"
        return ts_override, human
    now = datetime.now(ist)
    compact = now.strftime("%Y%m%d_%H%M%S")
    human = now.strftime("%Y-%m-%d %H:%M:%S")
    return compact, human


def set_header_style(ws, headers: List[str]):
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = WRAP_ALIGN
    ws.freeze_panes = "A2"


def auto_widths(ws):
    # Basic auto width: max length (capped)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        # heuristic widths
        width = min(max(12, max_len + 2), 80)
        ws.column_dimensions[col_letter].width = width


def apply_wrap_all(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGN


def build_workbook(rows: List[Dict[str, Any]], global_meta: Dict[str, Any], ip: str):
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Sheet 1: TestPlan (VISIBLE)
    ws1 = wb.create_sheet("TestPlan")
    set_header_style(ws1, TESTPLAN_COLUMNS)
    for obj in rows:
        ws1.append([
            obj.get("Index", ""),
            obj.get("SS / Module", ""),
            obj.get("Feature", ""),
            obj.get("Test Case Name", ""),
            obj.get("Test Description", ""),
            obj.get("Speed", ""),
            obj.get("Mode", ""),
            obj.get("Memory Start Offset", ""),
            obj.get("Memory End Offset", ""),
            obj.get("Remarks", ""),
            obj.get("Test Steps / Procedure", ""),
            obj.get("Impacted Registers", ""),
            obj.get("Validation / Acceptance Criteria", ""),
            obj.get("Code Generation (Required / Not)", ""),
        ])
    apply_wrap_all(ws1)
    auto_widths(ws1)

    # Sheet 2: MetaData (VERY HIDDEN)
    ws2 = wb.create_sheet("MetaData")

    # Global meta block at the top
    ws2.append(["Key", "Value"])
    h1 = ws2.cell(row=1, column=1)
    h2 = ws2.cell(row=1, column=2)
    h1.fill = BLUE_FILL; h2.fill = BLUE_FILL
    h1.font = WHITE_FONT; h2.font = WHITE_FONT
    h1.alignment = WRAP_ALIGN; h2.alignment = WRAP_ALIGN

    for k in ["owner", "repo", "branch", "IP_NAME", "source_sub_directory", "folder_count", "generated_timestamp"]:
        ws2.append([k, str(global_meta.get(k, ""))])

    ws2.append([])  # blank line

    # Per-test-case meta table
    start_row = ws2.max_row + 1
    for col in range(1, len(METADATA_COLUMNS) + 1):
        cell = ws2.cell(row=start_row, column=col, value=METADATA_COLUMNS[col-1])
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = WRAP_ALIGN

    for obj in rows:
        ws2.append([
            obj.get("Index", ""),
            obj.get("Test Case Name", ""),
            obj.get("Meta Test Description", ""),
            obj.get("Meta Test Steps / Procedure", ""),
            obj.get("Meta Impacted Registers", ""),
            obj.get("Meta Validation / Acceptance Criteria", ""),
            obj.get("Meta Headers", ""),
            obj.get("Meta Macros", ""),
            obj.get("Meta Arrays", ""),
        ])

    # Freeze first row per requirement
    ws2.freeze_panes = "A2"

    apply_wrap_all(ws2)
    auto_widths(ws2)

    # Very hidden sheet state
    ws2.sheet_state = 'veryHidden'

    return wb


def main():
    args = parse_args()

    # Prepare IST timestamps (use provided override for name; also compute human)
    compact_ts = args.timestamp  # expected format YYYYMMDD_HHMMSS
    # compute human-readable from provided compact_ts
    human_ts = f"{compact_ts[0:4]}-{compact_ts[4:6]}-{compact_ts[6:8]} {compact_ts[9:11]}:{compact_ts[11:13]}:{compact_ts[13:15]}"

    # Load JSON
    with open(args.json, "r", encoding="utf-8") as f:
        data = json.load(f)
    rows = ensure_list_of_objects(data)

    # Build global meta
    global_meta = {
        "owner": args.owner,
        "repo": args.repo,
        "branch": args.branch,
        "IP_NAME": args.ip,
        "source_sub_directory": args.source_subdir,
        "folder_count": args.folder_count,
        "generated_timestamp": human_ts,
    }

    wb = build_workbook(rows, global_meta, args.ip)

    # Ensure output directory exists
    out_dir = args.output_dir
    if not os.path.isabs(out_dir):
        out_dir = os.path.join(os.getcwd(), out_dir)
    os.makedirs(out_dir, exist_ok=True)

    filename = f"{args.ip}_TestPlan_{compact_ts}.xlsx"
    out_path = os.path.join(out_dir, filename)

    wb.save(out_path)

    # Print the path for CI to pick
    print(out_path)


if __name__ == "__main__":
    main()
