#!/usr/bin/env python3
import os
import json
import requests
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

TESTPLAN_HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_HEADERS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]


def ist_timestamp():
    # IST = UTC + 05:30
    ts = datetime.utcnow() + timedelta(hours=5, minutes=30)
    return ts.strftime("%Y%m%d_%H%M%S")


def fetch_json(json_url: str):
    if not json_url:
        raise ValueError("JSON_URL is not set")
    r = requests.get(json_url, timeout=60)
    r.raise_for_status()
    data = r.json()
    if not isinstance(data, list):
        raise ValueError("json_data must be an array of objects")
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"json_data element at index {i} is not an object")
    return data


def build_sheet(ws, headers, rows):
    # headers
    ws.append(headers)
    # bold headers
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold
    # freeze header row
    ws.freeze_panes = "A2"
    # data rows
    for row in rows:
        ws.append(row)


def main():
    json_url = os.environ.get("JSON_URL", "").strip()
    output_dir = os.environ.get("OUTPUT_DIR", "Test_Output/GPIO/TestPlan").strip()

    data = fetch_json(json_url)

    # Prepare rows maintaining order exactly as in input
    testplan_rows = []
    metadata_rows = []
    for obj in data:
        # Build TestPlan row (ensure missing keys map to empty string)
        testplan_rows.append([
            obj.get("Index", ""),
            obj.get("SS / Module", ""),
            obj.get("Feature", ""),
            obj.get("Test Case Name", ""),
            obj.get("Test Description", ""),
            obj.get("Speed", ""),
            obj.get("Mode", ""),
            obj.get("Memory Start Offset", ""),
            obj.get("Memory End Offset", ""),
            obj.get("Remarks", ""),
            obj.get("Test Steps / Procedure", ""),
            obj.get("Impacted Registers", ""),
            obj.get("Validation / Acceptance Criteria", ""),
            obj.get("Code Generation (Required / Not)", obj.get("Code Generation (Required / Not)", obj.get("Code Generation (Required / Not)", ""))),
        ])
        # Build MetaData row
        metadata_rows.append([
            obj.get("Index", ""),
            obj.get("Test Case Name", ""),
            obj.get("Meta Test Description", ""),
            obj.get("Meta Test Steps / Procedure", ""),
            obj.get("Meta Impacted Registers", ""),
            obj.get("Meta Validation / Acceptance Criteria", ""),
            obj.get("Meta Headers", ""),
            obj.get("Meta Macros", ""),
            obj.get("Meta Arrays", ""),
        ])

    wb = Workbook()
    # Create sheets in order
    ws_plan = wb.active
    ws_plan.title = "TestPlan"
    ws_meta = wb.create_sheet(title="MetaData")

    build_sheet(ws_plan, TESTPLAN_HEADERS, testplan_rows)
    build_sheet(ws_meta, METADATA_HEADERS, metadata_rows)

    # Set MetaData to VeryHidden
    ws_meta.sheet_state = "veryHidden"

    # Ensure output directory exists
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # IST timestamped filename
    fname = f"testplan_{ist_timestamp()}.xlsx"
    out_path = out_dir / fname

    wb.save(out_path.as_posix())
    print(f"Saved Excel: {out_path}")


if __name__ == "__main__":
    main()
