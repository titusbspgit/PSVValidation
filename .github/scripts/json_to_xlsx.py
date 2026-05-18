#!/usr/bin/env python3
import argparse
import json
import sys
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path

def load_rows(data):
    # Accept:
    # - list of objects
    # - single object
    # - single-key dict whose value is a list of objects
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        if len(data) == 1:
            only_val = next(iter(data.values()))
            if isinstance(only_val, list):
                rows = only_val
            else:
                rows = [data]
        else:
            rows = [data]
    else:
        raise ValueError("Unsupported JSON structure: top-level must be object or array")

    if not rows:
        raise ValueError("Empty JSON: no rows found")

    if not all(isinstance(r, dict) for r in rows):
        raise ValueError("Unsupported JSON structure: rows must be objects")

    return rows

def determine_headers(rows):
    headers = []
    seen = set()
    for r in rows:
        # preserve first-seen key order per row
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                headers.append(k)
    return headers

def to_cell_value(v):
    # Preserve original JSON values; leave None as blank
    # For non-scalar JSON (dict/list), serialize deterministically
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    # Deterministic JSON string for nested types
    return json.dumps(v, ensure_ascii=False, sort_keys=False)

def write_xlsx(rows, headers, out_path, sheet_name="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header (bold)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for r in rows:
        row_vals = [to_cell_value(r.get(h, "")) for h in headers]
        ws.append(row_vals)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-fit column widths (approximate)
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                length = 0
            else:
                length = len(str(val))
            if length > max_len:
                max_len = length
        # Add small padding; cap width
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def main():
    ap = argparse.ArgumentParser(description="Convert JSON to Excel (.xlsx) with deterministic schema and formatting.")
    ap.add_argument("--input", required=True, help="Path to JSON input file")
    ap.add_argument("--output", required=True, help="Output .xlsx file path")
    ap.add_argument("--sheet-name", default="Data", help="Worksheet name (default: Data)")
    args = ap.parse_args()

    try:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f, object_pairs_hook=OrderedDict)
    except Exception as e:
        print(f"ERROR: Failed to read/parse JSON: {e}", file=sys.stderr)
        sys.exit(2)

    try:
        rows = load_rows(data)
        headers = determine_headers(rows)
        write_xlsx(rows, headers, args.output, sheet_name=args.sheet_name)
    except Exception as e:
        print(f"ERROR: Conversion failed: {e}", file=sys.stderr)
        sys.exit(3)

if __name__ == "__main__":
    main()
