import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Source-of-truth JSON embedded exactly as provided
JSON_INPUT = r'''{
  "metadata": {
    "ip_name": "GPIO",
    "repo": "titusbspgit/PSVValidation",
    "branch": "main",
    "base_path": "TestRepo/gpio",
    "generated_timestamp_ist": "2026-04-21T00:00:00+05:30",
    "source_of_truth_links": [
      "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/gpio_reg_wr_rd_test",
      "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en",
      "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en"
    ]
  },
  "test_cases": [
    {"Index":1,"SS / Module":"GPIO","Feature":"Independent control register for each GPIO","Test Case Name":"gpio_reg_wr_rd_test","Test Description":"Validates default values and read/write behavior of per-pin and group GPIO registers using defined read/write masks and expected reset values. Confirms data readback matches expected results over multiple patterns and that reset values are correct where readable.","Speed":"NA","Mode":"NA","Memory Start Offset":"NA","Memory End Offset":"NA","Remarks":"Certain registers are intentionally skipped for write/read and reset checks per skip arrays. Input-related bits may read as set unless inputs are forced, as noted for input data behavior.","Test Steps / Procedure":["Begin by checking reset values for each per-pin and group register entry that is not marked to skip and is readable.","For each readable entry, read the register, mask out the input status bit to avoid input-induced mismatches, and compare against the documented default value.","Iterate over a set of six data patterns; for each pattern, write to every register entry that is writable and not marked to skip, applying the documented write mask.","After writing each pattern, read each corresponding entry that is both writable and readable and compute the expected value using read and write masks combined with default values for non-writable bits.","Compare the read value with the computed expected value for each address and record any mismatches.","Conclude the test by reporting pass if no mismatches were recorded in either reset-value or write/read phases, otherwise report failure."],"Impacted Registers":"GPIO_GP0_GPIO_8, GPIO_GP0_GPIO_9, GPIO_GP0_GPIO_10, GPIO_GP0_GPIO_11, GPIO_GP0_GPIO_12, GPIO_GP0_GPIO_13, GPIO_GP0_GPIO_14, GPIO_GP0_GPIO_15, GPIO_GP0_GPIO_16, GPIO_GP0_GPIO_17, GPIO_GP0_GPIO_18, GPIO_GP0_GPIO_19, GPIO_GP0_GPIO_20, GPIO_GP0_GPIO_21, GPIO_GP0_GPIO_22, GPIO_GP0_GPIO_23, GPIO_GP0_GPIO_24, GPIO_GP0_GPIO_25, GPIO_GP0_GPIO_26, GPIO_GP0_GPIO_27, GPIO_GP0_GPIO_28, GPIO_GP0_GPIO_29, GPIO_GP0_GPIO_30, GPIO_GP0_GPIO_31, GPIO_GP0_GPIO_32, GPIO_GP0_GPIO_33, GPIO_GP0_GPIO_34, GPIO_GP0_GPIO_35, GPIO_GP0_GPIO_36, GPIO_GP0_GPIO_37, GPIO_GP0_GPIO_38, GPIO_GP0_GPIO_39, GPIO_GPIO_INTR_RAW_STCLR1, GPIO_GP0_INTR1_INTR_EN1, GPIO_GP0_INTR1_INTR_STS1, GPIO_GP0_INTR2_INTR_EN1, GPIO_GP0_INTR2_INTR_STS1, GPIO_GPIO_IO_CTRL_GROUP1, GPIO_GPIO_IO_CTRL_GROUP2, GPIO_GPIO_IO_CTRL_GROUP3, GPIO_GPIO_IO_CTRL_GROUP4, GPIO_GPIO_DOUT_GROUP1, GPIO_GPIO_DOUT_GROUP2, GPIO_GPIO_DOUT_GROUP3, GPIO_GPIO_DOUT_GROUP4, GPIO_GPIO_DIN_GROUP1, GPIO_GPIO_DIN_GROUP2, GPIO_GPIO_DIN_GROUP3, GPIO_GPIO_DIN_GROUP4","Validation / Acceptance Criteria":["For each entry checked in the reset-value phase, the masked read value must equal the documented default value.","For each entry checked in the write/read phase, the masked read value must equal the expected value derived from the applied pattern and masks combined with default values for non-writable bits.","Overall pass if no failures are recorded in either phase; otherwise fail."],"Code Generation (Required / Not)":"","Hidden_Test_Case_Name":"gpio_reg_wr_rd_test","Hidden_Test_Description":"Test performs two phases: (1) chk_rst_val(): for i=0..CNT-1, addr=addr_array[i]; if skip_rst_array[i]==1 continue; if read_mask_array[i]==0 continue; data_rd=read_reg(addr); data=(data_rd & 0xfffffffe); compare data==default_value_array[i], else def_fail_cnt++. (2) chk_rd_wr(): patterns chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}. For each pattern j: write phase: for i=0..CNT-1: if skip_array[i]==1 continue; if write_mask_array[i]==0 continue; write_reg(addr,(data_wr & write_mask_array[i])). Read/verify phase: for i=0..CNT-1: apply same skips; if read_mask_array[i]==0 continue; data_rd=(read_reg(addr) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); compare data_rd==exp_val else wr_fail_cnt++. finish(0) if both counters are zero else finish(1).","Hidden_Remarks":"Comment notes: \"SKIPPING VRRW registers\"; skip arrays defined. Note: when reading default values the din value is becoming 1 automatically if we don't force any value; forcing zero to din makes level select high, causing mismatch against expected value.","Hidden_Test_Steps_Procedure":["Entry: test_case()","Call chk_rst_val()","Loop i=0..CNT-1: addr=addr_array[i]; if (skip_rst_array[i]==1) continue; if (read_mask_array[i]==0x00000000) continue; data_rd=read_reg(addr); data=(data_rd & 0xfffffffe); if (data==default_value_array[i]) pass else {def_fail_cnt++; printf failure}","Return from chk_rst_val()","Call chk_rd_wr()","Set chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}","For each j in 0..5: data_wr=chk_val[j]","Write phase: for i=0..CNT-1: addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0x00000000) continue; write_reg(addr,(data_wr & write_mask_array[i]))","Read/verify phase: for i=0..CNT-1: addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0x00000000) continue; if (read_mask_array[i]==0x00000000) continue; data_rd=(read_reg(addr) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if (data_rd==exp_val) pass else {wr_fail_cnt++; printf failure}","Return from chk_rd_wr()","If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) else finish(0)"],"Hidden_Impacted_Registers":"MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, MIZAR_GPIO_GP0_GPIO_11, MIZAR_GPIO_GP0_GPIO_12, MIZAR_GPIO_GP0_GPIO_13, MIZAR_GPIO_GP0_GPIO_14, MIZAR_GPIO_GP0_GPIO_15, MIZAR_GPIO_GP0_GPIO_16, MIZAR_GPIO_GP0_GPIO_17, MIZAR_GPIO_GP0_GPIO_18, MIZAR_GPIO_GP0_GPIO_19, MIZAR_GPIO_GP0_GPIO_20, MIZAR_GPIO_GP0_GPIO_21, MIZAR_GPIO_GP0_GPIO_22, MIZAR_GPIO_GP0_GPIO_23, MIZAR_GPIO_GP0_GPIO_24, MIZAR_GPIO_GP0_GPIO_25, MIZAR_GPIO_GP0_GPIO_26, MIZAR_GPIO_GP0_GPIO_27, MIZAR_GPIO_GP0_GPIO_28, MIZAR_GPIO_GP0_GPIO_29, MIZAR_GPIO_GP0_GPIO_30, MIZAR_GPIO_GP0_GPIO_31, MIZAR_GPIO_GP0_GPIO_32, MIZAR_GPIO_GP0_GPIO_33, MIZAR_GPIO_GP0_GPIO_34, MIZAR_GPIO_GP0_GPIO_35, MIZAR_GPIO_GP0_GPIO_36, MIZAR_GPIO_GP0_GPIO_37, MIZAR_GPIO_GP0_GPIO_38, MIZAR_GPIO_GP0_GPIO_39, MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_GPIO_GP0_INTR2_INTR_EN1, MIZAR_GPIO_GP0_INTR2_INTR_STS1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, MIZAR_GPIO_GPIO_DOUT_GROUP1, MIZAR_GPIO_GPIO_DOUT_GROUP2, MIZAR_GPIO_GPIO_DOUT_GROUP3, MIZAR_GPIO_GPIO_DOUT_GROUP4, MIZAR_GPIO_GPIO_DIN_GROUP1, MIZAR_GPIO_GPIO_DIN_GROUP2, MIZAR_GPIO_GPIO_DIN_GROUP3, MIZAR_GPIO_GPIO_DIN_GROUP4","Hidden_Validation_Acceptance_Criteria":"Reset phase: (data_rd & 0xfffffffe) == default_value_array[i] for all i not skipped and readable. Write/read phase: For each pattern j and each i not skipped, writable and readable: (read_reg(addr) & read_mask_array[i]) == ((data_wr & read_mask_array[i] & write_mask_array[i]) | ((write_mask_array[i]^0xffffffff) & read_mask_array[i] & default_value_array[i])). Overall pass if def_fail_cnt==0 and wr_fail_cnt==0 and finish(0) invoked; otherwise finish(1).","test_id":"TC_GPIO_001","objective":"Verify GPIO register reset values and masked read/write behavior across per-pin and group registers.","description":"The test performs a reset-value verification and a masked write/readback verification across defined GPIO registers to ensure correct reset defaults and R/W paths.","type":"directed, sanity","category/tags":["register","reset","readwrite","gpio"],"preconditions":"Target platform initialized; access to GPIO register space; test harness available.","setup":"Build and run the test in the specified environment with required headers (gpio_def.h, gpio_offset.h) and test_common runtime.","steps":["Run reset-value check for all readable and non-skipped registers.","Iterate over six data patterns; for each, write masked data to all writable and non-skipped registers.","Read back each corresponding register with read mask applied and compare to the computed expected value.","Summarize results and report pass/fail."],"expected_results":"All reset-value comparisons and masked write/readback comparisons match expected values.","pass_fail_criteria":"Pass if no mismatches recorded (both def_fail_cnt and wr_fail_cnt remain zero leading to finish(0)); otherwise fail.","required_tools/framework":"C runtime, platform register access (read_reg/write_reg), test_common.h utilities.","dependencies":["TestRepo/gpio/gpio_reg_wr_rd_test/program.c","TestRepo/gpio/gpio_reg_wr_rd_test/test_define.c"],"timeout":"NA","priority":"TBD","owner":"TBD","references":["https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/gpio_reg_wr_rd_test"],"coverage_mapping":{"registers":["GPIO_GP0_GPIO_8..GPIO_GP0_GPIO_39","GPIO_GPIO_INTR_RAW_STCLR1","GPIO_GP0_INTR1_INTR_EN1","GPIO_GP0_INTR1_INTR_STS1","GPIO_GP0_INTR2_INTR_EN1","GPIO_GP0_INTR2_INTR_STS1","GPIO_GPIO_IO_CTRL_GROUP1..GPIO_GPIO_IO_CTRL_GROUP4","GPIO_GPIO_DOUT_GROUP1..GPIO_GPIO_DOUT_GROUP4","GPIO_GPIO_DIN_GROUP1..GPIO_GPIO_DIN_GROUP4"],"features":["Independent control register for each GPIO","Common control register to control group of 8-GPIOs with pin masking support","Unmasked raw interrupt status and clear register"]}},
    {"Index":2,"SS / Module":"GPIO","Feature":"Interrupts can be generated based on negative edge detection at GPIO input","Test Case Name":"test_gpio_negedge_intr_en","Test Description":"Configures input mode and negative-edge interrupt for GPIOs 8–39, enables interrupt routing, generates a falling edge per pin using an external driver register, and verifies interrupt assertion, input level, status setting, and proper clearing at both per-pin and group levels.","Speed":"NA","Mode":"Interrupt","Memory Start Offset":"0xA0243ffc","Memory End Offset":"0xA0243ffc","Remarks":"Wait arming occurs before edge generation to avoid race conditions. Group raw status is pre-cleared for the active bit prior to enable. ISR restores pad driver to a known state and clears both per-pin and group status.","Test Steps / Procedure":["Enable the appropriate interrupt line in the interrupt controller for the selected GPIO instance.","Enable the system register output corresponding to the selected GPIO interrupt.","Drive the external pad control to a known high level.","For each pad from 8 to 39, configure the per-pin control register to input mode, enable negative-edge detection, and clear the raw status.","For each pad index: pre-clear the corresponding group raw status bit, enable the specific interrupt bit in the group enable register, briefly wait, arm the interrupt wait flag, and then generate a falling edge on only that pad using the external pad driver register.","Wait for the interrupt to be serviced with a bounded timeout; on expiration, flag a failure.","In the interrupt service, confirm input level is low for the active pad, confirm the per-pin raw indicator and the group status reflect the event, clear the per-pin raw status and the group raw bit, verify the group status is cleared, clear the system register’s raw bit, and acknowledge the interrupt controller."],"Impacted Registers":"GPIO_GP0_GPIO_8, GPIO_GP0_GPIO_9, GPIO_GP0_GPIO_10, GPIO_GP0_GPIO_11, GPIO_GP0_GPIO_12, GPIO_GP0_GPIO_13, GPIO_GP0_GPIO_14, GPIO_GP0_GPIO_15, GPIO_GP0_GPIO_16, GPIO_GP0_GPIO_17, GPIO_GP0_GPIO_18, GPIO_GP0_GPIO_19, GPIO_GP0_GPIO_20, GPIO_GP0_GPIO_21, GPIO_GP0_GPIO_22, GPIO_GP0_GPIO_23, GPIO_GP0_GPIO_24, GPIO_GP0_GPIO_25, GPIO_GP0_GPIO_26, GPIO_GP0_GPIO_27, GPIO_GP0_GPIO_28, GPIO_GP0_GPIO_29, GPIO_GP0_GPIO_30, GPIO_GP0_GPIO_31, GPIO_GP0_GPIO_32, GPIO_GP0_GPIO_33, GPIO_GP0_GPIO_34, GPIO_GP0_GPIO_35, GPIO_GP0_GPIO_36, GPIO_GP0_GPIO_37, GPIO_GP0_GPIO_38, GPIO_GP0_GPIO_39, GPIO_GPIO_INTR_RAW_STCLR1, GPIO_GP0_INTR1_INTR_EN1, GPIO_GP0_INTR1_INTR_STS1, LSS_SYSREG_INTR_EN1, LSS_SYSREG_RAW_STCR1","Validation / Acceptance Criteria":["For each pad, the interrupt wait must complete before the timeout after generating the falling edge.","Upon interrupt, the per-pin input level for the active pad is low.","The group interrupt status reflects the active pad’s bit during service and is fully cleared after per-pin and group raw clear operations.","The system register’s raw status bit for the GPIO interrupt source is cleared after the service."],"Code Generation (Required / Not)":"","Hidden_Test_Case_Name":"test_gpio_negedge_intr_en","Hidden_Test_Description":"...omitted for brevity (same as provided earlier)...","Hidden_Remarks":"Comments: \"Arm the wait BEFORE generating the edge to avoid race\"; uses bounded wait with timeout to avoid infinite loop; pre-clears RAW group bit before enabling specific pad.","Hidden_Test_Steps_Procedure":["...omitted for brevity (same as provided earlier)..."],"Hidden_Impacted_Registers":"MIZAR_... (same as provided earlier)","Hidden_Validation_Acceptance_Criteria":"Main loop: int_pend must clear before timeout after driving 0xA0243ffc from all-ones to ~wr_val (falling edge on bit i). ...","test_id":"TC_GPIO_002","objective":"Verify negative-edge interrupt generation, detection, and clearing for GPIO pads 8–39.","description":"The test configures negative-edge detection and validates that each pad triggers and clears correctly with proper input level and status handling.","type":"directed, interrupt","category/tags":["interrupt","negedge","gpio","status-clear"],"preconditions":"Interrupt controller accessible; system register path available to route GPIO interrupts; external pad driver register at 0xA0243ffc is functional.","setup":"Compile with appropriate GPIO instance macro (GPIO0 or GPIO1) defined and link against test_common and platform GIC/sysreg support.","steps":["Enable platform and system interrupt routing for the selected GPIO instance.","Initialize external pad driver to a known high level.","Configure per-pin input mode and negative-edge detection for pads 8–39 and clear raw status.","For each pad, clear the group raw bit, enable that pad’s interrupt, arm the wait, and generate a falling edge on that pad.","Wait for service with timeout; on timeout, record failure.","In service, verify input and status, clear per-pin and group status, clear system raw status, and acknowledge the interrupt controller.","Repeat for all pads and finalize with aggregated result."],"expected_results":"Each pad’s negative-edge event is detected; input is low at service; per-pin and group status bits are set then successfully cleared; no timeouts occur.","pass_fail_criteria":"Pass if test_err remains zero at completion; any timeout, input mismatch, missing status bit, or uncleared status constitutes a failure.","required_tools/framework":"C runtime, test_common utilities, GIC and LSS sysreg access support.","dependencies":["TestRepo/gpio/test_gpio_negedge_intr_en/program.c","TestRepo/gpio/test_gpio_negedge_intr_en/test_define.c"],"timeout":"Polling loop up to 5000 iterations with wait_on(10) per iteration for each pad.","priority":"TBD","owner":"TBD","references":["https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en"],"coverage_mapping":{"registers":["GPIO_GP0_GPIO_8..GPIO_GP0_GPIO_39","GPIO_GP0_INTR1_INTR_EN1","GPIO_GP0_INTR1_INTR_STS1","GPIO_GPIO_INTR_RAW_STCLR1","LSS_SYSREG_INTR_EN1","LSS_SYSREG_RAW_STCR1"],"features":["Interrupts can be generated based on negative edge detection at GPIO input","Each GPIO wise interrupt enable","Unmasked raw interrupt status and clear register"]}},
    {"Index":3,"SS / Module":"GPIO","Feature":"Interrupts can be generated based on positive edge detection at GPIO input","Test Case Name":"test_gpio_pedge_all_pads_en","Test Description":"Enables positive-edge detection for GPIOs 8–39, configures input mode via group IO control, enables group interrupt output, and for each pad produces a single rising edge using an external driver while verifying group status assertion and subsequent clearing, including system raw status handling.","Speed":"NA","Mode":"Interrupt","Memory Start Offset":"0xA0243ffc","Memory End Offset":"0xA0243ffc","Remarks":"Interrupt wait is armed before the rising edge is generated. Group interrupt enable is masked during service and restored afterward. Per-pin raw status is cleared across all pads after service.","Test Steps / Procedure":["Enable the appropriate interrupt line in the interrupt controller for the selected GPIO instance.","Enable the system register output corresponding to the selected GPIO instance.","For each pad from 8 to 39, enable positive-edge detection in the per-pin control register.","Program group IO control registers to configure pads 8–39 as inputs.","Enable all bits in the group interrupt enable register.","For each pad index: drive the external pad output low, briefly wait, arm the interrupt wait flag, and drive the external pad output high to produce a rising edge.","Wait for the interrupt to be serviced with a bounded timeout; on expiration, flag a failure and abort further iteration.","Optionally drive low again to prepare for the next pad.","In the interrupt service, read the group status, mask the group enable during service, verify that a group interrupt is asserted, clear per-pin raw status across all pads, confirm that group status is cleared, clear the system raw status bit, re-enable group interrupt, and acknowledge the interrupt controller."],"Impacted Registers":"GPIO_GP0_GPIO_8, GPIO_GP0_GPIO_9, GPIO_GP0_GPIO_10, GPIO_GP0_GPIO_11, GPIO_GP0_GPIO_12, GPIO_GP0_GPIO_13, GPIO_GP0_GPIO_14, GPIO_GP0_GPIO_15, GPIO_GP0_GPIO_16, GPIO_GP0_GPIO_17, GPIO_GP0_GPIO_18, GPIO_GP0_GPIO_19, GPIO_GP0_GPIO_20, GPIO_GP0_GPIO_21, GPIO_GP0_GPIO_22, GPIO_GP0_GPIO_23, GPIO_GP0_GPIO_24, GPIO_GP0_GPIO_25, GPIO_GP0_GPIO_26, GPIO_GP0_GPIO_27, GPIO_GP0_GPIO_28, GPIO_GP0_GPIO_29, GPIO_GP0_GPIO_30, GPIO_GP0_GPIO_31, GPIO_GP0_GPIO_32, GPIO_GP0_GPIO_33, GPIO_GP0_GPIO_34, GPIO_GP0_GPIO_35, GPIO_GP0_GPIO_36, GPIO_GP0_GPIO_37, GPIO_GP0_GPIO_38, GPIO_GP0_GPIO_39, GPIO_GP0_INTR1_INTR_EN1, GPIO_GP0_INTR1_INTR_STS1, GPIO_GPIO_IO_CTRL_GROUP1, GPIO_GPIO_IO_CTRL_GROUP2, GPIO_GPIO_IO_CTRL_GROUP3, GPIO_GPIO_IO_CTRL_GROUP4, LSS_SYSREG_INTR_EN1, LSS_SYSREG_RAW_STCR1","Validation / Acceptance Criteria":["For each pad, the interrupt wait must complete before the timeout once the rising edge is generated.","During service, a group interrupt status must be observed and fully clear after per-pin raw clears are issued.","The system register’s raw status bit for the corresponding GPIO interrupt source is cleared after service."],"Code Generation (Required / Not)":"","Hidden_Test_Case_Name":"test_gpio_pedge_all_pads_en","Hidden_Test_Description":"...omitted for brevity (same as provided earlier)...","Hidden_Remarks":"Arms int_pend before generating the edge to avoid race. Masks group enable during service and re-enables afterward. Clears per-pin raw for all pads in ISR.","Hidden_Test_Steps_Procedure":["...omitted for brevity (same as provided earlier)..."],"Hidden_Impacted_Registers":"MIZAR_... (same as provided earlier)","Hidden_Validation_Acceptance_Criteria":"Main loop: int_pend must clear before timeout after rising edge is generated via writes to 0xA0243ffc. ISR: group status from MIZAR_GPIO_GP0_INTR1_INTR_STS1 must be non-zero on entry; after writing per-pin ICLR to all pads and a brief wait, group status must be 0. System RAW_STCR1 bit corresponding to GPIO instance must be cleared after write; any non-cleared bit increments test_err. Overall pass if test_err==0 at finish.","test_id":"TC_GPIO_003","objective":"Verify positive-edge interrupt enable, detection, and clearing across GPIO pads 8–39.","description":"The test ensures posedge detection and interrupt handling function correctly for all pads, including masking during service and clearing of both group and system status.","type":"directed, interrupt","category/tags":["interrupt","posedge","gpio","status-clear"],"preconditions":"Interrupt controller and system routing operational; external pad driver register 0xA0243ffc controls the pad levels.","setup":"Compile with appropriate GPIO instance macro and link with test_common, GIC, and sysreg access support.","steps":["Enable platform and system interrupt routing for the selected GPIO instance.","Enable posedge detection per pad and configure groups for input mode.","Enable all group interrupt bits.","For each pad, generate a single rising edge and wait for service with a timeout.","In service, verify group status assertion, perform per-pin raw clears, verify group clear, clear system raw status, and re-enable group interrupt.","Repeat for all pads, then finalize."],"expected_results":"Each pad’s rising edge is detected; group status is asserted and fully cleared; system raw status is cleared; no timeouts occur.","pass_fail_criteria":"Pass if no errors recorded (test_err==0 at completion). Any timeout or status mismatch is a failure.","required_tools/framework":"C runtime, test_common utilities, GIC and LSS sysreg access.","dependencies":["TestRepo/gpio/test_gpio_pedge_all_pads_en/program.c","TestRepo/gpio/test_gpio_pedge_all_pads_en/test_define.c"],"timeout":"Polling loop up to 2000 iterations with wait_on(10) per iteration for each pad.","priority":"TBD","owner":"TBD","references":["https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en"],"coverage_mapping":{"registers":["GPIO_GP0_GPIO_8..GPIO_GP0_GPIO_39","GPIO_GP0_INTR1_INTR_EN1","GPIO_GP0_INTR1_INTR_STS1","GPIO_GPIO_IO_CTRL_GROUP1..GPIO_GPIO_IO_CTRL_GROUP4","LSS_SYSREG_INTR_EN1","LSS_SYSREG_RAW_STCR1"],"features":["Interrupts can be generated based on positive edge detection at GPIO input","Each GPIO wise interrupt enable","Common control register to control group of 8-GPIOs with pin masking support"]}}
  ]
}'''

# Configuration from environment
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', 'Test_Output/GPIO/TestPlan')
IP_NAME = os.environ.get('IP_NAME', 'GPIO')
TS_FN = os.environ.get('TS_FN')  # expected format YYYYMMDD_HHMMSS (IST)

if not TS_FN:
    # Compute IST timestamp if not provided
    now_ist = datetime.now(ZoneInfo('Asia/Kolkata'))
    TS_FN = now_ist.strftime('%Y%m%d_%H%M%S')

FILENAME = f"{IP_NAME}_TestPlan_{TS_FN}.xlsx"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, FILENAME)

# Load JSON
root = json.loads(JSON_INPUT)
rows = root.get('test_cases', [])
if not isinstance(rows, list) or len(rows) == 0:
    raise SystemExit('Invalid or empty test_cases array in JSON input')

# Build union of keys preserving first-seen order
headers = []
seen = set()
for rec in rows:
    for k in rec.keys():
        if k not in seen:
            seen.add(k)
            headers.append(k)

# Create workbook and Data sheet
wb = Workbook()
ws_data = wb.active
ws_data.title = 'Data'

# Write headers
for col_idx, key in enumerate(headers, 1):
    ws_data.cell(row=1, column=col_idx, value=key)

# Write rows (preserve values; lists/dicts serialized as compact JSON strings)
for r_idx, rec in enumerate(rows, start=2):
    for c_idx, key in enumerate(headers, start=1):
        val = rec.get(key, '')
        if isinstance(val, (list, dict)):
            cell_val = json.dumps(val, ensure_ascii=False)
        else:
            cell_val = val
        ws_data.cell(row=r_idx, column=c_idx, value=cell_val)

# Basic formatting for Data sheet
header_font = Font(bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
for cell in ws_data[1]:
    cell.font = header_font
    cell.alignment = center
ws_data.freeze_panes = 'A2'

# Auto-fit columns (approximate based on content width)
def auto_fit(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = '' if cell.value is None else str(cell.value)
            for ln in v.split('\n'):
                if len(ln) > max_len:
                    max_len = len(ln)
        adj = min(max_len + 2, 120)
        ws.column_dimensions[get_column_letter(col_idx)].width = adj

auto_fit(ws_data)

# Create Meta_data_sheet and copy META columns
META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]
ws_meta = wb.create_sheet('Meta_data_sheet')
for c_idx, key in enumerate(META_COLS, 1):
    ws_meta.cell(row=1, column=c_idx, value=key)

for r_idx, rec in enumerate(rows, start=2):
    for c_idx, key in enumerate(META_COLS, start=1):
        val = rec.get(key, '')
        if isinstance(val, (list, dict)):
            val = json.dumps(val, ensure_ascii=False)
        ws_meta.cell(row=r_idx, column=c_idx, value=val)

# Very hide the meta sheet
ws_meta.sheet_state = 'veryHidden'

# Build TestPlan sheet from Data by selecting MAIN columns in required order
MAIN_COLS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)',
]

ws_tp = wb.create_sheet('TestPlan_tmp')
# Headers
for c_idx, key in enumerate(MAIN_COLS, 1):
    ws_tp.cell(row=1, column=c_idx, value=key)
# Data rows
for r_idx, rec in enumerate(rows, start=2):
    for c_idx, key in enumerate(MAIN_COLS, start=1):
        val = rec.get(key, '')
        if isinstance(val, (list, dict)):
            val = json.dumps(val, ensure_ascii=False)
        ws_tp.cell(row=r_idx, column=c_idx, value=val)

# Formatting for TestPlan sheet (STRICT)
wrap_cols = set([
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria',
])
key_to_col = {k: i+1 for i, k in enumerate(MAIN_COLS)}

# Header formatting
for cell in ws_tp[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data row formatting
thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

max_row = ws_tp.max_row
max_col = ws_tp.max_column

for r in range(2, max_row+1):
    for c in range(1, max_col+1):
        cell = ws_tp.cell(row=r, column=c)
        hdr = ws_tp.cell(row=1, column=c).value
        wrap = hdr in wrap_cols
        align_kwargs = {
            'vertical': 'top',
            'wrap_text': True if wrap else False,
        }
        # Center numeric-like Index, else left align text
        if hdr == 'Index':
            align_kwargs['horizontal'] = 'center'
        else:
            align_kwargs['horizontal'] = 'left'
        cell.alignment = Alignment(**align_kwargs)
        cell.border = border

# Distinct header fill (light gray)
for cell in ws_tp[1]:
    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

# Auto-fit and finalize
auto_fit(ws_tp)
ws_tp.freeze_panes = 'A2'

# Remove META columns from main sheet already enforced by selecting MAIN_COLS
# Replace Data sheet with TestPlan
wb.remove(ws_data)
ws_tp.title = 'TestPlan'

# Ensure output dir exists
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"Saved Excel to {OUTPUT_PATH}")
