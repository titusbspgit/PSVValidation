#!/usr/bin/env python3
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

EXPECTED_HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
    "Source Files",
    "ip",
    "repository",
    "branch",
    "subdirectory",
    "generated_timestamp_IST",
]

WRAP_HEADERS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
}


def main():
    if len(sys.argv) < 2:
        print("Usage: reformat_single_sheet_excel.py <xlsx_path>")
        sys.exit(2)

    xlsx_path = sys.argv[1]

    wb = load_workbook(filename=xlsx_path)

    # Choose working sheet: prefer 'TestPlan', else active
    ws = wb["TestPlan"] if "TestPlan" in wb.sheetnames else wb.active

    # Enforce single-sheet: delete all other sheets
    for name in list(wb.sheetnames):
        if name != ws.title:
            del wb[name]

    # Ensure name is exactly 'TestPlan'
    ws.title = "TestPlan"

    # Read header row
    headers = [c.value if c.value is not None else "" for c in ws[1]]
    header_to_index = {}
    for i, h in enumerate(headers):
        if h not in header_to_index:
            header_to_index[h] = i + 1  # 1-based

    # Verify all expected headers exist
    missing = [h for h in EXPECTED_HEADERS if h not in header_to_index]
    if missing:
        print("ERROR: Missing expected headers:", ", ".join(missing))
        sys.exit(3)

    max_row = ws.max_row

    # Create a new in-memory sheet to reorder columns
    tmp = wb.create_sheet(title="__TMP__")

    # Write headers in expected order
    for col_idx, header in enumerate(EXPECTED_HEADERS, start=1):
        tmp.cell(row=1, column=col_idx, value=header)

    # Copy rows 2..max_row in expected order
    for r in range(2, max_row + 1):
        for col_idx, header in enumerate(EXPECTED_HEADERS, start=1):
            src_col = header_to_index[header]
            tmp.cell(row=r, column=col_idx, value=ws.cell(row=r, column=src_col).value)

    # Remove original working sheet and rename tmp -> TestPlan
    del wb[ws.title]
    tmp.title = "TestPlan"
    ws = tmp

    # Formatting (visual only)
    # Freeze header and enable filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Header style
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in ws[1]:
        c.font = header_font
        c.alignment = header_align

    # Data rows alignment and wrapping for specific columns
    top_align = Alignment(vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    left_align = Alignment(horizontal="left", vertical="top")

    wrap_align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Map header to column index after reorder
    hdr_to_col = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            # Index column centered, others default left/top
            if c == hdr_to_col.get("Index"):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Apply wrap for long text fields
    for name in WRAP_HEADERS:
        cidx = hdr_to_col.get(name)
        if cidx is None:
            continue
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=cidx).alignment = wrap_align_left

    # Thin borders around used range
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border

    # Autofit approx by content length with bounds
    for c in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if not isinstance(v, str):
                v = str(v)
            if len(v) > max_len:
                max_len = len(v)
        width = max(10, min(80, max_len + 2))
        ws.column_dimensions[get_column_letter(c)].width = width

    # Save in place
    wb.save(xlsx_path)


if __name__ == "__main__":
    main()
