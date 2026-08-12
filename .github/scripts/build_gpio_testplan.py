#!/usr/bin/env python3
import json
import os
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:  # pragma: no cover
    ZoneInfo = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---- Configuration (from task inputs) ----
IP_NAME = "GPIO"
OUTPUT_DIRECTORY = "Test_Output/GPIO/TestPlan/"
TIMEZONE = "Asia/Kolkata"  # IST (GMT+05:30)

# final_json EXACT from Agent 6; do not modify
FINAL_JSON = r'''[
  {
    "status": "success",
    "repo": "titusbspgit/PSVValidation",
    "branch": "main",
    "ist_timestamp_used": "2026-08-12 00:00:00 IST",
    "filename": "GPIO_TestPlan_20260812_000000.xlsx",
    "path": "Test_Output/GPIO/TestPlan/GPIO_TestPlan_20260812_000000.xlsx",
    "file_html_url": "https://github.com/titusbspgit/PSVValidation/blob/main/Test_Output/GPIO/TestPlan/GPIO_TestPlan_20260812_000000.xlsx",
    "commit": {
      "sha": "d9213eaed17dd2e391e57ae31d66fa5fb582254e",
      "message": "Add GPIO TestPlan (GPIO) generated on 2026-08-12 00:00:00 IST",
      "html_url": "https://github.com/titusbspgit/PSVValidation/commit/d9213eaed17dd2e391e57ae31d66fa5fb582254e",
      "author": {
        "name": "titusbspgit",
        "email": "52008376+titusbspgit@users.noreply.github.com",
        "date": "2026-08-12T08:42:44Z"
      }
    },
    "verification": {
      "present_on_main": true,
      "file_sha": "2009f2a3f824267e45b1dfe43c6319bdcd59dfa6"
    }
  },
  {
    "Index": "2",
    "SS / Module": "GPIO",
    "Feature": "NA",
    "Test Case Name": "test_gpio_level_sel_intr_en",
    "Test Description": "Validates level-sensitive interrupt behavior for GP0 GPIO lines 8–39. The test enables the SoC interrupt routing, programs each GPIO line as input with level-interrupt enabled, enables the group interrupt mask, triggers and waits for the interrupt service to complete, verifies raw and group interrupt status, clears the raw and group statuses, and checks that the system interrupt status is cleared. The sequence is executed for both active-high and active-low level selections.",
    "Meta Test Description": "The testcase configures platform interrupt routing and exercises GP0 GPIO level-interrupt generation across 32 GPIOs starting at MIZAR_GPIO_GP0_GPIO_8. Under GPIO0, GIC_EnableIRQ(87) is called; under GPIO1, GIC_EnableIRQ(88). System register interrupt is enabled via write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIOx_INTR) based on the build selection. For each i in [0..31], it programs MIZAR_GPIO_GP0_GPIO_8 + (i*4) with 0x00180000 (input mode, level interrupt enable with level-select = active-high per comment), enables the group mask by write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, (1<<i)), initializes 0xA0243ffc with 0xffffffff, sets int_pend=1, and waits in a while loop until int_pend is cleared by the interrupt handler. After completing the first pass, it repeats for active-low selection by programming MIZAR_GPIO_GP0_GPIO_8 + (i*4) with 0x00100000 and writing ~(1<<i) to 0xA0243ffc, again waiting for ISR completion. The Default_IRQHandler reads MIZAR_GPIO_GP0_GPIO_8 + (i*4), checks (rdata & 0x2) != 0x0 (raw level status set), verifies the group interrupt in MIZAR_GPIO_GP0_INTR1_INTR_STS1 has bit i set, clears the raw interrupt by writing 0x00110000 to MIZAR_GPIO_GP0_GPIO_8 + (i*4), delays, and reads back expecting rdata == 0x100001. It disables the group interrupt by writing 0x00000000 to MIZAR_GPIO_GP0_INTR1_INTR_EN1 and verifies MIZAR_GPIO_GP0_INTR1_INTR_STS1 == 0x0. It then clears the system raw status: under GPIO0, write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR) and confirm (read_reg(MIZAR_LSS_SYSREG_RAW_STCR1) & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0; similarly for GPIO1 with LSS_SYSREG_RAW_STCR1_GPIO1_INTR. The handler writes 0xA0243ffc = 0xffffffff and sets int_pend = 0 before exit. finish(test_err) is invoked to conclude the test.",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Assumes the appropriate GIC interrupt (87 for one instance or 88 for the alternate instance) is available and routed to the handler. Uses SRAM address 0xA0243ffc as a synchronization/flag location during the test. The procedure iterates over 32 GPIO lines starting at gp0_gpio_8.",
    "Test Steps / Procedure": "1. Enable the platform interrupt for the targeted GPIO instance via the interrupt controller.\n2. Enable system-level interrupt routing for the selected GPIO instance using the corresponding enable register.\n3. For each GPIO line from 8 to 39 (32 entries starting at gp0_gpio_8):\n   3.1 Configure the corresponding gp0_gpio_8+i register for input mode with level-interrupt enabled (active-high selection for the first pass).\n   3.2 Enable the corresponding bit in gp0_intr1_intr_en1 to unmask the interrupt for that line.\n   3.3 Initialize the SRAM flag at 0xA0243ffc and wait until the interrupt service completes.\n   3.4 Verify the raw level status for the line and confirm the group status bit is set in gp0_intr1_intr_sts1.\n   3.5 Clear the raw interrupt in the gp0_gpio_8+i register, verify the readback reflects the cleared condition, then disable the line in gp0_intr1_intr_en1 and confirm gp0_intr1_intr_sts1 is cleared.\n4. Repeat Step 3 for active-low selection by reprogramming each gp0_gpio_8+i register for active-low level-interrupt and re-validating raw and group statuses as above.\n5. After each interrupt handling, verify the system-level raw status is cleared using the appropriate system register.\n6. The test passes if all lines satisfy the checks with no errors reported at the end of execution.",
    "Meta Test Steps / Procedure": "Initialization:\n- test_err = 0; declare rdata, wr_val.\n- Conditional IRQ enable: ifdef GPIO0 => GIC_EnableIRQ(87); ifdef GPIO1 => GIC_EnableIRQ(88).\n- Enable sysreg interrupt: ifdef GPIO0 => write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR); ifdef GPIO1 => write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR).\n\nPass 1 (active-high per comment):\n- for (i = 0; i < 32; i++):\n  - wr_val = 1 << i;\n  - write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4), 0x00180000);\n  - wait_on(50);\n  - write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val);\n  - wait_on(10);\n  - write_reg(0xA0243ffc, 0xffffffff);\n  - int_pend = 1;\n  - while (int_pend == 1): print \"Waiting for interrupt\"; wait_on(10);\n- wait_on(100);\n\nPass 2 (active-low per comment):\n- for (i = 0; i < 32; i++):\n  - wr_val = 1 << i;\n  - write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4), 0x00100000);\n  - wait_on(50);\n  - write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val);\n  - wait_on(10);\n  - write_reg(0xA0243ffc, ~(wr_val));\n  - int_pend = 1;\n  - while (int_pend == 1): print \"Waiting for interrupt\"; wait_on(10);\n\nInterrupt handler (Default_IRQHandler):\n- wr_val = 1 << i;\n- int_pend = 0;\n- write_reg(0xA0243ffc, 0xffffffff);\n- rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4));\n- if ((rdata & 0x2) != 0x0):\n  - rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1);\n  - if ((rdata_grp & (1<<i)) != 0): success; else: print error; test_err++;\n  - write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4), 0x00110000);\n  - wait_on(20);\n  - rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i * 4));\n  - if (rdata == 0x100001): success; else: print error; test_err++;\n  - write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000);\n  - rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1);\n  - if (rdata_grp == 0x0): success; else: print error; test_err++;\n  - ifdef GPIO0:\n      write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR);\n      rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1);\n      if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0): success; else: print error; test_err++;\n    endif\n  - ifdef GPIO1:\n      write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR);\n      rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1);\n      if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0): success; else: print error; test_err++;\n    endif\n- else: print \"Interrupt Not occured\"; test_err++;\n- Conditional IRQ clear: ifdef GPIO0 => GIC_ClearIRQ(87); ifdef GPIO1 => GIC_ClearIRQ(88).\n\nCompletion:\n- finish(test_err);",
    "Impacted Registers": "gp0_gpio_18; gp0_gpio_8; gp0_intr1_intr_en1; gp0_intr1_intr_sts1; gp0_gpio_17",
    "Meta Impacted Registers": "MIZAR_LSS_SYSREG_INTR_EN1; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_INTR1_INTR_EN1; 0xA0243ffc; MIZAR_GPIO_GP0_INTR1_INTR_STS1; MIZAR_LSS_SYSREG_RAW_STCR1",
    "Validation / Acceptance Criteria": "For each GPIO line (8–39): an interrupt is observed when enabled; the corresponding group status bit in gp0_intr1_intr_sts1 is set; clearing the line’s raw status results in the expected cleared readback and the group status becomes 0 after disabling; and the system interrupt status is cleared. The test passes if no errors are reported at the end of execution.",
    "Meta Validation / Acceptance Criteria": "- In Default_IRQHandler, (read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)) & 0x2) must be nonzero to indicate raw level status is set; else: FAIL and test_err++.\n- read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) must have bit i set ((rdata_grp & (1<<i)) != 0); else: FAIL and test_err++.\n- After write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00110000) and wait_on(20), readback rdata must equal 0x100001; else: FAIL and test_err++.\n- After write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000), read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) must equal 0x0; else: FAIL and test_err++.\n- For GPIO0 build: after write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR), (read_reg(MIZAR_LSS_SYSREG_RAW_STCR1) & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) must be 0; else: FAIL and test_err++.\n- For GPIO1 build: after write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR), (read_reg(MIZAR_LSS_SYSREG_RAW_STCR1) & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) must be 0; else: FAIL and test_err++.\n- Overall PASS: test_err == 0 at finish().",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include<lss_sysreg.h>;#include<stdio.h>;#include<test_define.c>;#include<test_common.h>;#include<gpio/gpio_def.h>;#include<gpio/gpio_offset.h>",
    "Meta Macros": "#define CNT 49",
    "Meta Arrays": "const unsigned long int addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,MIZAR_GPIO_GP0_GPIO_10,MIZAR_GPIO_GP0_GPIO_11,MIZAR_GPIO_GP0_GPIO_12,MIZAR_GPIO_GP0_GPIO_13,MIZAR_GPIO_GP0_GPIO_14,MIZAR_GPIO_GP0_GPIO_15,};\n\nconst unsigned int default_value_array[49]={GPIO_GP0_GPIO_8_DEFAULT_VAL,GPIO_GP0_GPIO_9_DEFAULT_VAL,GPIO_GP0_GPIO_10_DEFAULT_VAL,GPIO_GP0_GPIO_11_DEFAULT_VAL,GPIO_GP0_GPIO_12_DEFAULT_VAL,GPIO_GP0_GPIO_13_DEFAULT_VAL,GPIO_GP0_GPIO_14_DEFAULT_VAL,GPIO_GP0_GPIO_15_DEFAULT_VAL,};\n\nconst unsigned int read_mask_array[49]={GPIO_GP0_GPIO_8_READ_MASK,GPIO_GP0_GPIO_9_READ_MASK,GPIO_GP0_GPIO_10_READ_MASK,GPIO_GP0_GPIO_11_READ_MASK,GPIO_GP0_GPIO_12_READ_MASK,GPIO_GP0_GPIO_13_READ_MASK,GPIO_GP0_GPIO_14_READ_MASK,GPIO_GP0_GPIO_15_READ_MASK,};\n\nconst unsigned int write_mask_array[49]={GPIO_GP0_GPIO_8_WRITE_MASK,GPIO_GP0_GPIO_9_WRITE_MASK,GPIO_GP0_GPIO_10_WRITE_MASK,GPIO_GP0_GPIO_11_WRITE_MASK,GPIO_GP0_GPIO_12_WRITE_MASK,GPIO_GP0_GPIO_13_WRITE_MASK,GPIO_GP0_GPIO_14_WRITE_MASK,GPIO_GP0_GPIO_15_WRITE_MASK,};\n\nconst int skip_array[49]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,};
  }
]'''


def make_dirs(path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)


def auto_col_widths(ws):
    # Determine reasonable column widths based on content length
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            # consider multi-line
            for line in val.splitlines() if val else [""]:
                max_len = max(max_len, len(line))
        width = max(12, min(max_len + 2, 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_header_format(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"


def wrap_all(ws):
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap


def build_workbook(final_json: list) -> tuple[Workbook, str]:
    meta = final_json[0]
    row_obj = final_json[1]

    # Prepare workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"
    ws_meta = wb.create_sheet("MetaData")

    # TestPlan sheet: headers in the insertion order of JSON keys
    headers = list(row_obj.keys())
    ws.append(headers)
    ws.append([row_obj.get(k, "") for k in headers])

    apply_header_format(ws)
    wrap_all(ws)
    auto_col_widths(ws)

    # MetaData sheet: single row of metadata (stringify nested values)
    meta_headers = list(meta.keys())
    meta_values = [json.dumps(meta[k], ensure_ascii=False) if isinstance(meta[k], (dict, list)) else meta[k] for k in meta_headers]
    ws_meta.append(meta_headers)
    ws_meta.append(meta_values)
    apply_header_format(ws_meta)
    wrap_all(ws_meta)
    auto_col_widths(ws_meta)
    ws_meta.sheet_state = "veryHidden"

    # Determine filename and path
    present = False
    try:
        present = bool(meta.get("verification", {}).get("present_on_main"))
    except Exception:
        present = False

    if present and meta.get("filename") and meta.get("path"):
        filename = meta["filename"]
        output_path = meta["path"]
    else:
        if ZoneInfo is None:
            ts = datetime.utcnow()
        else:
            ts = datetime.now(ZoneInfo(TIMEZONE))
        filename = f"{IP_NAME}_TestPlan_{ts:%Y%m%d}_{ts:%H%M%S}.xlsx"
        output_path = os.path.join(OUTPUT_DIRECTORY, filename).replace("\\", "/")

    make_dirs(output_path)
    wb.save(output_path)
    return wb, output_path


def main():
    data = json.loads(FINAL_JSON)
    _, out_path = build_workbook(data)
    print(f"Wrote Excel to: {out_path}")


if __name__ == "__main__":
    main()
