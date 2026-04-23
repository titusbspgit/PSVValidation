#!/usr/bin/env python3
import json
import sys
import os
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_TEXT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
DATA_CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def now_ist():
    if ZoneInfo is None:
        # Fallback to naive localtime; GitHub Actions has zoneinfo available on Python 3.11+
        return datetime.now()
    return datetime.now(ZoneInfo("Asia/Kolkata"))


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def stringify_value(val):
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)
    return val if val is not None else ""


def union_keys_preserve_order(rows):
    seen = []
    seen_set = set()
    for r in rows:
        for k in r.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def write_autofit(ws):
    # Approximate autofit: set column width based on max text length
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_len = 0
        for cell in col:
            try:
                text = str(cell.value) if cell.value is not None else ""
            except Exception:
                text = ""
            if len(text) > max_len:
                max_len = len(text)
        # Padding factor; cap to prevent excessively wide columns
        width = min(max_len + 2, 80)
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width


def style_table(ws, wrap_headers=None):
    # Header styling
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    # Data cell borders and alignment
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in r:
            c.border = THIN_BORDER
            # Default left/top alignment
            c.alignment = DATA_TEXT_ALIGN
    # Center the Index column if present
    try:
        idx_col = [c.value for c in ws[1]].index("Index") + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx_col).alignment = DATA_CENTER_ALIGN
    except ValueError:
        pass
    # Wrap specific columns
    if wrap_headers:
        headers = [c.value for c in ws[1]]
        for h in wrap_headers:
            if h in headers:
                col = headers.index(h) + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).alignment = DATA_TEXT_ALIGN


def build_workbook(test_cases, ip_name, out_dir):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"

    # Normalize schema
    keys_order = union_keys_preserve_order(test_cases)

    # Write headers
    for c, k in enumerate(keys_order, start=1):
        ws_data.cell(row=1, column=c, value=k)

    # Write rows
    for r_idx, row in enumerate(test_cases, start=2):
        for c_idx, k in enumerate(keys_order, start=1):
            ws_data.cell(row=r_idx, column=c_idx, value=stringify_value(row.get(k, "")))

    # Basic formatting on Data
    ws_data.freeze_panes = "A2"
    style_table(ws_data, wrap_headers=None)
    write_autofit(ws_data)

    # Meta_data_sheet creation
    ws_meta = wb.create_sheet("Meta_data_sheet")
    for c, k in enumerate(META_COLUMNS, start=1):
        ws_meta.cell(row=1, column=c, value=k)
    for r_idx, row in enumerate(test_cases, start=2):
        for c_idx, k in enumerate(META_COLUMNS, start=1):
            ws_meta.cell(row=r_idx, column=c_idx, value=stringify_value(row.get(k, "")))
    ws_meta.sheet_state = "veryHidden"

    # Prepare TestPlan sheet from Data
    ws_data.title = "TestPlan"

    # Rebuild TestPlan with only MAIN_COLUMNS in required order
    # Capture all current rows into memory
    headers = [cell.value for cell in ws_data[1]]
    data_rows = []
    for r in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
        data_rows.append({h: v for h, v in zip(headers, r)})

    # Clear the sheet
    ws_data.delete_rows(1, ws_data.max_row)

    # Write final headers
    for c, k in enumerate(MAIN_COLUMNS, start=1):
        ws_data.cell(row=1, column=c, value=k)

    # Write final rows (preserve values exactly; fill missing with blank)
    for r_idx, row_map in enumerate(data_rows, start=2):
        for c_idx, k in enumerate(MAIN_COLUMNS, start=1):
            ws_data.cell(row=r_idx, column=c_idx, value=stringify_value(row_map.get(k, "")))

    # Strict formatting for TestPlan
    ws_data.freeze_panes = "A2"
    style_table(ws_data, wrap_headers=WRAP_COLUMNS)
    write_autofit(ws_data)

    # Data validation for Code Generation (Required / Not)
    headers = [c.value for c in ws_data[1]]
    if "Code Generation (Required / Not)" in headers:
        col_idx = headers.index("Code Generation (Required / Not)") + 1
        col_letter = ws_data.cell(row=1, column=col_idx).column_letter
        dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showDropDown=True)
        dv.error = "Select only: Required or Not Required (or leave blank)."
        dv.errorTitle = "Invalid Selection"
        ws_data.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws_data.max_row}")

    # Build filename based on IST
    ts = now_ist()
    date_str = ts.strftime("%Y%m%d")
    time_str = ts.strftime("%H%M%S")
    file_name = f"{ip_name}_TestPlan_{date_str}_{time_str}.xlsx"

    ensure_dir(out_dir)
    # Write a timestamp helper for the workflow commit message
    with open(os.path.join(out_dir, ".last_ist_timestamp"), "w", encoding="utf-8") as f:
        f.write(ts.strftime("%Y-%m-%d %H:%M:%S"))

    out_path = os.path.join(out_dir, file_name)
    wb.save(out_path)
    print(out_path)
    return out_path


def main():
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--json", required=True, help="Path to JSON input with META_DATA and test_cases")
    p.add_argument("--ip", required=True, help="IP name for file naming")
    p.add_argument("--output-dir", required=True, help="Directory to save the final Excel")
    args = p.parse_args()

    data = load_json(args.json)

    if not isinstance(data, dict) or "test_cases" not in data or not isinstance(data["test_cases"], list) or len(data["test_cases"]) == 0:
        print("ERROR: Invalid or empty test_cases in JSON.", file=sys.stderr)
        sys.exit(2)

    build_workbook(data["test_cases"], args.ip, args.output_dir)


if __name__ == "__main__":
    main()
