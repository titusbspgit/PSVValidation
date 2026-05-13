import json
import os
import sys
from datetime import datetime, timezone, timedelta
from zipfile import is_zipfile
from typing import List, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# -------------------- Input JSON (embedded) --------------------
DATA: List[Dict] = [
  {
    "Index": 1,
    "SS / Module": "PCIE0 SII RC",
    "Feature": "writeAsRead",
    "Test Case Name": "pcie0_sii_rc_reg_wr_rd_test",
    "Test Description": "Verifies that registers report reset defaults and that writes update only writable fields while read-only fields remain unchanged.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "0xE68C0000",
    "Memory End Offset": "NA",
    "Remarks": "Non-readable entries are skipped. Non-writable entries are skipped. Entries in the skip list are skipped. The PHY reset control register is excluded from default checks. The soft reset is not executed.",
    "Test Steps / Procedure": "1) Read each readable SII register and compare with its documented default; exclude the PHY reset control register.\n2) For each data pattern, write to each writable SII register that is not skipped.\n3) Read back each affected register and compute the expected value using writable and read-only fields.\n4) Compare read-back with the expected value across all registers and patterns and record mismatches.\n5) Declare pass if no mismatches are found; otherwise declare fail.",
    "Impacted Registers": "SII_CFG_BAR0_START1\nSII_CFG_BAR0_START2\nSII_CFG_BAR0_LIMIT1\nSII_CFG_BAR0_LIMIT2\nSII_CFG_BAR1_START\nSII_CFG_BAR1_LIMIT1\nSII_CFG_BAR2_START1\nSII_CFG_BAR2_START2\nSII_CFG_BAR2_LIMIT1\nSII_CFG_BAR2_LIMIT2\nSII_CFG_BAR3_START\nSII_CFG_BAR3_LIMIT\nSII_CFG_BAR4_START1\nSII_CFG_BAR4_START2\nSII_CFG_BAR4_LIMIT1\nSII_CFG_BAR4_LIMIT2\nSII_CFG_BAR5_START\nSII_CFG_BAR5_LIMIT\nSII_PCIE0_CONFIG_INFO1\nSII_PCIE0_CONFIG_INFO2\nSII_PCIE0_GEN_CONTROL1\nSII_PCIE0_GEN_CONTROL2\nSII_PCIE0_GEN_CONTROL3\nSII_PCIE0_PM_CONTROL\nSII_PCIE0_CONTROL_PM_STS\nSII_PCIE0_TRANSMIT_HEADER1\nSII_PCIE0_TRANSMIT_HEADER2\nSII_PCIE0_TRANSMIT_HEADER3\nSII_PCIE0_TRANSMIT_HEADER4\nSII_PCIE0_TRANSMIT_REQ\nSII_PCIE0_RCV_MSG_HDR1\nSII_PCIE0_RCV_MSG_HDR2\nSII_PCIE0_RCV_MSG_HDR3\nSII_PCIE0_RCV_MSG_HDR4\nSII_PCIE0_RCV_MSG_STS\nSII_RCV_INTERRPUT_CTRL\nSII_CFG_EXP_ROM_START\nSII_CFG_EXP_ROM_LIMIT\nSII_CFG_EXP_ROM_INFO\nSII_CXPL_DEBUG_INFO1\nSII_CXPL_DEBUG_INFO2\nSII_CXPL_DEBUG_INFO_EI\nSII_PCIE0_TARGET_INFO1\nSII_PCIE0_TARGET_INFO2\nSII_PCIE0_CONTOLLER_ERROR_STATUS\nSII_PCIE0_CONTROLLER_INT_STS\nSII_PCIE0_CONTROLLER_INTERRUPT_CONTROL\nSII_PHY_RST_CONTROL\nSII_LINK_DEBUG_DATA\nSII_PCIE0_ERR_STS\nSII_PCIE0_ERR_INTERRUPT_CTRL\nSII_CFG_MSI_INT\nSII_LTR_MSG\nSII_LTR_MSG_LATENCY\nSII_APP_LTR_LATENCY\nSII_CFG_LTR_MAX_LATENCY\nSII_OBFF_CNTRL\nSII_SLV_AWMISC_INFO\nSII_SLV_AWMISC_INFO_HDR_34DW_HI\nSII_SLV_AWMISC_INFO_HDR_34DW_LO\nSII_SLV_MISC_INFO\nSII_SLV_MISC_RESP_INFO\nSII_MSTR_AWMISC_INFO_CNTRL\nSII_MSTR_AWMISC_INFO_1\nSII_MSTR_AWMISC_INFO_0\nSII_MSTR_AWMISC_INFO_HDR_34DW_HI\nSII_MSTR_AWMISC_INFO_HDR_34DW_LO\nSII_MSTR_ARMISC_INFO_CNTRL\nSII_MSTR_ARMISC_INFO_1\nSII_MSTR_ARMISC_INFO_0\nSII_MSTR_BMISC_RMISC_CPL_STAT_INFO\nSII_RADM_TIMEOUT_INFO\nSII_CFG_MSI_INFO\nSII_CFG_MSI_DATA\nSII_CFG_MSI_ADDR_HI\nSII_CFG_MSI_ADDR_LO\nSII_CFG_AER_INT_AND_PCIE0_CAP_INT_MSG\nSII_RTLH_RFC_DATA\nSII_APP_HDR_INFO\nSII_APP_HDR_LOG_3\nSII_APP_HDR_LOG_2\nSII_APP_HDR_LOG_1\nSII_APP_HDR_LOG_0\nSII_CFG_BUS_NUM\nSII_CFG_BR_CTRL_SERREN\nSII_APP_DEV_AND_BUS_NUM\nSII_PCIE0_CONTROLLER_INT_STS_1\nSII_PCIE0_CONTROLLER_INTERRUPT_CONTROL_1\nSII_APP_AND_SLOT_CONTROL_REG\nSII_DIAG_CTRL_BUS\nSII_CFG_REG_RO\nSII_CFG_ARI_FWD_EN\nSII_RADM_SLOT_PWR_PAYLOAD\nSII_DIAG_STATUS_BUS_0\nSII_DIAG_STATUS_BUS_1\nSII_DIAG_STATUS_BUS_2\nSII_DIAG_STATUS_BUS_3\nSII_DIAG_STATUS_BUS_4\nSII_DIAG_STATUS_BUS_5\nSII_DIAG_STATUS_BUS_6\nSII_DIAG_STATUS_BUS_7\nSII_DIAG_STATUS_BUS_8\nSII_DIAG_STATUS_BUS_9\nSII_DIAG_STATUS_BUS_10\nSII_DIAG_STATUS_BUS_11\nSII_DIAG_STATUS_BUS_12\nSII_DIAG_STATUS_BUS_13\nSII_DIAG_STATUS_BUS_14\nSII_DIAG_STATUS_BUS_15\nSII_DIAG_STATUS_BUS_16\nSII_DIAG_STATUS_BUS_17\nSII_DIAG_STATUS_BUS_18\nSII_DIAG_STATUS_BUS_19\nSII_RAM_PWR_CNTRL_0\nSII_RAM_PWR_CNTRL_1\nSII_SOFT_RESET_CTRL\nSII_CFG_MSI_PENDING_B\nSII_SMLH_LTSSM_STATE_TRAN_1\nSII_SMLH_LTSSM_STATE_TRAN_2\nSII_SMLH_LTSSM_STATE_TRAN_3\nSII_SMLH_LTSSM_STATE_TRAN_4\nSII_SMLH_LTSSM_STATE_TRAN_5\nSII_SMLH_LTSSM_STATE_TRAN_6\nSII_SMLH_LTSSM_STATE_TRAN_7\nSII_PHY_CONTROL_0\nSII_PHY_CONTROL_1\nSII_PHY_CONTROL_2\nSII_PHY_CONTROL_3\nSII_PHY_CONTROL_4\nSII_PHY_CONTROL_5\nSII_PHY_CONTROL_6\nSII_PHY_CONTROL_7\nSII_PHY_CONTROL_8\nSII_PHY_CONTROL_9\nSII_PHY_CONTROL_10\nSII_PHY_CONTROL_11\nSII_PHY_CONTROL_12\nSII_PHY_CONTROL_13\nSII_PHY_CONTROL_14\nSII_PHY_CONTROL_15\nSII_PHY_CONTROL_16\nSII_PHY_CONTROL_17\nSII_PHY_CONTROL_18\nSII_PHY_CONTROL_19\nSII_PHY_CONTROL_20\nSII_PHY_CONTROL_21\nSII_PHY_CONTROL_22\nSII_PHY_CONTROL_23\nSII_PHY_CONTROL_24\nSII_PHY_CONTROL_25\nSII_PHY_CONTROL_26\nSII_MSI_CTRL_IO\nSII_MSI_CTRL_INT_VEC"
  },
  {
    "Index": 2,
    "SS / Module": "PCIE1 SII RC",
    "Feature": "writeAsRead",
    "Test Case Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Test Description": "Verifies that registers report reset defaults and that writes update only writable fields while read-only fields remain unchanged.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "0xE68C1000",
    "Memory End Offset": "NA",
    "Remarks": "Non-readable entries are skipped. Non-writable entries are skipped. Entries in the skip list are skipped. The PHY reset control register is excluded from default checks. The soft reset is not executed.",
    "Test Steps / Procedure": "1) Read each readable SII register and compare with its documented default; exclude the PHY reset control register.\n2) For each data pattern, write to each writable SII register that is not skipped.\n3) Read back each affected register and compute the expected value using writable and read-only fields.\n4) Compare read-back with the expected value across all registers and patterns and record mismatches.\n5) Declare pass if no mismatches are found; otherwise declare fail.",
    "Impacted Registers": "SII_CFG_BAR0_START1\nSII_CFG_BAR0_START2\nSII_CFG_BAR0_LIMIT1\nSII_CFG_BAR0_LIMIT2\nSII_CFG_BAR1_START\nSII_CFG_BAR1_LIMIT1\nSII_CFG_BAR2_START1\nSII_CFG_BAR2_START2\nSII_CFG_BAR2_LIMIT1\nSII_CFG_BAR2_LIMIT2\nSII_CFG_BAR3_START\nSII_CFG_BAR3_LIMIT\nSII_CFG_BAR4_START1\nSII_CFG_BAR4_START2\nSII_CFG_BAR4_LIMIT1\nSII_CFG_BAR4_LIMIT2\nSII_CFG_BAR5_START\nSII_CFG_BAR5_LIMIT\nSII_PCIE1_CONFIG_INFO1\nSII_PCIE1_CONFIG_INFO2\nSII_PCIE1_GEN_CONTROL1\nSII_PCIE1_GEN_CONTROL2\nSII_PCIE1_GEN_CONTROL3\nSII_PCIE1_PM_CONTROL\nSII_PCIE1_CONTROL_PM_STS\nSII_PCIE1_TRANSMIT_HEADER1\nSII_PCIE1_TRANSMIT_HEADER2\nSII_PCIE1_TRANSMIT_HEADER3\nSII_PCIE1_TRANSMIT_HEADER4\nSII_PCIE1_TRANSMIT_REQ\nSII_PCIE1_RCV_MSG_HDR1\nSII_PCIE1_RCV_MSG_HDR2\nSII_PCIE1_RCV_MSG_HDR3\nSII_PCIE1_RCV_MSG_HDR4\nSII_PCIE1_RCV_MSG_STS\nSII_RCV_INTERRPUT_CTRL\nSII_CFG_EXP_ROM_START\nSII_CFG_EXP_ROM_LIMIT\nSII_CFG_EXP_ROM_INFO\nSII_CXPL_DEBUG_INFO1\nSII_CXPL_DEBUG_INFO2\nSII_CXPL_DEBUG_INFO_EI\nSII_PCIE1_TARGET_INFO1\nSII_PCIE1_TARGET_INFO2\nSII_PCIE1_CONTOLLER_ERROR_STATUS\nSII_PCIE1_CONTROLLER_INT_STS\nSII_PCIE1_CONTROLLER_INTERRUPT_CONTROL\nSII_PHY_RST_CONTROL\nSII_LINK_DEBUG_DATA\nSII_PCIE1_ERR_STS\nSII_PCIE1_ERR_INTERRUPT_CTRL\nSII_CFG_MSI_INT\nSII_LTR_MSG\nSII_LTR_MSG_LATENCY\nSII_APP_LTR_LATENCY\nSII_CFG_LTR_MAX_LATENCY\nSII_OBFF_CNTRL\nSII_SLV_AWMISC_INFO\nSII_SLV_AWMISC_INFO_HDR_34DW_HI\nSII_SLV_AWMISC_INFO_HDR_34DW_LO\nSII_SLV_MISC_INFO\nSII_SLV_MISC_RESP_INFO\nSII_MSTR_AWMISC_INFO_CNTRL\nSII_MSTR_AWMISC_INFO_1\nSII_MSTR_AWMISC_INFO_0\nSII_MSTR_AWMISC_INFO_HDR_34DW_HI\nSII_MSTR_AWMISC_INFO_HDR_34DW_LO\nSII_MSTR_ARMISC_INFO_CNTRL\nSII_MSTR_ARMISC_INFO_1\nSII_MSTR_ARMISC_INFO_0\nSII_MSTR_BMISC_RMISC_CPL_STAT_INFO\nSII_RADM_TIMEOUT_INFO\nSII_CFG_MSI_INFO\nSII_CFG_MSI_DATA\nSII_CFG_MSI_ADDR_HI\nSII_CFG_MSI_ADDR_LO\nSII_CFG_AER_INT_AND_PCIE1_CAP_INT_MSG\nSII_RTLH_RFC_DATA\nSII_APP_HDR_INFO\nSII_APP_HDR_LOG_3\nSII_APP_HDR_LOG_2\nSII_APP_HDR_LOG_1\nSII_APP_HDR_LOG_0\nSII_CFG_BUS_NUM\nSII_CFG_BR_CTRL_SERREN\nSII_APP_DEV_AND_BUS_NUM\nSII_PCIE1_CONTROLLER_INT_STS_1\nSII_PCIE1_CONTROLLER_INTERRUPT_CONTROL_1\nSII_APP_AND_SLOT_CONTROL_REG\nSII_DIAG_CTRL_BUS\nSII_CFG_REG_RO\nSII_CFG_ARI_FWD_EN\nSII_RADM_SLOT_PWR_PAYLOAD\nSII_DIAG_STATUS_BUS_0\nSII_DIAG_STATUS_BUS_1\nSII_DIAG_STATUS_BUS_2\nSII_DIAG_STATUS_BUS_3\nSII_DIAG_STATUS_BUS_4\nSII_DIAG_STATUS_BUS_5\nSII_DIAG_STATUS_BUS_6\nSII_DIAG_STATUS_BUS_7\nSII_DIAG_STATUS_BUS_8\nSII_DIAG_STATUS_BUS_9\nSII_DIAG_STATUS_BUS_10\nSII_DIAG_STATUS_BUS_11\nSII_DIAG_STATUS_BUS_12\nSII_DIAG_STATUS_BUS_13\nSII_DIAG_STATUS_BUS_14\nSII_DIAG_STATUS_BUS_15\nSII_DIAG_STATUS_BUS_16\nSII_DIAG_STATUS_BUS_17\nSII_DIAG_STATUS_BUS_18\nSII_DIAG_STATUS_BUS_19\nSII_RAM_PWR_CNTRL_0\nSII_RAM_PWR_CNTRL_1\nSII_SOFT_RESET_CTRL\nSII_CFG_MSI_PENDING_B\nSII_SMLH_LTSSM_STATE_TRAN_1\nSII_SMLH_LTSSM_STATE_TRAN_2\nSII_SMLH_LTSSM_STATE_TRAN_3\nSII_SMLH_LTSSM_STATE_TRAN_4\nSII_SMLH_LTSSM_STATE_TRAN_5\nSII_SMLH_LTSSM_STATE_TRAN_6\nSII_SMLH_LTSSM_STATE_TRAN_7\nSII_PHY_CONTROL_0\nSII_PHY_CONTROL_1\nSII_PHY_CONTROL_2\nSII_PHY_CONTROL_3\nSII_PHY_CONTROL_4\nSII_PHY_CONTROL_5\nSII_PHY_CONTROL_6\nSII_PHY_CONTROL_7\nSII_PHY_CONTROL_8\nSII_PHY_CONTROL_9\nSII_PHY_CONTROL_10\nSII_PHY_CONTROL_11\nSII_PHY_CONTROL_12\nSII_PHY_CONTROL_13\nSII_PHY_CONTROL_14\nSII_PHY_CONTROL_15\nSII_PHY_CONTROL_16\nSII_PHY_CONTROL_17\nSII_PHY_CONTROL_18\nSII_PHY_CONTROL_19\nSII_PHY_CONTROL_20\nSII_PHY_CONTROL_21\nSII_PHY_CONTROL_22\nSII_PHY_CONTROL_23\nSII_PHY_CONTROL_24\nSII_PHY_CONTROL_25\nSII_PHY_CONTROL_26\nSII_MSI_CTRL_IO\nSII_MSI_CTRL_INT_VEC"
  },
  {
    "Index": 3,
    "SS / Module": "PCIE",
    "Feature": "PF_TYPE0_HDR_DBI2i (DBI2 Shadow Block: PF PCI-Compatible Configuration Space Header Type0)",
    "Test Case Name": "pcie_cfg_wr_rd_test",
    "Test Description": "Verifies PCIe configuration by programming coherency controls, configuring BARs, polling readiness, and confirming completion.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "0xE6004100",
    "Memory End Offset": "NA",
    "Remarks": "Execution depends on compile-time role defines. Status is polled until ready. Completion requires a signature at a control register.",
    "Test Steps / Procedure": "1) Program DBI_DSP_COHERENCY_CONTROL_3_OFF for the first instance and then for the second instance. 2) Poll the status register until readiness is indicated. 3) For the first instance, read configuration space, then program all BAR registers with patterns and enable memory, I/O, and bus mastering. 4) For the second instance, read configuration space, then program all BAR registers with patterns and enable memory, I/O, and bus mastering. 5) Wait and poll the control register until the completion signature is observed, then end the test.",
    "Impacted Registers": "DBI_DSP_COHERENCY_CONTROL_3_OFF"
  }
]

IP_NAME = os.environ.get("IP_NAME", "PCIE")
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "Test_Output/PCIE/TestPlan")
BRANCH = os.environ.get("BRANCH", "main")

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
    "Hidden_Header_Includes",
    "Hidden_Macro_Define",
    "Hidden_Skip_Array_Definition",
]

WRAP_COLS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=False)
THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def validate_json(data):
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("JSON must be a non-empty array")
    for i, row in enumerate(data):
        if not isinstance(row, dict):
            raise ValueError(f"Row {i} is not an object")


def first_seen_keys(data):
    order = []
    for row in data:
        for k in row.keys():
            if k not in order:
                order.append(k)
    return order


def auto_width(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            text = "" if val is None else str(val)
            lines = text.split("\n")
            max_len = max((len(x) for x in lines), default=0)
            widths[idx] = max(widths.get(idx, 10), min(120, max_len + 2))
    for idx, w in widths.items():
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else _colname(idx)].width = w


def _colname(n):
    name = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(65 + r) + name
    return name


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = BORDER


def number_multiline(text: str) -> str:
    if text is None:
        return ""
    lines = [l.strip() for l in str(text).split("\n") if l.strip() != ""]
    return "\n".join([f"{i+1}. {l}" for i, l in enumerate(lines)])


def main():
    validate_json(DATA)
    keys_order = first_seen_keys(DATA)
    # Create workbook and Data sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers in first-seen order
    for c, key in enumerate(keys_order, start=1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL

    # Write rows
    for r, row in enumerate(DATA, start=2):
        for c, key in enumerate(keys_order, start=1):
            ws.cell(row=r, column=c, value=row.get(key, ""))

    ws.freeze_panes = "A2"

    # Create Meta_data_sheet with only present META columns
    meta_present = [k for k in META_COLUMNS if k in keys_order]
    meta = wb.create_sheet("Meta_data_sheet")
    if meta_present:
        for c, key in enumerate(meta_present, start=1):
            mc = meta.cell(row=1, column=c, value=key)
            mc.font = HEADER_FONT
            mc.alignment = CENTER
            mc.fill = BLUE_FILL
        for r, row in enumerate(DATA, start=2):
            for c, key in enumerate(meta_present, start=1):
                meta.cell(row=r, column=c, value=row.get(key, ""))
    meta.sheet_state = "veryHidden"

    # Normalize main sheet: remove META columns and reorder to MAIN order (preserving extras at end)
    cols_to_keep = [k for k in keys_order if k not in META_COLUMNS]
    final_order = [k for k in MAIN_COLUMNS if k in cols_to_keep]
    extras = [k for k in cols_to_keep if k not in final_order]
    final_order += extras

    # Rebuild Data sheet with final_order
    data_matrix = []
    for r in range(2, ws.max_row + 1):
        row_dict = {keys_order[c-1]: ws.cell(row=r, column=c).value for c in range(1, len(keys_order)+1)}
        data_matrix.append(row_dict)

    ws.delete_rows(1, ws.max_row)
    for c, key in enumerate(final_order, start=1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL
    for r, row in enumerate(data_matrix, start=2):
        for c, key in enumerate(final_order, start=1):
            ws.cell(row=r, column=c, value=row.get(key, ""))

    # Rename Data -> TestPlan
    ws.title = "TestPlan"

    # Wrapping, alignment and numbering in specific columns
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            header = ws.cell(row=1, column=c).value
            if header in WRAP_COLS:
                # Numbering for the two required columns
                if header in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                    cell.value = number_multiline(cell.value)
                cell.alignment = LEFT_WRAP
            else:
                # Align text left, numbers right/center
                try:
                    float(cell.value)
                    cell.alignment = RIGHT
                except (TypeError, ValueError):
                    if header == "Index":
                        cell.alignment = CENTER
                    else:
                        cell.alignment = LEFT

    # Data validation on Code Generation (Required / Not)
    if "Code Generation (Required / Not)" in header_map:
        col_idx = header_map["Code Generation (Required / Not)"]
        col_letter = _colname(col_idx)
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    auto_width(ws)

    # Row heights heuristic after wrapping
    wrap_idx = [header_map[h] for h in WRAP_COLS if h in header_map]
    for r in range(2, ws.max_row + 1):
        max_lines = 1
        for c in wrap_idx:
            val = ws.cell(row=r, column=c).value
            lines = 1 if val is None else str(val).count("\n") + 1
            max_lines = max(max_lines, lines)
        ws.row_dimensions[r].height = min(410, 14 * max_lines)

    apply_borders(ws)

    # Ensure no 'Data' sheet exists
    if "Data" in wb.sheetnames:
        del wb["Data"]

    # Build IST timestamped filename
    ist = timezone(timedelta(hours=5, minutes=30))
    ts = datetime.now(ist).strftime("%Y%m%d_%H%M%S")
    filename = f"{IP_NAME}_TestPlan_{ts}.xlsx"
    outdir = OUTPUT_DIR
    os.makedirs(outdir, exist_ok=True)
    outpath = os.path.join(outdir, filename)

    wb.save(outpath)

    # Validate as true XLSX (ZIP) and reload
    if not is_zipfile(outpath):
        print("XLSX validation failed: not a zip file", file=sys.stderr)
        sys.exit(2)
    try:
        load_wb = load_workbook(outpath)
        # Final sheet visibility check
        if "Data" in load_wb.sheetnames:
            print("Validation failed: Data sheet still present", file=sys.stderr)
            sys.exit(3)
        if "TestPlan" not in load_wb.sheetnames or "Meta_data_sheet" not in load_wb.sheetnames:
            print("Validation failed: required sheets missing", file=sys.stderr)
            sys.exit(4)
        load_wb.close()
    except Exception as e:
        print(f"XLSX reload failed: {e}", file=sys.stderr)
        sys.exit(5)

    # Emit summary for logs
    rows = len(DATA)
    cols = len(first_seen_keys(DATA))
    print(json.dumps({
        "status": "SUCCESS_LOCAL",
        "rows": rows,
        "cols": cols,
        "output": outpath,
    }))


if __name__ == "__main__":
    main()
