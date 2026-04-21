#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

META_HEADERS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_KEEP_HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}


def find_first_visible_sheet(wb: Workbook):
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") == "visible":
            return ws
    return wb.active


def build_header_index_map(ws):
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value is not None else "")
    index = {h: i + 1 for i, h in enumerate(headers)}
    return headers, index


def copy_columns_by_headers(src_ws, dst_ws, headers_to_copy):
    headers, hmap = build_header_index_map(src_ws)
    dst_col = 1
    max_row = src_ws.max_row
    for h in headers_to_copy:
        if h in hmap:
            col_idx = hmap[h]
            # write header
            dst_ws.cell(row=1, column=dst_col, value=h)
            for r in range(2, max_row + 1):
                val = src_ws.cell(row=r, column=col_idx).value
                dst_ws.cell(row=r, column=dst_col, value=val)
            dst_col += 1


def build_reordered_data(ws, keep_headers):
    headers, hmap = build_header_index_map(ws)
    present_headers = [h for h in keep_headers if h in hmap]
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = []
        for h in present_headers:
            cidx = hmap[h]
            row_vals.append(ws.cell(row=r, column=cidx).value)
        rows.append(row_vals)
    return present_headers, rows


def write_sheet(ws, headers, rows):
    # headers
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    # data
    for r, row_vals in enumerate(rows, start=2):
        for c, v in enumerate(row_vals, start=1):
            ws.cell(row=r, column=c, value=v)


def apply_formatting(ws, headers):
    # Header formatting
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.alignment = header_alignment
        cell.font = header_font

    # Data rows alignment and wrapping
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wrap_cols_idx = set()
    hmap = {h: i + 1 for i, h in enumerate(headers)}
    for h in WRAP_COLUMNS:
        if h in hmap:
            wrap_cols_idx.add(hmap[h])

    index_col = hmap.get("Index")

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            # vertical top for data rows
            align_kwargs = {"vertical": "top"}
            if c == index_col:
                align_kwargs.update({"horizontal": "center"})
            else:
                align_kwargs.update({"horizontal": "left"})
            if c in wrap_cols_idx:
                align_kwargs.update({"wrap_text": True})
            cell.alignment = Alignment(**align_kwargs)
            cell.border = border

    # Borders for header row as well
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).border = border

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto filter across used range
    ws.auto_filter.ref = ws.dimensions

    # Autofit column widths (approximation using text length)
    # Limit min and max widths
    min_w, max_w = 10, 80
    for idx, h in enumerate(headers, start=1):
        max_len = len(str(h)) if h is not None else 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx).value
            ln = len(str(v)) if v is not None else 0
            if ln > max_len:
                max_len = ln
        adj_w = min(max(min_w, max_len + 2), max_w)
        ws.column_dimensions[get_column_letter(idx)].width = adj_w


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="Path to the target Excel .xlsx file")
    args = ap.parse_args()

    xlsx_path = args.file

    wb = load_workbook(filename=xlsx_path)

    # STEP 2: Identify primary visible sheet as main
    main_ws = find_first_visible_sheet(wb)

    # STEP 3: Create META sheet and copy META columns
    meta_ws = wb.create_sheet("Meta_data_sheet")
    copy_columns_by_headers(main_ws, meta_ws, META_HEADERS)

    # STEP 4: Very hide META sheet
    meta_ws.sheet_state = "veryHidden"

    # STEP 5: Build reordered data for MAIN keep headers
    keep_headers_present, rows = build_reordered_data(main_ws, MAIN_KEEP_HEADERS)

    # Create a temporary sheet to hold reordered data
    tmp_ws = wb.create_sheet("TestPlan_TMP")
    write_sheet(tmp_ws, keep_headers_present, rows)

    # Remove the original main worksheet
    wb.remove(main_ws)

    # Rename TMP to TestPlan
    tmp_ws.title = "TestPlan"

    # STEP 5A: Apply formatting ONLY to TestPlan
    apply_formatting(tmp_ws, keep_headers_present)

    # STEP 6: Save back to the same file (overwrite)
    wb.save(xlsx_path)
