#!/usr/bin/env python3
# Deterministic Excel generation from embedded JSON using openpyxl
# Fulfills strict formatting, numbering, validation, and sheet visibility rules

import json
import os
import re
import sys
import zipfile
from copy import deepcopy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# --- Embedded JSON data (treated as a single-element array) ---
DATA = [
    {
        "Index": 0,
        "SS / Module": "GPIO",
        "Feature": "Writing 1 to iclr clears the interrupt raw status of respective GPIO pin",
        "Test Case Name": "test_gpio_nedge_random_pads_en",
        "Test Description": "Tests negative-edge interrupt handling on GPIO pads by enabling group interrupts, triggering an edge, and verifying status and clear behavior.",
        "Speed": "NA",
        "Mode": "ISR",
        "Memory Start Offset": "NA",
        "Memory End Offset": "NA",
        "Remarks": "Uses GPIO pads 8 to 39. Operates in interrupt mode on negative edge. Requires system interrupt enable.",
        "Test Steps / Procedure": "1) Enable INTR_EN1 for the relevant controller\n2) Configure GPIO_8 register for input and negative-edge behavior for the selected pad\n3) Enable the selected pad bit in INTR1_INTR_EN1\n4) Trigger a pad transition and wait for the interrupt to occur\n5) Read GPIO_8 to check input state and raw status\n6) Clear the pad interrupt in GPIO_8 and verify the cleared value\n7) Read INTR1_INTR_STS1 to confirm the bit was set and then cleared\n8) Clear system raw status in G10_intr_raw_stclr and verify it is cleared",
        "Impacted Registers": "INTR_EN1, GPIO_8, INTR1_INTR_EN1, INTR1_INTR_STS1, G10_intr_raw_stclr",
        "Validation / Acceptance Criteria": "1) Input state indicates expected transition → The value reflects the active level\n2) Raw status bit in the GPIO register is set on interrupt → Interrupt is detected\n3) Group status bit is set for the selected pad → Group interrupt is detected\n4) After clearing in the GPIO register → The GPIO status reflects cleared state\n5) Group status register becomes zero → Group interrupt is cleared\n6) System raw status indicates cleared → System status is cleared",
        "Code Generation (Required / Not)": "",
        "Hidden_Test_Case_Name": "test_gpio_nedge_random_pads_en",
        "Hidden_Test_Description": "Enables sysreg interrupt for GPIO0 or GPIO1. For enabling input mode and negedge interrupt for GPIOs 8-39. Randomly selects unique pad numbers, enables the gpio group interrupt, writes into sram location at 0xA0243ffc to drive the corresponding bit low then high, waits for the interrupt, and in the Default_IRQHandler checks DIN and raw interrupt status, verifies group interrupt status, clears the interrupt by setting the 16th bit of the GPIO register, and clears system raw status.",
        "Hidden_Remarks": "For enabling input mode and negedge interrupt for GPIOs 8-39. Clearing the interrupt raw status bit (16th bit of GPIO reg set to '1'). Writing into sram location.",
        "Hidden_Test_Steps_Procedure": "Initialization: If GPIO0 is defined, call GIC_EnableIRQ(87) and write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR). If GPIO1 is defined, call GIC_EnableIRQ(88) and write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR). Seed RNG with srand(time(NULL)). Write 0xFFFFFFFF to 0xA0243ffc. For i = 0..31: pick pad_num = rand() % 32; ensure uniqueness against arr[]; set wr_val = (1 << pad_num); write_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4), 0x00140000) to configure input mode and negedge; wait_on(50); write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, (1 << pad_num)) to enable group interrupt; wait_on(10); write_reg(0xA0243ffc, ~wr_val); wait_on(10); write_reg(0xA0243ffc, 0xFFFFFFFF); set int_pend = 1; while(int_pend == 1) { print \"Waiting for interrupt\"; wait_on(10); }. After loop, call finish(test_err). Default_IRQHandler: set wr_val = (1 << pad_num); set int_pend = 0; rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4)); if ((rdata & 0x1) != 0) { success log } else { print error; test_err++; }. If ((rdata & 0x2) != 0) { rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if ((rdata_grp & (1 << pad_num)) != 0) { success log } else { print \"ERROR: Group Interrupt not occured\"; test_err++; }. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000); write_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4), 0x00110001); wait_on(2); rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4)); if (rdata == 0x00100001) { success log } else { print error; test_err++; }. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if (rdata_grp == 0x0) { success log } else { print error; test_err++; }. If GPIO0 defined: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0) { success log } else { print error; test_err++; }. If GPIO1 defined: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0) { success log } else { print error; test_err++; }. } else { print \"Interrupt Not occured\"; test_err++; }. Finally, clear GIC: if GPIO0 defined GIC_ClearIRQ(87); if GPIO1 defined GIC_ClearIRQ(88).",
        "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR, LSS_SYSREG_INTR_EN1_GPIO1_INTR, MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR, LSS_SYSREG_RAW_STCR1_GPIO1_INTR",
        "Hidden_Validation_Acceptance_Criteria": "1) If (rdata & 0x1) != 0 in GPIO pad register during negedge then input state matches pad value; else increment test_err. 2) If (rdata & 0x2) != 0 then raw interrupt status is set; else print \"Interrupt Not occured\" and increment test_err. 3) If (read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) & (1<<pad_num)) != 0 then group interrupt is raised; else increment test_err. 4) After write_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4), 0x00110001) and wait, if readback equals 0x00100001 then interrupt clear is successful; else increment test_err. 5) If read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) == 0x0 then group interrupt is cleared; else increment test_err. 6) After writing LSS_SYSREG_RAW_STCR1_GPIO0_INTR or LSS_SYSREG_RAW_STCR1_GPIO1_INTR to MIZAR_LSS_SYSREG_RAW_STCR1, if corresponding bit reads back as 0 then system raw status is cleared; else increment test_err."
    }
]

# Constants
OUTPUT_PATH = os.path.join("Test_Output", "GPIO", "TestPlan", "testing.xlsx")
MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]
META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]
WRAP_COLS = [
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
]

BLUE_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP_CENTER = Alignment(horizontal="center", vertical="top")

BORDER_THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def fail(msg: str):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def validate_json(data):
    if not isinstance(data, list):
        fail("JSON root must be an array of objects")
    if len(data) == 0:
        fail("JSON array is empty")
    for i, row in enumerate(data):
        if not isinstance(row, dict):
            fail(f"JSON element at index {i} is not an object")


def union_keys_preserve_order(rows):
    seen = []
    seen_set = set()
    for row in rows:
        for k in row.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def split_items(text: str):
    if text is None:
        return []
    # Normalize newlines
    t = str(text).replace("\r\n", "\n").replace("\r", "\n")
    # First split by newline; then further split by semicolons when no newlines
    parts = []
    if "\n" in t:
        parts = [p.strip() for p in t.split("\n")]
    else:
        parts = [p.strip() for p in re.split(r";", t)]
    # Remove empties and leading bullets/numbers
    cleaned = []
    for p in parts:
        if not p:
            continue
        # Remove any leading bullet or numeric marker like '-', '•', '1)', '1.', etc.
        p = re.sub(r"^\s*[\-\u2022]*\s*(\d+[\.)]\s*)?", "", p)
        cleaned.append(p)
    return cleaned


def number_items(text: str):
    items = split_items(text)
    if not items:
        return text
    numbered = [f"{i+1}. {item}" for i, item in enumerate(items)]
    return "\n".join(numbered)


def apply_header_format(ws, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL


def apply_borders(ws, max_row, max_col):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER_THIN


def set_column_widths(ws, headers, data_rows):
    # Estimate width based on the longest line length in header or any cell in the column
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in data_rows:
            val = row.get(header, "")
            if val is None:
                val = ""
            for line in str(val).split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        width = min(80, max(10, int(max_len * 1.1) + 2))
        col_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col_letter].width = width


def set_row_heights(ws, wrap_cols_indexes):
    # Approximate row height based on the max number of lines across wrapped columns
    for r in range(2, ws.max_row + 1):
        max_lines = 1
        for c in wrap_cols_indexes:
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            lines = str(val).count("\n") + 1
            if lines > max_lines:
                max_lines = lines
        # Base height ~15 per line
        ws.row_dimensions[r].height = max(15, min(15 * max_lines, 409))


def main():
    # STEP 1: Validate JSON input
    validate_json(DATA)

    # STEP 2: Normalize schema (union of keys, first-seen order)
    all_keys = union_keys_preserve_order(DATA)

    # STEP 3: Create workbook and staging sheet 'Data'
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Write headers preserving order
    for c, key in enumerate(all_keys, start=1):
        ws.cell(row=1, column=c, value=key)

    # Write rows preserving exact values
    for r, obj in enumerate(DATA, start=2):
        for c, key in enumerate(all_keys, start=1):
            ws.cell(row=r, column=c, value=obj.get(key, None))

    # Base header formatting
    apply_header_format(ws, len(all_keys))

    # Base width auto-fit on staging
    set_column_widths(ws, all_keys, DATA)

    # STEP 5: Create Meta_data_sheet and copy META columns AS-IS
    meta_ws = wb.create_sheet("Meta_data_sheet")
    for c, key in enumerate(META_COLS, start=1):
        meta_ws.cell(row=1, column=c, value=key)
    for r, obj in enumerate(DATA, start=2):
        for c, key in enumerate(META_COLS, start=1):
            meta_ws.cell(row=r, column=c, value=obj.get(key, None))
    # STEP 6: Very Hidden meta sheet
    meta_ws.sheet_state = "veryHidden"

    # STEP 7: Rename 'Data' to 'TestPlan' and normalize in place
    ws.title = "TestPlan"

    # Build a lookup of row objects for in-place rewrite
    rows_objects = [deepcopy(obj) for obj in DATA]

    # Number items inside specified columns for TestPlan only (Meta retains raw)
    numbered_rows_objects = []
    for obj in rows_objects:
        new_obj = deepcopy(obj)
        for col in ["Test Steps / Procedure", "Validation / Acceptance Criteria"]:
            if col in new_obj and isinstance(new_obj[col], (str, int, float)):
                new_obj[col] = number_items(str(new_obj[col]))
        numbered_rows_objects.append(new_obj)

    # Rewrite header to MAIN_ORDER and rows accordingly, creating blanks where missing
    for c, key in enumerate(MAIN_ORDER, start=1):
        ws.cell(row=1, column=c, value=key)
    # Clear any extra columns beyond MAIN_ORDER after rewrite
    final_cols = len(MAIN_ORDER)

    for r, obj in enumerate(numbered_rows_objects, start=2):
        for c, key in enumerate(MAIN_ORDER, start=1):
            ws.cell(row=r, column=c, value=obj.get(key, None))

    # Delete any columns to the right of final order
    while ws.max_column > final_cols:
        ws.delete_cols(final_cols + 1)

    # Strict formatting on TestPlan
    apply_header_format(ws, final_cols)

    # Data rows alignment and wrapping
    wrap_indexes = []
    for c, key in enumerate(MAIN_ORDER, start=1):
        if key in WRAP_COLS:
            wrap_indexes.append(c)
    for r in range(2, ws.max_row + 1):
        for c in range(1, final_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c in wrap_indexes:
                cell.alignment = TOP_LEFT
            elif MAIN_ORDER[c - 1] == "Index":
                cell.alignment = TOP_CENTER
            else:
                # Default text cells left, top
                cell.alignment = TOP_LEFT

    # Borders on all populated cells in used range
    apply_borders(ws, ws.max_row, final_cols)

    # Auto-fit widths and row heights after wrapping
    set_column_widths(ws, MAIN_ORDER, numbered_rows_objects)
    set_row_heights(ws, wrap_indexes)

    # Data validation on Code Generation (Required / Not) for data rows only
    if "Code Generation (Required / Not)" in MAIN_ORDER:
        cg_col_index = MAIN_ORDER.index("Code Generation (Required / Not)") + 1
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showErrorMessage=True)
        start_cell = ws.cell(row=2, column=cg_col_index).coordinate
        end_cell = ws.cell(row=max(2, ws.max_row), column=cg_col_index).coordinate
        dv_range = f"{start_cell}:{end_cell}"
        dv.add(dv_range)
        ws.add_data_validation(dv)

    # STEP 7B: Enforce final sheet visibility and existence
    names = [s.title for s in wb.worksheets]
    if names.count("TestPlan") != 1 or names.count("Meta_data_sheet") != 1:
        fail("Workbook must contain exactly: TestPlan (visible) and Meta_data_sheet (Very Hidden)")

    # Ensure no sheet named 'Data' remains
    for s in wb.worksheets:
        if s.title == "Data":
            fail("Sheet named 'Data' exists after normalization")

    # STEP 8: Save to path
    out_path = OUTPUT_PATH
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

    # Validate as ZIP-based OOXML and by loading
    if not zipfile.is_zipfile(out_path):
        fail("Output .xlsx is not a valid ZIP-based OOXML file")
    try:
        _ = load_workbook(out_path, read_only=True, data_only=True)
    except Exception as e:
        fail(f"openpyxl reload validation failed: {e}")

    print("SUCCESS: Excel generated and validated:", out_path)


if __name__ == "__main__":
    main()
