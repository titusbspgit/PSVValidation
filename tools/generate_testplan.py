import json, os, re
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generate_workbook(rows):
    # rows: list of dicts; preserve key order from first row
    first = rows[0]
    testplan_cols = [k for k in first.keys() if not k.startswith("Meta ")]
    metadata_cols = [k for k in first.keys() if k.startswith("Meta ")]

    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"
    ws_meta = wb.create_sheet("MetaData")
    ws_meta.sheet_state = 'veryHidden'

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Headers for TestPlan
    for col_idx, key in enumerate(testplan_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill

    # Rows for TestPlan
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, key in enumerate(testplan_cols, start=1):
            v = item.get(key, "")
            ws.cell(row=r_idx, column=c_idx, value=v).alignment = wrap

    ws.freeze_panes = "A2"

    # Reasonable column widths
    width_defaults = 18
    long_fields = {
        "Test Description",
        "Meta Test Description",
        "Test Steps / Procedure",
        "Meta Test Steps / Procedure",
        "Remarks",
        "Validation / Acceptance Criteria",
        "Meta Validation / Acceptance Criteria",
        "Meta Arrays",
        "Meta Headers",
        "Meta Macros",
    }
    for c_idx, key in enumerate(testplan_cols, start=1):
        width = 60 if key in long_fields else width_defaults
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # Headers for MetaData
    for col_idx, key in enumerate(metadata_cols, start=1):
        cell = ws_meta.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill

    # Rows for MetaData
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, key in enumerate(metadata_cols, start=1):
            v = item.get(key, "")
            ws_meta.cell(row=r_idx, column=c_idx, value=v).alignment = wrap

    ws_meta.freeze_panes = "A2"
    for c_idx, key in enumerate(metadata_cols, start=1):
        width = 60 if key in long_fields else width_defaults
        ws_meta.column_dimensions[get_column_letter(c_idx)].width = width

    return wb


def main():
    out_dir = os.path.join("Test_Output", "GPIO", "TestPlan")
    os.makedirs(out_dir, exist_ok=True)

    with open(os.path.join("data", "final_json.json"), "r", encoding="utf-8") as f:
        rows = json.load(f)

    wb = generate_workbook(rows)

    # Determine IST timestamp (from file if present, else current IST)
    ts = None
    ts_file = os.path.join("data", "timestamp_ist.txt")
    if os.path.exists(ts_file):
        with open(ts_file, "r", encoding="utf-8") as tf:
            t = tf.read().strip()
            if re.fullmatch(r"\d{8}_\d{6}", t):
                ts = t
    if ts is None:
        ist_now = datetime.now(ZoneInfo("Asia/Kolkata"))
        ts = ist_now.strftime("%Y%m%d_%H%M%S")

    filename = f"GPIO_TestPlan_{ts}.xlsx"
    out_path = os.path.join(out_dir, filename)

    wb.save(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
