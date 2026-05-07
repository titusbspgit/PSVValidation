#!/usr/bin/env python3
import argparse, json, os, re, zipfile
from copy import deepcopy
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]

def load_json_array(path):
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, dict):
        # normalize TC1/TC2 -> array
        data = [v for k, v in sorted(data.items())]
    if not isinstance(data, list) or not data:
        raise SystemExit('JSON must be a non-empty array')
    # ensure dicts
    for i, v in enumerate(data):
        if not isinstance(v, dict):
            raise SystemExit(f'JSON element at {i} is not an object')
    return data


def read_existing_excel(path):
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    rows = []
    if 'TestPlan' not in wb.sheetnames:
        return rows
    ws = wb['TestPlan']
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        obj = {}
        for k, v in zip(header, r):
            if k is None:
                continue
            obj[str(k)] = v if v is not None else ''
        rows.append(obj)
    # Merge META if available
    if 'Meta_data_sheet' in wb.sheetnames:
        ms = wb['Meta_data_sheet']
        mheader = [c.value for c in next(ms.iter_rows(min_row=1, max_row=1))]
        meta_rows = []
        for r in ms.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in r):
                continue
            mobj = {}
            for k, v in zip(mheader, r):
                if k is None:
                    continue
                mobj[str(k)] = v if v is not None else ''
            meta_rows.append(mobj)
        by_name = {m.get('Hidden_Test_Case_Name'): m for m in meta_rows if m.get('Hidden_Test_Case_Name')}
        merged = []
        for row in rows:
            key = row.get('Test Case Name')
            m = by_name.get(key, {})
            mr = deepcopy(row)
            mr.update(m)
            merged.append(mr)
        rows = merged
    return rows


def union_keys(rows):
    seen = []
    sset = set()
    for r in rows:
        for k in r.keys():
            if k not in sset:
                sset.add(k)
                seen.append(k)
    return seen


def renumber(rows):
    for i, r in enumerate(rows, 1):
        r['Index'] = i
    return rows


def fill_missing(rows, keys):
    for r in rows:
        for k in keys:
            if k not in r or r[k] is None:
                r[k] = 'NA'
    return rows


def ensure_dirs(path):
    Path(os.path.dirname(path)).mkdir(parents=True, exist_ok=True)


def normalize_numbering(text):
    if text is None or text == '':
        return text
    lines = [ln.strip() for ln in str(text).splitlines() if ln is not None]
    lines = [ln for ln in lines if ln != '']
    out = []
    for i, ln in enumerate(lines, 1):
        ln2 = re.sub(r'^\s*(\d+[\)\.:\-])\s*', '', ln)
        out.append(f"{i}. {ln2}")
    return "\n".join(out)


def write_workbook(final_rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Stage: write union keys
    keys = union_keys(final_rows)
    # Ensure MAIN and META keys exist even if missing
    for col in MAIN_ORDER + META_COLS:
        if col not in keys:
            keys.append(col)

    # Header
    for ci, k in enumerate(keys, 1):
        ws.cell(row=1, column=ci, value=k)

    # Data rows (raw)
    for ri, row in enumerate(final_rows, 2):
        for ci, k in enumerate(keys, 1):
            ws.cell(row=ri, column=ci, value=row.get(k, ''))

    # Create META sheet with META_COLS as-is
    ms = wb.create_sheet('Meta_data_sheet')
    for ci, k in enumerate(META_COLS, 1):
        ms.cell(row=1, column=ci, value=k)
    for ri, row in enumerate(final_rows, 2):
        for ci, k in enumerate(META_COLS, 1):
            ms.cell(row=ri, column=ci, value=row.get(k, ''))
    ms.sheet_state = 'veryHidden'

    # Transform main sheet in-place: remove META cols and reorder to MAIN_ORDER
    # Read back current rows
    data_rows = []
    header_map = {k: idx+1 for idx, k in enumerate(keys)}
    nrows = ws.max_row
    for r in range(2, nrows+1):
        obj = {k: ws.cell(row=r, column=header_map[k]).value for k in keys}
        data_rows.append(obj)

    # Apply numbering formatting for two fields (visible sheet only)
    for r in data_rows:
        if 'Test Steps / Procedure' in r and r['Test Steps / Procedure']:
            r['Test Steps / Procedure'] = normalize_numbering(r['Test Steps / Procedure'])
        if 'Validation / Acceptance Criteria' in r and r['Validation / Acceptance Criteria']:
            r['Validation / Acceptance Criteria'] = normalize_numbering(r['Validation / Acceptance Criteria'])

    # Rebuild sheet as TestPlan with MAIN_ORDER only
    ws.title = 'TestPlan'
    ws.delete_rows(1, ws.max_row)
    # Headers
    for ci, k in enumerate(MAIN_ORDER, 1):
        ws.cell(row=1, column=ci, value=k)
    # Rows
    for ri, row in enumerate(data_rows, 2):
        for ci, k in enumerate(MAIN_ORDER, 1):
            ws.cell(row=ri, column=ci, value=row.get(k, ''))

    # Formatting
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='BDD7EE')  # light blue for readability

    # Freeze pane below header
    ws.freeze_panes = 'A2'

    # Apply header style
    max_col = ws.max_column
    max_row = ws.max_row
    for c in range(1, max_col+1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Data rows style
    for r in range(2, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            # left for text by default
            halign = 'left'
            if ws.cell(row=1, column=c).value == 'Index':
                halign = 'center'
            cell.alignment = Alignment(horizontal=halign, vertical='top', wrap_text=(ws.cell(row=1, column=c).value in ['Test Description','Remarks','Test Steps / Procedure','Validation / Acceptance Criteria']))
            cell.border = border

    # Column width autofit approximation
    for c in range(1, max_col+1):
        header = ws.cell(row=1, column=c).value or ''
        maxlen = len(str(header))
        for r in range(2, max_row+1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[chr(64+c)].width = min(maxlen + 2, 80)

    # Data validation for Code Generation (Required / Not)
    if 'Code Generation (Required / Not)' in MAIN_ORDER:
        col_idx = MAIN_ORDER.index('Code Generation (Required / Not)') + 1
        dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True)
        ws.add_data_validation(dv)
        if max_row >= 2:
            dv.add(f"{chr(64+col_idx)}2:{chr(64+col_idx)}{max_row}")

    # Safety: ensure only TestPlan and Meta_data_sheet exist
    for name in list(wb.sheetnames):
        if name not in ('TestPlan', 'Meta_data_sheet'):
            del wb[name]

    ensure_dirs(output_path)
    wb.save(output_path)

    # Validate XLSX container
    with zipfile.ZipFile(output_path, 'r') as zf:
        _ = zf.namelist()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--json', required=True)
    ap.add_argument('--output', required=True)
    args = ap.parse_args()

    input_rows = load_json_array(args.json)

    # Merge existing if present
    existing = read_existing_excel(args.output)
    final = existing + input_rows
    final = renumber(final)

    # Build union keys and fill missing with NA (for data integrity)
    keys = union_keys(final)
    for r in final:
        for k in keys:
            if k not in r or r[k] is None or r[k] == '':
                # Leave empty for optional, but use 'NA' to mark missing
                r.setdefault(k, 'NA')

    write_workbook(final, args.output)

if __name__ == '__main__':
    main()
