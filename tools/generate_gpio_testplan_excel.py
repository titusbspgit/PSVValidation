#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------------------
# Configuration / Inputs
# ---------------------------
OWNER = "titusbspgit"
REPO = "PSVValidation"
BRANCH = "main"
OUTPUT_DIRECTORY = "Test_Output/GPIO/TestPlan/"
IP_NAME = "GPIO"

# json_data embedded exactly as provided (do not alter)
json_text = r'''[
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "AHB 32-bit register interface",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Validates GPIO GP0 register block reset defaults and masked write/read-back behavior using polling. The test iterates over a list of unresolved register macros for addresses, skips entries based on skip arrays, verifies default values (with LSB masked to zero), writes multiple data patterns masked by per-register write masks, and checks read-back values masked by per-register read masks against computed expectations. Addresses are defined by unresolved register macro MIZAR_GPIO_GP0_GPIO_8 through MIZAR_GPIO_GP0_GPIO_27.",
    "Meta Test Description": "Objective: Verify reset default values and masked write/read-back behavior for GPIO GP0 registers using arrays of addresses and masks.\nInitialization: def_fail_cnt = 0; wr_fail_cnt = 0.\nDefault-value phase (chk_rst_val):\n- Loop i = 0..CNT-1:\n  - addr = addr_array[i] (MIZAR_GPIO_GP0_GPIO_8..MIZAR_GPIO_GP0_GPIO_27).\n  - If skip_rst_array[i] == 1: continue.\n  - If read_mask_array[i] == 0x00000000: continue.\n  - data_rd = read_reg(addr).\n  - data = (data_rd & 0xfffffffe).\n  - If data == default_value_array[i]: pass; else { def_fail_cnt++; print failure with addr, default_value_array[i], data, data_rd }.\nWrite/read-back phase (chk_rd_wr):\n- chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}.\n- For each j in 0..5: data_wr = chk_val[j].\n  - Write loop i = 0..CNT-1:\n    - addr = addr_array[i].\n    - If skip_array[i] == 1: continue.\n    - If write_mask_array[i] == 0x00000000: continue.\n    - write_reg(addr, (data_wr & write_mask_array[i])).\n  - Read/verify loop i = 0..CNT-1:\n    - addr = addr_array[i].\n    - If skip_array[i] == 1: continue.\n    - If write_mask_array[i] == 0x00000000: continue.\n    - If read_mask_array[i] == 0x00000000: continue.\n    - data_rd = (read_reg(addr) & read_mask_array[i]).\n    - wr_n = (write_mask_array[i] ^ 0xffffffff).\n    - exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]) ).\n    - If data_rd == exp_val: pass; else { wr_fail_cnt++; print mismatch }.\nCompletion (test_case):\n- chk_rst_val(); chk_rd_wr();\n- If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0).\nNote: soft_reset_chk() exists but is compiled out (#ifdef 0) and uses SOFT_RST_REG_ADDRESS and SOFT_RST_REG_DATA.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "During default value reads, din may become 1 automatically if not forced; forcing zero to din can drive level select high and cause read values to mismatch expected defaults.",
    "Test Steps / Procedure": "1. Initialize the test and counters; prepare address, default, read-mask, and write-mask arrays.\n2. For each entry where reset-read is enabled, read the register at UNRESOLVED(MIZAR_GPIO_GP0_GPIO_x), mask the LSB to zero, and compare with the corresponding default value. Record any mismatches.\n3. For each of six data patterns, write the pattern to every writable register address (UNRESOLVED(MIZAR_GPIO_GP0_GPIO_x)) using the per-register write mask, then read back using the per-register read mask.\n4. For each read-back, compute the expected value by combining the masked write data with masked preserved default bits (based on the inverse write mask) and compare. Record mismatches.\n5. Declare pass if no default mismatches and no write/read mismatches were recorded; otherwise declare fail.",
    "Meta Test Steps / Procedure": "- test_case():\n  - chk_rst_val();\n  - chk_rd_wr();\n  - if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0).\n- chk_rst_val():\n  - for (i = 0; i < CNT; i++):\n    - addr = addr_array[i];\n    - if (skip_rst_array[i] == 1) continue;\n    - if (read_mask_array[i] == 0x00000000) continue;\n    - data_rd = read_reg(addr);\n    - data = (data_rd & 0xfffffffe);\n    - if (data == default_value_array[i]) { /* PASS */ } else { def_fail_cnt++; printf failure with addr, default_value_array[i], data, data_rd; }\n- chk_rd_wr():\n  - unsigned int chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000};\n  - for (j = 0; j < 6; j++):\n    - data_wr = chk_val[j];\n    - // Write pass\n    - for (i = 0; i < CNT; i++):\n      - addr = addr_array[i];\n      - if (skip_array[i] == 1) continue;\n      - if (write_mask_array[i] == 0x00000000) continue;\n      - write_reg(addr, (data_wr & write_mask_array[i]));\n    - // Read/verify pass\n    - for (i = 0; i < CNT; i++):\n      - addr = addr_array[i];\n      - if (skip_array[i] == 1) continue;\n      - if (write_mask_array[i] == 0x00000000) continue;\n      - if (read_mask_array[i] == 0x00000000) continue;\n      - data_rd = (read_reg(addr) & read_mask_array[i]);\n      - wr_n = (write_mask_array[i] ^ 0xffffffff);\n      - exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]));\n      - if (data_rd == exp_val) { /* PASS */ } else { wr_fail_cnt++; printf mismatch with addr, exp_val, data_rd; }\n- soft_reset_chk(): (compiled out with #ifdef 0)\n  - default_value = read_reg(SOFT_RST_REG_ADDRESS);\n  - write_reg(SOFT_RST_REG_ADDRESS, SOFT_RST_REG_DATA);\n  - wait_on(1000);\n  - write_reg(SOFT_RST_REG_ADDRESS, default_value);\n  - wait_on(1000);",
    "Impacted Registers": "UNRESOLVED(MIZAR_GPIO_GP0_GPIO_8); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_9); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_10); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_11); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_12); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_13); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_14); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_15); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_16); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_17); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_18); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_19); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_20); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_21); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_22); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_23); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_24); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_25); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_26); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_27); UNRESOLVED(GPIO_GP0_GPIO_8_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_9_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_10_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_11_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_12_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_13_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_14_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_15_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_16_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_17_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_18_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_19_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_20_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_21_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_22_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_23_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_24_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_25_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_26_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_27_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_8_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_9_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_10_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_11_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_12_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_13_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_14_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_15_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_16_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_17_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_18_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_19_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_20_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_21_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_22_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_23_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_24_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_25_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_26_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_27_READ_MASK; UNRESOLVED(GPIO_GP0_GPIO_8_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_9_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_10_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_11_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_12_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_13_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_14_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_15_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_16_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_17_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_18_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_19_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_20_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_21_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_22_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_23_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_24_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_25_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_26_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_27_WRITE_MASK)",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27; GPIO_GP0_GPIO_8_DEFAULT_VAL; GPIO_GP0_GPIO_9_DEFAULT_VAL; GPIO_GP0_GPIO_10_DEFAULT_VAL; GPIO_GP0_GPIO_11_DEFAULT_VAL; GPIO_GP0_GPIO_12_DEFAULT_VAL; GPIO_GP0_GPIO_13_DEFAULT_VAL; GPIO_GP0_GPIO_14_DEFAULT_VAL; GPIO_GP0_GPIO_15_DEFAULT_VAL; GPIO_GP0_GPIO_16_DEFAULT_VAL; GPIO_GP0_GPIO_17_DEFAULT_VAL; GPIO_GP0_GPIO_18_DEFAULT_VAL; GPIO_GP0_GPIO_19_DEFAULT_VAL; GPIO_GP0_GPIO_20_DEFAULT_VAL; GPIO_GP0_GPIO_21_DEFAULT_VAL; GPIO_GP0_GPIO_22_DEFAULT_VAL; GPIO_GP0_GPIO_23_DEFAULT_VAL; GPIO_GP0_GPIO_24_DEFAULT_VAL; GPIO_GP0_GPIO_25_DEFAULT_VAL; GPIO_GP0_GPIO_26_DEFAULT_VAL; GPIO_GP0_GPIO_27_DEFAULT_VAL; GPIO_GP0_GPIO_8_READ_MASK; GPIO_GP0_GPIO_9_READ_MASK; GPIO_GP0_GPIO_10_READ_MASK; GPIO_GP0_GPIO_11_READ_MASK; GPIO_GP0_GPIO_12_READ_MASK; GPIO_GP0_GPIO_13_READ_MASK; GPIO_GP0_GPIO_14_READ_MASK; GPIO_GP0_GPIO_15_READ_MASK; GPIO_GP0_GPIO_16_READ_MASK; GPIO_GP0_GPIO_17_READ_MASK; GPIO_GP0_GPIO_18_READ_MASK; GPIO_GP0_GPIO_19_READ_MASK; GPIO_GP0_GPIO_20_READ_MASK; GPIO_GP0_GPIO_21_READ_MASK; GPIO_GP0_GPIO_22_READ_MASK; GPIO_GP0_GPIO_23_READ_MASK; GPIO_GP0_GPIO_24_READ_MASK; GPIO_GP0_GPIO_25_READ_MASK; GPIO_GP0_GPIO_26_READ_MASK; GPIO_GP0_GPIO_27_READ_MASK; GPIO_GP0_GPIO_8_WRITE_MASK; GPIO_GP0_GPIO_9_WRITE_MASK; GPIO_GP0_GPIO_10_WRITE_MASK; GPIO_GP0_GPIO_11_WRITE_MASK; GPIO_GP0_GPIO_12_WRITE_MASK; GPIO_GP0_GPIO_13_WRITE_MASK; GPIO_GP0_GPIO_14_WRITE_MASK; GPIO_GP0_GPIO_15_WRITE_MASK; GPIO_GP0_GPIO_16_WRITE_MASK; GPIO_GP0_GPIO_17_WRITE_MASK; GPIO_GP0_GPIO_18_WRITE_MASK; GPIO_GP0_GPIO_19_WRITE_MASK; GPIO_GP0_GPIO_20_WRITE_MASK; GPIO_GP0_GPIO_21_WRITE_MASK; GPIO_GP0_GPIO_22_WRITE_MASK; GPIO_GP0_GPIO_23_WRITE_MASK; GPIO_GP0_GPIO_24_WRITE_MASK; GPIO_GP0_GPIO_25_WRITE_MASK; GPIO_GP0_GPIO_26_WRITE_MASK; GPIO_GP0_GPIO_27_WRITE_MASK",
    "Validation / Acceptance Criteria": "Pass: No default-read mismatches and no write/read-back mismatches across all addressed registers. Default-read check masks the LSB to zero and must equal the expected default value. For each data pattern, the masked read-back must equal the expected value computed from the masked write data combined with preserved default bits (inverse of the write mask). Fail: Any mismatch increments counters and results in finish(1).",
    "Meta Validation / Acceptance Criteria": "- Default check: data_rd = read_reg(addr); data = (data_rd & 0xfffffffe); require data == default_value_array[i]. If false: def_fail_cnt++.\n- Write/read-back check: data_rd = (read_reg(addr) & read_mask_array[i]); wr_n = (write_mask_array[i] ^ 0xffffffff); exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); require data_rd == exp_val. If false: wr_fail_cnt++.\n- Final result: (def_fail_cnt > 0 || wr_fail_cnt > 0) -> finish(1); else finish(0).",
    "Code Generation (Required / Not)": "Not Required",
    "Meta Headers": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<gpio/gpio_def.h>\n#include<gpio/gpio_offset.h>",
    "Meta Macros": "#define CNT 49\n#define SOFT_RST_REG_ADDRESS 0x00000000\n#define SOFT_RST_REG_DATA 0x00000000",
    "Meta Arrays": "const unsigned long int addr_array[20]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,MIZAR_GPIO_GP0_GPIO_10,MIZAR_GPIO_GP0_GPIO_11,MIZAR_GPIO_GP0_GPIO_12,MIZAR_GPIO_GP0_GPIO_13,MIZAR_GPIO_GP0_GPIO_14,MIZAR_GPIO_GP0_GPIO_15,MIZAR_GPIO_GP0_GPIO_16,MIZAR_GPIO_GP0_GPIO_17,MIZAR_GPIO_GP0_GPIO_18,MIZAR_GPIO_GP0_GPIO_19,MIZAR_GPIO_GP0_GPIO_20,MIZAR_GPIO_GP0_GPIO_21,MIZAR_GPIO_GP0_GPIO_22,MIZAR_GPIO_GP0_GPIO_23,MIZAR_GPIO_GP0_GPIO_24,MIZAR_GPIO_GP0_GPIO_25,MIZAR_GPIO_GP0_GPIO_26,MIZAR_GPIO_GP0_GPIO_27,};\nconst unsigned int default_value_array[20]={GPIO_GP0_GPIO_8_DEFAULT_VAL,GPIO_GP0_GPIO_9_DEFAULT_VAL,GPIO_GP0_GPIO_10_DEFAULT_VAL,GPIO_GP0_GPIO_11_DEFAULT_VAL,GPIO_GP0_GPIO_12_DEFAULT_VAL,GPIO_GP0_GPIO_13_DEFAULT_VAL,GPIO_GP0_GPIO_14_DEFAULT_VAL,GPIO_GP0_GPIO_15_DEFAULT_VAL,GPIO_GP0_GPIO_16_DEFAULT_VAL,GPIO_GP0_GPIO_17_DEFAULT_VAL,GPIO_GP0_GPIO_18_DEFAULT_VAL,GPIO_GP0_GPIO_19_DEFAULT_VAL,GPIO_GP0_GPIO_20_DEFAULT_VAL,GPIO_GP0_GPIO_21_DEFAULT_VAL,GPIO_GP0_GPIO_22_DEFAULT_VAL,GPIO_GP0_GPIO_23_DEFAULT_VAL,GPIO_GP0_GPIO_24_DEFAULT_VAL,GPIO_GP0_GPIO_25_DEFAULT_VAL,GPIO_GP0_GPIO_26_DEFAULT_VAL,GPIO_GP0_GPIO_27_DEFAULT_VAL,};\nconst unsigned int read_mask_array[20]={GPIO_GP0_GPIO_8_READ_MASK,GPIO_GP0_GPIO_9_READ_MASK,GPIO_GP0_GPIO_10_READ_MASK,GPIO_GP0_GPIO_11_READ_MASK,GPIO_GP0_GPIO_12_READ_MASK,GPIO_GP0_GPIO_13_READ_MASK,GPIO_GP0_GPIO_14_READ_MASK,GPIO_GP0_GPIO_15_READ_MASK,GPIO_GP0_GPIO_16_READ_MASK,GPIO_GP0_GPIO_17_READ_MASK,GPIO_GP0_GPIO_18_READ_MASK,GPIO_GP0_GPIO_19_READ_MASK,GPIO_GP0_GPIO_20_READ_MASK,GPIO_GP0_GPIO_21_READ_MASK,GPIO_GP0_GPIO_22_READ_MASK,GPIO_GP0_GPIO_23_READ_MASK,GPIO_GP0_GPIO_24_READ_MASK,GPIO_GP0_GPIO_25_READ_MASK,GPIO_GP0_GPIO_26_READ_MASK,GPIO_GP0_GPIO_27_READ_MASK,};\nconst unsigned int write_mask_array[20]={GPIO_GP0_GPIO_8_WRITE_MASK,GPIO_GP0_GPIO_9_WRITE_MASK,GPIO_GP0_GPIO_10_WRITE_MASK,GPIO_GP0_GPIO_11_WRITE_MASK,GPIO_GP0_GPIO_12_WRITE_MASK,GPIO_GP0_GPIO_13_WRITE_MASK,GPIO_GP0_GPIO_14_WRITE_MASK,GPIO_GP0_GPIO_15_WRITE_MASK,GPIO_GP0_GPIO_16_WRITE_MASK,GPIO_GP0_GPIO_17_WRITE_MASK,GPIO_GP0_GPIO_18_WRITE_MASK,GPIO_GP0_GPIO_19_WRITE_MASK,GPIO_GP0_GPIO_20_WRITE_MASK,GPIO_GP0_GPIO_21_WRITE_MASK,GPIO_GP0_GPIO_22_WRITE_MASK,GPIO_GP0_GPIO_23_WRITE_MASK,GPIO_GP0_GPIO_24_WRITE_MASK,GPIO_GP0_GPIO_25_WRITE_MASK,GPIO_GP0_GPIO_26_WRITE_MASK,GPIO_GP0_GPIO_27_WRITE_MASK,};\nconst unsigned int skip_array[20]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,};\nconst unsigned int skip_rst_array[20]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,};\nunsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000};"
  }
]'''

json_data = json.loads(json_text)

# ---------------------------
# Helpers
# ---------------------------

def ist_now_str():
    if ZoneInfo is not None:
        tz = ZoneInfo("Asia/Kolkata")
        return datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    # Fallback without zoneinfo: use naive localtime
    return datetime.now().strftime("%Y%m%d_%H%M%S")

TESTPLAN_HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_HEADERS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")  # Blue
HEADER_FONT = Font(bold=True, color="FFFFFF")  # White
DATA_ALIGN = Alignment(wrap_text=True, vertical="top")

COLUMN_WIDTHS_TESTPLAN = [8, 14, 24, 22, 44, 10, 10, 18, 18, 28, 36, 36, 36, 22]
COLUMN_WIDTHS_METADATA = [8, 22, 44, 42, 42, 42, 36, 36, 36]


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Create MetaData sheet
    ws_meta = wb.create_sheet(title="MetaData")

    # Write headers
    ws.append(TESTPLAN_HEADERS)
    ws_meta.append(METADATA_HEADERS)

    # Style headers
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for cell in ws_meta[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Data rows
    for row in rows:
        ws.append([row.get(h, "") for h in TESTPLAN_HEADERS])
        ws_meta.append([
            row.get(h, "") for h in METADATA_HEADERS
        ])

    # Formatting: wrap text, top align for all data cells
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(TESTPLAN_HEADERS)):
        for c in r:
            c.alignment = DATA_ALIGN
    for r in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row, min_col=1, max_col=len(METADATA_HEADERS)):
        for c in r:
            c.alignment = DATA_ALIGN

    # Column widths
    for i, w in enumerate(COLUMN_WIDTHS_TESTPLAN, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for i, w in enumerate(COLUMN_WIDTHS_METADATA, start=1):
        ws_meta.column_dimensions[chr(64 + i)].width = w

    # Freeze first row
    ws.freeze_panes = "A2"
    ws_meta.freeze_panes = "A2"

    # Auto filters
    ws.auto_filter.ref = ws.dimensions
    ws_meta.auto_filter.ref = ws_meta.dimensions

    # MetaData VeryHidden
    ws_meta.sheet_state = "veryHidden"

    return wb


def validate_workbook(file_path, expected_rows):
    wb2 = load_workbook(file_path)
    assert "TestPlan" in wb2.sheetnames, "Missing TestPlan sheet"
    assert "MetaData" in wb2.sheetnames, "Missing MetaData sheet"
    ws = wb2["TestPlan"]
    ws_meta = wb2["MetaData"]
    # Check headers
    assert [c.value for c in ws[1]] == TESTPLAN_HEADERS, "TestPlan headers mismatch"
    assert [c.value for c in ws_meta[1]] == METADATA_HEADERS, "MetaData headers mismatch"
    # Check row counts (excluding header)
    data_rows = ws.max_row - 1 if ws.max_row else 0
    meta_rows = ws_meta.max_row - 1 if ws_meta.max_row else 0
    assert data_rows == expected_rows, f"TestPlan row count {data_rows} != {expected_rows}"
    assert meta_rows == expected_rows, f"MetaData row count {meta_rows} != {expected_rows}"
    # Check VeryHidden
    assert ws_meta.sheet_state == "veryHidden", "MetaData sheet not VeryHidden"


def main():
    # JSON validation: must be an array
    if not isinstance(json_data, list):
        raise SystemExit("json_data must be a JSON array")

    # Build workbook
    wb = build_workbook(json_data)

    # Ensure output directory
    out_dir = OUTPUT_DIRECTORY
    os.makedirs(out_dir, exist_ok=True)

    # Build filename with IST timestamp and IP_NAME rule
    ts = ist_now_str()
    filename = f"{IP_NAME}_TestPlan_{ts}.xlsx"
    out_path = os.path.join(out_dir, filename)

    # Save workbook
    wb.save(out_path)

    # Validate by reopening
    validate_workbook(out_path, expected_rows=len(json_data))

    # Print path for logs
    print(out_path)


if __name__ == "__main__":
    main()
