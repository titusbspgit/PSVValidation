import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

INPUT_JSON_PATH = 'tools/input_spi_testplan.json'
OUTPUT_DIR = 'Test_Output/SPI/TestPlan'
IP_NAME = 'SPI'

MAIN_COLS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]

WRAP_COLS = set([
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
])

HEADER_FILL = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')  # solid blue
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=False)
DATA_ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
DATA_ALIGN_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=False)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def load_json():
    with open(INPUT_JSON_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, dict) and 'TestCases' in data and isinstance(data['TestCases'], list):
        rows = data['TestCases']
    elif isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        rows = [data]
    else:
        raise SystemExit('Invalid JSON: unsupported root structure')
    return data, rows


def normalize_arrays(value, key):
    if isinstance(value, list):
        if key in ('Test Steps / Procedure', 'Hidden_Test_Steps_Procedure'):
            # Numbered newline-separated list
            return '\n'.join(f"{i+1}. {str(v)}" for i, v in enumerate(value))
        else:
            # Preserve as JSON string for any other arrays
            return json.dumps(value, ensure_ascii=False)
    return value


def build_union_keys(rows):
    keys = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    return keys


def auto_width(ws):
    # approximate column width based on max value length
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in col_cells:
            val = cell.value
            if val is None:
                continue
            s = str(val)
            # consider newlines
            for line in s.split('\n'):
                if len(line) > max_len:
                    max_len = len(line)
        # Add a little padding
        width = min(max(10, max_len + 2), 100)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and str(cell.value) != '':
                cell.border = THIN_BORDER


def adjust_row_heights(ws):
    base_height = 15
    for r in range(2, ws.max_row + 1):
        lines = 1
        for c in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=c).value
            if header in WRAP_COLS:
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                cnt = str(val).count('\n') + 1
                if cnt > lines:
                    lines = cnt
        ws.row_dimensions[r].height = base_height * lines


def create_workbook(data, rows):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = 'Data'

    # Build union keys preserving first appearance
    keys = build_union_keys(rows)

    # Write header (temporary; this sheet will be removed)
    for c, k in enumerate(keys, start=1):
        ws_data.cell(row=1, column=c, value=k)

    # Freeze top row
    ws_data.freeze_panes = 'A2'

    # Write data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, k in enumerate(keys, start=1):
            val = row.get(k, '')
            val = normalize_arrays(val, k)
            ws_data.cell(row=r_idx, column=c_idx, value=val)

    # Create Meta_data_sheet and copy META columns, no formatting per Stage1
    ws_meta = wb.create_sheet('Meta_data_sheet')
    for c, k in enumerate(META_COLS, start=1):
        ws_meta.cell(row=1, column=c, value=k)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, k in enumerate(META_COLS, start=1):
            val = row.get(k, '')
            val = normalize_arrays(val, k)
            ws_meta.cell(row=r_idx, column=c_idx, value=val)

    # Hide Meta sheet (veryHidden)
    ws_meta.sheet_state = 'veryHidden'

    # Prepare TestPlan sheet by filtering/remapping columns from Data
    ws_testplan = wb.create_sheet('TestPlan')
    for c, k in enumerate(MAIN_COLS, start=1):
        cell = ws_testplan.cell(row=1, column=c, value=k)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.fill = HEADER_FILL

    # Map from Data headers to their column indices for quick lookup
    header_to_idx = { ws_data.cell(row=1, column=ci).value: ci for ci in range(1, ws_data.max_column+1) }

    for r in range(2, ws_data.max_row + 1):
        for c, k in enumerate(MAIN_COLS, start=1):
            src_ci = header_to_idx.get(k, None)
            val = ''
            if src_ci is not None:
                val = ws_data.cell(row=r, column=src_ci).value
            ws_testplan.cell(row=r, column=c, value=val)

    # Formatting for TestPlan per strict rules
    ws_testplan.freeze_panes = 'A2'

    # Alignments and wrapping for data rows
    for r in range(2, ws_testplan.max_row + 1):
        for c in range(1, ws_testplan.max_column + 1):
            header = ws_testplan.cell(row=1, column=c).value
            cell = ws_testplan.cell(row=r, column=c)
            if header in WRAP_COLS:
                cell.alignment = DATA_ALIGN_WRAP
            elif header == 'Index':
                cell.alignment = DATA_ALIGN_CENTER
            else:
                cell.alignment = DATA_ALIGN_LEFT

    # Autofit and borders for visible TestPlan only
    auto_width(ws_testplan)
    apply_borders(ws_testplan)
    adjust_row_heights(ws_testplan)

    # Data validation for Code Generation (Required / Not)
    for c in range(1, ws_testplan.max_column + 1):
        if ws_testplan.cell(row=1, column=c).value == 'Code Generation (Required / Not)':
            col_letter = ws_testplan.cell(row=1, column=c).column_letter
            dv = DataValidation(type="list", formula1='"Required,Not Required"', allow_blank=True, showErrorMessage=True)
            ws_testplan.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{ws_testplan.max_row}")
            break

    # Remove the temporary Data sheet
    wb.remove(ws_data)

    return wb


def main():
    data, rows = load_json()
    wb = create_workbook(data, rows)

    # Compute IST timestamp
    now_ist = datetime.now(ZoneInfo('Asia/Kolkata'))
    ts_date = now_ist.strftime('%Y%m%d')
    ts_time = now_ist.strftime('%H%M%S')
    ts_human = now_ist.strftime('%Y-%m-%d %H:%M:%S')

    filename = f"{IP_NAME}_TestPlan_{ts_date}_{ts_time}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, filename)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb.save(out_path)

    # Write helper files for the workflow
    os.makedirs('tools', exist_ok=True)
    with open('tools/generated_filename.txt', 'w', encoding='utf-8') as f:
        f.write(out_path)
    with open('tools/commit_message.txt', 'w', encoding='utf-8') as f:
        f.write(f"Add SPI Test Plan (Excel) generated on {ts_human} IST")

    print(out_path)

if __name__ == '__main__':
    main()
