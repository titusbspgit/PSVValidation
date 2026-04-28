import json, os, sys
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

INPUT_JSON = os.environ.get('INPUT_JSON', 'tools/i2c_testplan_input.json')
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', 'Test_Output/I2C/TestPlan')
IP_NAME = os.environ.get('IP_NAME', 'I2C')
BRANCH = os.environ.get('BRANCH', 'main')
COMMIT_MESSAGE_TEMPLATE = os.environ.get('COMMIT_MESSAGE_TEMPLATE', 'Final formatted Excel generated from JSON input')

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

WRAP_COLS = {
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
}

BLUE_FILL = PatternFill(fill_type='solid', fgColor='2F5597')  # solid blue header
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def normalize_rows(json_data):
    if isinstance(json_data, dict) and 'test_cases' in json_data and isinstance(json_data['test_cases'], list):
        rows = json_data['test_cases']
    elif isinstance(json_data, list):
        rows = json_data
    elif isinstance(json_data, dict):
        rows = [json_data]
    else:
        print('Unsupported JSON structure', file=sys.stderr)
        sys.exit(1)

    # Collect keys in first-seen order across all rows
    seen = []
    seen_set = set()
    for r in rows:
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return rows, seen


def to_cell_value(v):
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def auto_width(ws):
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_len = 0
        for c in col:
            val = '' if c.value is None else str(c.value)
            max_len = max(max_len, len(val))
        width = min(120, max(10, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = THIN_BORDER


def create_meta_sheet(wb, data_rows):
    meta = wb.create_sheet('Meta_data_sheet')
    # header
    for j, h in enumerate(META_COLS, start=1):
        meta.cell(row=1, column=j, value=h)
    # data
    for i, r in enumerate(data_rows, start=2):
        for j, h in enumerate(META_COLS, start=1):
            meta.cell(row=i, column=j, value=to_cell_value(r.get(h, '')))
    # VeryHidden
    meta.sheet_state = 'veryHidden'


def build_testplan_sheet(wb, data_rows):
    ws = wb.active
    ws.title = 'Data'

    # Build union header order
    _, union_keys = normalize_rows({'test_cases': data_rows})

    # write headers
    for j, h in enumerate(union_keys, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # write rows
    for i, r in enumerate(data_rows, start=2):
        for j, h in enumerate(union_keys, start=1):
            ws.cell(row=i, column=j, value=to_cell_value(r.get(h, '')))

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    auto_width(ws)

    # Now create TestPlan with only MAIN_ORDER columns, removing META columns
    tp = wb.create_sheet('TestPlan')
    for j, h in enumerate(MAIN_ORDER, start=1):
        tp.cell(row=1, column=j, value=h)
    for i, r in enumerate(data_rows, start=2):
        for j, h in enumerate(MAIN_ORDER, start=1):
            tp.cell(row=i, column=j, value=to_cell_value(r.get(h, '')))

    # formatting for TestPlan
    # header formatting
    for j in range(1, len(MAIN_ORDER) + 1):
        c = tp.cell(row=1, column=j)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.fill = BLUE_FILL

    # data formatting
    for i in range(2, tp.max_row + 1):
        for j, h in enumerate(MAIN_ORDER, start=1):
            c = tp.cell(row=i, column=j)
            wrap = True if h in WRAP_COLS else False
            halign = 'left'
            if h == 'Index':
                halign = 'center'
            c.alignment = Alignment(horizontal=halign, vertical='top', wrap_text=wrap)

    tp.freeze_panes = 'A2'
    tp.auto_filter.ref = tp.dimensions

    auto_width(tp)
    apply_borders(tp)

    # Data validation for Code Generation (Required / Not)
    try:
        col_idx = MAIN_ORDER.index('Code Generation (Required / Not)') + 1
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
        dv.error = 'Select one of: Required, Blank, Not Required'
        dv.errorTitle = 'Invalid Selection'
        tp.add_data_validation(dv)
        dv.ranges.append(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{tp.max_row}")
    except ValueError:
        pass

    # Remove original Data (we keep a hidden copy as TestPlan already contains required columns)? Requirement says treat Data as main; rename to TestPlan then remove meta; but we recreated. Delete original Data to avoid confusion.
    wb.remove(ws)


def create_metadata_sheet(wb, meta):
    md = wb.create_sheet('Metadata')
    rows = [
        ('ip_name', meta.get('ip_name', '')),
        ('repo_owner', meta.get('repo_owner', '')),
        ('repo_name', meta.get('repo_name', '')),
        ('branch', meta.get('branch', '')),
        ('base_subdirectory', meta.get('base_subdirectory', '')),
        ('generation_timestamp_ist', get_ist_str()),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        md.cell(row=i, column=1, value=k)
        md.cell(row=i, column=2, value=v)

    start_row = len(rows) + 2
    md.cell(row=start_row, column=1, value='source_directories')
    headers = ['name', 'path', 'github_url']
    for j, h in enumerate(headers, start=1):
        md.cell(row=start_row + 1, column=j, value=h)
    for i, d in enumerate(meta.get('source_directories', []), start=start_row + 2):
        md.cell(row=i, column=1, value=d.get('name', ''))
        md.cell(row=i, column=2, value=d.get('path', ''))
        url = d.get('github_url', '')
        cell = md.cell(row=i, column=3, value=url)
        if url:
            cell.hyperlink = url
            cell.style = 'Hyperlink'


def create_traceability_sheet(wb, data_rows):
    tr = wb.create_sheet('Traceability')
    tr.cell(row=1, column=1, value='Test Case Name')
    tr.cell(row=1, column=2, value='source_path')
    r = 2
    for row in data_rows:
        name = row.get('Test Case Name', '')
        paths = row.get('source_paths', [])
        if isinstance(paths, list) and paths:
            for p in paths:
                tr.cell(row=r, column=1, value=name)
                c = tr.cell(row=r, column=2, value=p)
                if p:
                    c.hyperlink = p
                    c.style = 'Hyperlink'
                r += 1
        else:
            tr.cell(row=r, column=1, value=name)
            tr.cell(row=r, column=2, value='')
            r += 1


def get_ist_now():
    return datetime.now(ZoneInfo('Asia/Kolkata'))


def get_ist_str():
    return get_ist_now().strftime('%Y-%m-%d %H:%M:%S')


def main():
    data = load_json(INPUT_JSON)
    # Validate and extract rows
    if isinstance(data, dict) and 'test_cases' in data and isinstance(data['test_cases'], list):
        rows = data['test_cases']
    elif isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        rows = [data]
    else:
        print('FAILURE: Invalid JSON input structure', file=sys.stderr)
        sys.exit(1)

    # Build workbook
    wb = Workbook()
    create_meta_sheet(wb, rows)
    build_testplan_sheet(wb, rows)
    create_metadata_sheet(wb, data.get('metadata', {}))
    create_traceability_sheet(wb, rows)

    # Filename with IST
    ist = get_ist_now()
    fname = f"{IP_NAME}_TestPlan_{ist.strftime('%Y%m%d')}_{ist.strftime('%H%M%S')}.xlsx"

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, fname)
    wb.save(out_path)

    # Commit only the finalized Excel file
    msg = COMMIT_MESSAGE_TEMPLATE.replace('<IST YYYY-MM-DD HH:MM:SS>', ist.strftime('%Y-%m-%d %H:%M:%S'))

    os.system('git config user.name "github-actions[bot]"')
    os.system('git config user.email "41898282+github-actions[bot]@users.noreply.github.com"')
    os.system(f'git add "{out_path}"')
    # Avoid failing if nothing to commit
    rc = os.system(f'git commit -m "{msg}"')
    if rc == 0:
        os.system(f'git push origin {BRANCH}')
        print(f"Committed {out_path}")
    else:
        print('No changes to commit (file may already exist)')

    print(out_path)

if __name__ == '__main__':
    main()
