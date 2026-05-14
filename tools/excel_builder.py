#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
import os
import zipfile
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ====== Embedded JSON input (exactly as provided; no ellipses) ======
JSON_TEXT = r'''[
  {
    "Index": 1,
    "SS / Module": "PCIE0 SII RC",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie0_sii_rc_reg_wr_rd_test",
    "Test Description": "Verify PCIe root complex registers for correct reset values and masked write/read behavior. Ensure read-only bits remain unchanged while writable bits follow written patterns.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Unreadable locations are skipped. Unwritable locations are skipped. A skip list controls access. The reset control register is excluded. The reset helper is not used.",
    "Test Steps / Procedure": "1) Read each SII register reset value except the reset control register and compare it to the documented default.\n2) For six data patterns, write each pattern to all writable registers that are not in the skip list.\n3) For each register that is writable and readable and not skipped, read back and compute the expected value using the read and write masks with the documented default, then compare.\n4) Record any default mismatches or write/read mismatches and decide pass or fail based on counters.",
    "Impacted Registers": "SII_CFG_BAR0_START1\nSII_CFG_BAR0_START2\nSII_CFG_BAR0_LIMIT1\nSII_CFG_BAR0_LIMIT2\nSII_CFG_BAR1_START\nSII_CFG_BAR1_LIMIT1\nSII_CFG_BAR2_START1\nSII_CFG_BAR2_START2\nSII_CFG_BAR2_LIMIT1\nSII_CFG_BAR2_LIMIT2\nSII_CFG_BAR3_START\nSII_CFG_BAR3_LIMIT\nSII_CFG_BAR4_START1\nSII_CFG_BAR4_START2\nSII_CFG_BAR4_LIMIT1\nSII_CFG_BAR4_LIMIT2\nSII_CFG_BAR5_START\nSII_CFG_BAR5_LIMIT\nSII_PCIE0_CONFIG_INFO1\nSII_PCIE0_CONFIG_INFO2\nSII_PCIE0_GEN_CONTROL1\nSII_PCIE0_GEN_CONTROL2\nSII_PCIE0_GEN_CONTROL3\nSII_PCIE0_PM_CONTROL\nSII_PCIE0_CONTROL_PM_STS\nSII_PCIE0_TRANSMIT_HEADER1\nSII_PCIE0_TRANSMIT_HEADER2\nSII_PCIE0_TRANSMIT_HEADER3\nSII_PCIE0_TRANSMIT_HEADER4\nSII_PCIE0_TRANSMIT_REQ\nSII_PCIE0_RCV_MSG_HDR1\nSII_PCIE0_RCV_MSG_HDR2\nSII_PCIE0_RCV_MSG_HDR3\nSII_PCIE0_RCV_MSG_HDR4\nSII_PCIE0_RCV_MSG_STS\nSII_RCV_INTERRPUT_CTRL\nSII_CFG_EXP_ROM_START\nSII_CFG_EXP_ROM_LIMIT\nSII_CFG_EXP_ROM_INFO\nSII_CXPL_DEBUG_INFO1\nSII_CXPL_DEBUG_INFO2\nSII_CXPL_DEBUG_INFO_EI\nSII_PCIE0_TARGET_INFO1\nSII_PCIE0_TARGET_INFO2\nSII_PCIE0_CONTOLLER_ERROR_STATUS\nSII_PCIE0_CONTROLLER_INT_STS\nSII_PCIE0_CONTROLLER_INTERRUPT_CONTROL\nSII_PHY_RST_CONTROL\nSII_LINK_DEBUG_DATA\nSII_PCIE0_ERR_STS\nSII_PCIE0_ERR_INTERRUPT_CTRL\nSII_CFG_MSI_INT\nSII_LTR_MSG\nSII_LTR_MSG_LATENCY\nSII_APP_LTR_LATENCY\nSII_CFG_LTR_MAX_LATENCY\nSII_OBFF_CNTRL\nSII_SLV_AWMISC_INFO\nSII_SLV_AWMISC_INFO_HDR_34DW_HI\nSII_SLV_AWMISC_INFO_HDR_34DW_LO\nSII_SLV_MISC_INFO\nSII_SLV_MISC_RESP_INFO\nSII_MSTR_AWMISC_INFO_CNTRL\nSII_MSTR_AWMISC_INFO_1\nSII_MSTR_AWMISC_INFO_0\nSII_MSTR_AWMISC_INFO_HDR_34DW_HI\nSII_MSTR_AWMISC_INFO_HDR_34DW_LO\nSII_MSTR_ARMISC_INFO_CNTRL\nSII_MSTR_ARMISC_INFO_1\nSII_MSTR_ARMISC_INFO_0\nSII_MSTR_BMISC_RMISC_CPL_STAT_INFO\nSII_RADM_TIMEOUT_INFO\nSII_CFG_MSI_INFO\nSII_CFG_MSI_DATA\nSII_CFG_MSI_ADDR_HI\nSII_CFG_MSI_ADDR_LO\nSII_CFG_AER_INT_AND_PCIE0_CAP_INT_MSG\nSII_RTLH_RFC_DATA\nSII_APP_HDR_INFO\nSII_APP_HDR_LOG_3\nSII_APP_HDR_LOG_2\nSII_APP_HDR_LOG_1\nSII_APP_HDR_LOG_0\nSII_CFG_BUS_NUM\nSII_CFG_BR_CTRL_SERREN\nSII_APP_DEV_AND_BUS_NUM\nSII_PCIE0_CONTROLLER_INT_STS_1\nSII_PCIE0_CONTROLLER_INTERRUPT_CONTROL_1\nSII_APP_AND_SLOT_CONTROL_REG\nSII_DIAG_CTRL_BUS\nSII_CFG_REG_RO\nSII_CFG_ARI_FWD_EN\nSII_RADM_SLOT_PWR_PAYLOAD\nSII_DIAG_STATUS_BUS_0\nSII_DIAG_STATUS_BUS_1\nSII_DIAG_STATUS_BUS_2\nSII_DIAG_STATUS_BUS_3\nSII_DIAG_STATUS_BUS_4\nSII_DIAG_STATUS_BUS_5\nSII_DIAG_STATUS_BUS_6\nSII_DIAG_STATUS_BUS_7\nSII_DIAG_STATUS_BUS_8\nSII_DIAG_STATUS_BUS_9\nSII_DIAG_STATUS_BUS_10\nSII_DIAG_STATUS_BUS_11\nSII_DIAG_STATUS_BUS_12\nSII_DIAG_STATUS_BUS_13\nSII_DIAG_STATUS_BUS_14\nSII_DIAG_STATUS_BUS_15\nSII_DIAG_STATUS_BUS_16\nSII_DIAG_STATUS_BUS_17\nSII_DIAG_STATUS_BUS_18\nSII_DIAG_STATUS_BUS_19\nSII_RAM_PWR_CNTRL_0\nSII_RAM_PWR_CNTRL_1\nSII_SOFT_RESET_CTRL\nSII_CFG_MSI_PENDING_B\nSII_SMLH_LTSSM_STATE_TRAN_1\nSII_SMLH_LTSSM_STATE_TRAN_2\nSII_SMLH_LTSSM_STATE_TRAN_3\nSII_SMLH_LTSSM_STATE_TRAN_4\nSII_SMLH_LTSSM_STATE_TRAN_5\nSII_SMLH_LTSSM_STATE_TRAN_6\nSII_SMLH_LTSSM_STATE_TRAN_7\nSII_PHY_CONTROL_0\nSII_PHY_CONTROL_1\nSII_PHY_CONTROL_2\nSII_PHY_CONTROL_3\nSII_PHY_CONTROL_4\nSII_PHY_CONTROL_5\nSII_PHY_CONTROL_6\nSII_PHY_CONTROL_7\nSII_PHY_CONTROL_8\nSII_PHY_CONTROL_9\nSII_PHY_CONTROL_10\nSII_PHY_CONTROL_11\nSII_PHY_CONTROL_12\nSII_PHY_CONTROL_13\nSII_PHY_CONTROL_14\nSII_PHY_CONTROL_15\nSII_PHY_CONTROL_16\nSII_PHY_CONTROL_17\nSII_PHY_CONTROL_18\nSII_PHY_CONTROL_19\nSII_PHY_CONTROL_20\nSII_PHY_CONTROL_21\nSII_PHY_CONTROL_22\nSII_PHY_CONTROL_23\nSII_PHY_CONTROL_24\nSII_PHY_CONTROL_25\nSII_PHY_CONTROL_26\nSII_MSI_CTRL_IO\nSII_MSI_CTRL_INT_VEC",
    "Validation / Acceptance Criteria": "1) For each readable register (excluding SII_PHY_RST_CONTROL) → The reset value must match the documented default.\n2) For each writable and readable register not skipped after each pattern write → The read-back must equal the value computed using the read mask and write mask with preserved defaulted bits.\n3) Final result → Zero default mismatches and zero write/read mismatches indicate pass; any mismatch indicates fail.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie0_sii_rc_reg_wr_rd_test",
    "Hidden_Test_Description": "The test calls chk_rst_val() then chk_rd_wr(); if def_fail_cnt>0 or wr_fail_cnt>0 it finishes with failure, else success. chk_rst_val(): loops i=0..CNT-1, addr=addr_array[i]; if read_mask_array[i]==0x00000000, optionally prints and continues; if addr_array[i]==mizar_PCIE0_SII_PHY_RST_CONTROL, continue; else data_rd=read_reg(addr); compare data_rd with default_value_array[i]; on match, optionally prints under DEBUG_DISPLAY; on mismatch, increments def_fail_cnt and prints failure. chk_rd_wr(): defines chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}; for each j, sets data_wr=chk_val[j]; writing phase: for i=0..CNT-1, addr=addr_array[i]; if skip_array[i]==1, optionally prints and continues; if write_mask_array[i]==0x00000000, optionally prints and continues; else write_reg(addr,data_wr) and optionally prints; reading/verify phase: for i=0..CNT-1, addr=addr_array[i]; if skip_array[i]==1, continue; if write_mask_array[i]==0x00000000, continue; if read_mask_array[i]==0x00000000, continue; else data_rd=read_reg(addr); wr_n=(write_mask_array[i]^0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if data_rd==exp_val, optionally prints pass; else wr_fail_cnt++ and prints failure. soft_reset_chk() exists (reads SOFT_RST_REG_ADDRESS, writes SOFT_RST_REG_DATA, delays, and restores default) but is not invoked.",
    "Hidden_Remarks": "Addresses with read_mask_array[i]==0x00000000 are skipped as not readable. Addresses with write_mask_array[i]==0x00000000 are skipped for writing and verification. Entries where skip_array[i]==1 are skipped for both write and read. The address equal to mizar_PCIE0_SII_PHY_RST_CONTROL is skipped during default-value checking. The soft reset function is present but commented out in test_case().",
    "Hidden_Test_Steps_Procedure": "In test_case():\n- Call chk_rst_val(). If DEBUG_DISPLAY, print \"********* Default value check end ************\".\n- Call chk_rd_wr(). If DEBUG_DISPLAY, print \"********* Write & Read from registers end ************\".\n- //soft_reset_chk(); is commented out.\n- If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0).\n\nIn chk_rst_val():\n- For i=0..(CNT-1):\n  - addr = addr_array[i].\n  - If (read_mask_array[i] == 0x00000000): if DEBUG_DISPLAY print \"RST : This address 0x%x is not readable, hence skipped for reading\" and continue.\n  - If (addr_array[i] == mizar_PCIE0_SII_PHY_RST_CONTROL): continue.\n  - data_rd = read_reg(addr).\n  - If (data_rd == default_value_array[i]): if DEBUG_DISPLAY print \"RST : PASS Reading Default value from Address :0x%x Expected : 0x%x\\tRead_data : 0x%x\"; else def_fail_cnt++ and print \"RST : Failed Default value mismatch Addr :0x%x Expected : 0x%x\\tRead_data : 0x%x\".\n\nIn chk_rd_wr():\n- int chk_val[6] = {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}.\n- For j=0..5:\n  - data_wr = chk_val[j].\n  - Write loop (i=0..CNT-1):\n    - addr = addr_array[i].\n    - If (skip_array[i] == 1): if DEBUG_DISPLAY print skip message and continue.\n    - If (write_mask_array[i] == 0x00000000): if DEBUG_DISPLAY print not-writable message and continue.\n    - Else write_reg(addr,data_wr) and optionally print written address/data.\n  - Read/verify loop (i=0..CNT-1):\n    - addr = addr_array[i].\n    - If (skip_array[i] == 1) continue.\n    - If (write_mask_array[i] == 0x00000000) continue.\n    - If (read_mask_array[i] == 0x00000000) continue.\n    - data_rd = read_reg(addr).\n    - wr_n = (write_mask_array[i] ^ 0xffffffff).\n    - exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])).\n    - If (data_rd == exp_val): if DEBUG_DISPLAY print PASS; else wr_fail_cnt++ and print mismatch.",
    "Hidden_Impacted_Registers": "mizar_PCIE0_SII_CFG_BAR0_START1\nmizar_PCIE0_SII_CFG_BAR0_START2\nmizar_PCIE0_SII_CFG_BAR0_LIMIT1\nmizar_PCIE0_SII_CFG_BAR0_LIMIT2\nmizar_PCIE0_SII_CFG_BAR1_START\nmizar_PCIE0_SII_CFG_BAR1_LIMIT1\nmizar_PCIE0_SII_CFG_BAR2_START1\nmizar_PCIE0_SII_CFG_BAR2_START2\nmizar_PCIE0_SII_CFG_BAR2_LIMIT1\nmizar_PCIE0_SII_CFG_BAR2_LIMIT2\nmizar_PCIE0_SII_CFG_BAR3_START\nmizar_PCIE0_SII_CFG_BAR3_LIMIT\nmizar_PCIE0_SII_CFG_BAR4_START1\nmizar_PCIE0_SII_CFG_BAR4_START2\nmizar_PCIE0_SII_CFG_BAR4_LIMIT1\nmizar_PCIE0_SII_CFG_BAR4_LIMIT2\nmizar_PCIE0_SII_CFG_BAR5_START\nmizar_PCIE0_SII_CFG_BAR5_LIMIT\nmizar_PCIE0_SII_PCIE0_CONFIG_INFO1\nmizar_PCIE0_SII_PCIE0_CONFIG_INFO2\nmizar_PCIE0_SII_PCIE0_GEN_CONTROL1\nmizar_PCIE0_SII_PCIE0_GEN_CONTROL2\nmizar_PCIE0_SII_PCIE0_GEN_CONTROL3\nmizar_PCIE0_SII_PCIE0_PM_CONTROL\nmizar_PCIE0_SII_PCIE0_CONTROL_PM_STS\nmizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER1\nmizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2\nmizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3\nmizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER4\nmizar_PCIE0_SII_PCIE0_TRANSMIT_REQ\nmizar_PCIE0_SII_PCIE0_RCV_MSG_HDR1\nmizar_PCIE0_SII_PCIE0_RCV_MSG_HDR2\nmizar_PCIE0_SII_PCIE0_RCV_MSG_HDR3\nmizar_PCIE0_SII_PCIE0_RCV_MSG_HDR4\nmizar_PCIE0_SII_PCIE0_RCV_MSG_STS\nmizar_PCIE0_SII_RCV_INTERRPUT_CTRL\nmizar_PCIE0_SII_CFG_EXP_ROM_START\nmizar_PCIE0_SII_CFG_EXP_ROM_LIMIT\nmizar_PCIE0_SII_CFG_EXP_ROM_INFO\nmizar_PCIE0_SII_CXPL_DEBUG_INFO1\nmizar_PCIE0_SII_CXPL_DEBUG_INFO2\nmizar_PCIE0_SII_CXPL_DEBUG_INFO_EI\nmizar_PCIE0_SII_PCIE0_TARGET_INFO1\nmizar_PCIE0_SII_PCIE0_TARGET_INFO2\nmizar_PCIE0_SII_PCIE0_CONTOLLER_ERROR_STATUS\nmizar_PCIE0_SII_PCIE0_CONTROLLER_INT_STS\nmizar_PCIE0_SII_PCIE0_CONTROLLER_INTERRUPT_CONTROL\nmizar_PCIE0_SII_PHY_RST_CONTROL\nmizar_PCIE0_SII_LINK_DEBUG_DATA\nmizar_PCIE0_SII_PCIE0_ERR_STS\nmizar_PCIE0_SII_PCIE0_ERR_INTERRUPT_CTRL\nmizar_PCIE0_SII_CFG_MSI_INT\nmizar_PCIE0_SII_LTR_MSG\nmizar_PCIE0_SII_LTR_MSG_LATENCY\nmizar_PCIE0_SII_APP_LTR_LATENCY\nmizar_PCIE0_SII_CFG_LTR_MAX_LATENCY\nmizar_PCIE0_SII_OBFF_CNTRL\nmizar_PCIE0_SII_SLV_AWMISC_INFO\nmizar_PCIE0_SII_SLV_AWMISC_INFO_HDR_34DW_HI\nmizar_PCIE0_SII_SLV_AWMISC_INFO_HDR_34DW_LO\nmizar_PCIE0_SII_SLV_MISC_INFO\nmizar_PCIE0_SII_SLV_MISC_RESP_INFO\nmizar_PCIE0_SII_MSTR_AWMISC_INFO_CNTRL\nmizar_PCIE0_SII_MSTR_AWMISC_INFO_1\nmizar_PCIE0_SII_MSTR_AWMISC_INFO_0\nmizar_PCIE0_SII_MSTR_AWMISC_INFO_HDR_34DW_HI\nmizar_PCIE0_SII_MSTR_AWMISC_INFO_HDR_34DW_LO\nmizar_PCIE0_SII_MSTR_ARMISC_INFO_CNTRL\nmizar_PCIE0_SII_MSTR_ARMISC_INFO_1\nmizar_PCIE0_SII_MSTR_ARMISC_INFO_0\nmizar_PCIE0_SII_MSTR_BMISC_RMISC_CPL_STAT_INFO\nmizar_PCIE0_SII_RADM_TIMEOUT_INFO\nmizar_PCIE0_SII_CFG_MSI_INFO\nmizar_PCIE0_SII_CFG_MSI_DATA\nmizar_PCIE0_SII_CFG_MSI_ADDR_HI\nmizar_PCIE0_SII_CFG_MSI_ADDR_LO\nmizar_PCIE0_SII_CFG_AER_INT_AND_PCIE0_CAP_INT_MSG\nmizar_PCIE0_SII_RTLH_RFC_DATA\nmizar_PCIE0_SII_APP_HDR_INFO\nmizar_PCIE0_SII_APP_HDR_LOG_3\nmizar_PCIE0_SII_APP_HDR_LOG_2\nmizar_PCIE0_SII_APP_HDR_LOG_1\nmizar_PCIE0_SII_APP_HDR_LOG_0\nmizar_PCIE0_SII_CFG_BUS_NUM\nmizar_PCIE0_SII_CFG_BR_CTRL_SERREN\nmizar_PCIE0_SII_APP_DEV_AND_BUS_NUM\nmizar_PCIE0_SII_PCIE0_CONTROLLER_INT_STS_1\nmizar_PCIE0_SII_PCIE0_CONTROLLER_INTERRUPT_CONTROL_1\nmizar_PCIE0_SII_APP_AND_SLOT_CONTROL_REG\nmizar_PCIE0_SII_DIAG_CTRL_BUS\nmizar_PCIE0_SII_CFG_REG_RO\nmizar_PCIE0_SII_CFG_ARI_FWD_EN\nmizar_PCIE0_SII_RADM_SLOT_PWR_PAYLOAD\nmizar_PCIE0_SII_DIAG_STATUS_BUS_0\nmizar_PCIE0_SII_DIAG_STATUS_BUS_1\nmizar_PCIE0_SII_DIAG_STATUS_BUS_2\nmizar_PCIE0_SII_DIAG_STATUS_BUS_3\nmizar_PCIE0_SII_DIAG_STATUS_BUS_4\nmizar_PCIE0_SII_DIAG_STATUS_BUS_5\nmizar_PCIE0_SII_DIAG_STATUS_BUS_6\nmizar_PCIE0_SII_DIAG_STATUS_BUS_7\nmizar_PCIE0_SII_DIAG_STATUS_BUS_8\nmizar_PCIE0_SII_DIAG_STATUS_BUS_9\nmizar_PCIE0_SII_DIAG_STATUS_BUS_10\nmizar_PCIE0_SII_DIAG_STATUS_BUS_11\nmizar_PCIE0_SII_DIAG_STATUS_BUS_12\nmizar_PCIE0_SII_DIAG_STATUS_BUS_13\nmizar_PCIE0_SII_DIAG_STATUS_BUS_14\nmizar_PCIE0_SII_DIAG_STATUS_BUS_15\nmizar_PCIE0_SII_DIAG_STATUS_BUS_16\nmizar_PCIE0_SII_DIAG_STATUS_BUS_17\nmizar_PCIE0_SII_DIAG_STATUS_BUS_18\nmizar_PCIE0_SII_DIAG_STATUS_BUS_19\nmizar_PCIE0_SII_RAM_PWR_CNTRL_0\nmizar_PCIE0_SII_RAM_PWR_CNTRL_1\nmizar_PCIE0_SII_SOFT_RESET_CTRL\nmizar_PCIE0_SII_CFG_MSI_PENDING_B\nmizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_1\nmizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_2\nmizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_3\nmizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_4\nmizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_5\nmizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_6\nmizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_7\nmizar_PCIE0_SII_PHY_CONTROL_0\nmizar_PCIE0_SII_PHY_CONTROL_1\nmizar_PCIE0_SII_PHY_CONTROL_2\nmizar_PCIE0_SII_PHY_CONTROL_3\nmizar_PCIE0_SII_PHY_CONTROL_4\nmizar_PCIE0_SII_PHY_CONTROL_5\nmizar_PCIE0_SII_PHY_CONTROL_6\nmizar_PCIE0_SII_PHY_CONTROL_7\nmizar_PCIE0_SII_PHY_CONTROL_8\nmizar_PCIE0_SII_PHY_CONTROL_9\nmizar_PCIE0_SII_PHY_CONTROL_10\nmizar_PCIE0_SII_PHY_CONTROL_11\nmizar_PCIE0_SII_PHY_CONTROL_12\nmizar_PCIE0_SII_PHY_CONTROL_13\nmizar_PCIE0_SII_PHY_CONTROL_14\nmizar_PCIE0_SII_PHY_CONTROL_15\nmizar_PCIE0_SII_PHY_CONTROL_16\nmizar_PCIE0_SII_PHY_CONTROL_17\nmizar_PCIE0_SII_PHY_CONTROL_18\nmizar_PCIE0_SII_PHY_CONTROL_19\nmizar_PCIE0_SII_PHY_CONTROL_20\nmizar_PCIE0_SII_PHY_CONTROL_21\nmizar_PCIE0_SII_PHY_CONTROL_22\nmizar_PCIE0_SII_PHY_CONTROL_23\nmizar_PCIE0_SII_PHY_CONTROL_24\nmizar_PCIE0_SII_PHY_CONTROL_25\nmizar_PCIE0_SII_PHY_CONTROL_26\nmizar_PCIE0_SII_MSI_CTRL_IO\nmizar_PCIE0_SII_MSI_CTRL_INT_VEC"
  },
  {
    "Index": 2,
    "SS / Module": "PCIE1 SII RC",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Test Description": "Validate PCIe SII root complex register defaults and confirm masked write/read behavior across writable fields.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Unreadable registers are skipped. Unwritable registers are skipped. Specific addresses are excluded by a skip list. The reset control register is not checked. The reset helper is not executed.",
    "Test Steps / Procedure": "1) Read the reset value of each SII root complex register and compare with the documented default; skip unreadable and the reset control register.\n2) Write a set of data patterns to each writable SII root complex register that is not in the skip list.\n3) Read back each writable and readable SII root complex register and verify the masked result against the documented default.\n4) Record mismatches and determine the final result based on the absence of any failures.",
    "Impacted Registers": "SII_CFG_BAR0_START1\nSII_CFG_BAR0_START2\nSII_CFG_BAR0_LIMIT1\nSII_CFG_BAR0_LIMIT2\nSII_CFG_BAR1_START\nSII_CFG_BAR1_LIMIT1\nSII_CFG_BAR2_START1\nSII_CFG_BAR2_START2\nSII_CFG_BAR2_LIMIT1\nSII_CFG_BAR2_LIMIT2\nSII_CFG_BAR3_START\nSII_CFG_BAR3 LIMIT\nSII_CFG_BAR4 START1\nSII_CFG_BAR4 START2\nSII_CFG_BAR4 LIMIT1\nSII_CFG_BAR4 LIMIT2\nSII_CFG_BAR5 START\nSII_CFG_BAR5 LIMIT\nSII_PCIE1_CONFIG_INFO1\nSII_PCIE1_CONFIG_INFO2\nSII_PCIE1_GEN_CONTROL1\nSII_PCIE1_GEN_CONTROL2\nSII_PCIE1_GEN_CONTROL3\nSII_PCIE1_PM_CONTROL\nSII_PCIE1_CONTROL_PM_STS\nSII_PCIE1_TRANSMIT_HEADER1\nSII_PCIE1_TRANSMIT_HEADER2\nSII_PCIE1_TRANSMIT HEADER3\nSII_PCIE1_TRANSMIT HEADER4\nSII_PCIE1_TRANSMIT_REQ\nSII_PCIE1_RCV_MSG_HDR1\nSII_PCIE1_RCV_MSG_HDR2\nSII_PCIE1_RCV_MSG_HDR3\nSII_PCIE1_RCV_MSG_HDR4\nSII_PCIE1_RCV_MSG_STS\nSII_RCV_INTERRPUT_CTRL\nSII_CFG_EXP_ROM_START\nSII_CFG_EXP_ROM_LIMIT\nSII_CFG_EXP_ROM_INFO\nSII_CXPL_DEBUG_INFO1\nSII_CXPL_DEBUG_INFO2\nSII_CXPL_DEBUG_INFO_EI\nSII_PCIE1_TARGET_INFO1\nSII_PCIE1_TARGET_INFO2\nSII_PCIE1_CONTOLLER_ERROR_STATUS\nSII_PCIE1_CONTROLLER_INT_STS\nSII_PCIE1_CONTROLLER_INTERRUPT_CONTROL\nSII_PHY_RST_CONTROL\nSII_LINK_DEBUG_DATA\nSII_PCIE1_ERR_STS\nSII_PCIE1_ERR_INTERRUPT_CTRL\nSII_CFG_MSI_INT\nSII_LTR_MSG\nSII_LTR_MSG_LATENCY\nSII_APP_LTR_LATENCY\nSII_CFG_LTR_MAX_LATENCY\nSII_OBFF_CNTRL\nSII_SLV_AWMISC_INFO\nSII_SLV_AWMISC_INFO_HDR_34DW_HI\nSII_SLV_AWMISC_INFO_HDR_34DW_LO\nSII_SLV_MISC_INFO\nSII_SLV_MISC_RESP_INFO\nSII_MSTR_AWMISC_INFO_CNTRL\nSII_MSTR_AWMISC_INFO_1\nSII_MSTR_AWMISC_INFO_0\nSII_MSTR_AWMISC_INFO_HDR_34DW_HI\nSII_MSTR_AWMISC_INFO_HDR_34DW_LO\nSII_MSTR_ARMISC_INFO_CNTRL\nSII_MSTR_ARMISC_INFO_1\nSII_MSTR_ARMISC_INFO_0\nSII_MSTR_BMISC_RMISC_CPL_STAT_INFO\nSII_RADM_TIMEOUT_INFO\nSII_CFG_MSI_INFO\nSII_CFG_MSI_DATA\nSII_CFG_MSI_ADDR_HI\nSII_CFG_MSI_ADDR_LO\nSII_CFG_AER_INT AND PCIE1_CAP INT_MSG\nSII_RTLH_RFC_DATA\nSII_APP_HDR_INFO\nSII_APP_HDR_LOG_3\nSII_APP_HDR_LOG_2\nSII_APP_HDR_LOG_1\nSII_APP_HDR_LOG_0\nSII_CFG_BUS_NUM\nSII_CFG_BR_CTRL_SERREN\nSII_APP_DEV_AND_BUS_NUM\nSII_PCIE1_CONTROLLER INT_STS_1\nSII_PCIE1_CONTROLLER_INTERRUPT_CONTROL_1\nSII_APP_AND_SLOT_CONTROL_REG\nSII_DIAG_CTRL_BUS\nSII_CFG_REG_RO\nSII_CFG_ARI_FWD_EN\nSII_RADM_SLOT_PWR_PAYLOAD\nSII_DIAG_STATUS_BUS_0\nSII_DIAG_STATUS_BUS_1\nSII_DIAG_STATUS_BUS_2\nSII_DIAG_STATUS_BUS_3\nSII_DIAG_STATUS_BUS_4\nSII_DIAG_STATUS_BUS_5\nSII_DIAG_STATUS_BUS_6\nSII_DIAG_STATUS_BUS_7\nSII_DIAG_STATUS_BUS_8\nSII_DIAG_STATUS_BUS_9\nSII_DIAG_STATUS_BUS_10\nSII_DIAG_STATUS_BUS_11\nSII_DIAG_STATUS_BUS_12\nSII_DIAG_STATUS_BUS_13\nSII_DIAG_STATUS_BUS_14\nSII_DIAG_STATUS_BUS_15\nSII_DIAG_STATUS_BUS_16\nSII_DIAG_STATUS_BUS_17\nSII_DIAG_STATUS_BUS_18\nSII_DIAG_STATUS_BUS_19\nSII_RAM_PWR_CNTRL_0\nSII_RAM_PWR_CNTRL_1\nSII_SOFT_RESET_CTRL\nSII_CFG_MSI_PENDING_B\nSII_SMLH_LTSSM_STATE_TRAN_1\nSII_SMLH_LTSSM_STATE_TRAN_2\nSII_SMLH_LTSSM_STATE_TRAN_3\nSII_SMLH_LTSSM_STATE_TRAN_4\nSII_SMLH_LTSSM_STATE_TRAN_5\nSII_SMLH_LTSSM_STATE_TRAN_6\nSII_SMLH_LTSSM_STATE_TRAN_7\nSII_PHY_CONTROL_0\nSII_PHY_CONTROL_1\nSII_PHY_CONTROL_2\nSII_PHY_CONTROL_3\nSII_PHY_CONTROL_4\nSII_PHY_CONTROL_5\nSII_PHY_CONTROL_6\nSII_PHY_CONTROL_7\nSII_PHY_CONTROL_8\nSII_PHY_CONTROL_9\nSII_PHY_CONTROL_10\nSII_PHY_CONTROL_11\nSII_PHY_CONTROL_12\nSII_PHY_CONTROL_13\nSII_PHY_CONTROL_14\nSII_PHY_CONTROL_15\nSII_PHY_CONTROL_16\nSII_PHY_CONTROL_17\nSII_PHY_CONTROL_18\nSII_PHY_CONTROL_19\nSII_PHY_CONTROL_20\nSII_PHY_CONTROL_21\nSII_PHY_CONTROL_22\nSII_PHY_CONTROL_23\nSII_PHY_CONTROL_24\nSII_PHY_CONTROL_25\nSII_PHY_CONTROL_26\nSII_MSI_CTRL_IO\nSII_MSI_CTRL_INT_VEC",
    "Validation / Acceptance Criteria": "1) Default value check → Read value equals the documented default.\n2) Masked read-back after pattern write equals expected value.\n3) No failures → Pass.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Hidden_Test_Description": "test_case() calls chk_rst_val() then chk_rd_wr(); if def_fail_cnt>0 or wr_fail_cnt>0 it finishes with failure, else success. chk_rst_val(): loops i=0..CNT-1, addr=addr_array[i]; if read_mask_array[i]==0 skip; if addr_array[i]==mizar_PCIE1_SII_PHY_RST_CONTROL skip; data_rd=read_reg(addr); compare data_rd with default_value_array[i]; increment def_fail_cnt on mismatch. chk_rd_wr(): patterns chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}; for each pattern write to each address if skip_array[i]==0 and write_mask_array[i]!=0; then for each address if skip_array[i]==0 and write_mask_array[i]!=0 and read_mask_array[i]!=0, read_reg(addr); compute wr_n=(write_mask_array[i]^0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd==exp_val) pass; else wr_fail_cnt++ and print failure. soft_reset_chk() toggles SOFT_RST_REG_ADDRESS but is not invoked.",
    "Hidden_Remarks": "Addresses with read_mask_array[i]==0 are skipped. Addresses with write_mask_array[i]==0 are not written and skipped for verification. Entries where skip_array[i]==1 are skipped for both write and read. The address equal to mizar_PCIE1_SII_PHY_RST_CONTROL is excluded from default-value checking.",
    "Hidden_Test_Steps_Procedure": "1) In test_case(), call chk_rst_val(). 2) In chk_rst_val(): for i=0..CNT-1 set addr=addr_array[i]; if(read_mask_array[i]==0) continue; if(addr_array[i]==mizar_PCIE1_SII_PHY_RST_CONTROL) continue; data_rd=read_reg(addr); if(data_rd==default_value_array[i]) optional PASS print; else def_fail_cnt++ and print failure. 3) Back in test_case(), call chk_rd_wr(). 4) In chk_rd_wr(): define chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}; for each pattern j set data_wr=chk_val[j]; write phase: for i=0..CNT-1, addr=addr_array[i]; if(skip_array[i]==1) continue; if(write_mask_array[i]==0) continue; else write_reg(addr,data_wr). read/verify phase: for i=0..CNT-1, addr=addr_array[i]; if(skip_array[i]==1) continue; if(write_mask_array[i]==0) continue; if(read_mask_array[i]==0) continue; else data_rd=read_reg(addr); wr_n=(write_mask_array[i]^0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd==exp_val) optional PASS print; else wr_fail_cnt++ and print failure. 5) In test_case(), if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0).",
    "Hidden_Impacted_Registers": "mizar_PCIE1_SII_CFG_BAR0_START1\nmizar_PCIE1_SII_CFG_BAR0_START2\nmizar_PCIE1_SII_CFG_BAR0_LIMIT1\nmizar_PCIE1_SII_CFG_BAR0_LIMIT2\nmizar_PCIE1_SII_CFG_BAR1_START\nmizar_PCIE1_SII_CFG_BAR1_LIMIT1\nmizar_PCIE1_SII_CFG_BAR2_START1\nmizar_PCIE1_SII_CFG_BAR2_START2\nmizar_PCIE1_SII_CFG_BAR2_LIMIT1\nmizar_PCIE1_SII_CFG_BAR2_LIMIT2\nmizar_PCIE1_SII_CFG_BAR3_START\nmizar_PCIE1_SII_CFG_BAR3_LIMIT\nmizar_PCIE1_SII_CFG_BAR4_START1\nmizar_PCIE1_SII_CFG_BAR4_START2\nmizar_PCIE1_SII_CFG_BAR4_LIMIT1\nmizar_PCIE1_SII_CFG_BAR4_LIMIT2\nmizar_PCIE1_SII_CFG_BAR5_START\nmizar_PCIE1_SII_CFG_BAR5_LIMIT\nmizar_PCIE1_SII_PCIE1_CONFIG_INFO1\nmizar_PCIE1_SII_PCIE1_CONFIG_INFO2\nmizar_PCIE1_SII_PCIE1_GEN_CONTROL1\nmizar_PCIE1_SII_PCIE1_GEN_CONTROL2\nmizar_PCIE1_SII_PCIE1_GEN_CONTROL3\nmizar_PCIE1_SII_PCIE1_PM_CONTROL\nmizar_PCIE1_SII_PCIE1_CONTROL_PM_STS\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER1\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER4\nmizar_PCIE1_SII_PCIE1_TRANSMIT_REQ\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR1\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR2\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR3\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR4\nmizar_PCIE1_SII_PCIE1_RCV_MSG_STS\nmizar_PCIE1_SII_RCV_INTERRPUT_CTRL\nmizar_PCIE1_SII_CFG_EXP_ROM_START\nmizar_PCIE1_SII_CFG_EXP_ROM_LIMIT\nmizar_PCIE1_SII_CFG_EXP_ROM_INFO\nmizar_PCIE1_SII_CXPL_DEBUG_INFO1\nmizar_PCIE1_SII_CXPL_DEBUG_INFO2\nmizar_PCIE1_SII_CXPL_DEBUG_INFO_EI\nmizar_PCIE1_SII_PCIE1_TARGET_INFO1\nmizar_PCIE1_SII_PCIE1_TARGET_INFO2\nmizar_PCIE1_SII_PCIE1_CONTOLLER_ERROR_STATUS\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL\nmizar_PCIE1_SII_PHY_RST_CONTROL\nmizar_PCIE1_SII_LINK_DEBUG DATA\nmizar_PCIE1_SII_PCIE1_ERR_STS\nmizar_PCIE1_SII_PCIE1_ERR_INTERRUPT_CTRL\nmizar_PCIE1_SII_CFG_MSI_INT\nmizar_PCIE1_SII_LTR_MSG\nmizar_PCIE1_SII_LTR_MSG_LATENCY\nmizar_PCIE1_SII_APP_LTR_LATENCY\nmizar_PCIE1_SII_CFG_LTR_MAX_LATENCY\nmizar_PCIE1_SII_OBFF_CNTRL\nmizar_PCIE1_SII_SLV_AWMISC_INFO\nmizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_HI\nmizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_LO\nmizar_PCIE1_SII_SLV_MISC_INFO\nmizar_PCIE1_SII_SLV_MISC_RESP_INFO\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_CNTRL\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_1\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_0\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_HI\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_LO\nmizar_PCIE1_SII_MSTR_ARMISC_INFO_CNTRL\nmizar_PCIE1_SII_MSTR_ARMISC_INFO_1\nmizar_PCIE1_SII_MSTR_ARMISC_INFO_0\nmizar_PCIE1_SII_MSTR_BMISC_RMISC_CPL_STAT_INFO\nmizar_PCIE1_SII_RADM_TIMEOUT_INFO\nmizar_PCIE1_SII_CFG_MSI_INFO\nmizar_PCIE1_SII_CFG_MSI_DATA\nmizar_PCIE1_SII_CFG_MSI_ADDR_HI\nmizar_PCIE1_SII_CFG_MSI_ADDR_LO\nmizar_PCIE1_SII_CFG_AER_INT_AND_PCIE1_CAP_INT_MSG\nmizar_PCIE1_SII_RTLH_RFC_DATA\nmizar_PCIE1_SII_APP_HDR_INFO\nmizar_PCIE1_SII_APP_HDR_LOG_3\nmizar_PCIE1_SII_APP_HDR_LOG_2\nmizar_PCIE1_SII_APP_HDR_LOG_1\nmizar_PCIE1_SII_APP_HDR_LOG_0\nmizar_PCIE1_SII_CFG_BUS_NUM\nmizar_PCIE1_SII_CFG_BR_CTRL_SERREN\nmizar_PCIE1_SII_APP_DEV_AND BUS_NUM\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS_1\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL_1\nmizar_PCIE1_SII_APP_AND_SLOT_CONTROL_REG\nmizar_PCIE1_SII_DIAG_CTRL_BUS\nmizar_PCIE1_SII_CFG_REG_RO\nmizar_PCIE1_SII_CFG_ARI_FWD_EN\nmizar_PCIE1_SII_RADM_SLOT_PWR_PAYLOAD\nmizar_PCIE1_SII_DIAG_STATUS_BUS_0\nmizar_PCIE1_SII_DIAG_STATUS_BUS_1\nmizar_PCIE1_SII_DIAG_STATUS_BUS_2\nmizar_PCIE1_SII_DIAG_STATUS_BUS_3\nmizar_PCIE1_SII_DIAG_STATUS_BUS_4\nmizar_PCIE1_SII_DIAG_STATUS_BUS_5\nmizar_PCIE1_SII_DIAG_STATUS_BUS_6\nmizar_PCIE1_SII_DIAG_STATUS_BUS_7\nmizar_PCIE1_SII_DIAG_STATUS_BUS_8\nmizar_PCIE1_SII_DIAG_STATUS_BUS_9\nmizar_PCIE1_SII_DIAG_STATUS_BUS_10\nmizar_PCIE1_SII_DIAG_STATUS_BUS_11\nmizar_PCIE1_SII_DIAG_STATUS_BUS_12\nmizar_PCIE1_SII_DIAG_STATUS_BUS_13\nmizar_PCIE1_SII_DIAG_STATUS_BUS_14\nmizar_PCIE1_SII_DIAG_STATUS_BUS_15\nmizar_PCIE1_SII_DIAG_STATUS_BUS_16\nmizar_PCIE1_SII_DIAG_STATUS_BUS_17\nmizar_PCIE1_SII_DIAG_STATUS_BUS_18\nmizar_PCIE1_SII_DIAG_STATUS_BUS_19\nmizar_PCIE1_SII_RAM_PWR_CNTRL_0\nmizar_PCIE1_SII_RAM_PWR_CNTRL_1\nmizar_PCIE1_SII_SOFT_RESET_CTRL\nmizar_PCIE1_SII_CFG_MSI_PENDING_B\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_1\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_2\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_3\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_4\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_5\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_6\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_7\nmizar_PCIE1_SII_PHY_CONTROL_0\nmizar_PCIE1_SII_PHY_CONTROL_1\nmizar_PCIE1_SII_PHY_CONTROL_2\nmizar_PCIE1_SII_PHY_CONTROL_3\nmizar_PCIE1_SII_PHY_CONTROL_4\nmizar_PCIE1_SII_PHY_CONTROL_5\nmizar_PCIE1_SII_PHY_CONTROL_6\nmizar_PCIE1_SII_PHY_CONTROL_7\nmizar_PCIE1_SII_PHY_CONTROL_8\nmizar_PCIE1_SII_PHY_CONTROL_9\nmizar_PCIE1_SII_PHY_CONTROL_10\nmizar_PCIE1_SII_PHY_CONTROL_11\nmizar_PCIE1_SII_PHY_CONTROL_12\nmizar_PCIE1_SII_PHY_CONTROL_13\nmizar_PCIE1_SII_PHY_CONTROL_14\nmizar_PCIE1_SII_PHY_CONTROL_15\nmizar_PCIE1_SII_PHY_CONTROL_16\nmizar_PCIE1_SII_PHY_CONTROL_17\nmizar_PCIE1_SII_PHY_CONTROL_18\nmizar_PCIE1_SII_PHY_CONTROL_19\nmizar_PCIE1_SII_PHY_CONTROL_20\nmizar_PCIE1_SII_PHY_CONTROL_21\nmizar_PCIE1_SII_PHY_CONTROL_22\nmizar_PCIE1_SII_PHY_CONTROL_23\nmizar_PCIE1_SII_PHY_CONTROL_24\nmizar_PCIE1_SII_PHY_CONTROL_25\nmizar_PCIE1_SII_PHY_CONTROL_26\nmizar_PCIE1_SII_MSI_CTRL_IO\nmizar_PCIE1_SII_MSI_CTRL_INT_VEC",
    "Hidden_Validation_Acceptance_Criteria": "Default check: data_rd == default_value_array[i] for each i where read_mask_array[i]!=0 and addr_array[i]!=mizar_PCIE1_SII_PHY_RST_CONTROL; else increment def_fail_cnt and print failure. Write/read check: For each pattern in {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000} and for each i where skip_array[i]==0 and write_mask_array[i]!=0 and read_mask_array[i]!=0, after write_reg(addr_array[i],data_wr) then data_rd=read_reg(addr_array[i]); wr_n=(write_mask_array[i]^0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if data_rd==exp_val PASS; else wr_fail_cnt++ and print failure. Final: if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0).",
    "Hidden_Header_Includes": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<pcie1/pcie_sii_rc_def.h>\n#include<pcie1/pcie_sii_rc_offset.h>",
    "Hidden_Macro_Defines": "#define SOFT_RST_REG_ADDRESS\t0x00000000\n#define SOFT_RST_REG_DATA\t0x00000000\n#define MIZAR_PCIE1_SII_BASE     0xE68C1000\n#define CNT 153",
    "Hidden_Skip_Array_Definition": "const int skip_array[153]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,;}"
  },
  {
    "Index": 3,
    "SS / Module": "PCIE",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie_cfg_wr_rd_test",
    "Test Description": "Configures coherency settings and verifies basic PCIe configuration access with link readiness polling and a completion indication.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Behavior depends on compile-time options. Link must be ready. Final completion value is required.",
    "Test Steps / Procedure": "1) Program DBI_DSP_COHERENCY_CONTROL_3 on both controllers.\n2) Poll the SII status register until the link is reported ready.\n3) Program configuration registers including BAR registers and enable the command register.\n4) Wait until the completion indication register shows the expected value.",
    "Impacted Registers": "DBI_DSP_COHERENCY_CONTROL_3",
    "Validation / Acceptance Criteria": "1) Link status is ready → Continue configuration.\n2) Completion indication is observed → Test passes.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie_cfg_wr_rd_test",
    "Hidden_Test_Description": "The test writes 0x0 to 0xE6004100 and performs link training based on compile-time defines (DM0_RC, DM1_RC, DM0_EP, DM1_EP). It configures coherency by repeatedly setting bitfields (11–14, 3–6, 27–30, 19–22 with 0xF) in mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF using set_data(...) around read_reg()/write_reg() calls, with waits between sequences. It polls SII0 at offset 0xC0 until (data_rd & 0xD1) == 0xD1, and if DM1_RC is defined it also polls SII1 at 0xC0 until the same condition holds. It writes 0x11111111 to 0xE6004100 and waits. Under DM0_RC it calls mem_base_program_dm0_x4(), reads the first 10 configuration DWs via read_pcie_slv0_reg(i*4), writes 0xFFFFFFFF to BAR offsets 0x10..0x24, reads them back, then writes 0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000 to 0x10..0x24 respectively and reads back. It enables Memory/IO/Bus master by writing 0x7 to offset 0x4. Under DM1_RC it performs analogous steps using mem_base_program_dm1_x4() and the instance 1 config space accessors. Finally it waits until read_reg(0xE6004100) equals 0x12345678, then finishes with finish(0).",
    "Hidden_Remarks": "Behavior is gated by DM0_RC, DM1_RC, DM0_EP, DM1_EP compile-time defines. The function non_secure_prot_nic() is invoked before configuration accesses. Link readiness is determined by polling SII status offset 0xC0 for bits matching 0xD1. An external agent must write 0x12345678 to 0xE6004100 to satisfy the final handshake. DEBUG_DISPLAY controls diagnostic prints.",
    "Hidden_Test_Steps_Procedure": "1) write_reg(0xE6004100, 0x0). 2) Conditionally perform link_training_dm0_x4(4) or link_training_dm1_x4(4) for RC/EP variants based on DM0_RC, DM1_RC, DM0_EP, DM1_EP. 3) Configure coherency: for PCIE0, read_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF), set_data(...,11,14,0xF), set_data(...,3,6,0xF), write_reg(...); then read/set (27,30,0xF) and (19,22,0xF) and write back. Repeat the same sequences for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 4) wait_on(20); then for each of PCIE0 and PCIE1, read_reg of the same coherency control register and set the four bitfield ranges (11–14, 3–6, 27–30, 19–22) with 0xF before write_reg. 5) data_rd = read_sii0_reg(0xC0); while ((data_rd & 0xD1) != 0xD1) { data_rd = read_sii0_reg(0xC0); }. If DM1_RC, poll SII1 with read_sii1_reg(0xC0) until ((data_rd & 0xD1) == 0xD1). 6) write_reg(0xE6004100, 0x11111111); wait_on(15000). 7) If DM0_RC: call mem_base_program_dm0_x4(); wait_on(10); read first 10 DWs via read_pcie_slv0_reg(i*0x4) for i=0..9; write_pcie_slv0_reg 0x10..0x24 with 0xFFFFFFFF; read back 0x10..0x24; then write 0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000 to 0x10..0x24 respectively; read back; write_pcie_slv0_reg(0x4, 0x7). 8) If DM1_RC: call mem_base_program_dm1_x4(); read first 10 DWs via read_pcie_slv1_reg(i*0x4); write_pcie_slv1_reg(0x4, 0x7); write 0xFFFFFFFF to BAR offsets 0x10..0x24; read back; write 0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000 to 0x10..0x24; read back. 9) wait_on(10); loop reading data_rd = read_reg(0xE6004100) until data_rd == 0x12345678. 10) finish(0).",
    "Hidden_Impacted_Registers": "mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF\nmizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF",
    "Hidden_Validation_Acceptance_Criteria": "Link readiness: while(((data_rd)&(0xD1))!=0xD1) keep polling SII status at 0xC0; PASS when (data_rd & 0xD1) == 0xD1 for the required instance(s). Final handshake: read_reg(0xE6004100) must become 0x12345678; loop continues until equality holds, then PASS path executes finish(0). No explicit failure path is coded other than indefinite wait if conditions are not met.",
    "Hidden_Header_Includes": "#include <stdlib.h>\n#include <stdio.h>\n#include <test_common.h>\n#include \"pcie.h\"",
    "Hidden_Macro_Defines": "NA",
    "Hidden_Skip_Array_Definition": "NA"
  }
]'''

IP_NAME = "PCIE"
OUTPUT_DIR = os.path.join("Test_Output", IP_NAME, "TestPlan")

# Column definitions
META_COLUMNS_SPEC = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
    "Hidden_Header_Includes",
    # Accept both singular and plural to avoid dropping
    "Hidden_Macro_Define",
    "Hidden_Macro_Defines",
    "Hidden_Skip_Array_Definition",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")  # blue
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="top")
TOP = Alignment(vertical="top")
BORDER_THIN = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


def parse_json() -> List[Dict[str, Any]]:
    data = json.loads(JSON_TEXT)
    if not isinstance(data, list) or len(data) == 0:
        raise SystemExit("JSON input invalid or empty")
    return data


def union_keys_preserve_order(rows: List[Dict[str, Any]]) -> List[str]:
    seen = []
    s = set()
    for row in rows:
        for k in row.keys():
            if k not in s:
                s.add(k)
                seen.append(k)
    return seen


def build_workbook(rows: List[Dict[str, Any]], all_keys: List[str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write header
    for c, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = BORDER_THIN

    # Write rows
    for r, rec in enumerate(rows, start=2):
        for c, key in enumerate(all_keys, 1):
            val = rec.get(key, "")
            ws.cell(row=r, column=c, value=val).alignment = TOP
            ws.cell(row=r, column=c).border = BORDER_THIN

    ws.freeze_panes = "A2"

    # Auto-fit columns (approx by max len)
    for c, key in enumerate(all_keys, 1):
        max_len = len(str(key))
        for r in range(2, len(rows) + 2):
            v = ws.cell(row=r, column=c).value
            ln = len(str(v)) if v is not None else 0
            if ln > max_len:
                max_len = ln
        width = min(max_len + 2, 80)
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = width

    return wb


def create_meta_sheet(wb: Workbook, rows: List[Dict[str, Any]], all_keys: List[str]):
    meta_cols_present = [k for k in META_COLUMNS_SPEC if k in all_keys]
    ws = wb.create_sheet("Meta_data_sheet")

    # Header
    for c, key in enumerate(meta_cols_present, 1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = BORDER_THIN

    # Data
    for r, rec in enumerate(rows, start=2):
        for c, key in enumerate(meta_cols_present, 1):
            val = rec.get(key, "")
            ws.cell(row=r, column=c, value=val).alignment = TOP
            ws.cell(row=r, column=c).border = BORDER_THIN

    # Very Hidden
    ws.sheet_state = 'veryHidden'


def number_list(text: Any) -> Any:
    if text is None:
        return text
    s = str(text).strip()
    if s == "":
        return s
    parts = [p.strip() for p in s.splitlines() if p.strip()]
    if len(parts) <= 1:
        return s
    return "\n".join(f"{i+1}. {line}" for i, line in enumerate(parts))


def normalize_main_sheet(wb: Workbook, rows: List[Dict[str, Any]], all_keys: List[str]):
    ws = wb["Data"]

    meta_set = set(META_COLUMNS_SPEC)
    main_present = [k for k in MAIN_ORDER if k in all_keys]
    remaining = [k for k in all_keys if (k not in MAIN_ORDER and k not in meta_set)]
    final_cols = main_present + remaining

    ws.delete_rows(1, ws.max_row)

    # Header row
    for c, key in enumerate(final_cols, 1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = BORDER_THIN

    wrap_cols = {"Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"}
    name_to_col = {k: i+1 for i, k in enumerate(final_cols)}

    for r, rec in enumerate(rows, start=2):
        for key in final_cols:
            val = rec.get(key, "")
            if key in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                val = number_list(val)
            c = name_to_col[key]
            ws.cell(row=r, column=c, value=val)
            if key in wrap_cols:
                ws.cell(row=r, column=c).alignment = LEFT_WRAP
            elif key == "Index":
                ws.cell(row=r, column=c).alignment = CENTER
            else:
                ws.cell(row=r, column=c).alignment = LEFT
            ws.cell(row=r, column=c).border = BORDER_THIN

        # Approx row height
        max_lines = 1
        for key in ("Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"):
            if key in name_to_col:
                v = ws.cell(row=r, column=name_to_col[key]).value
                if v is not None:
                    lines = str(v).count("\n") + 1
                    if lines > max_lines:
                        max_lines = lines
        ws.row_dimensions[r].height = min(15 * max_lines, 200)

    ws.freeze_panes = "A2"

    # Auto-fit columns
    for c, key in enumerate(final_cols, 1):
        max_len = len(str(key))
        for r in range(2, len(rows) + 2):
            v = ws.cell(row=r, column=c).value
            ln = len(str(v)) if v is not None else 0
            if ln > max_len:
                max_len = ln
        width = min(max_len + 2, 100)
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = width

    # Data validation for Code Generation (Required / Not)
    if "Code Generation (Required / Not)" in name_to_col:
        col = name_to_col["Code Generation (Required / Not)"]
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=len(rows)+1, column=col).coordinate}")

    ws.title = "TestPlan"


def ensure_final_visibility(wb: Workbook):
    if "Data" in wb.sheetnames:
        del wb["Data"]


def save_xlsx_and_validate(wb: Workbook) -> str:
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    fname = f"{IP_NAME}_TestPlan_{now_ist.strftime('%Y%m%d')}_{now_ist.strftime('%H%M%S')}.xlsx"

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, fname)
    wb.save(out_path)

    with zipfile.ZipFile(out_path, 'r') as z:
        names = set(z.namelist())
        required = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
        if not required.issubset(names):
            raise SystemExit("XLSX validation failed: required parts missing")
    print(out_path)
    return out_path


def main():
    rows = parse_json()
    all_keys = union_keys_preserve_order(rows)

    wb = build_workbook(rows, all_keys)
    create_meta_sheet(wb, rows, all_keys)
    normalize_main_sheet(wb, rows, all_keys)
    ensure_final_visibility(wb)
    out_path = save_xlsx_and_validate(wb)

if __name__ == "__main__":
    main()
