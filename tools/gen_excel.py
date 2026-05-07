#!/usr/bin/env python3
import json, os, sys, zipfile, re
from copy import deepcopy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Constants
META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]
MAIN_COLUMNS_FINAL_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]
WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}
DV_COLUMN = "Code Generation (Required / Not)"
DV_ALLOWED = ["Required", "Blank", "Not Required"]

INPUT_JSON_PATH = os.getenv("INPUT_JSON_PATH", "Test_Output/GPIO/TestPlan/master_batch1.json")
OUTPUT_XLSX_PATH = os.getenv("OUTPUT_XLSX_PATH", "Test_Output/GPIO/TestPlan/GPIO_TestPlan_WORKING.xlsx")


def fail(msg: str):
    print(f"ERROR: {msg}")
    sys.exit(1)


def load_json_records(path: str):
    if not os.path.exists(path):
        fail(f"JSON input not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    # Accept either array or object keyed by TC1..TCn
    if isinstance(raw, dict):
        # Preserve insertion order; otherwise fall back to TC1..TCn sort
        keys = list(raw.keys())
        if not keys:
            fail("JSON object has no keys")
        records = [raw[k] for k in keys]
    elif isinstance(raw, list):
        records = raw
    else:
        fail("JSON root must be an object or array")
    if not records:
        fail("JSON input has zero records")
    # Ensure each record is a dict
    out = []
    for i, rec in enumerate(records, 1):
        if not isinstance(rec, dict):
            fail(f"Record {i} is not an object")
        out.append(deepcopy(rec))
    return out


def union_keys_preserve_order(records):
    seen = []
    s = set()
    for rec in records:
        for k in rec.keys():
            if k not in s:
                s.add(k)
                seen.append(k)
    return seen


def jsonify_cell(v):
    # Preserve scalars; deterministically stringify non-scalars
    if isinstance(v, (str, int, float)) or v is None:
        return v
    return json.dumps(v, ensure_ascii=False, separators=(",", ":"))


def write_data_sheet(ws, keys, records):
    # Header
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    # Rows
    for r, rec in enumerate(records, 2):
        for c, k in enumerate(keys, 1):
            v = jsonify_cell(rec.get(k, ""))
            ws.cell(row=r, column=c, value=v)


def estimate_col_widths(ws):
    # Simple deterministic width estimation
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                l = 0
            else:
                s = str(v)
                l = max((len(line) for line in s.splitlines()), default=0)
            if l > max_len:
                max_len = l
        width = min(max(10, max_len + 2), 120)
        ws.column_dimensions[col_letter].width = width


def create_meta_sheet(wb, records):
    ws = wb.create_sheet("Meta_data_sheet")
    # Determine which META columns exist in input (preserve defined order)
    meta_cols = [c for c in META_COLUMNS if any(c in rec for rec in records)]
    if not meta_cols:
        # Still create empty sheet per spec
        meta_cols = META_COLUMNS
    # Header
    for c, k in enumerate(meta_cols, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Rows
    for r, rec in enumerate(records, 2):
        for c, k in enumerate(meta_cols, 1):
            v = rec.get(k, "")
            ws.cell(row=r, column=c, value=jsonify_cell(v))
    # Very hidden
    ws.sheet_state = 'veryHidden'
    return ws, meta_cols


def normalize_testplan_sheet(ws, records, main_cols):
    # Build normalized rows for main sheet, removing META columns
    rows = []
    for rec in records:
        row = {k: rec.get(k, "") for k in main_cols}
        rows.append(row)
    # Clear existing sheet and rewrite headers + rows
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    for c, k in enumerate(main_cols, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(rows, 2):
        for c, k in enumerate(main_cols, 1):
            v = jsonify_cell(row.get(k, ""))
            ws.cell(row=r, column=c, value=v)


def apply_numbering_in_cell(s: str) -> str:
    if s is None:
        return ""
    text = str(s)
    # Split to logical items by newlines
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return text
    # Remove leading bullets/numbering like "- ", "1)", "1.", "•"
    cleaned = []
    for ln in lines:
        ln2 = re.sub(r"^([0-9]+[\).]|[-•])\s*", "", ln)
        cleaned.append(ln2)
    # Re-number 1..n with "."
    numbered = [f"{i+1}. {item}" for i, item in enumerate(cleaned)]
    return "\n".join(numbered)


def style_and_format_testplan(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    # Wrap text for specific columns
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, max_col+1)}
    wrap_cols_idx = [headers[k] for k in WRAP_COLUMNS if k in headers]
    for r in range(2, max_row+1):
        for c in wrap_cols_idx:
            cell = ws.cell(row=r, column=c)
            cell.value = apply_numbering_in_cell(cell.value) if ws.cell(row=1, column=c).value in {"Test Steps / Procedure", "Validation / Acceptance Criteria"} else ("" if cell.value is None else str(cell.value))
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    # Header style
    header_fill = PatternFill("solid", fgColor="4472C4")
    for c in range(1, max_col+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Data rows alignment
    idx_col = None
    for c in range(1, max_col+1):
        if ws.cell(row=1, column=c).value == "Index":
            idx_col = c
            break
    for r in range(2, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            if c == idx_col:
                cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
            else:
                # If looks numeric, right-align, else left
                v = cell.value
                if isinstance(v, (int, float)) or (isinstance(v, str) and v.isdigit()):
                    cell.alignment = Alignment(vertical="top", horizontal="right", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    # Thin borders for all populated cells
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            ws.cell(row=r, column=c).border = border
    # Estimate widths and row heights
    estimate_col_widths(ws)
    for r in range(2, max_row+1):
        # Estimate height by max line count among wrap columns
        max_lines = 1
        for c in wrap_cols_idx:
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            lines = str(v).count("\n") + 1
            if lines > max_lines:
                max_lines = lines
        ws.row_dimensions[r].height = min(15 * max_lines, 350)

    # Data validation for DV_COLUMN
    if DV_COLUMN in headers:
        col_idx = headers[DV_COLUMN]
        dv = DataValidation(type="list", formula1='"' + ", ".join(DV_ALLOWED) + '"', allow_blank=False, showDropDown=True)
        dv.error = "Select a value from the list"
        dv.prompt = "Choose one of: " + ", ".join(DV_ALLOWED)
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=max_row, column=col_idx).coordinate}")


def ensure_only_allowed_sheets(wb):
    names = [ws.title for ws in wb.worksheets]
    if names.count("TestPlan") != 1 or names.count("Meta_data_sheet") != 1:
        fail(f"Unexpected worksheets present: {names}")


def main():
    records = load_json_records(INPUT_JSON_PATH)
    # Create workbook and staging Data sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    keys = union_keys_preserve_order(records)
    write_data_sheet(ws, keys, records)
    estimate_col_widths(ws)

    # Create META sheet
    meta_ws, meta_cols = create_meta_sheet(wb, records)

    # Rename Data -> TestPlan and normalize main sheet
    ws = wb["Data"]
    ws.title = "TestPlan"
    normalize_testplan_sheet(ws, records, MAIN_COLUMNS_FINAL_ORDER)

    # Apply formatting and numbering rules
    style_and_format_testplan(ws)

    # Mandatory safety: ensure no sheet named 'Data' remains
    if "Data" in wb.sheetnames:
        # Delete if any stray
        del wb["Data"]
    ensure_only_allowed_sheets(wb)

    # Ensure output directory exists
    out_dir = os.path.dirname(OUTPUT_XLSX_PATH)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    # Save and validate as proper XLSX
    wb.save(OUTPUT_XLSX_PATH)
    if not zipfile.is_zipfile(OUTPUT_XLSX_PATH):
        fail("Produced file is not a valid XLSX ZIP")
    # Try opening with openpyxl again
    load_workbook(OUTPUT_XLSX_PATH).close()
    print(f"Saved XLSX: {OUTPUT_XLSX_PATH}")

if __name__ == "__main__":
    main()
