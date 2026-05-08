import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Constants
DATASET_PATH = Path("Test_Output/GPIO/TestPlan/dataset_GPIO.json")
OUTPUT_XLSX = Path("Test_Output/GPIO/TestPlan/GPIO_TestPlan_20260508_000000.xlsx")  # Locked name (IST at run start)
META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]
MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]
WRAP_COLS = {"Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"}

# Helpers
thin = Side(border_style="thin", color="000000")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="BDD7EE")  # light blue
header_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_top = Alignment(horizontal="center", vertical="top")


def renumber_block(text: str) -> str:
    if not text:
        return text
    lines = [ln.strip() for ln in str(text).splitlines() if ln.strip()]
    return "\n".join(f"{i+1}. {ln.lstrip('- ').lstrip('* ').lstrip('0123456789. ') }" for i, ln in enumerate(lines))


def autofit_ws(ws):
    col_widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            ln = 0 if val is None else len(str(val))
            col_widths[idx] = max(col_widths.get(idx, 0), ln)
    for idx, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(width + 2, 12), 80)


def apply_table_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_all


def main():
    # Load dataset
    data = json.loads(DATASET_PATH.read_text(encoding="utf-8"))
    assert isinstance(data, list) and len(data) > 0, "json_data must be a non-empty array"

    # Determine column order (union in first-seen order across rows)
    columns = []
    seen = set()
    for row in data:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)

    # Build workbook and Data sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header
    for c_idx, key in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=key)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill

    # Rows
    for r_idx, row in enumerate(data, start=2):
        for c_idx, key in enumerate(columns, start=1):
            val = row.get(key, "")
            # Mandatory numbering for two specific fields on main sheet later; keep raw here
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws.freeze_panes = "A2"
    autofit_ws(ws)

    # Create Meta_data_sheet and copy META columns as-is
    meta = wb.create_sheet("Meta_data_sheet")
    # Header
    for c_idx, key in enumerate(META_COLS, start=1):
        meta.cell(row=1, column=c_idx, value=key).font = header_font
    # Body
    for r_idx, row in enumerate(data, start=2):
        for c_idx, key in enumerate(META_COLS, start=1):
            meta.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
    meta.sheet_state = "veryHidden"

    # Transform Data -> TestPlan (remove META cols and reorder)
    ws.title = "TestPlan"

    # Build a mapping from column header to index
    header_map = {ws.cell(row=1, column=ci).value: ci for ci in range(1, ws.max_column + 1)}

    # Extract data for MAIN_ORDER only
    rows_out = []
    for r in range(2, ws.max_row + 1):
        out_row = []
        for key in MAIN_ORDER:
            ci = header_map.get(key)
            out_row.append(ws.cell(row=r, column=ci).value if ci else "")
        rows_out.append(out_row)

    # Clear ws and write MAIN_ORDER
    ws.delete_rows(1, ws.max_row)
    for c_idx, key in enumerate(MAIN_ORDER, start=1):
        cell = ws.cell(row=1, column=c_idx, value=key)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill

    for r_idx, out_row in enumerate(rows_out, start=2):
        for c_idx, val in enumerate(out_row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Mandatory numbering inside cells for two columns in TestPlan
    def col_index(name):
        for ci in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=ci).value == name:
                return ci
        return None

    for name in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
        ci = col_index(name)
        if ci:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=ci).value
                ws.cell(row=r, column=ci, value=renumber_block(v))

    # Formatting
    # Wrap text columns
    wrap_cols = {name: col_index(name) for name in WRAP_COLS}
    for name, ci in wrap_cols.items():
        if not ci:
            continue
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=ci).alignment = left_top

    # Header alignments already set; ensure center vertical
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).alignment = center

    # Numeric/text alignment
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.alignment = center_top
            elif cell.alignment.wrap_text:
                # keep left_top
                pass
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")

    # Borders
    apply_table_borders(ws)

    # Data validation for Code Generation (Required / Not)
    cg_ci = col_index("Code Generation (Required / Not)")
    if cg_ci:
        col_letter = ws.cell(row=1, column=cg_ci).column_letter
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showDropDown=True)
        dv.error = "Select one of: Required, Blank, Not Required"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    # Final validation: only TestPlan and Meta_data_sheet must exist
    assert "Data" not in [s.title for s in wb.worksheets], "Data sheet must not exist"

    # Ensure directories
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
