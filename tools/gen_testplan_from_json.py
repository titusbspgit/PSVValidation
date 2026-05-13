#!/usr/bin/env python3
import json
import sys
import os
import argparse
from datetime import datetime, timezone, timedelta
from zipfile import ZipFile
import subprocess

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
    'Hidden_Header_Includes',
    'Hidden_Macro_Define',
    'Hidden_Skip_Array_Definition'
]

WRAP_COLS = {
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
}

VALIDATION_COL = 'Code Generation (Required / Not)'
VALIDATION_LIST = ['Required', 'Blank', 'Not Required']

BLUE_FILL = PatternFill('solid', fgColor='4F81BD')
THIN = Side(style='thin', color='000000')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def ensure_numbered(text: str) -> str:
    if text is None:
        return ''
    lines = [l for l in str(text).splitlines() if l is not None]
    if not lines:
        return ''
    out = []
    n = 1
    for l in lines:
        ls = l.strip()
        # If already starts with N. or N) consider it numbered
        if len(ls) > 2 and (ls[0].isdigit()) and (ls[1] in {'.', ')'}):
            out.append(ls)
        else:
            out.append(f"{n}. {ls}")
            n += 1
    return "\n".join(out)


def union_keys_preserve_order(rows):
    seen = []
    idx = {}
    for r in rows:
        if isinstance(r, dict):
            for k in r.keys():
                if k not in idx:
                    idx[k] = True
                    seen.append(k)
    return seen


def auto_fit(ws):
    # approximate auto-fit by max string length
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            for l in s.splitlines():
                max_len = max(max_len, len(l))
        width = min(120, max(10, int(max_len * 0.9) + 2))
        ws.column_dimensions[col_letter].width = width


def apply_borders(ws):
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in r:
            c.border = BORDER_ALL


def set_header_style(ws):
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.fill = BLUE_FILL


def set_data_alignment(ws):
    # vertical top for all data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.alignment = Alignment(vertical='top', horizontal='left', wrap_text=c.alignment.wrap_text)
    # Center the Index column if present
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    if 'Index' in headers:
        col = headers.index('Index') + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)


def set_wraps(ws):
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    wrap_cols_idx = [i+1 for i, h in enumerate(headers) if h in WRAP_COLS]
    for r in range(2, ws.max_row + 1):
        for ci in wrap_cols_idx:
            c = ws.cell(row=r, column=ci)
            c.alignment = Alignment(wrap_text=True, vertical='top', horizontal=c.alignment.horizontal)


def adjust_row_heights(ws):
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    wrap_cols_idx = [i+1 for i, h in enumerate(headers) if h in WRAP_COLS]
    base = 15
    for r in range(2, ws.max_row + 1):
        lines = 1
        for ci in wrap_cols_idx:
            v = ws.cell(row=r, column=ci).value
            if v is None:
                continue
            cnt = str(v).count('\n') + 1
            lines = max(lines, cnt)
        ws.row_dimensions[r].height = base * lines


def add_validation(ws):
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    if VALIDATION_COL not in headers:
        return
    col = headers.index(VALIDATION_COL) + 1
    start = 2
    end = ws.max_row
    if end < start:
        return
    col_letter = get_column_letter(col)
    dv = DataValidation(type="list", formula1='"' + ", ".join(VALIDATION_LIST) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start}:{col_letter}{end}")


def validate_xlsx(path: str) -> None:
    with ZipFile(path, 'r') as z:
        # must contain core workbook parts
        assert 'xl/workbook.xml' in z.namelist(), 'workbook.xml missing'


def commit_file(path: str, commit_msg: str, author_name: str, author_email: str):
    subprocess.run(['git', 'config', '--local', 'user.name', author_name], check=True)
    subprocess.run(['git', 'config', '--local', 'user.email', author_email], check=True)
    subprocess.run(['git', 'add', path], check=True)
    # Only commit if there is a change staged
    diff = subprocess.run(['git', 'diff', '--cached', '--quiet'])
    if diff.returncode != 0:
        subprocess.run(['git', 'commit', '-m', commit_msg], check=True)
        subprocess.run(['git', 'push'], check=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--ip-name', required=True)
    ap.add_argument('--outdir', required=True)
    ap.add_argument('--commit-msg', required=True)
    ap.add_argument('--author-name', required=True)
    ap.add_argument('--author-email', required=True)
    args = ap.parse_args()

    # STEP 1: Read and validate JSON (array)
    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data, list) or len(data) == 0:
        print('ERROR: JSON must be a non-empty array', file=sys.stderr)
        sys.exit(1)

    # STEP 2: Normalize schema (union of keys, preserve first-seen order)
    key_order = union_keys_preserve_order(data)

    # STEP 3: Create workbook with single sheet named Data
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Headers: use key_order
    for c, k in enumerate(key_order, start=1):
        ws.cell(row=1, column=c, value=k)
    # Rows
    for r, obj in enumerate(data, start=2):
        for c, k in enumerate(key_order, start=1):
            ws.cell(row=r, column=c, value=obj.get(k, ''))

    # Base formatting
    ws.freeze_panes = 'A2'
    set_header_style(ws)
    auto_fit(ws)

    # STEP 5: Create Meta_data_sheet and copy META columns AS-IS if present
    meta_ws = wb.create_sheet('Meta_data_sheet')
    # Collect present META cols preserving META_COLS order
    present_meta = [k for k in META_COLS if k in key_order]
    if present_meta:
        for c, k in enumerate(present_meta, start=1):
            meta_ws.cell(row=1, column=c, value=k)
        rr = 2
        for obj in data:
            for c, k in enumerate(present_meta, start=1):
                meta_ws.cell(row=rr, column=c, value=obj.get(k, ''))
            rr += 1
    # Very hidden
    meta_ws.sheet_state = 'veryHidden'

    # STEP 7: Normalize MAIN sheet within the same Data sheet
    # Build final order for the main sheet
    final_headers = [h for h in MAIN_ORDER if h in key_order]
    # Also include any extra non-meta, non-main keys at the end preserving order
    extras = [k for k in key_order if (k not in META_COLS and k not in final_headers)]
    final_headers = final_headers + extras

    # Overwrite the sheet content to match final headers and row data
    # First, clear existing content by overwriting
    # Write headers
    for c, k in enumerate(final_headers, start=1):
        ws.cell(row=1, column=c, value=k)
    # Remove extra columns if any
    while ws.max_column > len(final_headers):
        ws.delete_cols(len(final_headers)+1)

    # Rebuild row data in the new order
    for r_idx in range(2, ws.max_row + 1):
        pass  # we will rewrite below
    # Fetch original objects aligned with rows
    for r, obj in enumerate(data, start=2):
        for c, k in enumerate(final_headers, start=1):
            ws.cell(row=r, column=c, value=obj.get(k, ''))

    # STEP 7A: Formatting
    # Wrap target columns, numbering for steps and validation inside cells
    hdr_to_idx = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column+1)}
    for field in ['Test Steps / Procedure', 'Validation / Acceptance Criteria']:
        if field in hdr_to_idx:
            ci = hdr_to_idx[field]
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=ci).value
                ws.cell(row=r, column=ci, value=ensure_numbered(val))

    set_wraps(ws)
    set_data_alignment(ws)
    apply_borders(ws)
    auto_fit(ws)
    adjust_row_heights(ws)

    # Rename Data to TestPlan
    ws.title = 'TestPlan'

    # Add strict validation on Code Generation (Required / Not)
    add_validation(ws)

    # STEP 7B: Ensure no sheet named Data exists
    for s in wb.sheetnames:
        if s == 'Data':
            print('ERROR: Data sheet still exists after normalization', file=sys.stderr)
            sys.exit(2)

    # STEP 8: Save with IST timestamp
    ist = timezone(timedelta(hours=5, minutes=30))
    ts = datetime.now(ist).strftime('%Y%m%d_%H%M%S')
    out_name = f"{args.ip_name}_TestPlan_{ts}.xlsx"
    out_path = os.path.join(args.outdir, out_name)
    os.makedirs(args.outdir, exist_ok=True)
    wb.save(out_path)

    # STEP 9/Validation: verify OOXML
    validate_xlsx(out_path)

    # STEP 10: Commit only the finalized Excel file
    commit_file(out_path, args.commit_msg, args.author_name, args.author_email)

    print(json.dumps({
        'status': 'SUCCESS',
        'rows': len(data),
        'cols': len(final_headers),
        'output': out_path
    }))

if __name__ == '__main__':
    main()
