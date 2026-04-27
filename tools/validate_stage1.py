#!/usr/bin/env python3
import sys
from argparse import ArgumentParser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

META_SHEET = "Meta_data_sheet"
TESTPLAN_SHEET = "TestPlan"


def read_output_path(file_path: str) -> str:
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read().strip()


def assert_true(cond: bool, msg: str):
    if not cond:
        print(f"VALIDATION FAILED: {msg}")
        sys.exit(1)


def main():
    ap = ArgumentParser()
    ap.add_argument('--output-path-file', required=True)
    args = ap.parse_args()

    xlsx_path = read_output_path(args.output_path_file)
    wb = load_workbook(xlsx_path)

    # 1) Sheets present
    assert_true(TESTPLAN_SHEET in wb.sheetnames, f"Missing sheet: {TESTPLAN_SHEET}")
    assert_true(META_SHEET in wb.sheetnames, f"Missing sheet: {META_SHEET}")

    meta = wb[META_SHEET]
    # Very Hidden check (openpyxl stores as 'veryHidden')
    assert_true(getattr(meta, 'sheet_state', '') == 'veryHidden', "Meta_data_sheet must be Very Hidden")

    ws = wb[TESTPLAN_SHEET]

    # 2) Header order
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    expected = MAIN_ORDER
    assert_true(headers[:len(expected)] == expected, f"Header order mismatch. Got {headers}")

    # 3) Wrap text enabled on data cells for specified columns
    header_to_col = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for name in WRAP_COLUMNS:
        assert_true(name in header_to_col, f"Missing wrap column: {name}")
        c = header_to_col[name]
        # Check first data row if exists
        if ws.max_row >= 2:
            al = ws.cell(row=2, column=c).alignment
            assert_true(bool(al and al.wrap_text), f"Wrap text not enabled for column: {name}")

    # 4) Header styling: bold, centered, blue fill, white font
    for c in range(1, len(expected) + 1):
        cell = ws.cell(row=1, column=c)
        assert_true(bool(cell.font and cell.font.bold), "Header not bold")
        assert_true(bool(cell.alignment and cell.alignment.horizontal == 'center' and cell.alignment.vertical == 'center'), "Header alignment incorrect")
        fg = getattr(cell.fill, 'fgColor', None)
        assert_true(fg is not None and (fg.rgb or fg.indexed is not None), "Header fill missing")
        # Normalize fg rgb; openpyxl may store ARGB, accept endswith '0070C0'
        rgb = (fg.rgb or '').upper()
        assert_true(rgb.endswith('0070C0'), f"Header fill color not blue 0070C0 (got {rgb})")
        # Font color white
        col = (cell.font.color.rgb if cell.font and cell.font.color else None)
        if col is not None:
            assert_true(col.upper().endswith('FFFFFF'), f"Header font color not white (got {col})")

    # 5) Borders thin on all populated cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            b = ws.cell(row=r, column=c).border
            assert_true(b is not None, "Missing border")
            for side in (b.left, b.right, b.top, b.bottom):
                assert_true(side is not None and side.style == 'thin', "Border not thin")

    # 6) Data validation for Code Generation
    code_col = header_to_col.get("Code Generation (Required / Not)")
    assert_true(code_col is not None, "Missing Code Generation column")
    col_letter = get_column_letter(code_col)
    found = False
    for dv in ws.data_validations.dataValidation:
        if isinstance(dv, DataValidation) and dv.type == 'list' and dv.formula1 == '"Required,Not Required,"':
            # Check if applies to our column
            for rng in dv.ranges:
                if str(rng).startswith(f"{col_letter}"):
                    found = True
                    break
        if found:
            break
    assert_true(found, "Data validation for Code Generation column not found or incorrect")

    print("VALIDATION PASSED")


if __name__ == '__main__':
    main()
