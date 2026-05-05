import os, re, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

MAIN_COLUMNS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

TESTPLAN_DIR = os.path.join('Test_Output', 'GPIO', 'TestPlan')


def find_latest_xlsx():
    if not os.path.isdir(TESTPLAN_DIR):
        raise FileNotFoundError(f'Missing directory: {TESTPLAN_DIR}')
    cand = [f for f in os.listdir(TESTPLAN_DIR) if f.startswith('GPIO_TestPlan_') and f.endswith('.xlsx')]
    if not cand:
        raise FileNotFoundError('No generated Excel found')
    cand.sort()  # filenames contain YYYYMMDD_HHMMSS, so lexical sort works
    return os.path.join(TESTPLAN_DIR, cand[-1])


def check_sheets(wb):
    names = wb.sheetnames
    if set(names) != {'TestPlan', 'Meta_data_sheet'}:
        raise AssertionError(f'Invalid sheet set: {names}')
    if 'Data' in names:
        raise AssertionError('Sheet named Data must not exist')
    # Visibility
    if wb['Meta_data_sheet'].sheet_state.lower() != 'veryhidden':
        raise AssertionError('Meta_data_sheet must be veryHidden')
    if wb['TestPlan'].sheet_state.lower() != 'visible':
        raise AssertionError('TestPlan must be visible')


def check_headers(ws):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    if headers != MAIN_COLUMNS:
        raise AssertionError(f'Header mismatch. Got {headers}')


def find_col(ws, header):
    for c in range(1, ws.max_column+1):
        if ws.cell(row=1, column=c).value == header:
            return c
    raise KeyError(f'Header not found: {header}')


def check_wrap_and_numbering(ws):
    for header in ('Test Steps / Procedure', 'Validation / Acceptance Criteria'):
        c = find_col(ws, header)
        for r in range(2, ws.max_row+1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if not val:
                continue
            # must be wrapped
            if not (cell.alignment and cell.alignment.wrapText):
                raise AssertionError(f'Wrap not enabled at {header} row {r}')
            lines = [x for x in str(val).splitlines() if x.strip()]
            if not lines:
                continue
            # numeric 1., 2., ...
            for i, line in enumerate(lines, start=1):
                if not re.match(rf'^{i}\.\s', line):
                    raise AssertionError(f'Line numbering missing at {header} row {r}: {line!r}')


def check_data_validation(ws):
    # Only one DV on Code Generation column
    dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
    if len(dvs) != 1:
        raise AssertionError(f'Expected exactly 1 data validation, found {len(dvs)}')
    dv = dvs[0]
    if dv.type != 'list' or (dv.formula1 or '').replace(' ', '') != '"Required,Blank,NotRequired"':
        raise AssertionError(f'Unexpected DV rule: type={dv.type}, formula1={dv.formula1}')
    # check range is only the Code Generation column rows 2..max_row
    col_idx = find_col(ws, 'Code Generation (Required / Not)')
    col_letter = get_column_letter(col_idx)
    expected_range = f'{col_letter}2:{col_letter}{ws.max_row}'
    # DataValidation can contain multiple ranges; combine and compare set
    ranges = set()
    for sqref in str(dv.sqref).split():
        ranges.add(sqref)
    if ranges != {expected_range}:
        raise AssertionError(f'DV range mismatch. Got {ranges}, expected {expected_range}')


def check_impacted_registers(ws):
    c = find_col(ws, 'Impacted Registers')
    for r in range(2, ws.max_row+1):
        val = ws.cell(row=r, column=c).value
        if not val:
            continue
        if 'MIZAR_' in str(val):
            raise AssertionError(f'Visible Impacted Registers contains macro at row {r}: {val!r}')


def main():
    xlsx = find_latest_xlsx()
    wb = load_workbook(xlsx, data_only=True)
    check_sheets(wb)
    ws = wb['TestPlan']
    check_headers(ws)
    check_wrap_and_numbering(ws)
    check_data_validation(ws)
    check_impacted_registers(ws)
    print(f'VALIDATION_OK: {os.path.basename(xlsx)}')

if __name__ == '__main__':
    main()
