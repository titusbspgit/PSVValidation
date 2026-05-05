#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os
import sys
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Embedded Test Plan JSON (source of truth for this workflow)
JSON_INPUT = r'''{
  "metadata": {
    "ip_name": "GPIO",
    "repo": "titusbspgit/PSVValidation",
    "branch": "main",
    "base_path": "TestRepo/gpio",
    "generation_timestamp": "AUTO_IST"
  },
  "testcases": [
    {
      "test_id": "TC1",
      "test_name": "gpio_reg_wr_rd_test",
      "objective": "GPIO register default and read/write behavior",
      "description": "Checks default values and verifies masked write/read behavior across configured GPIO registers.",
      "preconditions": ["Platform initialized", "GPIO block accessible"],
      "steps": [
        "Start the test",
        "For each configured register in the list, read the value and compare masked data with the expected default",
        "For each of six patterns, write the masked pattern to each writable register that is not skipped",
        "Read back each register with read mask applied and compute the expected value combining written and default bits",
        "Compare read data with the expected value for each register",
        "Report pass if no default or write/read mismatches are detected; otherwise report fail"
      ],
      "expected_results": [
        "Masked read equals the expected default value for each register",
        "Masked readback equals computed expected value for each pattern and register"
      ],
      "pass_fail_criteria": "Pass if both default and write/read checks have zero failures; otherwise fail",
      "registers": [
        {"name": "MIZAR_GPIO_GP0_GPIO_8"}, {"name": "MIZAR_GPIO_GP0_GPIO_9"}, {"name": "MIZAR_GPIO_GP0_GPIO_10"}, {"name": "MIZAR_GPIO_GP0_GPIO_11"}, {"name": "MIZAR_GPIO_GP0_GPIO_12"}, {"name": "MIZAR_GPIO_GP0_GPIO_13"}, {"name": "MIZAR_GPIO_GP0_GPIO_14"}, {"name": "MIZAR_GPIO_GP0_GPIO_15"}, {"name": "MIZAR_GPIO_GP0_GPIO_16"}, {"name": "MIZAR_GPIO_GP0_GPIO_17"}, {"name": "MIZAR_GPIO_GP0_GPIO_18"}, {"name": "MIZAR_GPIO_GP0_GPIO_19"}, {"name": "MIZAR_GPIO_GP0_GPIO_20"}, {"name": "MIZAR_GPIO_GP0_GPIO_21"}, {"name": "MIZAR_GPIO_GP0_GPIO_22"}, {"name": "MIZAR_GPIO_GP0_GPIO_23"}, {"name": "MIZAR_GPIO_GP0_GPIO_24"}, {"name": "MIZAR_GPIO_GP0_GPIO_25"}, {"name": "MIZAR_GPIO_GP0_GPIO_26"}, {"name": "MIZAR_GPIO_GP0_GPIO_27"}, {"name": "MIZAR_GPIO_GP0_GPIO_28"}, {"name": "MIZAR_GPIO_GP0_GPIO_29"}, {"name": "MIZAR_GPIO_GP0_GPIO_30"}, {"name": "MIZAR_GPIO_GP0_GPIO_31"}, {"name": "MIZAR_GPIO_GP0_GPIO_32"}, {"name": "MIZAR_GPIO_GP0_GPIO_33"}, {"name": "MIZAR_GPIO_GP0_GPIO_34"}, {"name": "MIZAR_GPIO_GP0_GPIO_35"}, {"name": "MIZAR_GPIO_GP0_GPIO_36"}, {"name": "MIZAR_GPIO_GP0_GPIO_37"}, {"name": "MIZAR_GPIO_GP0_GPIO_38"}, {"name": "MIZAR_GPIO_GP0_GPIO_39"},
        {"name": "MIZAR_GPIO_GPIO_INTR_RAW_STCLR1"},
        {"name": "MIZAR_GPIO_GP0_INTR1_INTR_EN1"},
        {"name": "MIZAR_GPIO_GP0_INTR1_INTR_STS1"},
        {"name": "MIZAR_GPIO_GP0_INTR2_INTR_EN1"},
        {"name": "MIZAR_GPIO_GP0_INTR2_INTR_STS1"},
        {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1"}, {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2"}, {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3"}, {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4"},
        {"name": "MIZAR_GPIO_GPIO_DOUT_GROUP1"}, {"name": "MIZAR_GPIO_GPIO_DOUT_GROUP2"}, {"name": "MIZAR_GPIO_GPIO_DOUT_GROUP3"}, {"name": "MIZAR_GPIO_GPIO_DOUT_GROUP4"},
        {"name": "MIZAR_GPIO_GPIO_DIN_GROUP1"}, {"name": "MIZAR_GPIO_GPIO_DIN_GROUP2"}, {"name": "MIZAR_GPIO_GPIO_DIN_GROUP3"}, {"name": "MIZAR_GPIO_GPIO_DIN_GROUP4"}
      ],
      "dependencies": ["Common GPIO access library"],
      "source_files": [
        {"path": "TestRepo/gpio/gpio_reg_wr_rd_test", "github_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/gpio_reg_wr_rd_test"}
      ],
      "tags": ["smoke", "register", "sanity"]
    },
    {
      "test_id": "TC2",
      "test_name": "test_gpio_negedge_intr_en",
      "objective": "GPIO negative-edge interrupt enable and raw/group status handling",
      "description": "Verifies negative-edge interrupts for GPIOs 8–39, including per-pin and group status handling and clearing.",
      "preconditions": ["Interrupt controller initialized", "GPIO pads 8–39 connected"],
      "steps": [
        "Enable the interrupt controller for the selected GPIO block",
        "Enable INTR_EN1 to route the GPIO interrupt to the system",
        "Drive the pad output register at 0xA0243ffc high to set a known level",
        "For GPIO_8 through GPIO_39, set input enable, negative-edge enable, and clear raw status",
        "For each bit, clear the group raw status and enable only that bit in INTR1_INTR_EN1",
        "Arm the wait flag, then create a falling edge by toggling 0xA0243ffc from all ones to all ones except the target bit",
        "Wait for the interrupt with a timeout; on timeout, record an error",
        "In the handler, restore pad output high via 0xA0243ffc and read the per-pin register",
        "Verify the input bit is low and the raw status is asserted; read INTR1_INTR_STS1 and confirm the corresponding bit is set",
        "Clear the per-pin raw status, clear the group raw status using gpio_intr_raw_stclr1, and verify INTR1_INTR_STS1 is zero",
        "Clear RAW_STCR1 for the selected GPIO interrupt source and clear the interrupt at the controller"
      ],
      "expected_results": [
        "Each pin’s interrupt must arrive within the configured timeout window",
        "DIN bit reads low after a falling edge and group status indicates the active bit",
        "After clearing per-pin and group raw status, the group status reads zero"
      ],
      "pass_fail_criteria": "No timeouts; correct DIN, raw, and group status behavior; proper clears",
      "registers": [
        {"name": "MIZAR_LSS_SYSREG_INTR_EN1"},
        {"name": "MIZAR_GPIO_GP0_GPIO_8"},
        {"name": "MIZAR_GPIO_GPIO_INTR_RAW_STCLR1"},
        {"name": "MIZAR_GPIO_GP0_INTR1_INTR_EN1"},
        {"name": "MIZAR_GPIO_GP0_INTR1_INTR_STS1"},
        {"name": "MIZAR_LSS_SYSREG_RAW_STCR1"}
      ],
      "dependencies": ["GIC driver", "SYSREG access"],
      "source_files": [
        {"path": "TestRepo/gpio/test_gpio_negedge_intr_en", "github_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en"}
      ],
      "tags": ["regression", "interrupt", "negedge", "pads", "per-pin"]
    },
    {
      "test_id": "TC3",
      "test_name": "test_gpio_pedge_all_pads_en",
      "objective": "GPIO positive-edge interrupt enable on all pads with group service",
      "description": "Enables positive-edge interrupts across GPIOs 8–39, verifies group status on interrupt, clears per-pin raw status, and checks system interrupt clear.",
      "preconditions": ["Interrupt controller initialized", "GPIO pads 8–39 connected"],
      "steps": [
        "Enable the interrupt controller for the selected GPIO block",
        "Enable INTR_EN1 to route the GPIO interrupt to the system",
        "For GPIO_8 through GPIO_39, set positive-edge enable for each pin",
        "Configure GPIO_IO_CTRL_GROUP1 to GPIO_IO_CTRL_GROUP4 for input mode",
        "Enable all bits in INTR1_INTR_EN1",
        "For each bit, drive the pad low at 0xA0243ffc, arm the wait flag, then drive high to generate a rising edge",
        "Wait for the interrupt with a timeout; on timeout, record an error and stop testing",
        "In the handler, read INTR1_INTR_STS1 and temporarily mask INTR1_INTR_EN1",
        "Confirm the group status indicates an interrupt occurred; otherwise record an error",
        "Clear per-pin raw status for all pins and verify INTR1_INTR_STS1 reads zero",
        "Clear RAW_STCR1 for the selected GPIO interrupt source and verify the status bit is cleared",
        "Re-enable INTR1_INTR_EN1 and clear the interrupt at the controller"
      ],
      "expected_results": [
        "An interrupt is observed within the timeout after a rising edge for each pin",
        "Group status is nonzero during service and zero after clears",
        "System interrupt status bit is cleared after RAW_STCR1 write"
      ],
      "pass_fail_criteria": "No timeouts; proper group status and clear behavior; proper system interrupt clear",
      "registers": [
        {"name": "MIZAR_LSS_SYSREG_INTR_EN1"},
        {"name": "MIZAR_GPIO_GP0_GPIO_8"},
        {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1"}, {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2"}, {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3"}, {"name": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4"},
        {"name": "MIZAR_GPIO_GP0_INTR1_INTR_EN1"},
        {"name": "MIZAR_GPIO_GP0_INTR1_INTR_STS1"},
        {"name": "MIZAR_LSS_SYSREG_RAW_STCR1"}
      ],
      "dependencies": ["GIC driver", "SYSREG access"],
      "source_files": [
        {"path": "TestRepo/gpio/test_gpio_pedge_all_pads_en", "github_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en"}
      ],
      "tags": ["regression", "interrupt", "posedge", "pads", "all-pads"]
    }
  ]
}'''

# Constants
MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]
META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

BLUE_FILL = PatternFill(fill_type="solid", start_color="FF4472C4", end_color="FF4472C4")
BORDER_THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
TOP_LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
TOP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=False)
RIGHT_TOP = Alignment(horizontal='right', vertical='top', wrap_text=False)


def now_ist():
    tz = ZoneInfo("Asia/Kolkata")
    return datetime.now(tz)


def to_ist_timestamp_parts(dt):
    return dt.strftime("%Y%m%d"), dt.strftime("%H%M%S")


def number_items(items):
    return "\n".join(f"{i+1}. {str(items[i])}" for i in range(len(items)))


def macro_to_reg(name: str) -> str:
    n = name
    if n.startswith("MIZAR_GPIO_GP0_GPIO_"):
        return n.replace("MIZAR_GPIO_GP0_", "")  # GPIO_X
    if n.startswith("MIZAR_GPIO_GPIO_"):
        return n.replace("MIZAR_GPIO_", "")     # GPIO_*
    if n == "MIZAR_GPIO_GP0_INTR1_INTR_EN1":
        return "INTR1_INTR_EN1"
    if n == "MIZAR_GPIO_GP0_INTR1_INTR_STS1":
        return "INTR1_INTR_STS1"
    if n == "MIZAR_GPIO_GP0_INTR2_INTR_EN1":
        return "INTR2_INTR_EN1"
    if n == "MIZAR_GPIO_GP0_INTR2_INTR_STS1":
        return "INTR2_INTR_STS1"
    if n.startswith("MIZAR_LSS_SYSREG_"):
        return n.replace("MIZAR_LSS_SYSREG_", "LSS_SYSREG_")
    if n.startswith("MIZAR_GPIO_GP0_"):
        return n.replace("MIZAR_GPIO_GP0_", "")
    return n


def build_rows(data):
    rows = []
    meta_rows = []
    testcases = data.get("testcases", [])
    if not isinstance(testcases, list) or len(testcases) == 0:
        raise ValueError("No testcases found in input JSON")

    for idx, tc in enumerate(testcases, start=1):
        tags = tc.get("tags", []) or []
        mode = "Interrupt" if any(t.lower() == "interrupt" for t in tags) else "NA"
        remarks = tc.get("pass_fail_criteria", "")
        steps = tc.get("steps", []) or []
        exp = tc.get("expected_results", []) or []
        exp_all = exp + ([remarks] if remarks else [])
        regs_raw = [macro_to_reg(r.get("name", "")) for r in (tc.get("registers", []) or []) if r.get("name")]
        # Deduplicate preserving order
        seen = set()
        regs = []
        for r in regs_raw:
            if r not in seen:
                regs.append(r)
                seen.add(r)
        regs_str = ", ".join(regs)

        row = {
            "Index": idx,
            "SS / Module": data.get("metadata", {}).get("ip_name", ""),
            "Feature": tc.get("objective", ""),
            "Test Case Name": tc.get("test_name", ""),
            "Test Description": tc.get("description", ""),
            "Speed": tc.get("speed", "NA") if tc.get("speed") else "NA",
            "Mode": mode,
            "Memory Start Offset": tc.get("memory_start_offset", ""),
            "Memory End Offset": tc.get("memory_end_offset", ""),
            "Remarks": remarks,
            "Test Steps / Procedure": number_items(steps),
            "Impacted Registers": regs_str,
            "Validation / Acceptance Criteria": number_items(exp_all),
            "Code Generation (Required / Not)": "",
        }
        rows.append(row)

        meta_row = {
            "Hidden_Test_Case_Name": tc.get("test_name", ""),
            "Hidden_Test_Description": tc.get("description", ""),
            "Hidden_Remarks": remarks,
            "Hidden_Test_Steps_Procedure": "\n".join(steps),
            "Hidden_Impacted_Registers": regs_str,
            "Hidden_Validation_Acceptance_Criteria": "\n".join(exp_all),
        }
        meta_rows.append(meta_row)

    return rows, meta_rows


def autofit_columns(ws):
    # Estimate width by max string length per column (capped)
    max_width = {}
    for row in ws.iter_rows(values_only=True):
        for col_idx, value in enumerate(row, start=1):
            text = "" if value is None else str(value)
            w = min(max(10, len(text) + 2), 120)
            if col_idx not in max_width:
                max_width[col_idx] = w
            else:
                if w > max_width[col_idx]:
                    max_width[col_idx] = w
    for col_idx, width in max_width.items():
        ws.column_dimensions[chr(64+col_idx) if col_idx<=26 else (chr(64+(col_idx-1)//26) + chr(65+((col_idx-1)%26)))] .width = width


def apply_borders(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN


def find_col_index(ws, header_name):
    for c in range(1, ws.max_column+1):
        if (ws.cell(row=1, column=c).value or '') == header_name:
            return c
    return None


def build_workbook(data_rows, meta_rows, ip_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"  # authoritative staging

    # Write combined columns for staging: include MAIN + META (ensures we can copy META exactly), then only MAIN kept
    staging_headers = MAIN_COLUMNS + META_COLUMNS
    ws.append(staging_headers)
    for r, m in zip(data_rows, meta_rows):
        row_vals = [r.get(col, "") for col in MAIN_COLUMNS] + [m.get(col, "") for col in META_COLUMNS]
        ws.append(row_vals)

    # Base formatting on Data
    for c in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL
    ws.freeze_panes = "A2"

    # Create META sheet and copy META columns as-is
    meta = wb.create_sheet("Meta_data_sheet")
    meta.append(META_COLUMNS)
    for m in meta_rows:
        meta.append([m.get(col, "") for col in META_COLUMNS])
    # Very hidden
    meta.sheet_state = 'veryHidden'

    # Normalize MAIN sheet on the same worksheet: remove META columns (keep only MAIN order)
    # We'll reconstruct rows keeping only MAIN columns
    # First, rename sheet to TestPlan as required
    ws.title = "TestPlan"

    # Now enforce wrap and alignments for specific columns
    wrap_cols = {"Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"}
    for r in range(2, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            hdr = ws.cell(row=1, column=c).value
            cell = ws.cell(row=r, column=c)
            if hdr in wrap_cols:
                cell.alignment = TOP_LEFT_WRAP
            elif hdr == "Index":
                cell.alignment = CENTER
            else:
                cell.alignment = TOP_LEFT

    # Since staging has MAIN+META, we must drop META columns from visible sheet while preserving data
    # Strategy: create a temporary in-memory table for MAIN, clear sheet, write MAIN only
    main_table = []
    main_table.append(MAIN_COLUMNS)
    for r in range(2, ws.max_row+1):
        row = []
        values = {ws.cell(row=1, column=c).value: ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)}
        for col in MAIN_COLUMNS:
            row.append(values.get(col, ""))
        main_table.append(row)
    # Clear all cells
    ws.delete_rows(1, ws.max_row)
    # Write back MAIN only
    for ridx, row in enumerate(main_table, start=1):
        ws.append(row)
        if ridx == 1:
            for c in range(1, len(row)+1):
                cell = ws.cell(row=1, column=c)
                cell.font = HEADER_FONT
                cell.alignment = CENTER
                cell.fill = BLUE_FILL
    ws.freeze_panes = "A2"

    # Data validation for Code Generation (Required / Not)
    col_idx = find_col_index(ws, "Code Generation (Required / Not)")
    if col_idx:
        last_row = ws.max_row
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        dv = DataValidation(type="list", formula1='"Required, Blank, Not Required"', allow_blank=True, showErrorMessage=True)
        dv.error = "Select one of: Required, Blank, Not Required"
        dv.promptTitle = "Code Generation"
        dv.prompt = "Choose requirement"
        rng = f"{col_letter}2:{col_letter}{last_row}"
        dv.add(rng)
        ws.add_data_validation(dv)

    # Borders and alignment again for final structure
    for r in range(2, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            hdr = ws.cell(row=1, column=c).value
            if hdr in {"Index"}:
                cell.alignment = CENTER
            elif hdr in wrap_cols:
                cell.alignment = TOP_LEFT_WRAP
            else:
                cell.alignment = TOP_LEFT
    apply_borders(ws)

    # Autofit
    autofit_columns(ws)

    # Final visibility enforcement: only TestPlan and Meta_data_sheet
    names = [s.title for s in wb.worksheets]
    if "Data" in names:
        # Delete if still present
        for s in wb.worksheets:
            if s.title == "Data":
                wb.remove(s)
                break

    # Validate workbook structure
    allowed = set(["TestPlan", "Meta_data_sheet"])
    current = set([s.title for s in wb.worksheets])
    if not (allowed == current):
        raise RuntimeError(f"Unexpected worksheets present: {current}")

    return wb


def validate_xlsx_binary(data: bytes) -> None:
    with ZipFile(BytesIO(data), 'r') as z:
        # Must contain key OOXML parts
        required_parts = {"[Content_Types].xml", "xl/workbook.xml"}
        names = set(z.namelist())
        for part in required_parts:
            if part not in names:
                raise ValueError(f"XLSX validation failed: missing {part}")


def save_and_commit(wb, ip_name: str):
    ist_now = now_ist()
    ymd, hms = to_ist_timestamp_parts(ist_now)
    file_name = f"{ip_name}_TestPlan_{ymd}_{hms}.xlsx"
    out_dir = os.path.join("Test_Output", ip_name, "TestPlan")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, file_name)

    # Save to bytes for validation
    bio = BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    validate_xlsx_binary(data)

    # Write to disk
    with open(out_path, 'wb') as f:
        f.write(data)

    # Commit using git
    os.system('git config user.email "github-actions[bot]@users.noreply.github.com"')
    os.system('git config user.name "github-actions[bot]"')
    os.system(f"git add '{out_path}'")
    # Use [skip ci] to prevent re-trigger
    msg = "Add GPIO TestPlan Excel generated by Stage1 (IST timestamped) [skip ci]"
    rc = os.system(f"git commit -m \"{msg}\"")
    # If nothing to commit, still exit 0
    os.system("git push origin HEAD:main")

    print(out_path)


def main():
    data = json.loads(JSON_INPUT)
    ip_name = data.get("metadata", {}).get("ip_name", "GPIO")
    rows, meta_rows = build_rows(data)
    wb = build_workbook(rows, meta_rows, ip_name)
    save_and_commit(wb, ip_name)


if __name__ == '__main__':
    main()
