#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stage1 deterministic generator: Convert provided Test Plan JSON to Excel (.xlsx) strictly following
Ag-Emb-Mpsoc-Stage1 rules and commit-ready structure.

Key behaviors:
- Validate JSON structure (object with non-empty 'tests' array)
- Build Data sheet from union of keys (first-seen order), one row per test
- Create Meta_data_sheet with META columns; set sheet_state to veryHidden (no formatting applied)
- Rename Data -> TestPlan, drop META columns, enforce MAIN column order ONLY
- Visual formatting only on TestPlan: blue header fill, bold, wraps, borders, alignment
- Add drop-down list to 'Code Generation (Required / Not)' with values: Required, Not Required, and allow blank
- Set workbook metadata 'created' to IST timestamp provided by --timestamp (YYYYMMDD_HHMMSS)
- Filename: <IP_NAME>_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx where timestamp is IST provided by --timestamp
- Output path: Test_Output/<IP_NAME>/TestPlan/

No mutation of values. Arrays/objects are JSON-encoded to preserve exact content.
"""

import argparse
import json
from copy import deepcopy
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Embedded JSON input (exactly as provided)
TEST_PLAN_JSON_STR = r'''{
  "ip_name": "SPI",
  "repo": "titusbspgit/PSVValidation",
  "branch": "main",
  "subdirectory": "TestRepo/spi",
  "generated_timestamp_ist": "2026-04-27T00:00:00+05:30",
  "tests": [
    {
      "Index": 1,
      "SS / Module": "SPI",
      "Feature": "Full duplex, four-wire synchronous transfers",
      "Test Case Name": "spi_pio_full_duplex",
      "Test Description": "Validates full-duplex operation using interrupts by configuring the controller, performing a handshake, and repeatedly handling receive and transmit FIFO interrupts to move data through the data register. The test concludes by checking the final status value.",
      "Speed": "NA",
      "Mode": "Interrupt",
      "Memory Start Offset": "0xA0000018",
      "Memory End Offset": "0xA000001C",
      "Remarks": "The selected SPI instance is compile-time controlled. Optional debug printing may be enabled. FIFO service counts depend on configured thresholds.",
      "Test Steps / Procedure": [
        "Entry: invoke the test entrypoint.",
        "Enable the system interrupt for the selected SPI instance via the system register at 0xA000001C, then enable the corresponding interrupt controller line.",
        "Configure the SPI controller.",
        "Perform the SPI VIP handshake.",
        "Repeat eight times: set the loop flag; wait in a delay loop until the flag is cleared by an interrupt service routine; then increment a data counter.",
        "Interrupt path: read the masked interrupt status register; if the receive FIFO condition is indicated, read the required number of entries from the data register; otherwise if the transmit FIFO condition is indicated, write the required number of entries to the data register.",
        "Clear the system interrupt via the system register at 0xA0000018 and clear the interrupt controller line.",
        "Obtain the final status and end the test with that status."
      ],
      "Impacted Registers": "DATA_REG, MIS",
      "Validation / Acceptance Criteria": [
        "The test passes when the final status value equals zero.",
        "Receive or transmit FIFO interrupts occur such that the wait loop exits as the flag is cleared.",
        "If the final status value is nonzero, the test fails."
      ],
      "Code Generation (Required / Not)": "",
      "Hidden_Test_Case_Name": "spi_pio_full_duplex",
      "Hidden_Test_Description": "Initializes SPI, enables SysReg SPIx interrupt (0xA000001C) and GIC line, calls spi_cntrl_config(), then spi_vip_handshake(). For j=0..7: sets int_pend=1 and busy-waits with wait_on(10) until Default_IRQHandler sets int_pend=0. In Default_IRQHandler: sets data_addr=MIZAR_SPI_DATA_REG, mis_addr=MIZAR_SPI_MIS; reads MaskedInterrupt=read_reg(MIZAR_SPI_MIS); if (MaskedInterrupt & 0x2)==0x2, reads SPI_RX_FIFO_THLD entries from MIZAR_SPI_DATA_REG; else if ((MaskedInterrupt & 0x1)==0x1), writes SPI_TX_FIFO_THLD entries of 'count' into MIZAR_SPI_DATA_REG; then writes 0xA0000018 with the SPI instance mask and calls GIC_ClearIRQ(). After the loop, err1=spi_vip_scbd_status(); finish(err1).",
      "Hidden_Remarks": "SPI instance selected by preprocessor (SPI0..SPI3). DEBUG_DISPALY controls prints. FIFO service depths use SPI_RX_FIFO_THLD and SPI_TX_FIFO_THLD.",
      "Hidden_Test_Steps_Procedure": [
        "int test_case() entry.",
        "#ifdef SPI0: write_reg(0xA000001C, 0x01000000); GIC_EnableIRQ(76);",
        "#ifdef SPI1: write_reg(0xA000001C, 0x02000000); GIC_EnableIRQ(77);",
        "#ifdef SPI2: write_reg(0xA000001C, 0x04000000); GIC_EnableIRQ(78);",
        "#ifdef SPI3: write_reg(0xA000001C, 0x08000000); GIC_EnableIRQ(79);",
        "Call spi_cntrl_config();",
        "Call spi_vip_handshake();",
        "Loop j=0..7: set int_pend=1; while(int_pend){ wait_on(10); } then count++.",
        "After loop: err1 = spi_vip_scbd_status(); finish(err1).",
        "ISR Default_IRQHandler(): set int_pend=0; data_addr = MIZAR_SPI_DATA_REG; mis_addr = MIZAR_SPI_MIS;",
        "Read MaskedInterrupt = read_reg(MIZAR_SPI_MIS);",
        "If (MaskedInterrupt & 0x2)==0x2: for i=0..SPI_RX_FIFO_THLD-1: rd_data = read_reg(MIZAR_SPI_DATA_REG);",
        "Else if (MaskedInterrupt & 0x1)==0x1: for j=0..SPI_TX_FIFO_THLD-1: write_reg(MIZAR_SPI_DATA_REG, count);",
        "#ifdef SPI0: write_reg(0xA0000018, 0x01000000); GIC_ClearIRQ(76);",
        "#ifdef SPI1: write_reg(0xA0000018, 0x02000000); GIC_ClearIRQ(77);",
        "#ifdef SPI2: write_reg(0xA0000018, 0x04000000); GIC_ClearIRQ(78);",
        "#ifdef SPI3: write_reg(0xA0000018, 0x08000000); GIC_ClearIRQ(79);"
      ],
      "Hidden_Impacted_Registers": "MIZAR_SPI_DATA_REG, MIZAR_SPI_MIS",
      "Hidden_Validation_Acceptance_Criteria": [
        "Interrupt status check: if ((MaskedInterrupt & 0x2)==0x2) then RX FIFO path executes; else if ((MaskedInterrupt & 0x1)==0x1) then TX FIFO path executes.",
        "Loop exits only when ISR sets int_pend=0.",
        "finish(err1) is invoked; PASS if err1==0; FAIL otherwise."
      ]
    },
    {
      "Index": 2,
      "SS / Module": "SPI",
      "Feature": "Two independent DMA interfaces are available – between TX FIFO and system DMAC, and one between the RX FIFO and system DMAC",
      "Test Case Name": "spi_pio_rx_dma_tx",
      "Test Description": "Configures the controller and DMA to transmit using the DMA engine while using interrupts to coordinate progress. Test data are preloaded into memory, transferred to the data register by DMA, and completion is observed via interrupts before the final status is checked.",
      "Speed": "NA",
      "Mode": "DMA, Interrupt",
      "Memory Start Offset": "0xA0000018",
      "Memory End Offset": "0xA1700054",
      "Remarks": "DMA and interrupt settings are controlled by constants, with multiple DMA-related registers at 0xA1700008 through 0xA1700054 written prior to configuration. The interrupt mask register is written to enable and then disable the relevant interrupt.",
      "Test Steps / Procedure": [
        "Entry: invoke the test entrypoint.",
        "Enable the system interrupt for the selected SPI instance via the system register at 0xA000001C, then enable the corresponding interrupt controller line.",
        "Initialize DMA-related control registers at addresses 0xA1700008 through 0xA1700054.",
        "Prepare DMA parameters including channel number, source memory address, transfer count, direction, and request source.",
        "Preload the source memory region with a sequence of data words.",
        "Configure the SPI controller.",
        "Perform the SPI VIP handshake.",
        "Set the destination to the data register and configure the DMA transfer from memory to the data register.",
        "Disable DMA to apply settings, then enable the transmit interrupt in the mask register.",
        "Set the loop flag; wait in a delay loop until the flag is cleared by an interrupt service routine.",
        "Interrupt path: read the masked interrupt status register; if the transmit condition is indicated, disable interrupts via the mask register.",
        "Clear the system interrupt via the system register at 0xA0000018 and clear the interrupt controller line.",
        "Obtain the final status and end the test with that status."
      ],
      "Impacted Registers": "DATA_REG, INTMSK, MIS",
      "Validation / Acceptance Criteria": [
        "The test passes when the final status value equals zero.",
        "A transmit interrupt occurs to allow the wait loop to complete, and the interrupt is then masked.",
        "If the final status value is nonzero, the test fails."
      ],
      "Code Generation (Required / Not)": "",
      "Hidden_Test_Case_Name": "spi_pio_rx_dma_tx",
      "Hidden_Test_Description": "Enables SysReg SPIx interrupt (0xA000001C) and GIC line, writes a series of DMA-related system registers at 0xA1700008..0xA1700054 with 0x1, sets DMA params (ch_num=0, src_addr=0xA0243E6C, src_xcnt=8, tx_rx=0, tc_intr_en=0, spi_mst=0, src_req=SPI_TX_SRC_REQ). Preloads memory at src_addr with 8 words (0xaaaaaaa1+i). Calls spi_cntrl_config() and spi_vip_handshake(). Sets dst_addr=MIZAR_SPI_DATA_REG; calls dma_config(ch_num, src_addr, dst_addr, src_xcnt, tx_rx, tc_intr_en, src_req, spi_mst); calls dma_disable(). Enables TX interrupt by write_reg(MIZAR_SPI_IMSC,0x1); sets int_pend=1 and busy-waits with wait_on(5) until ISR clears flag. ISR reads read_reg(MIZAR_SPI_MIS) into MaskedInterrupt; if ((MaskedInterrupt & 0x1)==0x1) then prints debug and writes write_reg(MIZAR_SPI_IMSC,0x0) to disable interrupt. Then writes to 0xA0000018 and calls GIC_ClearIRQ(). After loop: err1=spi_vip_scbd_status(); finish(err1).",
      "Hidden_Remarks": "Uses DMAC/system registers at 0xA1700008..0xA1700054; enables and then disables SPI TX interrupt via MIZAR_SPI_IMSC; src_req uses SPI_TX_SRC_REQ; no explicit RX FIFO reads in ISR.",
      "Hidden_Test_Steps_Procedure": [
        "int test_case() entry.",
        "#ifdef SPI0: write_reg(0xA000001C, 0x01000000); GIC_EnableIRQ(76);",
        "#ifdef SPI1: write_reg(0xA000001C, 0x02000000); GIC_EnableIRQ(77);",
        "#ifdef SPI2: write_reg(0xA000001C, 0x04000000); GIC_EnableIRQ(78);",
        "#ifdef SPI3: write_reg(0xA000001C, 0x08000000); GIC_EnableIRQ(79);",
        "Initialize DMA/system regs: write_reg(0xA1700008,1); 0xA170000C=1; 0xA1700014=1; 0xA1700018=1; 0xA170001C=1; 0xA1700020=1; 0xA1700024=1; 0xA1700028=1; 0xA170002C=1; 0xA1700030=1; 0xA1700034=1; 0xA1700038=1; 0xA170003C=1; 0xA1700044=1; 0xA1700048=1; 0xA1700050=1; 0xA1700054=1;",
        "Set ch_num=0; src_addr=0xA0243E6C; src_xcnt=8; tx_rx=0; tc_intr_en=0; spi_mst=0; src_req=SPI_TX_SRC_REQ;",
        "for i=0..7: write_reg(src_addr + i*4, 0xaaaaaaa1 + i);",
        "Call spi_cntrl_config();",
        "Call spi_vip_handshake();",
        "dst_addr = MIZAR_SPI_DATA_REG;",
        "dma_config(ch_num, src_addr, dst_addr, src_xcnt, tx_rx, tc_intr_en, src_req, spi_mst);",
        "dma_disable();",
        "write_reg(MIZAR_SPI_IMSC, 0x1);",
        "int_pend=1; while(int_pend){ wait_on(5); }",
        "ISR Default_IRQHandler(): int_pend=0; data_addr=MIZAR_SPI_DATA_REG; mis_addr=MIZAR_SPI_MIS; MaskedInterrupt=read_reg(MIZAR_SPI_MIS);",
        "If ((MaskedInterrupt & 0x1)==0x1): write_reg(MIZAR_SPI_IMSC,0x0);",
        "#ifdef SPI0: write_reg(0xA0000018, 0x01000000); GIC_ClearIRQ(76);",
        "#ifdef SPI1: write_reg(0xA0000018, 0x02000000); GIC_ClearIRQ(77);",
        "#ifdef SPI2: write_reg(0xA0000018, 0x04000000); GIC_ClearIRQ(78);",
        "#ifdef SPI3: write_reg(0xA0000018, 0x08000000); GIC_ClearIRQ(79);",
        "After loop: err1=spi_vip_scbd_status(); finish(err1)."
      ],
      "Hidden_Impacted_Registers": "MIZAR_SPI_DATA_REG, MIZAR_SPI_IMSC, MIZAR_SPI_MIS",
      "Hidden_Validation_Acceptance_Criteria": [
        "Interrupt status check: if ((MaskedInterrupt & 0x1)==0x1) then TX interrupt path executes and interrupt is masked by writing MIZAR_SPI_IMSC=0.",
        "Loop exits only when ISR sets int_pend=0.",
        "finish(err1) is invoked; PASS if err1==0; FAIL otherwise."
      ]
    },
    {
      "Index": 3,
      "SS / Module": "SPI",
      "Feature": "Register default values and R/W behavior",
      "Test Case Name": "spi_reg_wr_rd_test",
      "Test Description": "Performs register default-value verification and masked write/readback checks across a defined register list. For multiple data patterns, writable fields are updated and the expected readback is computed using the read and write masks before comparing to the actual value. The test completes by reporting success only if all checks pass.",
      "Speed": "NA",
      "Mode": "NA",
      "Memory Start Offset": "NA",
      "Memory End Offset": "NA",
      "Remarks": "Non-readable and non-writable registers are explicitly skipped according to provided masks. The covered register list and their default, read, and write masks are sourced from the associated arrays.",
      "Test Steps / Procedure": [
        "Entry: invoke the test entrypoint.",
        "Default-value phase: iterate through the configured register list; for each readable register, read the value and compare against the documented default; count failures.",
        "Write/readback phase: for each data pattern, iterate through the configured register list; for each writable register, write the pattern; then re-iterate to read back and compute the expected value using the read and write masks and defaults; compare and count mismatches.",
        "Conclude by reporting pass if both failure counts are zero; otherwise report fail."
      ],
      "Impacted Registers": "RSM_CR0, RSM_CR1, RSM_DMA_CR, RSM_BAUD_RATE, RSM_TIMEOUT, RSM_FIFO, RSM_DMA, TX_FIFO_LEVEL_REG, RX_FIFO_LEVEL_REG, TRQ_LEVEL_REG, STS, INTMSK, RIS, MIS, IC, RSM_TRQ_DATA_REG, RSM_TXTEST_DATA_REG, DATA_REG, RSM_TS_CTRL, SPI_TLEAD, SPI_TTRAIL, SPI_TIDLE",
      "Validation / Acceptance Criteria": [
        "For each readable register, the initial value must match the documented default.",
        "For each writable field, the readback must equal the masked combination of the written pattern and preserved default bits according to the read and write masks.",
        "The test passes only when both the default-value failure count and the write-read mismatch count are zero."
      ],
      "Code Generation (Required / Not)": "",
      "Hidden_Test_Case_Name": "spi_reg_wr_rd_test",
      "Hidden_Test_Description": "test_case(): calls chk_rst_val() then chk_rd_wr(); if (def_fail_cnt>0 || wr_fail_cnt>0) finish(1) else finish(0). chk_rst_val(): for i=0..CNT-1, addr=addr_array[i]; if (read_mask_array[i]==0) continue; data_rd=read_reg(addr); if (data_rd==default_value_array[i]) PASS else def_fail_cnt++. chk_rd_wr(): for each pattern in {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}: write phase: for i=0..CNT-1, if skip_array[i]==1 continue; if (write_mask_array[i]==0) continue; else write_reg(addr_array[i], data_wr). readback phase: for i=0..CNT-1, if skip_array[i]==1 continue; if (write_mask_array[i]==0) continue; if (read_mask_array[i]==0) continue; data_rd=read_reg(addr_array[i]); wr_n=(write_mask_array[i]^0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if (data_rd==exp_val) PASS else wr_fail_cnt++.",
      "Hidden_Remarks": "Covers 22 registers from addr_array[] using masks from read_mask_array[] and write_mask_array[]. Skips non-readable/non-writable entries and any with skip_array[i]==1.",
      "Hidden_Test_Steps_Procedure": [
        "int test_case() entry.",
        "Call chk_rst_val().",
        "In chk_rst_val(): for i=0..CNT-1: addr=addr_array[i]; if (read_mask_array[i]==0) continue; data_rd=read_reg(addr); compare with default_value_array[i]; on mismatch def_fail_cnt++.",
        "Call chk_rd_wr().",
        "In chk_rd_wr(): For each pattern in {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}:",
        "Write phase: for i=0..CNT-1: addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0) continue; else write_reg(addr, data_wr).",
        "Readback phase: for i=0..CNT-1: addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0) continue; if (read_mask_array[i]==0) continue; data_rd=read_reg(addr); wr_n=(write_mask_array[i]^0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); compare; on mismatch wr_fail_cnt++.",
        "Return to test_case(): if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0)."
      ],
      "Hidden_Impacted_Registers": "MIZAR_SPI_RSM_CR0, MIZAR_SPI_RSM_CR1, MIZAR_SPI_RSM_DMA_CR, MIZAR_SPI_RSM_BAUD_RATE, MIZAR_SPI_RSM_TIMEOUT, MIZAR_SPI_RSM_FIFO, MIZAR_SPI_RSM_DMA, MIZAR_SPI_TX_FIFO_LEVEL_REG, MIZAR_SPI_RX_FIFO_LEVEL_REG, MIZAR_SPI_TRQ_LEVEL_REG, MIZAR_SPI_STS, MIZAR_SPI_IMSC, MIZAR_SPI_RIS, MIZAR_SPI_MIS, MIZAR_SPI_IC, MIZAR_SPI_RSM_TRQ_DATA_REG, MIZAR_SPI_RSM_TXTEST_DATA_REG, MIZAR_SPI_DATA_REG, MIZAR_SPI_RSM_TS_CTRL, MIZAR_SPI_SPI_TLEAD, MIZAR_SPI_SPI_TTRAIL, MIZAR_SPI_SPI_TIDLE",
      "Hidden_Validation_Acceptance_Criteria": [
        "Default check: if (data_rd == default_value_array[i]) PASS else def_fail_cnt++.",
        "Write/readback expected: exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | ((write_mask_array[i]^0xffffffff) & read_mask_array[i] & default_value_array[i])); PASS if data_rd == exp_val else wr_fail_cnt++.",
        "Final decision: finish(0) if def_fail_cnt==0 && wr_fail_cnt==0; else finish(1)."
      ]
    }
  ]
}'''

MAIN_COLUMNS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLUMNS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--timestamp', required=True, help='IST timestamp in YYYYMMDD_HHMMSS')
    p.add_argument('--ip-name', required=True, help='IP name for file naming (must be SPI)')
    return p.parse_args()


def json_value_to_cell(v):
    # Preserve original data without mutation
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        return v
    # For arrays/objects, keep JSON string to avoid altering semantics
    return json.dumps(v, ensure_ascii=False)


def compute_union_keys_preserve_order(objs):
    seen = []
    sset = set()
    for o in objs:
        for k in o.keys():
            if k not in sset:
                sset.add(k)
                seen.append(k)
    return seen


def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            for line in str(v).split('\n'):
                if len(line) > max_len:
                    max_len = len(line)
        width = min(max(10, max_len + 2), 90)
        ws.column_dimensions[col[0].column_letter].width = width


def apply_table_borders(ws):
    thin = Side(border_style='thin', color='FF000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def build_workbook(data_tests, ist_timestamp_str):
    wb = Workbook()
    # Set workbook metadata created to IST timestamp
    try:
        created_dt = datetime.strptime(ist_timestamp_str, '%Y%m%d_%H%M%S')
        wb.properties.created = created_dt
    except Exception:
        pass

    ws = wb.active
    ws.title = 'Data'

    # Normalize schema for Data sheet
    union_keys = compute_union_keys_preserve_order(data_tests)

    # Base Data sheet
    for c, key in enumerate(union_keys, start=1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for r, item in enumerate(data_tests, start=2):
        for c, key in enumerate(union_keys, start=1):
            ws.cell(row=r, column=c, value=json_value_to_cell(item.get(key, '')))
    ws.freeze_panes = 'A2'

    # Meta_data_sheet (no formatting)
    meta_ws = wb.create_sheet('Meta_data_sheet')
    for c, key in enumerate(META_COLUMNS, start=1):
        meta_ws.cell(row=1, column=c, value=key)
    for r, item in enumerate(data_tests, start=2):
        for c, key in enumerate(META_COLUMNS, start=1):
            meta_ws.cell(row=r, column=c, value=json_value_to_cell(item.get(key, '')))
    meta_ws.sheet_state = 'veryHidden'

    # Transform Data -> TestPlan, drop META columns, enforce MAIN column order ONLY
    header_row = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    keep_headers = [h for h in header_row if h not in META_COLUMNS]
    final_headers = [h for h in MAIN_COLUMNS if h in keep_headers]

    tmp = wb.create_sheet('TMP')
    for c, key in enumerate(final_headers, start=1):
        cell = tmp.cell(row=1, column=c, value=key)
        cell.font = Font(bold=True)
    idx_map = {header_row[i]: i + 1 for i in range(len(header_row))}
    for r in range(2, ws.max_row + 1):
        for c, key in enumerate(final_headers, start=1):
            src_col = idx_map.get(key)
            val = ws.cell(row=r, column=src_col).value if src_col else ''
            tmp.cell(row=r, column=c, value=val)

    # Replace Data with TestPlan
    wb.remove(ws)
    tmp.title = 'TestPlan'
    ws = tmp

    # Strict formatting for TestPlan
    header_font = Font(bold=True, color='FFFFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='FF1F4E79')

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill

    wrap_cols = {
        'Test Description',
        'Remarks',
        'Test Steps / Procedure',
        'Validation / Acceptance Criteria'
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            header = ws.cell(row=1, column=cell.column).value
            if header in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            elif header == 'Index':
                cell.alignment = Alignment(vertical='top', horizontal='center')
            else:
                cell.alignment = Alignment(vertical='top', horizontal='left')

    ws.freeze_panes = 'A2'
    autofit_columns(ws)
    apply_table_borders(ws)

    # Data validation for Code Generation column
    code_gen_col = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == 'Code Generation (Required / Not)':
            code_gen_col = i
            break
    if code_gen_col is not None and ws.max_row >= 2:
        dv = DataValidation(type='list', formula1='"Required,Not Required"', allow_blank=True, showErrorMessage=True)
        rng = f"{ws.cell(row=2, column=code_gen_col).coordinate}:{ws.cell(row=ws.max_row, column=code_gen_col).coordinate}"
        dv.add(rng)
        ws.add_data_validation(dv)

    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--timestamp', required=True)
    ap.add_argument('--ip-name', required=True)
    args = ap.parse_args()

    # Validate JSON
    try:
        parsed = json.loads(TEST_PLAN_JSON_STR)
    except Exception as e:
        raise SystemExit(f'Invalid JSON input: {e}')
    if not isinstance(parsed, dict) or 'tests' not in parsed or not isinstance(parsed['tests'], list) or len(parsed['tests']) == 0:
        raise SystemExit('Invalid or empty JSON: expecting an object with non-empty "tests" array')

    tests = deepcopy(parsed['tests'])

    wb = build_workbook(tests, args.timestamp)

    out_dir = Path('Test_Output') / args.ip_name / 'TestPlan'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"{args.ip_name}_TestPlan_{args.timestamp}.xlsx"
    out_path = out_dir / out_name

    wb.save(out_path.as_posix())
    print(out_path.as_posix())


if __name__ == '__main__':
    main()
