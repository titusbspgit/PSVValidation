import json
import os
import zipfile
from datetime import datetime, timezone, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ----------------------
# Input JSON (array)
# ----------------------
json_rows = [
  {
    "Index": "1",
    "SS / Module": "I2C",
    "Feature": "I2C register reset and read/write compliance",
    "Test Case Name": "i2c_reg_rd_wr_test",
    "Test Description": "This test verifies default reset values of I2C/SMBus registers and validates masked write/read behavior across all defined registers.",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Addresses marked non-readable or non-writable are explicitly skipped. Addresses listed in skip_array are not accessed.",
    "Test Steps / Procedure": "1) Initialize the test and perform default register value checks across the defined device registers.\n2) For each test pattern, write to each writable device register and then read back to validate masked write/read behavior.\n3) Determine the result: pass if no mismatches are found; otherwise fail.",
    "Impacted Registers": "DEV_CTRL, TSFR_CTRL, SLV_ADDR, TGT_SLV_ADDR, i2c_mstr_code, I2C_BYTE_CNT, SF_HCNT, SF_LCNT, I2C_HS_HCNT, I2C_HS_LCNT, RIS, MASK_INTR, INTR_STS, INTR_CLR, TAS, TX_FIFO_THLD, RX_FIFO_THLD, DMA_CTRL, FF, TX_FIFO_LVL, RX_FIFO_LVL, I2C_MSTR_STS, I2C_FLTR_SEL, I2C_CURRENT_BYTECNT, I2C_SMB_SFTRST, SMB_HST_STS, SMB_HST_CTRL, SMB_HST_CMD, SMB_HST_DATA0, SMB_HST_DATA1, SMB_HST_BLOCK_DATA, SMB_PEC_DATA, SMB_SLAVE_WDATA, SMB_SLAVE_CMD, SMB_SLAVE_CTS, SMB_SLV, SMB_NOTIFY_ADDR, SMB_NOTIFY_LOW_BYTE, SMB_NOTIFY_HIGH_BYTE, SMB_DATA_HLDTIME, SMB_TIMEOUT_CNT, SMB_TMEXT_CNT, I2CSMB_DATA_SETUP",
    "Validation / Acceptance Criteria": "- All readable registers must equal their documented default values; any mismatch results in failure.\n- For each writable register and test pattern, the read value must match the expected masked composition; any mismatch results in failure.\n- Overall test passes only if both default value checks and write/read checks report zero failures.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "i2c_reg_rd_wr_test",
    "Hidden_Test_Description": "The test invokes chk_rst_val() to iterate over CNT addresses (addr_array). For each address: if read_mask_array[i] == 0, it skips read; else reads via read_reg(addr) and compares to default_value_array[i], incrementing def_fail_cnt on mismatch. Then chk_rd_wr() iterates over six patterns in chk_val. For each pattern and for each address: if skip_array[i] == 1, skip; if write_mask_array[i] == 0, skip; else write_reg(addr, data_wr). Next, for each address: if skip or not writable or not readable, skip; else read_reg(addr) and compute expected value as ((data_wr & read_mask & write_mask) | ((~write_mask) & read_mask & default_value)), compare and increment wr_fail_cnt on mismatch. Finally, if any failures occurred, finish(1); else finish(0).",
    "Hidden_Remarks": "Registers with read_mask_array[i] == 0 are skipped for reads; registers with write_mask_array[i] == 0 are skipped for writes and subsequent read checks; addresses with skip_array[i] == 1 are skipped entirely.",
    "Hidden_Test_Steps_Procedure": "Entry point: test_case()\n1. Call chk_rst_val().\n   1.1 Initialize loop index i from 0 to CNT-1.\n   1.2 For each i: addr = addr_array[i].\n       - If read_mask_array[i] == 0x00000000: print skip message (if enabled) and continue to next i.\n       - Else: data_rd = read_reg(addr).\n         - If (data_rd == default_value_array[i]): optionally print PASS.\n         - Else: increment def_fail_cnt and print failure details.\n   1.3 Exit loop after i == CNT.\n2. Call chk_rd_wr().\n   2.1 Initialize local variables and chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0x00000000, 0xA5A5A5A5, 0xffff0000}.\n   2.2 For each j in [0..5]:\n       - data_wr = chk_val[j].\n       - Write phase: For i in [0..CNT-1]:\n         a) addr = addr_array[i].\n         b) If skip_array[i] == 1: optionally print skip message and continue.\n         c) If write_mask_array[i] == 0x00000000: optionally print not-writable message and continue.\n         d) Else: write_reg(addr, data_wr); optionally print write details.\n       - Read/verify phase: For i in [0..CNT-1]:\n         a) addr = addr_array[i].\n         b) If skip_array[i] == 1: optionally print skip and continue.\n         c) If write_mask_array[i] == 0x00000000: optionally print not-writable and continue.\n         d) If read_mask_array[i] == 0x00000000: optionally print not-readable and continue.\n         e) Else: data_rd = read_reg(addr);\n            wr_n = (write_mask_array[i] ^ 0xffffffff);\n            exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]));\n            If (data_rd == exp_val): optionally print PASS; else: wr_fail_cnt++ and print failure details.\n   2.3 Complete all j iterations and return.\n3. Decision and finish:\n   - If (def_fail_cnt > 0 || wr_fail_cnt > 0): finish(1); else finish(0).\nTiming: No explicit waits/delays in executed path. Soft reset routine (with waits) is defined but not invoked.\nRegister accesses (macro usage via arrays): READ=read_reg(addr_array[i]); WRITE=write_reg(addr_array[i], data_wr).",
    "Hidden_Impacted_Registers": "MIZAR_I2C_DEV_CTRL, MIZAR_I2C_TSFR_CTRL, MIZAR_I2C_SLV_ADDR, MIZAR_I2C_TGT_SLV_ADDR, MIZAR_I2C_I2C_MSTR_CODE, MIZAR_I2C_I2C_BYTE_CNT, MIZAR_I2C_SF_HCNT, MIZAR_I2C_SF_LCNT, MIZAR_I2C_I2C_HS_HCNT, MIZAR_I2C_I2C_HS_LCNT, MIZAR_I2C_RIS, MIZAR_I2C_MASK_INTR, MIZAR_I2C_INTR_STS, MIZAR_I2C_INTR_CLR, MIZAR_I2C_TAS, MIZAR_I2C_TX_FIFO_THLD, MIZAR_I2C_RX_FIFO_THLD, MIZAR_I2C_DMA_CTRL, MIZAR_I2C_FF, MIZAR_I2C_TX_FIFO_LVL, MIZAR_I2C_RX_FIFO_LVL, MIZAR_I2C_I2C_MSTR_STS, MIZAR_I2C_I2C_FLTR_SEL, MIZAR_I2C_I2C_CURRENT_BYTECNT, MIZAR_I2C_I2C_SMB_SFTRST, MIZAR_I2C_SMB_HST_STS, MIZAR_I2C_SMB_HST_CTRL, MIZAR_I2C_SMB_HST_CMD, MIZAR_I2C_SMB_HST_DATA0, MIZAR_I2C_SMB_HST_DATA1, MIZAR_I2C_SMB_HST_BLOCK_DATA, MIZAR_I2C_SMB_PEC_DATA, MIZAR_I2C_SMB_SLAVE_WDATA, MIZAR_I2C_SMB_SLAVE_CMD, MIZAR_I2C_SMB_SLAVE_CTS, MIZAR_I2C_SMB_SLV, MIZAR_I2C_SMB_NOTIFY_ADDR, MIZAR_I2C_SMB_NOTIFY_LOW_BYTE, MIZAR_I2C_SMB_NOTIFY_HIGH_BYTE, MIZAR_I2C_SMB_DATA_HLDTIME, MIZAR_I2C_SMB_TIMEOUT_CNT, MIZAR_I2C_SMB_TMEXT_CNT, MIZAR_I2C_I2CSMB_DATA_SETUP",
    "Hidden_Validation_Acceptance_Criteria": "Default value check: data_rd == default_value_array[i] for all i where read_mask_array[i] != 0. Write/read check: For each pattern and each i where write_mask_array[i] != 0 and read_mask_array[i] != 0 and skip_array[i] == 0, read_reg(addr_array[i]) must equal ((data_wr & read_mask_array[i] & write_mask_array[i]) | ((~write_mask_array[i]) & read_mask_array[i] & default_value_array[i])). Overall PASS: def_fail_cnt == 0 and wr_fail_cnt == 0; else FAIL."
  },
  {
    "Index": "2",
    "SS / Module": "I2C",
    "Feature": "I2C DMA Logic, i2c_dma_logic",
    "Test Case Name": "test_i2c0_mst_i2c1_slv_dma_hs_md",
    "Test Description": "This test transfers a block of data from I2C0 master to I2C1 slave using DMA and verifies data integrity and interrupt handling.",
    "Speed": "NA",
    "Mode": "DMA, Interrupt, Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "LSS NIC slave interfaces are configured as non-secure by writing to multiple system registers before enabling I2C and DMA. The test waits for the I2C0 transfer-complete interrupt and then verifies that both I2C and system interrupt status are cleared.",
    "Test Steps / Procedure": "1) Configure system registers to allow non-secure access and enable I2C interrupts in the system controller.\n2) Initialize I2C0 and I2C1: clear pending interrupts, set control and address parameters, program byte count, and configure FIFO thresholds.\n3) Configure timing parameters and interrupt mask, and enable DMA on I2C0.\n4) Preload source SRAM with a 5-word pattern and configure DMA channel 0 to move data into the I2C0 block data register.\n5) Start the I2C0 transfer, poll the current byte count until it reaches zero, and wait for the transfer-complete interrupt.\n6) In the interrupt flow, confirm the transfer-complete status and clear both the I2C and system interrupt sources.\n7) Enable DMA on I2C1 and configure DMA channel 1 to move received data from the I2C1 block data register to destination SRAM.\n8) After DMA completion, compare the 5-word source and destination buffers in SRAM and report pass or fail.",
    "Impacted Registers": "DEV_CTRL, TSFR_CTRL, SLV_ADDR, TGT_SLV_ADDR, I2C_MSTR_CODE, I2C_BYTE_CNT, SF_HCNT, SF_LCNT, I2C_HS_HCNT, I2C_HS_LCNT, RIS, MASK_INTR, INTR_STS, INTR_CLR, TAS, TX_FIFO_THLD, RX_FIFO_THLD, DMA_CTRL, FF, TX_FIFO_LVL, RX_FIFO_LVL, I2C_MSTR_STS, I2C_FLTR_SEL, I2C_CURRENT_BYTECNT, I2C_SMB_SFTRST, SMB_HST_STS, SMB_HST_CTRL, SMB_HST_CMD, SMB_HST_DATA0, SMB_HST_DATA1, SMB_HST_BLOCK_DATA, SMB_PEC_DATA, SMB_SLAVE_WDATA, SMB_SLAVE_CMD, SMB_SLAVE_CTS, SMB_SLV, SMB_NOTIFY_ADDR, SMB_NOTIFY_LOW_BYTE, SMB_NOTIFY_HIGH_BYTE, SMB_DATA_HLDTIME, SMB_TIMEOUT_CNT, SMB_TMEXT_CNT, I2CSMB_DATA_SETUP, CH0_CTRL, CH0_SRC_ADDR, CH0_DEST_ADDR, CH0_SRC_XCNT, CH0_SRC_XMDFY, CH0_DEST_XMDFY, CH0_SRC_REQ, DMA_CH_EN, CH1_CTRL, CH1_SRC_ADDR, CH1_DEST_ADDR, CH1_SRC_XCNT, CH1_SRC_XMDFY, CH1_DEST_XMDFY, CH1_SRC_REQ, TC_INTR_EN, INTR_EN0, RAW_STCR0",
    "Validation / Acceptance Criteria": "- Transfer-complete interrupt must assert during the run, and both the peripheral and system interrupt status must clear successfully; otherwise the test fails.\n- After both DMA operations finish, the five data words read from the destination SRAM must match the five words written to the source SRAM; any mismatch fails the test.\n- The current byte count must reach zero before the interrupt is handled; failure to reach zero indicates an error.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_i2c0_mst_i2c1_slv_dma_hs_md",
    "Hidden_Test_Description": "The program sets int_pend=1, configures a range of LSS NIC registers at 0xA1700008..0xA1700054 to 0x1, enables GIC IRQs 80 and 81, and writes MIZAR_LSS_SYSREG_INTR_EN0 with I2C0 and I2C1 interrupt enables. It prints a banner. I2C configuration: write MIZAR_I2C0_FF and MIZAR_I2C1_FF to 0x3, clear MIZAR_I2C0_INTR_CLR and MIZAR_I2C1_INTR_CLR, set MIZAR_I2C0_DEV_CTRL and MIZAR_I2C1_DEV_CTRL to 0x38A, program MIZAR_I2C1_SLV_ADDR (0x37) and MIZAR_I2C0_TGT_SLV_ADDR (0x6E), set I2C byte counts to 0x5, set SF_LCNT/H_CNT for I2C0 and HS_LCNT/H_CNT, set MIZAR_I2C0_I2C_MSTR_CODE (0x2), and mask interrupts with MIZAR_I2C0_MASK_INTR. Set TX/RX FIFO thresholds for both I2C0 and I2C1 to 0x5. Enable MIZAR_I2C0_DMA_CTRL=0x2. DMA in: src_addr=SRAM_ADDR_1 (0xA0243F00), dest_addr=MIZAR_I2C0_SMB_HST_BLOCK_DATA; preload five 32-bit words into SRAM starting at SRAM_ADDR_1; program DMA CH0 registers: CH0_CTRL=0x8028028, CH0_SRC_ADDR=SRAM_ADDR_1, CH0_DEST_ADDR=I2C0 block data, CH0_SRC_XCNT=0x5, CH0_SRC_XMDFY=0x4, CH0_DEST_XMDFY=0x0, CH0_SRC_REQ=0x5; then DMA_CH_EN=0x1; call dma_disable() which waits while DMA_CH_EN is nonzero, polling until zero. Start I2C transfer by writing MIZAR_I2C0_TSFR_CTRL=0x2. Poll MIZAR_I2C0_I2C_CURRENT_BYTECNT until it reads 0, with wait_on(100) between reads. Wait while(int_pend) { print and wait_on(10); }. On interrupt, Default_IRQHandler sets int_pend=0, reads MIZAR_I2C0_INTR_STS; if ==0x0010, clears MIZAR_I2C0_INTR_CLR bit 0x10, clears GIC IRQs, writes MIZAR_LSS_SYSREG_RAW_STCR0 with LSS_SYSREG_RAW_STCR0_I2C0_INTERRUPT, waits 100, re-reads MIZAR_I2C0_INTR_STS and checks LSS_SYSREG RAW_STCR0 & LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT; requires both ==0x00; else increments test_err. If initial int_status !=0x0010 it increments test_err. After ISR: set MIZAR_I2C1_DMA_CTRL=0x1, program DMA CH1 to move from MIZAR_I2C1_SMB_HST_BLOCK_DATA to SRAM_ADDR_2 (0xA0243E00): CH1_CTRL=0x8024028, CH1_SRC_ADDR=I2C1 block data, CH1_DEST_ADDR=SRAM_ADDR_2, CH1_SRC_XCNT=0x5, CH1_SRC_XMDFY=0x0, CH1_DEST_XMDFY=0x4, CH1_SRC_REQ=0x6; DMA_CH_EN=0x2; call dma_disable(). Compare five words from SRAM_ADDR_1 and SRAM_ADDR_2; print success/failure and increment test_err on mismatches. wait_on(100) and finish(test_err).",
    "Hidden_Remarks": "Non-secure LSS NIC configuration is required prior to I2C/DMA operations. The test assumes interrupt line mapping for I2C0/1 to IRQs 80 and 81. Transfer complete is detected by I2C0 interrupt status 0x0010 and must be cleared in both I2C and system status registers.",
    "Hidden_Test_Steps_Procedure": "Entry point: test_case()\n1. Initialize and security setup:\n   1.1 int_pend = 1.\n   1.2 Write 0x1 to system addresses 0xA1700008, 0xA170000C, 0xA1700014, 0xA1700018, 0xA170001C, 0xA1700020, 0xA1700024, 0xA1700028, 0xA170002C, 0xA1700030, 0xA1700034, 0xA1700038, 0xA170003C, 0xA1700044, 0xA1700048, 0xA1700050, 0xA1700054.\n   1.3 GIC_EnableIRQ(80), GIC_EnableIRQ(81).\n   1.4 write_reg(MIZAR_LSS_SYSREG_INTR_EN0, LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT | LSS_SYSREG_INTR_EN0_I2C1_INTERRUPT).\n2. I2C configuration (both instances):\n   2.1 write_reg(MIZAR_I2C0_FF, 0x3); write_reg(MIZAR_I2C1_FF, 0x3).\n   2.2 write_reg(MIZAR_I2C0_INTR_CLR, 0xFFFFFFFF); write_reg(MIZAR_I2C1_INTR_CLR, 0xFFFFFFFF).\n   2.3 write_reg(MIZAR_I2C0_DEV_CTRL, 0x38A); write_reg(MIZAR_I2C1_DEV_CTRL, 0x38A).\n   2.4 write_reg(MIZAR_I2C1_SLV_ADDR, 0x37); write_reg(MIZAR_I2C0_TGT_SLV_ADDR, 0x6E).\n   2.5 write_reg(MIZAR_I2C0_I2C_BYTE_CNT, 0x5); write_reg(MIZAR_I2C1_I2C_BYTE_CNT, 0x5).\n   2.6 write_reg(MIZAR_I2C0_SF_LCNT, 0x2C); write_reg(MIZAR_I2C0_SF_HCNT, 0x18).\n   2.7 write_reg(MIZAR_I2C0_I2C_HS_LCNT, 0xC); write_reg(MIZAR_I2C0_I2C_HS_HCNT, 0x8).\n   2.8 write_reg(MIZAR_I2C0_I2C_MSTR_CODE, 0x2).\n   2.9 write_reg(MIZAR_I2C0_MASK_INTR, 0xFFFFFFEF).\n   2.10 write_reg(MIZAR_I2C0_TX_FIFO_THLD, 0x5); write_reg(MIZAR_I2C1_TX_FIFO_THLD, 0x5).\n   2.11 write_reg(MIZAR_I2C0_RX_FIFO_THLD, 0x5); write_reg(MIZAR_I2C1_RX_FIFO_THLD, 0x5).\n   2.12 write_reg(MIZAR_I2C0_DMA_CTRL, 0x2).\n3. DMA CH0 setup (SRAM to I2C0 block data):\n   3.1 src_addr = SRAM_ADDR_1 (0xA0243F00); dest_addr = MIZAR_I2C0_SMB_HST_BLOCK_DATA.\n   3.2 For i=0..4: tx_data[i] = i*5; write_reg(0xA0243F00 + i*4, tx_data[i]).\n   3.3 write_reg(MIZAR_DMA_CH0_CTRL, 0x8028028).\n   3.4 write_reg(MIZAR_DMA_CH0_SRC_ADDR, 0xA0243F00).\n   3.5 write_reg(MIZAR_DMA_CH0_DEST_ADDR, MIZAR_I2C0_SMB_HST_BLOCK_DATA).\n   3.6 write_reg(MIZAR_DMA_CH0_SRC_XCNT, 0x5).\n   3.7 write_reg(MIZAR_DMA_CH0_SRC_XMDFY, 0x4).\n   3.8 write_reg(MIZAR_DMA_CH0_DEST_XMDFY, 0x0).\n   3.9 write_reg(MIZAR_DMA_CH0_SRC_REQ, 0x5).\n   3.10 write_reg(MIZAR_DMA_DMA_CH_EN, 0x1).\n   3.11 Call dma_disable(): while(read_reg(MIZAR_DMA_DMA_CH_EN)) { wait_on(100); }.\n4. Start I2C transfer and poll:\n   4.1 write_reg(MIZAR_I2C0_TSFR_CTRL, 0x2).\n   4.2 data_rd = read_reg(MIZAR_I2C0_I2C_CURRENT_BYTECNT).\n       While (data_rd != 0): read_reg(MIZAR_I2C0_I2C_CURRENT_BYTECNT) into data_rd; print; wait_on(100).\n   4.3 while(int_pend) { print; wait_on(10); }.\n5. Interrupt handler (Default_IRQHandler):\n   5.1 int_pend = 0.\n   5.2 int_status = read_reg(MIZAR_I2C0_INTR_STS).\n       If (int_status == 0x0010):\n         a) write_reg(MIZAR_I2C0_INTR_CLR, 0x00000010).\n         b) GIC_ClearIRQ(80); GIC_ClearIRQ(81).\n         c) write_reg(MIZAR_LSS_SYSREG_RAW_STCR0, LSS_SYSREG_RAW_STCR0_I2C0_INTERRUPT).\n         d) wait_on(100).\n         e) int_status = read_reg(MIZAR_I2C0_INTR_STS).\n            int_status_lss = read_reg(MIZAR_LSS_SYSREG_RAW_STCR0) & LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT.\n            If (int_status == 0x00 && int_status_lss == 0x00): print success; else: test_err++.\n       Else: test_err++ and print error.\n6. DMA CH1 setup (I2C1 block data to SRAM):\n   6.1 write_reg(MIZAR_I2C1_DMA_CTRL, 0x1).\n   6.2 src_addr = MIZAR_I2C1_SMB_HST_BLOCK_DATA; dest_addr = SRAM_ADDR_2 (0xA0243E00).\n   6.3 write_reg(MIZAR_DMA_CH1_CTRL, 0x8024028).\n   6.4 write_reg(MIZAR_DMA_CH1_SRC_ADDR, MIZAR_I2C1_SMB_HST_BLOCK_DATA).\n   6.5 write_reg(MIZAR_DMA_CH1_DEST_ADDR, 0xA0243E00).\n   6.6 write_reg(MIZAR_DMA_CH1_SRC_XCNT, 0x5).\n   6.7 write_reg(MIZAR_DMA_CH1_SRC_XMDFY, 0x0).\n   6.8 write_reg(MIZAR_DMA_CH1_DEST_XMDFY, 0x4).\n   6.9 write_reg(MIZAR_DMA_CH1_SRC_REQ, 0x6).\n   6.10 write_reg(MIZAR_DMA_DMA_CH_EN, 0x2).\n   6.11 Call dma_disable(): while(read_reg(MIZAR_DMA_DMA_CH_EN)) { wait_on(100); }.\n7. Data integrity check:\n   7.1 For i=0..4: data_sent = read_reg(0xA0243F00 + i*4); data_rcvd = read_reg(0xA0243E00 + i*4); if equal: print success; else: print failure and test_err++.\n8. Finalize: wait_on(100); finish(test_err).",
    "Hidden_Impacted_Registers": "MIZAR_I2C0_DEV_CTRL, MIZAR_I2C0_TSFR_CTRL, MIZAR_I2C0_SLV_ADDR, MIZAR_I2C0_TGT_SLV_ADDR, MIZAR_I2C0_I2C_MSTR_CODE, MIZAR_I2C0_I2C_BYTE_CNT, MIZAR_I2C0_SF_HCNT, MIZAR_I2C0_SF_LCNT, MIZAR_I2C0_I2C_HS_HCNT, MIZAR_I2C0_I2C_HS_LCNT, MIZAR_I2C0_RIS, MIZAR_I2C0_MASK_INTR, MIZAR_I2C0_INTR_STS, MIZAR_I2C0_INTR_CLR, MIZAR_I2C0_TAS, MIZAR_I2C0_TX_FIFO_THLD, MIZAR_I2C0_RX_FIFO_THLD, MIZAR_I2C0_DMA_CTRL, MIZAR_I2C0_FF, MIZAR_I2C0_TX_FIFO_LVL, MIZAR_I2C0_RX_FIFO_LVL, MIZAR_I2C0_I2C_MSTR_STS, MIZAR_I2C0_I2C_FLTR_SEL, MIZAR_I2C0_I2C_CURRENT_BYTECNT, MIZAR_I2C0_I2C_SMB_SFTRST, MIZAR_I2C0_SMB_HST_STS, MIZAR_I2C0_SMB_HST_CTRL, MIZAR_I2C0_SMB_HST_CMD, MIZAR_I2C0_SMB_HST_DATA0, MIZAR_I2C0_SMB_HST_DATA1, MIZAR_I2C0_SMB_HST_BLOCK_DATA, MIZAR_I2C0_SMB_PEC_DATA, MIZAR_I2C0_SMB_SLAVE_WDATA, MIZAR_I2C0_SMB_SLAVE_CMD, MIZAR_I2C0_SMB_SLAVE_CTS, MIZAR_I2C0_SMB_SLV, MIZAR_I2C0_SMB_NOTIFY_ADDR, MIZAR_I2C0_SMB_NOTIFY_LOW_BYTE, MIZAR_I2C0_SMB_NOTIFY_HIGH_BYTE, MIZAR_I2C0_SMB_DATA_HLDTIME, MIZAR_I2C0_SMB_TIMEOUT_CNT, MIZAR_I2C0_SMB_TMEXT_CNT, MIZAR_I2C0_I2CSMB_DATA_SETUP, MIZAR_I2C1_DEV_CTRL, MIZAR_I2C1_TSFR_CTRL, MIZAR_I2C1_SLV_ADDR, MIZAR_I2C1_TGT_SLV_ADDR, MIZAR_I2C1_I2C_MSTR_CODE, MIZAR_I2C1_I2C_BYTE_CNT, MIZAR_I2C1_SF_HCNT, MIZAR_I2C1_SF_LCNT, MIZAR_I2C1_RIS, MIZAR_I2C1_MASK_INTR, MIZAR_I2C1_INTR_STS, MIZAR_I2C1_INTR_CLR, MIZAR_I2C1_TAS, MIZAR_I2C1_TX_FIFO_THLD, MIZAR_I2C1_RX_FIFO_THLD, MIZAR_I2C1_DMA_CTRL, MIZAR_I2C1_FF, MIZAR_I2C1_TX_FIFO_LVL, MIZAR_I2C1_RX_FIFO_LVL, MIZAR_I2C1_I2C_MSTR_STS, MIZAR_I2C1_I2C_FLTR_SEL, MIZAR_I2C1_I2C_CURRENT_BYTECNT, MIZAR_I2C1_I2C_SMB_SFTRST, MIZAR_I2C1_SMB_HST_STS, MIZAR_I2C1_SMB_HST_CTRL, MIZAR_I2C1_SMB_HST_CMD, MIZAR_I2C1_SMB_HST_DATA0, MIZAR_I2C1_SMB_HST_DATA1, MIZAR_I2C1_SMB_HST_BLOCK_DATA, MIZAR_I2C1_SMB_PEC_DATA, MIZAR_I2C1_SMB_SLAVE_WDATA, MIZAR_I2C1_SMB_SLAVE_CMD, MIZAR_I2C1_SMB_SLAVE_CTS, MIZAR_I2C1_SMB_SLV, MIZAR_I2C1_SMB_NOTIFY_ADDR, MIZAR_I2C1_SMB_NOTIFY_LOW_BYTE, MIZAR_I2C1_SMB_NOTIFY_HIGH_BYTE, MIZAR_I2C1_SMB_DATA_HLDTIME, MIZAR_I2C1_SMB_TIMEOUT_CNT, MIZAR_I2C1_SMB_TMEXT_CNT, MIZAR_I2C1_I2CSMB_DATA_SETUP, MIZAR_DMA_CH0_CTRL, MIZAR_DMA_CH0_SRC_ADDR, MIZAR_DMA_CH0_DEST_ADDR, MIZAR_DMA_CH0_SRC_XCNT, MIZAR_DMA_CH0_SRC_XMDFY, MIZAR_DMA_CH0_DEST_XMDFY, MIZAR_DMA_CH0_SRC_REQ, MIZAR_DMA_DMA_CH_EN, MIZAR_DMA_CH1_CTRL, MIZAR_DMA_CH1_SRC_ADDR, MIZAR_DMA_CH1_DEST_ADDR, MIZAR_DMA_CH1_SRC_XCNT, MIZAR_DMA_CH1_SRC_XMDFY, MIZAR_DMA_CH1_DEST_XMDFY, MIZAR_DMA_CH1_SRC_REQ, MIZAR_DMA_TC_INTR_EN, MIZAR_LSS_SYSREG_INTR_EN0, MIZAR_LSS_SYSREG_RAW_STCR0, LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT, LSS_SYSREG_INTR_EN0_I2C1_INTERRUPT",
    "Hidden_Validation_Acceptance_Criteria": "1) Interrupt: int_status from MIZAR_I2C0_INTR_STS must be 0x0010; after clearing MIZAR_I2C0_INTR_CLR with 0x10 and writing MIZAR_LSS_SYSREG_RAW_STCR0, both MIZAR_I2C0_INTR_STS and (MIZAR_LSS_SYSREG_RAW_STCR0 & LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT) must read 0x00. Any deviation increments test_err.\n2) Data integrity: For i=0..4, values read from SRAM_ADDR_2+(i*4) must equal values from SRAM_ADDR_1+(i*4). Mismatch increments test_err.\n3) Progress: MIZAR_I2C0_I2C_CURRENT_BYTECNT must reach 0 before ISR concludes; loop waits with wait_on(100)."
  },
  {
    "Index": "3",
    "SS / Module": "I2C",
    "Feature": "I2C DMA Logic, i2c_dma_logic",
    "Test Case Name": "test_i2c0_mst_i2c1_slv_dma_st_md",
    "Test Description": "This test performs a DMA-based transfer from I2C0 master to I2C1 slave in standard mode and verifies interrupt handling and data integrity.",
    "Speed": "Standard mode",
    "Mode": "DMA, Interrupt, Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "LSS NIC slave interfaces are configured to non-secure before running I2C and DMA operations. The test requires the transfer-complete interrupt to be raised and cleared correctly.",
    "Test Steps / Procedure": "1) Configure system registers for non-secure access and enable I2C interrupts in the system controller.\n2) Initialize I2C0 and I2C1: clear interrupts, set control and address registers, configure byte count and FIFO thresholds.\n3) Program standard-mode timing parameters and enable DMA on I2C0.\n4) Load a 5-word pattern into the source SRAM and configure DMA channel 0 to move it to the I2C0 block data register.\n5) Start the I2C0 transfer, poll the current byte count until it is zero, and wait for the transfer-complete interrupt.\n6) In the interrupt flow, confirm the transfer-complete condition and clear the interrupt in both the I2C peripheral and system controller.\n7) Enable DMA on I2C1 and configure DMA channel 1 to move the received data into destination SRAM.\n8) After DMA completes, compare the 5-word source and destination SRAM buffers and report pass or fail.",
    "Impacted Registers": "DEV_CTRL, TSFR_CTRL, SLV_ADDR, TGT_SLV_ADDR, I2C_MSTR_CODE, I2C_BYTE_CNT, SF_HCNT, SF_LCNT, RIS, MASK_INTR, INTR_STS, INTR_CLR, TAS, TX_FIFO_THLD, RX_FIFO_THLD, DMA_CTRL, FF, TX_FIFO_LVL, RX_FIFO_LVL, I2C_MSTR_STS, I2C_FLTR_SEL, I2C_CURRENT_BYTECNT, I2C_SMB_SFTRST, SMB_HST_STS, SMB_HST_CTRL, SMB_HST_CMD, SMB_HST_DATA0, SMB_HST_DATA1, SMB_HST_BLOCK_DATA, SMB_PEC_DATA, SMB_SLAVE_WDATA, SMB_SLAVE_CMD, SMB_SLAVE_CTS, SMB_SLV, SMB_NOTIFY_ADDR, SMB_NOTIFY_LOW_BYTE, SMB_NOTIFY_HIGH_BYTE, SMB_DATA_HLDTIME, SMB_TIMEOUT_CNT, SMB_TMEXT_CNT, I2CSMB_DATA_SETUP, CH0_CTRL, CH0_SRC_ADDR, CH0_DEST_ADDR, CH0_SRC_XCNT, CH0_SRC_XMDFY, CH0_DEST_XMDFY, CH0_SRC_REQ, DMA_CH_EN, CH1_CTRL, CH1_SRC_ADDR, CH1_DEST_ADDR, CH1_SRC_XCNT, CH1_SRC_XMDFY, CH1_DEST_XMDFY, CH1_SRC_REQ, INTR_EN0, RAW_STCR0",
    "Validation / Acceptance Criteria": "- The transfer-complete interrupt must occur and be cleared in both the I2C peripheral and system controller; otherwise the test fails.\n- After both DMA operations complete, the five destination SRAM words must exactly match the five source SRAM words; any mismatch fails the test.\n- The current byte count must reach zero before interrupt handling completes; otherwise report failure.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_i2c0_mst_i2c1_slv_dma_st_md",
    "Hidden_Test_Description": "The program configures a series of system addresses (0xA1700008..0xA1700054) to 0x1, enables GIC IRQs 80 and 81, and enables I2C0/I2C1 interrupts via MIZAR_LSS_SYSREG_INTR_EN0. I2C setup: write MIZAR_I2C0_FF and MIZAR_I2C1_FF = 0x3, clear MIZAR_I2C0_INTR_CLR and MIZAR_I2C1_INTR_CLR with 0xFFFFFFFF, set MIZAR_I2C0_DEV_CTRL and MIZAR_I2C1_DEV_CTRL = 0x382, set MIZAR_I2C1_SLV_ADDR = 0x37 and MIZAR_I2C0_TGT_SLV_ADDR = 0x6E, set MIZAR_I2C0_I2C_BYTE_CNT and MIZAR_I2C1_I2C_BYTE_CNT = 0x5. Configure standard mode timing: MIZAR_I2C0_SF_LCNT=0x12C and MIZAR_I2C0_SF_HCNT=0xC8. Mask I2C0 interrupts with MIZAR_I2C0_MASK_INTR = 0xFFFFFFEF. Set TX/RX FIFO thresholds for both I2C0/I2C1 to 0x5. Enable I2C0 DMA via MIZAR_I2C0_DMA_CTRL = 0x2. DMA CH0 is programmed: CH0_CTRL=0x8028028, CH0_SRC_ADDR=SRAM_ADDR_1 (0xA0243FC0), CH0_DEST_ADDR=MIZAR_I2C0_SMB_HST_BLOCK_DATA, CH0_SRC_XCNT=0x5, CH0_SRC_XMDFY=0x4, CH0_DEST_XMDFY=0x0, CH0_SRC_REQ=0x5, DMA_CH_EN=0x1, then dma_disable() busy-waits until DMA_CH_EN becomes 0. The transfer starts by writing MIZAR_I2C0_TSFR_CTRL=0x2. It polls MIZAR_I2C0_I2C_CURRENT_BYTECNT until zero with wait_on(5000) delays. Then int_pend is set to 1 and it waits while(int_pend) with wait_on(10) until IRQ occurs. In Default_IRQHandler: reads MIZAR_I2C0_INTR_STS; if 0x0010, clears MIZAR_I2C0_INTR_CLR bit 0x10, clears GIC IRQs 80/81, writes MIZAR_LSS_SYSREG_RAW_STCR0 with LSS_SYSREG_RAW_STCR0_I2C0_INTERRUPT, waits 100, re-reads MIZAR_I2C0_INTR_STS and LSS_SYSREG RAW_STCR0 & LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT must both be 0x00; otherwise increments test_err. If not 0x0010, increments test_err. After ISR: set MIZAR_I2C1_DMA_CTRL=0x1 and program DMA CH1 to move from MIZAR_I2C1_SMB_HST_BLOCK_DATA to SRAM_ADDR_2 (0xA0243FE0): CH1_CTRL=0x8024028, CH1_SRC_ADDR=MIZAR_I2C1_SMB_HST_BLOCK_DATA, CH1_DEST_ADDR=0xA0243FE0, CH1_SRC_XCNT=0x5, CH1_SRC_XMDFY=0x0, CH1_DEST_XMDFY=0x4, CH1_SRC_REQ=0x6, DMA_CH_EN=0x2; then dma_disable() waits until channel enable clears. Finally, it compares five words at SRAM_ADDR_1 and SRAM_ADDR_2; mismatches increment test_err; then waits 100 and finish(test_err).",
    "Hidden_Remarks": "Non-secure configuration of system NIC is required. The test assumes IRQ mapping for I2C0/1 and uses a specific transfer-complete status value (0x0010) for validation.",
    "Hidden_Test_Steps_Procedure": "Entry point: test_case()\n1. Security and interrupt setup:\n   1.1 Write 0x1 to system addresses 0xA1700008..0xA1700054 (specific listed offsets in code).\n   1.2 GIC_EnableIRQ(80); GIC_EnableIRQ(81).\n   1.3 write_reg(MIZAR_LSS_SYSREG_INTR_EN0, LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT | LSS_SYSREG_INTR_EN0_I2C1_INTERRUPT).\n2. I2C configuration for standard mode:\n   2.1 write_reg(MIZAR_I2C0_FF, 0x3); write_reg(MIZAR_I2C1_FF, 0x3).\n   2.2 write_reg(MIZAR_I2C0_INTR_CLR, 0xFFFFFFFF); write_reg(MIZAR_I2C1_INTR_CLR, 0xFFFFFFFF).\n   2.3 write_reg(MIZAR_I2C0_DEV_CTRL, 0x382); write_reg(MIZAR_I2C1_DEV_CTRL, 0x382).\n   2.4 write_reg(MIZAR_I2C1_SLV_ADDR, 0x37); write_reg(MIZAR_I2C0_TGT_SLV_ADDR, 0x6E).\n   2.5 write_reg(MIZAR_I2C0_I2C_BYTE_CNT, 0x5); write_reg(MIZAR_I2C1_I2C_BYTE_CNT, 0x5).\n   2.6 write_reg(MIZAR_I2C0_SF_LCNT, 0x12C); write_reg(MIZAR_I2C0_SF_HCNT, 0xC8).\n   2.7 write_reg(MIZAR_I2C0_MASK_INTR, 0xFFFFFFEF).\n   2.8 write_reg(MIZAR_I2C0_TX_FIFO_THLD, 0x5); write_reg(MIZAR_I2C1_TX_FIFO_THLD, 0x5).\n   2.9 write_reg(MIZAR_I2C0_RX_FIFO_THLD, 0x5); write_reg(MIZAR_I2C1_RX_FIFO_THLD, 0x5).\n   2.10 write_reg(MIZAR_I2C0_DMA_CTRL, 0x2).\n3. DMA CH0 setup and transmit preload:\n   3.1 src_addr = 0xA0243FC0; dest_addr = MIZAR_I2C0_SMB_HST_BLOCK_DATA.\n   3.2 For i=0..4: tx_data[i] = i*5; write_reg(0xA0243FC0 + i*4, tx_data[i]).\n   3.3 write_reg(MIZAR_DMA_CH0_CTRL, 0x8028028).\n   3.4 write_reg(MIZAR_DMA_CH0_SRC_ADDR, 0xA0243FC0).\n   3.5 write_reg(MIZAR_DMA_CH0_DEST_ADDR, MIZAR_I2C0_SMB_HST_BLOCK_DATA).\n   3.6 write_reg(MIZAR_DMA_CH0_SRC_XCNT, 0x5).\n   3.7 write_reg(MIZAR_DMA_CH0_SRC_XMDFY, 0x4).\n   3.8 write_reg(MIZAR_DMA_CH0_DEST_XMDFY, 0x0).\n   3.9 write_reg(MIZAR_DMA_CH0_SRC_REQ, 0x5).\n   3.10 write_reg(MIZAR_DMA_DMA_CH_EN, 0x1).\n   3.11 dma_disable(): while(read_reg(MIZAR_DMA_DMA_CH_EN)) { wait_on(100); }.\n4. Start transfer and poll:\n   4.1 write_reg(MIZAR_I2C0_TSFR_CTRL, 0x2).\n   4.2 Poll MIZAR_I2C0_I2C_CURRENT_BYTECNT until 0 with wait_on(5000) between reads.\n   4.3 int_pend = 1; while(int_pend) { print; wait_on(10); }.\n5. Interrupt service (Default_IRQHandler):\n   5.1 int_pend = 0; int_status = read_reg(MIZAR_I2C0_INTR_STS).\n       If (int_status == 0x0010):\n         a) write_reg(MIZAR_I2C0_INTR_CLR, 0x00000010).\n         b) GIC_ClearIRQ(80); GIC_ClearIRQ(81).\n         c) write_reg(MIZAR_LSS_SYSREG_RAW_STCR0, LSS_SYSREG_RAW_STCR0_I2C0_INTERRUPT).\n         d) wait_on(100).\n         e) Verify: read_reg(MIZAR_I2C0_INTR_STS) == 0x00 and (read_reg(MIZAR_LSS_SYSREG_RAW_STCR0) & LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT) == 0x00; else test_err++.\n       Else: test_err++.\n6. DMA CH1 setup (receive to SRAM):\n   6.1 write_reg(MIZAR_I2C1_DMA_CTRL, 0x1).\n   6.2 src_addr = MIZAR_I2C1_SMB_HST_BLOCK_DATA; dest_addr = 0xA0243FE0.\n   6.3 write_reg(MIZAR_DMA_CH1_CTRL, 0x8024028).\n   6.4 write_reg(MIZAR_DMA_CH1_SRC_ADDR, MIZAR_I2C1_SMB_HST_BLOCK_DATA).\n   6.5 write_reg(MIZAR_DMA_CH1_DEST_ADDR, 0xA0243FE0).\n   6.6 write_reg(MIZAR_DMA_CH1_SRC_XCNT, 0x5).\n   6.7 write_reg(MIZAR_DMA_CH1_SRC_XMDFY, 0x0).\n   6.8 write_reg(MIZAR_DMA_CH1_DEST_XMDFY, 0x4).\n   6.9 write_reg(MIZAR_DMA_CH1_SRC_REQ, 0x6).\n   6.10 write_reg(MIZAR_DMA_DMA_CH_EN, 0x2).\n   6.11 dma_disable(): while(read_reg(MIZAR_DMA_DMA_CH_EN)) { wait_on(100); }.\n7. Compare data:\n   7.1 For i=0..4: read source at 0xA0243FC0 + i*4 and destination at 0xA0243FE0 + i*4; if equal print success else increment test_err.\n8. Finalize: wait_on(100); finish(test_err).",
    "Hidden_Impacted_Registers": "MIZAR_I2C0_DEV_CTRL, MIZAR_I2C0_TSFR_CTRL, MIZAR_I2C0_SLV_ADDR, MIZAR_I2C0_TGT_SLV_ADDR, MIZAR_I2C0_I2C_MSTR_CODE, MIZAR_I2C0_I2C_BYTE_CNT, MIZAR_I2C0_SF_HCNT, MIZAR_I2C0_SF_LCNT, MIZAR_I2C0_RIS, MIZAR_I2C0_MASK_INTR, MIZAR_I2C0_INTR_STS, MIZAR_I2C0_INTR_CLR, MIZAR_I2C0_TAS, MIZAR_I2C0_TX_FIFO_THLD, MIZAR_I2C0_RX_FIFO_THLD, MIZAR_I2C0_DMA_CTRL, MIZAR_I2C0_FF, MIZAR_I2C0_TX_FIFO_LVL, MIZAR_I2C0_RX_FIFO_LVL, MIZAR_I2C0_I2C_MSTR_STS, MIZAR_I2C0_I2C_FLTR_SEL, MIZAR_I2C0_I2C_CURRENT_BYTECNT, MIZAR_I2C0_I2C_SMB_SFTRST, MIZAR_I2C0_SMB_HST_STS, MIZAR_I2C0_SMB_HST_CTRL, MIZAR_I2C0_SMB_HST_CMD, MIZAR_I2C0_SMB_HST_DATA0, MIZAR_I2C0_SMB_HST_DATA1, MIZAR_I2C0_SMB_HST_BLOCK_DATA, MIZAR_I2C0_SMB_PEC_DATA, MIZAR_I2C0_SMB_SLAVE_WDATA, MIZAR_I2C0_SMB_SLAVE_CMD, MIZAR_I2C0_SMB_SLAVE_CTS, MIZAR_I2C0_SMB_SLV, MIZAR_I2C0_SMB_NOTIFY_ADDR, MIZAR_I2C0_SMB_NOTIFY_LOW_BYTE, MIZAR_I2C0_SMB_NOTIFY_HIGH_BYTE, MIZAR_I2C0_SMB_DATA_HLDTIME, MIZAR_I2C0_SMB_TIMEOUT_CNT, MIZAR_I2C0_SMB_TMEXT_CNT, MIZAR_I2C0_I2CSMB_DATA_SETUP, MIZAR_I2C1_DEV_CTRL, MIZAR_I2C1_TSFR_CTRL, MIZAR_I2C1_SLV_ADDR, MIZAR_I2C1_TGT_SLV_ADDR, MIZAR_I2C1_I2C_MSTR_CODE, MIZAR_I2C1_I2C_BYTE_CNT, MIZAR_I2C1_SF_HCNT, MIZAR_I2C1_SF_LCNT, MIZAR_I2C1_RIS, MIZAR_I2C1_MASK_INTR, MIZAR_I2C1_INTR_STS, MIZAR_I2C1_INTR_CLR, MIZAR_I2C1_TAS, MIZAR_I2C1_TX_FIFO_THLD, MIZAR_I2C1_RX_FIFO_THLD, MIZAR_I2C1_DMA_CTRL, MIZAR_I2C1_FF, MIZAR_I2C1_TX_FIFO_LVL, MIZAR_I2C1_RX_FIFO_LVL, MIZAR_I2C1_I2C_MSTR_STS, MIZAR_I2C1_I2C_FLTR_SEL, MIZAR_I2C1_I2C_CURRENT_BYTECNT, MIZAR_I2C1_I2C_SMB_SFTRST, MIZAR_I2C1_SMB_HST_STS, MIZAR_I2C1_SMB_HST_CTRL, MIZAR_I2C1_SMB_HST_CMD, MIZAR_I2C1_SMB_HST_DATA0, MIZAR_I2C1_SMB_HST_DATA1, MIZAR_I2C1_SMB_HST_BLOCK_DATA, MIZAR_I2C1_SMB_PEC_DATA, MIZAR_I2C1_SMB_SLAVE_WDATA, MIZAR_I2C1_SMB_SLAVE_CMD, MIZAR_I2C1_SMB_SLAVE_CTS, MIZAR_I2C1_SMB_SLV, MIZAR_I2C1_SMB_NOTIFY_ADDR, MIZAR_I2C1_SMB_NOTIFY_LOW_BYTE, MIZAR_I2C1_SMB_NOTIFY_HIGH_BYTE, MIZAR_I2C1_SMB_DATA_HLDTIME, MIZAR_I2C1_SMB_TIMEOUT_CNT, MIZAR_I2C1_SMB_TMEXT_CNT, MIZAR_I2C1_I2CSMB_DATA_SETUP, MIZAR_DMA_CH0_CTRL, MIZAR_DMA_CH0_SRC_ADDR, MIZAR_DMA_CH0_DEST_ADDR, MIZAR_DMA_CH0_SRC_XCNT, MIZAR_DMA_CH0_SRC_XMDFY, MIZAR_DMA_CH0_DEST_XMDFY, MIZAR_DMA_CH0_SRC_REQ, MIZAR_DMA_DMA_CH_EN, MIZAR_DMA_CH1_CTRL, MIZAR_DMA_CH1_SRC_ADDR, MIZAR_DMA_CH1_DEST_ADDR, MIZAR_DMA_CH1_SRC_XCNT, MIZAR_DMA_CH1_SRC_XMDFY, MIZAR_DMA_CH1_DEST_XMDFY, MIZAR_DMA_CH1_SRC_REQ, MIZAR_LSS_SYSREG_INTR_EN0, MIZAR_LSS_SYSREG_RAW_STCR0, LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT, LSS_SYSREG_INTR_EN0_I2C1_INTERRUPT",
    "Hidden_Validation_Acceptance_Criteria": "1) Interrupt: MIZAR_I2C0_INTR_STS must show 0x0010; after clearing with MIZAR_I2C0_INTR_CLR and writing MIZAR_LSS_SYSREG_RAW_STCR0, both peripheral and system status must read 0x00. Failure increments test_err.\n2) Data integrity: Values at 0xA0243FE0..(n) must match values at 0xA0243FC0..(n) for five words; mismatch increments test_err.\n3) Progress: MIZAR_I2C0_I2C_CURRENT_BYTECNT must become 0 before ISR is processed."
  }
]

# Stage1 column definitions
META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

# Ensure JSON is valid and non-empty
if not isinstance(json_rows, list) or not json_rows:
    raise SystemExit("Invalid or empty JSON input")

# Normalize keys: ensure all rows have the same keys (union); preserve order as provided in first row extended by any missing keys
seen_keys = []
for k in json_rows[0].keys():
    seen_keys.append(k)
for row in json_rows[1:]:
    for k in row.keys():
        if k not in seen_keys:
            seen_keys.append(k)

for row in json_rows:
    for k in seen_keys:
        row.setdefault(k, "")

# Create workbook and the 'Data' sheet
wb = Workbook()
ws = wb.active
ws.title = 'Data'

# Write headers (union key order)
for c, key in enumerate(seen_keys, start=1):
    ws.cell(row=1, column=c, value=key)

# Write data rows exactly as values
for r, row in enumerate(json_rows, start=2):
    for c, key in enumerate(seen_keys, start=1):
        ws.cell(row=r, column=c, value=row.get(key, ""))

# Base formatting: bold header, freeze top row, basic width estimation
header_font = Font(bold=True)
for c in range(1, len(seen_keys) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = header_font

ws.freeze_panes = 'A2'

# Auto-fit column widths (approx)
for c, key in enumerate(seen_keys, start=1):
    max_len = len(str(key))
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=c).value
        if val is None:
            continue
        max_len = max(max_len, len(str(val)))
    ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(80, max(12, max_len * 0.9))

# Create Meta_data_sheet and copy META columns AS-IS (order preserved)
meta = wb.create_sheet('Meta_data_sheet')
for c, key in enumerate(META_COLS, start=1):
    meta.cell(row=1, column=c, value=key)

for r, row in enumerate(json_rows, start=2):
    for c, key in enumerate(META_COLS, start=1):
        meta.cell(row=r, column=c, value=row.get(key, ""))

# Hide meta sheet (very hidden)
meta.sheet_state = 'veryHidden'

# Transform Data sheet into TestPlan main sheet
# 1) Remove META columns and reorder to MAIN_ORDER on the same sheet
# Build a matrix for TestPlan
main_headers = MAIN_ORDER
main_rows = []
for row in json_rows:
    main_rows.append([row.get(h, "") for h in main_headers])

# Clear current Data sheet and write reordered content
ws.delete_rows(1, ws.max_row)
for c, key in enumerate(main_headers, start=1):
    ws.cell(row=1, column=c, value=key)
for r, vals in enumerate(main_rows, start=2):
    for c, v in enumerate(vals, start=1):
        ws.cell(row=r, column=c, value=v)

# Rename 'Data' to 'TestPlan'
ws.title = 'TestPlan'

# Safety check: ensure no sheet named 'Data' remains
for sh in list(wb.sheetnames):
    if sh == 'Data':
        # If a stray sheet named Data exists, remove it
        del wb['Data']

# Strict formatting for TestPlan
wrap_cols = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
left_top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
right_top = Alignment(horizontal='right', vertical='top', wrap_text=False)
center_top = Alignment(horizontal='center', vertical='top', wrap_text=False)

# Header styling
for c in range(1, len(main_headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True)
    cell.alignment = center
    cell.fill = blue_fill

# Apply wrapping and alignments to data rows
header_index = {h: i + 1 for i, h in enumerate(main_headers)}
for r in range(2, ws.max_row + 1):
    for h, c in header_index.items():
        cell = ws.cell(row=r, column=c)
        if h in wrap_cols:
            cell.alignment = left_top_wrap
        elif h == 'Index':
            cell.alignment = center_top
        else:
            # default left/top for text
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

# Numbering inside specific columns (Test Steps / Procedure and Validation / Acceptance Criteria)
for col_name in ["Test Steps / Procedure", "Validation / Acceptance Criteria"]:
    c = header_index[col_name]
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=c).value
        if val is None:
            continue
        # Split on lines and normalize numbering as 1., 2., ...
        lines = [ln.strip() for ln in str(val).splitlines() if ln.strip()]
        if lines:
            numbered = []
            for i, ln in enumerate(lines, start=1):
                # Remove any leading bullets or numbers
                lns = ln
                # common bullet patterns
                for pat in ["- ", "• ", "* ", "-", "•", "*", "\u2022 "]:
                    if lns.startswith(pat):
                        lns = lns[len(pat):].strip()
                # patterns like 1) or 1. or a) etc
                if len(lns) > 1 and (lns[0].isdigit() and (lns[1:2] in [')', '.'])):
                    lns = lns[2:].strip()
                numbered.append(f"{i}. {lns}")
            ws.cell(row=r, column=c, value="\n".join(numbered))
            ws.cell(row=r, column=c).alignment = left_top_wrap

# Thin borders for all populated cells
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).border = border

# Autofilter on header row
ws.auto_filter.ref = ws.dimensions

# Recompute column widths after wrapping (approx)
for c, key in enumerate(main_headers, start=1):
    max_len = len(str(key))
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=c).value
        if val is None:
            continue
        for line in str(val).splitlines():
            max_len = max(max_len, len(line))
    ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(100, max(14, max_len * 0.9))

# Approximate row heights based on wrapped content lines
for r in range(2, ws.max_row + 1):
    lines = 1
    for col in wrap_cols:
        c = header_index[col]
        val = ws.cell(row=r, column=c).value
        if val:
            lines = max(lines, len(str(val).splitlines()))
    ws.row_dimensions[r].height = min(200, 14 * lines)

# Data validation ONLY on 'Code Generation (Required / Not)'
code_col = header_index["Code Generation (Required / Not)"]
dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
dv.error = "Select one of: Required, Blank, Not Required"
dv.errorTitle = "Invalid selection"
ws.add_data_validation(dv)
dv.add(f"{ws.cell(row=2, column=code_col).coordinate}:{ws.cell(row=ws.max_row, column=code_col).coordinate}")

# Compute IST timestamp for filename
ist = timezone(timedelta(hours=5, minutes=30))
now_ist = datetime.now(ist)
filename = f"I2C_TestPlan_{now_ist.strftime('%Y%m%d')}_{now_ist.strftime('%H%M%S')}.xlsx"
output_dir = Path("Test_Output/I2C/TestPlan")
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / filename

# Save workbook
wb.save(output_path)

# Validate as a true OOXML ZIP-based XLSX
required_entries = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
with zipfile.ZipFile(output_path, 'r') as zf:
    names = set(zf.namelist())
    if not required_entries.issubset(names):
        raise SystemExit("XLSX validation failed: missing OOXML core entries")

# Persist generated path for workflow consumption
with open('.generated_excel_path.txt', 'w') as f:
    f.write(str(output_path))

print(f"Generated Excel: {output_path}")
