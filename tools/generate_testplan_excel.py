#!/usr/bin/env python3
import json
import os
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Inputs (constants per task)
OWNER = "titusbspgit"
REPO = "PSVValidation"
BRANCH = "main"
OUTPUT_DIRECTORY = os.path.join("Test_Output", "GPIO", "TestPlan")
IP_NAME = "GPIO"
COMMIT_CHANGES = True

# Exact final aggregated JSON from Agent 6 (unchanged)
FINAL_JSON_STR = r'''[
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "NA",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Verifies the default reset values and masked write/read behavior of gp0_gpio_8 through gp0_gpio_30. For each register, the test first checks the default value (when readable). Then, for a set of predefined data patterns, it writes masked values to writable registers, reads back using the read mask, and confirms that writable fields reflect the programmed value while non-writable fields retain their defaults. The test reports PASS only if all checks across all targeted registers succeed.",
    "Meta Test Description": "The test executes in two phases: (1) default/reset value checks and (2) masked write-read validation. Phase 1 (chk_rst_val): For i in [0..CNT-1], take addr = addr_array[i]. If skip_rst_array[i] == 1, continue. If read_mask_array[i] == 0x00000000, continue. Read data_rd = read_reg(addr); compute data = (data_rd & 0xfffffffe) to mask off the LSB; compare data with default_value_array[i]. On mismatch, increment def_fail_cnt and print failure log. Phase 2 (chk_rd_wr): Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}. For each pattern data_wr in chk_val: (a) Write pass: For each i, addr = addr_array[i]; if skip_array[i] == 1, continue; if write_mask_array[i] == 0x00000000, continue; else write_reg(addr, (data_wr & write_mask_array[i])). (b) Read/verify pass: For each i, addr = addr_array[i]; if skip_array[i] == 1, continue; if write_mask_array[i] == 0x00000000, continue; if read_mask_array[i] == 0x00000000, continue; else read masked value data_rd = (read_reg(addr) & read_mask_array[i]). Compute wr_n = (write_mask_array[i] ^ 0xffffffff). Compute expected value exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])). If data_rd == exp_val, PASS log; else increment wr_fail_cnt and print failure log. At the end of test_case(), if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) else finish(0). A soft reset check helper (soft_reset_chk) is present but disabled with #ifdef 0. The test relies on arrays of addresses, default values, and read/write masks in test_define.c and on read_reg/write_reg/finish/wait_on from the test framework.",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Some registers are intentionally skipped for reset and/or write-read checks per skip_rst_array and skip_array. The test depends on read_reg, write_reg, wait_on, and finish implementations from the test framework. A known behavior noted in comments: reading default values can force an input level high unless driven low, which may affect comparisons. The optional soft reset path is disabled.",
    "Test Steps / Procedure": "1. Read each target register (gp0_gpio_8 through gp0_gpio_30) that is marked readable and not flagged to be skipped for reset checks; verify the read value (with the documented LSB mask applied) matches its default value.\n2. For each of the predefined data patterns (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555, 0xF5F5F5F5, 0xA5A5A5A5, 0xFFFF0000), program all writable and non-skipped registers with the pattern masked by their write masks.\n3. Read back each targeted register using its read mask and verify that writable fields reflect the programmed pattern and non-writable fields retain their default values.\n4. Record any mismatches found during default checks or masked write-read checks.\n5. Declare the test PASS only if no mismatches are detected across all registers and all data patterns; otherwise declare FAIL.",
    "Meta Test Steps / Procedure": "Initialization: def_fail_cnt = 0; wr_fail_cnt = 0. Call chk_rst_val(): for (i = 0; i < CNT; i++): addr = addr_array[i]; if (skip_rst_array[i] == 1) continue; if (read_mask_array[i] == 0x00000000) continue; data_rd = read_reg(addr); data = (data_rd & 0xfffffffe); if (data == default_value_array[i]) PASS log; else def_fail_cnt++ and print failure log with addr, expected default, data, and data_rd. Call chk_rd_wr(): define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}; For each j in [0..5]: data_wr = chk_val[j]; Write loop: for (i = 0; i < CNT; i++): addr = addr_array[i]; if (skip_array[i] == 1) continue; if (write_mask_array[i] == 0x00000000) continue; else write_reg(addr, (data_wr & write_mask_array[i])). Read/verify loop: for (i = 0; i < CNT; i++): addr = addr_array[i]; if (skip_array[i] == 1) continue; if (write_mask_array[i] == 0x00000000) continue; if (read_mask_array[i] == 0x00000000) continue; else data_rd = (read_reg(addr) & read_mask_array[i]); wr_n = (write_mask_array[i] ^ 0xffffffff); exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if (data_rd == exp_val) PASS log; else wr_fail_cnt++ and print failure log with addr, exp_val, data_rd. Test completion: if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) else finish(0). soft_reset_chk() contains disabled steps to write/read a reset register with waits (#ifdef 0).",
    "Impacted Registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10; gp0_gpio_11; gp0_gpio_12; gp0_gpio_13; gp0_gpio_14; gp0_gpio_15; gp0_gpio_16; gp0_gpio_17; gp0_gpio_18; gp0_gpio_19; gp0_gpio_20; gp0_gpio_21; gp0_gpio_22; gp0_gpio_23; gp0_gpio_24; gp0_gpio_25; gp0_gpio_26; gp0_gpio_27; gp0_gpio_28; gp0_gpio_29; gp0_gpio_30",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27; MIZAR_GPIO_GP0_GPIO_28; MIZAR_GPIO_GP0_GPIO_29; MIZAR_GPIO_GP0_GPIO_30",
    "Validation / Acceptance Criteria": "PASS if: (1) all readable, non-skipped registers match their documented default values; and (2) for each data pattern, every writable and readable register reads back the expected masked value (writable bits reflect the pattern; non-writable bits retain defaults). FAIL if any mismatch is detected in either phase.",
    "Meta Validation / Acceptance Criteria": "Default check: data = (read_reg(addr_array[i]) & 0xfffffffe) must equal default_value_array[i] for all i where skip_rst_array[i] == 0 and read_mask_array[i] != 0. Write-read check for each data_wr in {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}: for all i where skip_array[i] == 0, write_mask_array[i] != 0, and read_mask_array[i] != 0, compute wr_n = (write_mask_array[i] ^ 0xffffffff) and exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])). The readback data_rd = (read_reg(addr_array[i]) & read_mask_array[i]) must equal exp_val. def_fail_cnt and wr_fail_cnt accumulate failures; final result: finish(0) if both counters are zero, else finish(1).",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>; #include <stdlib.h>; #include \"test_common.h\"; #include \"test_define.c\"; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>",
    "Meta Macros": "#define SOFT_RST_REG_ADDRESS\t0x00000000; #define SOFT_RST_REG_DATA\t0x00000000; #define CNT 49",
    "Meta Arrays": "const unsigned long int addr_array[49]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,MIZAR_GPIO_GP0_GPIO_10,MIZAR_GPIO_GP0_GPIO_11,MIZAR_GPIO_GP0_GPIO_12,MIZAR_GPIO_GP0_GPIO_13,MIZAR_GPIO_GP0_GPIO_14,MIZAR_GPIO_GP0_GPIO_15,MIZAR_GPIO_GP0_GPIO_16,MIZAR_GPIO_GP0_GPIO_17,MIZAR_GPIO_GP0_GPIO_18,MIZAR_GPIO_GP0_GPIO_19,MIZAR_GPIO_GP0_GPIO_20,MIZAR_GPIO_GP0_GPIO_21,MIZAR_GPIO_GP0_GPIO_22,MIZAR_GPIO_GP0_GPIO_23,MIZAR_GPIO_GP0_GPIO_24,MIZAR_GPIO_GP0_GPIO_25,MIZAR_GPIO_GP0_GPIO_26,MIZAR_GPIO_GP0_GPIO_27,MIZAR_GPIO_GP0_GPIO_28,MIZAR_GPIO_GP0_GPIO_29,MIZAR_GPIO_GP0_GPIO_30,};\n\nconst unsigned int default_value_array[49]={GPIO_GP0_GPIO_8_DEFAULT_VAL,GPIO_GP0_GPIO_9_DEFAULT_VAL,GPIO_GP0_GPIO_10_DEFAULT_VAL,GPIO_GP0_GPIO_11_DEFAULT_VAL,GPIO_GP0_GPIO_12_DEFAULT_VAL,GPIO_GP0_GPIO_13_DEFAULT_VAL,GPIO_GP0_GPIO_14_DEFAULT_VAL,GPIO_GP0_GPIO_15_DEFAULT_VAL,GPIO_GP0_GPIO_16_DEFAULT_VAL,GPIO_GP0_GPIO_17_DEFAULT_VAL,GPIO_GP0_GPIO_18_DEFAULT_VAL,GPIO_GP0_GPIO_19_DEFAULT_VAL,GPIO_GP0_GPIO_20_DEFAULT_VAL,GPIO_GP0_GPIO_21_DEFAULT_VAL,GPIO_GP0_GPIO_22_DEFAULT_VAL,GPIO_GP0_GPIO_23_DEFAULT_VAL,GPIO_GP0_GPIO_24_DEFAULT_VAL,GPIO_GP0_GPIO_25_DEFAULT_VAL,GPIO_GP0_GPIO_26_DEFAULT_VAL,GPIO_GP0_GPIO_27_DEFAULT_VAL,GPIO_GP0_GPIO_28_DEFAULT_VAL,GPIO_GP0_GPIO_29_DEFAULT_VAL,GPIO_GP0_GPIO_30_DEFAULT_VAL,};\n\nconst unsigned int read_mask_array[49]={GPIO_GP0_GPIO_8_READ_MASK,GPIO_GP0_GPIO_9_READ_MASK,GPIO_GP0_GPIO_10_READ_MASK,GPIO_GP0_GPIO_11_READ_MASK,GPIO_GP0_GPIO_12_READ_MASK,GPIO_GP0_GPIO_13_READ_MASK,GPIO_GP0_GPIO_14_READ_MASK,GPIO_GP0_GPIO_15_READ_MASK,GPIO_GP0_GPIO_16_READ_MASK,GPIO_GP0_GPIO_17_READ_MASK,GPIO_GP0_GPIO_18_READ_MASK,GPIO_GP0_GPIO_19_READ_MASK,GPIO_GP0_GPIO_20_READ_MASK,GPIO_GP0_GPIO_21_READ_MASK,GPIO_GP0_GPIO_22_READ_MASK,GPIO_GP0_GPIO_23_READ_MASK,GPIO_GP0_GPIO_24_READ_MASK,GPIO_GP0_GPIO_25_READ_MASK,GPIO_GP0_GPIO_26_READ_MASK,GPIO_GP0_GPIO_27_READ_MASK,GPIO_GP0_GPIO_28_READ_MASK,GPIO_GP0_GPIO_29_READ_MASK,GPIO_GP0_GPIO_30_READ_MASK,};\n\nconst unsigned int write_mask_array[49]={GPIO_GP0_GPIO_8_WRITE_MASK,GPIO_GP0_GPIO_9_WRITE_MASK,GPIO_GP0_GPIO_10_WRITE_MASK,GPIO_GP0_GPIO_11_WRITE_MASK,GPIO_GP0_GPIO_12_WRITE_MASK,GPIO_GP0_GPIO_13_WRITE_MASK,GPIO_GP0_GPIO_14_WRITE_MASK,GPIO_GP0_GPIO_15_WRITE_MASK,GPIO_GP0_GPIO_16_WRITE_MASK,GPIO_GP0_GPIO_17_WRITE_MASK,GPIO_GP0_GPIO_18_WRITE_MASK,GPIO_GP0_GPIO_19_WRITE_MASK,GPIO_GP0_GPIO_20_WRITE_MASK,GPIO_GP0_GPIO_21_WRITE_MASK,GPIO_GP0_GPIO_22_WRITE_MASK,GPIO_GP0_GPIO_23_WRITE_MASK,GPIO_GP0_GPIO_24_WRITE_MASK,GPIO_GP0_GPIO_25_WRITE_MASK,GPIO_GP0_GPIO_26_WRITE_MASK,GPIO_GP0_GPIO_27_WRITE_MASK,GPIO_GP0_GPIO_28_WRITE_MASK,GPIO_GP0_GPIO_29_WRITE_MASK,GPIO_GP0_GPIO_30_WRITE_MASK,};\n\nconst unsigned int skip_array[49]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,1,1,1,1,1,1,1,1,0,0,0,0,};\nconst unsigned int skip_rst_array[49]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,1,1,1,1,1,1,1,1,1,1,1,};"
  }
]'''


def main():
    # Parse JSON exactly as given
    data = json.loads(FINAL_JSON_STR)
    if not isinstance(data, list) or not data:
        raise SystemExit("final_json is empty or invalid")

    # Compute IST timestamp
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    date_part = now_ist.strftime('%Y%m%d')
    time_part = now_ist.strftime('%H%M%S')
    filename = f"{IP_NAME}_TestPlan_{date_part}_{time_part}.xlsx"

    # Prepare workbook with two sheets
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"
    meta = wb.create_sheet("MetaData")
    meta.sheet_state = 'veryHidden'

    # Header order from the first item
    first_obj = data[0]
    headers = list(first_obj.keys())

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical='top')

    # Write headers
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    # Write data rows preserving order
    for obj in data:
        row = [obj.get(h, "") for h in headers]
        ws.append(row)

    # Apply wrap to all cells and set reasonable column widths
    col_widths = [len(str(h)) for h in headers]
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c_idx, cell in enumerate(r, start=1):
            cell.alignment = wrap
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > col_widths[c_idx-1]:
                col_widths[c_idx-1] = val_len
    for i, w in enumerate(col_widths, start=1):
        adj = min(max(w + 2, 12), 60)
        ws.column_dimensions[get_column_letter(i)].width = adj

    # Freeze first row
    ws.freeze_panes = "A2"

    # MetaData sheet: two columns Key, Value
    meta_headers = ["Key", "Value"]
    meta.append(meta_headers)
    for col_idx in range(1, 3):
        cell = meta.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    meta_rows = [
        ("owner", OWNER),
        ("repo", REPO),
        ("branch", BRANCH),
        ("output_directory", OUTPUT_DIRECTORY),
        ("IP_NAME", IP_NAME),
        ("commit_changes", str(COMMIT_CHANGES).lower()),
        ("timestamp_IST", now_ist.strftime('%Y-%m-%d %H:%M:%S IST')),
        ("intended_filename", filename),
        ("intended_commit_message", f"Add {IP_NAME} TestPlan (IST {now_ist.strftime('%Y-%m-%d %H:%M:%S IST')})"),
        ("final_json", FINAL_JSON_STR),
    ]
    for k, v in meta_rows:
        meta.append([k, v])
    # Style meta sheet and widths
    meta_cols = [0, 0]
    for r in meta.iter_rows(min_row=1, max_row=meta.max_row, min_col=1, max_col=2):
        for c_idx, cell in enumerate(r, start=1):
            cell.alignment = wrap
            l = len(str(cell.value)) if cell.value is not None else 0
            if l > meta_cols[c_idx-1]:
                meta_cols[c_idx-1] = l
    for i, w in enumerate(meta_cols, start=1):
        meta.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 16), 100)

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIRECTORY, filename)

    # Save workbook
    wb.save(out_path)

    print(out_path)

if __name__ == "__main__":
    main()
