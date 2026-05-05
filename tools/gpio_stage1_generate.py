#!/usr/bin/env python3
import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import zipfile

# Input JSON (embedded deterministically)
JSON_TEXT = r'''{
  "metadata": {
    "ip_name": "GPIO",
    "repo": "titusbspgit/PSVValidation",
    "branch": "main",
    "subdir": "TestRepo/gpio",
    "generated_ts_ist": "AUTO-GENERATE-IST",
    "source_items": [
      "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/gpio_reg_wr_rd_test/main.c",
      "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/test_gpio_negedge_intr_en/main.c",
      "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/test_gpio_pedge_all_pads_en/main.c"
    ]
  },
  "tests": [
    {
      "Index": "1",
      "SS / Module": "GPIO",
      "Feature": "AHB 32-bit register interface",
      "Test Case Name": "gpio_reg_wr_rd_test",
      "Test Description": "Verifies GPIO register default values and masked read/write behavior across individual and group registers.",
      "Speed": "NA",
      "Mode": "NA",
      "Memory Start Offset": "NA",
      "Memory End Offset": "NA",
      "Remarks": "When reading default values, DIN can become 1 if not forced; forcing zero to DIN may drive level select high, causing a mismatch with expected values.",
      "Test Steps / Procedure": "1) For each register (GPIO_8 to GPIO_39, GPIO_INTR_RAW_STCLR1, INTR1_INTR_EN1, INTR1_INTR_STS1, INTR2_INTR_EN1, INTR2_INTR_STS1, GPIO_IO_CTRL_GROUP1 to GPIO_IO_CTRL_GROUP4, GPIO_DOUT_GROUP1 to GPIO_DOUT_GROUP4, GPIO_DIN_GROUP1 to GPIO_DIN_GROUP4), read the value and compare its masked value against the documented default.\n2) For each data pattern, write the masked value to every writable register and read back with the read mask.\n3) For each readback, compute the expected value by combining written bits with preserved default bits per mask and compare.\n4) Report any mismatch and declare failure if any default or write/read check fails.",
      "Impacted Registers": ["GPIO_8","GPIO_9","GPIO_10","GPIO_11","GPIO_12","GPIO_13","GPIO_14","GPIO_15","GPIO_16","GPIO_17","GPIO_18","GPIO_19","GPIO_20","GPIO_21","GPIO_22","GPIO_23","GPIO_24","GPIO_25","GPIO_26","GPIO_27","GPIO_28","GPIO_29","GPIO_30","GPIO_31","GPIO_32","GPIO_33","GPIO_34","GPIO_35","GPIO_36","GPIO_37","GPIO_38","GPIO_39","GPIO_INTR_RAW_STCLR1","INTR1_INTR_EN1","INTR1_INTR_STS1","INTR2_INTR_EN1","INTR2_INTR_STS1","GPIO_IO_CTRL_GROUP1","GPIO_IO_CTRL_GROUP2","GPIO_IO_CTRL_GROUP3","GPIO_IO_CTRL_GROUP4","GPIO_DOUT_GROUP1","GPIO_DOUT_GROUP2","GPIO_DOUT_GROUP3","GPIO_DOUT_GROUP4","GPIO_DIN_GROUP1","GPIO_DIN_GROUP2","GPIO_DIN_GROUP3","GPIO_DIN_GROUP4"],
      "Validation / Acceptance Criteria": "1) Default value check passes when each readable, non-skipped register's masked read equals its documented default value. Any mismatch is a failure.\n2) Write/read check passes when each readable and writable register's masked read equals the expected value derived from written data and masks; any mismatch is a failure.\n3) The overall test passes only if no default or write/read mismatches are reported.",
      "Source Files": [
        {"role": "main_test", "url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/gpio_reg_wr_rd_test/main.c"}
      ],
      "Macro Aliases": [
        "MIZAR_GPIO_GP0_GPIO_8","MIZAR_GPIO_GP0_GPIO_9","MIZAR_GPIO_GP0_GPIO_10","MIZAR_GPIO_GP0_GPIO_11","MIZAR_GPIO_GP0_GPIO_12","MIZAR_GPIO_GP0_GPIO_13","MIZAR_GPIO_GP0_GPIO_14","MIZAR_GPIO_GP0_GPIO_15","MIZAR_GPIO_GP0_GPIO_16","MIZAR_GPIO_GP0_GPIO_17","MIZAR_GPIO_GP0_GPIO_18","MIZAR_GPIO_GP0_GPIO_19","MIZAR_GPIO_GP0_GPIO_20","MIZAR_GPIO_GP0_GPIO_21","MIZAR_GPIO_GP0_GPIO_22","MIZAR_GPIO_GP0_GPIO_23","MIZAR_GPIO_GP0_GPIO_24","MIZAR_GPIO_GP0_GPIO_25","MIZAR_GPIO_GP0_GPIO_26","MIZAR_GPIO_GP0_GPIO_27","MIZAR_GPIO_GP0_GPIO_28","MIZAR_GPIO_GP0_GPIO_29","MIZAR_GPIO_GP0_GPIO_30","MIZAR_GPIO_GP0_GPIO_31","MIZAR_GPIO_GP0_GPIO_32","MIZAR_GPIO_GP0_GPIO_33","MIZAR_GPIO_GP0_GPIO_34","MIZAR_GPIO_GP0_GPIO_35","MIZAR_GPIO_GP0_GPIO_36","MIZAR_GPIO_GP0_GPIO_37","MIZAR_GPIO_GP0_GPIO_38","MIZAR_GPIO_GP0_GPIO_39","MIZAR_GPIO_GPIO_INTR_RAW_STCLR1","MIZAR_GPIO_GP0_INTR1_INTR_EN1","MIZAR_GPIO_GP0_INTR1_INTR_STS1","MIZAR_GPIO_GP0_INTR2_INTR_EN1","MIZAR_GPIO_GP0_INTR2_INTR_STS1","MIZAR_GPIO_GPIO_IO_CTRL_GROUP1","MIZAR_GPIO_GPIO_IO_CTRL_GROUP2","MIZAR_GPIO_GPIO_IO_CTRL_GROUP3","MIZAR_GPIO_GPIO_IO_CTRL_GROUP4","MIZAR_GPIO_GPIO_DOUT_GROUP1","MIZAR_GPIO_GPIO_DOUT_GROUP2","MIZAR_GPIO_GPIO_DOUT_GROUP3","MIZAR_GPIO_GPIO_DOUT_GROUP4","MIZAR_GPIO_GPIO_DIN_GROUP1","MIZAR_GPIO_GPIO_DIN_GROUP2","MIZAR_GPIO_GPIO_DIN_GROUP3","MIZAR_GPIO_GPIO_DIN_GROUP4"
      ]
    },
    {
      "Index": "2",
      "SS / Module": "GPIO",
      "Feature": "Negative edge interrupt enable",
      "Test Case Name": "test_gpio_negedge_intr_en",
      "Test Description": "Validates falling-edge interrupt generation and clearing for GPIO pins 8 through 39.",
      "Speed": "NA",
      "Mode": "Interrupt",
      "Memory Start Offset": "0xA0243ffc",
      "Memory End Offset": "0xA0243ffc",
      "Remarks": "Interrupts are routed through system registers and GIC. A bounded wait with timeout is used to avoid hangs. Pad levels are driven by a memory-mapped pad control at 0xA0243ffc.",
      "Test Steps / Procedure": "1) Enable interrupt in INTR_EN1 for the selected GPIO instance.\n2) Drive the external pad output port high to set a known level.\n3) For each pin (8 to 39), configure the corresponding GPIO_x register to enable negative-edge detection and clear raw status.\n4) For each pin, clear the group raw status in GPIO_INTR_RAW_STCLR1 and enable the pin in INTR1_INTR_EN1.\n5) Arm the wait, then create a falling edge by toggling the pad output from high to low for the active bit.\n6) Wait until the interrupt is observed or until timeout expires.\n7) In the interrupt handler, verify the pin input is low, confirm the group status in INTR1_INTR_STS1 for the active bit, clear the per‑pin raw (via GPIO_x) and the group raw (GPIO_INTR_RAW_STCLR1), verify the group status is cleared, clear RAW_STCR1, and clear the GIC interrupt.",
      "Impacted Registers": ["GPIO_8","GPIO_9","GPIO_10","GPIO_11","GPIO_12","GPIO_13","GPIO_14","GPIO_15","GPIO_16","GPIO_17","GPIO_18","GPIO_19","GPIO_20","GPIO_21","GPIO_22","GPIO_23","GPIO_24","GPIO_25","GPIO_26","GPIO_27","GPIO_28","GPIO_29","GPIO_30","GPIO_31","GPIO_32","GPIO_33","GPIO_34","GPIO_35","GPIO_36","GPIO_37","GPIO_38","GPIO_39","GPIO_INTR_RAW_STCLR1","INTR1_INTR_EN1","INTR1_INTR_STS1","INTR_EN1","RAW_STCR1"],
      "Validation / Acceptance Criteria": "1) Each falling edge results in an interrupt before the timeout expires; otherwise it is a failure.\n2) In the handler, the pin input reads low and the group status shows the active bit; both must be true to pass for that pin.\n3) After clearing per‑pin and group raw status, the group status reads zero; otherwise it is a failure.\n4) System interrupt status is cleared in RAW_STCR1 and the interrupt is cleared at the GIC.",
      "Source Files": [
        {"role": "main_test", "url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/test_gpio_negedge_intr_en/main.c"}
      ],
      "Macro Aliases": [
        "MIZAR_LSS_SYSREG_INTR_EN1","LSS_SYSREG_INTR_EN1_GPIO0_INTR","LSS_SYSREG_INTR_EN1_GPIO1_INTR","MIZAR_GPIO_GP0_GPIO_8","MIZAR_GPIO_GPIO_INTR_RAW_STCLR1","MIZAR_GPIO_GP0_INTR1_INTR_EN1","MIZAR_GPIO_GP0_INTR1_INTR_STS1","MIZAR_LSS_SYSREG_RAW_STCR1","LSS_SYSREG_RAW_STCR1_GPIO0_INTR","LSS_SYSREG_RAW_STCR1_GPIO1_INTR"
      ]
    },
    {
      "Index": "3",
      "SS / Module": "GPIO",
      "Feature": "Positive edge interrupt enable",
      "Test Case Name": "test_gpio_pedge_all_pads_en",
      "Test Description": "Validates rising-edge interrupt generation for GPIO pins 8 through 39 with per-pin raw clear and group status handling.",
      "Speed": "NA",
      "Mode": "Interrupt",
      "Memory Start Offset": "0xA0243ffc",
      "Memory End Offset": "0xA0243ffc",
      "Remarks": "Interrupts are routed through system registers and GIC. A timeout prevents infinite waits. Group interrupt is masked during handler service and re-enabled afterward.",
      "Test Steps / Procedure": "1) Enable interrupt in INTR_EN1 for the selected GPIO instance.\n2) For each pin (8 to 39), set GPIO_x to enable rising-edge detection.\n3) Configure input mode using GPIO_IO_CTRL_GROUP1 to GPIO_IO_CTRL_GROUP4.\n4) Enable all pins in INTR1_INTR_EN1.\n5) For each pin, drive the pad low, arm the wait, then drive high to create a rising edge.\n6) Wait for the interrupt or timeout.\n7) In the interrupt handler, check group status in INTR1_INTR_STS1, mask INTR1_INTR_EN1, clear per‑pin raw via GPIO_x, verify group status clears to zero, clear RAW_STCR1, re‑enable INTR1_INTR_EN1, and clear the GIC interrupt.",
      "Impacted Registers": ["GPIO_8","GPIO_9","GPIO_10","GPIO_11","GPIO_12","GPIO_13","GPIO_14","GPIO_15","GPIO_16","GPIO_17","GPIO_18","GPIO_19","GPIO_20","GPIO_21","GPIO_22","GPIO_23","GPIO_24","GPIO_25","GPIO_26","GPIO_27","GPIO_28","GPIO_29","GPIO_30","GPIO_31","GPIO_32","GPIO_33","GPIO_34","GPIO_35","GPIO_36","GPIO_37","GPIO_38","GPIO_39","GPIO_IO_CTRL_GROUP1","GPIO_IO_CTRL_GROUP2","GPIO_IO_CTRL_GROUP3","GPIO_IO_CTRL_GROUP4","INTR1_INTR_EN1","INTR1_INTR_STS1","INTR_EN1","RAW_STCR1"],
      "Validation / Acceptance Criteria": "1) Each rising edge triggers an interrupt before the timeout; otherwise it is a failure.\n2) Group status shows a pending interrupt upon entry and reads zero after per‑pin and group raw clears; otherwise it is a failure.\n3) System interrupt status in RAW_STCR1 is cleared and the GIC interrupt is acknowledged; failures increment the error count.",
      "Source Files": [
        {"role": "main_test", "url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/test_gpio_pedge_all_pads_en/main.c"}
      ],
      "Macro Aliases": [
        "MIZAR_GPIO_GP0_GPIO_8","MIZAR_GPIO_GPIO_IO_CTRL_GROUP1","MIZAR_GPIO_GPIO_IO_CTRL_GROUP2","MIZAR_GPIO_GPIO_IO_CTRL_GROUP3","MIZAR_GPIO_GPIO_IO_CTRL_GROUP4","MIZAR_GPIO_GP0_INTR1_INTR_EN1","MIZAR_GPIO_GP0_INTR1_INTR_STS1","MIZAR_LSS_SYSREG_INTR_EN1","MIZAR_LSS_SYSREG_RAW_STCR1","LSS_SYSREG_INTR_EN1_GPIO0_INTR","LSS_SYSREG_INTR_EN1_GPIO1_INTR","LSS_SYSREG_RAW_STCR1_GPIO0_INTR","LSS_SYSREG_RAW_STCR1_GPIO1_INTR"
      ]
    }
  ],
  "parsed_files": [
    {"url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/gpio_reg_wr_rd_test/main.c", "role": "main_test"},
    {"url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/test_gpio_negedge_intr_en/main.c", "role": "main_test"},
    {"url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/test_gpio_pedge_all_pads_en/main.c", "role": "main_test"}
  ]
}'''


def ensure_dir(p):
    os.makedirs(p, exist_ok=True)


def collect_union_keys(rows):
    seen = []
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen.append(k)
    return seen


def json_to_string(v):
    if isinstance(v, (list, dict)):
        return json.dumps(v, separators=(",", ":"))
    return v


def renumber_multiline(text):
    if not isinstance(text, str):
        return text
    lines = [ln for ln in text.splitlines() if ln.strip() != ""]
    out = []
    for i, ln in enumerate(lines, 1):
        # strip existing leading numbering/bullets
        s = ln.strip()
        # generic trim of leading markers (digits + ) . - : or bullet markers)
        while True:
            if len(s) > 1 and s[0].isdigit():
                j = 1
                while j < len(s) and s[j].isdigit():
                    j += 1
                if j < len(s) and s[j] in ") .-:":
                    j += 1
                    while j < len(s) and s[j] == ' ':
                        j += 1
                    s = s[j:]
                    break
            if s.startswith(('- ', '* ')):
                s = s[2:]
                break
            break
        out.append(f"{i}. {s}")
    return "\n".join(out) if out else text


def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = cell.value
                l = len(str(v)) if v is not None else 0
                max_len = max(max_len, l)
            except Exception:
                pass
        width = min(max(10, int(max_len * 1.2) + 2), 60)
        ws.column_dimensions[col_letter].width = width


def apply_borders(ws):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def set_row_heights(ws, wrap_cols_idx):
    default = 15
    for r in range(2, ws.max_row + 1):
        lines = 1
        for c in wrap_cols_idx:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                lines = max(lines, v.count("\n") + 1)
        ws.row_dimensions[r].height = min(300, default * lines)


def main():
    data = json.loads(JSON_TEXT)
    tests = data.get("tests", [])
    if not isinstance(tests, list) or not tests:
        raise SystemExit("ERROR: tests array missing or empty")

    # Phase 1 — JSON to Excel (staging on 'Data')
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    union_keys = collect_union_keys(tests)
    # Write header
    for ci, key in enumerate(union_keys, 1):
        ws.cell(row=1, column=ci, value=key)
    # Write rows
    for ri, row in enumerate(tests, 2):
        for ci, key in enumerate(union_keys, 1):
            val = row.get(key, "")
            ws.cell(row=ri, column=ci, value=json_to_string(val))

    # Base formatting on 'Data'
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = "A2"
    autofit_columns(ws)

    # Phase 2 — Meta_data_sheet (Very Hidden)
    meta_cols = [
        "Hidden_Test_Case_Name",
        "Hidden_Test_Description",
        "Hidden_Remarks",
        "Hidden_Test_Steps_Procedure",
        "Hidden_Impacted_Registers",
        "Hidden_Validation_Acceptance_Criteria",
    ]
    ws_meta = wb.create_sheet(title="Meta_data_sheet")
    for ci, key in enumerate(meta_cols, 1):
        ws_meta.cell(row=1, column=ci, value=key)
    for ri, row in enumerate(tests, 2):
        # Preserve hidden values exactly; also store Macro Aliases as Hidden_Impacted_Registers if explicit hidden value not provided
        hidden_vals = {k: row.get(k, "") for k in meta_cols}
        if not hidden_vals.get("Hidden_Impacted_Registers") and row.get("Macro Aliases"):
            mac = row.get("Macro Aliases")
            if isinstance(mac, list):
                hidden_vals["Hidden_Impacted_Registers"] = ", ".join(str(x) for x in mac)
            else:
                hidden_vals["Hidden_Impacted_Registers"] = str(mac)
        for ci, key in enumerate(meta_cols, 1):
            ws_meta.cell(row=ri, column=ci, value=hidden_vals.get(key, ""))
    ws_meta.sheet_state = 'veryHidden'

    # Rename 'Data' -> 'TestPlan' and reorder/normalize columns on same sheet
    ws.title = "TestPlan"
    main_order = [
        "Index",
        "SS / Module",
        "Feature",
        "Test Case Name",
        "Test Description",
        "Speed",
        "Mode",
        "Memory Start Offset",
        "Memory End Offset",
        "Remarks",
        "Test Steps / Procedure",
        "Impacted Registers",
        "Validation / Acceptance Criteria",
        "Code Generation (Required / Not)",
    ]

    # Snapshot current data
    headers = [c.value for c in ws[1]]
    records = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        for c in range(1, ws.max_column + 1):
            rec[headers[c - 1]] = ws.cell(row=r, column=c).value
        records.append(rec)

    # Clear sheet and write main columns
    ws.delete_rows(1, ws.max_row)
    for ci, key in enumerate(main_order, 1):
        ws.cell(row=1, column=ci, value=key)
    for ri, rec in enumerate(records, 2):
        for ci, key in enumerate(main_order, 1):
            v = rec.get(key, "")
            if key in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                v = renumber_multiline(v)
            if key == "Impacted Registers":
                try:
                    arr = json.loads(v) if isinstance(v, str) else v
                    if isinstance(arr, list):
                        v = ", ".join(str(x) for x in arr)
                except Exception:
                    pass
            ws.cell(row=ri, column=ci, value=v)

    # Strict formatting on 'TestPlan'
    blue_fill = PatternFill("solid", fgColor="FFB7DEE8")
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.fill = blue_fill

    wrap_cols = ["Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"]
    wrap_cols_idx = []
    for ci, key in enumerate(main_order, 1):
        if key in wrap_cols:
            wrap_cols_idx.append(ci)
    for r in range(2, ws.max_row + 1):
        for ci, key in enumerate(main_order, 1):
            cell = ws.cell(row=r, column=ci)
            if key in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            elif key == "Index":
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")

    autofit_columns(ws)
    set_row_heights(ws, wrap_cols_idx)
    ws.auto_filter.ref = ws.dimensions
    apply_borders(ws)
    ws.freeze_panes = "A2"

    # Data Validation for Code Generation (Required / Not)
    if ws.max_row >= 2:
        try:
            codegen_col = main_order.index("Code Generation (Required / Not)") + 1
            dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
            dv.error = "Select a value from the list"
            dv.errorTitle = "Invalid Input"
            ws.add_data_validation(dv)
            dv.ranges.add(f"{ws.cell(row=2, column=codegen_col).coordinate}:{ws.cell(row=ws.max_row, column=codegen_col).coordinate}")
        except ValueError:
            pass

    # Mandatory safety: ensure no sheet named 'Data'
    if any(sh.title == 'Data' for sh in wb.worksheets):
        for sh in list(wb.worksheets):
            if sh.title == 'Data':
                wb.remove(sh)
        if any(sh.title == 'Data' for sh in wb.worksheets):
            raise SystemExit("ERROR: lingering 'Data' sheet could not be removed")

    # Phase 3 — Save & Validate
    ist = datetime.now(ZoneInfo('Asia/Kolkata'))
    datestr = ist.strftime('%Y%m%d_%H%M%S')
    out_dir = os.path.join('Test_Output', 'GPIO', 'TestPlan')
    ensure_dir(out_dir)
    out_name = f"GPIO_TestPlan_{datestr}.xlsx"
    out_path = os.path.join(out_dir, out_name)

    wb.save(out_path)

    with zipfile.ZipFile(out_path, 'r') as zf:
        names = set(zf.namelist())
        required = {'[Content_Types].xml', 'xl/workbook.xml'}
        if not required.issubset(names):
            raise SystemExit("ERROR: XLSX validation failed — missing core members")

    with open('.stage1_generated_path.txt', 'w', encoding='utf-8') as f:
        f.write(out_path)
    print(f"Generated Excel at {out_path}")


if __name__ == '__main__':
    main()
