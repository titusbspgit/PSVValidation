import json, os, zipfile, re
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Meta and final column definitions
META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

FINAL_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

HEADER_FILL = PatternFill('solid', fgColor='FF4472C4')  # solid blue
HEADER_FONT = Font(bold=True, color='FFFFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_ALIGN_TEXT = Alignment(horizontal='left', vertical='top', wrap_text=False)
DATA_ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
DATA_ALIGN_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=False)
THIN = Side(style='thin', color='FF000000')
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def normalize_and_number_block(text: str) -> str:
    """Split multiline text into logical items, strip any existing numbering like '1) ', '1. ', '- ', then re-number as 'n. ' within one wrapped cell."""
    if text is None:
        return ''
    items = []
    for raw in str(text).splitlines():
        s = raw.strip()
        if not s:
            continue
        # Remove leading numbering patterns: '1) ', '1. ', '1 - ', '- ', '* ', etc.
        s = re.sub(r'^(\d+)[\.)\-: ]+\s*', '', s)
        s = re.sub(r'^[\-*•]+\s*', '', s)
        items.append(s)
    if not items:
        return ''
    return "\n".join(f"{i}. {itm}" for i, itm in enumerate(items, start=1))


def auto_width(ws):
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        max_len = len(str(header)) if header is not None else 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            s = '' if v is None else str(v)
            if s:
                for line in s.splitlines():
                    max_len = max(max_len, len(line))
        width = max(12, min(max_len + 2, 100))
        ws.column_dimensions[get_column_letter(c)].width = width


def apply_borders(ws):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = BORDER_THIN


def set_row_heights(ws):
    base = 15
    for r in range(1, ws.max_row + 1):
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            s = '' if v is None else str(v)
            lines = s.count('\n') + 1 if s else 1
            max_lines = max(max_lines, lines)
        ws.row_dimensions[r].height = min(409, base * max_lines)


def validate_xlsx(path: str):
    if not zipfile.is_zipfile(path):
        raise SystemExit('Saved file is not a valid XLSX ZIP')
    with zipfile.ZipFile(path, 'r') as z:
        names = set(z.namelist())
        req = {'[Content_Types].xml', 'xl/workbook.xml'}
        if not req.issubset(names):
            raise SystemExit('XLSX structure invalid')


def main():
    # STEP 1: Validate JSON Input
    with open('tools/inline_json_input.json', 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    if not isinstance(json_data, list) or not json_data:
        raise SystemExit('Invalid JSON: not a non-empty array')

    # STEP 2: Normalize Tabular Schema (preserve first-seen key order across all rows)
    key_order = []
    seen = set()
    for row in json_data:
        if not isinstance(row, dict):
            raise SystemExit('Invalid row: must be object')
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                key_order.append(k)

    # PHASE 1 — Generate Base Workbook with single sheet 'Data'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Header row
    for c, k in enumerate(key_order, start=1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.fill = HEADER_FILL
    ws.freeze_panes = 'A2'

    # Data rows
    for r, row in enumerate(json_data, start=2):
        for c, k in enumerate(key_order, start=1):
            ws.cell(row=r, column=c, value=row.get(k, ''))

    auto_width(ws)
    apply_borders(ws)

    # PHASE 2 — META sheet creation
    meta_ws = wb.create_sheet('Meta_data_sheet')
    for c, k in enumerate(META_COLS, start=1):
        meta_ws.cell(row=1, column=c, value=k).font = Font(bold=True)
    for r, row in enumerate(json_data, start=2):
        for c, k in enumerate(META_COLS, start=1):
            meta_ws.cell(row=r, column=c, value=row.get(k, ''))
    meta_ws.sheet_state = 'veryHidden'

    # STEP 7 — Normalize MAIN sheet in-place: rename Data -> TestPlan and reorder columns
    ws.title = 'TestPlan'

    # Capture existing data from Data(TestPlan now)
    current_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows_data = []
    for r in range(2, ws.max_row + 1):
        rows_data.append({current_headers[c - 1]: ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)})

    # Clear and rebuild with FINAL_ORDER (excluding META columns by construction)
    ws.delete_rows(1, ws.max_row)

    for c, h in enumerate(FINAL_ORDER, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.fill = HEADER_FILL

    for r, rd in enumerate(rows_data, start=2):
        for c, h in enumerate(FINAL_ORDER, start=1):
            ws.cell(row=r, column=c, value=rd.get(h, ''))

    # Strict formatting rules
    wrap_cols = {'Test Description', 'Remarks', 'Test Steps / Procedure', 'Validation / Acceptance Criteria'}
    for r in range(2, ws.max_row + 1):
        for c, h in enumerate(FINAL_ORDER, start=1):
            if h in wrap_cols:
                ws.cell(row=r, column=c).alignment = DATA_ALIGN_WRAP
            elif h == 'Index':
                ws.cell(row=r, column=c).alignment = DATA_ALIGN_CENTER
            else:
                ws.cell(row=r, column=c).alignment = DATA_ALIGN_TEXT

    auto_width(ws)
    set_row_heights(ws)
    apply_borders(ws)

    # Mandatory numbering inside cells for two columns (normalize any existing numbering)
    for r in range(2, ws.max_row + 1):
        for col_name in ['Test Steps / Procedure', 'Validation / Acceptance Criteria']:
            idx = FINAL_ORDER.index(col_name) + 1
            old = ws.cell(row=r, column=idx).value
            ws.cell(row=r, column=idx, value=normalize_and_number_block(old))

    # Data validation only for Code Generation (Required / Not)
    if 'Code Generation (Required / Not)' in FINAL_ORDER:
        cg_col = FINAL_ORDER.index('Code Generation (Required / Not)') + 1
        dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(cg_col)}2:{get_column_letter(cg_col)}{ws.max_row}")

    # STEP 7B — Enforce final sheet visibility
    names = [s.title for s in wb.worksheets]
    if 'Data' in names:
        for s in wb.worksheets:
            if s.title == 'Data':
                wb.remove(s)
    names2 = [s.title for s in wb.worksheets]
    if set(names2) != {'TestPlan', 'Meta_data_sheet'}:
        raise SystemExit(f'Unexpected worksheets present: {names2}')

    # PHASE 3 — SAVE & VALIDATE
    out_dir = os.path.join('Test_Output', 'GPIO', 'TestPlan')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'newfile.xlsx')
    wb.save(out_path)

    validate_xlsx(out_path)

    # Print for logs
    print(f'Generated: {out_path}')


if __name__ == '__main__':
    main()
