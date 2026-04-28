#!/usr/bin/env python3
import json, os, sys
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INPUT_JSON_PATH = os.path.join('TestPlan_Input', 'I2C_TestPlan.json')
OUTPUT_DIR = os.path.join('Test_Output', 'I2C', 'TestPlan')
IP_NAME = 'I2C'

MAIN_COLS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]

def read_json(path):
    if not os.path.isfile(path):
        print(f"ERROR: JSON file not found at {path}", file=sys.stderr)
        sys.exit(1)
    with open(path, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
        except Exception as e:
            print(f"ERROR: Failed to parse JSON: {e}", file=sys.stderr)
            sys.exit(1)
    if not isinstance(data, dict) or 'Test_Cases' not in data or not isinstance(data['Test_Cases'], list) or len(data['Test_Cases']) == 0:
        print("ERROR: JSON must be an object with non-empty 'Test_Cases' array", file=sys.stderr)
        sys.exit(1)
    return data


def ist_now_str():
    if ZoneInfo is None:
        # Fallback to fixed offset if zoneinfo is not available
        from datetime import timezone, timedelta
        ist = timezone(timedelta(hours=5, minutes=30))
        now = datetime.now(ist)
    else:
        now = datetime.now(ZoneInfo('Asia/Kolkata'))
    return now.strftime('%Y-%m-%d %H:%M:%S IST'), now.strftime('%Y%m%d'), now.strftime('%H%M%S')


def auto_width(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            val = '' if v is None else str(v)
            widths[i] = max(widths.get(i, 0), len(val))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 12), 80)


def build_all_keys(records):
    seen = []
    seen_set = set()
    for rec in records:
        if isinstance(rec, dict):
            for k in rec.keys():
                if k not in seen_set:
                    seen.append(k)
                    seen_set.add(k)
    return seen


def normalize_value(v):
    if isinstance(v, (dict, list)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return v


def write_sheet(ws, headers, rows):
    # Header
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    # Rows
    for r_idx, rec in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=normalize_value(rec.get(h, '')))
    # Basic formatting
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align
    ws.freeze_panes = 'A2'
    auto_width(ws)


def apply_testplan_format(ws):
    # Header style
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    max_col = ws.max_column
    max_row = ws.max_row

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Wrap text for specific columns
    wrap_cols = {
        'Test Description',
        'Remarks',
        'Test Steps / Procedure',
        'Validation / Acceptance Criteria'
    }
    # Alignment
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            hdr = ws.cell(row=1, column=c).value
            # Borders
            cell.border = border
            # Vertical alignment
            if hdr == 'Index':
                cell.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
            else:
                if hdr in wrap_cols:
                    cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
                else:
                    # Default text columns left
                    cell.alignment = Alignment(vertical='top', horizontal='left')

    auto_width(ws)
    # Attempt to auto-fit row heights after wrapping
    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = None

    # Data validation for Code Generation (Required / Not)
    if 'Code Generation (Required / Not)' in header_map:
        cg_col = header_map['Code Generation (Required / Not)']
        col_letter = get_column_letter(cg_col)
        dv = DataValidation(type='list', formula1='"Required,Not Required"', allow_blank=True, showErrorMessage=True)
        dv.error = 'Select only: Required or Not Required (or leave blank)'
        dv.errorTitle = 'Invalid Selection'
        ws.add_data_validation(dv)
        dv.ranges.append(f"{col_letter}2:{col_letter}{max_row}")


def main():
    data = read_json(INPUT_JSON_PATH)
    records = data['Test_Cases']

    # Phase 1: Build base workbook and Data sheet
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = 'Data'

    all_keys = build_all_keys(records)
    write_sheet(ws_data, all_keys, records)

    # Phase 2: Meta_data_sheet (Very Hidden)
    ws_meta = wb.create_sheet('Meta_data_sheet')
    meta_rows = []
    for rec in records:
        meta_rows.append({k: rec.get(k, '') for k in META_COLS})
    write_sheet(ws_meta, META_COLS, meta_rows)
    ws_meta.sheet_state = 'veryHidden'

    # Phase 2: TestPlan sheet (only MAIN_COLS, exact order)
    ws_data.title = 'TestPlan'  # rename Data to TestPlan
    # Create a clean TestPlan with only required columns in order
    ws_tp = wb.create_sheet('TestPlan_tmp')
    main_rows = []
    for rec in records:
        main_rows.append({k: rec.get(k, '') for k in MAIN_COLS})
    write_sheet(ws_tp, MAIN_COLS, main_rows)
    apply_testplan_format(ws_tp)
    # Remove old TestPlan and rename tmp
    del wb['TestPlan']
    ws_tp.title = 'TestPlan'

    # Phase 2: Auxiliary sheets
    ws_over = wb.create_sheet('Overview')
    ist_human, ymd, hms = ist_now_str()
    overview = [
        ('IP_NAME', IP_NAME),
        ('repo_url', 'https://github.com/titusbspgit/PSVValidation'),
        ('branch', 'main'),
        ('IST_timestamp', ist_human),
    ]
    for r, (k, v) in enumerate(overview, start=1):
        ws_over.cell(row=r, column=1, value=k)
        ws_over.cell(row=r, column=2, value=v)
    auto_width(ws_over)

    ws_cat = wb.create_sheet('Test Catalog')
    cat_cols = ['Index', 'Test Case Name', 'Feature', 'Speed', 'Mode']
    write_sheet(ws_cat, cat_cols, main_rows)

    ws_dep = wb.create_sheet('Dependencies')
    dep_cols = ['Index', 'Test Case Name', 'Impacted Registers']
    write_sheet(ws_dep, dep_cols, main_rows)

    ws_exec = wb.create_sheet('Execution')
    exec_cols = ['Index', 'Test Case Name', 'Test Steps / Procedure']
    write_sheet(ws_exec, exec_cols, main_rows)

    ws_tags = wb.create_sheet('Tags/Traceability')
    tags_cols = ['Index', 'Test Case Name', 'Tags']
    write_sheet(ws_tags, tags_cols, [{'Index': r.get('Index',''), 'Test Case Name': r.get('Test Case Name',''), 'Tags': ''} for r in main_rows])

    # Phase 3: Save Final Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ist_human, ymd, hms = ist_now_str()
    filename = f"{IP_NAME}_TestPlan_{ymd}_{hms}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(out_path)

    # Persist metadata for workflow commit message
    with open(os.path.join(OUTPUT_DIR, '.latest'), 'w', encoding='utf-8') as f:
        f.write(f"filename={filename}\n")
        f.write(f"ist={ist_human}\n")

    print(f"Generated: {out_path}")

if __name__ == '__main__':
    main()
