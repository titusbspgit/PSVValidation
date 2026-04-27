from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
IP_NAME = "SPI"
OUTPUT_DIR = os.path.join("Test_Output", "SPI", "TestPlan")

# Discovered items from repo (top-level under TestRepo/spi)
TEST_ITEMS = [
    {
        "id": "SPI_TC_001",
        "name": "spi_pio_full_duplex",
        "objective": "Validate SPI full-duplex data transfer using PIO for both TX and RX.",
        "expected": "Data integrity verified for simultaneous TX/RX across configured frames. No framing or overrun errors."
    },
    {
        "id": "SPI_TC_002",
        "name": "spi_pio_rx_dma_tx",
        "objective": "Validate RX via PIO while TX path uses DMA for burst transfers.",
        "expected": "DMA-driven TX completes without underruns; RX PIO captures all frames with correct ordering and content."
    },
    {
        "id": "SPI_TC_003",
        "name": "spi_reg_wr_rd_test",
        "objective": "Basic register access sanity: write/read back key SPI control/status registers.",
        "expected": "All writable registers reflect programmed values; status bits toggle as expected."
    }
]

REQUIREMENTS = [
    ("SPI_REQ_001", "Full-duplex operation"),
    ("SPI_REQ_002", "DMA support for TX path"),
    ("SPI_REQ_003", "Programmable registers accessible and sticky as defined"),
    ("SPI_REQ_004", "PIO operation for TX/RX"),
    ("SPI_REQ_005", "Error-free frame transfer and status reporting"),
]

TRACEABILITY = [
    ("SPI_REQ_001", "SPI_TC_001"),
    ("SPI_REQ_004", "SPI_TC_001"),
    ("SPI_REQ_002", "SPI_TC_002"),
    ("SPI_REQ_001", "SPI_TC_002"),
    ("SPI_REQ_005", "SPI_TC_002"),
    ("SPI_REQ_003", "SPI_TC_003"),
]

RISKS = [
    ("R-001", "DMA/PIO concurrency issues", "Medium", "Staggered starts; add stress tests; monitor FIFO levels"),
    ("R-002", "Clock polarity/phase mismatch (CPOL/CPHA)", "Low", "Enumerate modes in separate tests; add assertions"),
    ("R-003", "Data width/config mismatch", "Low", "Parameterize tests; add configuration validation"),
]

# Helpers

def ist_now():
    return datetime.now(ZoneInfo("Asia/Kolkata"))


def ts_for_filename(dt):
    return dt.strftime("%Y%m%d_%H%M%S")


def apply_header_style(ws, header_row=1):
    header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
    bold_font = Font(bold=True)
    center = Alignment(vertical="center")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center
    thin = Side(style='thin', color='FF999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = border


def autosize(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(val))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max(12, max_length + 2), 60)

# Build workbook

def build_workbook(now_ist):
    wb = Workbook()

    # Overview (reuse default sheet)
    ws_over = wb.active
    ws_over.title = "Overview"
    ws_over.append(["Field", "Value"])
    ws_over.append(["Title", f"{IP_NAME} Test Plan"])
    ws_over.append(["IP", IP_NAME])
    ws_over.append(["Generated (IST)", now_ist.strftime("%Y-%m-%d %H:%M:%S")])
    ws_over.append(["Scope", "Validation of SPI core focusing on PIO full-duplex, DMA-assisted TX, and register access sanity."])
    ws_over.append(["Out of Scope", "Electrical/analog characteristics; external device compliance beyond protocol behavior."])
    apply_header_style(ws_over)
    autosize(ws_over)

    # Features
    ws_feat = wb.create_sheet("Features")
    ws_feat.append(["Feature/Requirement", "Description"])
    feats = [
        ("Full-duplex PIO", "Simultaneous TX/RX via PIO"),
        ("DMA TX path", "Offload TX to DMA while RX via PIO"),
        ("Register Access", "RW/RO registers per spec; reset values"),
        ("Frame Formats", "Configurable frame size, CPOL/CPHA"),
        ("Error Handling", "Overrun/underrun flags; status bits"),
    ]
    for f in feats:
        ws_feat.append(list(f))
    apply_header_style(ws_feat)
    autosize(ws_feat)

    # Test Matrix
    ws_mat = wb.create_sheet("Test Matrix")
    ws_mat.append(["Test ID", "Test Name", "Objective", "Pre-reqs/Config", "Steps (summary)", "Expected Result", "Priority", "Type", "Owner", "Status"]) 
    for t in TEST_ITEMS:
        ws_mat.append([
            t["id"],
            t["name"],
            t["objective"],
            "Basic SPI init; loopback or slave device available",
            "Program mode; run transfer; capture/compare",
            t["expected"],
            "P1",
            "Directed",
            "TBD",
            "TBD",
        ])
    apply_header_style(ws_mat)
    autosize(ws_mat)

    # Testcases (details)
    ws_tc = wb.create_sheet("Testcases")
    ws_tc.append(["Test ID", "Name", "Pre-conditions", "Procedure", "Checks", "Logs/Artifacts"]) 
    details = {
        "SPI_TC_001": (
            "Loopback or paired device; set CPOL/CPHA; frame size 8/16",
            "Send N frames while reading concurrently (PIO/PIO)",
            "RX matches TX; no overruns; status flags correct",
        ),
        "SPI_TC_002": (
            "DMA configured for TX; RX via PIO; burst length >= FIFO depth",
            "Start DMA TX; poll/ISR for completion; read RX PIO",
            "No TX underrun; RX sequence correct; completion interrupt raised",
        ),
        "SPI_TC_003": (
            "SPI accessible; reset applied",
            "Write key control regs; read-back; toggle enables; read status",
            "RW regs echo values; RO unaffected; reset defaults match spec",
        ),
    }
    for t in TEST_ITEMS:
        pre, proc, chk = details[t["id"]]
        ws_tc.append([t["id"], t["name"], pre, proc, chk, "paths to logs TBD"])
    apply_header_style(ws_tc)
    autosize(ws_tc)

    # Traceability
    ws_tr = wb.create_sheet("Traceability")
    ws_tr.append(["Requirement ID", "Requirement", "Covered By (Test ID)"])
    # Build a map for requirement description
    req_desc = {rid: desc for rid, desc in REQUIREMENTS}
    for rid, tcid in TRACEABILITY:
        ws_tr.append([rid, req_desc.get(rid, ""), tcid])
    apply_header_style(ws_tr)
    autosize(ws_tr)

    # Execution
    ws_exe = wb.create_sheet("Execution")
    ws_exe.append(["Run Date (IST)", "Test ID", "Status", "Notes", "Log Link"])
    apply_header_style(ws_exe)
    autosize(ws_exe)

    # Risks
    ws_risk = wb.create_sheet("Risks")
    ws_risk.append(["Risk ID", "Risk", "Severity", "Mitigation"])
    for r in RISKS:
        ws_risk.append(list(r))
    apply_header_style(ws_risk)
    autosize(ws_risk)

    # Changelog
    ws_chg = wb.create_sheet("Changelog")
    ws_chg.append(["Date (IST)", "Author", "Change"])
    ws_chg.append([now_ist.strftime("%Y-%m-%d %H:%M:%S"), "auto", "Initial template with Overview, Features, Test Matrix, Testcases, Traceability, Execution, Risks, Changelog."])
    apply_header_style(ws_chg)
    autosize(ws_chg)

    return wb


def main():
    now = ist_now()
    ts = ts_for_filename(now)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = build_workbook(now)
    # Timestamped output
    fname_ts = f"{IP_NAME}_TestPlan_{ts}.xlsx"
    out_ts = os.path.join(OUTPUT_DIR, fname_ts)
    wb.save(out_ts)
    # Stable latest link
    out_latest = os.path.join(OUTPUT_DIR, f"{IP_NAME}_TestPlan_latest.xlsx")
    wb.save(out_latest)
    print(f"Generated: {out_ts}")
    print(f"Updated: {out_latest}")

if __name__ == "__main__":
    main()
