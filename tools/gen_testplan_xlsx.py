#!/usr/bin/env python3
import argparse, json, os, sys, re, zipfile
from datetime import datetime, timedelta, timezone
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_COLUMNS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLUMNS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
    'Hidden_Header_Includes',
    'Hidden_Macro_Define',  # Note: if not present in JSON, remains blank by design
    'Hidden_Skip_Array_Definition'
]

ALLOWED_VALIDATION_VALUES = 'Required,Blank,Not Required'

BLUE_FILL = PatternFill(fill_type='solid', fgColor='FF4472C4')
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)
RIGHT_TOP = Alignment(horizontal='right', vertical='top', wrap_text=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def parse_args():
    ap = argparse.ArgumentParser(description='Generate formatted Excel TestPlan from JSON input')
    ap.add_argument('--json-data', help='JSON array payload as string', default='')
    ap.add_argument('--json-file', help='Path to JSON file (optional)', default='')
    ap.add_argument('--output-dir', required=True, help='Output directory inside repository')
    ap.add_argument('--output-filename', default='', help='Output filename; if blank, auto-named by IST time')
    ap.add_argument('--ip-name', default='PCIE', help='IP name for filename prefix')
    return ap.parse_args()


def ist_now_str():
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    return now_ist.strftime('%Y%m%d'), now_ist.strftime('%H%M%S')


def load_json(json_data_str, json_file):
    payload = None
    if json_data_str:
        try:
            payload = json.loads(json_data_str)
        except Exception as e:
            print(f'ERROR: Invalid JSON in --json-data: {e}', file=sys.stderr)
            sys.exit(2)
    elif json_file:
        with open(json_file, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    else:
        print('ERROR: Provide --json-data or --json-file', file=sys.stderr)
        sys.exit(2)

    if not isinstance(payload, list) or len(payload) == 0:
        print('ERROR: JSON must be a non-empty array of objects', file=sys.stderr)
        sys.exit(2)
    for i, rec in enumerate(payload):
        if not isinstance(rec, dict):
            print(f'ERROR: JSON element at index {i} is not an object', file=sys.stderr)
            sys.exit(2)
    return payload


def union_keys(records):
    seen = []
    seen_set = set()
    for rec in records:
        for k in rec.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def auto_widths(ws):
    max_len = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            text = '' if val is None else str(val)
            l = len(text)
            if l > max_len.get(idx, 0):
                max_len[idx] = l
    for idx, l in max_len.items():
        # heuristic: character count + padding
        ws.column_dimensions[chr(64+idx) if idx <= 26 else _col_letter(idx)].width = min(120, max(10, l + 2))


def _col_letter(n):
    s = ''
    while n:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s


def renumber_block(text):
    if text is None:
        return ''
    s = str(text).replace('\r', '')
    lines = [ln for ln in s.split('\n')]
    out = []
    k = 1
    for ln in lines:
        t = ln.strip()
        if not t:
            continue
        # remove existing numeric/bullet prefixes like '1)', '1.', '(1)'
        t = re.sub(r'^\s*(\(?\d+\)?[\.|\)]\s*)', '', t)
        out.append(f"{k}. {t}")
        k += 1
    return '\n'.join(out) if out else s


def build_workbook(records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Phase 1: write Data sheet
    headers = union_keys(records)
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL
    ws.freeze_panes = 'A2'

    for rec in records:
        row = [rec.get(h, '') for h in headers]
        ws.append(row)

    auto_widths(ws)

    # Phase 2: Create META sheet
    meta = wb.create_sheet('Meta_data_sheet')
    for j, h in enumerate(META_COLUMNS, start=1):
        cell = meta.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL
    for i, rec in enumerate(records, start=2):
        for j, h in enumerate(META_COLUMNS, start=1):
            meta.cell(row=i, column=j, value=rec.get(h, ''))
    meta.sheet_state = 'veryHidden'
    auto_widths(meta)

    # Phase 2: Normalize MAIN sheet on the same original sheet
    ws.title = 'TestPlan'

    # Build mapping from header to column index
    header_to_idx = {h: i+1 for i, h in enumerate(headers)}

    # Determine columns to keep and order
    keep_headers = MAIN_COLUMNS

    # Create a new header row in place (row 1)
    # First, clear row 1
    for c in range(1, ws.max_column+1):
        ws.cell(row=1, column=c, value=None)
    for col_idx, h in enumerate(keep_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=1, column=col_idx).font = HEADER_FONT
        ws.cell(row=1, column=col_idx).alignment = CENTER
        ws.cell(row=1, column=col_idx).fill = BLUE_FILL

    # Rebuild data rows according to keep_headers
    data_rows = []
    for r in range(2, ws.max_row+1):
        rec_map = {}
        for h, idx in header_to_idx.items():
            rec_map[h] = ws.cell(row=r, column=idx).value
        data_rows.append([rec_map.get(h, '') for h in keep_headers])

    # Clear existing data (rows 2..end)
    ws.delete_rows(2, ws.max_row)

    # Write normalized rows
    for row in data_rows:
        ws.append(row)

    # Enable wrap text for specified columns
    wrap_cols = set([
        'Test Description', 'Remarks', 'Test Steps / Procedure', 'Validation / Acceptance Criteria'
    ])
    header_index = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}

    for r in range(2, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            val = ws.cell(row=1, column=c).value
            cell = ws.cell(row=r, column=c)
            if val in wrap_cols:
                cell.alignment = LEFT_TOP
            else:
                # Index and numeric-like: center/right; everything else left
                if val == 'Index' or isinstance(cell.value, (int, float)):
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT_TOP

    # Mandatory numbering inside specific cells (TestPlan only)
    for col_name in ['Test Steps / Procedure', 'Validation / Acceptance Criteria']:
        if col_name in header_index:
            cidx = header_index[col_name]
            for r in range(2, ws.max_row+1):
                cell = ws.cell(row=r, column=cidx)
                cell.value = renumber_block(cell.value)
                cell.alignment = LEFT_TOP

    # Apply thin borders to all populated cells
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    auto_widths(ws)

    # Data validation ONLY on Code Generation (Required / Not)
    if 'Code Generation (Required / Not)' in header_index and ws.max_row >= 2:
        col = header_index['Code Generation (Required / Not)']
        col_letter = _col_letter(col)
        dv = DataValidation(type='list', formula1=f'"{ALLOWED_VALIDATION_VALUES}"', allow_blank=True, showErrorMessage=True)
        dv.error = 'Select a value from the list'
        dv.errorTitle = 'Invalid Entry'
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{ws.max_row}')

    # Final visibility enforcement: only TestPlan (visible) and Meta_data_sheet (Very Hidden)
    # Ensure no sheet named 'Data' exists
    if 'Data' in wb.sheetnames:
        # Delete if still exists for any reason
        ws_data = wb['Data']
        wb.remove(ws_data)

    if set(wb.sheetnames) != {'TestPlan', 'Meta_data_sheet'}:
        # Reorder to exactly two sheets if extras somehow appeared
        for name in list(wb.sheetnames):
            if name not in {'TestPlan', 'Meta_data_sheet'}:
                wb.remove(wb[name])

    # Save workbook
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    # Validate as true XLSX
    validate_xlsx(output_path)


def validate_xlsx(path):
    if not os.path.isfile(path):
        print(f'ERROR: Output file not found: {path}', file=sys.stderr)
        sys.exit(3)
    if not zipfile.is_zipfile(path):
        print('ERROR: File is not a valid OOXML ZIP', file=sys.stderr)
        sys.exit(3)
    with zipfile.ZipFile(path, 'r') as zf:
        required = {'[Content_Types].xml', 'xl/workbook.xml'}
        names = set(zf.namelist())
        missing = required - names
        if missing:
            print(f'ERROR: Missing OOXML parts: {missing}', file=sys.stderr)
            sys.exit(3)
    wb = load_workbook(path, data_only=True)
    if 'TestPlan' not in wb.sheetnames or 'Meta_data_sheet' not in wb.sheetnames:
        print('ERROR: Required sheets missing after save', file=sys.stderr)
        sys.exit(3)
    if 'Data' in wb.sheetnames:
        print('ERROR: Sheet named Data still exists', file=sys.stderr)
        sys.exit(3)
    meta = wb['Meta_data_sheet']
    if getattr(meta, 'sheet_state', '') != 'veryHidden':
        print('ERROR: Meta_data_sheet is not Very Hidden', file=sys.stderr)
        sys.exit(3)
    ws = wb['TestPlan']
    # Validate single data validation on the specific column only
    dvs = getattr(ws, 'data_validations', None)
    if dvs is None or len(dvs.dataValidation) != 1:
        print('ERROR: Data validation not applied exactly once', file=sys.stderr)
        sys.exit(3)
    dv = list(dvs.dataValidation)[0]
    if dv.type != 'list' or (dv.formula1 or '').strip('\"') != ALLOWED_VALIDATION_VALUES:
        print('ERROR: Data validation list mismatch', file=sys.stderr)
        sys.exit(3)


def main():
    args = parse_args()
    records = load_json(args.json_data, args.json_file)

    # Derive filename if not provided
    if args.output_filename:
        filename = args.output_filename
    else:
        date_s, time_s = ist_now_str()
        filename = f"{args.ip_name}_TestPlan_{date_s}_{time_s}.xlsx"

    out_dir = args.output_dir
    if not out_dir.endswith('/'):
        out_dir += '/'
    output_path = os.path.join(out_dir, filename)
    # Normalize path to repo root
    output_path = os.path.normpath(output_path)

    # Build workbook
    build_workbook(records, output_path)

    # Emit sidecar with output path for the workflow commit step
    with open('.gen_last_output.json', 'w', encoding='utf-8') as f:
        json.dump({'output_path': output_path, 'rows': len(records), 'columns': len(union_keys(records))}, f)

    print(f"SUCCESS: Generated {output_path}")

if __name__ == '__main__':
    main()
