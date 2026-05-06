#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate a true binary .xlsx Test Plan from JSON input following strict Stage1 rules:
- JSON array (or dict of TC* objects) -> rows
- Preserve first-seen key order; union across inconsistent keys with blanks
- Create Data sheet, then reorganize the SAME sheet to TestPlan
- Create Very Hidden Meta_data_sheet with META columns
- Strict formatting, numbering, data validation (single column only)
- Final workbook contains ONLY: TestPlan (visible) and Meta_data_sheet (veryHidden)
- Validate as real Office Open XML (.xlsx) using zipfile
"""
import argparse
import json
import os
import sys
import zipfile
from collections import OrderedDict
from copy import deepcopy
from datetime import datetime, timedelta, timezone

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as e:
    print(f"ERROR: openpyxl not available: {e}", file=sys.stderr)
    sys.exit(2)

META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

VALIDATION_COL = "Code Generation (Required / Not)"
VALIDATION_LIST = ["Required", "Blank", "Not Required"]


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--json", required=True, help="Path to input JSON file (array or dict of TCs)")
    p.add_argument("--output-dir", required=True, help="Directory to write the final Excel")
    p.add_argument("--file-name", required=True, help="Output Excel file name. Supports placeholders <YYYYMMDD>_<HHMMSS>")
    p.add_argument("--ist-timestamp", action="store_true", help="Use IST now to fill placeholders in file name")
    return p.parse_args()


def load_json_array(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Allow dict of TC* -> object
    if isinstance(data, dict):
        # If looks like {"TC1": {...}, "TC2": {...}}
        values = list(data.values())
        if all(isinstance(v, dict) for v in values):
            return values
        else:
            raise ValueError("JSON dict must contain only testcase objects")
    if not isinstance(data, list):
        raise ValueError("JSON must be an array of objects or a dict of testcase objects")
    for i, row in enumerate(data):
        if not isinstance(row, dict):
            raise ValueError(f"JSON array element at index {i} is not an object")
    if len(data) == 0:
        raise ValueError("JSON array is empty")
    return data


def ist_now_str():
    ist = timezone(timedelta(hours=5, minutes=30))
    return datetime.now(ist).strftime("%Y%m%d_%H%M%S")


def resolve_output_path(output_dir, file_name, use_ist):
    name = file_name
    if use_ist and ("<YYYYMMDD>_<HHMMSS>" in name):
        name = name.replace("<YYYYMMDD>_<HHMMSS>", ist_now_str())
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, name)


def union_keys_preserve_order(rows):
    seen = OrderedDict()
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen[k] = None
    return list(seen.keys())


def autosize_columns(ws):
    # Compute max length per column
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_len = 0
        for cell in col:
            val = cell.value
            if val is None:
                continue
            s = str(val)
            for line in s.split("\n"):
                max_len = max(max_len, len(line))
        # width heuristic
        width = min(100, max(10, int(max_len * 1.1)))
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width


def add_borders(ws):
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in r:
            c.border = border


def number_items(text):
    if text is None:
        return text
    s = str(text).strip()
    if not s:
        return s
    # Split on explicit newlines. If no newline, return as-is.
    lines = s.split('\n')
    if len(lines) == 1:
        # already single line; keep as-is
        return s
    numbered = []
    n = 1
    for ln in lines:
        ln_s = ln.strip()
        if not ln_s:
            continue
        # Strip any existing leading bullets/numbers
        numbered.append(f"{n}. {ln_s}")
        n += 1
    return "\n".join(numbered) if numbered else s


def create_workbook(rows, output_path):
    # Phase 1: Data sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Determine columns (union preserve first-seen)
    all_cols = union_keys_preserve_order(rows)

    # Header
    ws.append(all_cols)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill("solid", fgColor="4F81BD")  # blue
    for c in ws[1]:
        c.font = header_font
        c.alignment = header_align
        c.fill = header_fill

    # Rows
    for row in rows:
        ws.append([row.get(k, "") for k in all_cols])

    # Freeze header
    ws.freeze_panes = "A2"

    # Base autosize
    autosize_columns(ws)

    # Phase 2: Meta_data_sheet creation
    ws_meta = wb.create_sheet(title="Meta_data_sheet")
    # Copy META columns AS-IS in the order defined
    meta_existing = [c for c in META_COLS if c in all_cols]
    if meta_existing:
        ws_meta.append(meta_existing)
        for row in rows:
            ws_meta.append([row.get(k, "") for k in meta_existing])
    # Very Hidden
    ws_meta.sheet_state = 'veryHidden'

    # Normalize main sheet on the SAME sheet: remove META cols and reorder to MAIN_ORDER
    # Build new column order for TestPlan
    main_existing = [c for c in MAIN_ORDER if c in all_cols]
    # Include any non-main, non-meta trailing columns (preserve first-seen order)
    trailing = [c for c in all_cols if c not in META_COLS and c not in MAIN_ORDER]
    final_order = main_existing + trailing

    # Build a map from header -> index in Data
    header_idx = {ws.cell(row=1, column=i+1).value: i for i in range(ws.max_column)}

    # Create a temporary 2D array of values in final order
    data_matrix = []
    data_matrix.append(final_order)
    for r in range(2, ws.max_row + 1):
        row_vals = []
        for col_name in final_order:
            idx = header_idx.get(col_name, None)
            if idx is None:
                row_vals.append("")
            else:
                row_vals.append(ws.cell(row=r, column=idx+1).value)
        data_matrix.append(row_vals)

    # Clear the Data sheet and rewrite in-place
    ws.delete_rows(1, ws.max_row)
    for r in data_matrix:
        ws.append(r)

    # Rename Data -> TestPlan (same sheet)
    ws.title = "TestPlan"

    # Enable wrap text for WRAP_COLS
    # Find column indices by header
    header_pos = {ws.cell(row=1, column=i+1).value: i+1 for i in range(ws.max_column)}
    wrap_cols_idx = [header_pos[c] for c in WRAP_COLS if c in header_pos]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in wrap_cols_idx:
            cell = row[col_idx-1]
            # Numbering for specific columns
            if ws.cell(row=1, column=col_idx).value in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                cell.value = number_items(cell.value)
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    # Reapply header formatting to TestPlan
    for c in ws[1]:
        c.font = header_font
        c.alignment = header_align
        c.fill = header_fill

    # General data alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Default left align, top vertical
            if cell.alignment is None:
                cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

    # Data validation ONLY for Code Generation (Required / Not)
    if VALIDATION_COL in header_pos:
        col_letter = ws.cell(row=1, column=header_pos[VALIDATION_COL]).column_letter
        dv = DataValidation(type="list", formula1=f'"{",".join(VALIDATION_LIST)}"', allow_blank=True, showErrorMessage=True)
        # Apply to data rows
        dv_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        dv.ranges.append(dv_range)
        ws.add_data_validation(dv)

    # Borders
    add_borders(ws)

    # Autosize after wrapping
    autosize_columns(ws)

    # Safety: only two sheets allowed
    for sh in list(wb.sheetnames):
        if sh not in ("TestPlan", "Meta_data_sheet"):
            del wb[sh]

    # Save
    wb.save(output_path)

    # Validate as ZIP-based OOXML
    if not zipfile.is_zipfile(output_path):
        raise ValueError("Generated file is not a valid ZIP-based .xlsx")
    with zipfile.ZipFile(output_path, 'r') as zf:
        if "[Content_Types].xml" not in zf.namelist():
            raise ValueError("[Content_Types].xml missing in .xlsx; invalid workbook")

    return output_path


def main():
    args = parse_args()
    rows = load_json_array(args.json)
    out_path = resolve_output_path(args.output_dir, args.file_name, args.ist_timestamp)
    path = create_workbook(rows, out_path)
    print(json.dumps({
        "status": "SUCCESS",
        "rows": len(rows),
        "output_path": path
    }))


if __name__ == "__main__":
    main()
