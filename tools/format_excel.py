#!/usr/bin/env python3
import argparse
import os
from datetime import datetime

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)',
]

WRAP_COLS = set([
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria',
])


def ist_now():
    if ZoneInfo is not None:
        return datetime.now(ZoneInfo('Asia/Kolkata'))
    # Fallback: IST = UTC+5:30
    return datetime.utcnow()


def make_output_name(rule: str) -> str:
    now = ist_now()
    yyyy = now.strftime('%Y')
    mm = now.strftime('%m')
    dd = now.strftime('%d')
    HH = now.strftime('%H')
    MM = now.strftime('%M')
    SS = now.strftime('%S')
    name = rule.replace('<YYYYMMDD>', f'{yyyy}{mm}{dd}').replace('<HHMMSS>', f'{HH}{MM}{SS}')
    return name


def pick_latest_xlsx(input_dir: str) -> str:
    xlsxs = [f for f in os.listdir(input_dir) if f.lower().endswith('.xlsx')]
    if not xlsxs:
        return ''
    # Deterministic: pick lexicographically largest filename
    xlsxs.sort()
    return os.path.join(input_dir, xlsxs[-1])


def get_header_map(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None and v != '':
            headers[str(v)] = col
    return headers


def copy_columns(src_ws, dst_ws, columns, header_map):
    col_idx = 1
    for name in columns:
        if name not in header_map:
            continue
        src_col = header_map[name]
        # header
        dst_ws.cell(row=1, column=col_idx, value=name)
        # data
        for r in range(2, src_ws.max_row + 1):
            dst_ws.cell(row=r, column=col_idx, value=src_ws.cell(row=r, column=src_col).value)
        col_idx += 1


def set_very_hidden(ws):
    try:
        ws.sheet_state = 'veryHidden'
    except Exception:
        # If not supported, keep as hidden (best effort without data change)
        ws.sheet_state = 'hidden'


def autofit_columns(ws):
    # Approximate autofit: compute max display length; cap width
    max_width = 100
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        width = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row, col).value
            if val is None:
                l = 0
            else:
                s = str(val)
                l = max([len(part) for part in s.split('\n')])
            if l > width:
                width = l
        width = min(max(10, width + 2), max_width)
        ws.column_dimensions[get_column_letter(col)].width = width


def autofit_row_heights(ws):
    base_height = 15
    for row in range(2, ws.max_row + 1):
        max_lines = 1
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if v is None:
                continue
            s = str(v)
            lines = s.count('\n') + 1
            if lines > max_lines:
                max_lines = lines
        ws.row_dimensions[row].height = base_height * max_lines


def apply_formatting(ws):
    # Header style
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_top_left = Alignment(horizontal='left', vertical='top', wrap_text=False)
    data_top_center = Alignment(horizontal='center', vertical='top', wrap_text=False)
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='DDDDDD')

    # Wrap selected columns
    header_names = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    wrap_mask = [ (name in WRAP_COLS) for name in header_names ]

    # Header row
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all
        cell.fill = header_fill

    # Data rows
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            # Horizontal alignment rule
            if header_names[c-1] == 'Index':
                cell.alignment = data_top_center
            else:
                # Wrap only for specified columns
                if wrap_mask[c-1]:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = data_top_left
            cell.border = border_all

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Autofit
    autofit_columns(ws)
    autofit_row_heights(ws)


def process(input_dir: str, output_rule: str):
    src_path = pick_latest_xlsx(input_dir)
    if not src_path or not os.path.exists(src_path):
        print('No source .xlsx found; nothing to do')
        return 0

    wb = load_workbook(src_path)

    # Identify first visible worksheet as main
    main_ws = None
    for ws in wb.worksheets:
        if getattr(ws, 'sheet_state', 'visible') == 'visible':
            main_ws = ws
            break
    if main_ws is None:
        # Fallback to first worksheet
        main_ws = wb.worksheets[0]

    # Build header map from main
    header_map = get_header_map(main_ws)

    # Create Meta_data_sheet and copy meta columns present
    if 'Meta_data_sheet' in [s.title for s in wb.worksheets]:
        meta_ws = wb['Meta_data_sheet']
        # Clear existing content
        for row in meta_ws[1:meta_ws.max_row]:
            for cell in row:
                cell.value = None
        for col in range(1, meta_ws.max_column + 1):
            col_letter = get_column_letter(col)
            meta_ws.column_dimensions[col_letter].width = None
    else:
        meta_ws = wb.create_sheet('Meta_data_sheet')

    copy_columns(main_ws, meta_ws, META_COLS, header_map)
    set_very_hidden(meta_ws)

    # Rename main sheet to TestPlan
    main_ws.title = 'TestPlan'

    # Build TestPlan in specified order using only columns that exist
    plan_ws = Workbook().active
    plan_ws.title = 'TestPlan_tmp'
    keep_cols = [name for name in MAIN_ORDER if name in header_map]
    copy_columns(main_ws, plan_ws, keep_cols, header_map)

    # Remove old TestPlan and replace
    wb.remove(wb['TestPlan'])
    wb._add_sheet(plan_ws)
    wb['TestPlan_tmp'].title = 'TestPlan'

    # Apply formatting to TestPlan only
    apply_formatting(wb['TestPlan'])

    # Save
    out_name = make_output_name(output_rule)
    out_path = os.path.join(input_dir, out_name)
    wb.save(out_path)
    print(f'Wrote: {out_path}')
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input-dir', required=True)
    ap.add_argument('--output-rule', required=True)
    args = ap.parse_args()
    return process(args.input_dir, args.output_rule)

if __name__ == '__main__':
    raise SystemExit(main())
