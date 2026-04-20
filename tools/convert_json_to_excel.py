#!/usr/bin/env python3
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except Exception as e:
    print(f"ERROR: openpyxl not available: {e}", file=sys.stderr)
    sys.exit(2)

def to_cell(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if value is None:
        return ""
    return str(value)

def main():
    ap = argparse.ArgumentParser(description="Convert JSON (single object or array of objects) to a single-sheet Excel file")
    ap.add_argument("--input", required=True, help="Path to input JSON file")
    ap.add_argument("--output", required=True, help="Path to output .xlsx file")
    ap.add_argument("--sheet-name", default="TestPlan", help="Worksheet name (default: TestPlan)")
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        raw = in_path.read_text(encoding="utf-8")
        data = json.loads(raw)
    except Exception as e:
        print(f"ERROR: Failed to read/parse JSON: {e}", file=sys.stderr)
        sys.exit(2)

    # Normalize to tabular rows per requirements
    rows = []
    headers = []

    if isinstance(data, list):
        # Union of keys in first-seen order
        seen = {}
        for obj in data:
            if not isinstance(obj, dict):
                print("ERROR: Array elements must be JSON objects", file=sys.stderr)
                sys.exit(2)
            for k in obj.keys():
                if k not in seen:
                    seen[k] = None
        headers = list(seen.keys())
        for obj in data:
            row = [to_cell(obj.get(h, "")) for h in headers]
            rows.append(row)
    elif isinstance(data, dict):
        headers = list(data.keys())  # preserves insertion order
        rows.append([to_cell(data.get(h, "")) for h in headers])
    else:
        print("ERROR: Unsupported JSON structure (must be object or array of objects)", file=sys.stderr)
        sys.exit(2)

    # Build workbook
    wb = Workbook()
    # Remove default sheet to ensure exactly one sheet we control
    default_ws = wb.active
    wb.remove(default_ws)
    ws = wb.create_sheet(title=args.sheet_name)

    # Header row
    bold_font = Font(bold=True)
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = bold_font

    # Data rows
    for r, row_vals in enumerate(rows, start=2):
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=r, column=c, value=val)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Naive auto-fit based on character length
    from openpyxl.utils import get_column_letter
    col_widths = [len(str(h)) for h in headers]
    for row_vals in rows:
        for idx, val in enumerate(row_vals):
            l = len(str(val)) if val is not None else 0
            if l > col_widths[idx]:
                col_widths[idx] = l
    for i, w in enumerate(col_widths, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = min(max(w + 2, 10), 120)

    try:
        wb.save(out_path)
    except Exception as e:
        print(f"ERROR: Failed to save Excel file: {e}", file=sys.stderr)
        sys.exit(2)

    print(f"Saved Excel to {out_path}")

if __name__ == "__main__":
    main()
