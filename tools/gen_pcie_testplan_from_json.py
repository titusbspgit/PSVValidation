#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse
import json
import os
from datetime import datetime, timezone, timedelta
from zipfile import ZipFile
from io import BytesIO
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Embedded JSON data (FULL_JSON_STRUCTURE array of 3 objects)
JSON_DATA = [
  {
    "Index": 1,
    "SS / Module": "PCIE0 SII RC",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie0_sii_rc_reg_wr_rd_test",
    "Test Description": "Verify that the PCIe SII Root Complex registers power up with documented defaults and that writable fields retain written values while read-only fields remain unchanged.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Unreadable addresses are skipped. Unwritable addresses are skipped. The SII_PHY_RST_CONTROL register is excluded from default checks. The optional soft reset is not executed.",
    "Test Steps / Procedure": "1) Read each register in the PCIe SII Root Complex and compare with its documented default.\n2) Exclude SII_PHY_RST_CONTROL from default checks.\n3) Write test patterns to each writable register.\n4) Read back each readable register and derive the expected value using the read and write masks with the defaults.\n5) Count any mismatches and determine pass or fail at the end.",
    "Impacted Registers": "SII_CFG_BAR0_START1, SII_CFG_BAR0_START2, SII_CFG_BAR0_LIMIT1, SII_CFG_BAR0_LIMIT2, SII_PHY_RST_CONTROL, SII_SOFT_RESET_CTRL, SII_MSI_CTRL_IO, SII_MSI_CTRL_INT_VEC",
    "Validation / Acceptance Criteria": "1) Reset read equals the documented default for each readable register → Pass.\n2) Post-write read equals the expected masked value for each readable and writable register → Pass.\n3) No mismatches in default or write-read checks → Pass; any mismatch → Fail.",
    "Code Generation (Required / Not)": "",
    "Header Includes": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<pcie0/pcie_sii_rc_def.h>\n#include<pcie0/pcie_sii_rc_offset.h>",
    "Macro Defines": "#define SOFT_RST_REG_ADDRESS\t0x00000000\n#define SOFT_RST_REG_DATA\t0x00000000\n#define MIZAR_PCIE0_SII_BASE     0xE68C0000\n#define CNT 153",
    "Skip Array Definition / Declaration": "const int skip_array[153]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,}",
    "Hidden_Test_Case_Name": "pcie0_sii_rc_reg_wr_rd_test",
    "Hidden_Test_Description": "test_case() calls chk_rst_val() then chk_rd_wr(). DEFAULT VALUE_CHECK: For i in [0..CNT-1], addr=addr_array[i]; if read_mask_array[i]==0x00000000, print skip and continue; if addr_array[i]==mizar_PCIE0_SII_PHY_RST_CONTROL, continue; data_rd=read_reg(addr); if data_rd==default_value_array[i], optionally print PASS; else increment def_fail_cnt and print failure. WRITE & READ CHECK: int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}. For each pattern j, data_wr=chk_val[j]; Write phase: for i in [0..CNT-1], addr=addr_array[i]; if skip_array[i]==1 continue; if write_mask_array[i]==0x00000000 continue; else write_reg(addr,data_wr). Read phase: for i in [0..CNT-1], if skip_array[i]==1 continue; if write_mask_array[i]==0x00000000 continue; if read_mask_array[i]==0x00000000 continue; else data_rd=read_reg(addr); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if data_rd==exp_val print PASS; else wr_fail_cnt++ and print failure. End: if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0). soft_reset_chk() exists but is commented out in test_case().",
    "Hidden_Remarks": "Registers with read_mask_array[i]==0x00000000 are skipped as not readable. Registers with write_mask_array[i]==0x00000000 are skipped as not writable. The address equal to mizar_PCIE0_SII_PHY_RST_CONTROL is excluded from default value checks. The optional soft reset sequence using SOFT_RST_REG_ADDRESS and SOFT_RST_REG_DATA is present but not executed.",
    "Hidden_Test_Steps_Procedure": "Entry: test_case(). 1) Call chk_rst_val(): for(i=0;i<CNT;i++){ addr=addr_array[i]; if(read_mask_array[i]==0x00000000){ print skip; continue; } if(addr_array[i]==mizar_PCIE0_SII_PHY_RST_CONTROL){ continue; } data_rd=read_reg(addr); if(data_rd==default_value_array[i]){ optional PASS print; } else { def_fail_cnt++; print failure; } } 2) Call chk_rd_wr(): Define chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}. For(j=0;j<6;j++){ data_wr=chk_val[j]; Write phase: for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ print skip; continue; } if(write_mask_array[i]==0x00000000){ print skip; continue; } else { write_reg(addr,data_wr); optional print; } } Read/verify phase: for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ print skip; continue; } if(write_mask_array[i]==0x00000000){ print skip; continue; } if(read_mask_array[i]==0x00000000){ print skip; continue; } else { data_rd=read_reg(addr); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd==exp_val){ optional PASS print; } else { wr_fail_cnt++; print failure; } } } } 3) Finalize: if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0). Note: soft_reset_chk() (reads SOFT_RST_REG_ADDRESS, writes SOFT_RST_REG_DATA, waits, restores) exists but is not invoked.",
    "Hidden_Impacted_Registers": "mizar_PCIE0_SII_CFG_BAR0_START1,mizar_PCIE0_SII_CFG_BAR0_START2,mizar_PCIE0_SII_CFG_BAR0_LIMIT1,mizar_PCIE0_SII_CFG_BAR0_LIMIT2,mizar_PCIE0_SII_CFG_BAR1_START,mizar_PCIE0_SII_CFG_BAR1_LIMIT1,mizar_PCIE0_SII_CFG_BAR2_START1,mizar_PCIE0_SII_CFG_BAR2_START2,mizar_PCIE0_SII_CFG_BAR2_LIMIT1,mizar_PCIE0_SII_CFG_BAR2_LIMIT2,mizar_PCIE0_SII_CFG_BAR3_START,mizar_PCIE0_SII_CFG_BAR3_LIMIT,mizar_PCIE0_SII_CFG_BAR4_START1,mizar_PCIE0_SII_CFG_BAR4_START2,mizar_PCIE0_SII_CFG_BAR4_LIMIT1,mizar_PCIE0_SII_CFG_BAR4_LIMIT2,mizar_PCIE0_SII_CFG_BAR5_START,mizar_PCIE0_SII_CFG_BAR5_LIMIT,mizar_PCIE0_SII_PCIE0_CONFIG_INFO1,mizar_PCIE0_SII_PCIE0_CONFIG_INFO2,mizar_PCIE0_SII_PCIE0_GEN_CONTROL1,mizar_PCIE0_SII_PCIE0_GEN_CONTROL2,mizar_PCIE0_SII_PCIE0_GEN_CONTROL3,mizar_PCIE0_SII_PCIE0_PM_CONTROL,mizar_PCIE0_SII_PCIE0_CONTROL_PM_STS,mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER1,mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER2,mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER3,mizar_PCIE0_SII_PCIE0_TRANSMIT_HEADER4,mizar_PCIE0_SII_PCIE0_TRANSMIT_REQ,mizar_PCIE0_SII_PCIE0_RCV_MSG_HDR1,mizar_PCIE0_SII_PCIE0_RCV_MSG_HDR2,mizar_PCIE0_SII_PCIE0_RCV_MSG_HDR3,mizar_PCIE0_SII_PCIE0_RCV_MSG_HDR4,mizar_PCIE0_SII_PCIE0_RCV_MSG_STS,mizar_PCIE0_SII_RCV_INTERRPUT_CTRL,mizar_PCIE0_SII_CFG_EXP_ROM_START,mizar_PCIE0_SII_CFG_EXP_ROM_LIMIT,mizar_PCIE0_SII_CFG_EXP_ROM_INFO,mizar_PCIE0_SII_CXPL_DEBUG_INFO1,mizar_PCIE0_SII_CXPL_DEBUG_INFO2,mizar_PCIE0_SII_CXPL_DEBUG_INFO_EI,mizar_PCIE0_SII_PCIE0_TARGET_INFO1,mizar_PCIE0_SII_PCIE0_TARGET_INFO2,mizar_PCIE0_SII_PCIE0_CONTOLLER_ERROR_STATUS,mizar_PCIE0_SII_PCIE0_CONTROLLER_INT_STS,mizar_PCIE0_SII_PCIE0_CONTROLLER_INTERRUPT_CONTROL,mizar_PCIE0_SII_PHY_RST_CONTROL,mizar_PCIE0_SII_LINK_DEBUG_DATA,mizar_PCIE0_SII_PCIE0_ERR_STS,mizar_PCIE0_SII_PCIE0_ERR_INTERRUPT_CTRL,mizar_PCIE0_SII_CFG_MSI_INT,mizar_PCIE0_SII_LTR_MSG,mizar_PCIE0_SII_LTR_MSG_LATENCY,mizar_PCIE0_SII_APP_LTR_LATENCY,mizar_PCIE0_SII_CFG_LTR_MAX_LATENCY,mizar_PCIE0_SII_OBFF_CNTRL,mizar_PCIE0_SII_SLV_AWMISC_INFO,mizar_PCIE0_SII_SLV_AWMISC_INFO_HDR_34DW_HI,mizar_PCIE0_SII_SLV_AWMISC_INFO_HDR_34DW_LO,mizar_PCIE0_SII_SLV_MISC_INFO,mizar_PCIE0_SII_SLV_MISC_RESP_INFO,mizar_PCIE0_SII_MSTR_AWMISC_INFO_CNTRL,mizar_PCIE0_SII_MSTR_AWMISC_INFO_1,mizar_PCIE0_SII_MSTR_AWMISC_INFO_0,mizar_PCIE0_SII_MSTR_AWMISC_INFO_HDR_34DW_HI,mizar_PCIE0_SII_MSTR_AWMISC_INFO_HDR_34DW_LO,mizar_PCIE0_SII_MSTR_ARMISC_INFO_CNTRL,mizar_PCIE0_SII_MSTR_ARMISC_INFO_1,mizar_PCIE0_SII_MSTR_ARMISC_INFO_0,mizar_PCIE0_SII_MSTR_BMISC_RMISC_CPL_STAT_INFO,mizar_PCIE0_SII_RADM_TIMEOUT_INFO,mizar_PCIE0_SII_CFG_MSI_INFO,mizar_PCIE0_SII_CFG_MSI_DATA,mizar_PCIE0_SII_CFG_MSI_ADDR_HI,mizar_PCIE0_SII_CFG_MSI_ADDR_LO,mizar_PCIE0_SII_CFG_AER_INT_AND_PCIE0_CAP_INT_MSG,mizar_PCIE0_SII_RTLH_RFC_DATA,mizar_PCIE0_SII_APP_HDR_INFO,mizar_PCIE0_SII_APP_HDR_LOG_3,mizar_PCIE0_SII_APP_HDR_LOG_2,mizar_PCIE0_SII_APP_HDR_LOG_1,mizar_PCIE0_SII_APP_HDR_LOG_0,mizar_PCIE0_SII_CFG_BUS_NUM,mizar_PCIE0_SII_CFG_BR_CTRL_SERREN,mizar_PCIE0_SII_APP_DEV_AND_BUS_NUM,mizar_PCIE0_SII_PCIE0_CONTROLLER_INT_STS_1,mizar_PCIE0_SII_PCIE0_CONTROLLER_INTERRUPT_CONTROL_1,mizar_PCIE0_SII_APP_AND_SLOT_CONTROL_REG,mizar_PCIE0_SII_DIAG_CTRL_BUS,mizar_PCIE0_SII_CFG_REG_RO,mizar_PCIE0_SII_CFG_ARI_FWD_EN,mizar_PCIE0_SII_RADM_SLOT_PWR_PAYLOAD,mizar_PCIE0_SII_DIAG_STATUS_BUS_0,mizar_PCIE0_SII_DIAG_STATUS_BUS_1,mizar_PCIE0_SII_DIAG_STATUS_BUS_2,mizar_PCIE0_SII_DIAG_STATUS_BUS_3,mizar_PCIE0_SII_DIAG_STATUS_BUS_4,mizar_PCIE0_SII_DIAG_STATUS_BUS_5,mizar_PCIE0_SII_DIAG_STATUS_BUS_6,mizar_PCIE0_SII_DIAG_STATUS_BUS_7,mizar_PCIE0_SII_DIAG_STATUS_BUS_8,mizar_PCIE0_SII_DIAG_STATUS_BUS_9,mizar_PCIE0_SII_DIAG_STATUS_BUS_10,mizar_PCIE0_SII_DIAG_STATUS_BUS_11,mizar_PCIE0_SII_DIAG_STATUS_BUS_12,mizar_PCIE0_SII_DIAG_STATUS_BUS_13,mizar_PCIE0_SII_DIAG_STATUS_BUS_14,mizar_PCIE0_SII_DIAG_STATUS_BUS_15,mizar_PCIE0_SII_DIAG_STATUS_BUS_16,mizar_PCIE0_SII_DIAG_STATUS_BUS_17,mizar_PCIE0_SII_DIAG_STATUS_BUS_18,mizar_PCIE0_SII_DIAG_STATUS_BUS_19,mizar_PCIE0_SII_RAM_PWR_CNTRL_0,mizar_PCIE0_SII_RAM_PWR_CNTRL_1,mizar_PCIE0_SII_SOFT_RESET_CTRL,mizar_PCIE0_SII_CFG_MSI_PENDING_B,mizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_1,mizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_2,mizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_3,mizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_4,mizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_5,mizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_6,mizar_PCIE0_SII_SMLH_LTSSM_STATE_TRAN_7,mizar_PCIE0_SII_PHY_CONTROL_0,mizar_PCIE0_SII_PHY_CONTROL_1,mizar_PCIE0_SII_PHY_CONTROL_2,mizar_PCIE0_SII_PHY_CONTROL_3,mizar_PCIE0_SII_PHY_CONTROL_4,mizar_PCIE0_SII_PHY_CONTROL_5,mizar_PCIE0_SII_PHY_CONTROL_6,mizar_PCIE0_SII_PHY_CONTROL_7,mizar_PCIE0_SII_PHY_CONTROL_8,mizar_PCIE0_SII_PHY_CONTROL_9,mizar_PCIE0_SII_PHY_CONTROL_10,mizar_PCIE0_SII_PHY_CONTROL_11,mizar_PCIE0_SII_PHY_CONTROL_12,mizar_PCIE0_SII_PHY_CONTROL_13,mizar_PCIE0_SII_PHY_CONTROL_14,mizar_PCIE0_SII_PHY_CONTROL_15,mizar_PCIE0_SII_PHY_CONTROL_16,mizar_PCIE0_SII_PHY_CONTROL_17,mizar_PCIE0_SII_PHY_CONTROL_18,mizar_PCIE0_SII_PHY_CONTROL_19,mizar_PCIE0_SII_PHY_CONTROL_20,mizar_PCIE0_SII_PHY_CONTROL_21,mizar_PCIE0_SII_PHY_CONTROL_22,mizar_PCIE0_SII_PHY_CONTROL_23,mizar_PCIE0_SII_PHY_CONTROL_24,mizar_PCIE0_SII_PHY_CONTROL_25,mizar_PCIE0_SII_PHY_CONTROL_26,mizar_PCIE0_SII_MSI_CTRL_IO,mizar_PCIE0_SII_MSI_CTRL_INT_VEC",
    "Hidden_Validation_Acceptance_Criteria": "Default value check: For each i in [0..CNT-1], if read_mask_array[i]==0x00000000 then skip; if addr_array[i]==mizar_PCIE0_SII_PHY_RST_CONTROL then skip; else read_reg(addr_array[i]) and compare to default_value_array[i]; PASS when equal, else increment def_fail_cnt and log failure. Write/read check: For each pattern in {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}, write to each address where skip_array[i]!=1 and write_mask_array[i]!=0x00000000; then for each such address where read_mask_array[i]!=0x00000000, read back and compute wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); PASS when data_rd==exp_val, else increment wr_fail_cnt and log failure. Final status: finish(0) when def_fail_cnt==0 and wr_fail_cnt==0; otherwise finish(1).",
    "Hidden_Header_Includes": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<pcie0/pcie_sii_rc_def.h>\n#include<pcie0/pcie_sii_rc_offset.h>",
    "Hidden_Macro_Defines": "#define SOFT_RST_REG_ADDRESS\t0x00000000\n#define SOFT_RST_REG_DATA\t0x00000000\n#define MIZAR_PCIE0_SII_BASE     0xE68C0000\n#define CNT 153",
    "Hidden_Skip_Array_Definition": "const int skip_array[153]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,}"
  },
  {
    "Index": 2,
    "SS / Module": "PCIE1 SII RC",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Test Description": "Verify PCIe1 SII Root Complex registers for reset defaults and masked write-read behavior across all addresses using defined read/write masks and default arrays. The flow performs a default value check, then iterates multiple data patterns to validate that writable fields accept values while read-only fields retain defaults; optional soft reset routine exists but is not executed in the main flow.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Unreadable addresses (read_mask_array[i] == 0x00000000) are skipped. Unwritable addresses (write_mask_array[i] == 0x00000000) are skipped. Address equal to mizar_PCIE1_SII_PHY_RST_CONTROL is excluded from default checks. Optional soft reset sequence is defined but commented out in test_case().",
    "Test Steps / Procedure": "Entry: test_case(). 1) Call chk_rst_val(). In chk_rst_val(): for(i=0;i<CNT;i++){ addr=addr_array[i]; if(read_mask_array[i]==0x00000000){ optional print and continue; } if(addr_array[i]==mizar_PCIE1_SII_PHY_RST_CONTROL){ continue; } data_rd=read_reg(addr); if(data_rd==default_value_array[i]){ optional PASS print; } else { def_fail_cnt++; printf failure; } }. 2) Call chk_rd_wr(). In chk_rd_wr(): int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}; for(j=0;j<6;j++){ data_wr=chk_val[j]; // Write phase: for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ optional print and continue; } if(write_mask_array[i]==0x00000000){ optional print and continue; } else { write_reg(addr,data_wr); optional print; } } // Read/verify phase: for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ optional print and continue; } if(write_mask_array[i]==0x00000000){ optional print and continue; } if(read_mask_array[i]==0x00000000){ optional print and continue; } else { data_rd=read_reg(addr); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd==exp_val){ optional PASS print; } else { wr_fail_cnt++; printf failure; } } } }. 3) Finalize: if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0). Note: soft_reset_chk() is not called in test_case(); its logic is default_value=read_reg(SOFT_RST_REG_ADDRESS); write_reg(SOFT_RST_REG_ADDRESS,SOFT_RST_REG_DATA); wait_on(1000); write_reg(SOFT_RST_REG_ADDRESS,default_value); wait_on(1000).",
    "Impacted Registers": "SII_CFG_BAR0_START1, SII_CFG_BAR0_START2, SII_CFG_BAR0_LIMIT1, SII_CFG_BAR0_LIMIT2, SII_CFG_BAR1_START, SII_CFG_BAR1_LIMIT1, SII_CFG_BAR2_START1, SII_CFG_BAR2_START2, SII_CFG_BAR2_LIMIT1, SII_CFG_BAR2_LIMIT2, SII_CFG_BAR3_START, SII_CFG_BAR3_LIMIT, BAR4_MASK_REG, BAR4_MASK_REG, BAR4_MASK_REG, BAR4_MASK_REG, BAR5_MASK_REG, BAR5_MASK_REG, SII_PCIE1_CONFIG_INFO1, SII_PCIE1_CONFIG_INFO2, SII_PCIE1_GEN_CONTROL1, SII_PCIE1_GEN_CONTROL2, SII_PCIE1_GEN_CONTROL3, SII_PCIE1_PM_CONTROL, SII_PCIE1_CONTROL_PM_STS, SII_PCIE1_TRANSMIT_HEADER1, SII_PCIE1_TRANSMIT_HEADER2, SII_PCIE1_TRANSMIT_HEADER3, SII_PCIE1_TRANSMIT_HEADER4, SII_PCIE1_TRANSMIT_REQ, SII_PCIE1_RCV_MSG_HDR1, SII_PCIE1_RCV_MSG_HDR2, SII_PCIE1_RCV_MSG_HDR3, SII_PCIE1_RCV_MSG_HDR4, SII_PCIE1_RCV_MSG_STS, SII_RCV_INTERRPUT_CTRL, EXP_ROM_BAR_MASK_REG, EXP_ROM_BAR_MASK_REG, EXP_ROM_BAR_MASK_REG, SII_CXPL_DEBUG_INFO1, SII_CXPL_DEBUG_INFO2, SII_CXPL_DEBUG_INFO_EI, SII_PCIE1_TARGET_INFO1, SII_PCIE1_TARGET_INFO2, SII_PCIE1_CONTOLLER_ERROR_STATUS, SII_PCIE1_CONTROLLER_INT_STS, SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL, SII_PHY_RST_CONTROL, SII_LINK_DEBUG_DATA, SII_PCIE1_ERR_STS, SII_PCIE1_ERR_INTERRUPT_CTRL, SII_CFG_MSI_INT, SII_LTR_MSG, SII_LTR_MSG_LATENCY, SII_APP_LTR_LATENCY, SII_CFG_LTR_MAX_LATENCY, SII_OBFF_CNTRL, SII_SLV_AWMISC_INFO, SII_SLV_AWMISC_INFO_HDR_34DW_HI, SII_SLV_AWMISC_INFO_HDR_34DW_LO, SII_SLV_MISC_INFO, SII_SLV_MISC_RESP_INFO, SII_MSTR_AWMISC_INFO_CNTRL, SII_MSTR_AWMISC_INFO_1, SII_MSTR_AWMISC_INFO_0, SII_MSTR_AWMISC_INFO_HDR_34DW_HI, SII_MSTR_AWMISC_INFO_HDR_34DW_LO, SII_MSTR_ARMISC_INFO_CNTRL, SII_MSTR_ARMISC_INFO_1, SII_MSTR_ARMISC_INFO_0, SII_MSTR_BMISC_RMISC_CPL_STAT_INFO, SII_RADM_TIMEOUT_INFO, SII_CFG_MSI_INFO, SII_CFG_MSI_DATA, SII_CFG_MSI_ADDR_HI, SII_CFG_MSI_ADDR_LO, SII_CFG_AER_INT AND_PCIE1_CAP_INT_MSG, SII_RTLH_RFC_DATA, SII_APP_HDR_INFO, SII_APP_HDR_LOG_3, SII_APP_HDR_LOG_2, SII_APP_HDR_LOG_1, SII_APP_HDR_LOG_0, SII_CFG_BUS_NUM, SII_CFG_BR_CTRL_SERREN, SII_APP_DEV_AND_BUS_NUM, SII_PCIE1_CONTROLLER_INT_STS_1, SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL_1, SII_APP_AND_SLOT_CONTROL_REG, SII_DIAG_CTRL_BUS, SII_CFG_REG_RO, SII_CFG_ARI_FWD_EN, SII_RADM_SLOT_PWR_PAYLOAD, SII_DIAG_STATUS_BUS_0, SII_DIAG_STATUS_BUS_1, SII_DIAG_STATUS_BUS_2, SII_DIAG_STATUS_BUS_3, SII_DIAG_STATUS_BUS_4, SII_DIAG_STATUS_BUS_5, SII_DIAG_STATUS_BUS_6, SII_DIAG_STATUS_BUS_7, SII_DIAG_STATUS_BUS_8, SII_DIAG_STATUS_BUS_9, SII_DIAG_STATUS_BUS_10, SII_DIAG_STATUS_BUS_11, SII_DIAG_STATUS_BUS_12, SII_DIAG_STATUS_BUS_13, SII_DIAG_STATUS_BUS_14, SII_DIAG_STATUS_BUS_15, SII_DIAG_STATUS_BUS_16, SII_DIAG_STATUS_BUS_17, SII_DIAG_STATUS_BUS_18, SII_DIAG_STATUS_BUS_19, SII_RAM_PWR_CNTRL_0, SII_RAM_PWR_CNTRL_1, SII_SOFT_RESET_CTRL, SII_CFG_MSI_PENDING_B, SII_SMLH_LTSSM_STATE_TRAN_1, SII_SMLH_LTSSM_STATE_TRAN_2, SII_SMLH_LTSSM_STATE_TRAN_3, SII_SMLH_LTSSM_STATE_TRAN_4, SII_SMLH_LTSSM_STATE_TRAN_5, SII_SMLH_LTSSM_STATE_TRAN_6, SII_SMLH_LTSSM_STATE_TRAN_7, SII_PHY_CONTROL_0, SII_PHY_CONTROL_1, SII_PHY_CONTROL_2, SII_PHY_CONTROL_3, SII_PHY_CONTROL_4, SII_PHY_CONTROL_5, SII_PHY_CONTROL_6, SII_PHY_CONTROL_7, SII_PHY_CONTROL_8, SII_PHY_CONTROL_9, SII_PHY_CONTROL_10, SII_PHY_CONTROL_11, SII_PHY_CONTROL_12, SII_PHY_CONTROL_13, SII_PHY_CONTROL_14, SII_PHY_CONTROL_15, SII_PHY_CONTROL_16, SII_PHY_CONTROL_17, SII_PHY_CONTROL_18, SII_PHY_CONTROL_19, SII_PHY_CONTROL_20, SII_PHY_CONTROL_21, SII_PHY_CONTROL_22, SII_PHY_CONTROL_23, SII_PHY_CONTROL_24, SII_PHY_CONTROL_25, SII_PHY_CONTROL_26, SII_MSI_CTRL_IO, SII_MSI_CTRL_INT_VEC",
    "Validation / Acceptance Criteria": "Default value check: For each i in [0..CNT-1], if read_mask_array[i]==0x00000000 then skip; if addr_array[i]==mizar_PCIE1_SII_PHY_RST_CONTROL then skip; else read_reg(addr_array[i]) and compare to default_value_array[i]; PASS when equal, else increment def_fail_cnt and print failure. Write/read check: For each pattern in {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}, write to each address where skip_array[i]!=1 and write_mask_array[i]!=0x00000000; then for each such address where read_mask_array[i]!=0x00000000, read back and compute wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); PASS when data_rd==exp_val, else increment wr_fail_cnt and print failure. Final result: finish(0) if def_fail_cnt==0 and wr_fail_cnt==0; otherwise finish(1).",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Hidden_Test_Description": "Verify PCIe1 SII Root Complex registers for reset defaults and masked write-read behavior across all addresses using defined read/write masks and default arrays. The flow performs a default value check, then iterates multiple data patterns to validate that writable fields accept values while read-only fields retain defaults; optional soft reset routine exists but is not executed in the main flow.",
    "Hidden_Remarks": "Unreadable addresses (read_mask_array[i] == 0x00000000) are skipped. Unwritable addresses (write_mask_array[i] == 0x00000000) are skipped. Address equal to mizar_PCIE1_SII_PHY_RST_CONTROL is excluded from default checks. Optional soft reset sequence is defined but commented out in test_case().",
    "Hidden_Test_Steps_Procedure": "Entry: test_case(). 1) Call chk_rst_val(). In chk_rst_val(): for(i=0;i<CNT;i++){ addr=addr_array[i]; if(read_mask_array[i]==0x00000000){ optional print and continue; } if(addr_array[i]==mizar_PCIE1_SII_PHY_RST_CONTROL){ continue; } data_rd=read_reg(addr); if(data_rd==default_value_array[i]){ optional PASS print; } else { def_fail_cnt++; printf failure; } }. 2) Call chk_rd_wr(). In chk_rd_wr(): int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}; for(j=0;j<6;j++){ data_wr=chk_val[j]; // Write phase: for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ optional print and continue; } if(write_mask_array[i]==0x00000000){ optional print and continue; } else { write_reg(addr,data_wr); optional print; } } // Read/verify phase: for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ optional print and continue; } if(write_mask_array[i]==0x00000000){ optional print and continue; } if(read_mask_array[i]==0x00000000){ optional print and continue; } else { data_rd=read_reg(addr); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd==exp_val){ optional PASS print; } else { wr_fail_cnt++; printf failure; } } } }. 3) Finalize: if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0). Note: soft_reset_chk() is not called in test_case(); its logic is default_value=read_reg(SOFT_RST_REG_ADDRESS); write_reg(SOFT_RST_REG_ADDRESS,SOFT_RST_REG_DATA); wait_on(1000); write_reg(SOFT_RST_REG_ADDRESS,default_value); wait_on(1000).",
    "Hidden_Impacted_Registers": "mizar_PCIE1_SII_CFG_BAR0_START1,mizar_PCIE1_SII_CFG_BAR0_START2,mizar_PCIE1_SII_CFG_BAR0_LIMIT1,mizar_PCIE1_SII_CFG_BAR0_LIMIT2,mizar_PCIE1_SII_CFG_BAR1_START,mizar_PCIE1_SII_CFG_BAR1_LIMIT1,mizar_PCIE1_SII_CFG_BAR2_START1,mizar_PCIE1_SII_CFG_BAR2_START2,mizar_PCIE1_SII_CFG_BAR2_LIMIT1,mizar_PCIE1_SII_CFG_BAR2_LIMIT2,mizar_PCIE1_SII_CFG_BAR3_START,mizar_PCIE1_SII_CFG_BAR3_LIMIT,mizar_PCIE1_SII_CFG_BAR4_START1,mizar_PCIE1_SII_CFG_BAR4_START2,mizar_PCIE1_SII_CFG_BAR4_LIMIT1,mizar_PCIE1_SII_CFG_BAR4_LIMIT2,mizar_PCIE1_SII_CFG_BAR5_START,mizar_PCIE1_SII_CFG_BAR5_LIMIT,mizar_PCIE1_SII_PCIE1_CONFIG_INFO1,mizar_PCIE1_SII_PCIE1_CONFIG_INFO2,mizar_PCIE1_SII_PCIE1_GEN_CONTROL1,mizar_PCIE1_SII_PCIE1_GEN_CONTROL2,mizar_PCIE1_SII_PCIE1_GEN_CONTROL3,mizar_PCIE1_SII_PCIE1_PM_CONTROL,mizar_PCIE1_SII_PCIE1_CONTROL_PM_STS,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER1,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER4,mizar_PCIE1_SII_PCIE1_TRANSMIT_REQ,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR1,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR2,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR3,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR4,mizar_PCIE1_SII_PCIE1_RCV_MSG_STS,mizar_PCIE1_SII_RCV_INTERRPUT_CTRL,mizar_PCIE1_SII_CFG_EXP_ROM_START,mizar_PCIE1_SII_CFG_EXP_ROM_LIMIT,mizar_PCIE1_SII_CFG_EXP_ROM_INFO,mizar_PCIE1_SII_CXPL_DEBUG_INFO1,mizar_PCIE1_SII_CXPL_DEBUG_INFO2,mizar_PCIE1_SII_CXPL_DEBUG_INFO_EI,mizar_PCIE1_SII_PCIE1_TARGET_INFO1,mizar_PCIE1_SII_PCIE1_TARGET_INFO2,mizar_PCIE1_SII_PCIE1_CONTOLLER_ERROR_STATUS,mizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS,mizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL,mizar_PCIE1_SII_PHY_RST_CONTROL,mizar_PCIE1_SII_LINK_DEBUG DATA,mizar_PCIE1_SII_PCIE1_ERR_STS,mizar_PCIE1_SII_PCIE1_ERR_INTERRUPT_CTRL,mizar_PCIE1_SII_CFG_MSI_INT,mizar_PCIE1_SII_LTR_MSG,mizar_PCIE1_SII_LTR_MSG_LATENCY,mizar_PCIE1_SII_APP_LTR_LATENCY,mizar_PCIE1_SII_CFG_LTR_MAX_LATENCY,mizar_PCIE1_SII_OBFF_CNTRL,mizar_PCIE1_SII_SLV_AWMISC_INFO,mizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_HI,mizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_LO,mizar_PCIE1_SII_SLV_MISC_INFO,mizar_PCIE1_SII_SLV_MISC_RESP_INFO,mizar_PCIE1_SII_MSTR_AWMISC_INFO_CNTRL,mizar_PCIE1_SII_MSTR_AWMISC_INFO_1,mizar_PCIE1_SII_MSTR_AWMISC_INFO_0,mizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_HI,mizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_LO,mizar_PCIE1_SII_MSTR_ARMISC_INFO_CNTRL,mizar_PCIE1_SII_MSTR_ARMISC_INFO_1,mizar_PCIE1_SII_MSTR_ARMISC_INFO_0,mizar_PCIE1_SII_MSTR_BMISC_RMISC_CPL_STAT_INFO,mizar_PCIE1_SII_RADM_TIMEOUT_INFO,mizar_PCIE1_SII_CFG_MSI_INFO,mizar_PCIE1_SII_CFG_MSI_DATA,mizar_PCIE1_SII_CFG_MSI_ADDR_HI,mizar_PCIE1_SII_CFG_MSI_ADDR_LO,mizar_PCIE1_SII_CFG_AER_INT_AND_PCIE1_CAP_INT_MSG,mizar_PCIE1_SII_RTLH_RFC_DATA,mizar_PCIE1_SII_APP_HDR_INFO,mizar_PCIE1_SII_APP_HDR_LOG_3,mizar_PCIE1_SII_APP_HDR_LOG_2,mizar_PCIE1_SII_APP_HDR_LOG_1,mizar_PCIE1_SII_APP_HDR_LOG_0,mizar_PCIE1_SII_CFG_BUS_NUM,mizar_PCIE1_SII_CFG_BR_CTRL_SERREN,mizar_PCIE1_SII_APP_DEV_AND_BUS_NUM,mizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS_1,mizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL_1,mizar_PCIE1_SII_APP_AND_SLOT_CONTROL_REG,mizar_PCIE1_SII_DIAG_CTRL_BUS,mizar_PCIE1_SII_CFG_REG_RO,mizar_PCIE1_SII_CFG_ARI_FWD_EN,mizar_PCIE1_SII_RADM_SLOT_PWR_PAYLOAD,mizar_PCIE1_SII_DIAG_STATUS_BUS_0,mizar_PCIE1_SII_DIAG_STATUS_BUS_1,mizar_PCIE1_SII_DIAG_STATUS_BUS_2,mizar_PCIE1_SII_DIAG_STATUS_BUS_3,mizar_PCIE1_SII_DIAG_STATUS_BUS_4,mizar_PCIE1_SII_DIAG_STATUS_BUS_5,mizar_PCIE1_SII_DIAG_STATUS_BUS_6,mizar_PCIE1_SII_DIAG_STATUS_BUS_7,mizar_PCIE1_SII_DIAG_STATUS_BUS_8,mizar_PCIE1_SII_DIAG_STATUS_BUS_9,mizar_PCIE1_SII_DIAG_STATUS_BUS_10,mizar_PCIE1_SII_DIAG_STATUS_BUS_11,mizar_PCIE1_SII_DIAG_STATUS_BUS_12,mizar_PCIE1_SII_DIAG_STATUS_BUS_13,mizar_PCIE1_SII_DIAG_STATUS_BUS_14,mizar_PCIE1_SII_DIAG_STATUS_BUS_15,mizar_PCIE1_SII_DIAG_STATUS_BUS_16,mizar_PCIE1_SII_DIAG_STATUS_BUS_17,mizar_PCIE1_SII_DIAG_STATUS_BUS_18,mizar_PCIE1_SII_DIAG_STATUS_BUS_19,mizar_PCIE1_SII_RAM_PWR_CNTRL_0,mizar_PCIE1_SII_RAM_PWR_CNTRL_1,mizar_PCIE1_SII_SOFT_RESET_CTRL,mizar_PCIE1_SII_CFG_MSI_PENDING_B,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_1,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_2,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_3,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_4,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_5,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_6,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_7,mizar_PCIE1_SII_PHY_CONTROL_0,mizar_PCIE1_SII_PHY_CONTROL_1,mizar_PCIE1_SII_PHY_CONTROL_2,mizar_PCIE1_SII_PHY_CONTROL_3,mizar_PCIE1_SII_PHY_CONTROL_4,mizar_PCIE1_SII_PHY_CONTROL_5,mizar_PCIE1_SII_PHY_CONTROL_6,mizar_PCIE1_SII_PHY_CONTROL_7,mizar_PCIE1_SII_PHY_CONTROL_8,mizar_PCIE1_SII_PHY_CONTROL_9,mizar_PCIE1_SII_PHY_CONTROL_10,mizar_PCIE1_SII_PHY_CONTROL_11,mizar_PCIE1_SII_PHY_CONTROL_12,mizar_PCIE1_SII_PHY_CONTROL_13,mizar_PCIE1_SII_PHY_CONTROL_14,mizar_PCIE1_SII_PHY_CONTROL_15,mizar_PCIE1_SII_PHY_CONTROL_16,mizar_PCIE1_SII_PHY_CONTROL_17,mizar_PCIE1_SII_PHY_CONTROL_18,mizar_PCIE1_SII_PHY_CONTROL_19,mizar_PCIE1_SII_PHY_CONTROL_20,mizar_PCIE1_SII_PHY_CONTROL_21,mizar_PCIE1_SII_PHY_CONTROL_22,mizar_PCIE1_SII_PHY_CONTROL_23,mizar_PCIE1_SII_PHY_CONTROL_24,mizar_PCIE1_SII_PHY_CONTROL_25,mizar_PCIE1_SII_PHY_CONTROL_26,mizar_PCIE1_SII_MSI_CTRL_IO,mizar_PCIE1_SII_MSI_CTRL_INT_VEC",
    "Hidden_Validation_Acceptance_Criteria": "Default value check: For each i in [0..CNT-1], if read_mask_array[i]==0x00000000 then skip; if addr_array[i]==mizar_PCIE1_SII_PHY_RST_CONTROL then skip; else read_reg(addr_array[i]) and compare to default_value_array[i]; PASS when equal, else increment def_fail_cnt and print failure. Write/read check: For each pattern in {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}, write to each address where skip_array[i]!=1 and write_mask_array[i]!=0x00000000; then for each such address where read_mask_array[i]!=0x00000000, read back and compute wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); PASS when data_rd==exp_val, else increment wr_fail_cnt and print failure. Final result: finish(0) if def_fail_cnt==0 and wr_fail_cnt==0; otherwise finish(1).",
    "Hidden_Header_Includes": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<pcie1/pcie_sii_rc_def.h>\n#include<pcie1/pcie_sii_rc_offset.h>",
    "Hidden_Macro_Defines": "#define SOFT_RST_REG_ADDRESS\t0x00000000\n#define SOFT_RST_REG_DATA\t0x00000000\n#define MIZAR_PCIE1_SII_BASE     0xE68C1000\n#define CNT 153",
    "Hidden_Skip_Array_Definition": "const int skip_array[153]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,}"
  },
  {
    "Index": 3,
    "SS / Module": "PCIE CFG/DBI",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie_cfg_wr_rd_test",
    "Test Description": "Configure PCIe coherency control and verify link readiness. Perform configuration space programming and complete with a handshake confirmation.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Execution path depends on compile-time selection for instance and role. Link must be ready before configuration steps. Test completes when handshake value is observed.",
    "Test Steps / Procedure": "1) Program DBI_DSP_COHERENCY_CONTROL_3_OFF fields for both instances. 2) Poll the SII status until link ready. 3) Program configuration space base address and control registers. 4) Poll the handshake register until the expected signature is observed.",
    "Impacted Registers": "DBI_DSP_COHERENCY_CONTROL_3_OFF",
    "Validation / Acceptance Criteria": "1) SII status indicates link ready → Proceed to next step. 2) Handshake register matches the expected signature → Test passes.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie_cfg_wr_rd_test",
    "Hidden_Test_Description": "Test flow programs coherency and configuration registers, checks PCIe link readiness, and finalizes after a handshake. Steps: write_reg(0xE6004100,0x0); optional link training based on DM0_RC/DM1_RC/DM0_EP/DM1_EP; DBI coherency programming for both instances using mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF via set_data on bit ranges (11–14), (3–6), (27–30), (19–22); wait_on(20); consolidated writes to the same register fields; read SII0 status at offset 0xC0; call non_secure_prot_nic(); poll while ((data_rd & 0xD1) != 0xD1); if DM1_RC, poll SII1 status similarly; write_reg(0xE6004100,0x11111111); wait_on(15000); if DM0_RC: mem_base_program_dm0_x4(); wait_on(10); read_pcie_slv0_reg for i=0..9; write_pcie_slv0_reg BAR offsets 0x10,0x14,0x18,0x1c,0x20,0x24 with 0xFFFFFFFF; read them back; write the same BARs with 0x0,0x4,0x20000000,0x40000000,0x60000000,0x80000000; read back; enable command by write_pcie_slv0_reg(0x4,0x7). If DM1_RC: mem_base_program_dm1_x4(); read_pcie_slv1_reg for i=0..9; enable command write_pcie_slv1_reg(0x4,0x7); write/read BARs 0x10..0x24 with 0xFFFFFFFF then programmed values; wait_on(10); poll read_reg(0xE6004100) until it equals 0x12345678; finish(0).",
    "Hidden_Remarks": "Compile-time flags (DM0_RC, DM1_RC, DM0_EP, DM1_EP) control which instance and role execute. Link readiness is required before configuration transactions, checked via SII status mask 0xD1 at offset 0xC0. A memory-mapped handshake register at 0xE6004100 must reach 0x12345678 before finishing.",
    "Hidden_Test_Steps_Procedure": "1) write_reg(0xE6004100,0x0). 2) If DM0_RC then link_training_dm0_x4(4); if DM1_RC then link_training_dm1_x4(4); if DM0_EP then link_training_dm0_x4(4); if DM1_EP then link_training_dm1_x4(4). 3) Coherency programming for PCIe0: rd_wr_data1 = set_data(read_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF),11,14,0xf); rd_wr_data1 = set_data(rd_wr_data1,3,6,0xf); write_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF,rd_wr_data1); then rd_wr_data1 = set_data(read_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF),27,30,0xf); rd_wr_data1 = set_data(rd_wr_data1,19,22,0xf); write_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF,rd_wr_data1). 4) Coherency programming for PCIe1: same sequence using mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF. 5) wait_on(20). 6) Consolidated coherency writes: For PCIe0 and PCIe1, set bits (11–14), (3–6), (27–30), (19–22) cumulatively and write back to the respective DBI coherency control register. 7) data_rd = read_sii0_reg(0xC0); call non_secure_prot_nic(); while(((data_rd)&(0xD1))!=0xD1){ data_rd = read_sii0_reg(0xC0); }. 8) If DM1_RC: data_rd = read_sii1_reg(0xC0); while(((data_rd)&(0xD1))!=0xD1){ data_rd = read_sii1_reg(0xC0); }. 9) write_reg(0xE6004100,0x11111111); wait_on(15000). 10) If DM0_RC: mem_base_program_dm0_x4(); wait_on(10); for(i=0;i<10;i++){ rd_wr_data1 = read_pcie_slv0_reg(i*0x4); } write_pcie_slv0_reg(0x10,0xFFFFFFFF); write_pcie_slv0_reg(0x14,0xFFFFFFFF); write_pcie_slv0_reg(0x18,0xFFFFFFFF); write_pcie_slv0_reg(0x1c,0xFFFFFFFF); write_pcie_slv0_reg(0x20,0xFFFFFFFF); write_pcie_slv0_reg(0x24,0xFFFFFFFF); read_pcie_slv0_reg(0x10); read_pcie_slv0_reg(0x14); read_pcie_slv0_reg(0x18); read_pcie_slv0_reg(0x1c); read_pcie_slv0_reg(0x20); read_pcie_slv0_reg(0x24); write_pcie_slv0_reg(0x10,0x0); write_pcie_slv0_reg(0x14,0x4); write_pcie_slv0_reg(0x18,0x20000000); write_pcie_slv0_reg(0x1c,0x40000000); write_pcie_slv0_reg(0x20,0x60000000); write_pcie_slv0_reg(0x24,0x80000000); read_pcie_slv0_reg(0x10); read_pcie_slv0_reg(0x14); read_pcie_slv0_reg(0x18); read_pcie_slv0_reg(0x1c); read_pcie_slv0_reg(0x20); read_pcie_slv0_reg(0x24); write_pcie_slv0_reg(0x4,0x7). 11) If DM1_RC: mem_base_program_dm1_x4(); for(i=0;i<10;i++){ rd_wr_data1 = read_pcie_slv1_reg(i*0x4); } write_pcie_slv1_reg(0x4,0x7); write_pcie_slv1_reg(0x10,0xFFFFFFFF); write_pcie_slv1_reg(0x14,0xFFFFFFFF); write_pcie_slv1_reg(0x18,0xFFFFFFFF); write_pcie_slv1_reg(0x1c,0xFFFFFFFF); write_pcie_slv1_reg(0x20,0xFFFFFFFF); write_pcie_slv1_reg(0x24,0xFFFFFFFF); read_pcie_slv1_reg(0x10); read_pcie_slv1_reg(0x14); read_pcie_slv1_reg(0x18); read_pcie_slv1_reg(0x1c); read_pcie_slv1_reg(0x20); read_pcie_slv1_reg(0x24); write_pcie_slv1_reg(0x10,0x0); write_pcie_slv1_reg(0x14,0x4); write_pcie_slv1_reg(0x18,0x20000000); write_pcie_slv1_reg(0x1c,0x40000000); write_pcie_slv1_reg(0x20,0x60000000); write_pcie_slv1_reg(0x24,0x80000000); read_pcie_slv1_reg(0x10); read_pcie_slv1_reg(0x14); read_pcie_slv1_reg(0x18); read_pcie_slv1_reg(0x1c); read_pcie_slv1_reg(0x20); read_pcie_slv1_reg(0x24). 12) wait_on(10); data_rd = read_reg(0xE6004100); while(data_rd != 0x12345678){ wait_on(5); data_rd = read_reg(0xE6004100); } finish(0).",
    "Hidden_Impacted_Registers": "mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF,mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF",
    "Hidden_Validation_Acceptance_Criteria": "Link readiness: while(((data_rd)&(0xD1))!=0xD1) loops on SII status (0xC0) until ((value & 0xD1) == 0xD1), then proceed. Handshake: poll read_reg(0xE6004100) until it equals 0x12345678; only then finish(0).",
    "Hidden_Header_Includes": "#include <stdlib.h>\n#include <stdio.h>\n#include <test_common.h>\n#include \"pcie.h\"",
    "Hidden_Macro_Defines": "NA",
    "Hidden_Skip_Array_Definition": "NA"
  }
]

MAIN_COLS = [
  "Index",
  "SS / Module",
  "Feature",
  "Test Case Name",
  "Test Description",
  "Speed",
  "Mode",
  "Memory Start Offset",
  "Memory End Offset",
  "Remarks",
  "Test Steps / Procedure",
  "Impacted Registers",
  "Validation / Acceptance Criteria",
  "Code Generation (Required / Not)",
]

META_COLS_ORDER = [
  "Hidden_Test_Case_Name",
  "Hidden_Test_Description",
  "Hidden_Remarks",
  "Hidden_Test_Steps_Procedure",
  "Hidden_Impacted_Registers",
  "Hidden_Validation_Acceptance_Criteria",
  "Hidden_Header_Includes",
  "Hidden_Macro_Defines",
  "Hidden_Skip_Array_Definition",
]

BLUE_FILL = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
BORDER_THIN = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_TOP_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=False)
CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=False)


def ordered_union_keys(records):
    seen = []
    sset = set()
    for obj in records:
        for k in obj.keys():
            if k not in sset:
                sset.add(k)
                seen.append(k)
    return seen


def ensure_numbering(text: str) -> str:
    if not text:
        return ""
    # Split on existing newlines; trim whitespace; drop empties
    parts = [p.strip() for p in str(text).replace("\r", "").split("\n")]
    parts = [p for p in parts if p]
    # Renumber strictly with 1., 2., 3.
    numbered = []
    for i, p in enumerate(parts, start=1):
        # Remove any existing numeric prefixes like '1)', '1.' at the start
        cleaned = p
        # Simple heuristic: strip leading digits and punctuation
        while cleaned and (cleaned[0].isdigit() or cleaned[0] in [')', '.', '-', ':']):
            cleaned = cleaned[1:].lstrip()
        numbered.append(f"{i}. {cleaned}")
    return "\n".join(numbered)


def auto_fit_columns(ws):
    from openpyxl.utils import get_column_letter
    max_width = {}
    for r in ws.iter_rows(values_only=True):
        for idx, cell in enumerate(r, start=1):
            val = "" if cell is None else str(cell)
            w = len(val)
            if w > max_width.get(idx, 0):
                max_width[idx] = w
    for idx, width in max_width.items():
        col = get_column_letter(idx)
        # cap width to avoid overly wide columns
        ws.column_dimensions[col].width = min(max(10, width + 2), 120)


def apply_borders(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value) != "":
                cell.border = BORDER_THIN


def build_workbook(data_records, output_path):
    if not isinstance(data_records, list) or len(data_records) == 0:
        raise ValueError("JSON input must be a non-empty array of objects")

    # STEP 2 — Normalize Tabular Schema
    headers = ordered_union_keys(data_records)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # STEP 3 — Base sheet write
    ws.append(headers)
    for obj in data_records:
        row = [obj.get(k, "") for k in headers]
        ws.append(row)

    # Base formatting
    for c in ws[1]:
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.fill = BLUE_FILL
    ws.freeze_panes = "A2"
    auto_fit_columns(ws)

    # STEP 5 — Create META sheet
    meta_ws = wb.create_sheet("Meta_data_sheet")
    present_meta_cols = [c for c in META_COLS_ORDER if c in headers]
    meta_ws.append(present_meta_cols)
    for obj in data_records:
        meta_ws.append([obj.get(c, "") for c in present_meta_cols])
    # Very hidden
    meta_ws.sheet_state = "veryHidden"

    # STEP 7 — Normalize MAIN sheet in-place
    # Rename Data -> TestPlan
    ws.title = "TestPlan"

    # Build reordered (and reduced) data for TestPlan sheet
    # Remove any META and non-main columns: keep only MAIN_COLS in given order
    keep_cols = [c for c in MAIN_COLS]

    # Rebuild the sheet content on the same worksheet
    ws.delete_rows(1, ws.max_row)
    ws.append(keep_cols)
    for obj in data_records:
        row_obj = {}
        for k in keep_cols:
            row_obj[k] = obj.get(k, "")
        # Mandatory numbering inside cells for specific columns
        for k in ["Test Steps / Procedure", "Validation / Acceptance Criteria"]:
            if k in row_obj and isinstance(row_obj[k], str):
                row_obj[k] = ensure_numbering(row_obj[k])
        ws.append([row_obj.get(k, "") for k in keep_cols])

    # STEP 7A — FORMAT MAIN SHEET
    # Header row
    for c in ws[1]:
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.fill = BLUE_FILL

    # Data rows formatting
    # Wrap text for specified columns
    wrap_cols = set(["Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"])
    col_index = {name: idx+1 for idx, name in enumerate(keep_cols)}

    for r in range(2, ws.max_row + 1):
        for name, idx in col_index.items():
            cell = ws.cell(row=r, column=idx)
            if name in wrap_cols:
                cell.alignment = LEFT_TOP_WRAP
            elif name == "Index":
                cell.alignment = CENTER_TOP
            else:
                cell.alignment = LEFT_TOP

    auto_fit_columns(ws)
    apply_borders(ws)

    # STEP 7A — Data Validation ONLY for Code Generation (Required / Not)
    if "Code Generation (Required / Not)" in col_index:
        start_row = 2
        end_row = ws.max_row
        col_letter = chr(ord('A') + col_index["Code Generation (Required / Not)"] - 1)
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
        dv.error = "Select a value from the list"
        dv.errorTitle = "Invalid Selection"
        dv.prompt = "Choose: Required, Blank, or Not Required"
        dv.promptTitle = "Code Generation (Required / Not)"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

    # STEP 7B — Enforce final sheet visibility: ensure no 'Data' sheet exists
    for s in list(wb.sheetnames):
        if s == "Data":
            del wb[s]

    # STEP 8 — Save Final Excel File
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    # STEP 11 — Validate OOXML ZIP structure
    with ZipFile(output_path, 'r') as zf:
        members = zf.namelist()
        if '[Content_Types].xml' not in members or 'xl/workbook.xml' not in members:
            raise RuntimeError('XLSX validation failed: missing core OOXML parts')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--output-dir', required=True)
    parser.add_argument('--output-name', required=True)
    args = parser.parse_args()

    # PHASE 1 — STEP 1: Validate JSON Input
    if not isinstance(JSON_DATA, list) or len(JSON_DATA) == 0:
        raise SystemExit('FAIL: json_data must be a non-empty array')

    # PHASE 1 — STEP 2..3 and PHASE 2 and PHASE 3 within build_workbook
    output_path = os.path.join(args.output_dir, args.output_name)
    build_workbook(JSON_DATA, output_path)

    print(f"WROTE: {output_path}")

if __name__ == '__main__':
    main()
