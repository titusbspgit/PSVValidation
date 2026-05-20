#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate a REAL .xlsx Test Plan workbook from embedded JSON.
- Sheet "TestPlan" (visible) with required columns
- Sheet "MetaData" (VERY HIDDEN) with meta columns
- Header row bold, first row frozen
- File name: testplan_<timestamp>.xlsx where timestamp is IST (YYYYMMDD_HHMMSS)

Usage:
  python tools/generate_testplan_xlsx.py --output-dir Test_Output
"""
import argparse
import json
import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

# Embedded aggregated JSON (verbatim). Do not modify.
JSON_TEXT = r'''[
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "GPIO Register Reset and R/W Verification",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Validate that GPIO_8 to GPIO_27 registers power-on/reset to their documented default values (ignoring bit 0) and that writable fields accept data patterns while read-only fields retain their default values. Verify read/write behavior using write and read masks with multiple data patterns.",
    "Meta Test Description": "The test performs two phases: (1) Default value verification and (2) Write/Read verification. Phase 1 iterates i=0..CNT-1 over addr_array[i] and, if skip_rst_array[i]==1 or read_mask_array[i]==0, skips. Otherwise reads data_rd=read_reg(addr), masks out bit0 via data=(data_rd & 0xfffffffe), and compares against default_value_array[i]; increments def_fail_cnt on mismatch. Phase 2 iterates over six test patterns {0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}. For each pattern, it writes (data_wr & write_mask_array[i]) to each address where skip_array[i]==0 and write_mask_array[i]!=0. Then reads back data_rd=(read_reg(addr) & read_mask_array[i]) for each address where read_mask_array[i]!=0 and write_mask_array[i]!=0. It computes wr_n=(write_mask_array[i] ^ 0xffffffff) and expected value exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])). If data_rd != exp_val, increments wr_fail_cnt. At end, the test calls finish(1) if (def_fail_cnt>0 || wr_fail_cnt>0) else finish(0). A soft reset check function is present but compiled out (#ifdef 0).",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Bit 0 is intentionally ignored during reset-value comparison for all targeted GPIO registers. Only GPIO_8 through GPIO_27 are targeted in this test iteration. Registers marked non-readable or non-writable by masks are skipped accordingly. The soft reset check is disabled in the current test build.",
    "Test Steps / Procedure": "1. For each GPIO register from GPIO_8 to GPIO_27 that is readable and not marked as skipped for reset verification, read the register and compare its value (with bit 0 ignored) against its documented reset default value. 2. For each of the defined test data patterns, write the pattern to each GPIO register from GPIO_8 to GPIO_27 that is writable and not marked as skipped, applying the write mask to limit writes to writable bits. 3. Read back each affected GPIO register that is readable and writable, applying the read mask to the returned value. 4. For each readback, confirm that writable bits match the written pattern and that non-writable (read-only) bits remain at their reset default values. 5. Record any mismatches as failures and report overall test status as PASS only if no default-value or write/read mismatches are observed.",
    "Meta Test Steps / Procedure": "Initialization: def_fail_cnt=0; wr_fail_cnt=0. Phase 1 (chk_rst_val): for (i=0;i<CNT;i++){ addr=addr_array[i]; if (skip_rst_array[i]==1) continue; if (read_mask_array[i]==0x00000000) continue; data_rd=read_reg(addr); data=(data_rd & 0xfffffffe); if (data==default_value_array[i]) PASS else {def_fail_cnt++; log fail with addr, expected, read, and raw data_rd}; }. Phase 2 (chk_rd_wr): unsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}; for (each data_wr in chk_val){ // write pass: for (i=0;i<CNT;i++){ addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0x00000000) continue; write_reg(addr, (data_wr & write_mask_array[i])); } // read/verify pass: for (i=0;i,CNT;i++){ addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0x00000000) continue; if (read_mask_array[i]==0x00000000) continue; data_rd=(read_reg(addr) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if (data_rd==exp_val) PASS else {wr_fail_cnt++; log fail addr/expected/read}; } }. Completion: if (def_fail_cnt>0 || wr_fail_cnt>0) finish(1); else finish(0). Note: soft_reset_chk() code is disabled by #ifdef 0.",
    "Impacted Registers": "GPIO_8, GPIO_9, GPIO_10, GPIO_11, GPIO_12, GPIO_13, GPIO_14, GPIO_15, GPIO_16, GPIO_17, GPIO_18, GPIO_19, GPIO_20, GPIO_21, GPIO_22, GPIO_23, GPIO_24, GPIO_25, GPIO_26, GPIO_27",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, MIZAR_GPIO_GP0_GPIO_11, MIZAR_GPIO_GP0_GPIO_12, MIZAR_GPIO_GP0_GPIO_13, MIZAR_GPIO_GP0_GPIO_14, MIZAR_GPIO_GP0_GPIO_15, MIZAR_GPIO_GP0_GPIO_16, MIZAR_GPIO_GP0_GPIO_17, MIZAR_GPIO_GP0_GPIO_18, MIZAR_GPIO_GP0_GPIO_19, MIZAR_GPIO_GP0_GPIO_20, MIZAR_GPIO_GP0_GPIO_21, MIZAR_GPIO_GP0_GPIO_22, MIZAR_GPIO_GP0_GPIO_23, MIZAR_GPIO_GP0_GPIO_24, MIZAR_GPIO_GP0_GPIO_25, MIZAR_GPIO_GP0_GPIO_26, MIZAR_GPIO_GP0_GPIO_27, GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL, GPIO_GP0_GPIO_11_DEFAULT_VAL, GPIO_GP0_GPIO_12_DEFAULT_VAL, GPIO_GP0_GPIO_13_DEFAULT_VAL, GPIO_GP0_GPIO_14_DEFAULT_VAL, GPIO_GP0_GPIO_15_DEFAULT_VAL, GPIO_GP0_GPIO_16_DEFAULT_VAL, GPIO_GP0_GPIO_17_DEFAULT_VAL, GPIO_GP0_GPIO_18_DEFAULT_VAL, GPIO_GP0_GPIO_19_DEFAULT_VAL, GPIO_GP0_GPIO_20_DEFAULT_VAL, GPIO_GP0_GPIO_21_DEFAULT_VAL, GPIO_GP0_GPIO_22_DEFAULT_VAL, GPIO_GP0_GPIO_23_DEFAULT_VAL, GPIO_GP0_GPIO_24_DEFAULT_VAL, GPIO_GP0_GPIO_25_DEFAULT_VAL, GPIO_GP0_GPIO_26_DEFAULT_VAL, GPIO_GP0_GPIO_27_DEFAULT_VAL, GPIO_GP0_GPIO_8_READ_MASK, GPIO_GP0_GPIO_9_READ_MASK, GPIO_GP0_GPIO_10_READ_MASK, GPIO_GP0_GPIO_11_READ_MASK, GPIO_GP0_GPIO_12_READ_MASK, GPIO_GP0_GPIO_13_READ_MASK, GPIO_GP0_GPIO_14_READ_MASK, GPIO_GP0_GPIO_15_READ_MASK, GPIO_GP0_GPIO_16_READ_MASK, GPIO_GP0_GPIO_17_READ_MASK, GPIO_GP0_GPIO_18_READ_MASK, GPIO_GP0_GPIO_19_READ_MASK, GPIO_GP0_GPIO_20_READ_MASK, GPIO_GP0_GPIO_21_READ_MASK, GPIO_GP0_GPIO_22_READ_MASK, GPIO_GP0_GPIO_23_READ_MASK, GPIO_GP0_GPIO_24_READ_MASK, GPIO_GP0_GPIO_25_READ_MASK, GPIO_GP0_GPIO_26_READ_MASK, GPIO_GP0_GPIO_27_READ_MASK, GPIO_GP0_GPIO_8_WRITE_MASK, GPIO_GP0_GPIO_9_WRITE_MASK, GPIO_GP0_GPIO_10_WRITE_MASK, GPIO_GP0_GPIO_11_WRITE_MASK, GPIO_GP0_GPIO_12_WRITE_MASK, GPIO_GP0_GPIO_13_WRITE_MASK, GPIO_GP0_GPIO_14_WRITE_MASK, GPIO_GP0_GPIO_15_WRITE_MASK, GPIO_GP0_GPIO_16_WRITE_MASK, GPIO_GP0_GPIO_17_WRITE_MASK, GPIO_GP0_GPIO_18_WRITE_MASK, GPIO_GP0_GPIO_19_WRITE_MASK, GPIO_GP0_GPIO_20_WRITE_MASK, GPIO_GP0_GPIO_21_WRITE_MASK, GPIO_GP0_GPIO_22_WRITE_MASK, GPIO_GP0_GPIO_23_WRITE_MASK, GPIO_GP0_GPIO_24_WRITE_MASK, GPIO_GP0_GPIO_25_WRITE_MASK, GPIO_GP0_GPIO_26_WRITE_MASK, GPIO_GP0_GPIO_27_WRITE_MASK",
    "Validation / Acceptance Criteria": "PASS if: (a) For all readable GPIO_8..GPIO_27 registers not marked as skipped in reset checks, the read value with bit 0 ignored matches the documented reset default; and (b) For each data pattern, all readable and writable GPIO_8..GPIO_27 registers return readback values where writable bits match the pattern and non-writable bits remain at their reset defaults. FAIL otherwise.",
    "Meta Validation / Acceptance Criteria": "Default check: For each i, if (skip_rst_array[i]==0 && read_mask_array[i]!=0) then data_rd=read_reg(addr_array[i]); data=(data_rd & 0xfffffffe); Expect data == default_value_array[i]; else def_fail_cnt++. Write/Read check: For each pattern data_wr in {0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}, for each i with skip_array[i]==0 and write_mask_array[i]!=0 perform write_reg(addr_array[i], (data_wr & write_mask_array[i])); then if (read_mask_array[i]!=0) read data_rd=(read_reg(addr_array[i]) & read_mask_array[i]); compute wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); Expect data_rd==exp_val; else wr_fail_cnt++. Final: finish(0) only if (def_fail_cnt==0 && wr_fail_cnt==0) else finish(1).",
    "Code Generation (Required / Not)": "NA",
    "Meta Headers": "#include <stdio.h>, #include <stdlib.h>, #include \"test_common.h\", #include \"test_define.c\", #include <gpio/gpio_def.h>, #include <gpio/gpio_offset.h>",
    "Meta Macros": "#define CNT 49; #define SOFT_RST_REG_ADDRESS 0x00000000; #define SOFT_RST_REG_DATA 0x00000000",
    "Meta Arrays": "addr_array[20] = {MIZAR_GPIO_GP0_GPIO_8..MIZAR_GPIO_GP0_GPIO_27}; default_value_array[20] = {GPIO_GP0_GPIO_8_DEFAULT_VAL..GPIO_GP0_GPIO_27_DEFAULT_VAL}; read_mask_array[20] = {GPIO_GP0_GPIO_8_READ_MASK..GPIO_GP0_GPIO_27_READ_MASK}; write_mask_array[20] = {GPIO_GP0_GPIO_8_WRITE_MASK..GPIO_GP0_GPIO_27_WRITE_MASK}; skip_array[20] = all zeros; skip_rst_array[20] = all zeros",
    "ip_name": "GPIO",
    "test_case_id": "gpio_reg_wr_rd_test",
    "test_case_name": "gpio_reg_wr_rd_test",
    "folder_path": "TestRepo/gpio/gpio_reg_wr_rd_test",
    "github_folder_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/gpio_reg_wr_rd_test",
    "repo": "titusbspgit/PSVValidation",
    "branch": "main"
  },
  {
    "ip_name": "GPIO",
    "test_case_id": "test_gpio_negedge_intr_en",
    "test_case_name": "test_gpio_negedge_intr_en",
    "folder_path": "TestRepo/gpio/test_gpio_negedge_intr_en",
    "github_folder_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en",
    "repo": "titusbspgit/PSVValidation",
    "branch": "main"
  },
  {
    "Index": "3",
    "SS / Module": "GPIO",
    "Feature": "GPIO Positive-Edge Interrupt (all pads)",
    "Test Case Name": "test_gpio_pedge_all_pads_en",
    "Test Description": "Enable positive-edge interrupts for GPIO pads 8–39, configure GPIO I/O control groups for proper direction, enable group interrupt, then for each pad generate a single rising edge using an external stimulus at 0xA0243FFC, verify interrupt reception via group status, clear per-pin raw status for all pads, confirm group status clears, and clear SoC raw interrupt status.",
    "Meta Test Description": "The test enables the relevant GIC IRQ (IRQ 87 when built for GPIO0 or IRQ 88 when built for GPIO1) and sets the SoC-level interrupt enable in LSS_SYSREG_INTR_EN1 for the selected GPIO instance. It programs GPIO_8..GPIO_39 with 0x00020000 (bit17=1) to enable positive-edge detection per pad. After a short wait, it sets GPIO_IO_CTRL_GROUP1..GPIO_IO_CTRL_GROUP4 to 0x000000FF (direction configuration per design comment), waits, and then enables all group interrupt masks by writing 0xFFFFFFFF to INTR1_INTR_EN1. For i=0..31, it drives the external stimulus at 0xA0243FFC low, arms int_pend=1, then drives it high to create a single rising edge. It polls with a timeout (2000 iterations with wait_on(10)) for int_pend to be cleared by the ISR. On timeout, it logs an error, increments test_err, and breaks. After each iteration it optionally drives the stimulus low again and waits. The Default_IRQHandler sets int_pend=0, reads INTR1_INTR_STS1, and masks group output (INTR1_INTR_EN1=0) during service. If any bit in INTR1_INTR_STS1 is set, it proceeds; otherwise logs an error and increments test_err. It clears per-pin raw interrupt status by writing 0x00010000 to each pad register GPIO_8 + (j*4) for j=0..31, waits, and verifies INTR1_INTR_STS1 reads 0x0; otherwise logs error and increments test_err. It clears the SoC raw status in LSS_SYSREG_RAW_STCR1 for the selected GPIO instance and verifies the corresponding bit is cleared; otherwise logs error and increments test_err. Finally, it re-enables INTR1_INTR_EN1 (0xFFFFFFFF) and clears the GIC IRQ (87 or 88). The test completes by calling finish(test_err).",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "0xA0243FFC",
    "Memory End Offset": "0xA0243FFC",
    "Remarks": "Ensure LSS_SYSREG_INTR_EN1 is configured for the target GPIO instance and the corresponding GIC IRQ line (87 or 88) is enabled. Configure GPIO_IO_CTRL_GROUP1..GPIO_IO_CTRL_GROUP4 as required before enabling interrupts. Group interrupt output is masked during ISR service and must be re-enabled afterward. SoC-level raw status in LSS_SYSREG_RAW_STCR1 must be cleared to deassert the interrupt. External stimulus at 0xA0243FFC must be accessible to generate rising edges on pads 8–39.",
    "Test Steps / Procedure": "1. Enable the SoC-level interrupt for the selected GPIO instance in LSS_SYSREG_INTR_EN1 and enable the corresponding GIC IRQ line (87 or 88). 2. Configure GPIO_8..GPIO_39 for positive-edge interrupt detection. 3. Configure GPIO_IO_CTRL_GROUP1..GPIO_IO_CTRL_GROUP4 for the intended I/O direction. 4. Enable the group interrupt mask by writing all ones to INTR1_INTR_EN1. 5. For each pad in 8–39: a) Drive the external stimulus at 0xA0243FFC low, arm the interrupt pending flag, then drive it high to create a rising edge. b) Wait for the ISR to complete within the configured timeout. 6. In the ISR: verify that INTR1_INTR_STS1 indicates an active interrupt, temporarily mask the group output, clear per-pin raw status for all pads, confirm INTR1_INTR_STS1 clears to zero, clear the SoC raw status in LSS_SYSREG_RAW_STCR1, then re-enable INTR1_INTR_EN1 and clear the GIC IRQ. 7. Report PASS if all pads generate and service interrupts within timeout and all status clears succeed; otherwise FAIL.",
    "Meta Test Steps / Procedure": "Initialization: test_err=0. If GPIO0 build: GIC_EnableIRQ(87); write_reg(LSS_SYSREG_INTR_EN1, (GPIO0 interrupt enable bit via build define)); If GPIO1 build: GIC_EnableIRQ(88); write_reg(LSS_SYSREG_INTR_EN1, (GPIO1 interrupt enable bit via build define)). Configure pads: for (i=0..31) write_reg(GPIO_8 + i*4, 0x00020000); wait_on(10). Configure I/O: write_reg(GPIO_IO_CTRL_GROUP1, 0x000000FF); write_reg(GPIO_IO_CTRL_GROUP2, 0x000000FF); write_reg(GPIO_IO_CTRL_GROUP3, 0x000000FF); write_reg(GPIO_IO_CTRL_GROUP4, 0x000000FF); wait_on(10). Enable group masks: write_reg(INTR1_INTR_EN1, 0xFFFFFFFF). For each i=0..31: write_reg(0xA0243FFC, 0x00000000); wait_on(10); int_pend=1; write_reg(0xA0243FFC, 0xFFFFFFFF); timeout=2000; while ((int_pend==1) && (--timeout>0)) wait_on(10); if (timeout==0) { printf timeout; test_err++; break; } write_reg(0xA0243FFC, 0x00000000); wait_on(10). ISR (Default_IRQHandler): wr_val=(1<<i); int_pend=0; rdata_grp=read_reg(INTR1_INTR_STS1); write_reg(INTR1_INTR_EN1, 0x00000000); if ((rdata_grp & 0xFFFFFFFF)==0) { printf error; test_err++; } for (j=0..31) write_reg(GPIO_8 + j*4, 0x00010000); wait_on(2); rdata_grp=read_reg(INTR1_INTR_STS1); if (rdata_grp!=0x0) { printf error; test_err++; } If GPIO0: write_reg(LSS_SYSREG_RAW_STCR1, (GPIO0 bit via build define)); rdata=read_reg(LSS_SYSREG_RAW_STCR1); if ((rdata & (GPIO0 bit))!=0) { printf error; test_err++; } If GPIO1: write_reg(LSS_SYSREG_RAW_STCR1, (GPIO1 bit via build define)); rdata=read_reg(LSS_SYSREG_RAW_STCR1); if ((rdata & (GPIO1 bit))!=0) { printf error; test_err++; } write_reg(INTR1_INTR_EN1, 0xFFFFFFFF); If GPIO0: GIC_ClearIRQ(87); If GPIO1: GIC_ClearIRQ(88). Completion: finish(test_err).",
    "Impacted Registers": "GPIO_8..GPIO_39, INTR1_INTR_EN1, INTR1_INTR_STS1, GPIO_IO_CTRL_GROUP1, GPIO_IO_CTRL_GROUP2, GPIO_IO_CTRL_GROUP3, GPIO_IO_CTRL_GROUP4, LSS_SYSREG_INTR_EN1, LSS_SYSREG_RAW_STCR1",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8..MIZAR_GPIO_GP0_GPIO_39, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR, LSS_SYSREG_INTR_EN1_GPIO1_INTR, MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR, LSS_SYSREG_RAW_STCR1_GPIO1_INTR",
    "Validation / Acceptance Criteria": "PASS if each pad 8–39 triggers a positive-edge interrupt within the timeout, INTR1_INTR_STS1 indicates an active interrupt, per-pin raw status clearing results in INTR1_INTR_STS1 reading zero, and the SoC raw interrupt status in LSS_SYSREG_RAW_STCR1 is cleared. Any timeout or status-clear failure is FAIL.",
    "Meta Validation / Acceptance Criteria": "Timeout: While loop must exit by ISR clearing int_pend before timeout reaches zero; else record error. Group status: After rising edge, (read(INTR1_INTR_STS1) & 0xFFFFFFFF)!=0; else error. Clear per-pin: After writing 0x00010000 to all GPIO_8..GPIO_39, read(INTR1_INTR_STS1)==0x0; else error. SoC status: After write to LSS_SYSREG_RAW_STCR1 for the active GPIO instance, (read(LSS_SYSREG_RAW_STCR1) & instance_bit)==0; else error. Final result: finish(0) only when test_err==0.",
    "Code Generation (Required / Not)": "NA",
    "Meta Headers": "#include <lss_sysreg.h>, #include <stdio.h>, #include <test_define.c>, #include <test_common.h>",
    "Meta Macros": "#define CNT 49",
    "Meta Arrays": "addr_array[20]={MIZAR_GPIO_GP0_GPIO_8..MIZAR_GPIO_GP0_GPIO_27}; default_value_array[20]={GPIO_GP0_GPIO_8_DEFAULT_VAL..GPIO_GP0_GPIO_27_DEFAULT_VAL}; read_mask_array[20]={GPIO_GP0_GPIO_8_READ_MASK..GPIO_GP0_GPIO_27_READ_MASK}; write_mask_array[20]={GPIO_GP0_GPIO_8_WRITE_MASK..GPIO_GP0_GPIO_27_WRITE_MASK}; skip_array[20]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0}",
    "ip_name": "GPIO",
    "test_case_id": "test_gpio_pedge_all_pads_en",
    "test_case_name": "test_gpio_pedge_all_pads_en",
    "folder_path": "TestRepo/gpio/test_gpio_pedge_all_pads_en",
    "github_folder_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en",
    "repo": "titusbspgit/PSVValidation",
    "branch": "main"
  }
]'''

TESTPLAN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_COLUMNS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]


def build_workbook(rows, output_dir):
    # Compute IST timestamp for filename
    now_utc = datetime.now(timezone.utc)
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = now_utc.astimezone(ist)
    ts = now_ist.strftime('%Y%m%d_%H%M%S')

    filename = f"testplan_{ts}.xlsx"
    out_path = os.path.join(output_dir, filename)

    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # TestPlan sheet (visible)
    ws_plan = wb.create_sheet("TestPlan")
    bold_font = Font(bold=True)

    # Header
    for c, key in enumerate(TESTPLAN_COLUMNS, start=1):
        cell = ws_plan.cell(row=1, column=c, value=key)
        cell.font = bold_font
    ws_plan.freeze_panes = 'A2'

    # Rows
    for r_idx, obj in enumerate(rows, start=2):
        for c, key in enumerate(TESTPLAN_COLUMNS, start=1):
            val = obj.get(key, "")
            ws_plan.cell(row=r_idx, column=c, value=val)

    # MetaData sheet (VERY HIDDEN)
    ws_meta = wb.create_sheet("MetaData")
    for c, key in enumerate(METADATA_COLUMNS, start=1):
        cell = ws_meta.cell(row=1, column=c, value=key)
        cell.font = bold_font
    ws_meta.freeze_panes = 'A2'

    for r_idx, obj in enumerate(rows, start=2):
        for c, key in enumerate(METADATA_COLUMNS, start=1):
            val = obj.get(key, "")
            ws_meta.cell(row=r_idx, column=c, value=val)

    # VERY HIDDEN state
    ws_meta.sheet_state = 'veryHidden'

    # Ensure output dir exists
    os.makedirs(output_dir, exist_ok=True)
    wb.save(out_path)

    return out_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--output-dir', required=True, help='Directory to write the Excel file into (will be created if missing)')
    args = parser.parse_args()

    try:
        data = json.loads(JSON_TEXT)
        if not isinstance(data, list):
            raise ValueError('json_data must be a JSON array')
    except Exception as e:
        raise SystemExit(f"Invalid json_data: {e}")

    out_path = build_workbook(data, args.output_dir)
    print(f"Wrote Excel: {out_path}")


if __name__ == '__main__':
    main()
