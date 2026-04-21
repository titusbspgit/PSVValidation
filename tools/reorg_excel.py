#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_COLS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

WRAP_COLS = {
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
}


def get_header_map(ws):
    header_map = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str):
            header_map[v] = col
        else:
            # Non-string headers: preserve by stringifying consistently
            header_map[str(v) if v is not None else ''] = col
    return header_map


def copy_meta_sheet(src_ws, wb):
    # Remove existing Meta_data_sheet if present to keep deterministic output
    if 'Meta_data_sheet' in wb.sheetnames:
        del wb['Meta_data_sheet']
    meta_ws = wb.create_sheet('Meta_data_sheet')
    # Write headers in the specified order
    for c_idx, h in enumerate(META_COLS, start=1):
        meta_ws.cell(row=1, column=c_idx, value=h)

    hmap = get_header_map(src_ws)
    # Copy rows as-is, mapping by header names
    for r in range(2, src_ws.max_row + 1):
        out_row = []
        for h in META_COLS:
            col = hmap.get(h)
            val = src_ws.cell(row=r, column=col).value if col else None
            out_row.append(val)
        for c_idx, val in enumerate(out_row, start=1):
            meta_ws.cell(row=r, column=c_idx, value=val)

    # Very hide this sheet
    meta_ws.sheet_state = 'veryHidden'


def build_testplan_sheet(src_ws, wb):
    # Rename source to TestPlan first (if already exists, temporarily rename old one)
    if src_ws.title != 'TestPlan':
        src_ws.title = 'TestPlan'

    # Create a temp sheet to enforce exact column order and removal of META columns
    tmp_name = 'TestPlan_tmp'
    if tmp_name in wb.sheetnames:
        del wb[tmp_name]
    tmp = wb.create_sheet(tmp_name)

    # Headers in specified order
    for c_idx, h in enumerate(MAIN_COLS, start=1):
        tmp.cell(row=1, column=c_idx, value=h)

    hmap = get_header_map(src_ws)

    for r in range(2, src_ws.max_row + 1):
        row_vals = []
        for h in MAIN_COLS:
            col = hmap.get(h)
            val = src_ws.cell(row=r, column=col).value if col else None
            row_vals.append(val)
        for c_idx, val in enumerate(row_vals, start=1):
            tmp.cell(row=r, column=c_idx, value=val)

    # Remove original TestPlan and rename tmp
    del wb['TestPlan']
    tmp.title = 'TestPlan'
    return wb['TestPlan']


def format_testplan(ws):
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    # Header formatting
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Data formatting
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            header = ws.cell(row=1, column=c).value
            # Alignment rules
            if header == 'Index':
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=(header in WRAP_COLS))
            else:
                h_align = 'left'
                wrap = (header in WRAP_COLS)
                cell.alignment = Alignment(horizontal=h_align, vertical='top', wrap_text=wrap)
            # Borders
            cell.border = border

    # Column width approximation (auto-fit like)
    for c in range(1, max_col + 1):
        header = ws.cell(row=1, column=c).value
        max_len = len(str(header)) if header is not None else 0
        for r in range(2, max_row + 1):
            v = ws.cell(row=r, column=c).value
            l = len(str(v)) if v is not None else 0
            if l > max_len:
                max_len = l
        # scale: roughly 1 unit ~ one character, add padding
        width = min(max_len + 2, 100)
        ws.column_dimensions[get_column_letter(c)].width = width

    # Do not explicitly set row heights; let Excel auto-size on open


def find_primary_visible_sheet(wb):
    for ws in wb.worksheets:
        if ws.sheet_state == 'visible':
            return ws
    return wb.active


def main():
    parser = argparse.ArgumentParser(description='Reorganize Excel into MAIN(TestPlan) and very-hidden META sheet.')
    parser.add_argument('--input', required=True, help='Path to input .xlsx inside repo')
    parser.add_argument('--output', required=True, help='Path for output .xlsx inside repo')
    args = parser.parse_args()

    # Load workbook
    wb = load_workbook(args.input)

    # Validate .xlsx by virtue of load success; proceed with first visible sheet as main
    src_ws = find_primary_visible_sheet(wb)

    # Create META sheet (very hidden)
    copy_meta_sheet(src_ws, wb)

    # Normalize and format MAIN sheet
    main_ws = build_testplan_sheet(src_ws, wb)
    format_testplan(main_ws)

    # Save to output
    wb.save(args.output)


if __name__ == '__main__':
    main()
