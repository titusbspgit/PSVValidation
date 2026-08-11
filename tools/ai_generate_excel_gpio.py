#!/usr/bin/env python3
import json, os, sys
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.sheetproperties import SheetProperties

IP_NAME = "GPIO"
OUTPUT_DIR = "Test_Output/GPIO/TestPlan"
AGG_JSON_PATH = os.path.join(OUTPUT_DIR, "ai_aggregated.json")

def ist_now():
    ist = timezone(timedelta(hours=5, minutes=30))
    return datetime.now(ist)

def make_filename():
    ts = ist_now().strftime("%Y%m%d_%H%M%S")
    return f"{IP_NAME}_TestPlan_{ts}.xlsx"

HEADERS = [
    "Index","SS / Module","Feature","Test Case Name","Test Description",
    "Meta Test Description","Speed","Mode","Memory Start Offset","Memory End Offset",
    "Remarks","Test Steps / Procedure","Meta Test Steps / Procedure","Impacted Registers",
    "Meta Impacted Registers","Validation / Acceptance Criteria","Meta Validation / Acceptance Criteria",
    "Code Generation (Required / Not)","Meta Headers","Meta Macros","Meta Arrays"
]

META_HEADERS = ["Key","Value"]


def apply_format(ws):
    # Bold + fill header
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Freeze first row
    ws.freeze_panes = "A2"
    # Wrap and set widths
    for col in range(1, len(HEADERS)+1):
        ws.column_dimensions[get_column_letter(col)].width = 24
    # Set default alignment for data rows
    for row in ws.iter_rows(min_row=2, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def very_hide(ws):
    ws.sheet_state = 'veryHidden'


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"
    apply_format(ws)
    # Write rows in order, preserve keys exactly as provided via HEADERS
    r = 2
    for item in rows:
        for c, key in enumerate(HEADERS, 1):
            ws.cell(row=r, column=c, value=item.get(key, "NA"))
        r += 1

    # MetaData sheet
    md = wb.create_sheet("MetaData")
    md_props = SheetProperties()
    very_hide(md)
    # Minimal metadata
    md.cell(row=1, column=1, value="Key").font = Font(bold=True)
    md.cell(row=1, column=2, value="Value").font = Font(bold=True)
    md.cell(row=2, column=1, value="IP_NAME")
    md.cell(row=2, column=2, value=IP_NAME)
    md.cell(row=3, column=1, value="GeneratedAtIST")
    md.cell(row=3, column=2, value=ist_now().strftime("%Y-%m-%d %H:%M:%S IST"))
    return wb


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(AGG_JSON_PATH, 'r', encoding='utf-8') as f:
        rows = json.load(f)
    wb = build_workbook(rows)
    fname = make_filename()
    out_path = os.path.join(OUTPUT_DIR, fname)
    wb.save(out_path)
    print(out_path)

if __name__ == "__main__":
    main()
