#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os
from pathlib import Path
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Aggregated JSON (exact input)
DATA_JSON = r'''[
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "NA",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Validate basic register access for the GPIO GP0 registers referenced by the test using read/write flows.",
    "Meta Test Description": "The test reads a soft-reset/control register (SOFT_RST_REG_ADDRESS) in program.c, then performs register accesses on multiple GPIO GP0 register macros (MIZAR_GPIO_GP0_GPIO_8 through MIZAR_GPIO_GP0_GPIO_27) referenced in test_define.c. Specific per-register operations are not detailed in the provided sources. Intent implied by the testcase name is to perform register write/read validation over the referenced GPIO GP0 registers.",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "NA",
    "Test Steps / Procedure": "1. Read the reset/control register used by the test and record the value.\n2. For each GPIO GP0 register referenced by the test (pins 8 through 27), exercise the register access flow used by the test (e.g., write then read back, as applicable).\n3. Record observed values for each access and note any anomalies.",
    "Meta Test Steps / Procedure": "1) program.c: read from SOFT_RST_REG_ADDRESS; capture the returned value. No compare or mask operations specified in the provided inputs.\n2) test_define.c: access the following GPIO GP0 register macros in sequence: MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, MIZAR_GPIO_GP0_GPIO_11, MIZAR_GPIO_GP0_GPIO_12, MIZAR_GPIO_GP0_GPIO_13, MIZAR_GPIO_GP0_GPIO_14, MIZAR_GPIO_GP0_GPIO_15, MIZAR_GPIO_GP0_GPIO_16, MIZAR_GPIO_GP0_GPIO_17, MIZAR_GPIO_GP0_GPIO_18, MIZAR_GPIO_GP0_GPIO_19, MIZAR_GPIO_GP0_GPIO_20, MIZAR_GPIO_GP0_GPIO_21, MIZAR_GPIO_GP0_GPIO_22, MIZAR_GPIO_GP0_GPIO_23, MIZAR_GPIO_GP0_GPIO_24, MIZAR_GPIO_GP0_GPIO_25, MIZAR_GPIO_GP0_GPIO_26, MIZAR_GPIO_GP0_GPIO_27. Operation types for these macros are not specified in the inputs (marked as unknown).\n3) No explicit loops, conditions, bitwise operations, waits, interrupts, or assertions are provided in the inputs; treat accesses as straightforward register operations for the referenced macros.",
    "Impacted Registers": "NA",
    "Meta Impacted Registers": "SOFT_RST_REG_ADDRESS; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27",
    "Validation / Acceptance Criteria": "NA",
    "Meta Validation / Acceptance Criteria": "NA",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "NA",
    "Meta Macros": "#define MIZAR_GPIO_BASE 0xA001A000; #define MIZAR_GPIO_GP0_GPIO_8 MIZAR_GPIO_BASE + GPIO_GP0_GPIO_8_OFFSET; #define GPIO_GP0_GPIO_8_OFFSET 0x0; #define GPIO_GP0_GPIO_8_DEFAULT_VAL 0x00100000; #define GPIO_GP0_GPIO_8_VALID_MASK 0x003F0003; #define GPIO_GP0_GPIO_8_WRITE_MASK 0x003F0000",
    "Meta Arrays": "NA"
  }
]'''

IP_NAME = "GPIO"
OUTPUT_DIR = Path("Test_Output") / IP_NAME / "TestPlan"


def compute_ist_timestamp():
    ist = timezone(timedelta(hours=5, minutes=30))
    return datetime.now(ist).strftime("%Y%m%d_%H%M%S")


def split_headers_preserve_order(row_dict):
    keys_in_order = list(row_dict.keys())
    test_headers = [k for k in keys_in_order if not k.startswith("Meta ")]
    meta_headers = [k for k in keys_in_order if k.startswith("Meta ")]
    return test_headers, meta_headers


def apply_formatting(ws, headers):
    header_fill = PatternFill(fill_type="solid", fgColor="FFD9E1F2")  # light blue
    header_font = Font(bold=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Header styles
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = wrap_align

    # Body wrap
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.alignment = wrap_align

    # Freeze first row
    ws.freeze_panes = "A2"

    # Reasonable column widths
    long_hints = {"Description", "Steps", "Impacted", "Validation", "Macros", "Headers", "Arrays"}
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        base = 18
        if any(h in header for h in long_hints):
            base = 60
        if "Steps" in header:
            base = 80
        width = max(base, min(max_len + 2, 80))
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_workbook(data_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"
    ws_meta = wb.create_sheet(title="MetaData")
    ws_meta.sheet_state = "veryHidden"

    # Determine headers from the first row (order-preserving)
    test_headers, meta_headers = split_headers_preserve_order(data_rows[0])

    # Write headers
    ws.append(test_headers)
    ws_meta.append(meta_headers)

    # Write rows preserving order and data exactly
    for row in data_rows:
        ws.append([row.get(h, "") for h in test_headers])
        ws_meta.append([row.get(h, "") for h in meta_headers])

    # Apply formatting
    apply_formatting(ws, test_headers)
    apply_formatting(ws_meta, meta_headers)

    return wb


def main():
    data = json.loads(DATA_JSON)
    assert isinstance(data, list) and len(data) > 0, "Aggregated JSON must contain at least one row"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    ts = compute_ist_timestamp()
    filename = f"{IP_NAME}_TestPlan_{ts}.xlsx"
    output_path = OUTPUT_DIR / filename

    wb = build_workbook(data)
    wb.save(output_path)

    print(str(output_path))


if __name__ == "__main__":
    main()
