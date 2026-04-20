#!/usr/bin/env python3
import json
import os
from pathlib import Path
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo  # pragma: no cover

from openpyxl import Workbook
from openpyxl.styles import Font


def dumps_compact(obj):
    return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))


def main():
    repo_root = Path(__file__).resolve().parents[2]
    json_path = repo_root / "tools/json/gpio_testplan.json"
    if not json_path.exists():
        raise SystemExit(f"JSON not found: {json_path}")

    with json_path.open('r', encoding='utf-8') as f:
        data = json.load(f)

    # Determine columns in strict order
    columns = ["ip", "repo", "generated_at", "assumptions", "tests", "notes", "gaps"]

    # Prepare row values, serializing arrays/objects compactly
    row = []
    for key in columns:
        val = data.get(key, None)
        if isinstance(val, (dict, list)):
            row.append(dumps_compact(val))
        elif val is None:
            row.append("")
        else:
            # Preserve exact text representation
            row.append(str(val))

    # Create workbook with exactly one sheet named 'TestPlan'
    wb = Workbook()
    ws = wb.active
    ws.title = "TestPlan"

    # Header formatting: bold, freeze top row
    header_font = Font(bold=True)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
    ws.freeze_panes = "A2"

    # Write data row
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    # Column width autosize (approximate)
    for col_idx, name in enumerate(columns, start=1):
        values = [name, str(row[col_idx-1])]
        max_len = max(len(v) for v in values if v is not None)
        ws.column_dimensions[chr(64+col_idx)].width = min(max_len + 2, 120)

    # Compute IST timestamp and filename
    ist = ZoneInfo("Asia/Kolkata")
    now_ist = datetime.now(ist)
    ymd = now_ist.strftime("%Y%m%d")
    hms = now_ist.strftime("%H%M%S")
    pretty = now_ist.strftime("%Y-%m-%d %H:%M:%S IST")

    filename = f"gpio_TestPlan_{ymd}_{hms}.xlsx"
    out_dir = repo_root / "Test_Output" / "GPIO" / "TestPlan"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename

    wb.save(out_path)

    print(f"Generated Excel: {out_path}")

    # Expose outputs to GitHub Actions
    gh_out = os.environ.get("GITHUB_OUTPUT")
    if gh_out:
        with open(gh_out, "a", encoding="utf-8") as f:
            f.write(f"filename={filename}\n")
            f.write(f"ist_pretty={pretty}\n")


if __name__ == "__main__":
    main()
