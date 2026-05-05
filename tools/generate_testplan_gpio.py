import json, os, sys, zipfile
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:
    from backports.zoneinfo import ZoneInfo  # type: ignore

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Configuration
INPUT_JSON_PATH = os.path.join('tools', 'gpio_testplan_input.json')
OUTPUT_DIR = os.path.join('Test_Output', 'GPIO', 'TestPlan')
IP_NAME = 'GPIO'
TZ = ZoneInfo('Asia/Kolkata')

MAIN_COLUMNS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLUMNS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]

WRAP_COLUMNS = set([
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
])

HEADER_FILL = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')  # Blue
HEADER_FONT = Font(bold=True, color='FFFFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=False)
DATA_ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
DATA_ALIGN_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=False)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def load_json(path: str):
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, dict) and 'test_cases' in data and isinstance(data['test_cases'], list):
        meta = data.get('metadata', {})
        rows = data['test_cases']
    elif isinstance(data, list):
        meta = {}
        rows = data
    else:
        raise ValueError('Invalid JSON structure: expected array or object with test_cases')
    if not rows:
        raise ValueError('Empty test_cases array')
    return meta, rows


def union_keys_preserve_order(rows):
    seen = []
    sset = set()
    for row in rows:
        for k in row.keys():
            if k not in sset:
                sset.add(k)
                seen.append(k)
    # Ensure META columns exist in union (preserve order appending if missing)
    for k in META_COLUMNS:
        if k not in sset:
            seen.append(k)
            sset.add(k)
    return seen


def to_cell_value(v):
    if isinstance(v, list):
        return ', '.join(str(x) for x in v)
    return v


def write_data_sheet(wb, headers, rows):
    # Remove default sheet if present
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)
    ws = wb.create_sheet('Data')
    # Header
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    # Rows
    for r_idx, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            v = to_cell_value(row.get(h, ''))
            cell = ws.cell(row=r_idx, column=c, value=v)
            # Alignment defaults; wrap adjusted later
            if h == 'Index' and isinstance(v, (int, float)):
                cell.alignment = DATA_ALIGN_CENTER
            else:
                cell.alignment = DATA_ALIGN_LEFT
            cell.border = THIN_BORDER
    ws.freeze_panes = 'A2'
    # Approx auto-fit columns based on max length
    autofit_columns(ws)
    return ws


def autofit_columns(ws):
    col_widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            if value is None:
                l = 0
            else:
                s = str(value)
                l = max(len(line) for line in s.splitlines()) if s else 0
            col_widths[idx] = max(col_widths.get(idx, 0), l)
    for idx, width in col_widths.items():
        adj = min(max(width * 1.2 + 2, 10), 80)
        ws.column_dimensions[get_column_letter(idx)].width = adj


def create_meta_sheet(wb, meta, headers, rows):
    ws = wb.create_sheet('Meta_data_sheet')
    # Write top-level metadata block
    md_items = [
        ('IP_NAME', meta.get('ip_name', IP_NAME)),
        ('SOURCE_REPO', meta.get('source_repo', '')),
        ('SUBDIRECTORY', meta.get('subdirectory', '')),
        ('GENERATION_TIMESTAMP_IST', meta.get('generation_timestamp_ist', '')),
    ]
    for r, (k, v) in enumerate(md_items, start=1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
    start_row = len(md_items) + 2
    # Write META columns header
    for c, h in enumerate(META_COLUMNS, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    # Write META rows
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c, h in enumerate(META_COLUMNS, start=1):
            v = to_cell_value(row.get(h, ''))
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.alignment = DATA_ALIGN_LEFT
            cell.border = THIN_BORDER
    autofit_columns(ws)
    # Very hidden
    ws.sheet_state = 'veryHidden'
    return ws


def rebuild_as_testplan(ws, headers, rows):
    # Build in-memory matrix for MAIN columns only
    data_matrix = []
    for row in rows:
        rec = []
        for h in MAIN_COLUMNS:
            rec.append(to_cell_value(row.get(h, '')))
        data_matrix.append(rec)
    # Clear existing content
    ws.delete_rows(1, ws.max_row)
    # Write new header
    for c, h in enumerate(MAIN_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    # Write data rows
    for r_idx, rec in enumerate(data_matrix, start=2):
        for c, v in enumerate(rec, start=1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            # Alignment rules
            h = MAIN_COLUMNS[c-1]
            if h == 'Index' and isinstance(v, (int, float)):
                cell.alignment = DATA_ALIGN_CENTER
            elif h in WRAP_COLUMNS:
                cell.alignment = DATA_ALIGN_LEFT_WRAP
            else:
                cell.alignment = DATA_ALIGN_LEFT
            cell.border = THIN_BORDER
    # Wrap text columns and renumber list items
    wrap_and_number_columns(ws)
    # Header formatting already done; Freeze panes
    ws.freeze_panes = 'A2'
    # Data validation for Code Generation column
    apply_codegen_validation(ws)
    # Autofit again and adjust row heights
    autofit_columns(ws)
    adjust_row_heights(ws)


def wrap_and_number_columns(ws):
    header_to_col = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    for h in WRAP_COLUMNS:
        if h not in header_to_col:
            continue
        col = header_to_col[h]
        for r in range(2, ws.max_row+1):
            cell = ws.cell(row=r, column=col)
            val = cell.value
            if val is None:
                continue
            lines = [x.strip() for x in str(val).splitlines() if x.strip()]
            if not lines:
                continue
            # Renumber with 1., 2., ...
            renum = []
            for i, line in enumerate(lines, start=1):
                # Strip any leading bullets/numbers
                stripped = line
                # common patterns: '1)', '1.', '-', '*'
                if len(stripped) > 2 and (stripped[1] in ').') and stripped[0].isdigit():
                    stripped = stripped[2:].strip()
                elif stripped[:2] in ('- ', '* '):
                    stripped = stripped[2:].strip()
                renum.append(f"{i}. {stripped}")
            new_text = "\n".join(renum)
            cell.value = new_text
            cell.alignment = DATA_ALIGN_LEFT_WRAP


def apply_codegen_validation(ws):
    header_to_col = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    h = 'Code Generation (Required / Not)'
    if h not in header_to_col:
        return
    col = header_to_col[h]
    col_letter = get_column_letter(col)
    dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showDropDown=True)
    dv.error = 'Select one of: Required, Blank, Not Required'
    dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")


def adjust_row_heights(ws):
    # Estimate row height by number of wrapped lines across WRAP_COLUMNS
    header_to_col = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    wrap_cols = [header_to_col[h] for h in WRAP_COLUMNS if h in header_to_col]
    base_height = 15
    for r in range(2, ws.max_row+1):
        max_lines = 1
        for c in wrap_cols:
            val = ws.cell(row=r, column=c).value
            if val:
                lines = str(val).count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
        ws.row_dimensions[r].height = base_height * max(1, max_lines)


def set_sheet_visibility_and_rename(wb):
    ws = wb['Data']
    ws.title = 'TestPlan'
    # Ensure no stray 'Data' sheet
    if 'Data' in wb.sheetnames:
        ds = wb['Data']
        wb.remove(ds)


def validate_xlsx(path):
    # Check ZIP structure
    with zipfile.ZipFile(path, 'r') as zf:
        if '[Content_Types].xml' not in zf.namelist():
            raise ValueError('Not a valid OOXML (missing [Content_Types].xml)')
    # Re-open with openpyxl
    _ = load_workbook(path)


def main():
    meta, rows = load_json(INPUT_JSON_PATH)
    # Update generation timestamp in IST for metadata sheet
    now_ist = datetime.now(TZ)
    meta = dict(meta or {})
    meta['ip_name'] = meta.get('ip_name') or IP_NAME
    meta['generation_timestamp_ist'] = now_ist.strftime('%Y-%m-%d %H:%M:%S IST')

    headers = union_keys_preserve_order(rows)

    wb = Workbook()
    ws_data = write_data_sheet(wb, headers, rows)

    # Create META sheet
    create_meta_sheet(wb, meta, headers, rows)

    # Rebuild Data sheet as TestPlan with MAIN columns only
    rebuild_as_testplan(ws_data, headers, rows)

    # Rename to TestPlan and ensure no 'Data' remains
    set_sheet_visibility_and_rename(wb)

    # Ensure only allowed sheets exist: TestPlan (visible) and Meta_data_sheet (veryHidden)
    allowed = set(['TestPlan', 'Meta_data_sheet'])
    for name in list(wb.sheetnames):
        if name not in allowed:
            ws = wb[name]
            wb.remove(ws)
    # Safety check
    if 'TestPlan' not in wb.sheetnames or 'Meta_data_sheet' not in wb.sheetnames:
        raise RuntimeError('Final sheet set invalid')

    # Save file with IST timestamp
    fname = f"{IP_NAME}_TestPlan_{now_ist.strftime('%Y%m%d')}_{now_ist.strftime('%H%M%S')}.xlsx"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, fname)
    wb.save(out_path)

    # Validate OOXML
    validate_xlsx(out_path)

    print(out_path)

if __name__ == '__main__':
    main()
