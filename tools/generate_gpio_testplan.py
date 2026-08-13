import os
import json
import subprocess
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def main():
    ip_name = os.getenv('IP_NAME', 'GPIO')
    output_dir = os.getenv('OUTPUT_DIR', 'Test_Output/GPIO/TestPlan')

    ist = ZoneInfo('Asia/Kolkata')
    now_ist = datetime.now(ist)
    ts = now_ist.strftime('%Y%m%d_%H%M%S')
    filename = f"{ip_name}_TestPlan_{ts}.xlsx"
    out_path = os.path.join(output_dir, filename)

    # Create workbook
    wb = Workbook()

    # TestPlan sheet
    ws = wb.active
    ws.title = 'TestPlan'

    headers = [
        'TestID', 'TestName', 'Description', 'Preconditions', 'Steps',
        'ExpectedResult', 'Priority', 'Owner', 'Tags'
    ]

    # Write headers
    ws.append(headers)

    # Formatting
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    wrap_align = Alignment(wrap_text=True, vertical='top')

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align

    # Freeze first row
    ws.freeze_panes = 'A2'

    # Reasonable column widths
    widths = [10, 24, 40, 28, 40, 32, 10, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # MetaData very hidden sheet containing final_json exactly as-is
    md = wb.create_sheet('MetaData')
    final_json = []  # exactly as provided
    md['A1'] = json.dumps(final_json, separators=(',', ':'))
    md.sheet_state = 'veryHidden'

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    wb.save(out_path)

    # Commit and push
    subprocess.run(['git', 'config', 'user.email', 'actions@github.com'], check=True)
    subprocess.run(['git', 'config', 'user.name', 'github-actions'], check=True)
    subprocess.run(['git', 'add', out_path], check=True)
    commit_msg = f"Add {ip_name} TestPlan {ts} (auto-generated)"
    # It's okay if there are no changes (e.g., rerun) — handle gracefully
    try:
        subprocess.run(['git', 'commit', '-m', commit_msg], check=True)
    except subprocess.CalledProcessError:
        print('No changes to commit.')
    subprocess.run(['git', 'push'], check=True)

if __name__ == '__main__':
    main()
