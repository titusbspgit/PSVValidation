#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Deterministic fallback generator for GPIO_TestPlan_WORKING.xlsx (Batch 1)
- Builds Excel directly from embedded JSON (3 rows) per strict rules
- If an existing Excel exists at the target path, merges existing TestPlan rows with incoming rows
- Applies formatting, Meta sheet extraction, and validations exactly as specified
"""
import json
import os
import zipfile
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ========================
# Embedded JSON (array)
# ========================
JSON_INPUT = r'''[
  {"Index": 1, "SS / Module": "GPIO", "Feature": "interrupts can be generated based on negative edge detection at GPIO input", "Test Case Name": "test_gpio_nedge_random_pads_en", "Test Description": "Verifies negative-edge interrupt behavior on randomly selected GPIO pads and checks that status and clear operations work as expected.", "Speed": "NA", "Mode": "Interrupt", "Memory Start Offset": "0xA0243ffc", "Memory End Offset": "0xA0243ffc", "Remarks": "Uses conditional compilation to select interrupt source (GPIO0 or GPIO1). Writes to SRAM location 0xA0243ffc to drive input changes. Enables input mode and negative-edge interrupt for GPIOs in the range 8–39 as stated in code comments.", "Test Steps / Procedure": "1) If GPIO0 is selected, enable interrupt ID 87 in the interrupt controller and enable the corresponding system interrupt in INTR_EN1; if GPIO1 is selected, enable interrupt ID 88 and enable the corresponding system interrupt in INTR_EN1.\n2) Initialize the random generator and set the SRAM word at 0xA0243ffc to all ones to establish a known state.\n3) For each of 32 iterations, select a unique random pad index from 0–31; if a duplicate is selected, repeat the iteration.\n4) For the chosen pad, program gp0_intr2_intr_en1 plus the pad offset to set input mode and negative-edge detection, then wait briefly.\n5) Enable the pad’s group interrupt by setting its bit in INTR1_INTR_EN1.\n6) Toggle the SRAM word at 0xA0243ffc to drive a falling edge on the selected pad, then restore it to all ones.\n7) Wait until the pending flag clears, indicating the interrupt handler has executed.\n8) In the interrupt handler, read the pad’s control/status register (gp0_intr2_intr_en1 plus the pad offset) and confirm the input bit reflects the expected value and the raw interrupt bit is set.\n9) Read INTR1_INTR_STS1 and confirm that the selected pad’s status bit is set.\n10) Disable the group interrupt in INTR1_INTR_EN1, write the clear value to the pad’s control/status register (gp0_intr2_intr_en1 plus the pad offset), and verify the readback value matches the expected cleared state.\n11) Read INTR1_INTR_STS1 again and confirm the group status is cleared.\n12) If GPIO0 is selected, clear the system raw status in RAW_STCR1 for the GPIO0 source and confirm it reads back cleared; if GPIO1 is selected, perform the same for the GPIO1 source.\n13) Finally, clear the corresponding interrupt in the interrupt controller (ID 87 for GPIO0 or ID 88 for GPIO1).", "Impacted Registers": ["gp0_intr2_intr_en1", "INTR1_INTR_EN1", "INTR1_INTR_STS1", "INTR_EN1", "RAW_STCR1"], "Validation / Acceptance Criteria": "- The selected pad’s input bit reads as expected and the raw interrupt bit is set in the per-pad register; pass if both conditions are true, otherwise fail.\n- The selected pad’s status bit is set in INTR1_INTR_STS1 after the edge; pass if set, otherwise fail.\n- After clearing, the per-pad register readback matches the expected cleared value; pass if it matches, otherwise fail.\n- The group status in INTR1_INTR_STS1 reads zero after clear; pass if zero, otherwise fail.\n- The system raw status in RAW_STCR1 for the selected source reads cleared after the clear operation; pass if cleared, otherwise fail.", "Code Generation (Required / Not)": "", "Hidden_Test_Case_Name": "test_gpio_nedge_random_pads_en", "Hidden_Test_Description": "Test programs GPIO input mode and negative-edge interrupt for randomly selected pads (GPIOs 8–39 implied by code comment), enables group interrupt, toggles an SRAM location to generate a falling edge, waits for the interrupt, and validates per-pad DIN bit, raw interrupt status, group interrupt status, and clearing sequences including system register raw status. Source: https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_random_pads_en/program.c", "Hidden_Remarks": "Conditional blocks use GPIO0 and GPIO1 to choose interrupt IDs 87 or 88 and related system register bits. The code writes to 0xA0243ffc as an SRAM location to generate pad input transitions. A comment states enabling input mode and negedge interrupt for GPIOs 8–39. Source evidence only within folder.", "Hidden_Test_Steps_Procedure": "File references: program.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_random_pads_en/program.c), test_define.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_random_pads_en/test_define.c), Makefile (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_random_pads_en/Makefile).\n1. test_case entry prints \"test_case\" to console.\n2. Set test_err = 0.\n3. If GPIO0 is defined: call GIC_EnableIRQ(87).\n4. If GPIO1 is defined: call GIC_EnableIRQ(88).\n5. If GPIO0 is defined: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR).\n6. If GPIO1 is defined: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR).\n7. Call srand(time(NULL)).\n8. write_reg(0xA0243ffc, 0xffffffff).\n9. For i from 0 to < 32 do:\n   9.1. Set pad_num = rand() % 32.\n   9.2. For j from 0 to <= i-1 do: if (pad_num == arr[j]) break.\n   9.3. If (i == j):\n       9.3.1. arr[i] = pad_num.\n       9.3.2. wr_val = 1 << pad_num.\n       9.3.3. write_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4), 0x00140000) to enable input mode and negedge interrupt per comment.\n       9.3.4. wait_on(50).\n       9.3.5. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 1 << pad_num) to enable the GPIO group interrupt.\n       9.3.6. wait_on(10).\n       9.3.7. write_reg(0xA0243ffc, ~(wr_val)).\n       9.3.8. wait_on(10).\n       9.3.9. write_reg(0xA0243ffc, 0xffffffff).\n       9.3.10. Set int_pend = 1.\n       9.3.11. While (int_pend == 0x1):\n           9.3.11.1. printf(\"Waiting for interrupt\\n\").\n           9.3.11.2. wait_on(10).\n   9.4. Else (duplicate pad selected): set i = i - 1 (retry iteration).\n10. Call finish(test_err).\n11. Default_IRQHandler entry (interrupt handler):\n    11.1. Compute wr_val = 1 << pad_num.\n    11.2. Set int_pend = 0x0.\n    11.3. rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4)).\n    11.4. If ((rdata & 0x1) != 0): (success message under DEBUG_DISPLAY only; no action when not defined).\n    11.5. Else: printf(\"ERROR: GPIO_NUM = %0d Default_IRQHandler:: DIN value does not match with the Pad_value read_data = %0x\\n\", pad_num, rdata); test_err++.\n    11.6. If ((rdata & 0x2) != 0x0):\n        11.6.1. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1).\n        11.6.2. If ((rdata_grp & (1 << pad_num)) != 0): (success message under DEBUG_DISPLAY only).\n        11.6.3. Else: printf(\"ERROR: Group Interrupt not occured\\n\"); test_err = test_err + 1.\n        11.6.4. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000) to disable group interrupt.\n        11.6.5. write_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4), 0x00110001) to clear the interrupt (per comment: 16th bit set to '1').\n        11.6.6. wait_on(2).\n        11.6.7. rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4)).\n        11.6.8. If (rdata == 0x100001): (success message under DEBUG_DISPLAY only).\n        11.6.9. Else: printf(\"ERROR : Interrupt clear failed : Interrupt value = %x\\n\", rdata); test_err = test_err + 1.\n        11.6.10. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1).\n        11.6.11. If (rdata_grp == 0x0): (success message under DEBUG_DISPLAY only).\n        11.6.12. Else: printf(\"ERROR : Group Interrupt clear failed: Interrupt value:%x\\n\", rdata_grp); test_err = test_err + 1.\n        11.6.13. If GPIO0 is defined:\n            11.6.13.1. write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR).\n            11.6.13.2. rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1).\n            11.6.13.3. If ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0): (success message under DEBUG_DISPLAY only).\n            11.6.13.4. Else: printf(\"sysreg status not cleared : %0x\\n\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++.\n        11.6.14. If GPIO1 is defined:\n            11.6.14.1. write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR).\n            11.6.14.2. rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1).\n            11.6.14.3. If ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0): (success message under DEBUG_DISPLAY only).\n            11.6.14.4. Else: printf(\"sysreg status not cleared : %0x\\n\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++.\n    11.7. Else: printf(\"Interrupt Not occured\\n\"); test_err++.\n    11.8. If GPIO0 is defined: GIC_ClearIRQ(87).\n    11.9. If GPIO1 is defined: GIC_ClearIRQ(88).\nNotes: DEBUG_DISPLAY-related printf statements are compiled only if DEBUG_DISPLAY is defined; Makefile in this folder defines DEBUG_RW_MSG only.", "Hidden_Impacted_Registers": ["MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_LSS_SYSREG_RAW_STCR1", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1"], "Hidden_Validation_Acceptance_Criteria": ["If ((rdata & 0x1) != 0) after reading MIZAR_GPIO_GP0_GPIO_8 + (pad_num*4), treat as success; else print error and increment test_err.", "If ((rdata & 0x2) != 0x0), proceed to read group status; else print \"Interrupt Not occured\" and increment test_err.", "After read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1), if (rdata_grp & (1<<pad_num)) != 0, success; else print error and increment test_err.", "After disabling group interrupt and clearing per-pad register, if readback rdata == 0x100001, success; else print error and increment test_err.", "After clearing, if read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) == 0x0, success; else print error and increment test_err.", "If GPIO0: after write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR), then readback ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0) indicates success; else print error and increment test_err.", "If GPIO1: after write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR), then readback ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0) indicates success; else print error and increment test_err."]},
  {"Index": 2, "SS / Module": "GPIO", "Feature": "neie: Negative edge interrupt enable; 1 - Enable interrupt when falling edge (neg edge) is detected on gpio pin; 0 - Disable interrupt on neg-edge detection;", "Test Case Name": "test_gpio_nedge_walking_zeros_pattern", "Test Description": "Validates falling-edge triggered interrupts across a sequence of GPIO inputs using a walking‑zeros stimulus and verifies status and clear behavior.", "Speed": "NA", "Mode": "Interrupt", "Memory Start Offset": "0xA0243ffc", "Memory End Offset": "0xA0243ffc", "Remarks": "The test drives input changes using memory at 0xA0243ffc. It iterates across 32 GPIOs. One of two GPIO interrupt sources and corresponding interrupt IDs is selected conditionally at build time. Group IO control registers are set to input mode for four groups.", "Test Steps / Procedure": "1) Enable the relevant interrupt ID in the interrupt controller.\n2) Enable the system interrupt source through gp0_intr2_intr_en1.\n3) For each of 32 GPIOs, configure the per‑GPIO control to trigger on a falling edge using gp0_intr2_intr_en1.\n4) Set GPIO_IO_CTRL_GROUP1, GPIO_IO_CTRL_GROUP2, GPIO_IO_CTRL_GROUP3, and GPIO_IO_CTRL_GROUP4 so the GPIOs operate in input mode.\n5) Enable all GPIO group interrupts via gp0_intr2_intr_en1.\n6) For each GPIO index, write all ones to 0xA0243ffc, then clear only the target bit to create a falling edge, and finally restore all ones; wait for the interrupt to be serviced.\n7) In the interrupt handler, read the per‑GPIO control and confirm the input status reflects the edge and the raw status bit is set.\n8) Read INTR1_INTR_STS1 and confirm the corresponding group status bit is set.\n9) Clear the per‑GPIO raw status in gp0_intr2_intr_en1 and verify the readback shows the cleared state.\n10) Read INTR1_INTR_STS1 again and confirm the group status is cleared.\n11) Clear the selected system raw interrupt status in RAW_STCR1 and verify it reads cleared.\n12) Clear the pending interrupt in the interrupt controller.", "Impacted Registers": ["gp0_intr2_intr_en1", "INTR1_INTR_STS1", "RAW_STCR1", "GPIO_IO_CTRL_GROUP1", "GPIO_IO_CTRL_GROUP2", "GPIO_IO_CTRL_GROUP3", "GPIO_IO_CTRL_GROUP4"], "Validation / Acceptance Criteria": "1) After a falling edge, the per‑GPIO input status shows the expected active state and the raw status is set; pass if both are observed.\n2) INTR1_INTR_STS1 shows the bit corresponding to the active GPIO set after the edge; pass if set.\n3) After clearing the per‑GPIO raw status, the per‑GPIO readback matches the expected cleared value; pass if matched.\n4) After the clear, INTR1_INTR_STS1 reads zero for the group; pass if zero.\n5) After clearing the system raw status in RAW_STCR1, the corresponding source bit reads cleared; pass if cleared.", "Code Generation (Required / Not)": "", "Hidden_Test_Case_Name": "test_gpio_nedge_walking_zeros_pattern", "Hidden_Test_Description": "Test configures negative-edge interrupt on GPIOs 8–39, sets group IO controls to input mode, enables all group interrupts, then applies a walking‑zeros pattern via memory at 0xA0243ffc to generate falling edges per GPIO. It waits for an interrupt each time and validates per‑GPIO DIN bit, raw interrupt status, group interrupt status, and clear sequences including system raw status. Sources: program.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_walking_zeros_pattern/program.c), test_define.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_walking_zeros_pattern/test_define.c), Makefile (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_walking_zeros_pattern/Makefile).", "Hidden_Remarks": "Conditional compilation selects which interrupt ID to enable (87 or 88) and which system interrupt bit to set/clear. The comment states enabling input mode and negative-edge interrupt for GPIOs 8–39. Memory address 0xA0243ffc is used to drive input transitions for each GPIO index. A commented line to disable group interrupt is present but not executed.", "Hidden_Test_Steps_Procedure": "Files: program.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_walking_zeros_pattern/program.c), test_define.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_walking_zeros_pattern/test_define.c), Makefile (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_nedge_walking_zeros_pattern/Makefile).\n1. test_case entry; local variables declared: gpio_number, test_err, i; extern int_pend.\n2. If GPIO0 defined: call GIC_EnableIRQ(87).\n3. If GPIO1 defined: call GIC_EnableIRQ(88).\n4. Set test_err = 0.\n5. If GPIO0 defined: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR) to enable system interrupt for GPIO0.\n6. If GPIO1 defined: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR) to enable system interrupt for GPIO1.\n7. For i from 0 to 31 inclusive (32 iterations): write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00040000) to enable negative-edge interrupt (comment: 17th bit as '1').\n8. wait_on(10).\n9. Write group IO control: write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, 0x000000FF);\n10. write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, 0x000000FF);\n11. write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, 0x000000FF);\n12. write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, 0x000000FF). Comment indicates enabling input mode and negedge interrupt for GPIOs 8–39.\n13. wait_on(10).\n14. Enable group interrupt: write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0xFFFFFFFF).\n15. For i from 0 to 31 inclusive (32 iterations):\n   15.1. wr_val = 1 << i.\n   15.2. write_reg(0xA0243ffc, 0xFFFFFFFF).\n   15.3. wait_on(30).\n   15.4. write_reg(0xA0243ffc, ~(wr_val)).\n   15.5. wait_on(30).\n   15.6. int_pend = 1.\n   15.7. while (int_pend == 1): printf(\"Waiting for interrupt\"); wait_on(10).\n16. Call finish(test_err).\n17. Default_IRQHandler entry: local variables j, rdata, rdata_grp, wr_val declared.\n18. wr_val = 1 << i.\n19. int_pend = 0.\n20. write_reg(0xA0243ffc, 0xFFFFFFFF).\n21. If DEBUG_DISPLAY defined: printf(\"Entered into default IRQ Handler!! with pad value = %d\", i).\n22. rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)).\n23. If ((rdata & 0x1) != 0):\n   23.1. If DEBUG_DISPLAY defined: printf success message for DIN value match.\n24. Else:\n   24.1. printf(\"ERROR: ... DIN value does not match ... read_data = %0x\", rdata).\n   24.2. test_err++.\n25. If ((rdata & 0x2) != 0x0):\n   25.1. If DEBUG_DISPLAY defined: printf success message for raw interrupt raised.\n   25.2. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1).\n   25.3. If ((rdata_grp & (1<<i)) != 0):\n        25.3.1. If DEBUG_DISPLAY defined: printf success message for group interrupt raised.\n       Else:\n        25.3.2. printf(\"ERROR: Group Interrupt not occured\"); test_err = test_err + 1.\n   25.4. write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00110001) to clear raw status (comment: 16th bit set to '1').\n   25.5. wait_on(2).\n   25.6. rdata = read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)).\n   25.7. If (rdata == 0x100001): If DEBUG_DISPLAY defined: printf success for clear; Else:\n        25.7.1. printf(\"ERROR : Interrupt clear failed : Interrupt value = %x\", rdata); test_err = test_err + 1. (This else corresponds to rdata != 0x100001)\n   25.8. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1).\n   25.9. If (rdata_grp == 0x0): If DEBUG_DISPLAY defined: printf success for group clear; Else:\n        25.9.1. printf(\"ERROR : Group Interrupt clear failed: Interrupt value:%x\", rdata_grp); test_err = test_err + 1. (This else corresponds to rdata_grp != 0x0)\n   25.10. If GPIO0 defined:\n        25.10.1. write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR).\n        25.10.2. rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1).\n        25.10.3. If ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0): If DEBUG_DISPLAY defined: printf success; Else:\n              25.10.3.1. printf(\"sysreg status not cleared : %0x\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++ (This else corresponds to non-zero bit).\n   25.11. If GPIO1 defined:\n        25.11.1. write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR).\n        25.11.2. rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1).\n        25.11.3. If ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0): If DEBUG_DISPLAY defined: printf success; Else:\n              25.11.3.1. printf(\"sysreg status not cleared : %0x\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++.\n26. Else (i.e., (rdata & 0x2) == 0): printf(\"Interrupt Not occured\"); test_err++.\n27. If GPIO0 defined: GIC_ClearIRQ(87).\n28. If GPIO1 defined: GIC_ClearIRQ(88).\nNotes:\n- DEBUG_DISPLAY prints are compiled only if DEBUG_DISPLAY is defined; Makefile defines DEBUG_RW_MSG only.\n- The commented line // write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1,0x00000000); is not executed.", "Hidden_Impacted_Registers": ["MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_RAW_STCR1"], "Hidden_Validation_Acceptance_Criteria": ["If ((rdata & 0x1) != 0) after reading MIZAR_GPIO_GP0_GPIO_8 + (i*4), treat as success; else print error and increment test_err.", "If ((rdata & 0x2) != 0x0), proceed; else print \"Interrupt Not occured\" and increment test_err.", "After read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1), if (rdata_grp & (1<<i)) != 0, success; else print error and increment test_err.", "After clearing by write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00110001) and reading back, if rdata == 0x100001, success; else print error and increment test_err.", "After clearing, if read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) == 0x0, success; else print error and increment test_err.", "If GPIO0: after write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR), then ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0) indicates success; else print error and increment test_err.", "If GPIO1: after write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR), then ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0) indicates success; else print error and increment test_err."]},
  {"Index": 3, "SS / Module": "GPIO", "Feature": "neie: Negative edge interrupt enable; 1 - Enable interrupt when falling edge (neg edge) is detected on gpio pin; 0 - Disable interrupt on neg-edge detection", "Test Case Name": "test_gpio_negedge_all_pads_en", "Test Description": "Validates negative-edge interrupt behavior when enabled across all GPIO pads, including group status behavior and clear sequences.", "Speed": "NA", "Mode": "Interrupt", "Memory Start Offset": "0xA0243ffc", "Memory End Offset": "0xA0243ffc", "Remarks": "Input mode is enabled for GPIOs 8–39. Negative-edge detection is enabled per GPIO using the documented bit. The raw interrupt status is cleared by writing the documented bit. A shared SRAM word at 0xA0243ffc is toggled to generate input transitions. The handler disables group interrupts before clearing.", "Test Steps / Procedure": "1) Enable the relevant interrupt in the interrupt controller for the selected GPIO source.\n2) Enable the system interrupt source in INTR_EN1 for the selected GPIO source.\n3) For each of 32 GPIOs starting at the first per-GPIO control register (GP0_GPIO_8), enable negative-edge detection.\n4) Set GPIO_IO_CTRL_GROUP1, GPIO_IO_CTRL_GROUP2, GPIO_IO_CTRL_GROUP3, and GPIO_IO_CTRL_GROUP4 for input mode.\n5) Enable all group interrupt bits in gp0_intr2_intr_en1.\n6) For each GPIO index, write all ones to 0xA0243ffc, then write all zeros to generate falling edges, and wait until the interrupt is serviced.\n7) In the interrupt handler, read GP0_INTR1_INTR_STS1 and verify that at least one bit is set, then disable group interrupts in gp0_intr2_intr_en1.\n8) Clear the per-GPIO raw status by writing the clear value to each per-GPIO control register.\n9) Read GP0_INTR1_INTR_STS1 again and verify that the value is zero.\n10) Clear the system raw interrupt status for the selected source in RAW_STCR1 and verify it is cleared.\n11) Clear the pending interrupt in the interrupt controller.", "Impacted Registers": ["INTR_EN1", "RAW_STCR1", "GP0_GPIO_8", "gp0_intr2_intr_en1", "GP0_INTR1_INTR_STS1", "GPIO_IO_CTRL_GROUP1", "GPIO_IO_CTRL_GROUP2", "GPIO_IO_CTRL_GROUP3", "GPIO_IO_CTRL_GROUP4"], "Validation / Acceptance Criteria": "1) After generating falling edges, the group interrupt status shows at least one bit set; pass if set, otherwise fail.\n2) After clearing per-GPIO raw status, the group interrupt status reads zero; pass if zero, otherwise fail.\n3) After clearing the system raw status, the corresponding source bit in RAW_STCR1 reads cleared; pass if cleared, otherwise fail.", "Code Generation (Required / Not)": "", "Hidden_Test_Case_Name": "test_gpio_negedge_all_pads_en", "Hidden_Test_Description": "Test enables negative-edge interrupt per GPIO, configures GPIOs 8–39 for input mode, enables group interrupts, generates falling edges via 0xA0243ffc, then validates group interrupt status and performs clear sequences including system RAW status. Sources: program.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_negedge_all_pads_en/program.c), test_define.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_negedge_all_pads_en/test_define.c), Makefile (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_negedge_all_pads_en/Makefile).", "Hidden_Remarks": "- Comments specify bit semantics: negative-edge enable uses the 18th bit as '1'; raw status clear uses the 16th bit as '1'.\n- Input mode is enabled for GPIOs 8–39 via group IO control writes.\n- A shared SRAM location 0xA0243ffc is used to drive GPIO inputs high then low.\n- The handler disables group interrupts before clearing per-GPIO raw status.\n- Build-time switches (GPIO0/GPIO1) determine which interrupt ID and system interrupt bits are used.\n- DEBUG_DISPLAY prints are conditional; Makefile defines DEBUG_RW_MSG only.", "Hidden_Test_Steps_Procedure": "Files: program.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_negedge_all_pads_en/program.c), test_define.c (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_negedge_all_pads_en/test_define.c), Makefile (https://github.com/titusbspgit/PSVValidation/blob/main/TestRepo/gpio/test_gpio_negedge_all_pads_en/Makefile).\n1. test_case entry; declare rdata, wr_val; set test_err = 0.\n2. If GPIO0 defined: GIC_EnableIRQ(87); if GPIO1 defined: GIC_EnableIRQ(88).\n3. If GPIO0 defined: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR); if GPIO1 defined: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR).\n4. For i = 0..31: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00040000); comment: enabling negedge interrupt (18th bit as '1').\n5. wait_on(10).\n6. Enable input mode for GPIOs 8–39: write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, 0x000000FF); write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, 0x000000FF); write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, 0x000000FF); write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, 0x000000FF).\n7. wait_on(10).\n8. Enable group interrupt: write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0xFFFFFFFF).\n9. For i = 0..31:\n   9.1. wr_val = 1 << i; (not used further in this loop).\n   9.2. write_reg(0xA0243ffc, 0xFFFFFFFF).\n   9.3. wait_on(30).\n   9.4. write_reg(0xA0243ffc, 0x00000000).\n   9.5. int_pend = 1.\n   9.6. while (int_pend == 1): printf(\"Waiting for interrupt\"); wait_on(10).\n10. finish(test_err).\n11. Default_IRQHandler entry: declare j, rdata, rdata_grp, wr_val; wr_val = 1 << i; int_pend = 0; write_reg(0xA0243ffc, 0xFFFFFFFF); if DEBUG_DISPLAY: print entry message.\n12. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000) to disable group interrupts.\n13. If ((rdata_grp & 0x0FFFFFFF) != 0): if DEBUG_DISPLAY: print success; else: do nothing.\n14. Else: printf(\"ERROR: Group Interrupt not occured\"); test_err = test_err + 1.\n15. For i = 0..31: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00110001); comment: clearing raw status (16th bit as '1').\n16. wait_on(2).\n17. rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1).\n18. If (rdata_grp == 0x0): if DEBUG_DISPLAY: print success.\n19. Else: printf(\"ERROR : Group Interrupt clear failed: Interrupt value:%x\\n\", rdata_grp); test_err = test_err + 1.\n20. If GPIO0 defined: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0): if DEBUG_DISPLAY: print success; else: do nothing; else: printf(\"sysreg status not cleared : %0x\\n\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++.\n21. If GPIO1 defined: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0): if DEBUG_DISPLAY: print success; else: do nothing; else: printf(\"sysreg status not cleared : %0x\\n\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++.\n22. If GPIO0 defined: GIC_ClearIRQ(87); if GPIO1 defined: GIC_ClearIRQ(88).", "Hidden_Impacted_Registers": ["MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_LSS_SYSREG_RAW_STCR1", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4"], "Hidden_Validation_Acceptance_Criteria": ["If ((rdata_grp & 0x0FFFFFFF) != 0) after read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1), treat as success for group interrupt occurrence; else print \"ERROR: Group Interrupt not occured\" and increment test_err.", "After clearing per-GPIO raw status, if read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) == 0x0, treat as success; else print error and increment test_err.", "If GPIO0: after write_reg/read_reg on MIZAR_LSS_SYSREG_RAW_STCR1, ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) == 0) indicates success; else print error and increment test_err.", "If GPIO1: after write_reg/read_reg on MIZAR_LSS_SYSREG_RAW_STCR1, ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) == 0) indicates success; else print error and increment test_err."]}
]'''

OUTPUT_FILE = Path('Test_Output/GPIO/TestPlan/GPIO_TestPlan_WORKING.xlsx')
OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

MAIN_COLUMNS_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]
META_COLUMNS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]

BLUE_FILL = PatternFill(start_color='FF4F81BD', end_color='FF4F81BD', fill_type='solid')
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
WRAP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
)


def to_cell_value(v):
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def parse_json_input():
    try:
        data = json.loads(JSON_INPUT)
    except Exception as e:
        raise SystemExit(f"Invalid embedded JSON: {e}")
    if not isinstance(data, list) or len(data) == 0:
        raise SystemExit("Embedded JSON must be a non-empty array of objects")
    return data


def read_existing_rows_from_testplan(xlsx_path):
    rows = []
    if not xlsx_path.exists():
        return rows
    try:
        wb = load_workbook(xlsx_path)
        if 'TestPlan' not in wb.sheetnames:
            return rows
        ws = wb['TestPlan']
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all((cell is None or str(cell).strip() == '') for cell in r):
                continue
            row = {}
            for k, v in zip(headers, r):
                row[k] = v
            rows.append(row)
    except Exception:
        # On any failure reading, treat as no existing rows
        rows = []
    return rows


def build_union_keys(rows):
    seen = []
    for obj in rows:
        for k in obj.keys():
            if k not in seen:
                seen.append(k)
    return seen


def write_base_workbook(rows, union_keys):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'  # staging sheet

    # Header
    for col_idx, key in enumerate(union_keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL

    # Rows
    for r_idx, obj in enumerate(rows, start=2):
        for c_idx, key in enumerate(union_keys, start=1):
            ws.cell(row=r_idx, column=c_idx, value=to_cell_value(obj.get(key, '')))

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Basic autofit (approximate)
    for col_idx, key in enumerate(union_keys, start=1):
        max_len = len(str(key))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(120, max(10, max_len + 2))

    # Apply thin borders
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(union_keys)):
        for cell in row:
            cell.border = THIN_BORDER

    return wb, ws


def create_meta_sheet(wb, rows):
    ws_meta = wb.create_sheet('Meta_data_sheet')
    # Headers
    for col_idx, key in enumerate(META_COLUMNS, start=1):
        cell = ws_meta.cell(row=1, column=col_idx, value=key)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL
    # Data
    for r_idx, obj in enumerate(rows, start=2):
        for c_idx, key in enumerate(META_COLUMNS, start=1):
            ws_meta.cell(row=r_idx, column=c_idx, value=to_cell_value(obj.get(key, '')))
    # Very hidden
    ws_meta.sheet_state = 'veryHidden'


def normalize_main_sheet_inplace(wb):
    ws = wb['Data']
    # Build current headers
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # Determine which headers to keep (remove META, keep only MAIN order)
    main_set = set(MAIN_COLUMNS_ORDER)
    keep_headers = [h for h in headers if (h in main_set)]
    # Ensure final order strictly matches MAIN_COLUMNS_ORDER; include blanks for missing
    final_headers = MAIN_COLUMNS_ORDER[:]

    # Build a map from header -> column index in current sheet
    idx_map = {h: (headers.index(h) + 1) for h in headers if h in keep_headers}

    # Create a temporary in-memory table for final sheet
    final_table = []
    for r in range(2, ws.max_row + 1):
        row_dict = {}
        empty = True
        for h in final_headers:
            if h in idx_map:
                val = ws.cell(row=r, column=idx_map[h]).value
            else:
                val = ''
            if val not in (None, ''):
                empty = False
            row_dict[h] = val
        if not empty:
            final_table.append(row_dict)

    # Clear current sheet content
    ws.delete_rows(1, ws.max_row)

    # Write headers in final order
    for c, h in enumerate(final_headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = BLUE_FILL

    # Write data
    for r_idx, row_dict in enumerate(final_table, start=2):
        for c_idx, h in enumerate(final_headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row_dict.get(h, ''))

    # Rename sheet to TestPlan
    ws.title = 'TestPlan'

    # Wrap and align columns
    wrap_cols = {
        'Test Description',
        'Remarks',
        'Test Steps / Procedure',
        'Validation / Acceptance Criteria'
    }
    col_letter_by_header = {h: None for h in final_headers}
    for c_idx, h in enumerate(final_headers, start=1):
        col_letter_by_header[h] = ws.cell(row=1, column=c_idx).column_letter
        # Column width estimation
        max_len = len(str(h))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=c_idx).value
            if v is None:
                continue
            s = str(v)
            # Account for wrapped lines
            for part in s.split('\n'):
                if len(part) > max_len:
                    max_len = len(part)
        ws.column_dimensions[col_letter_by_header[h]].width = min(120, max(10, max_len + 2))

    # Header formatting already applied; ensure borders and row heights
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(final_headers)):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.alignment = CENTER

    # Data rows alignment
    idx_col = final_headers.index('Index') + 1 if 'Index' in final_headers else None
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(final_headers) + 1):
            cell = ws.cell(row=r, column=c)
            header = final_headers[c - 1]
            if header in wrap_cols:
                cell.alignment = WRAP_LEFT
            else:
                if idx_col is not None and c == idx_col:
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        # Row height heuristics based on wrapped text
        wrap_text = []
        for header in wrap_cols:
            if header in final_headers:
                c_idx = final_headers.index(header) + 1
                v = ws.cell(row=r, column=c_idx).value
                if v is not None:
                    wrap_text.append(str(v))
        lines = 1
        for t in wrap_text:
            lines = max(lines, t.count('\n') + 1)
        ws.row_dimensions[r].height = min(409, 15 * lines)

    # Numbering inside specified cells
    def renumber_cell_text(v):
        if v is None:
            return v
        s = str(v).strip()
        if not s:
            return s
        parts = [p.strip() for p in s.split('\n') if p.strip()]
        if len(parts) <= 1:
            return s
        numbered = []
        for i, p in enumerate(parts, start=1):
            # Remove leading bullets like -, *, digits), digits., etc.
            q = p
            # common patterns
            for pref in ['- ', '* ']:
                if q.startswith(pref):
                    q = q[len(pref):].strip()
            # remove N) or N.
            if len(q) > 2 and q[0].isdigit() and (q[1] in [')', '.']):
                q = q[2:].strip()
            numbered.append(f"{i}. {q}")
        return "\n".join(numbered)

    for col_name in ['Test Steps / Procedure', 'Validation / Acceptance Criteria']:
        if col_name in final_headers:
            c_idx = final_headers.index(col_name) + 1
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=c_idx).value
                ws.cell(row=r, column=c_idx, value=renumber_cell_text(v))

    # Data validation on Code Generation (Required / Not)
    if 'Code Generation (Required / Not)' in final_headers and ws.max_row >= 2:
        c_idx = final_headers.index('Code Generation (Required / Not)') + 1
        start_cell = ws.cell(row=2, column=c_idx).coordinate
        end_cell = ws.cell(row=ws.max_row, column=c_idx).coordinate
        dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True, showDropDown=True)
        dv.error = 'Invalid value. Allowed: Required, Blank, Not Required'
        dv.prompt = 'Select one of: Required, Blank, Not Required'
        ws.add_data_validation(dv)
        dv.add(f"{start_cell}:{end_cell}")

    # Safety check: no sheet named 'Data' must remain
    if 'Data' in wb.sheetnames:
        # This can only happen if rename failed; try deleting
        try:
            del wb['Data']
        except Exception as e:
            raise SystemExit(f"Validation failed: residual 'Data' sheet could not be removed: {e}")


def validate_xlsx_binary(path: Path):
    required = [
        '[Content_Types].xml',
        '_rels/.rels',
        'xl/workbook.xml',
    ]
    with zipfile.ZipFile(path, 'r') as zf:
        names = zf.namelist()
        for req in required:
            if req not in names:
                raise SystemExit(f"XLSX validation failed: missing {req}")
        # At least one worksheet and docProps
        if not any(n.startswith('xl/worksheets/') for n in names):
            raise SystemExit("XLSX validation failed: no worksheets found")
        if not any(n.startswith('docProps/') for n in names):
            raise SystemExit("XLSX validation failed: missing docProps")


def main():
    incoming = parse_json_input()
    # Pre-processing merge if file exists
    existing_rows = read_existing_rows_from_testplan(OUTPUT_FILE)
    merged_rows = existing_rows + incoming if existing_rows else incoming

    union_keys = build_union_keys(merged_rows)

    wb, _ = write_base_workbook(merged_rows, union_keys)
    create_meta_sheet(wb, merged_rows)
    normalize_main_sheet_inplace(wb)

    # Final save
    wb.save(OUTPUT_FILE)

    # Validate XLSX as ZIP-based OOXML
    validate_xlsx_binary(OUTPUT_FILE)

    print(f"SUCCESS: Wrote {OUTPUT_FILE} with {len(merged_rows)} rows and {len(union_keys)} columns")


if __name__ == '__main__':
    main()
