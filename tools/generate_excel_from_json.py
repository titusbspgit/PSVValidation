#!/usr/bin/env python3
import argparse
import json
import os
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font
except Exception as e:
    print(f"ERROR: openpyxl not available: {e}")
    raise


def parse_args():
    p = argparse.ArgumentParser(description="Convert JSON test plan to single-sheet Excel (.xlsx)")
    p.add_argument('--input', required=True, help='Path to input JSON file')
    p.add_argument('--output-dir', required=True, help='Directory to place generated Excel file')
    p.add_argument('--ip-name', required=True, help='IP name for filename rule')
    return p.parse_args()


def validate_and_extract(json_obj: Any) -> List[Dict[str, Any]]:
    if json_obj is None:
        raise ValueError('Invalid JSON: empty content')
    if isinstance(json_obj, dict):
        tests = json_obj.get('tests')
        if tests is None or not isinstance(tests, list) or len(tests) == 0:
            raise ValueError('Invalid JSON: missing or empty "tests" array')
        return tests
    elif isinstance(json_obj, list):
        if len(json_obj) == 0:
            raise ValueError('Invalid JSON: empty array')
        return json_obj
    else:
        raise ValueError('Unsupported JSON structure: must be an object with "tests" array or an array of objects')


def normalize_headers(rows: List[Dict[str, Any]]) -> List[str]:
    headers: List[str] = []
    seen = set()
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError('Each test entry must be an object')
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                headers.append(k)
    return headers


def to_cell_value(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    # lists/dicts/others -> stable JSON text preserving insertion order
    try:
        return json.dumps(v, ensure_ascii=False, separators=(',', ':'), sort_keys=False)
    except Exception:
        return str(v)


def autosize_columns(ws):
    from openpyxl.utils import get_column_letter
    max_widths = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for j, cell in enumerate(row, start=1):
            text = "" if cell is None else str(cell)
            max_widths[j] = max(max_widths.get(j, 0), len(text))
    for col_idx, width in max_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(width + 2, 80)


def build_workbook(rows: List[Dict[str, Any]], headers: List[str]) -> 'openpyxl.workbook.workbook.Workbook':
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Header
    ws.append(headers)
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    # Rows
    for row in rows:
        ws.append([to_cell_value(row.get(h, "")) for h in headers])

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-size columns
    autosize_columns(ws)

    return wb


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def ist_now_string() -> str:
    ist = timezone(timedelta(hours=5, minutes=30))
    now = datetime.now(ist)
    return now.strftime('%Y%m%d_%H%M%S')


def main():
    args = parse_args()
    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)

    rows = validate_and_extract(data)
    headers = normalize_headers(rows)

    wb = build_workbook(rows, headers)

    ensure_dir(args.output_dir)
    ts = ist_now_string()
    outfile = os.path.join(args.output_dir, f"{args.ip_name}_TestPlan_{ts}.xlsx")
    wb.save(outfile)
    print(f"Generated: {outfile}")


if __name__ == '__main__':
    main()
