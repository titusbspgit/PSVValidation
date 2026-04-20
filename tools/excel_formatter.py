#!/usr/bin/env python3
import argparse
import glob
import json
import os
from datetime import datetime
import pytz
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def autosize(ws):
    for col_cells in ws.columns:
        length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            v = '' if cell.value is None else str(cell.value)
            if len(v) > length:
                length = len(v)
        width = min(max(10, length + 2), 100)
        ws.column_dimensions[col_letter].width = width


def format_sheet(ws):
    # Bold header, freeze, basic alignment
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
    ws.freeze_panes = 'A2'
    autosize(ws)


def find_latest_xlsx(input_dir: str) -> str:
    files = sorted(glob.glob(os.path.join(input_dir, '*.xlsx')))
    if not files:
        raise FileNotFoundError('No .xlsx files found in input directory')
    return files[-1]


def build_output_name(ip_name: str, tz_name: str, out_dir: str) -> str:
    tz = pytz.timezone(tz_name)
    now = datetime.now(tz)
    out_base = f"{ip_name}_TestPlan_{now.strftime('%Y%m%d')}_{now.strftime('%H%M%S')}_IST.xlsx"
    return os.path.join(out_dir, out_base)


def main():
    ap = argparse.ArgumentParser(description='Format latest Excel and write IST-named copy.')
    ap.add_argument('--input-dir', required=True)
    ap.add_argument('--ip-name', required=True)
    ap.add_argument('--output-dir', required=True)
    ap.add_argument('--sheet-name', default='Data')
    ap.add_argument('--tz', default='Asia/Kolkata')
    args = ap.parse_args()

    latest = find_latest_xlsx(args.input_dir)
    wb = load_workbook(latest)
    # Ensure single-sheet named as requested
    if args.sheet_name in wb.sheetnames:
        ws = wb[args.sheet_name]
    else:
        # If not present, rename the first sheet
        ws = wb.active
        ws.title = args.sheet_name
    format_sheet(ws)

    os.makedirs(args.output_dir, exist_ok=True)
    out_path = build_output_name(args.ip_name, args.tz, args.output_dir)
    wb.save(out_path)
    print(f"Wrote formatted workbook: {out_path}")


if __name__ == '__main__':
    main()
