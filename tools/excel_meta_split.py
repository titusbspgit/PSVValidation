import os
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

INPUT_FILE = os.environ.get('INPUT_FILE')
OUTPUT_FILE = os.environ.get('OUTPUT_FILE')
COMMIT_CHANGES = os.environ.get('COMMIT_CHANGES', 'false').lower() == 'true'
COMMIT_MESSAGE = os.environ.get('COMMIT_MESSAGE', 'TestPlan Generated and Pushed to Github')

# Strict column definitions
META_HEADERS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_HEADERS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)',
]

WRAP_HEADERS = set([
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria',
])

if not INPUT_FILE or not INPUT_FILE.lower().endswith('.xlsx'):
    print('FAILURE: Missing or invalid INPUT_FILE (.xlsx required)')
    sys.exit(1)

if OUTPUT_FILE and not OUTPUT_FILE.lower().endswith('.xlsx'):
    print('FAILURE: OUTPUT_FILE must have .xlsx extension')
    sys.exit(1)

if not os.path.exists(INPUT_FILE):
    print('FAILURE: Input file not found at %s' % INPUT_FILE)
    sys.exit(1)

# Load workbook
wb = load_workbook(INPUT_FILE)

# Identify primary visible sheet (main)
main_ws = None
for name in wb.sheetnames:
    ws = wb[name]
    if ws.sheet_state == 'visible':
        main_ws = ws
        break
if main_ws is None:
    main_ws = wb.active

# STEP 3 — Create META Sheet and copy META columns
if 'Meta_data_sheet' in wb.sheetnames:
    # Remove existing to ensure deterministic result
    ws_old = wb['Meta_data_sheet']
    wb.remove(ws_old)
meta_ws = wb.create_sheet('Meta_data_sheet')

# Build header map for main sheet
header_map = {}
max_col = main_ws.max_column
max_row = main_ws.max_row
for col in range(1, max_col + 1):
    v = main_ws.cell(row=1, column=col).value
    if v is not None and str(v) not in header_map:
        header_map[str(v)] = col

# Copy only existing META headers, preserving order and values exactly
meta_dst_col = 1
for hdr in META_HEADERS:
    if hdr in header_map:
        src_col = header_map[hdr]
        for row in range(1, max_row + 1):
            meta_ws.cell(row=row, column=meta_dst_col).value = main_ws.cell(row=row, column=src_col).value
        meta_dst_col += 1

# STEP 4 — Hide META Sheet (very hidden)
meta_ws.sheet_state = 'veryHidden'

# STEP 5 — Normalize MAIN Sheet (Data)
# Rename main sheet to TestPlan (preserve values)
original_main_title = main_ws.title
main_ws.title = 'TestPlan'

# Create a new sheet with only approved MAIN columns in exact order, values unchanged
temp_ws = wb.create_sheet('TestPlan_tmp')

# Collect columns that exist in main in desired order
existing_main_cols = []
for hdr in MAIN_HEADERS:
    if hdr in header_map:
        existing_main_cols.append((hdr, header_map[hdr]))

# Write headers and data to temp_ws
for dst_c, (hdr, src_c) in enumerate(existing_main_cols, start=1):
    # Header
    temp_ws.cell(row=1, column=dst_c).value = hdr
    # Data rows
    for r in range(2, max_row + 1):
        temp_ws.cell(row=r, column=dst_c).value = main_ws.cell(row=r, column=src_c).value

# Remove old TestPlan and replace with temp
wb.remove(main_ws)
new_main = temp_ws
new_main.title = 'TestPlan'

# STEP 5A — Format MAIN Sheet (TestPlan)
# Header style
header_font = Font(bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=False)
center_top = Alignment(horizontal='center', vertical='top', wrap_text=False)
wrap_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

thin = Side(style='thin', color='000000')
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color='FFECECEC', end_color='FFECECEC', fill_type='solid')

max_row = new_main.max_row
max_col = new_main.max_column

# Freeze panes at B2 (first row and first column frozen)
new_main.freeze_panes = 'B2'

# Build index of headers in new_main
nm_header_map = {}
for c in range(1, max_col + 1):
    v = new_main.cell(row=1, column=c).value
    if v is not None and str(v) not in nm_header_map:
        nm_header_map[str(v)] = c

# Apply header formatting and borders
for c in range(1, max_col + 1):
    cell = new_main.cell(row=1, column=c)
    cell.font = header_font
    cell.alignment = center
    cell.fill = header_fill
    cell.border = border_thin

# Apply data row formatting and borders
for r in range(2, max_row + 1):
    for c in range(1, max_col + 1):
        cell = new_main.cell(row=r, column=c)
        # Alignment rules
        hdr = new_main.cell(row=1, column=c).value
        if hdr == 'Index':
            cell.alignment = center_top
        elif hdr in WRAP_HEADERS:
            cell.alignment = wrap_left_top
        else:
            cell.alignment = left_top
        cell.border = border_thin

# Auto-fit column widths deterministically based on content length
for c in range(1, max_col + 1):
    header = new_main.cell(row=1, column=c).value
    max_len = len(str(header)) if header is not None else 0
    for r in range(2, max_row + 1):
        v = new_main.cell(row=r, column=c).value
        l = len(str(v)) if v is not None else 0
        if l > max_len:
            max_len = l
    # Scale: approx. characters width + padding, clamp to [10, 120]
    width = min(max(10, max_len + 2), 120)
    new_main.column_dimensions[get_column_letter(c)].width = width

# Note: Excel computes row heights on open for wrapped text; openpyxl does not auto-fit row heights.

# Ensure output directory exists
out_dir = os.path.dirname(OUTPUT_FILE)
if out_dir and not os.path.exists(out_dir):
    os.makedirs(out_dir, exist_ok=True)

# Save workbook
wb.save(OUTPUT_FILE)

print('SUCCESS: Processed and saved to %s' % OUTPUT_FILE)

# Optional commit back to repo (for Actions environment)
if COMMIT_CHANGES:
    try:
        import subprocess
        subprocess.run(['git', 'config', 'user.email', 'github-actions[bot]@users.noreply.github.com'], check=True)
        subprocess.run(['git', 'config', 'user.name', 'github-actions[bot]'], check=True)
        subprocess.run(['git', 'add', OUTPUT_FILE], check=True)
        subprocess.run(['git', 'commit', '-m', COMMIT_MESSAGE], check=True)
        subprocess.run(['git', 'push'], check=True)
        print('SUCCESS: Changes committed and pushed')
    except Exception as e:
        print('FAILURE: Commit/push failed: %s' % str(e))
        sys.exit(1)
