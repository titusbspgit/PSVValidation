#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os
import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# JSON payload provided directly (do not modify)
JSON_PAYLOAD = r'''{
  "ip_name": "GPIO",
  "version": "",
  "repo_context": {
    "owner": "titusbspgit",
    "repo": "PSVValidation",
    "branch": "main",
    "base_path": "TestRepo/gpio",
    "source_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio"
  },
  "test_cases": [
    {
      "Index": 1,
      "SS / Module": "GPIO",
      "Feature": "neie — Negative edge interrupt enable",
      "Test Case Name": "test_gpio_negedge_intr_en",
      "Test Description": "Validates falling-edge interrupt behavior per pin across GPIO[8..39], including per-pin raw status latching, group masked status, and proper clearing via per-pin and group clear registers with system interrupt routing.",
      "Speed": "NA",
      "Mode": "Interrupt",
      "Memory Start Offset": "0xA0243ffc",
      "Memory End Offset": "0xA0243ffc",
      "Remarks": "The wait is armed before generating the edge to prevent races. Bounded timeouts avoid infinite waits. Pins are driven high to a known state before configuration, and restored to high within the handler.",
      "Test Steps / Procedure": [
        "Entry: test_case()",
        "If instance 0 is targeted, enable interrupt ID 87; if instance 1 is targeted, enable interrupt ID 88.",
        "Enable system register interrupt routing for the selected GPIO instance via INTR_EN1 (write the corresponding enable bit).",
        "Write 0xFFFFFFFF to memory address 0xA0243ffc to drive all pads high (establish known state).",
        "For each index i from 0 to 31:",
        "- Compute per-pin register address as GPIO_8 + (i * 4).",
        "- WRITE: GPIO_8 + (i*4) with bits to set input mode, enable falling-edge interrupt, and clear raw status.",
        "- Delay: wait_on(10 cycles).",
        "For each index i from 0 to 31:",
        "- Compute mask wr_val = (1 << i).",
        "- WRITE: GPIO_GPIO_INTR_RAW_STCLR1 with wr_val to pre-clear any pending raw status.",
        "- WRITE: GPIO_GP0_INTR1_INTR_EN1 with wr_val to enable interrupt for the current pin only.",
        "- Delay: wait_on(10 cycles).",
        "- Arm wait flag for ISR completion (int_pend = 1).",
        "- Generate falling edge on the current pin:",
        "-- WRITE: 0xA0243ffc = 0xFFFFFFFF (drive high).",
        "-- Delay: wait_on(30 cycles).",
        "-- WRITE: 0xA0243ffc = bitwise NOT of wr_val (drive the selected pin low).",
        "- Start bounded wait for ISR completion with timeout = 5000:",
        "-- While int_pend is asserted and timeout > 0: wait_on(10 cycles), decrement timeout.",
        "- If timeout expires, log a timeout error and increment error count.",
        "Exit: finish(test_err).",
        "ISR Entry: Default_IRQHandler()",
        "Clear wait flag: set int_pend = 0.",
        "Restore pads to high: WRITE 0xA0243ffc = 0xFFFFFFFF.",
        "READ: GPIO_8 + (i*4) into rdata.",
        "Check per-pin input after falling edge: if DIN field indicates high, increment error count.",
        "If the per-pin raw status bit indicates an interrupt:",
        "- READ: GPIO_GP0_INTR1_INTR_STS1 into rdata_grp.",
        "- Verify group status contains the current pin bit; if not, increment error count.",
        "- Clear per-pin raw: WRITE GPIO_8 + (i*4) with bits to keep input mode and clear raw status.",
        "- Clear group raw status: WRITE GPIO_GPIO_INTR_RAW_STCLR1 with (1 << i).",
        "- READ: GPIO_GP0_INTR1_INTR_STS1 and verify it is 0x0; if not, increment error count.",
        "- Clear system raw status for the selected instance via RAW_STCR1 (write the corresponding clear bit).",
        "- Clear the interrupt in the interrupt controller for the selected instance."
      ],
      "Impacted Registers": "INTR_EN1, GPIO_8, GPIO_GPIO_INTR_RAW_STCLR1, GPIO_GP0_INTR1_INTR_EN1, GPIO_GP0_INTR1_INTR_STS1, RAW_STCR1",
      "Validation / Acceptance Criteria": [
        "For each pin, the bounded wait for the interrupt does not expire.",
        "After the falling edge, the input field of the per-pin register indicates low.",
        "The group masked interrupt status shows the current pin’s bit set during service and reads as zero after clearing.",
        "System raw status for the selected instance is cleared during service.",
        "The test passes when the accumulated error count is zero at completion."
      ],
      "Code Generation (Required / Not)": "",
      "Hidden_Test_Case_Name": "test_gpio_negedge_intr_en",
      "Hidden_Test_Description": "test_case() configures GPIO[8..39] for input and negative-edge interrupt (neie), pre-clears raw with write_reg(MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, 1<<i), enables one bit at MIZAR_GPIO_GP0_INTR1_INTR_EN1, generates a falling edge via writes to 0xA0243ffc, and waits for Default_IRQHandler() to clear int_pend. Default_IRQHandler() verifies DIN==0 (rdata & 0x1 == 0), checks raw (rdata & 0x2 != 0), validates group status via read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1), clears per-pin raw by write_reg(MIZAR_GPIO_GP0_GPIO_8 + i*4, (1<<20)|(1<<16)) and group raw via write_reg(MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, 1<<i), rechecks group clear, then clears MIZAR_LSS_SYSREG_RAW_STCR1 and calls GIC_ClearIRQ(87/88).",
      "Hidden_Remarks": "int_pend is set before edge generation to avoid races. Timeouts (5000 iterations with wait_on(10)) bound ISR wait. Pads are driven to 0xFFFFFFFF at start and in ISR. Conditional paths depend on GPIO0/GPIO1 build-time defines for GIC and sysreg routing.",
      "Hidden_Test_Steps_Procedure": [
        "test_case():",
        "#ifdef GPIO0 GIC_EnableIRQ(87); #endif",
        "#ifdef GPIO1 GIC_EnableIRQ(88); #endif",
        "#ifdef GPIO0 write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR); #endif",
        "#ifdef GPIO1 write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR); #endif",
        "write_reg(0xA0243ffc, 0xffffffff);",
        "for (i=0;i<32;i++){ addr1=MIZAR_GPIO_GP0_GPIO_8 + (i*4); write_reg(addr1, (1u<<20)|(1u<<18)|(1u<<16)); wait_on(10); }",
        "for (i=0;i<32;i++){ wr_val=1u<<i; write_reg(MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, wr_val); write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val); wait_on(10); int_pend=1; write_reg(0xA0243ffc,0xffffffff); wait_on(30); write_reg(0xA0243ffc, ~wr_val); unsigned int timeout=5000; while(int_pend && timeout--){ wait_on(10);} if(timeout==0){ printf(\"ERROR: Timeout waiting for GPIO%u negedge interrupt\\n\",(unsigned)(i+8)); test_err++; } }",
        "finish(test_err);",
        "Default_IRQHandler():",
        "unsigned int local_wr=1u<<i; int_pend=0; write_reg(0xA0243ffc,0xffffffff); raddr=MIZAR_GPIO_GP0_GPIO_8 + (i*4); rdata=read_reg(raddr);",
        "if((rdata & 0x1) != 0){ test_err++; }",
        "if((rdata & 0x2) != 0x0){ rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if((rdata_grp & local_wr)==0){ test_err++; } raddr2=MIZAR_GPIO_GP0_GPIO_8 + (i*4); write_reg(raddr2,(1u<<20)|(1u<<16)); write_reg(MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, local_wr); rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if(rdata_grp != 0x0){ test_err++; } #ifdef GPIO0 write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); GIC_ClearIRQ(87); #endif #ifdef GPIO1 write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); GIC_ClearIRQ(88); #endif } else { test_err++; }"
      ],
      "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN1, MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_LSS_SYSREG_RAW_STCR1",
      "Hidden_Validation_Acceptance_Criteria": "No timeout in the wait loop (while(int_pend && timeout--)). In ISR, (rdata & 0x1) == 0 (DIN low) and ((rdata & 0x2) != 0) indicates raw set; (read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) & (1<<i)) != 0 before clear, then equals 0 after clearing via per-pin write and MIZAR_GPIO_GPIO_INTR_RAW_STCLR1. test_err remains 0; finish(test_err) results in pass.",
      "path": "TestRepo/gpio/test_gpio_negedge_intr_en",
      "url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_negedge_intr_en",
      "objective": "Exercise negative-edge interrupt enable on each pin, confirm raw and group interrupt status behavior, and validate clear sequences and interrupt routing.",
      "preconditions": "Interrupt controller configured; build targets GPIO instance 0 or 1; platform supports memory-mapped pad drive at 0xA0243ffc.",
      "expected_results": "Each generated falling edge triggers an interrupt within timeout, input reads low, group status sets and then clears to zero, and system raw status is cleared.",
      "input_parameters": "",
      "dependencies": [
        "lss_sysreg.h",
        "test_define.c",
        "test_common.h",
        "gpio/gpio_def.h",
        "gpio/gpio_offset.h"
      ],
      "tags": []
    },
    {
      "Index": 2,
      "SS / Module": "GPIO",
      "Feature": "peie — Positive edge interrupt enable",
      "Test Case Name": "test_gpio_pedge_all_pads_en",
      "Test Description": "Enables rising-edge interrupts for GPIO[8..39], drives a rising edge per pin, validates group interrupt status assertion and complete clear, and verifies system raw status clearing and interrupt controller servicing.",
      "Speed": "NA",
      "Mode": "Interrupt",
      "Memory Start Offset": "0xA0243ffc",
      "Memory End Offset": "0xA0243ffc",
      "Remarks": "Interrupt wait is armed before generating the rising edge to avoid races. Group interrupt is masked during service, and then re-enabled after clearing.",
      "Test Steps / Procedure": [
        "Entry: test_case()",
        "If instance 0 is targeted, enable interrupt ID 87; if instance 1 is targeted, enable interrupt ID 88.",
        "Enable system register interrupt routing for the selected instance via INTR_EN1 (write the corresponding enable bit).",
        "For i = 0..31: WRITE per-pin configuration at GPIO_8 + (i*4) to enable rising-edge interrupt.",
        "Delay: wait_on(10 cycles).",
        "Configure GPIOs 8–39 for input mode: WRITE GPIO_GPIO_IO_CTRL_GROUP1 = 0x000000FF; WRITE GPIO_GPIO_IO_CTRL_GROUP2 = 0x000000FF; WRITE GPIO_GPIO_IO_CTRL_GROUP3 = 0x000000FF; WRITE GPIO_GPIO_IO_CTRL_GROUP4 = 0x000000FF.",
        "Delay: wait_on(10 cycles).",
        "Enable all 32 group interrupts: WRITE GPIO_GP0_INTR1_INTR_EN1 = 0xFFFFFFFF.",
        "For each i from 0 to 31:",
        "- Prepare low level: WRITE 0xA0243ffc = 0x00000000.",
        "- Delay: wait_on(10 cycles).",
        "- Arm wait flag for ISR completion (int_pend = 1).",
        "- Generate a rising edge: WRITE 0xA0243ffc = 0xFFFFFFFF.",
        "- Start bounded wait for ISR completion with timeout = 2000:",
        "-- While int_pend is asserted and timeout > 0: wait_on(10 cycles), decrement timeout.",
        "- If timeout expires, log error and increment error count; break loop.",
        "- Optionally drive low again for next iteration: WRITE 0xA0243ffc = 0x00000000; wait_on(10 cycles).",
        "Exit: finish(test_err).",
        "ISR Entry: Default_IRQHandler()",
        "Compute wr_val = (1 << i) and clear wait flag: int_pend = 0.",
        "READ group masked status: READ GPIO_GP0_INTR1_INTR_STS1 into rdata_grp.",
        "Mask group interrupt during service: WRITE GPIO_GP0_INTR1_INTR_EN1 = 0x00000000.",
        "If group masked status is non-zero, continue; else log error and increment error count.",
        "Clear per-pin raw: for j = 0..31: WRITE GPIO_8 + (j*4) with bit to clear raw status.",
        "Delay: wait_on(2 cycles).",
        "Verify group clear: READ GPIO_GP0_INTR1_INTR_STS1 and expect 0x0; otherwise increment error count.",
        "For the selected instance: WRITE RAW_STCR1 with the corresponding bit to clear system raw; READ RAW_STCR1 back and verify the bit is cleared; otherwise increment error count.",
        "Re-enable group interrupt: WRITE GPIO_GP0_INTR1_INTR_EN1 = 0xFFFFFFFF.",
        "Clear the interrupt in the interrupt controller for the selected instance."
      ],
      "Impacted Registers": "INTR_EN1, GPIO_8, GPIO_GPIO_IO_CTRL_GROUP1, GPIO_GPIO_IO_CTRL_GROUP2, GPIO_GPIO_IO_CTRL_GROUP3, GPIO_GPIO_IO_CTRL_GROUP4, GPIO_GP0_INTR1_INTR_EN1, GPIO_GP0_INTR1_INTR_STS1, RAW_STCR1",
      "Validation / Acceptance Criteria": [
        "For each pin, the interrupt arrives before the timeout expires.",
        "During service, the group masked status is non-zero and reads back as zero after per-pin raw clears.",
        "System raw status reads as cleared after writing the clear bit.",
        "The test passes when the accumulated error count is zero at completion."
      ],
      "Code Generation (Required / Not)": "",
      "Hidden_Test_Case_Name": "test_gpio_pedge_all_pads_en",
      "Hidden_Test_Description": "test_case() enables posedge (peie) on GPIO[8..39] by write_reg(MIZAR_GPIO_GP0_GPIO_8 + i*4, 0x00020000), sets DOE via group IO_CTRL regs, enables all via write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0xFFFFFFFF), and for i=0..31 drives 0xA0243ffc low then high to create a rising edge while int_pend is armed. Default_IRQHandler() reads MIZAR_GPIO_GP0_INTR1_INTR_STS1, masks group enable, clears all per-pin raw via writes (0x00010000) to MIZAR_GPIO_GP0_GPIO_8 + j*4, confirms group clear, clears system raw via MIZAR_LSS_SYSREG_RAW_STCR1 and reads back, then re-enables group and clears GIC IRQ.",
      "Hidden_Remarks": "Group masked during service (write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000)) to prevent nested interrupts. Rising edge generated by writing 0xA0243ffc from 0x00000000 to 0xFFFFFFFF while int_pend is 1. Timeout loop (2000 with wait_on(10)).",
      "Hidden_Test_Steps_Procedure": [
        "test_case():",
        "#ifdef GPIO0 GIC_EnableIRQ(87); #endif",
        "#ifdef GPIO1 GIC_EnableIRQ(88); #endif",
        "#ifdef GPIO0 write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR); #endif",
        "#ifdef GPIO1 write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR); #endif",
        "for(i=0;i<32;i++){ write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00020000); }",
        "wait_on(10);",
        "write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, 0x000000FF);",
        "write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, 0x000000FF);",
        "write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, 0x000000FF);",
        "write_reg(MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, 0x000000FF);",
        "wait_on(10);",
        "write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0xFFFFFFFF);",
        "for(i=0;i<32;i++){ write_reg(0xA0243ffc,0x00000000); wait_on(10); int_pend=1; write_reg(0xA0243ffc,0xFFFFFFFF); int timeout=2000; while((int_pend==1) && (--timeout>0)){ wait_on(10); } if(timeout==0){ printf(\"ERROR: Timeout waiting for GPIO IRQ at i=%u\\n\", i); test_err++; break; } write_reg(0xA0243ffc,0x00000000); wait_on(10); }",
        "finish(test_err);",
        "Default_IRQHandler():",
        "unsigned int wr_val = 1 << i; int_pend=0;",
        "rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1);",
        "write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000);",
        "if((rdata_grp & 0xffffffff) != 0) { /* OK */ } else { printf(\"ERROR: Group Interrupt not occured\\n\"); test_err++; }",
        "for(j=0;j<32;j++){ write_reg(MIZAR_GPIO_GP0_GPIO_8 + (j*4), 0x00010000); }",
        "wait_on(2);",
        "rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if(rdata_grp == 0x0){ /* OK */ } else { printf(\"ERROR : Group Interrupt clear failed: Interrupt value:%x\\n\", rdata_grp); test_err++; }",
        "#ifdef GPIO0 write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR) != 0){ printf(\"sysreg status not cleared : %0x\\n\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++; } #endif",
        "#ifdef GPIO1 write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); rdata = read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR) != 0){ printf(\"sysreg status not cleared : %0x\\n\", MIZAR_LSS_SYSREG_RAW_STCR1); test_err++; } #endif",
        "write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0xFFFFFFFF);",
        "#ifdef GPIO0 GIC_ClearIRQ(87); #endif",
        "#ifdef GPIO1 GIC_ClearIRQ(88); #endif"
      ],
      "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN1, MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_LSS_SYSREG_RAW_STCR1",
      "Hidden_Validation_Acceptance_Criteria": "For each pin: timeout loop terminates via ISR (int_pend cleared). Group status read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1) is non-zero during service and becomes 0 after writes to MIZAR_GPIO_GP0_GPIO_8+(j*4) with 0x00010000 and wait_on(2). System raw clear is verified by reading back MIZAR_LSS_SYSREG_RAW_STCR1 after writing the corresponding clear bit. test_err remains 0; finish(test_err) indicates pass.",
      "path": "TestRepo/gpio/test_gpio_pedge_all_pads_en",
      "url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/test_gpio_pedge_all_pads_en",
      "objective": "Exercise positive-edge interrupt enable for all pins, verify group interrupt behavior and clear sequence, and confirm system raw status handling.",
      "preconditions": "Interrupt controller configured; build targets GPIO instance 0 or 1; platform supports memory-mapped pad drive at 0xA0243ffc.",
      "expected_results": "Each rising edge triggers an interrupt within timeout; group status asserts and de-asserts after clear; system raw status reads as cleared.",
      "input_parameters": "",
      "dependencies": [
        "lss_sysreg.h",
        "test_define.c",
        "test_common.h",
        "gpio/gpio_def.h",
        "gpio/gpio_offset.h"
      ],
      "tags": []
    },
    {
      "Index": 3,
      "SS / Module": "GPIO",
      "Feature": "AHB 32-bit register interface",
      "Test Case Name": "gpio_reg_wr_rd_test",
      "Test Description": "Performs default-value reads and masked write/read checks across GPIO control and status registers using defined address, mask, and skip tables.",
      "Speed": "NA",
      "Mode": "NA",
      "Memory Start Offset": "NA",
      "Memory End Offset": "NA",
      "Remarks": "Default-value comparison masks out bit[0] before compare. Not-readable or not-writable registers are skipped based on masks. Additional skip tables exclude VRRW and certain group registers as documented in the source.",
      "Test Steps / Procedure": [
        "Entry: test_case()",
        "Invoke default value check routine.",
        "- For each entry in the address table:",
        "-- If the reset-skip table marks the entry, skip reading.",
        "-- If the read mask is zero, skip reading.",
        "-- READ: the register address from the table.",
        "-- Mask out bit[0] and compare the value with the default table entry; increment default-fail count on mismatch.",
        "Invoke masked write/read routine.",
        "- For each of six data patterns (0xFFFFFFFF, 0xAAAAAAAA, 0x55555555, 0xF5F5F5F5, 0xA5A5A5A5, 0xFFFF0000):",
        "-- For each entry in the address table:",
        "--- If the write-skip table marks the entry, skip write.",
        "--- If the write mask is zero, skip write.",
        "--- WRITE: masked pattern to the register.",
        "-- For each entry in the address table:",
        "--- If skip table marks the entry, or write mask is zero, or read mask is zero, skip readback.",
        "--- READ: masked value from the register.",
        "--- Compute expected value from the pattern, read mask, write mask, and default value; increment write-fail count if mismatch.",
        "Complete: Report PASS if both default-fail and write-fail counts are zero; otherwise report FAIL."
      ],
      "Impacted Registers": "GPIO_8, GPIO_GP0_INTR1_INTR_EN1, GPIO_GP0_INTR1_INTR_STS1, GPIO_GP0_INTR2_INTR_EN1, GPIO_GP0_INTR2_INTR_STS1, GPIO_GPIO_INTR_RAW_STCLR1, GPIO_GPIO_IO_CTRL_GROUP1, GPIO_GPIO_IO_CTRL_GROUP2, GPIO_GPIO_IO_CTRL_GROUP3, GPIO_GPIO_IO_CTRL_GROUP4, GPIO_GPIO_DOUT_GROUP1, GPIO_GPIO_DOUT_GROUP2, GPIO_GPIO_DOUT_GROUP3, GPIO_GPIO_DOUT_GROUP4, GPIO_GPIO_DIN_GROUP1, GPIO_GPIO_DIN_GROUP2, GPIO_GPIO_GPIO_DIN_GROUP3, GPIO_GPIO_DIN_GROUP4",
      "Validation / Acceptance Criteria": [
        "Default-value phase: for each non-skipped, readable register, the masked read equals the corresponding default table entry.",
        "Write/read phase: for each non-skipped register, masked readback equals the expected value derived from pattern, read mask, write mask, and default value.",
        "Test passes when both default-fail count and write-fail count are zero at completion."
      ],
      "Code Generation (Required / Not)": "",
      "Hidden_Test_Case_Name": "gpio_reg_wr_rd_test",
      "Hidden_Test_Description": "test_case() calls chk_rst_val() and chk_rd_wr(). chk_rst_val(): iterates i=0..CNT-1 over addr_array[], skips per skip_rst_array or read_mask==0, then reads via read_reg(addr), masks with 0xfffffffe, and compares to default_value_array[i], incrementing def_fail_cnt on mismatch. chk_rd_wr(): for each pattern in {0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}, writes masked values to each addr if write_mask!=0 and not skipped (skip_array==0), then reads back masked values where allowed, computing exp_val=((data_wr & read_mask & write_mask) | ((~write_mask) & read_mask & default_value)) and incrementing wr_fail_cnt on mismatch. finish(1) if any fail counters > 0 else finish(0).",
      "Hidden_Remarks": "Comments document skipping VRRW registers and that, when reading default values, DIN may become 1 if not forced; forcing DIN low affects bit-level selection and can alter expected reads.",
      "Hidden_Test_Steps_Procedure": [
        "test_case():",
        "chk_rst_val();",
        "chk_rd_wr();",
        "if(def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0);",
        "chk_rst_val():",
        "for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_rst_array[i]==1){ continue; } if(read_mask_array[i]==0){ continue; } data_rd=read_reg(addr); data=(data_rd & 0xfffffffe); if(data==default_value_array[i]){/*pass*/} else { def_fail_cnt++; printf(\"RST : Failed Default value mismatch Addr :0x%x Expected : 0x%x\\tRead_data : 0x%x\\tDATA : 0x%x\\n\",addr,default_value_array[i],data,data_rd);} }",
        "chk_rd_wr():",
        "unsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000};",
        "for(j=0;j<6;j++){ data_wr=chk_val[j]; for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ continue; } if(write_mask_array[i]==0){ continue; } write_reg(addr,(data_wr & write_mask_array[i])); } for(i=0;i<CNT;i++){ addr=addr_array[i]; if(skip_array[i]==1){ continue; } if(write_mask_array[i]==0){ continue; } if(read_mask_array[i]==0){ continue; } data_rd=(read_reg(addr) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd == exp_val){ /*pass*/ } else { wr_fail_cnt++; printf(\"Read_write : Failed : Write Read mismatch For Address %x, Expected value=0x%x\\tRead value=0x%x\\n\",addr,exp_val ,data_rd); } } }"
      ],
      "Hidden_Impacted_Registers": "MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, MIZAR_GPIO_GP0_GPIO_11, MIZAR_GPIO_GP0_GPIO_12, MIZAR_GPIO_GP0_GPIO_13, MIZAR_GPIO_GP0_GPIO_14, MIZAR_GPIO_GP0_GPIO_15, MIZAR_GPIO_GP0_GPIO_16, MIZAR_GPIO_GP0_GPIO_17, MIZAR_GPIO_GP0_GPIO_18, MIZAR_GPIO_GP0_GPIO_19, MIZAR_GPIO_GP0_GPIO_20, MIZAR_GPIO_GP0_GPIO_21, MIZAR_GPIO_GP0_GPIO_22, MIZAR_GPIO_GP0_GPIO_23, MIZAR_GPIO_GP0_GPIO_24, MIZAR_GPIO_GP0_GPIO_25, MIZAR_GPIO_GP0_GPIO_26, MIZAR_GPIO_GP0_GPIO_27, MIZAR_GPIO_GP0_GPIO_28, MIZAR_GPIO_GP0_GPIO_29, MIZAR_GPIO_GP0_GPIO_30, MIZAR_GPIO_GP0_GPIO_31, MIZAR_GPIO_GP0_GPIO_32, MIZAR_GPIO_GP0_GPIO_33, MIZAR_GPIO_GP0_GPIO_34, MIZAR_GPIO_GP0_GPIO_35, MIZAR_GPIO_GP0_GPIO_36, MIZAR_GPIO_GP0_GPIO_37, MIZAR_GPIO_GP0_GPIO_38, MIZAR_GPIO_GP0_GPIO_39, MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_GPIO_GP0_INTR2_INTR_EN1, MIZAR_GPIO_GP0_INTR2_INTR_STS1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, MIZAR_GPIO_GPIO_DOUT_GROUP1, MIZAR_GPIO_GPIO_DOUT_GROUP2, MIZAR_GPIO_GPIO_DOUT_GROUP3, MIZAR_GPIO_GPIO_DOUT_GROUP4, MIZAR_GPIO_GPIO_DIN_GROUP1, MIZAR_GPIO_GPIO_DIN_GROUP2, MIZAR_GPIO_GPIO_DIN_GROUP3, MIZAR_GPIO_GPIO_DIN_GROUP4"
    }
  ],
  "META_DATA_NOTE": "Hidden_* fields preserve raw macro names and function identifiers. Main column fields are rewritten and macro-to-register name replacement is applied where mapping was available. No assumptions beyond code and comments were introduced."
}'''

# Constants per Stage1 formatting rules
META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_COLS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)',
]

OUTPUT_DIR = os.path.join('Test_Output', 'GPIO', 'TestPlan')


def _json_to_rows():
    obj = json.loads(JSON_PAYLOAD)
    # If top-level contains test_cases list, use it; otherwise, use array/object as provided
    if isinstance(obj, dict) and 'test_cases' in obj and isinstance(obj['test_cases'], list):
        rows = obj['test_cases']
        ip_name = obj.get('ip_name', 'OUTPUT')
    elif isinstance(obj, list):
        rows = obj
        ip_name = 'OUTPUT'
    else:
        rows = [obj]
        ip_name = obj.get('ip_name', 'OUTPUT') if isinstance(obj, dict) else 'OUTPUT'
    return ip_name, rows


def _normalize_schema(rows):
    seen = []
    seen_set = set()
    for r in rows:
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    # Ensure META and MAIN keys appear if missing (as empty columns)
    for k in META_COLS + MAIN_COLS:
        if k not in seen_set:
            seen.append(k)
            seen_set.add(k)
    return seen


def _cell_value(v):
    if isinstance(v, (dict, list)):
        # Preserve exact JSON textual form for complex types
        return json.dumps(v, ensure_ascii=False)
    return v


def _write_data_sheet(wb, headers, rows):
    ws = wb.active
    ws.title = 'Data'
    # Write header
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Write rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            val = ''
            if isinstance(row, dict) and h in row:
                val = _cell_value(row[h])
            ws.cell(row=r_idx, column=c_idx, value=val)
    ws.freeze_panes = 'A2'
    # Basic width autofit approximation
    _autofit_columns(ws)
    return ws


def _create_meta_sheet(wb, headers, rows):
    ws = wb.create_sheet('Meta_data_sheet')
    # Build meta headers present
    for c, h in enumerate(META_COLS, start=1):
        ws.cell(row=1, column=c, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(META_COLS, start=1):
            val = ''
            if isinstance(row, dict) and h in row:
                val = _cell_value(row[h])
            ws.cell(row=r_idx, column=c_idx, value=val)
    # Mark very hidden
    ws.sheet_state = 'veryHidden'
    return ws


def _build_testplan_sheet(wb, data_ws, rows):
    ws = wb.create_sheet('TestPlan_tmp')
    # Write headers in the strict order
    for c, h in enumerate(MAIN_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='FFEFEFEF', end_color='FFEFEFEF', fill_type='solid')
    # Write data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(MAIN_COLS, start=1):
            val = ''
            if isinstance(row, dict) and h in row:
                val = _cell_value(row[h])
            ws.cell(row=r_idx, column=c_idx, value=val)
    # Apply formatting to TestPlan only
    _format_testplan(ws)
    # Remove original Data sheet and rename tmp to TestPlan
    wb.remove(data_ws)
    ws.title = 'TestPlan'
    return ws


def _format_testplan(ws):
    # Wrap text for specified columns
    wrap_cols = set([
        'Test Description', 'Remarks', 'Test Steps / Procedure', 'Validation / Acceptance Criteria'
    ])
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Apply alignment and borders
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header formatting already bold+center; ensure borders
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data rows
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            header = ws.cell(row=1, column=c).value
            if header in wrap_cols:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif header == 'Index':
                cell.alignment = Alignment(horizontal='center', vertical='top')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top')
            cell.border = border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Autofit columns (approximate) and row heights
    _autofit_columns(ws)
    _autoheight_rows(ws, wrap_cols)


def _autofit_columns(ws):
    # Approximate best-fit based on text length
    max_len = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            s = str(val) if val is not None else ''
            max_len[c] = max(max_len.get(c, 0), len(s))
    for c, l in max_len.items():
        width = min(max(10, l + 2), 80)  # clamp
        ws.column_dimensions[get_column_letter(c)].width = width


def _autoheight_rows(ws, wrap_cols):
    # Heuristic: estimate lines by length vs column width
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_widths = {c: ws.column_dimensions[get_column_letter(c)].width or 10 for c in range(1, ws.max_column + 1)}
    base_height = 15
    for r in range(2, ws.max_row + 1):
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            h = header[c - 1]
            val = ws.cell(row=r, column=c).value
            if h in wrap_cols and val is not None:
                s = str(val)
                approx_chars_per_line = max(1, int(col_widths[c]))
                lines = 0
                for part in s.split('\n'):
                    lines += max(1, (len(part) // approx_chars_per_line) + 1)
                if lines > max_lines:
                    max_lines = lines
        ws.row_dimensions[r].height = min(300, base_height * max_lines)


def main():
    ip_name, rows = _json_to_rows()
    headers = _normalize_schema(rows)

    wb = Workbook()
    data_ws = _write_data_sheet(wb, headers, rows)
    _create_meta_sheet(wb, headers, rows)
    _build_testplan_sheet(wb, data_ws, rows)

    # IST timestamp for filename
    tz = pytz.timezone('Asia/Kolkata')
    now = datetime.datetime.now(tz)
    fname = f"{ip_name}_TestPlan_{now:%Y%m%d_%H%M%S}.xlsx"

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, fname)

    wb.save(out_path)
    print(out_path)

if __name__ == '__main__':
    main()
