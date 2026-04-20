#!/usr/bin/env python3
import sys
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

META_COLUMNS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_COLUMNS_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

WRAP_COLUMNS = set([
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria',
])


def parse_args():
    p = argparse.ArgumentParser(description='Reorganize Excel TestPlan: move META cols to very-hidden sheet and format MAIN sheet.')
    p.add_argument('--input', required=True, help='Path to source .xlsx file')
    p.add_argument('--output', required=True, help='Path to save modified .xlsx file')
    return p.parse_args()


def get_first_visible_sheet(wb):
    for ws in wb.worksheets:
        if ws.sheet_state == 'visible':
            return ws
    return wb.active


def build_header_index_map(ws):
    header_map = {}
    for j, cell in enumerate(ws[1], start=1):
        v = cell.value
        if v is None:
            continue
        header_map[str(v)] = j
    return header_map


def copy_meta_columns(main_ws, wb):
    header_map = build_header_index_map(main_ws)
    missing = [c for c in META_COLUMNS if c not in header_map]
    if missing:
        sys.stderr.write('Missing META columns: ' + ', '.join(missing) + '\n')
        sys.exit(2)

    # Determine observed left-to-right order on main sheet
    ordered_meta = sorted(META_COLUMNS, key=lambda c: header_map[c])

    # Create or clear Meta_data_sheet
    if 'Meta_data_sheet' in wb.sheetnames:
        meta_ws = wb['Meta_data_sheet']
        # clear existing content
        if meta_ws.max_row:
            meta_ws.delete_rows(1, meta_ws.max_row)
        if meta_ws.max_column:
            meta_ws.delete_cols(1, meta_ws.max_column)
    else:
        meta_ws = wb.create_sheet('Meta_data_sheet')

    # Copy headers and values exactly
    for new_col_idx, col_name in enumerate(ordered_meta, start=1):
        meta_ws.cell(row=1, column=new_col_idx, value=col_name)
        src_col_idx = header_map[col_name]
        for r in range(2, main_ws.max_row + 1):
            meta_ws.cell(row=r, column=new_col_idx, value=main_ws.cell(row=r, column=src_col_idx).value)

    # Very hide sheet
    meta_ws.sheet_state = 'veryHidden'

    return ordered_meta


def remove_columns_by_names(ws, col_names):
    # Build fresh map
    header_map = build_header_index_map(ws)
    # Sort indices in descending order to delete safely
    indices = sorted([header_map[name] for name in col_names], reverse=True)
    for idx in indices:
        ws.delete_cols(idx)


def reorder_and_filter_main(ws, wb):
    # After META deletion, rebuild the header map
    header_map = build_header_index_map(ws)
    missing_main = [c for c in MAIN_COLUMNS_ORDER if c not in header_map]
    if missing_main:
        sys.stderr.write('Missing MAIN columns: ' + ', '.join(missing_main) + '\n')
        sys.exit(3)

    max_row = ws.max_row

    # Create a temporary sheet and copy only allowed columns in the exact order
    tmp_name = 'TestPlan_tmp'
    if tmp_name in wb.sheetnames:
        del wb[tmp_name]
    tmp_ws = wb.create_sheet(tmp_name)

    # Write headers
    for j, col_name in enumerate(MAIN_COLUMNS_ORDER, start=1):
        tmp_ws.cell(row=1, column=j, value=col_name)

    for r in range(2, max_row + 1):
        for j, col_name in enumerate(MAIN_COLUMNS_ORDER, start=1):
            src_col = header_map[col_name]
            tmp_ws.cell(row=r, column=j, value=ws.cell(row=r, column=src_col).value)

    # Delete original ws content by replacing sheet entirely
    title = ws.title
    del wb[title]
    tmp_ws.title = title

    return wb[title]


def apply_formatting(ws):
    # Header formatting
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin')
    medium = Side(style='medium')

    ncols = len(MAIN_COLUMNS_ORDER)
    nrows = ws.max_row

    # Apply header styles and medium bottom border
    for j in range(1, ncols + 1):
        c = ws.cell(row=1, column=j)
        c.font = header_font
        c.alignment = header_align
        c.border = Border(left=thin, right=thin, top=thin, bottom=medium)

    # Data rows formatting
    for r in range(2, nrows + 1):
        for j in range(1, ncols + 1):
            hdr = ws.cell(row=1, column=j).value
            h_align = 'left'
            if hdr == 'Index':
                h_align = 'center'
            elif hdr in ('Memory Start Offset', 'Memory End Offset'):
                h_align = 'right'
            # Wrap for specific columns
            wrap = True if hdr in WRAP_COLUMNS else False
            c = ws.cell(row=r, column=j)
            c.alignment = Alignment(horizontal=h_align, vertical='top', wrap_text=wrap)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set wrap on header of wrap columns as well (optional, not required)
    for j in range(1, ncols + 1):
        hdr = ws.cell(row=1, column=j).value
        if hdr in WRAP_COLUMNS:
            c = ws.cell(row=1, column=j)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Autofit columns based on content length (approximation)
    for j in range(1, ncols + 1):
        max_len = 0
        for r in range(1, nrows + 1):
            val = ws.cell(row=r, column=j).value
            if val is None:
                continue
            s = str(val)
            # account for multi-line
            for line in s.split('\n'):
                if len(line) > max_len:
                    max_len = len(line)
        adjusted_width = min(max(10, max_len + 2), 80)
        ws.column_dimensions[get_column_letter(j)].width = adjusted_width

    # Ensure row heights are auto (Excel will auto-fit on open when wrap is set and height is default)
    for r in range(1, nrows + 1):
        ws.row_dimensions[r].height = None


def main():
    args = parse_args()

    if not args.input.lower().endswith('.xlsx'):
        sys.stderr.write('Input file must be .xlsx\n')
        sys.exit(1)

    wb = load_workbook(args.input)

    # Identify the primary visible sheet
    main_ws = get_first_visible_sheet(wb)

    # Create META sheet and copy columns
    ordered_meta = copy_meta_columns(main_ws, wb)

    # Hide META sheet is done inside copy function

    # Rename main sheet to TestPlan
    main_ws.title = 'TestPlan'

    # Remove META columns from TestPlan
    remove_columns_by_names(wb['TestPlan'], ordered_meta)

    # Keep and order only allowed MAIN columns
    testplan_ws = reorder_and_filter_main(wb['TestPlan'], wb)

    # Apply formatting ONLY to TestPlan
    apply_formatting(testplan_ws)

    # Save the workbook to output
    wb.save(args.output)


if __name__ == '__main__':
    main()
