#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate USB TestPlan Excel (.xlsx) from embedded JSON.
- Creates two sheets: TestPlan (visible) and MetaData (veryHidden)
- Applies formatting (bold blue header, white font, wrap text, freeze first row)
- Saves to Test_Output/USB/TestPlan/USB_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx using IST time
- Writes the relative output path to generated_excel_path.txt (for the workflow to commit)
"""
import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Embedded input parameters
IP_NAME = "USB"
OWNER = "titusbspgit"
REPO = "PSVValidation"
SOURCE_PATH = "TestRepo/usb"
OUTPUT_DIR = os.path.join("Test_Output", "USB", "TestPlan")

# Exact JSON data (do not modify to preserve content)
JSON_DATA = [
  {
    "Index": "1",
    "SS / Module": "USB",
    "Feature": "DWC USB3 Memory Map [Source: usb_autoreg(RegSpec).csv]",
    "Test Case Name": "basic_test",
    "Test Description": "Smoke test that exercises DWC USB3 memory map access by performing basic MMIO reads and a write-readback sequence on USB registers, then exits without explicit value checks.",
    "Meta Test Description": "C function test_case prints a banner, performs MMIO reads from 0xA0000000 and 0xA001706C using read_reg(), logs the values, writes 0xDEADBEEF to 0xA0240000 via write_reg(), reads back from 0xA0240000 using read_reg(), logs the value, and calls finish(0). No conditional checks or assertions are implemented.",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "0xA0000000",
    "Memory End Offset": "0xA0240000",
    "Remarks": "NA",
    "Test Steps / Procedure": "1) Start the USB basic smoke test and initialize console logging.\n2) Read from a USB register in the DWC USB3 memory map and log the value.\n3) Read from another USB register in the DWC USB3 memory map and log the value.\n4) Write a known test pattern to a USB register in the DWC USB3 memory map and read it back to confirm access path functionality (observational logging only).\n5) Finish the test without explicit pass/fail comparison.",
    "Meta Test Steps / Procedure": "- printf(\"[C-Programme] Hello world\\n\");\n- int rd_data = 0;\n- rd_data = read_reg(0xA0000000); printf(\"THE READ DATA is %X\\n\", rd_data);\n- rd_data = read_reg(0xA001706C); printf(\"THE READ DATA1 is %X\\n\", rd_data);\n- write_reg(0xA0240000, 0xDEADBEEF);\n- rd_data = read_reg(0xA0240000); printf(\"THE READ DATA2 is %X\\n\", rd_data);\n- finish(0);",
    "Impacted Registers": "DWC USB3 Memory Map (register at 0xA0000000); DWC USB3 Memory Map (register at 0xA001706C); DWC USB3 Memory Map (register at 0xA0240000)",
    "Meta Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "Pass if the MMIO read/write operations to DWC USB3 memory map registers complete without exceptions and the program reaches finish(0). No explicit value comparison is implemented.",
    "Meta Validation / Acceptance Criteria": "No asserts or conditional checks present. Successful termination via finish(0) after executing read_reg/write_reg calls implies pass. Any fault or failure to reach finish(0) implies fail.",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>\n#include <stdlib.h>\n#include<usb/usb_def.h>\n#include<usb/usb_offsets.h>",
    "Meta Macros": "NA",
    "Meta Arrays": "NA"
  },
  {
    "Index": "2",
    "SS / Module": "USB",
    "Feature": "USB Host Enumeration (Low Speed)",
    "Test Case Name": "usb_host_enumeration_ls",
    "Test Description": "Configures the DWC USB3/xHCI host controller, PHY and port, sets up event/command rings and device/context structures, enables interrupts and runs the host, performs port reset, issues Enable Slot and Address Device commands, then executes standard EP0 control transfers (GET descriptors and SET configuration) to enumerate a low-speed device, and completes upon observing event completions.",
    "Meta Test Description": "test_case initializes NIC and enables all IRQs. It programs controller global registers (GCTL, GFLADJ, GUCTL), PIPE (GUSB3PIPECTL at BASE+0xc2c0) and PHY control (BASE+0xc200). It reads HCSPARAMS1/SUPTPRT2_DW2/SUPTPRT3_DW2 and PORTSC, enables wake bits (WCE/WDE/WOE) and writes PORTSC with 0xe0002a0. It reads DBOFF, sets up the Event Ring Segment Table at Event_Ring_Segment_Table pointing to Default_Event_Ring_Array with size 0x30. It reads HCSPARAMS2 and PAGESIZE. It programs Scratchpad_Buffer_Array (SCRATCHPAD0/1) and Device_Context_Base_Address_Array entries to point to scratchpads and Device_Context_Array. It sets the command ring (CRCR_LO=Default_Command_Ring+1, CRCR_HI=0), configures MaxSlotsEn (CONFIG=0x10 then 0x110), sets DCBAAP, ERSTSZ=1, ERDP to Default_Event_Ring_Array, ERSTBA to Event_Ring_Segment_Table, IMOD=0, IMAN=2, USBCMD=0x4, enables system-level interrupt (SYSREG_INTR_EN0=0x80000000), and starts the controller (USBCMD=0x5). With int_pend set to 1, it waits until cleared (intended via ISR) before acknowledging USBSTS and re-enabling IMAN; it reads PORTSC (expecting connect). It advances ERDP, performs port reset via PORTSC writes (0xe0006f1 then 0xe220200), acknowledges USBSTS/IMAN, and waits on int_pend again. It issues an Enable Slot command TRB at Default_Command_Ring and advances ERDP. It then writes PORTSC=0xe200e01, rings the doorbell (DB=0) and waits for completion (int_pend). It builds the Default_Input_Context (Input Control Flags=0x3; Slot/EP0 contexts with values including EP0_TR_Dequeue_Pointer|1 and interval=0x08). It issues Address Device with BSR=1 (0x01002e01), acknowledges, advances ERDP, rings DB, waits for completion, reads event completion at Default_Event_Ring_Array+0x30, then reuses the input context and issues Address Device with BSR=0 (0x01002c01); acknowledges, advances ERDP, rings DB, waits and reads completion at Default_Event_Ring_Array+0x40. It calls enumeration(), which programs EP0 transfer ring TRBs to perform: GET Device Descriptor (bmRequestType=0x80, bRequest=GET_DESCRIPTOR 0x06, wLength=0x12) with data and status stages; GET Configuration Descriptor (0x02000680, length 0x9) with data and status; SET Configuration (SETUP 0x00010900, status 0x00011023); GET Configuration again sequences with lengths 0x8 and then 0x18; it writes a run register at USB_BASE+0x484=0x1 and polls an event completion at Default_Event_Ring_Array+0x110 until non-zero. After enumeration returns, the test prints a message and calls finish(0). Default_IRQHandler reads SYSREG MSK_STS0 and RAW_STCR0, and if bit 31 is set, writes IMAN=0x1 and clears RAW_STCR0 bit 31; it then clears GIC IRQ 84.",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Requires pre-initialized event ring, scratchpad, device context, and transfer ring buffer addresses; controller interrupts must be enabled to progress events.",
    "Test Steps / Procedure": "1) Initialize the NIC and enable all required interrupts for the USB controller and GIC.\n2) Configure controller and PHY: program global control, frame length adjust, uctl, PIPE, and PHY control; enable wake and port features.\n3) Initialize xHCI data structures: program Event Ring Segment Table, ERDP, ERST size/base, scratchpads, Device Context Base Address Array, and Command Ring.\n4) Configure host parameters: set MaxSlots and CIE, program DCBAAP, and start the controller with interrupts enabled.\n5) Handle initial events: wait for and acknowledge status; confirm port connection; perform a port reset sequence and acknowledge statuses.\n6) Issue Enable Slot command and process the resulting event; advance the event ring dequeue pointer as needed.\n7) Build the input context for Slot/EP0 and issue Address Device (first with BSR=1, then with BSR=0), ringing the doorbell and handling completions between commands.\n8) Perform EP0 control transfers for enumeration: GET Device Descriptor, GET Configuration Descriptor, SET Configuration, and follow-up GET Configuration sequences.\n9) Poll for the final event completion on the event ring to confirm the control transfers have completed.\n10) Declare the test passed upon successful event completions and orderly termination.",
    "Meta Test Steps / Procedure": "- nic_programming();\n- GIC_EnableAllIRQ();\n- write_reg(MIZAR_USB_GCTL, set_data(read_reg(MIZAR_USB_GCTL), 0xFFFFFFFF, 0x30c11234));\n- write_reg(MIZAR_USB_GFLADJ, set_data(read_reg(MIZAR_USB_GFLADJ), 0xFFFFFFFF, 0x0a87f000));\n- write_reg(MIZAR_USB_GUCTL, set_data(read_reg(MIZAR_USB_GUCTL), 0xFFFFFFFF, 0x02000010));\n- rd_data = read_reg(MIZAR_USB_BASE+0xc2c0); write_reg(MIZAR_USB_BASE+0xc2c0, 0x010c0002);\n- rd_data = read_reg(MIZAR_USB_BASE+0xc200); write_reg(MIZAR_USB_BASE+0xc200, 0x00102407);\n- port_count = read_reg(MIZAR_USB_HCSPARAMS1); rd_data = read_reg(MIZAR_USB_SUPTPRT2_DW2); rd_data = read_reg(MIZAR_USB_SUPTPRT3_DW2); rd_data = read_reg(MIZAR_USB_PORTSC_20);\n- write_reg(MIZAR_USB_PORTSC_20, set_data(read_reg(MIZAR_USB_PORTSC_20), USB_PORTSC_20_WCE, 1));\n- write_reg(MIZAR_USB_PORTSC_20, set_data(read_reg(MIZAR_USB_PORTSC_20), USB_PORTSC_20_WDE, 1));\n- write_reg(MIZAR_USB_PORTSC_20, set_data(read_reg(MIZAR_USB_PORTSC_20), USB_PORTSC_20_WOE, 1));\n- write_reg(MIZAR_USB_PORTSC_20, 0x0e0002a0);\n- db_offset = read_reg(MIZAR_USB_DBOFF);\n- write_reg(Event_Ring_Segment_Table, Default_Event_Ring_Array);\n- write_reg(Event_Ring_Segment_Table + DWORD, 0x0);\n- write_reg(Event_Ring_Segment_Table + 2*DWORD, 0x30);\n- rd_data = read_reg(MIZAR_USB_HCSPARAMS2); rd_data = read_reg(MIZAR_USB_PAGESIZE);\n- write_reg(Scratchpad_Buffer_Array, SCRATCHPAD0); write_reg(Scratchpad_Buffer_Array + DWORD, 0x0);\n- write_reg(Scratchpad_Buffer_Array + 2*DWORD, SCRATCHPAD1); write_reg(Scratchpad_Buffer_Array + 3*DWORD, 0x0);\n- write_reg(Device_Context_Base_Address_Array, Scratchpad_Buffer_Array); write_reg(Device_Context_Base_Address_Array + DWORD, 0x0);\n- write_reg(Device_Context_Base_Address_Array + 2*DWORD, Device_Context_Array + 0x100); write_reg(Device_Context_Base_Address_Array + 3*DWORD, 0x0);\n- write_reg(Device_Context_Base_Address_Array + 4*DWORD, Device_Context_Array + 0x0d00); write_reg(Device_Context_Base_Address_Array + 5*DWORD, 0x0);\n- write_reg(MIZAR_USB_CRCR_LO, Default_Command_Ring + 0x1); write_reg(MIZAR_USB_CRCR_HI, 0x0);\n- write_reg(MIZAR_USB_CONFIG, 0x10); rd_data = read_reg(MIZAR_USB_CONFIG); write_reg(MIZAR_USB_CONFIG, 0x110);\n- write_reg(MIZAR_USB_DCBAAP_LO, Device_Context_Base_Address_Array); write_reg(MIZAR_USB_DCBAAP_HI, 0x0);\n- write_reg(MIZAR_USB_ERSTSZ, 0x1);\n- write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array); write_reg(MIZAR_USB_ERDP_HI, 0x0);\n- write_reg(MIZAR_USB_ERSTBA_LO, Event_Ring_Segment_Table); write_reg(MIZAR_USB_ERSTBA_HI, 0x0);\n- write_reg(MIZAR_USB_IMOD, 0x0); write_reg(MIZAR_USB_IMAN, 0x2);\n- write_reg(MIZAR_USB_USBCMD, 0x4); write_reg(MIZAR_LSS_SYSREG_INTR_EN0, 0x80000000); write_reg(MIZAR_USB_USBCMD, 0x5);\n- int_pend = 1; while (int_pend) { wait_on(100); }\n- usb_status = read_reg(MIZAR_USB_USBSTS); write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2); write_reg(MIZAR_USB_ERDP_HI, 0x0);\n- port_status = read_reg(MIZAR_USB_PORTSC_20);\n- write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array + 0x18);\n- write_reg(MIZAR_USB_PORTSC_20, 0x0e0006f1);\n- write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2);\n- write_reg(MIZAR_USB_PORTSC_20, 0x0e220200);\n- port_status = read_reg(MIZAR_USB_PORTSC_20);\n- int_pend = 1; while (int_pend) { wait_on(100); }\n- write_reg(Default_Command_Ring + 0x0, 0x0); write_reg(Default_Command_Ring + 0x4, 0x0); write_reg(Default_Command_Ring + 0x8, 0x0); write_reg(Default_Command_Ring + 0xc, 0x00002401);\n- usb_status = read_reg(MIZAR_USB_USBSTS); write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2); write_reg(MIZAR_USB_ERDP_HI, 0x0);\n- write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array + 0x28);\n- write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2);\n- write_reg(MIZAR_USB_PORTSC_20, 0x0e200e01);\n- write_reg(MIZAR_USB_DB, 0x0);\n- int_pend = 1; while (int_pend) { wait_on(100); }\n- write_reg(Default_Input_Context, 0x0); write_reg(Default_Input_Context + DWORD, 0x3);\n- write_reg(Default_Input_Context + 0x40, 0x08200000); write_reg(Default_Input_Context + 0x44, 0x00010000);\n- write_reg(Default_Input_Context + 0x80, 0x00); write_reg(Default_Input_Context + 0x84, 0x00080020);\n- write_reg(Default_Input_Context + 0x88, EP0_TR_Dequeue_Pointer | 0x1);\n- write_reg(Default_Input_Context + 0x90, 0x08);\n- input_context_address = (Default_Input_Context) + 0x0;\n- write_reg(Default_Command_Ring + 0x10, input_context_address); write_reg(Default_Command_Ring + 0x1c, 0x01002e01);\n- write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2); write_reg(MIZAR_USB_ERDP_HI, 0x0); write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array + 0x38);\n- write_reg(MIZAR_USB_DB, 0x0);\n- int_pend = 1; while (int_pend) { wait_on(100); }\n- event_completion = read_reg(Default_Event_Ring_Array + 0x30);\n- write_reg(Default_Input_Context + DWORD, 0x3);\n- write_reg(Default_Input_Context + 0x40, 0x08200000); write_reg(Default_Input_Context + 0x44, 0x00010000);\n- write_reg(Default_Input_Context + 0x80, 0x00); write_reg(Default_Input_Context + 0x84, 0x00080020);\n- write_reg(Default_Input_Context + 0x88, EP0_TR_Dequeue_Pointer | 0x1); write_reg(Default_Input_Context + 0x90, 0x08);\n- write_reg(Default_Command_Ring + 0x20, input_context_address); write_reg(Default_Command_Ring + 0x2c, 0x01002c01);\n- write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2); write_reg(MIZAR_USB_ERDP_HI, 0x0); write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array + 0x48); write_reg(MIZAR_USB_DB, 0x0);\n- int_pend = 1; while (int_pend) { wait_on(100); }\n- event_completion = read_reg(Default_Event_Ring_Array + 0x40);\n- enumeration();\n- printf(\"after enumeration\\n\");\n- finish(0);\n\nDefault_IRQHandler:\n- rd_data = read_reg(MIZAR_LSS_SYSREG_MSK_STS0); rd_data = read_reg(MIZAR_LSS_SYSREG_RAW_STCR0);\n- if (rd_data && 0x80000000) { write_reg(MIZAR_USB_IMAN, 0x1); write_reg(MIZAR_LSS_SYSREG_RAW_STCR0, 0x80000000); }\n- GIC_ClearIRQ(84);\n\nenumeration():\n- // GET Device Descriptor\n- write_reg(EP0_TR_Dequeue_Pointer + 0x00, 0x01000680); write_reg(EP0_TR_Dequeue_Pointer + 0x04, 0x00120000); write_reg(EP0_TR_Dequeue_Pointer + 0x08, 0x08); write_reg(EP0_TR_Dequeue_Pointer + 0x0c, 0x00030861);\n- write_reg(EP0_TR_Dequeue_Pointer + 0x10, EP0_TR_Dequeue_Pointer + 0x200); write_reg(EP0_TR_Dequeue_Pointer + 0x18, 0x12); write_reg(EP0_TR_Dequeue_Pointer + 0x1c, 0x00010c27);\n- write_reg(EP0_TR_Dequeue_Pointer + 0x2c, 0x00001023);\n- // GET Configuration Descriptor (short)\n- write_reg(EP0_TR_Dequeue_Pointer + 0x30, 0x02000680); write_reg(EP0_TR_Dequeue_Pointer + 0x34, 0x00090000); write_reg(EP0_TR_Dequeue_Pointer + 0x38, 0x08); write_reg(EP0_TR_Dequeue_Pointer + 0x3c, 0x00030861);\n- write_reg(EP0_TR_Dequeue_Pointer + 0x40, EP0_TR_Dequeue_Pointer + 0x240); write_reg(EP0_TR_Dequeue_Pointer + 0x48, 0x9); write_reg(EP0_TR_Dequeue_Pointer + 0x4c, 0x00010c25);\n- write_reg(EP0_TR_Dequeue_Pointer + 0x5c, 0x00001023);\n- // SET Configuration\n- write_reg(EP0_TR_Dequeue_Pointer + 0x60, 0x00010900); write_reg(EP0_TR_Dequeue_Pointer + 0x64, 0x00000000); write_reg(EP0_TR_Dequeue_Pointer + 0x68, 0x08); write_reg(EP0_TR_Dequeue_Pointer + 0x6c, 0x00000841);\n- write_reg(EP0_TR_Dequeue_Pointer + 0x7c, 0x00011023);\n- // GET Configuration (follow-ups)\n- write_reg(EP0_TR_Dequeue_Pointer + 0x80, 0x02000680); write_reg(EP0_TR_Dequeue_Pointer + 0x84, 0x00090000); write_reg(EP0_TR_Dequeue_Pointer + 0x88, 0x08); write_reg(EP0_TR_Dequeue_Pointer + 0x8c, 0x00030861);\n- write_reg(EP0_TR_Dequeue_Pointer + 0x90, EP0_TR_Dequeue_Pointer + 0x500); write_reg(EP0_TR_Dequeue_Pointer + 0x98, 0x8); write_reg(EP0_TR_Dequeue_Pointer + 0x9c, 0x00010c27); write_reg(EP0_TR_Dequeue_Pointer + 0xac, 0x00001023);\n- write_reg(EP0_TR_Dequeue_Pointer + 0xb0, 0x02000680); write_reg(EP0_TR_Dequeue_Pointer + 0xb4, 0x00180000); write_reg(EP0_TR_Dequeue_Pointer + 0xb8, 0x08); write_reg(EP0_TR_Dequeue_Pointer + 0xbc, 0x00030861);\n- write_reg(EP0_TR_Dequeue_Pointer + 0xc0, EP0_TR_Dequeue_Pointer + 0x560); write_reg(EP0_TR_Dequeue_Pointer + 0xc8, 0x18); write_reg(EP0_TR_Dequeue_Pointer + 0xcc, 0x00010c25); write_reg(EP0_TR_Dequeue_Pointer + 0xdc, 0x00001023);\n- write_reg(MIZAR_USB_BASE + 0x484, 0x1);\n- event_completion = read_reg(Default_Event_Ring_Array + 0x110); while (event_completion == 0) { wait_on(4); event_completion = read_reg(Default_Event_Ring_Array + 0x110); }
]

TESTPLAN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_COLUMNS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

def sort_key(item):
    try:
        return int(item.get("Index", 0))
    except Exception:
        return 0


def create_workbook(data: list) -> tuple[Workbook, str]:
    # Time in IST
    ist = ZoneInfo("Asia/Kolkata")
    now = datetime.now(ist)
    ts = now.strftime("%Y%m%d_%H%M%S")
    filename = f"{IP_NAME}_TestPlan_{ts}.xlsx"

    wb = Workbook()

    # Visible sheet: TestPlan
    ws_plan = wb.active
    ws_plan.title = "TestPlan"

    # Hidden (veryHidden) sheet: MetaData
    ws_meta = wb.create_sheet("MetaData")

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    # Write headers
    ws_plan.append(TESTPLAN_COLUMNS)
    ws_meta.append(METADATA_COLUMNS)

    for cell in ws_plan[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_top
    for cell in ws_meta[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_top

    # Write rows preserving order by Index (numeric sort)
    data_sorted = sorted(data, key=sort_key)

    for item in data_sorted:
        # TestPlan row
        row_plan = [item.get(col, "") for col in TESTPLAN_COLUMNS]
        ws_plan.append(row_plan)
        # MetaData row
        row_meta = [item.get(col, "") for col in METADATA_COLUMNS]
        ws_meta.append(row_meta)

    # Global metadata block in MetaData sheet (requested extra info)
    start_row = ws_meta.max_row + 2
    ws_meta.cell(row=start_row, column=1, value="Key").fill = header_fill
    ws_meta.cell(row=start_row, column=1).font = header_font
    ws_meta.cell(row=start_row, column=1).alignment = wrap_top
    ws_meta.cell(row=start_row, column=2, value="Value").fill = header_fill
    ws_meta.cell(row=start_row, column=2).font = header_font
    ws_meta.cell(row=start_row, column=2).alignment = wrap_top

    meta_extra = [
        ("IP_NAME", IP_NAME),
        ("SourceRepo", f"{OWNER}/{REPO}"),
        ("SourcePath", SOURCE_PATH),
        ("GenerationTimestamp(IST)", now.strftime("%Y-%m-%d %H:%M:%S %Z")),
        ("RecordCount", len(data_sorted)),
    ]
    r = start_row + 1
    for k, v in meta_extra:
        ws_meta.cell(row=r, column=1, value=k).alignment = wrap_top
        ws_meta.cell(row=r, column=2, value=str(v)).alignment = wrap_top
        r += 1

    # Formatting: wrap text for all cells and set reasonable column widths
    for ws in (ws_plan, ws_meta):
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = wrap_top
        # Set widths
        widths = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                widths[cell.column_letter] = max(widths.get(cell.column_letter, 10), min(len(val) + 2, 80))
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    # Set MetaData sheet to very hidden
    ws_meta.sheet_state = 'veryHidden'

    return wb, os.path.join(OUTPUT_DIR, filename)


def main():
    # STEP 1: Validate JSON
    if not isinstance(JSON_DATA, list) or any(not isinstance(x, dict) for x in JSON_DATA):
        raise SystemExit("Invalid json_data: must be a list of objects")

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # STEP 2-4: Build and save workbook
    wb, rel_path = create_workbook(JSON_DATA)
    wb.save(rel_path)

    # Record generated path for the workflow commit step
    with open('generated_excel_path.txt', 'w', encoding='utf-8') as f:
        f.write(rel_path.replace('\\\\', '/').replace('\\', '/'))

    print(f"Generated Excel at: {rel_path}")


if __name__ == "__main__":
    main()
