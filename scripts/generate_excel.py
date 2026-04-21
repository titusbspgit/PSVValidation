#!/usr/bin/env python3
import json, os, sys, math
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def json_to_str(v):
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v


def auto_fit(ws):
    max_width = {}
    for r in ws.iter_rows(values_only=True):
        for ci, val in enumerate(r, start=1):
            s = "" if val is None else str(val)
            # consider multi-line values
            width = 0
            for line in s.splitlines() if s else [""]:
                width = max(width, len(line))
            width = min(width, 120)
            if ci not in max_width or width > max_width[ci]:
                max_width[ci] = width
    for ci, w in max_width.items():
        # Add padding and clamp to reasonable bounds
        ws.column_dimensions[get_column_letter(ci)].width = max(10, min(w + 2, 80))


def add_borders(ws):
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = ws.max_row
    max_col = ws.max_column
    for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in r:
            cell.border = border


def adjust_row_heights(tp, wrap_headers):
    # Estimate row height based on content length and column width
    header_map = {tp.cell(row=1, column=ci).value: ci for ci in range(1, tp.max_column + 1)}
    wrap_cols = [header_map[h] for h in wrap_headers if h in header_map]
    default_height = 15.0  # points, typical default
    for ri in range(2, tp.max_row + 1):
        max_lines = 1
        for ci in wrap_cols:
            cell = tp.cell(row=ri, column=ci)
            text = "" if cell.value is None else str(cell.value)
            if not text:
                continue
            width_chars = tp.column_dimensions[get_column_letter(ci)].width or 10
            # Effective characters per line (conservative, account for padding)
            eff = max(5, int(width_chars) - 2)
            # Count explicit line breaks and wrapped lines
            parts = text.splitlines() if text else [""]
            lines = 0
            for p in parts:
                if not p:
                    lines += 1
                else:
                    lines += max(1, math.ceil(len(p) / eff))
            max_lines = max(max_lines, lines)
        # Cap maximum height for safety
        tp.row_dimensions[ri].height = min(300, max(default_height, default_height * max_lines * 1.1))


def main():
    repo_root = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.abspath(os.path.join(repo_root, '..'))
    in_path = os.path.join(repo_root, 'data', 'input_gpio.json')
    with open(in_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    tests = data.get('tests', [])
    if not isinstance(tests, list) or not tests:
        print('ERROR: tests array missing or empty', file=sys.stderr)
        sys.exit(2)
    # Build header order (union, first-seen)
    header_order = []
    for row in tests:
        if not isinstance(row, dict):
            print('ERROR: each test must be an object', file=sys.stderr)
            sys.exit(2)
        for k in row.keys():
            if k not in header_order:
                header_order.append(k)
    meta_cols = [
        'Hidden_Test_Case_Name',
        'Hidden_Test_Description',
        'Hidden_Remarks',
        'Hidden_Test_Steps_Procedure',
        'Hidden_Impacted_Registers',
        'Hidden_Validation_Acceptance_Criteria',
    ]
    main_cols = [
        'Index','SS / Module','Feature','Test Case Name','Test Description','Speed','Mode',
        'Memory Start Offset','Memory End Offset','Remarks','Test Steps / Procedure',
        'Impacted Registers','Validation / Acceptance Criteria','Code Generation (Required / Not)'
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    # Headers (base sheet)
    for ci, h in enumerate(header_order, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    # Rows
    for ri, row in enumerate(tests, start=2):
        for ci, h in enumerate(header_order, start=1):
            v = row.get(h, '')
            ws.cell(row=ri, column=ci, value=json_to_str(v))
    auto_fit(ws)
    # Create Meta sheet (no formatting per spec)
    meta_ws = wb.create_sheet('Meta_data_sheet')
    present_meta = [c for c in meta_cols if c in header_order]
    for ci, h in enumerate(present_meta, start=1):
        meta_ws.cell(row=1, column=ci, value=h)
    # Copy values
    for ri in range(2, ws.max_row + 1):
        for ci, h in enumerate(present_meta, start=1):
            src_idx = header_order.index(h) + 1
            meta_ws.cell(row=ri, column=ci, value=ws.cell(row=ri, column=src_idx).value)
    # Very hidden
    meta_ws.sheet_state = 'veryHidden'
    # Build TestPlan sheet with only approved columns, ordered
    testplan_tmp = wb.create_sheet('TestPlan_tmp')
    present_main = [c for c in main_cols if c in header_order]
    for ci, h in enumerate(present_main, start=1):
        cell = testplan_tmp.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for ri in range(2, ws.max_row + 1):
        for ci, h in enumerate(present_main, start=1):
            src_idx = header_order.index(h) + 1
            testplan_tmp.cell(row=ri, column=ci, value=ws.cell(row=ri, column=src_idx).value)
    # Delete original Data and rename tmp → TestPlan
    del wb['Data']
    testplan_tmp.title = 'TestPlan'
    tp = wb['TestPlan']
    # Strict formatting (visual only) applied ONLY to TestPlan
    wrap_cols = {
        'Test Description', 'Remarks', 'Test Steps / Procedure', 'Validation / Acceptance Criteria'
    }
    # Header already bold+centered; align data cells
    for ri in range(2, tp.max_row + 1):
        for ci in range(1, tp.max_column + 1):
            h = tp.cell(row=1, column=ci).value
            cell = tp.cell(row=ri, column=ci)
            if h == 'Index':
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=False)
            elif h in ('Memory Start Offset','Memory End Offset'):
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=False)
            elif h in wrap_cols:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
    auto_fit(tp)
    add_borders(tp)
    adjust_row_heights(tp, wrap_cols)
    # Save to output path with IST timestamp
    tz = ZoneInfo('Asia/Kolkata')
    now = datetime.now(tz)
    ip_env = os.environ.get('IP_NAME')
    ip_from_json = str(data.get('ip','')).upper() if data.get('ip') else 'IP'
    ip_name = (ip_env or ip_from_json or 'IP').upper()
    out_dir = os.path.join(repo_root, 'Test_Output', 'GPIO', 'TestPlan')
    os.makedirs(out_dir, exist_ok=True)
    fname = f"{ip_name}_TestPlan_{now.strftime('%Y%m%d')}_{now.strftime('%H%M%S')}.xlsx"
    out_path = os.path.join(out_dir, fname)
    wb.save(out_path)
    # Write repo-relative path for git add
    rel = os.path.relpath(out_path, repo_root)
    with open(os.path.join(repo_root, 'scripts', 'output_path.txt'), 'w', encoding='utf-8') as outf:
        outf.write(rel)
    print(f"Saved: {rel}")


if __name__ == '__main__':
    main()
