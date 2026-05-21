#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generates a real .xlsx from aggregated Test Plan JSON with two sheets:
- TestPlan (visible)
- MetaData (VERY HIDDEN)

Filename: testplan_<YYYYMMDD_HHMMSS IST>.xlsx
Output directory: Test_Output

Requirements: openpyxl
"""
import os
from pathlib import Path
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:
    ZoneInfo = None

from openpyxl import Workbook
from openpyxl.styles import Font

# ---- Aggregated JSON data (preserve exactly) ----
JSON_DATA = [
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "GPIO Register R/W and Reset-Default Validation",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Validate GPIO GP0 register reset defaults and masked write/read-back behavior for registers corresponding to pins 8 through 27. For each readable register, confirm the reset value (with input bit masked). For each writable register, write test patterns masked by the write mask, then read back and verify only writable/readable bits change while others retain default state.",
    "Meta Test Description": "The test performs two phases:\n1) Reset-default verification: Iterate over the configured GPIO GP0 register address list (for pins 8..27). For each index i, if skip_rst_array[i] == 1, skip. If read_mask_array[i] == 0, skip as non-readable. Otherwise read data_rd = read_reg(addr_array[i]) and compute data = (data_rd & 0xfffffffe) to mask out the LSB input bit. Compare data to default_value_array[i]. On mismatch, increment def_fail_cnt and log failure; otherwise log pass.\n2) Write/read-back verification using six patterns {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}. For each pattern data_wr, perform a write phase over all indices i: if skip_array[i] == 1 or write_mask_array[i] == 0, skip; else write write_reg(addr_array[i], (data_wr & write_mask_array[i])). Then a read/verify phase: for each i, if skip_array[i] == 1 or write_mask_array[i] == 0 or read_mask_array[i] == 0, skip; else read data_rd = (read_reg(addr_array[i]) & read_mask_array[i]); compute wr_n = (write_mask_array[i] ^ 0xffffffff) and expected value exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])). If data_rd != exp_val, increment wr_fail_cnt and log failure; otherwise log pass. Test ends with finish(0) if def_fail_cnt == 0 and wr_fail_cnt == 0, else finish(1).",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Configured for GPIO GP0 registers corresponding to pins 8..27. No registers are marked to be skipped in current skip arrays. The default value check masks off the least significant bit (0xfffffffe). A soft reset check routine exists but is disabled. Note: CNT is defined as 49 while the configured arrays have 20 entries, which may cause out-of-bounds access unless CNT matches the array sizes.",
    "Test Steps / Procedure": "1) Initialize test and load the configured GPIO GP0 register list for pins 8 through 27 with associated default, read-mask, and write-mask data.\n2) For each listed register: if it is readable, read the reset value (with the input bit masked) and verify it matches the expected default value; otherwise skip non-readable entries.\n3) For each predefined test pattern, write the pattern to all writable registers using their write masks.\n4) For each register written in step 3, read back the value (respecting the read mask) and verify that writable/readable bits match the written pattern while non-writable bits retain their default values.\n5) Declare PASS if all default-value checks and all masked write/read-back checks pass for all applicable registers; otherwise declare FAIL.",
    "Meta Test Steps / Procedure": "- test_case():\n  - Call chk_rst_val().\n  - Call chk_rd_wr().\n  - If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0).\n- chk_rst_val():\n  - For i = 0 .. CNT-1:\n    - addr = addr_array[i].\n    - If skip_rst_array[i] == 1: log skip and continue.\n    - If read_mask_array[i] == 0x00000000: log non-readable and continue.\n    - data_rd = read_reg(addr).\n    - data = (data_rd & 0xfffffffe).\n    - If data == default_value_array[i]: log PASS; else def_fail_cnt++ and log mismatch (addr, expected, read, raw).\n- chk_rd_wr():\n  - Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0xf5f5f5f5, 0xA5A5A5A5, 0xffff0000}.\n  - For j = 0..5:\n    - data_wr = chk_val[j].\n    - Write phase: For i = 0..CNT-1:\n      - addr = addr_array[i].\n      - If skip_array[i] == 1: continue.\n      - If write_mask_array[i] == 0x00000000: continue.\n      - Else write_reg(addr, (data_wr & write_mask_array[i])).\n    - Read/verify phase: For i = 0..CNT-1:\n      - addr = addr_array[i].\n      - If skip_array[i] == 1: continue.\n      - If write_mask_array[i] == 0x00000000: continue.\n      - If read_mask_array[i] == 0x00000000: continue.\n      - data_rd = (read_reg(addr) & read_mask_array[i]).\n      - wr_n = (write_mask_array[i] ^ 0xffffffff).\n      - exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]) ).\n      - If data_rd == exp_val: log PASS; else wr_fail_cnt++ and log mismatch (addr, expected, read).\n- Disabled routine soft_reset_chk(): writes and restores a soft reset register value with waits, but is excluded by preprocessor condition.",
    "Impacted Registers": "gp0_gpio_8",
    "Meta Impacted Registers": "Address macros: MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_GPIO_9, MIZAR_GPIO_GP0_GPIO_10, MIZAR_GPIO_GP0_GPIO_11, MIZAR_GPIO_GP0_GPIO_12, MIZAR_GPIO_GP0_GPIO_13, MIZAR_GPIO_GP0_GPIO_14, MIZAR_GPIO_GP0_GPIO_15, MIZAR_GPIO_GP0_GPIO_16, MIZAR_GPIO_GP0_GPIO_17, MIZAR_GPIO_GP0_GPIO_18, MIZAR_GPIO_GP0_GPIO_19, MIZAR_GPIO_GP0_GPIO_20, MIZAR_GPIO_GP0_GPIO_21, MIZAR_GPIO_GP0_GPIO_22, MIZAR_GPIO_GP0_GPIO_23, MIZAR_GPIO_GP0_GPIO_24, MIZAR_GPIO_GP0_GPIO_25, MIZAR_GPIO_GP0_GPIO_26, MIZAR_GPIO_GP0_GPIO_27. Mask/default macros: GPIO_GP0_GPIO_8_DEFAULT_VAL, GPIO_GP0_GPIO_9_DEFAULT_VAL, GPIO_GP0_GPIO_10_DEFAULT_VAL, GPIO_GP0_GPIO_11_DEFAULT_VAL, GPIO_GP0_GPIO_12_DEFAULT_VAL, GPIO_GP0_GPIO_13_DEFAULT_VAL, GPIO_GP0_GPIO_14_DEFAULT_VAL, GPIO_GP0_GPIO_15_DEFAULT_VAL, GPIO_GP0_GPIO_16_DEFAULT_VAL, GPIO_GP0_GPIO_17_DEFAULT_VAL, GPIO_GP0_GPIO_18_DEFAULT_VAL, GPIO_GP0_GPIO_19_DEFAULT_VAL, GPIO_GP0_GPIO_20_DEFAULT_VAL, GPIO_GP0_GPIO_21_DEFAULT_VAL, GPIO_GP0_GPIO_22_DEFAULT_VAL, GPIO_GP0_GPIO_23_DEFAULT_VAL, GPIO_GP0_GPIO_24_DEFAULT_VAL, GPIO_GP0_GPIO_25_DEFAULT_VAL, GPIO_GP0_GPIO_26_DEFAULT_VAL, GPIO_GP0_GPIO_27_DEFAULT_VAL, GPIO_GP0_GPIO_8_READ_MASK, GPIO_GP0_GPIO_9_READ_MASK, GPIO_GP0_GPIO_10_READ_MASK, GPIO_GP0_GPIO_11_READ_MASK, GPIO_GP0_GPIO_12_READ_MASK, GPIO_GP0_GPIO_13_READ_MASK, GPIO_GP0_GPIO_14_READ_MASK, GPIO_GP0_GPIO_15_READ_MASK, GPIO_GP0_GPIO_16_READ_MASK, GPIO_GP0_GPIO_17_READ_MASK, GPIO_GP0_GPIO_18_READ_MASK, GPIO_GP0_GPIO_19_READ_MASK, GPIO_GP0_GPIO_20_READ_MASK, GPIO_GP0_GPIO_21_READ_MASK, GPIO_GP0_GPIO_22_READ_MASK, GPIO_GP0_GPIO_23_READ_MASK, GPIO_GP0_GPIO_24_READ_MASK, GPIO_GP0_GPIO_25_READ_MASK, GPIO_GP0_GPIO_26_READ_MASK, GPIO_GP0_GPIO_27_READ_MASK, GPIO_GP0_GPIO_8_WRITE_MASK, GPIO_GP0_GPIO_9_WRITE_MASK, GPIO_GP0_GPIO_10_WRITE_MASK, GPIO_GP0_GPIO_11_WRITE_MASK, GPIO_GP0_GPIO_12_WRITE_MASK, GPIO_GP0_GPIO_13_WRITE_MASK, GPIO_GP0_GPIO_14_WRITE_MASK, GPIO_GP0_GPIO_15_WRITE_MASK, GPIO_GP0_GPIO_16_WRITE_MASK, GPIO_GP0_GPIO_17_WRITE_MASK, GPIO_GP0_GPIO_18_WRITE_MASK, GPIO_GP0_GPIO_19_WRITE_MASK, GPIO_GP0_GPIO_20_WRITE_MASK, GPIO_GP0_GPIO_21_WRITE_MASK, GPIO_GP0_GPIO_22_WRITE_MASK, GPIO_GP0_GPIO_23_WRITE_MASK, GPIO_GP0_GPIO_24_WRITE_MASK, GPIO_GP0_GPIO_25_WRITE_MASK, GPIO_GP0_GPIO_26_WRITE_MASK, GPIO_GP0_GPIO_27_WRITE_MASK",
    "Validation / Acceptance Criteria": "PASS if all applicable GPIO GP0 registers (pins 8..27) meet both criteria: (a) the readable reset value (with the input bit masked) equals the expected default value; (b) for each test pattern, the read-back value equals the expected masked composition where writable/readable bits reflect the pattern and non-writable bits preserve default values. FAIL if any default or write/read-back comparison fails.",
    "Meta Validation / Acceptance Criteria": "- Default check criterion: For each i where read_mask_array[i] != 0 and skip_rst_array[i] == 0: (read_reg(addr_array[i]) & 0xfffffffe) == default_value_array[i].\n- Write/read-back check criterion: For each pattern data_wr and each i where write_mask_array[i] != 0, read_mask_array[i] != 0, and skip_array[i] == 0: Let data_rd = (read_reg(addr_array[i]) & read_mask_array[i]) and wr_n = (write_mask_array[i] ^ 0xffffffff). Expected exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])). Require data_rd == exp_val.\n- Overall result: finish(0) if def_fail_cnt == 0 and wr_fail_cnt == 0; else finish(1).",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>, #include <stdlib.h>, #include \"test_common.h\", #include \"test_define.c\", #include <gpio/gpio_def.h>, #include <gpio/gpio_offset.h>",
    "Meta Macros": "#define CNT 49; #define SOFT_RST_REG_ADDRESS 0x00000000; #define SOFT_RST_REG_DATA 0x00000000",
    "Meta Arrays": "skip_array[20] = {0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0}; skip_rst_array[20] = {0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0}."
  },
  {
    "Index": "2",
    "SS / Module": "GPIO",
    "Feature": "GPIO Negative-edge Interrupt Enable/Validation",
    "Test Case Name": "test_gpio_negedge_intr_en",
    "Test Description": "Verify negative-edge interrupt generation, detection, and clearing for GPIO GP0 pins 8 through 39 using the ISR path. The test enables the GPIO interrupt at the system level, configures each per-pin control, enables per-line interrupt, generates a falling edge on each line, expects the ISR to trigger, validates per-pin and group status, clears the interrupt, and confirms the status is cleared.",
    "Meta Test Description": "The test initializes GIC and system interrupt enable for the selected GPIO instance (GPIO0 uses IRQ 87; GPIO1 uses IRQ 88). A stimulus register at 0xA0243ffc is used to drive GPIO lines. First, the stimulus is set high (0xffffffff). Then for i = 0..31, the per-pin configuration register at base GP0_GPIO_8 + i*4 is programmed with bits [20,18,16] set, and a short wait is performed. Next, for i = 0..31: clear any pending raw interrupt for bit i in gpio_intr_raw_stclr1; enable the corresponding per-line interrupt bit i in GP0_INTR1_INTR_EN1; wait; set int_pend = 1; re-drive stimulus high then drive it low only for bit i (write ~wr_val to 0xA0243ffc) to create a falling edge; poll for int_pend to be cleared by the ISR with a timeout. On timeout, increment test_err. The Default_IRQHandler clears int_pend, drives the stimulus back high (0xffffffff), reads per-pin register GP0_GPIO_8 + i*4, expects input bit0 to read 0 and interrupt bit1 to be set; reads group status GP0_INTR1_INTR_STS1 and expects bit i set; writes GP0_GPIO_8 + i*4 with bits [20,16] set to reconfigure/ack, clears raw interrupt for bit i in gpio_intr_raw_stclr1, re-reads GP0_INTR1_INTR_STS1 and expects it to be 0; clears the system RAW_STCR1 for the corresponding GPIO interrupt and clears the GIC IRQ. Any mismatch increments test_err. Test ends with finish(test_err).",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Requires system interrupt enable (INTR_EN1) and RAW status clear (RAW_STCR1). GIC IRQ 87 or 88 must be available depending on the GPIO instance. External stimulus at address 0xA0243ffc must control the GPIO lines to generate falling edges. Assumes 32 consecutive per-pin configuration registers starting from GP0_GPIO_8 (offset increment of 4 bytes).",
    "Test Steps / Procedure": "1) Enable the appropriate GIC interrupt for the target GPIO instance and enable the GPIO interrupt at the system level (INTR_EN1).\n2) Drive the external stimulus register (0xA0243ffc) to set all GPIO lines high as the initial state.\n3) For each of the 32 GPIO lines (pins 8..39), configure the per-pin control register (starting at GP0_GPIO_8 + i*4) to enable input/interrupt as required.\n4) For each GPIO line i: clear any pending raw interrupt bit i in gpio_intr_raw_stclr1, then enable bit i in GP0_INTR1_INTR_EN1.\n5) Generate a falling edge on line i by toggling the external stimulus (0xA0243ffc) from high to low only on bit i, and wait for the ISR to run; declare a timeout failure if the ISR does not clear the pending flag within the allowed time.\n6) In the ISR, verify the per-pin input reads low and the per-pin interrupt indication is set; verify the group status (GP0_INTR1_INTR_STS1) bit i is set; then clear the per-line raw interrupt and confirm the group status is cleared; finally clear the system RAW status and the GIC interrupt.\n7) Pass if all lines trigger correctly and all validations and clears succeed; otherwise fail.",
    "Meta Test Steps / Procedure": "- test_case():\n  - test_err = 0.\n  - If GPIO0: GIC_EnableIRQ(87); write INTR_EN1 with the GPIO0 interrupt enable field.\n  - If GPIO1: GIC_EnableIRQ(88); write INTR_EN1 with the GPIO1 interrupt enable field.\n  - write_reg(0xA0243ffc, 0xffffffff) to drive stimulus high.\n  - For (i = 0; i < 32; i++):\n    - addr1 = GP0_GPIO_8 + i*4.\n    - write_reg(addr1, (1<<20) | (1<<18) | (1<<16)).\n    - wait_on(10).\n  - For (i = 0; i < 32; i++):\n    - wr_val = 1u << i.\n    - write_reg(gpio_intr_raw_stclr1, wr_val) to clear raw status bit i.\n    - write_reg(GP0_INTR1_INTR_EN1, wr_val) to enable per-line interrupt i.\n    - wait_on(10).\n    - int_pend = 1.\n    - write_reg(0xA0243ffc, 0xffffffff); wait_on(30); write_reg(0xA0243ffc, ~wr_val) to create a falling edge on line i.\n    - timeout = 5000; while (int_pend && timeout--) wait_on(10).\n    - If (timeout == 0): log timeout for GPIO(i+8); test_err++.\n  - finish(test_err).\n- Default_IRQHandler():\n  - local_wr = 1u << i; int_pend = 0.\n  - write_reg(0xA0243ffc, 0xffffffff) to restore stimulus high.\n  - raddr = GP0_GPIO_8 + i*4; rdata = read_reg(raddr).\n  - If ((rdata & 0x1) != 0): test_err++ (input should be low after edge).\n  - If ((rdata & 0x2) != 0x0):\n    - rdata_grp = read_reg(GP0_INTR1_INTR_STS1); if ((rdata_grp & local_wr) == 0) test_err++.\n    - write_reg(GP0_GPIO_8 + i*4, (1<<20) | (1<<16)).\n    - write_reg(gpio_intr_raw_stclr1, local_wr).\n    - rdata_grp = read_reg(GP0_INTR1_INTR_STS1); if (rdata_grp != 0x0) test_err++.\n    - Clear system and GIC: write RAW_STCR1 with the corresponding GPIO interrupt clear field; GIC_ClearIRQ(87 or 88).\n  - Else: test_err++.",
    "Impacted Registers": "INTR_EN1, GP0_GPIO_8, GP0_INTR1_INTR_EN1, GP0_INTR1_INTR_STS1, gpio_intr_raw_stclr1, RAW_STCR1",
    "Meta Impacted Registers": "MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR, LSS_SYSREG_INTR_EN1_GPIO1_INTR, MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GPIO_INTR_RAW_STCLR1, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR, LSS_SYSREG_RAW_STCR1_GPIO1_INTR",
    "Validation / Acceptance Criteria": "PASS if for each GPIO line (pins 8..39) a negative-edge stimulus triggers an interrupt handled by the ISR, the per-pin input reads low and the interrupt indication is set, the group interrupt status reflects the active line, and after clearing, the group status becomes zero. Any timeout waiting for the ISR or any mismatch in per-pin or group status results in FAIL.",
    "Meta Validation / Acceptance Criteria": "- Timeout check: After generating the falling edge for bit i, the ISR must run and clear int_pend before timeout expires; else test_err++.\n- Per-pin checks in ISR: (rdata & 0x1) == 0; (rdata & 0x2) != 0x0.\n- Group status set: (read_reg(GP0_INTR1_INTR_STS1) & (1<<i)) != 0.\n- Clear sequence: write_reg(gpio_intr_raw_stclr1, 1<<i); then read_reg(GP0_INTR1_INTR_STS1) == 0.\n- System/GIC clear: write RAW_STCR1 GPIO interrupt clear field and GIC_ClearIRQ(87/88).\n- Overall: finish(0) → PASS; any increment to test_err → FAIL.",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>, #include <lss_sysreg.h>, #include \\\"test_define.c\\\", #include <test_common.h>, #include <gpio/gpio_def.h>, #include <gpio/gpio_offset.h>",
    "Meta Macros": "#define CNT 49",
    "Meta Arrays": "skip_array[20] = {0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0}."
  },
  {
    "Index": "3",
    "SS / Module": "GPIO",
    "Feature": "GPIO Positive-edge Interrupt Enablement",
    "Test Case Name": "test_gpio_pedge_all_pads_en",
    "Test Description": "Enable positive-edge interrupt detection on all GPIO pads/lines and validate that a low-to-high transition on each line is detected and can be cleared.",
    "Meta Test Description": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Requires the ability to generate low-to-high transitions on GPIO lines for validation. Ensure system interrupt routing is configured if using ISR-based observation.",
    "Test Steps / Procedure": "1) Configure all GPIO pads for input and interrupt detection.\n2) Enable positive-edge interrupt for all pads using the pedge_intr_en register.\n3) Apply a low-to-high transition on each GPIO line one at a time.\n4) Verify that the corresponding interrupt/status is asserted for the active line.\n5) Clear the interrupt/status and confirm it is deasserted.\n6) Repeat steps 3–5 for all pads and record pass/fail.",
    "Meta Test Steps / Procedure": "NA",
    "Impacted Registers": "pedge_intr_en, GPIO_8",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_INTR1_PEDGE_EN1, MIZAR_GPIO_GP0_GPIO_8",
    "Validation / Acceptance Criteria": "PASS if enabling positive-edge detection results in interrupts/status for each pad upon a low-to-high transition and the condition can be cleared for all pads; otherwise FAIL.",
    "Meta Validation / Acceptance Criteria": "NA",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "NA",
    "Meta Macros": "NA",
    "Meta Arrays": "NA"
  }
]

TESTPLAN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_COLUMNS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]


def bold_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)


def write_sheet(ws, columns, rows):
    ws.append(columns)
    for item in rows:
        ws.append([item.get(col, "") for col in columns])
    ws.freeze_panes = "A2"  # Freeze first row


def main():
    # Step 1: Validate JSON
    if not isinstance(JSON_DATA, list) or not all(isinstance(x, dict) for x in JSON_DATA):
        raise SystemExit("json_data must be a list of objects")

    # Step 2/3: Create workbook and sheets
    wb = Workbook()
    ws_plan = wb.active
    ws_plan.title = "TestPlan"
    write_sheet(ws_plan, TESTPLAN_COLUMNS, JSON_DATA)
    bold_header(ws_plan)

    ws_meta = wb.create_sheet("MetaData")
    write_sheet(ws_meta, METADATA_COLUMNS, JSON_DATA)
    bold_header(ws_meta)

    # VERY HIDDEN metadata sheet
    ws_meta.sheet_state = "veryHidden"

    # Step 4: Save file with IST timestamp
    tz = ZoneInfo("Asia/Kolkata") if ZoneInfo else None
    now = datetime.now(tz) if tz else datetime.utcnow()
    ts = now.strftime("%Y%m%d_%H%M%S")
    out_dir = Path("Test_Output")
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = f"testplan_{ts}.xlsx"
    out_path = out_dir / filename
    wb.save(out_path)
    print(f"Generated: {out_path}")


if __name__ == "__main__":
    main()
