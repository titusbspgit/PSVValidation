#!/usr/bin/env python3
# Deterministic Stage1 fallback generator for PCIE TestPlan Excel
# - Generates a true binary .xlsx using openpyxl
# - Applies strict Stage1 formatting and validation rules
# - Commits ONLY the finalized Excel to the repository (branch main)

import json
import os
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# -----------------------------
# INPUT CONSTANTS (deterministic)
# -----------------------------
IP_NAME = "PCIE"
OUTPUT_DIR = Path("Test_Output/PCIE/TestPlan/")
BRANCH = os.getenv("GITHUB_REF_NAME", "main")
COMMIT_MESSAGE_PREFIX = "[Stage1] Add PCIE TestPlan Excel (IST: {stamp})"

# Latest Test Plan JSON (array of rows). Using the generated content from TestPlan-Gen step.
# NOTE: JSON is an array as required. Each object -> one row.
json_data = [
  {
    "Index": "1",
    "SS / Module": "PCIE",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie0_dbi_dsp_reg_wr_rd_test",
    "Test Description": "Verifies default values and masked write/read behavior of PCIe DBI DSP registers using multiple data patterns.",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Addresses marked not readable or not writable are skipped. Registers listed in the skip array are not exercised. Default value checking is skipped for DBI_DSP_CAP_ID_NXT_PTR_REG, DBI_DSP_DEVICE_CONTROL_DEVICE_STATUS, and DBI_DSP_PL_DEBUG1_OFF.",
    "Test Steps / Procedure": "1) Read each readable DBI DSP register and compare the value with its documented reset default; skip CAP_ID_NXT_PTR_REG, DEVICE_CONTROL_DEVICE_STATUS, and PL_DEBUG1_OFF.\n2) For six predefined data patterns, write each pattern to every writable DBI DSP register.\n3) After each write pass, read back each readable and writable register and compute the expected masked value; compare the read value against the expected result.\n4) Record any mismatches during default checks or write/read verification.\n5) Report pass if no mismatches are found; otherwise report fail.",
    "Impacted Registers": "DBI_DSP_CAP_ID_NXT_PTR_REG; DBI_DSP_DEVICE_CONTROL_DEVICE_STATUS; DBI_DSP_PL_DEBUG1_OFF; DBI_DSP_TYPE1_DEV_ID_VEND_ID_REG; DBI_DSP_LINK_CONTROL2_LINK_STATUS2_REG",
    "Validation / Acceptance Criteria": "- Pass when all readable registers match their documented default values and all masked write/read comparisons match the expected values across all six data patterns.\n- Fail if any default value mismatch or any masked write/read mismatch is observed.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie0_dbi_dsp_reg_wr_rd_test",
    "Hidden_Test_Description": "Default value check and masked write/read verification for PCIe0 DBI DSP register set over six data patterns.",
    "Hidden_Remarks": "Addresses with read_mask_array[i] == 0x00000000 are skipped for reading. Addresses with write_mask_array[i] == 0x00000000 are skipped for writing. Addresses with skip_array[i] == 1 are skipped for both write and read. Default value checking in chk_rst_val() is explicitly skipped for mizar_PCIE0_DBI_DSP_CAP_ID_NXT_PTR_REG, mizar_PCIE0_DBI_DSP_DEVICE_CONTROL_DEVICE_STATUS, and mizar_PCIE0_DBI_DSP_PL_DEBUG1_OFF.",
    "Hidden_Test_Steps_Procedure": "Entry Point: test_case()\n1) Call chk_rst_val()\n   1.1) Loop entry: for (i = 0; i < CNT; i++)\n        - addr = addr_array[i]\n        - If read_mask_array[i] == 0x00000000: continue (skip read)\n        - If (addr == mizar_PCIE0_DBI_DSP_CAP_ID_NXT_PTR_REG) || (addr == mizar_PCIE0_DBI_DSP_DEVICE_CONTROL_DEVICE_STATUS) || (addr == mizar_PCIE0_DBI_DSP_PL_DEBUG1_OFF): continue (skip default check)\n        - Operation: READ -> data_rd = read_reg(addr) // register macro source: addr_array[i]\n        - Condition: if (data_rd == default_value_array[i]) pass; else { def_fail_cnt++; printf failure }\n   1.2) Loop exit condition: i reaches CNT (i == CNT)\n2) Call chk_rd_wr()\n   2.1) Initialize local patterns: int chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0x00000000, 0xA5A5A5A5, 0xffff0000}\n   2.2) Outer loop entry: for (j = 0; j < 6; j++)\n        - data_wr = chk_val[j]\n        - Write pass loop: for (i = 0; i < CNT; i++)\n          * addr = addr_array[i]\n          * If skip_array[i] == 1: continue (skip)\n          * If write_mask_array[i] == 0x00000000: continue (skip write)\n          * Operation: WRITE -> write_reg(addr, data_wr) // register macro source: addr_array[i]\n        - Read/verify pass loop: for (i = 0; i < CNT; i++)\n          * addr = addr_array[i]\n          * If skip_array[i] == 1: continue\n          * If write_mask_array[i] == 0x00000000: continue (skip read-back)\n          * If read_mask_array[i] == 0x00000000: continue (skip read-back)\n          * Operation: READ -> data_rd = read_reg(addr) // register macro source: addr_array[i]\n          * Compute: wr_n = (write_mask_array[i] ^ 0xffffffff)\n          * Compute expected: exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]))\n          * Condition: if (data_rd == exp_val) pass; else { wr_fail_cnt++; printf failure }\n   2.3) Outer loop exit condition: j reaches 6 (j == 6)\n3) Final result in test_case()\n   - If (def_fail_cnt > 0 || wr_fail_cnt > 0): finish(1) // fail\n   - Else: finish(0) // pass\nNotes:\n- Timing: No explicit delays in executed path (wait_on() exists only in soft_reset_chk() which is not called).\n- All register accesses use the macro-defined addresses from addr_array[i].\n- Explicitly referenced macros during logic: mizar_PCIE0_DBI_DSP_CAP_ID_NXT_PTR_REG, mizar_PCIE0_DBI_DSP_DEVICE_CONTROL_DEVICE_STATUS, mizar_PCIE0_DBI_DSP_PL_DEBUG1_OFF.",
    "Hidden_Impacted_Registers": "mizar_PCIE0_DBI_DSP_CAP_ID_NXT_PTR_REG; mizar_PCIE0_DBI_DSP_DEVICE_CONTROL_DEVICE_STATUS; mizar_PCIE0_DBI_DSP_PL_DEBUG1_OFF; mizar_PCIE0_DBI_DSP_TYPE1_DEV_ID_VEND_ID_REG; mizar_PCIE0_DBI_DSP_LINK_CONTROL2_LINK_STATUS2_REG",
    "Hidden_Validation_Acceptance_Criteria": "1) For each i where read_mask_array[i] != 0 and not one of the three skipped addresses, read_reg(addr_array[i]) must equal default_value_array[i]; else def_fail_cnt++.\n2) For each pattern j and each i where skip_array[i] == 0 and write_mask_array[i] != 0 and read_mask_array[i] != 0: read_reg(addr_array[i]) must equal exp_val computed as ((data_wr & read_mask_array[i] & write_mask_array[i]) | ((write_mask_array[i] ^ 0xffffffff) & read_mask_array[i] & default_value_array[i])); else wr_fail_cnt++.\n3) PASS if (def_fail_cnt == 0 && wr_fail_cnt == 0) resulting in finish(0); otherwise FAIL with finish(1)."
  }
]

# Column definitions
META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]
MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

# Utilities
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')  # Excel blue


def ensure_json_array(js):
    if not isinstance(js, list) or len(js) == 0:
        raise SystemExit("JSON input must be a non-empty array of objects")
    for row in js:
        if not isinstance(row, dict):
            raise SystemExit("Each JSON array element must be an object (dict)")


def union_keys_preserve_order(rows):
    seen = []
    sset = set()
    for r in rows:
        for k in r.keys():
            if k not in sset:
                seen.append(k)
                sset.add(k)
    return seen


def autosize_columns(ws):
    col_widths = {}
    for r in ws.iter_rows(values_only=True):
        for idx, val in enumerate(r, start=1):
            txt = "" if val is None else str(val)
            w = max(3, min(80, len(txt) + 2))
            col_widths[idx] = max(col_widths.get(idx, 0), w)
    for idx, width in col_widths.items():
        ws.column_dimensions[chr(64+idx) if idx <= 26 else (chr(64 + (idx-1)//26) + chr(64 + (idx-1)%26 + 1))].width = width


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def enumerate_multiline(text):
    if text is None:
        return ""
    parts = [ln.strip() for ln in str(text).splitlines()]
    parts = [p for p in parts if p]
    out = []
    for i, p in enumerate(parts, 1):
        out.append(f"{i}. {p}")
    return "\n".join(out) if out else ""


def validate_xlsx(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            names = set(zf.namelist())
            required = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
            return required.issubset(names)
    except Exception:
        return False


def get_ist_now():
    ist = timezone(timedelta(hours=5, minutes=30))
    return datetime.now(ist)


def main():
    ensure_json_array(json_data)

    # Normalize schema order
    key_order = union_keys_preserve_order(json_data)

    # Create workbook with a single authoritative sheet 'Data'
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers
    for c, key in enumerate(key_order, start=1):
        ws.cell(row=1, column=c, value=key)

    # Write rows preserving values exactly; fill missing with blank
    for r_idx, row in enumerate(json_data, start=2):
        for c, key in enumerate(key_order, start=1):
            ws.cell(row=r_idx, column=c, value=row.get(key, ""))

    # Base formatting: bold headers, freeze top row, auto filter, auto-fit widths
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)

    # Create META sheet and copy META columns AS-IS
    meta = wb.create_sheet("Meta_data_sheet")
    # Header
    for c, key in enumerate(META_COLS, start=1):
        meta.cell(row=1, column=c, value=key).font = Font(bold=True)
    # Data
    for r_idx, row in enumerate(json_data, start=2):
        for c, key in enumerate(META_COLS, start=1):
            meta.cell(row=r_idx, column=c, value=row.get(key, ""))
    # Very hidden
    meta.sheet_state = 'veryHidden'

    # STEP: Normalize MAIN sheet (operate on same original sheet) -> rename to TestPlan
    ws.title = "TestPlan"

    # Remove META columns from TestPlan and reorder remaining to MAIN_ORDER
    # Build a mapping from header to column index first
    headers = [cell.value for cell in ws[1]]
    # Determine indices to keep in MAIN_ORDER; add missing MAIN columns as empty later
    keep_set = set(MAIN_ORDER)
    # Build temporary table for TestPlan with only MAIN_ORDER columns
    rows_table = []
    # header row
    rows_table.append(MAIN_ORDER[:])
    # data rows
    for r in range(2, ws.max_row + 1):
        row_map = {}
        for c, h in enumerate(headers, start=1):
            val = ws.cell(row=r, column=c).value
            row_map[h] = val
        new_row = [row_map.get(h, "") for h in MAIN_ORDER]
        rows_table.append(new_row)

    # Clear existing sheet and write back
    ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(rows_table, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Format header row: bold, center H/V, blue fill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = HEADER_FILL

    # Enable wrapping for specific columns and set alignments
    wrap_cols = {
        "Test Description",
        "Remarks",
        "Test Steps / Procedure",
        "Validation / Acceptance Criteria",
    }
    # Map header names to column indices
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Numbering inside cells for Test Steps / Procedure & Validation / Acceptance Criteria
    for r in range(2, ws.max_row + 1):
        for h in ["Test Steps / Procedure", "Validation / Acceptance Criteria"]:
            c = header_map.get(h)
            if c:
                orig = ws.cell(row=r, column=c).value
                ws.cell(row=r, column=c, value=enumerate_multiline(orig))

    # Set cell alignments and wrapping rules
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            cell = ws.cell(row=r, column=c)
            if h in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            elif h == "Index":
                cell.alignment = Alignment(vertical='top', horizontal='center')
            else:
                cell.alignment = Alignment(vertical='top', horizontal='left')

    # Re-apply freeze, filter, autosize, borders on TestPlan
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)
    apply_borders(ws)

    # Data validation ONLY for Code Generation (Required / Not)
    code_col = header_map.get("Code Generation (Required / Not)")
    if code_col:
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showDropDown=True)
        dv.error = "Select a value from the list: Required, Blank, Not Required"
        dv.errorTitle = "Invalid Selection"
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=code_col).coordinate}:{ws.cell(row=max(2, ws.max_row), column=code_col).coordinate}")

    # Safety check: only TestPlan (visible) and Meta_data_sheet (veryHidden) must exist
    if any(sh.title == 'Data' for sh in wb.worksheets):
        # Delete any stray 'Data' sheet if present
        for sh in wb.worksheets:
            if sh.title == 'Data':
                wb.remove(sh)

    allowed = {"TestPlan", "Meta_data_sheet"}
    if set(sh.title for sh in wb.worksheets) - allowed:
        raise SystemExit("Unexpected extra worksheets present after normalization")

    # Compute IST timestamp and filename per rule
    ist_now = get_ist_now()
    file_stamp = ist_now.strftime("%Y%m%d_%H%M%S")
    file_name = f"{IP_NAME}_TestPlan_{file_stamp}.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / file_name

    # Save workbook
    wb.save(out_path)

    # Validate XLSX structure
    if not validate_xlsx(out_path):
        raise SystemExit("Generated file failed OOXML validation (fallback should be retried)")

    # Commit ONLY the finalized Excel file
    # Configure git identity (GitHub Actions bot)
    subprocess.run(["git", "config", "user.name", "github-actions[bot]"], check=True)
    subprocess.run(["git", "config", "user.email", "41898282+github-actions[bot]@users.noreply.github.com"], check=True)

    subprocess.run(["git", "add", str(out_path)], check=True)
    commit_msg = COMMIT_MESSAGE_PREFIX.format(stamp=ist_now.strftime("%Y-%m-%d %H:%M:%S"))
    # Commit if there is a change
    rc = subprocess.run(["git", "diff", "--cached", "--quiet"]).returncode
    if rc != 0:
        subprocess.run(["git", "commit", "-m", commit_msg], check=True)
        # Push using the token already configured by actions/checkout
        subprocess.run(["git", "push", "origin", BRANCH], check=True)
        print(f"::notice ::Committed {out_path} with message: {commit_msg}")
    else:
        print("::notice ::No changes to commit (file may already exist with identical content)")

    # Emit summary
    print(f"Generated: {out_path}")
    print(f"Rows: {len(json_data)}")
    print(f"Columns: {len(MAIN_ORDER)} (main) + {len(META_COLS)} (meta)")


if __name__ == "__main__":
    main()
