#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import os
import re
import sys
import zipfile
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# FULL_JSON_STRUCTURE embedded exactly as provided (do not modify)
FULL_JSON_STRUCTURE = r'''[
  {
    "Index": 1,
    "SS / Module": "GPIO",
    "Feature": "Interrupts can be generated based on negative edge detection at GPIO input",
    "Test Case Name": "test_gpio_nedge_random_pads_en",
    "Test Description": "Verify negative-edge interrupt behavior on GPIO inputs for randomly selected pads and ensure interrupt status is set and cleared correctly.",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "0xA0243ffc",
    "Memory End Offset": "0xA0243ffc",
    "Remarks": "Requires system interrupt enable for the selected GPIO block. Uses a flag to synchronize the interrupt handler. Stimulus is applied via a memory-mapped interface. Pads are selected randomly without repetition.",
    "Test Steps / Procedure": "1) Enable the system interrupt for the targeted GPIO block in the system register.\n2) Enable the platform interrupt controller line for the selected GPIO block.\n3) Initialize the stimulus memory location to the default high value.\n4) For each unique pad in the target range, configure the per-pin control register for input with negative-edge interrupt.\n5) Enable the corresponding bit in the GPIO group interrupt enable register for the selected pad.\n6) Drive the input to generate a falling edge using the stimulus interface, then restore it to the default state.\n7) Wait until the interrupt handler clears the pending flag.\n8) In the handler, read the per-pin register and verify the input state and raw interrupt indicator.\n9) Read the GPIO group interrupt status register and verify the bit for the selected pad is set.\n10) Disable the corresponding bit in the GPIO group interrupt enable register and clear the per-pin interrupt through the per-pin control register.\n11) Verify the per-pin status indicates the interrupt is cleared and the GPIO group interrupt status register is cleared.\n12) Clear the system raw status bit in the system register and verify it is cleared.\n13) Clear the platform interrupt controller line and repeat for all selected pads.",
    "Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "1) GPIO input state indicates the expected value after the event → Pass if the input state bit is set as expected.\n2) GPIO raw interrupt bit is asserted after the falling edge → Pass if the per-pin raw interrupt indicator is set.\n3) GPIO group interrupt status reflects the selected pad → Pass if the corresponding group status bit is set.\n4) After clearing, the per-pin interrupt status is cleared → Pass if the per-pin status indicates cleared.\n5) After clearing, the GPIO group interrupt status is zero → Pass if the group status register reads as cleared.\n6) System raw status bit for the GPIO interrupt is cleared → Pass if the system status bit reads as cleared.\n7) No unexpected \"Interrupt Not occurred\" path is taken → Pass if no error increments occur and the test error counter remains zero at finish.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_gpio_nedge_random_pads_en",
    "Hidden_Test_Description": "Enable sysreg interrupts for GPIO0/1. Initialize stimulus at 0xA0243ffc to 0xffffffff. For 32 iterations, pick a unique random pad_num in [0..31] mapping to GPIO_8..GPIO_39. Configure MIZAR_GPIO_GP0_GPIO_8 + (pad_num*4) with 0x00140000 (input mode and negedge interrupt). Enable group interrupt by writing MIZAR_GPIO_GP0_INTR1_INTR_EN1 with (1<<pad_num). Write ~(1<<pad_num) to 0xA0243ffc then 0xffffffff to generate a falling edge. Set int_pend=1 and busy-wait until the ISR clears it. In Default_IRQHandler: set int_pend=0; read the per-pin register; if ((rdata & 0x1) != 0) then OK else error. If ((rdata & 0x2) != 0) then read MIZAR_GPIO_GP0_INTR1_INTR_STS1 and require (bit pad_num) set else error. Disable group enable (write 0 to MIZAR_GPIO_GP0_INTR1_INTR_EN1). Clear per-pin interrupt by writing 0x00110001 to the per-pin register, wait, read back must be 0x100001 else error. Require group status becomes 0x0 else error. Clear sysreg raw status bit by writing MIZAR_LSS_SYSREG_RAW_STCR1 with LSS_SYSREG_RAW_STCR1_GPIO{0|1}_INTR and verify readback bit cleared else error. Clear GIC IRQ line 87 or 88. Else path: if raw interrupt bit not set, print \"Interrupt Not occured\" and increment error. Finish with finish(test_err).",
    "Hidden_Remarks": "Compile-time defines GPIO0 or GPIO1 select which system interrupt bit and GIC line are used. Uses int_pend flag for synchronization between main loop and ISR. Uses address 0xA0243ffc to stimulate input transitions and resets it to 0xffffffff between tests. Tests GPIO pads 8–39 via base MIZAR_GPIO_GP0_GPIO_8 + (pad_num * 4) with unique random selection. Uses wait_on delays (50/10/2) around configuration and stimulus. Clears per-pin raw interrupt by setting the 16th bit via write value 0x00110001 after disabling group interrupt enable.",
    "Hidden_Test_Steps_Procedure": "Initialize: test_err=0. If GPIO0: GIC_EnableIRQ(87) and write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR). If GPIO1: GIC_EnableIRQ(88) and write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR). Seed rand. write_reg(0xA0243ffc, 0xffffffff). For i=0..31: pick pad_num=rand()%32; ensure uniqueness vs arr[0..i-1]; if duplicate, decrement i and continue. For unique pad: arr[i]=pad_num; wr_val=1<<pad_num; write_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num*4), 0x00140000); wait_on(50); write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 1<<pad_num); wait_on(10); write_reg(0xA0243ffc, ~(wr_val)); wait_on(10); write_reg(0xA0243ffc, 0xffffffff); int_pend=1; while(int_pend==1){printf(\"Waiting for interrupt\"); wait_on(10);} End for. Call finish(test_err). Default_IRQHandler: wr_val=1<<pad_num; int_pend=0; rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num*4)). If ((rdata & 0x1)!=0) then OK else print error and test_err++. If ((rdata & 0x2)!=0): rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if ((rdata_grp & (1<<pad_num))!=0) OK else print \"Group Interrupt not occured\" and test_err++. write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000). write_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num*4), 0x00110001); wait_on(2); rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8 + (pad_num*4)); if (rdata==0x100001) OK else print error and test_err++. rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if (rdata_grp==0x0) OK else print error and test_err++. If GPIO0: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); rdata=read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR)==0) OK else print error and test_err++. If GPIO1: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); rdata=read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR)==0) OK else print error and test_err++. Else (if ((rdata & 0x2)==0)): print \"Interrupt Not occured\" and test_err++. Finally, if GPIO0: GIC_ClearIRQ(87); if GPIO1: GIC_ClearIRQ(88).",
    "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR, LSS_SYSREG_INTR_EN1_GPIO1_INTR, MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR, LSS_SYSREG_RAW_STCR1_GPIO1_INTR",
    "Hidden_Validation_Acceptance_Criteria": "In ISR: (rdata & 0x1)!=0 is success else error++. (rdata & 0x2)!=0 indicates raw interrupt raised; else print \"Interrupt Not occured\" and error++. Group status read MIZAR_GPIO_GP0_INTR1_INTR_STS1 must have (1<<pad_num) set; else error++. After write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0x00000000) and per-pin clear write 0x00110001, per-pin readback must equal 0x100001; else error++. Group status must read 0x0; else error++. Sysreg raw status clear: after writing MIZAR_LSS_SYSREG_RAW_STCR1 with LSS_SYSREG_RAW_STCR1_GPIO{0|1}_INTR, the corresponding bit must read cleared; else error++. Final Pass if finish(test_err) is called with test_err==0."
  },
  {
    "Index": 2,
    "SS / Module": "GPIO",
    "Feature": "Interrupts can be generated based on negative edge detection at GPIO input",
    "Test Case Name": "test_gpio_nedge_walking_zeros_pattern",
    "Test Description": "Validate negative-edge interrupt operation across all GPIO pads using a walking zeros stimulus and confirm per-pin and group interrupt behavior including clear sequences.",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "0xA0243ffc",
    "Memory End Offset": "0xA0243ffc",
    "Remarks": "Requires enabling the correct system interrupt for the selected GPIO instance. Uses a shared flag to synchronize with the interrupt handler. Stimulus is driven through a memory-mapped register. Applies a walking zeros pattern across all pads.",
    "Test Steps / Procedure": "1) Enable the system interrupt for the target GPIO instance in the system register.\n2) Enable the corresponding interrupt line in the interrupt controller.\n3) For each pad, configure the per-pin control register to enable negative-edge detection.\n4) Configure the GPIO I/O control group registers to set input mode for all pads under test.\n5) Enable the GPIO group interrupt for all pads through the group interrupt enable register.\n6) Initialize the stimulus register to the default high value.\n7) For each pad index, drive a falling edge for only that pad using the stimulus register and wait for the interrupt.\n8) In the interrupt handler, read the per-pin control/status register and verify input and raw interrupt indicators.\n9) Read the GPIO group interrupt status register and verify the bit for the current pad is set.\n10) Clear the per-pin interrupt status via the per-pin control register and verify the readback value.\n11) Verify the GPIO group interrupt status register is cleared.\n12) Clear the system raw status bit for the GPIO interrupt and verify it is cleared.\n13) Clear the interrupt controller line and proceed to the next pad.",
    "Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "1) Input status after the event → Pass if the input state bit indicates the expected value.\n2) Raw interrupt after the falling edge → Pass if the per-pin raw interrupt indicator is set.\n3) Group interrupt status → Pass if the corresponding bit for the active pad is set.\n4) Per-pin interrupt clear → Pass if the per-pin readback matches the expected cleared value.\n5) Group interrupt clear → Pass if the group status register reads as cleared.\n6) System raw status clear → Pass if the system status bit for the GPIO interrupt is cleared.\n7) No unexpected missing interrupt path → Pass if no error increments occur and the error counter remains zero at completion.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_gpio_nedge_walking_zeros_pattern",
    "Hidden_Test_Description": "Enable GIC for GPIO0 (IRQ 87) or GPIO1 (IRQ 88). Enable system register interrupt via write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR or LSS_SYSREG_INTR_EN1_GPIO1_INTR). For i=0..31: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00040000) to enable negedge (17th bit). wait_on(10). Set input mode for GPIOs 8–39 by writing 0x000000FF to MIZAR_GPIO_GPIO_IO_CTRL_GROUP1..4. wait_on(10). Enable group interrupt: write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, 0xFFFFFFFF). For i=0..31: wr_val=1<<i; write_reg(0xA0243ffc, 0xFFFFFFFF); wait_on(30); write_reg(0xA0243ffc, ~(wr_val)); wait_on(30); int_pend=1; while(int_pend==1){printf(\"Waiting for interrupt\"); wait_on(10);} finish(test_err). In Default_IRQHandler: wr_val=1<<i; int_pend=0; write_reg(0xA0243ffc, 0xFFFFFFFF); rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)); if ((rdata & 0x1)!=0) OK else print error and test_err++. If ((rdata & 0x2)!=0) then rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if ((rdata_grp & (1<<i))!=0) OK else print error and test_err++. Clear per-pin interrupt: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00110001); wait_on(2); rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)); if (rdata==0x100001) OK else error++. rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if (rdata_grp==0x0) OK else error++. If GPIO0: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); read back and verify bit cleared else error++. If GPIO1: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); read back and verify bit cleared else error++. Else (if no raw): print \"Interrupt Not occured\" and test_err++. Finally, GIC_ClearIRQ(87) or GIC_ClearIRQ(88).",
    "Hidden_Remarks": "Compile-time switches GPIO0/GPIO1 select the system interrupt bit and GIC line. A global int_pend flag is used to synchronize main and ISR. The stimulus is driven via memory address 0xA0243ffc using a walking zeros pattern. wait_on delays (10/30/2) are used around configuration and stimulus transitions. GPIO I/O control group registers are set to 0x000000FF to enable input mode for each group.",
    "Hidden_Test_Steps_Procedure": "If GPIO0: GIC_EnableIRQ(87); else if GPIO1: GIC_EnableIRQ(88). Write MIZAR_LSS_SYSREG_INTR_EN1 with LSS_SYSREG_INTR_EN1_GPIO0_INTR (GPIO0) or LSS_SYSREG_INTR_EN1_GPIO1_INTR (GPIO1). For i=0..31: write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00040000) to enable negedge (17th bit). wait_on(10). Write 0x000000FF to MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, GROUP2, GROUP3, GROUP4. wait_on(10). Write MIZAR_GPIO_GP0_INTR1_INTR_EN1 with 0xFFFFFFFF. For i=0..31: wr_val=1<<i; write_reg(0xA0243ffc, 0xFFFFFFFF); wait_on(30); write_reg(0xA0243ffc, ~(wr_val)); wait_on(30); int_pend=1; while(int_pend==1){ printf(\"Waiting for interrupt\"); wait_on(10);} After loop, finish(test_err). Default_IRQHandler: wr_val=1<<i; int_pend=0; write_reg(0xA0243ffc, 0xFFFFFFFF); rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)); if ((rdata & 0x1) != 0) pass else print error and test_err++. if ((rdata & 0x2) != 0) { rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if ((rdata_grp & (1<<i)) != 0) pass else print \"Group Interrupt not occured\" and test_err++; write_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4), 0x00110001); wait_on(2); rdata=read_reg(MIZAR_GPIO_GP0_GPIO_8 + (i*4)); if (rdata==0x100001) pass else error++; rdata_grp=read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if (rdata_grp==0x0) pass else error++; if GPIO0: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); rdata=read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO0_INTR)==0) pass else error++; if GPIO1: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); rdata=read_reg(MIZAR_LSS_SYSREG_RAW_STCR1); if ((rdata & LSS_SYSREG_RAW_STCR1_GPIO1_INTR)==0) pass else error++; } else { printf(\"Interrupt Not occured\"); test_err++; } If GPIO0: GIC_ClearIRQ(87); if GPIO1: GIC_ClearIRQ(88).",
    "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR, LSS_SYSREG_INTR_EN1_GPIO1_INTR, MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR, LSS_SYSREG_RAW_STCR1_GPIO1_INTR",
    "Hidden_Validation_Acceptance_Criteria": "In ISR: (rdata & 0x1)!=0 indicates DIN is as expected; else error++. (rdata & 0x2)!=0 indicates raw interrupt raised; else print \"Interrupt Not occured\" and error++. Group status MIZAR_GPIO_GP0_INTR1_INTR_STS1 must have (1<<i) set; else error++. After clearing per-pin with 0x00110001, readback must equal 0x100001; else error++. Group status must read 0x0; else error++. After writing MIZAR_LSS_SYSREG_RAW_STCR1 with the appropriate GPIO interrupt bit, the corresponding bit must be cleared on readback; else error++. Final Pass if finish(test_err) is called with test_err==0."
  },
  {
    "Index": 3,
    "SS / Module": "GPIO",
    "Feature": "Interrupts can be generated based on negative edge detection at GPIO input",
    "Test Case Name": "test_gpio_negedge_intr_en",
    "Test Description": "Verify negative-edge interrupt operation for GPIO inputs by configuring each pin, generating a falling edge stimulus, and validating per-pin and group interrupt behavior including clear sequences.",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "0xA0243ffc",
    "Memory End Offset": "0xA0243ffc",
    "Remarks": "Uses compile-time selection for the GPIO instance. Synchronizes main and ISR using a shared flag. Stimulus transitions are applied through a memory-mapped location. Per-pin interrupt is cleared by a defined write sequence.",
    "Test Steps / Procedure": "1) Enable the system interrupt for the target GPIO instance in the system register.\n2) Enable the corresponding interrupt line in the interrupt controller.\n3) Initialize the stimulus memory location to the default high value.\n4) For each pad index, configure the per-pin control register for input with negative-edge interrupt enabled.\n5) Enable the corresponding bit in the GPIO group interrupt enable register for the current pad.\n6) Drive the input high, then generate a falling edge for the current pad using the stimulus interface.\n7) Wait until the interrupt handler clears the pending flag.\n8) In the handler, read the per-pin control/status register and verify input state and raw interrupt indication.\n9) Read the GPIO group interrupt status register and verify the bit for the current pad is set.\n10) Clear the per-pin interrupt status through the per-pin control register and verify the readback value.\n11) Verify the GPIO group interrupt status register is cleared.\n12) Clear the system raw status for the GPIO interrupt and clear the interrupt controller line.\n13) Repeat the sequence for all pad indices.",
    "Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "1) Input status after the event → Pass if the input state bit indicates the expected value.\n2) Raw interrupt after the falling edge → Pass if the per-pin raw interrupt indicator is set.\n3) Group interrupt status → Pass if the corresponding bit for the active pad is set.\n4) Per-pin interrupt clear → Pass if the per-pin readback matches the expected cleared value.\n5) Group interrupt clear → Pass if the group status register reads as cleared.\n6) System raw status clear and IRQ line → Pass if the system status is cleared and the line is acknowledged.\n7) No missing-interrupt path taken → Pass if no error increments occur and the final error counter is zero.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_gpio_negedge_intr_en",
    "Hidden_Test_Description": "Initialize: test_err=0. If GPIO0: GIC_EnableIRQ(87) and write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR). If GPIO1: GIC_EnableIRQ(88) and write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR). Write stimulus location 0xA0243ffc = 0xffffffff. For i=0..31: addr1 = MIZAR_GPIO_GP0_GPIO_8 + (i*4); write_reg(addr1, 0x00140000) to program input mode and enable negedge (per comment). wait_on(50). wr_val = 1<<i. Enable group interrupt bit with write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val). wait_on(10). Write 0xA0243ffc = 0xffffffff; wait_on(30); write_reg(0xA0243ffc, ~(wr_val)) to generate falling edge; set int_pend=1; while(int_pend){wait_on(10);} After loop, finish(test_err). Default_IRQHandler: int_pend=0; write_reg(0xA0243ffc, 0xffffffff). raddr = MIZAR_GPIO_GP0_GPIO_8 + (i*4); rdata = read_reg(raddr); If ((rdata & 0x1) != 0x0) pass else test_err++. If ((rdata & 0x2) != 0x0) then rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if ((rdata_grp & (wr_val)) != 0) pass else test_err++; raddr2 = MIZAR_GPIO_GP0_GPIO_8 + (i*4); write_reg(raddr2, 0x00110001); rdata = read_reg(raddr2); if (rdata == 0x100001) pass else test_err++; rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if (rdata_grp == 0x0) pass else test_err++; If GPIO0: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); GIC_ClearIRQ(87). If GPIO1: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); GIC_ClearIRQ(88). Else branch (if (rdata & 0x2) == 0): test_err++.",
    "Hidden_Remarks": "Compile-time defines GPIO0 or GPIO1 select the system interrupt bit and GIC line. A global flag int_pend is used to synchronize main flow and ISR. Stimulus is driven and reset using 0xA0243ffc with high then inverted single-bit patterns. Per-pin raw interrupt is cleared by writing 0x00110001 to the per-pin register and verifying 0x100001 readback. Group interrupt enable uses wr_val with a single active bit per iteration. Delays via wait_on are inserted around configuration and stimulus.",
    "Hidden_Test_Steps_Procedure": "test_err=0; If GPIO0: GIC_EnableIRQ(87); If GPIO1: GIC_EnableIRQ(88); If GPIO0: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR); If GPIO1: write_reg(MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO1_INTR); write_reg(0xA0243ffc, 0xffffffff); For i=0..31: addr1 = MIZAR_GPIO_GP0_GPIO_8 + (i*4); write_reg(addr1, 0x00140000); wait_on(50); wr_val = 1<<i; write_reg(MIZAR_GPIO_GP0_INTR1_INTR_EN1, wr_val); wait_on(10); write_reg(0xA0243ffc, 0xffffffff); wait_on(30); write_reg(0xA0243ffc, ~(wr_val)); int_pend=1; while(int_pend){ wait_on(10);} After loop: finish(test_err). Default_IRQHandler: int_pend=0; write_reg(0xA0243ffc, 0xffffffff); raddr = MIZAR_GPIO_GP0_GPIO_8 + (i*4); rdata = read_reg(raddr); if ((rdata & 0x1) != 0x0) { } else { test_err++; } if ((rdata & 0x2) != 0x0) { rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if ((rdata_grp & (wr_val)) != 0) { } else { test_err++; } raddr2 = MIZAR_GPIO_GP0_GPIO_8 + (i*4); write_reg(raddr2, 0x00110001); rdata = read_reg(raddr2); if (rdata == 0x100001) { } else { test_err++; } rdata_grp = read_reg(MIZAR_GPIO_GP0_INTR1_INTR_STS1); if (rdata_grp == 0x0) { } else { test_err++; } If GPIO0: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO0_INTR); GIC_ClearIRQ(87); If GPIO1: write_reg(MIZAR_LSS_SYSREG_RAW_STCR1, LSS_SYSREG_RAW_STCR1_GPIO1_INTR); GIC_ClearIRQ(88); } else { test_err++; }",
    "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN1, LSS_SYSREG_INTR_EN1_GPIO0_INTR, LSS_SYSREG_INTR_EN1_GPIO1_INTR, MIZAR_GPIO_GP0_GPIO_8, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_INTR1_INTR_STS1, MIZAR_LSS_SYSREG_RAW_STCR1",
    "Hidden_Validation_Acceptance_Criteria": "Pass conditions: (1) (rdata & 0x1) != 0x0 after ISR entry; (2) (rdata & 0x2) != 0x0 indicating raw interrupt; (3) MIZAR_GPIO_GP0_INTR1_INTR_STS1 has bit wr_val set; (4) After write 0x00110001 to per-pin register, readback equals 0x100001; (5) MIZAR_GPIO_GP0_INTR1_INTR_STS1 reads 0x0 after clear; (6) System raw status is cleared by writing MIZAR_LSS_SYSREG_RAW_STCR1 with the GPIO instance bit and acknowledging the IRQ line; (7) No path where raw interrupt is absent (else branch) is taken and test_err remains zero at finish."
  },
  {
    "Index": 4,
    "SS / Module": "GPIO",
    "Feature": "doe field common io mode control register(Group IO control) along with mask bits",
    "Test Case Name": "test_gpio_op_mode_all_pad_en",
    "Test Description": "Enable output mode on GPIO pads 8–39 via group IO control, then for each pad drive output high and low using the per-pin register and verify pad value by reading a status register. Errors are counted if pad value does not match the driven state.",
    "Speed": "NA",
    "Mode": "loops",
    "Memory Start Offset": "0xA0243ffc",
    "Memory End Offset": "0xA0243ffc",
    "Remarks": "Pads 8–39 are configured for output using group IO control registers. All group1 interrupt enables are set but interrupts are not functionally used for validation. Compile-time options GPIO0 or GPIO1 determine which GIC line is enabled and cleared. Pad state is verified by reading a status register at 0xA0243ffc. A default IRQ handler increments error count if triggered.",
    "Test Steps / Procedure": "1) Enable the interrupt line for the selected GPIO instance in the interrupt controller.\n2) Program the GPIO IO control group registers to configure pads 8–39 as outputs.\n3) Enable the GPIO interrupt group for all pads.\n4) For each pad index from 0 to 31, set the per-pin control/status register to drive output high, then verify pad state via the status register; set the per-pin control/status register to drive output low, then verify pad state via the status register.\n5) Clear the interrupt line for the selected GPIO instance after each verification.\n6) Finish the test and report the error count.",
    "Impacted Registers": "GPIO IO control group registers, GPIO interrupt enable register, GPIO per-pin control/status registers, Pad status register (0xA0243ffc)",
    "Validation / Acceptance Criteria": "1) When output is driven high for a pad → The corresponding bit in the status register is set.\n2) When output is driven low for a pad → The corresponding bit in the status register is cleared.\n3) No unexpected interrupt handler entry occurs → The error counter is not incremented by the default handler.\n4) Final test result → Pass if the error counter is zero at test completion.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_gpio_op_mode_all_pad_en",
    "Hidden_Test_Description": "Enable output mode on GPIO pads 8–39 via group IO control, then for each pad drive output high and low using the per-pin register and verify pad value by reading a status register. Errors are counted if pad value does not match the driven state.",
    "Hidden_Remarks": "Pads 8–39 are configured for output using group IO control registers. All group1 interrupt enables are set but interrupts are not functionally used for validation. Compile-time options GPIO0 or GPIO1 determine which GIC line is enabled and cleared. Pad state is verified by reading a status register at 0xA0243ffc. A default IRQ handler increments error count if triggered.",
    "Hidden_Test_Steps_Procedure": "Initialize: test_err = 0. If GPIO0 then GIC_EnableIRQ(87); if GPIO1 then GIC_EnableIRQ(88). Write MIZAR_GPIO_GPIO_IO_CTRL_GROUP1 = 0x00FF00FF; MIZAR_GPIO_GPIO_IO_CTRL_GROUP2 = 0x00FF00FF; MIZAR_GPIO_GPIO_IO_CTRL_GROUP3 = 0x00FF00FF; MIZAR_GPIO_GPIO_IO_CTRL_GROUP4 = 0x00FF00FF. wait_on(10). Write MIZAR_GPIO_GP0_INTR1_INTR_EN1 = 0xFFFFFFFF. For i = 0..31: write MIZAR_GPIO_GP0_GPIO_8 + (i*4) = 0x00200000; set gp0_flag_dout_one = 1; call check_for_pad_value(i). wait_on(20). write MIZAR_GPIO_GP0_GPIO_8 + (i*4) = 0x00000000; set gp0_flag_dout_zero = 1; call check_for_pad_value(i). After loop, finish(test_err). In check_for_pad_value(gpio_pad_num): rdata = read_reg(0xA0243ffc). If gp0_flag_dout_one == 1: if ((rdata & (1 << gpio_pad_num)) != 0) then success else print error and test_err++; set gp0_flag_dout_one = 0. If gp0_flag_dout_zero == 1: if ((rdata & (1 << gpio_pad_num)) == 0) then success else print error and test_err++; set gp0_flag_dout_zero = 0. If GPIO0 then GIC_ClearIRQ(87); if GPIO1 then GIC_ClearIRQ(88). Default_IRQHandler: print error message and increment test_err.",
    "Hidden_Impacted_Registers": "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1, MIZAR_GPIO_GPIO_IO_CTRL_GROUP2, MIZAR_GPIO_GPIO_IO_CTRL_GROUP3, MIZAR_GPIO_GPIO_IO_CTRL_GROUP4, MIZAR_GPIO_GP0_INTR1_INTR_EN1, MIZAR_GPIO_GP0_GPIO_8",
    "Hidden_Validation_Acceptance_Criteria": "If gp0_flag_dout_one == 1 then require (read_reg(0xA0243ffc) & (1 << gpio_pad_num)) != 0 → pass; else log error and increment test_err. If gp0_flag_dout_zero == 1 then require (read_reg(0xA0243ffc) & (1 << gpio_pad_num)) == 0 → pass; else log error and increment test_err. Default_IRQHandler triggers → increment test_err. Final pass if finish(test_err) is called with test_err == 0."
  }
]'''

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)',
]


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--ip-name', required=True)
    p.add_argument('--output-dir', required=True)
    return p.parse_args()


def load_json():
    try:
        data = json.loads(FULL_JSON_STRUCTURE)
    except Exception as e:
        sys.exit(f'Invalid JSON input: {e}')
    if not isinstance(data, list) or not data:
        sys.exit('Invalid JSON: must be a non-empty array')
    return data


def union_keys(rows):
    seen = []
    for o in rows:
        for k in o.keys():
            if k not in seen:
                seen.append(k)
    return seen


def create_base_workbook(rows, keys):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    # Header
    for c, key in enumerate(keys, 1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = Font(bold=True)
    # Rows
    for r, obj in enumerate(rows, 2):
        for c, key in enumerate(keys, 1):
            ws.cell(row=r, column=c, value=obj.get(key, ''))
    # Freeze top row
    ws.freeze_panes = 'A2'
    # Autofit columns (approx)
    for c, key in enumerate(keys, 1):
        maxlen = len(str(key))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            for ln in str(v).split('\n'):
                maxlen = max(maxlen, len(ln))
        ws.column_dimensions[get_column_letter(c)].width = min(max(10, maxlen + 2), 100)
    return wb, ws


def add_meta_sheet(wb, rows):
    ws = wb.create_sheet('Meta_data_sheet')
    for c, key in enumerate(META_COLS, 1):
        ws.cell(row=1, column=c, value=key).font = Font(bold=True)
    for r, obj in enumerate(rows, 2):
        for c, key in enumerate(META_COLS, 1):
            ws.cell(row=r, column=c, value=obj.get(key, ''))
    ws.sheet_state = 'veryHidden'


def normalize_testplan(ws):
    # Extract
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    data_rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i+1).value for i, h in enumerate(headers)}
        data_rows.append(row)

    # Rename in place
    ws.title = 'TestPlan'
    # Clear and write MAIN_ORDER
    ws.delete_rows(1, ws.max_row)
    for c, key in enumerate(MAIN_ORDER, 1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = Font(bold=True)

    # Data rows in order, enforce numbering
    for r, obj in enumerate(data_rows, 2):
        for c, key in enumerate(MAIN_ORDER, 1):
            val = obj.get(key, '')
            if key in ('Test Steps / Procedure', 'Validation / Acceptance Criteria') and isinstance(val, str):
                def fixnum(line: str) -> str:
                    return re.sub(r'^(\s*)(\d+)\)', r'\1\2.', line)
                val = '\n'.join(fixnum(ln) for ln in val.split('\n'))
            ws.cell(row=r, column=c, value=val)

    # Formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    left_top = Alignment(horizontal='left', vertical='top', wrap_text=False)
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header styling
    for c in range(1, len(MAIN_ORDER) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    # Data rows styling
    idx_col = MAIN_ORDER.index('Index') + 1
    for r in range(2, ws.max_row + 1):
        for c, key in enumerate(MAIN_ORDER, 1):
            cell = ws.cell(row=r, column=c)
            if key in ('Test Description', 'Remarks', 'Test Steps / Procedure', 'Validation / Acceptance Criteria'):
                cell.alignment = left_wrap
            elif c == idx_col:
                cell.alignment = center
            else:
                cell.alignment = left_top
            cell.border = border

    # Autofit columns and adjust row heights
    for c, key in enumerate(MAIN_ORDER, 1):
        maxlen = len(str(key))
        maxlines = 1
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            lines = s.split('\n')
            maxlines = max(maxlines, len(lines))
            for ln in lines:
                maxlen = max(maxlen, len(ln))
        ws.column_dimensions[get_column_letter(c)].width = min(max(12, maxlen + 2), 100)
        if maxlines > 1:
            for r in range(2, ws.max_row + 1):
                ws.row_dimensions[r].height = max(15, 15 * maxlines)

    # Data validation for Code Generation (Required / Not)
    code_col = MAIN_ORDER.index('Code Generation (Required / Not)') + 1
    if ws.max_row >= 2:
        dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=False, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(code_col)}2:{get_column_letter(code_col)}{ws.max_row}")

    ws.freeze_panes = 'A2'


def ensure_no_data_sheet(wb):
    for w in list(wb.worksheets):
        if w.title == 'Data':
            wb.remove(w)


def save_and_validate(wb, ip_name: str, out_dir: str) -> str:
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%Y%m%d_%H%M%S')
    fname = f"{ip_name}_TestPlan_{ts}.xlsx"
    fpath = os.path.join(out_dir, fname)
    wb.save(fpath)
    # Validate OOXML
    if not zipfile.is_zipfile(fpath):
        sys.exit('XLSX validation failed: not a ZIP-based OOXML file')
    _ = load_workbook(fpath)
    print(f"Generated: {fpath}")
    return fpath


def main():
    args = parse_args()
    rows = load_json()
    keys = union_keys(rows)
    wb, ws = create_base_workbook(rows, keys)
    add_meta_sheet(wb, rows)
    normalize_testplan(ws)
    ensure_no_data_sheet(wb)
    path = save_and_validate(wb, args.ip_name, args.output_dir)
    print(path)

if __name__ == '__main__':
    main()
