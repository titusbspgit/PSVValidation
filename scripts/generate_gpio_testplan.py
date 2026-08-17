#!/usr/bin/env python3
# Deterministic fallback generator for GPIO TestPlan Excel
# - Creates a real .xlsx with two sheets: TestPlan (visible) and MetaData (very hidden)
# - Preserves row order and data exactly
# - Applies formatting (bold headers, header fill, wrap text, freeze first row, column widths)
# - Filenames include IST timestamp per <IP_NAME>_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx

import json
import os
from datetime import datetime, timedelta, timezone
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:
    ZoneInfo = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Inputs (static for this automation)
OWNER = "titusbspgit"
REPO = "PSVValidation"
BRANCH = "main"
OUTPUT_DIRECTORY = "Test_Output/GPIO/TestPlan/"
IP_NAME = "GPIO"
FINAL_JSON_PATH = "data/final_json_gpio.txt"

# Read EXACT aggregated JSON from file (DO NOT MODIFY CONTENT)
with open(FINAL_JSON_PATH, "r", encoding="utf-8") as f:
    final_json_str = f.read()

# Parse JSON for TestPlan sheet (raises if invalid)
records = json.loads(final_json_str)

# Determine IST timestamp
if ZoneInfo is not None:
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
else:
    # Fallback: approximate IST by adding 19800 seconds (5.5 hours) to UTC
    now_ist = datetime.utcnow().replace(tzinfo=timezone.utc) + timedelta(seconds=19800)

stamp = now_ist.strftime("%Y%m%d_%H%M%S")
filename = f"{IP_NAME}_TestPlan_{stamp}.xlsx"
output_dir = OUTPUT_DIRECTORY
output_path = os.path.join(output_dir, filename)

# Ensure output directory exists
os.makedirs(output_dir, exist_ok=True)

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "TestPlan"

# Headers from first record (preserve order)
headers = list(records[0].keys())

header_font = Font(bold=True)
header_fill = PatternFill(fill_type='solid', start_color='FFDDEBF7', end_color='FFDDEBF7')
wrap = Alignment(wrap_text=True, vertical='top')

# Write header row
for col_idx, key in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=key)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

# Write data rows preserving order (values as strings)
for row_idx, rec in enumerate(records, start=2):
    for col_idx, key in enumerate(headers, start=1):
        val = rec.get(key, "")
        c = ws.cell(row=row_idx, column=col_idx, value=str(val))
        c.alignment = wrap

# Freeze first row
ws.freeze_panes = 'A2'

# Reasonable column widths
for col_idx, key in enumerate(headers, start=1):
    max_len = len(str(key))
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        l = len(str(v))
        if l > max_len:
            max_len = l
    width = max(15, min(max_len + 2, 100))
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# MetaData very hidden sheet with exact JSON string
ws_meta = wb.create_sheet("MetaData")
ws_meta.cell(row=1, column=1, value=final_json_str).alignment = wrap
# Optionally include context (owner/repo/branch/path)
ws_meta.cell(row=2, column=1, value=f"owner: {OWNER}")
ws_meta.cell(row=3, column=1, value=f"repo: {REPO}")
ws_meta.cell(row=4, column=1, value=f"branch: {BRANCH}")
ws_meta.cell(row=5, column=1, value=f"output_path: {output_path}")
try:
    ws_meta.sheet_state = 'veryHidden'
except Exception:
    pass

# Save workbook
wb.save(output_path)

print(f"Generated: {output_path}")
