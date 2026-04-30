#!/usr/bin/env python3
import argparse, json, os, sys, zipfile, datetime
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as e:
    print(f"FATAL: openpyxl not available: {e}", file=sys.stderr)
    sys.exit(2)

META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

DV_COL = "Code Generation (Required / Not)"
DV_ALLOWED = ["Required", "Blank", "Not Required"]

BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)


def to_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        # Preserve raw order and content without numbering
        return "\n".join(str(x) if x is not None else "" for x in v)
    return str(v)


def to_numbered(v: Any) -> str:
    # Number only if value is a list or contains newline-separated items
    if isinstance(v, list):
        items = [str(x) for x in v]
    else:
        s = to_text(v)
        if "\n" in s:
            items = s.split("\n")
        else:
            items = [s] if s else []
    return "\n".join([f"{i+1}. {item}" for i, item in enumerate(items)])


def compute_schema(rows: List[Dict[str, Any]]) -> List[str]:
    seen = []
    sset = set()
    for row in rows:
        for k in row.keys():
            if k not in sset:
                sset.add(k)
                seen.append(k)
    return seen


def auto_width(ws):
    # Approximate auto-fit based on max length in each column
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            for line in s.split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)


def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def normalize_main_sheet(ws, headers: List[str]):
    # Reorder and remove META columns on the SAME sheet; rename to TestPlan before this.
    # Build a new 2D array in memory then write back into same sheet.
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    old_headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]
    # Map old header index
    idx_map = {h: i for i, h in enumerate(old_headers)}
    # Prepare new header row
    new_headers = [h for h in MAIN_ORDER if h in headers]
    # Create new grid
    new_grid = [new_headers]
    for r in data_rows:
        new_row = []
        for h in new_headers:
            i = idx_map.get(h, None)
            val = r[i] if (i is not None and i < len(r)) else None
            # Numbering for specific fields
            if h in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                new_row.append(to_numbered(val))
            else:
                new_row.append(val if val is not None else "")
        new_grid.append(new_row)
    # Clear existing ws
    for _ in range(ws.max_row):
        ws.delete_rows(1)
    # Write new grid
    for ri, row in enumerate(new_grid, 1):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-json", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--ip-name", required=True)
    ap.add_argument("--timestamp", required=False)
    args = ap.parse_args()

    with open(args.input_json, "r", encoding="utf-8") as f:
        raw = f.read()
        try:
            data = json.loads(raw)
        except Exception as e:
            print(f"ERROR: Invalid JSON input: {e}", file=sys.stderr)
            sys.exit(1)

    # Accept array or mapping {TC1: {...}, TC2: {...}}
    if isinstance(data, dict):
        # Preserve insertion order
        rows = [data[k] for k in data.keys()]
    elif isinstance(data, list):
        rows = data
    else:
        print("ERROR: JSON must be an array or an object mapping to row objects", file=sys.stderr)
        sys.exit(1)

    if not rows:
        print("ERROR: Empty JSON array", file=sys.stderr)
        sys.exit(1)

    # Normalize schema
    headers = compute_schema(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = BLUE_FILL

    # Write rows preserving values exactly (lists become newline-joined text here only)
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            v = row.get(h, "")
            ws.cell(row=ri, column=ci, value=to_text(v))

    ws.freeze_panes = "A2"
    auto_width(ws)

    # Create META sheet and copy META columns as-is, unnumbered
    meta_ws = wb.create_sheet("Meta_data_sheet")
    for ci, h in enumerate(META_COLUMNS, 1):
        meta_ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(META_COLUMNS, 1):
            meta_ws.cell(row=ri, column=ci, value=to_text(row.get(h, "")))
    # Very Hidden
    meta_ws.sheet_state = 'veryHidden'

    # Normalize MAIN sheet: rename Data->TestPlan, remove META columns, reorder to MAIN_ORDER on same sheet
    ws.title = "TestPlan"
    normalize_main_sheet(ws, headers)

    # Apply formatting to TestPlan
    # Header row style (already bold+blue for headers placed now via normalize), ensure again
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = BLUE_FILL
    # Wrap text for specific columns
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for h in WRAP_COLS:
        if h in header_map:
            col_idx = header_map[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    # Default align for data rows
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).alignment is None or not ws.cell(row=r, column=c).alignment.wrap_text:
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top", horizontal="left")
    auto_width(ws)

    # Borders for all populated cells
    apply_borders(ws)

    # Data validation for DV_COL
    if DV_COL in header_map:
        col_idx = header_map[DV_COL]
        # Create a list validation "Required,Blank,Not Required"
        dv = DataValidation(type="list", formula1='"' + ",".join(DV_ALLOWED) + '"', allow_blank=True)
        dv.error = "Select a value from the list"
        dv.prompt = "Choose: Required, Blank, or Not Required"
        ws.add_data_validation(dv)
        rng = openpyxl.utils.get_column_letter(col_idx) + "2:" + openpyxl.utils.get_column_letter(col_idx) + str(ws.max_row)
        dv.add(rng)

    # Safety check: only TestPlan and Meta_data_sheet
    if any(s.title == "Data" for s in wb.worksheets):
        print("ERROR: Sheet named 'Data' remains after normalization", file=sys.stderr)
        sys.exit(3)

    # Save
    os.makedirs(args.output_dir, exist_ok=True)
    if args.timestamp:
        ts = args.timestamp
    else:
        # IST timestamp
        ts = (datetime.datetime.utcnow() + datetime.timedelta(hours=5, minutes=30)).strftime('%Y%m%d_%H%M%S')
    out_name = f"{args.ip_name}_TestPlan_{ts}.xlsx"
    out_path = os.path.join(args.output_dir, out_name)

    wb.save(out_path)

    # Validate ZIP-based OOXML
    if not zipfile.is_zipfile(out_path):
        print("ERROR: Saved file is not a valid XLSX (ZIP) container", file=sys.stderr)
        sys.exit(4)

    # Smoke load
    try:
        openpyxl.load_workbook(out_path)
    except Exception as e:
        print(f"ERROR: Unable to reload workbook: {e}", file=sys.stderr)
        sys.exit(5)

    print(out_path)

if __name__ == "__main__":
    main()
