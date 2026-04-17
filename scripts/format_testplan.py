#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

META_HEADERS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_HEADERS_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

NUMERIC_CENTER_COLUMNS = {"Index", "Memory Start Offset", "Memory End Offset"}


def detect_main_sheet(wb):
    # Prefer a visible sheet named 'TestPlan', else first visible sheet
    if "TestPlan" in wb.sheetnames and wb["TestPlan"].sheet_state == "visible":
        return wb["TestPlan"]
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") == "visible":
            return ws
    # Fallback to active
    return wb.active


def header_map(ws):
    m = {}
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            m[str(v).strip()] = col
    return m


def copy_meta_sheet(wb, main_ws):
    if "Meta_data_sheet" in wb.sheetnames:
        wb.remove(wb["Meta_data_sheet"])
    meta_ws = wb.create_sheet("Meta_data_sheet")

    hmap = header_map(main_ws)
    write_col = 1
    for col_name in META_HEADERS:
        if col_name in hmap:
            src_col = hmap[col_name]
            # Write header
            meta_ws.cell(row=1, column=write_col, value=col_name)
            # Copy all rows from main sheet for that column
            for r in range(2, main_ws.max_row + 1):
                meta_ws.cell(row=r, column=write_col, value=main_ws.cell(row=r, column=src_col).value)
            write_col += 1
    # Very hidden
    meta_ws.sheet_state = "veryHidden"


def build_testplan_sheet(wb, main_ws):
    # Create a new working sheet to enforce column order and drop disallowed columns
    temp_name = "__TMP_TESTPLAN__"
    if temp_name in wb.sheetnames:
        wb.remove(wb[temp_name])
    new_ws = wb.create_sheet(temp_name)

    hmap = header_map(main_ws)

    # Determine which of the allowed MAIN headers are present, preserving the prescribed order
    present_headers = [h for h in MAIN_HEADERS_ORDER if h in hmap]

    # Write headers in the specified order (only those present)
    for c, h in enumerate(present_headers, start=1):
        new_ws.cell(row=1, column=c, value=h)

    # Copy row values without modification
    for r in range(2, main_ws.max_row + 1):
        for c, h in enumerate(present_headers, start=1):
            src_col = hmap[h]
            src_cell = main_ws.cell(row=r, column=src_col)
            dst_cell = new_ws.cell(row=r, column=c, value=src_cell.value)
            # Preserve number format (visual; value is unchanged)
            dst_cell.number_format = src_cell.number_format

    # Replace old main sheet with new one named 'TestPlan'
    # Remove original Meta_data_sheet is handled elsewhere; here we only swap the main
    old_title = main_ws.title
    wb.remove(main_ws)
    new_ws.title = "TestPlan"
    return wb["TestPlan"]


def apply_formatting(ws):
    # Header formatting
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_row = ws.max_row
    max_col = ws.max_column

    # Determine column header names for alignment rules
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]

    # Column index by header
    col_by_header = {str(h).strip(): idx for idx, h in enumerate(headers, start=1) if h is not None}

    # Set wrap_text for specific columns
    for name in WRAP_COLUMNS:
        if name in col_by_header:
            col = col_by_header[name]
            for r in range(1, max_row + 1):
                ws.cell(row=r, column=col).alignment = Alignment(
                    horizontal=("center" if r == 1 else "left"),
                    vertical=("center" if r == 1 else "top"),
                    wrap_text=True,
                )

    # Header row formatting and borders
    thin = Side(style="thin", color="000000")
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = all_border

    # Data rows alignment and borders
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            hdr = ws.cell(row=1, column=c).value
            if hdr in NUMERIC_CENTER_COLUMNS:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=cell.alignment.wrap_text if cell.alignment else False)
            else:
                # If previously set wrap_text for specific columns, keep it; else default left/top
                wrap = cell.alignment.wrap_text if cell.alignment else False
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
            cell.border = all_border

    # Approximate autofit for columns
    for c in range(1, max_col + 1):
        max_length = 0
        col_letter = get_column_letter(c)
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            text = str(val)
            # Estimate visual length; headers a bit larger
            length = len(text)
            if r == 1:
                length += 2
            if length > max_length:
                max_length = length
        # convert char length to Excel column width (roughly)
        width = min(max(10, max_length + 2), 80)
        ws.column_dimensions[col_letter].width = width

    # Approximate row height after wrapping: estimate lines per wrapped columns
    base_height = 15  # default approx
    for r in range(2, max_row + 1):
        max_lines = 1
        for name in WRAP_COLUMNS:
            if name in col_by_header:
                c = col_by_header[name]
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                text = str(val)
                # Estimate chars per line based on set column width
                col_letter = get_column_letter(c)
                col_width = ws.column_dimensions[col_letter].width or 10
                chars_per_line = max(int(col_width - 2), 5)
                est_lines = 1
                if chars_per_line > 0:
                    est_lines = (len(text) // chars_per_line) + 1
                # account for explicit line breaks
                est_lines = max(est_lines, text.count("\n") + 1)
                if est_lines > max_lines:
                    max_lines = est_lines
        ws.row_dimensions[r].height = base_height * max_lines


def process(input_path: str, output_path: str):
    if not input_path.lower().endswith(".xlsx"):
        raise ValueError("Input file must be .xlsx")

    wb = load_workbook(input_path)

    main_ws = detect_main_sheet(wb)

    # Build/refresh Meta sheet from the current main sheet BEFORE reordering/removing columns
    copy_meta_sheet(wb, main_ws)

    # Rebuild the main sheet to contain only allowed MAIN columns in required order
    testplan_ws = build_testplan_sheet(wb, main_ws)

    # Apply formatting strictly to TestPlan only
    apply_formatting(testplan_ws)

    # Ensure Meta_data_sheet remains very hidden and no formatting applied
    if "Meta_data_sheet" in wb.sheetnames:
        wb["Meta_data_sheet"].sheet_state = "veryHidden"

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Format TestPlan Excel deterministically (no data changes)")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    process(args.input, args.output)


if __name__ == "__main__":
    main()
