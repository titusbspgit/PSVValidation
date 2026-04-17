import os, math
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font

FILE_PATH = os.getenv('FILE_PATH', 'Test_Output/GPIO/TestPlan/GPIO_TestPlan_1.xlsx')
OUTPUT_FILE_PATH = os.getenv('OUTPUT_FILE_PATH', FILE_PATH)
OUTPUT_FILE_NAME = os.getenv('OUTPUT_FILE_NAME', os.path.basename(OUTPUT_FILE_PATH))
COMMIT_MESSAGE = os.getenv('COMMIT_MESSAGE', ' TestPlan Generated and Pushed to Github')

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

WRAP_COLS = [
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria',
]

NUMERIC_COL_NAMES = set([
    'Index',
    'Speed',
    'Memory Start Offset',
    'Memory End Offset',
])

def find_primary_visible_sheet(wb):
    for ws in wb.worksheets:
        if getattr(ws, 'sheet_state', 'visible') == 'visible':
            return ws
    return wb.active


def header_map(ws):
    m = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip() != '':
            m[v] = c
    return m


def copy_meta_sheet(wb, main_ws):
    # Remove existing Meta_data_sheet if present to avoid duplicates
    if 'Meta_data_sheet' in wb.sheetnames:
        std = wb['Meta_data_sheet']
        wb.remove(std)
    meta_ws = wb.create_sheet('Meta_data_sheet')

    hmap = header_map(main_ws)
    write_col = 1
    for name in META_COLS:
        if name in hmap:
            src_col = hmap[name]
            # Header
            meta_ws.cell(row=1, column=write_col, value=name)
            # Values
            for r in range(2, main_ws.max_row + 1):
                meta_ws.cell(row=r, column=write_col, value=main_ws.cell(row=r, column=src_col).value)
            write_col += 1
    # Very hidden
    meta_ws.sheet_state = 'veryHidden'


def remove_meta_from_main(main_ws):
    hmap = header_map(main_ws)
    cols = [hmap[name] for name in META_COLS if name in hmap]
    for idx in sorted(cols, reverse=True):
        main_ws.delete_cols(idx, 1)


def reorder_main_to_allowed(main_ws):
    # Rename main to TestPlan
    main_ws.title = 'TestPlan'

    # Capture data for allowed headers
    hmap = header_map(main_ws)
    selected = [h for h in MAIN_ORDER if h in hmap]

    # Extract data rows
    data_rows = []
    for r in range(2, main_ws.max_row + 1):
        row_vals = [main_ws.cell(row=r, column=hmap[h]).value for h in selected]
        data_rows.append(row_vals)

    # Clear current sheet content
    if main_ws.max_column:
        main_ws.delete_cols(1, main_ws.max_column)

    # Write headers
    for c, h in enumerate(selected, start=1):
        main_ws.cell(row=1, column=c, value=h)

    # Write data
    for r_idx, row_vals in enumerate(data_rows, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            main_ws.cell(row=r_idx, column=c_idx, value=val)


def apply_formatting(main_ws):
    # Build maps again after reordering
    hmap = header_map(main_ws)

    # Header styling
    for c in range(1, main_ws.max_column + 1):
        cell = main_ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Determine wrap columns and numeric columns by header name
    wrap_cols_idx = [hmap[h] for h in WRAP_COLS if h in hmap]
    numeric_cols_idx = [hmap[h] for h in NUMERIC_COL_NAMES if h in hmap]

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Data rows alignment and borders
    for r in range(2, main_ws.max_row + 1):
        for c in range(1, main_ws.max_column + 1):
            cell = main_ws.cell(row=r, column=c)
            # Vertical top for all data rows
            align_kwargs = {'vertical': 'top'}
            # Horizontal alignment
            if c in numeric_cols_idx or isinstance(cell.value, (int, float)):
                align_kwargs['horizontal'] = 'center'
            else:
                align_kwargs['horizontal'] = 'left'
            # Wrap for specified columns
            if c in wrap_cols_idx:
                align_kwargs['wrap_text'] = True
            cell.alignment = Alignment(**align_kwargs)
            # Borders for all populated cells (including header handled later)
            if cell.value is not None or r == 2:  # ensure first data row gets borders even if empty
                cell.border = border

    # Borders for header row
    for c in range(1, main_ws.max_column + 1):
        cell = main_ws.cell(row=1, column=c)
        cell.border = border

    # Autofit columns based on max text length
    col_widths = {}
    for c in range(1, main_ws.max_column + 1):
        max_len = 0
        for r in range(1, main_ws.max_row + 1):
            v = main_ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        # approximate width in characters
        width = max(10, min(80, max_len + 2))
        col_letter = get_column_letter(c)
        main_ws.column_dimensions[col_letter].width = width
        col_widths[c] = width

    # Autofit row heights after wrapping
    base_h = 15
    for r in range(2, main_ws.max_row + 1):
        max_lines = 1
        for c in wrap_cols_idx:
            v = main_ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            width = col_widths.get(c, 10)
            # estimate characters per line roughly equals column width
            per_line = max(1, int(width))
            est_lines = max(s.count('\n') + 1, math.ceil(len(s) / per_line))
            if est_lines > max_lines:
                max_lines = est_lines
        main_ws.row_dimensions[r].height = base_h * max_lines


def main():
    # STEP 1 + 2: Fetch/validate handled by GitHub checkout; confirm .xlsx via extension
    if not FILE_PATH.lower().endswith('.xlsx'):
        raise SystemExit('File is not .xlsx: ' + FILE_PATH)
    if not os.path.exists(FILE_PATH):
        raise SystemExit('File not found: ' + FILE_PATH)

    wb = load_workbook(FILE_PATH)
    main_ws = find_primary_visible_sheet(wb)

    # STEP 3 + 4: Meta sheet
    copy_meta_sheet(wb, main_ws)

    # STEP 5: Normalize main sheet
    remove_meta_from_main(main_ws)
    reorder_main_to_allowed(main_ws)

    # STEP 5A: Formatting only on TestPlan
    apply_formatting(main_ws)

    # STEP 6: Save workbook
    # Ensure directory exists
    out_dir = os.path.dirname(OUTPUT_FILE_PATH)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    wb.save(OUTPUT_FILE_PATH)

if __name__ == '__main__':
    main()
