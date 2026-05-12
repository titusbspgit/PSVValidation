import json, os, sys, zipfile
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:
    from tzdata import ZoneInfo  # type: ignore
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

IP_NAME = os.environ.get('IP_NAME', 'GPIO')
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', 'Test_Output/GPIO/TestPlan')
OUTPUT_FILENAME_PREFIX = os.environ.get('OUTPUT_FILENAME_PREFIX', f'{IP_NAME}_TestPlan')
JSON_PATH = 'data/gpio_full_json.json'

MAIN_COL_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]

WRAP_COLS = [
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
]

VALIDATION_COL = 'Code Generation (Required / Not)'
VALIDATION_LIST = 'Required,Blank, Not Required'.replace(' ', '')  # keep commas tight for Excel list
VALIDATION_LIST = 'Required,Blank,Not Required'

def fail(msg: str):
    print(f'[ERROR] {msg}', file=sys.stderr)
    sys.exit(1)

# STEP 1: Validate JSON Input
if not os.path.exists(JSON_PATH):
    fail(f'Missing JSON file: {JSON_PATH}')
with open(JSON_PATH, 'r', encoding='utf-8') as f:
    try:
        data = json.load(f)
    except Exception as e:
        fail(f'Invalid JSON: {e}')
if not isinstance(data, list) or len(data) == 0:
    fail('JSON must be a non-empty array of objects')

# STEP 2: Normalize Tabular Schema (preserve first-seen key order union)
ordered_keys = []
seen = set()
for row in data:
    if not isinstance(row, dict):
        fail('Each JSON array item must be an object')
    for k in row.keys():
        if k not in seen:
            seen.add(k)
            ordered_keys.append(k)

# STEP 3: Generate Base Excel Workbook with single sheet named Data
wb = Workbook()
ws = wb.active
ws.title = 'Data'

# Headers
for col_idx, key in enumerate(ordered_keys, start=1):
    cell = ws.cell(row=1, column=col_idx, value=key)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
# Freeze header
ws.freeze_panes = 'A2'

# Rows
for r_idx, row in enumerate(data, start=2):
    for c_idx, key in enumerate(ordered_keys, start=1):
        ws.cell(row=r_idx, column=c_idx, value=row.get(key, ''))

# Base auto width (rough estimate)
def set_column_widths(sheet):
    for c_idx, key in enumerate(ordered_keys, start=1):
        max_len = len(str(key))
        for r in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r, column=c_idx).value
            if val is None:
                continue
            val_len = len(str(val))
            if val_len > max_len:
                max_len = val_len
        sheet.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 80)

set_column_widths(ws)

# STEP 5: Create META sheet and copy META columns as-is
meta = wb.create_sheet('Meta_data_sheet')
for c_idx, key in enumerate(META_COLS, start=1):
    meta.cell(row=1, column=c_idx, value=key).font = Font(bold=True)

for r_idx, row in enumerate(data, start=2):
    for c_idx, key in enumerate(META_COLS, start=1):
        meta.cell(row=r_idx, column=c_idx, value=row.get(key, ''))

# Hide META as Very Hidden
meta.sheet_state = 'veryHidden'

# STEP 7: Normalize MAIN sheet in-place: remove META cols and reorder to MAIN_COL_ORDER
# Build a mapping from current headers to column indices
header_to_index = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

# Extract current data as dict rows to avoid in-sheet column deletion complexity
rows_dict = []
for r in range(2, ws.max_row + 1):
    d = {}
    for key, col in header_to_index.items():
        d[key] = ws.cell(row=r, column=col).value
    rows_dict.append(d)

# Rewrite the same worksheet with only MAIN_COL_ORDER columns
ws.delete_rows(1, ws.max_row)
for c_idx, key in enumerate(MAIN_COL_ORDER, start=1):
    ws.cell(row=1, column=c_idx, value=key)
for r_idx, d in enumerate(rows_dict, start=2):
    for c_idx, key in enumerate(MAIN_COL_ORDER, start=1):
        ws.cell(row=r_idx, column=c_idx, value=d.get(key, ''))

# Rename Data -> TestPlan
ws.title = 'TestPlan'

# STEP 7A: Strict formatting
header_font = Font(bold=True)
header_fill = PatternFill(fill_type='solid', fgColor='4472C4')  # blue
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
left = Alignment(horizontal='left', vertical='top', wrap_text=False)
right = Alignment(horizontal='right', vertical='top', wrap_text=False)
thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header formatting
for c_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=c_idx)
    cell.font = header_font
    cell.alignment = center
    cell.fill = header_fill
    cell.border = border

# In-cell numbering for two columns
wrap_headers = set(WRAP_COLS)
hdr_to_col = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

def ensure_numbering(text):
    if text is None:
        return ''
    s = str(text)
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    if not lines:
        return s
    # If already numbered 1. / 1) etc, keep as-is; else add numbering
    prefixed = True
    for ln in lines:
        parts = ln.split(maxsplit=1)
        if not parts:
            continue
        head = parts[0]
        if not (head.rstrip('.)').isdigit()):
            prefixed = False
            break
    if prefixed:
        return '\n'.join(lines)
    return '\n'.join([f"{i+1}. {ln}" for i, ln in enumerate(lines)])

# Data rows formatting and numbering
for r in range(2, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        cell = ws.cell(row=r, column=c)
        if header in WRAP_COLS:
            cell.value = ensure_numbering(cell.value) if header in [
                'Test Steps / Procedure', 'Validation / Acceptance Criteria'] else cell.value
            cell.alignment = left_wrap
        elif header == 'Index':
            cell.alignment = center
        elif isinstance(cell.value, (int, float)):
            cell.alignment = right
        else:
            cell.alignment = left
        cell.border = border

# Adjust column widths (recompute on final layout)
for c in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=c).value
    max_len = len(str(header))
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=c).value
        if val is None:
            continue
        val_len = max(len(str(val_part)) for val_part in str(val).splitlines())
        if val_len > max_len:
            max_len = val_len
    ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 100)

# Apply data validation ONLY to Code Generation (Required / Not)
if VALIDATION_COL in hdr_to_col:
    col_idx = hdr_to_col[VALIDATION_COL]
    col_letter = get_column_letter(col_idx)
    dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

# STEP 7B: Enforce final sheet visibility (only TestPlan visible, Meta_data_sheet veryHidden)
for s in list(wb.sheetnames):
    if s not in ['TestPlan', 'Meta_data_sheet']:
        del wb[s]

# STEP 8: Save with IST timestamped filename
ist = ZoneInfo('Asia/Kolkata')
now = datetime.now(ist)
ts = now.strftime('%Y%m%d_%H%M%S')
filename = f"{OUTPUT_FILENAME_PREFIX}_{ts}.xlsx"
out_path = os.path.join(OUTPUT_DIR, filename)
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(out_path)

# STEP 11: Validate as true XLSX OOXML
ok = False
with zipfile.ZipFile(out_path, 'r') as z:
    names = set(z.namelist())
    ok = ('[Content_Types].xml' in names) and any(n.startswith('xl/') for n in names)
if not ok:
    fail('Saved file is not a valid XLSX (missing OOXML parts)')

print(f'Wrote Excel: {out_path}')
