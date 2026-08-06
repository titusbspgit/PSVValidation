#!/usr/bin/env python3
import json
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Input JSON data (exactly as provided; DO NOT MODIFY)
JSON_TEXT = r'''[
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "Supports AHB 32-bit for the programming of CSR",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Verifies default values and masked write-readback of a set of GPIO GP0 registers. For each register in the address list (unresolved register macro MIZAR_GPIO_GP0_GPIO_8 through unresolved register macro MIZAR_GPIO_GP0_GPIO_27), the test checks the masked default value, then writes six test patterns using the write mask and validates the masked readback against the expected value. Test passes if no default or write-read mismatches are detected.",
    "Meta Test Description": "Objective: Validate default reset values and masked write-read functionality for a group of GPIO GP0 CSRs.\nInitialization: def_fail_cnt=0; wr_fail_cnt=0.\nDefault value check (chk_rst_val): For i in [0..CNT-1], addr=addr_array[i]. If skip_rst_array[i]==1 or read_mask_array[i]==0x00000000, continue. Read data_rd=read_reg(addr); mask LSB: data=(data_rd & 0xfffffffe); compare data with default_value_array[i]; on mismatch, increment def_fail_cnt and print failure.\nWrite-read check (chk_rd_wr): For each of 6 patterns in chk_val={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}, set data_wr=chk_val[j]. Write phase: For i in [0..CNT-1], addr=addr_array[i]; if skip_array[i]==1 or write_mask_array[i]==0x00000000, continue; write_reg(addr,(data_wr & write_mask_array[i])). Read/validate phase: For i in [0..CNT-1], skip if skip_array[i]==1 or write_mask_array[i]==0x00000000 or read_mask_array[i]==0x00000000; read masked value data_rd=(read_reg(addr) & read_mask_array[i]); compute wr_n=(write_mask_array[i] ^ 0xffffffff); expected exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if (data_rd != exp_val) increment wr_fail_cnt and print mismatch.\nCompletion: if(def_fail_cnt>0 || wr_fail_cnt>0) finish(1) else finish(0).\nNotes: soft_reset_chk() present but disabled with #ifdef 0; defines SOFT_RST_REG_ADDRESS and SOFT_RST_REG_DATA are unused in active flow. Arrays used: addr_array[], default_value_array[], read_mask_array[], write_mask_array[], skip_array[], skip_rst_array[].",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Known constraint noted in source: when reading default values, DIN may become 1 automatically if not forced; forcing zero to DIN can drive level select high, potentially affecting default readback expectations. Ensure GPIO block is clocked and out of reset before running the test.",
    "Test Steps / Procedure": "1. Initialize failure counters and load address, default, read-mask, write-mask, and skip arrays.\n2. For each address in the list (UNRESOLVED(MIZAR_GPIO_GP0_GPIO_8) through UNRESOLVED(MIZAR_GPIO_GP0_GPIO_27)), if reset-check skip is set or read mask is zero, skip the default check.\n3. Read each address and compare the masked value (LSB forced to 0) against its expected default; record failures.\n4. For each of six data patterns, iterate all addresses: if write skip is set or write mask is zero, skip the write; otherwise write the pattern masked by the write mask.\n5. For each written address with nonzero read mask, read back masked by the read mask.\n6. Compute expected value as the combination of written bits (where writable and readable) and preserved default bits (where not writable but readable); compare to the read value and record mismatches.\n7. After all patterns and addresses, declare PASS if both default and write-read failure counters are zero; otherwise FAIL.",
    "Meta Test Steps / Procedure": "test_case():\n- Call chk_rst_val(). If DEBUG_DISPLAY, print end of default check.\n- Call chk_rd_wr(). If DEBUG_DISPLAY, print end of write-read check.\n- If (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1); else finish(0).\n\nchk_rst_val():\n- for(i=0; i<CNT; i++): addr=addr_array[i]; if(skip_rst_array[i]==1) continue; if(read_mask_array[i]==0x00000000) continue; data_rd=read_reg(addr); data=(data_rd & 0xfffffffe); if(data==default_value_array[i]) PASS else { def_fail_cnt++; printf failure with addr, expected default_value_array[i], data, data_rd }.\n\nchk_rd_wr():\n- chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}.\n- for(j=0; j<6; j++): data_wr=chk_val[j].\n  Write phase: for(i=0; i<CNT; i++): addr=addr_array[i]; if(skip_array[i]==1) continue; if(write_mask_array[i]==0x00000000) continue; write_reg(addr,(data_wr & write_mask_array[i])).\n  Read/validate phase: for(i=0; i<CNT; i++): addr=addr_array[i]; if(skip_array[i]==1) continue; if(write_mask_array[i]==0x00000000) continue; if(read_mask_array[i]==0x00000000) continue; data_rd=(read_reg(addr) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd==exp_val) PASS else { wr_fail_cnt++; printf mismatch with addr, exp_val, data_rd }.",
    "Impacted Registers": "UNRESOLVED(MIZAR_GPIO_GP0_GPIO_8); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_9); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_10); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_11); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_12); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_13); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_14); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_15); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_16); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_17); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_18); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_19); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_20); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_21); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_22); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_23); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_24); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_25); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_26); UNRESOLVED(MIZAR_GPIO_GP0_GPIO_27); UNRESOLVED(GPIO_GP0_GPIO_8_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_9_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_10_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_11_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_12_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_13_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_14_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_15_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_16_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_17_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_18_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_19_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_20_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_21_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_22_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_23_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_24_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_25_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_26_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_27_DEFAULT_VAL); UNRESOLVED(GPIO_GP0_GPIO_8_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_9_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_10_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_11_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_12_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_13_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_14_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_15_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_16_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_17_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_18_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_19_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_20_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_21_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_22_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_23_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_24_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_25_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_26_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_27_READ_MASK); UNRESOLVED(GPIO_GP0_GPIO_8_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_9_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_10_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_11_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_12_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_13_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_14_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_15_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_16_WRITE_MASK); UNRESOLVED(GPIO_GP0_GPIO_17_WRITE_MASK; ... truncated for brevity in description, use the exact array from json_data",
    "Meta Impacted Registers": "MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27; GPIO_GP0_GPIO_8_DEFAULT_VAL; GPIO_GP0_GPIO_9_DEFAULT_VAL; GPIO_GP0_GPIO_10_DEFAULT_VAL; GPIO_GP0_GPIO_11_DEFAULT_VAL; GPIO_GP0_GPIO_12_DEFAULT_VAL; GPIO_GP0_GPIO_13_DEFAULT_VAL; GPIO_GP0_GPIO_14_DEFAULT_VAL; GPIO_GP0_GPIO_15_DEFAULT_VAL; GPIO_GP0_GPIO_16_DEFAULT_VAL; GPIO_GP0_GPIO_17_DEFAULT_VAL; GPIO_GP0_GPIO_18_DEFAULT_VAL; GPIO_GP0_GPIO_19_DEFAULT_VAL; GPIO_GP0_GPIO_20_DEFAULT_VAL; GPIO_GP0_GPIO_21_DEFAULT_VAL; GPIO_GP0_GPIO_22_DEFAULT_VAL; GPIO_GP0_GPIO_23_DEFAULT_VAL; GPIO_GP0_GPIO_24_DEFAULT_VAL; GPIO_GP0_GPIO_25_DEFAULT_VAL; GPIO_GP0_GPIO_26_DEFAULT_VAL; GPIO_GP0_GPIO_27_DEFAULT_VAL; GPIO_GP0_GPIO_8_READ_MASK; GPIO_GP0_GPIO_9_READ_MASK; GPIO_GP0_GPIO_10_READ_MASK; GPIO_GP0_GPIO_11_READ_MASK; GPIO_GP0_GPIO_12_READ_MASK; GPIO_GP0_GPIO_13_READ_MASK; GPIO_GP0_GPIO_14_READ_MASK; GPIO_GP0_GPIO_15_READ_MASK; GPIO_GP0_GPIO_16_READ_MASK; GPIO_GP0_GPIO_17_READ_MASK; GPIO_GP0_GPIO_18_READ_MASK; GPIO_GP0_GPIO_19_READ_MASK; GPIO_GP0_GPIO_20_READ_MASK; GPIO_GP0_GPIO_21_READ_MASK; GPIO_GP0_GPIO_22_READ_MASK; GPIO_GP0_GPIO_23_READ_MASK; GPIO_GP0_GPIO_24_READ_MASK; GPIO_GP0_GPIO_25_READ_MASK; GPIO_GP0_GPIO_26_READ_MASK; GPIO_GP0_GPIO_27_READ_MASK; GPIO_GP0_GPIO_8_WRITE_MASK; GPIO_GP0_GPIO_9_WRITE_MASK; GPIO_GP0_GPIO_10_WRITE_MASK; GPIO_GP0_GPIO_11_WRITE_MASK; GPIO_GP0_GPIO_12_WRITE_MASK; GPIO_GP0_GPIO_13_WRITE_MASK; GPIO_GP0_GPIO_14_WRITE_MASK; GPIO_GP0_GPIO_15_WRITE_MASK; GPIO_GP0_GPIO_16_WRITE_MASK; GPIO_GP0_GPIO_17_WRITE_MASK; GPIO_GP0_GPIO_18_WRITE_MASK; GPIO_GP0_GPIO_19_WRITE_MASK; GPIO_GP0_GPIO_20_WRITE_MASK; GPIO_GP0_GPIO_21_WRITE_MASK; GPIO_GP0_GPIO_22_WRITE_MASK; GPIO_GP0_GPIO_23_WRITE_MASK; GPIO_GP0_GPIO_24_WRITE_MASK; GPIO_GP0_GPIO_25_WRITE_MASK; GPIO_GP0_GPIO_26_WRITE_MASK; GPIO_GP0_GPIO_27_WRITE_MASK",
    "Validation / Acceptance Criteria": "Pass if all default checks match expected values (with LSB masked to 0) and all masked write-readback comparisons equal the expected values for all six patterns across all UNRESOLVED(MIZAR_GPIO_GP0_GPIO_8) through UNRESOLVED(MIZAR_GPIO_GP0_GPIO_27) addresses; otherwise fail.",
    "Meta Validation / Acceptance Criteria": "Default check: data_rd=read_reg(addr_array[i]); data=(data_rd & 0xfffffffe); expect data == default_value_array[i]; on mismatch: def_fail_cnt++.\nWrite-read check: For each pattern data_wr, write_reg(addr_array[i], (data_wr & write_mask_array[i])) when write_mask_array[i] != 0 and skip_array[i]==0. On readback: data_rd=(read_reg(addr_array[i]) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); expect data_rd == exp_val; on mismatch: wr_fail_cnt++.\nFinal: finish(0) if (def_fail_cnt==0 && wr_fail_cnt==0); else finish(1).",
    "Code Generation (Required / Not)": "Not Required",
    "Meta Headers": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<gpio/gpio_def.h> \n#include<gpio/gpio_offset.h> ",
    "Meta Macros": "#define SOFT_RST_REG_ADDRESS 0x00000000\n#define SOFT_RST_REG_DATA 0x00000000\n#define CNT 49",
    "Meta Arrays": "const unsigned long int addr_array[20]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,MIZAR_GPIO_GP0_GPIO_10,MIZAR_GPIO_GP0_GPIO_11,MIZAR_GPIO_GP0_GPIO_12,MIZAR_GPIO_GP0_GPIO_13,MIZAR_GPIO_GP0_GPIO_14,MIZAR_GPIO_GP0_GPIO_15,MIZAR_GPIO_GP0_GPIO_16,MIZAR_GPIO_GP0_GPIO_17,MIZAR_GPIO_GP0_GPIO_18,MIZAR_GPIO_GP0_GPIO_19,MIZAR_GPIO_GP0_GPIO_20,MIZAR_GPIO_GP0_GPIO_21,MIZAR_GPIO_GP0_GPIO_22,MIZAR_GPIO_GP0_GPIO_23,MIZAR_GPIO_GP0_GPIO_24,MIZAR_GPIO_GP0_GPIO_25,MIZAR_GPIO_GP0_GPIO_26,MIZAR_GPIO_GP0_GPIO_27,};\nconst unsigned int default_value_array[20]={GPIO_GP0_GPIO_8_DEFAULT_VAL,GPIO_GP0_GPIO_9_DEFAULT_VAL,GPIO_GP0_GPIO_10_DEFAULT_VAL,GPIO_GP0_GPIO_11_DEFAULT_VAL,GPIO_GP0_GPIO_12_DEFAULT_VAL,GPIO_GP0_GPIO_13_DEFAULT_VAL,GPIO_GP0_GPIO_14_DEFAULT_VAL,GPIO_GP0_GPIO_15_DEFAULT_VAL,GPIO_GP0_GPIO_16_DEFAULT_VAL,GPIO_GP0_GPIO_17_DEFAULT_VAL,GPIO_GP0_GPIO_18_DEFAULT_VAL,GPIO_GP0_GPIO_19_DEFAULT_VAL,GPIO_GP0_GPIO_20_DEFAULT_VAL,GPIO_GP0_GPIO_21_DEFAULT_VAL,GPIO_GP0_GPIO_22_DEFAULT_VAL,GPIO_GP0_GPIO_23_DEFAULT_VAL,GPIO_GP0_GPIO_24_DEFAULT_VAL,GPIO_GP0_GPIO_25_DEFAULT_VAL,GPIO_GP0_GPIO_26_DEFAULT_VAL,GPIO_GP0_GPIO_27_DEFAULT_VAL,};\nconst unsigned int read_mask_array[20]={GPIO_GP0_GPIO_8_READ_MASK,GPIO_GP0_GPIO_9_READ_MASK,GPIO_GP0_GPIO_10_READ_MASK,GPIO_GP0_GPIO_11_READ_MASK,GPIO_GP0_GPIO_12_READ_MASK,GPIO_GP0_GPIO_13_READ_MASK,GPIO_GP0_GPIO_14_READ_MASK,GPIO_GP0_GPIO_15_READ_MASK,GPIO_GP0_GPIO_16_READ_MASK,GPIO_GP0_GPIO_17_READ_MASK,GPIO_GP0_GPIO_18_READ_MASK,GPIO_GP0_GPIO_19_READ_MASK,GPIO_GP0_GPIO_20_READ_MASK,GPIO_GP0_GPIO_21_READ_MASK,GPIO_GP0_GPIO_22_READ_MASK,GPIO_GP0_GPIO_23_READ_MASK,GPIO_GP0_GPIO_24_READ_MASK,GPIO_GP0_GPIO_25_READ_MASK,GPIO_GP0_GPIO_26_READ_MASK,GPIO_GP0_GPIO_27_READ_MASK,};\nconst unsigned int write_mask_array[20]={GPIO_GP0_GPIO_8_WRITE_MASK,GPIO_GP0_GPIO_9_WRITE_MASK,GPIO_GP0_GPIO_10_WRITE_MASK,GPIO_GP0_GPIO_11_WRITE_MASK,GPIO_GP0_GPIO_12_WRITE_MASK,GPIO_GP0_GPIO_13_WRITE_MASK,GPIO_GP0_GPIO_14_WRITE_MASK,GPIO_GP0_GPIO_15_WRITE_MASK,GPIO_GP0_GPIO_16_WRITE_MASK,GPIO_GP0_GPIO_17_WRITE_MASK,GPIO_GP0_GPIO_18_WRITE_MASK,GPIO_GP0_GPIO_19_WRITE_MASK,GPIO_GP0_GPIO_20_WRITE_MASK,GPIO_GP0_GPIO_21_WRITE_MASK,GPIO_GP0_GPIO_22_WRITE_MASK,GPIO_GP0_GPIO_23_WRITE_MASK,GPIO_GP0_GPIO_24_WRITE_MASK,GPIO_GP0_GPIO_25_WRITE_MASK,GPIO_GP0_GPIO_26_WRITE_MASK,GPIO_GP0_GPIO_27_WRITE_MASK,};\nconst unsigned int skip_array[20]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,};\nconst unsigned int skip_rst_array[20]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,};"
  }
]'''

json_data = json.loads(JSON_TEXT)

# Fixed headers (order matters)
TESTPLAN_HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_HEADERS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

# Column widths for better readability (approximate)
TESTPLAN_WIDTHS = [8, 14, 28, 24, 80, 10, 12, 22, 22, 36, 80, 70, 60, 28]
METADATA_WIDTHS = [8, 24, 80, 80, 70, 60, 40, 36, 80]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DATA_ALIGNMENT = Alignment(wrapText=True, vertical="top")


def apply_header_format(ws, widths):
    for col_idx, title in enumerate([cell.value for cell in ws[1]], start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = DATA_ALIGNMENT
        # Set width
        if col_idx <= len(widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx-1]
    # Freeze header row and apply filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_rows(ws, headers, objs):
    for r_idx, obj in enumerate(objs, start=2):
        row_vals = [obj.get(h, "") for h in headers]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = DATA_ALIGNMENT


def create_workbook(objs, output_path: Path):
    wb = Workbook()
    ws_tp = wb.active
    ws_tp.title = "TestPlan"

    # Create MetaData as the second sheet
    ws_md = wb.create_sheet(title="MetaData")

    # Write headers
    ws_tp.append(TESTPLAN_HEADERS)
    ws_md.append(METADATA_HEADERS)

    # Write data rows aligned
    write_rows(ws_tp, TESTPLAN_HEADERS, objs)
    write_rows(ws_md, METADATA_HEADERS, objs)

    # Format headers and cells
    apply_header_format(ws_tp, TESTPLAN_WIDTHS)
    apply_header_format(ws_md, METADATA_WIDTHS)

    # VeryHidden MetaData
    ws_md.sheet_state = 'veryHidden'

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save
    wb.save(output_path)

    # Reopen and validate
    wb2 = load_workbook(output_path, data_only=True)
    if "TestPlan" not in wb2.sheetnames or "MetaData" not in wb2.sheetnames:
        raise RuntimeError("Required sheets missing")
    md = wb2["MetaData"]
    if md.sheet_state != 'veryHidden':
        raise RuntimeError("MetaData is not VeryHidden")

    # Validate headers
    tp = wb2["TestPlan"]
    tp_headers = [tp.cell(row=1, column=i+1).value for i in range(len(TESTPLAN_HEADERS))]
    md_headers = [md.cell(row=1, column=i+1).value for i in range(len(METADATA_HEADERS))]
    if tp_headers != TESTPLAN_HEADERS:
        raise RuntimeError("TestPlan headers mismatch")
    if md_headers != METADATA_HEADERS:
        raise RuntimeError("MetaData headers mismatch")

    # Validate row counts
    expected = len(objs)
    actual_tp = tp.max_row - 1 if tp.max_row else 0
    actual_md = md.max_row - 1 if md.max_row else 0
    if actual_tp != expected or actual_md != expected:
        raise RuntimeError(f"Row count mismatch: expected {expected}, got TestPlan={actual_tp}, MetaData={actual_md}")


def main():
    objs = json_data
    # Timestamp in Asia/Kolkata
    ts = datetime.now(ZoneInfo("Asia/Kolkata")).strftime('%Y%m%d_%H%M%S')
    ip_name = "GPIO"
    filename = f"{ip_name}_TestPlan_{ts}.xlsx"
    out_path = Path("Test_Output/GPIO/TestPlan") / filename
    create_workbook(objs, out_path)
    print(f"Generated: {out_path}")


if __name__ == "__main__":
    main()
