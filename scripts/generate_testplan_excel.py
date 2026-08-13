#!/usr/bin/env python3
import argparse
import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

TESTPLAN_HEADERS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_HEADERS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

COL_WIDTHS_TESTPLAN = {
    "A": 8,   # Index
    "B": 18,  # SS / Module
    "C": 20,  # Feature
    "D": 30,  # Test Case Name
    "E": 50,  # Test Description
    "F": 10,  # Speed
    "G": 12,  # Mode
    "H": 22,  # Memory Start Offset
    "I": 22,  # Memory End Offset
    "J": 30,  # Remarks
    "K": 60,  # Test Steps / Procedure
    "L": 40,  # Impacted Registers
    "M": 60,  # Validation / Acceptance Criteria
    "N": 28,  # Code Generation (Required / Not)
}

COL_WIDTHS_METADATA = {
    "A": 8,   # Index
    "B": 30,  # Test Case Name
    "C": 50,  # Meta Test Description
    "D": 60,  # Meta Test Steps / Procedure
    "E": 40,  # Meta Impacted Registers
    "F": 60,  # Meta Validation / Acceptance Criteria
    "G": 30,  # Meta Headers
    "H": 30,  # Meta Macros
    "I": 30,  # Meta Arrays
}


def load_json_array(path: str):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("json_data must be an array of objects")
    return data


def apply_header_style(ws, headers):
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    align = Alignment(wrap_text=True, vertical="top")
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align


def set_column_widths(ws, width_map):
    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width


def write_rows(ws, headers, rows):
    align = Alignment(wrap_text=True, vertical="top")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = align


def build_rows(data):
    # Maintain order; if keys missing, leave blank
    testplan_rows = []
    metadata_rows = []
    for item in data:
        if not isinstance(item, dict):
            raise ValueError("Each array element must be an object")
        # Normal fields
        tp = {h: item.get(h, "") for h in TESTPLAN_HEADERS}
        # Meta fields
        md = {h: item.get(h, "") for h in METADATA_HEADERS}
        testplan_rows.append(tp)
        metadata_rows.append(md)
    return testplan_rows, metadata_rows


def generate_excel(json_path: str, out_dir: str, ip_name: str) -> str:
    data = load_json_array(json_path)
    testplan_rows, metadata_rows = build_rows(data)

    wb = Workbook()
    ws_tp = wb.active
    ws_tp.title = "TestPlan"
    ws_md = wb.create_sheet("MetaData")

    # Headers and formatting
    apply_header_style(ws_tp, TESTPLAN_HEADERS)
    apply_header_style(ws_md, METADATA_HEADERS)

    # Data
    write_rows(ws_tp, TESTPLAN_HEADERS, testplan_rows)
    write_rows(ws_md, METADATA_HEADERS, metadata_rows)

    # Formatting: column widths, wrap already set per cell, freeze panes
    set_column_widths(ws_tp, COL_WIDTHS_TESTPLAN)
    set_column_widths(ws_md, COL_WIDTHS_METADATA)
    ws_tp.freeze_panes = "A2"
    ws_md.freeze_panes = "A2"

    # Very hidden MetaData sheet
    ws_md.sheet_state = "veryHidden"

    # Ensure output directory exists
    os.makedirs(out_dir, exist_ok=True)

    # IST timestamp
    ist = ZoneInfo("Asia/Kolkata")
    timestamp = datetime.now(ist).strftime("%Y%m%d_%H%M%S")
    filename = f"{ip_name}_TestPlan_{timestamp}.xlsx"
    out_path = os.path.join(out_dir, filename)

    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Generate TestPlan Excel from JSON")
    parser.add_argument("--json-file", required=True, help="Path to JSON array file")
    parser.add_argument("--out-dir", required=True, help="Output directory for the Excel file")
    parser.add_argument("--ip-name", required=True, help="IP name for filename prefix")
    args = parser.parse_args()

    out_path = generate_excel(args.json_file, args.out_dir, args.ip_name)
    print(f"Generated: {out_path}")


if __name__ == "__main__":
    main()
