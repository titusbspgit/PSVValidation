import json, os, sys, zipfile
from copy import deepcopy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_COLUMNS = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

META_COLUMNS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria'
]

WRAP_COLUMNS = {
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria'
}

BLUE_FILL = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def load_json_input(path):
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, dict) and all(str(k).upper().startswith('TC') for k in data.keys()):
        # Convert {"TC1": {...}, ...} -> [{...}, ...]
        arr = []
        for key in sorted(data.keys(), key=lambda x: int(''.join(filter(str.isdigit, x)) or 0)):
            arr.append(data[key])
        return arr
    if isinstance(data, list):
        if not all(isinstance(x, dict) for x in data):
            raise SystemExit('JSON array must contain objects only')
        return data
    raise SystemExit('Invalid JSON input format')


def read_existing_excel(path):
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    if 'TestPlan' not in wb.sheetnames:
        return []
    ws = wb['TestPlan']
    # Read visible headers
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        obj = {headers[i]: ('' if r[i] is None else r[i]) for i in range(len(headers))}
        rows.append(obj)
    # Read META if present and merge per row by index
    if 'Meta_data_sheet' in wb.sheetnames:
        mws = wb['Meta_data_sheet']
        m_headers = [c.value for c in mws[1]]
        m_rows = []
        for r in mws.iter_rows(min_row=2, values_only=True):
            m_rows.append({m_headers[i]: ('' if r[i] is None else r[i]) for i in range(len(m_headers))})
        for i in range(min(len(rows), len(m_rows))):
            rows[i].update({k: m_rows[i].get(k, '') for k in META_COLUMNS})
    return rows


def union_keys(objs):
    seen = []
    for o in objs:
        for k in o.keys():
            if k not in seen:
                seen.append(k)
    return seen


def ensure_all_keys(objs, keys):
    for o in objs:
        for k in keys:
            if k not in o:
                o[k] = ''


def write_meta_sheet(wb, final_json):
    mws = wb.create_sheet('Meta_data_sheet')
    # Header
    for c, h in enumerate(META_COLUMNS, start=1):
        mws.cell(row=1, column=c, value=h)
    # Rows
    for r, row in enumerate(final_json, start=2):
        for c, h in enumerate(META_COLUMNS, start=1):
            mws.cell(row=r, column=c, value=row.get(h, ''))
    # Very hidden
    mws.sheet_state = 'veryHidden'


def apply_base_formatting(ws):
    # Freeze top row
    ws.freeze_panes = 'A2'
    # Header formatting
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = BLUE_FILL
        cell.border = THIN_BORDER
    # Borders and alignment for data rows
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if ws.cell(row=1, column=c).value in WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            else:
                cell.alignment = Alignment(vertical='top', horizontal='left')
    # Center Index column if present
    for c in range(1, max_col + 1):
        if ws.cell(row=1, column=c).value == 'Index':
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=c).alignment = Alignment(vertical='top', horizontal='center')
            break
    # Approximate auto-fit column widths
    for c in range(1, max_col + 1):
        header = ws.cell(row=1, column=c).value or ''
        max_len = len(str(header))
        for r in range(2, max_row + 1):
            val = ws.cell(row=r, column=c).value
            l = len(str(val)) if val is not None else 0
            if l > max_len:
                max_len = l
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(120, max(10, max_len + 2))


def reorder_to_main_sheet(wb, data_rows):
    # Operate on the existing 'Data' sheet in place
    ws = wb['Data']
    # Build desired order present in data
    desired = [c for c in MAIN_COLUMNS if c in data_rows[0].keys()]
    # Prepare rows for writing
    rows = []
    rows.append(desired)
    for row in data_rows:
        rows.append([row.get(k, '') for k in desired])
    # Clear and rewrite
    ws.delete_rows(1, ws.max_row)
    for r_idx, row_vals in enumerate(rows, start=1):
        for c_idx, v in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    # Rename to TestPlan
    ws.title = 'TestPlan'
    apply_base_formatting(ws)
    # Data validation on Code Generation column
    try:
        headers = [cell.value for cell in ws[1]]
        if 'Code Generation (Required / Not)' in headers:
            col_idx = headers.index('Code Generation (Required / Not)') + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True)
            ws.add_data_validation(dv)
            if ws.max_row >= 2:
                dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
    except Exception as e:
        # Non-fatal: continue without DV
        pass


def main():
    json_path = os.environ.get('JSON_INPUT_PATH')
    out_dir = os.environ.get('OUTPUT_FILE_PATH')
    out_name = os.environ.get('OUTPUT_FILE_NAME')
    if not json_path or not out_dir or not out_name:
        raise SystemExit('Missing required environment variables')

    incoming = load_json_input(json_path)

    # Merge with existing Excel if present
    out_full = os.path.join(out_dir, out_name)
    existing = read_existing_excel(out_full)

    # FINAL_JSON = existing + incoming
    final_json = existing + incoming

    # Reindex
    for i, row in enumerate(final_json, start=1):
        row['Index'] = i

    # Union of keys for staging sheet
    all_keys = union_keys(final_json)

    # Ensure all keys present
    ensure_all_keys(final_json, all_keys)

    # Build workbook
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Write staging headers
    for c, h in enumerate(all_keys, start=1):
        ws.cell(row=1, column=c, value=h)
    # Write rows
    for r, row in enumerate(final_json, start=2):
        for c, h in enumerate(all_keys, start=1):
            ws.cell(row=r, column=c, value=row.get(h, ''))

    # Create META sheet
    write_meta_sheet(wb, final_json)

    # Normalize main sheet (rename Data -> TestPlan, remove to MAIN columns only and format)
    reorder_to_main_sheet(wb, final_json)

    # Safety check: only TestPlan and Meta_data_sheet should remain, and no sheet named 'Data'
    if 'Data' in wb.sheetnames:
        # Attempt to delete if any leftover
        del wb['Data']
    if set(wb.sheetnames) - set(['TestPlan', 'Meta_data_sheet']):
        # If any unexpected sheet remains, fail to trigger workflow error
        raise SystemExit('Unexpected worksheets present after normalization')

    wb.save(out_full)

    # Validate as ZIP-based Office Open XML
    if not zipfile.is_zipfile(out_full):
        raise SystemExit('Output is not a valid XLSX (zip) file')

if __name__ == '__main__':
    main()
