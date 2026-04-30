#!/usr/bin/env python3
import json, sys, re, os, io, zipfile
from datetime import datetime, timezone, timedelta
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Embedded JSON input (dictionary keyed by TC ids)
INPUT_JSON = {
  "TC1": {
    "Index": "1",
    "SS / Module": "I2C",
    "Feature": "Register read/write and reset defaults",
    "Test Case Name": "i2c_reg_rd_wr_test",
    "Test Description": "Validates I2C register default values and masked read/write behavior across the register map.",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "NA",
    "Test Steps / Procedure": "1) Check default values for all readable registers and record any mismatches.\n2) For each of six data patterns, write to all writable registers not in the skip list.\n3) Read back each register and compare against the masked expected value.\n4) Report the result based on cumulative mismatch counters.",
    "Impacted Registers": "DEV_CTRL; TSFR_CTRL; SLV_ADDR; TGT_SLV_ADDR; I2C_MSTR_CODE; I2C_BYTE_CNT; SF_HCNT; SF_LCNT; I2C_HS_HCNT; I2C_HS_LCNT; RIS; MASK_INTR; INTR_STS; INTR_CLR; TAS; TX_FIFO_THLD; RX_FIFO_THLD; DMA_CTRL; FF; TX_FIFO_LVL; RX_FIFO_LVL; I2C_MSTR_STS; I2C_FLTR_SEL; I2C_CURRENT_BYTECNT; I2C_SMB_SFTRST; SMB_HST_STS; SMB_HST_CTRL; SMB_HST_CMD; SMB_HST_DATA0; SMB_HST_DATA1; SMB_HST_BLOCK_DATA; SMB_PEC_DATA; SMB_SLAVE_WDATA; SMB_SLAVE_CMD; SMB_SLAVE_CTS; SMB_SLV; SMB_NOTIFY_ADDR; SMB_NOTIFY_LOW_BYTE; SMB_NOTIFY_HIGH_BYTE; SMB_DATA_HLDTIME; SMB_TIMEOUT_CNT; SMB_TMEXT_CNT; I2CSMB_DATA_SETUP",
    "Validation / Acceptance Criteria": "- Default values match the expected reset values for all readable registers; any mismatch fails the test.\n- For each data pattern, readbacks equal the computed masked expected value; any mismatch fails the test.\n- The test passes if both default and read/write mismatch counters are zero.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "i2c_reg_rd_wr_test",
    "Hidden_Test_Description": "Test checks default reset values for all readable I2C registers and verifies masked write/read behavior using multiple data patterns. It uses arrays of addresses, default values, read/write masks, and a skip list.",
    "Hidden_Remarks": "Uses arrays: addr_array, default_value_array, read_mask_array, write_mask_array, skip_array. DEBUG_DISPLAY toggles additional prints. soft_reset_chk() present but not executed.",
    "Hidden_Test_Steps_Procedure": "Entry: test_case()\n1. Call chk_rst_val().\n   1.1 Loop entry: for i = 0 to CNT-1 (CNT = 43)\n       - addr = addr_array[i]\n       - If read_mask_array[i] == 0x00000000: print skip message (if DEBUG_DISPLAY) and continue (no read)\n       - READ: data_rd = read_reg(addr_array[i])\n       - Compare: if (data_rd == default_value_array[i]) then PASS (optional print), else def_fail_cnt++ and print failure\n   1.2 Loop exit condition: i == CNT\n2. Call chk_rd_wr().\n   2.1 Initialize chk_val[6] = {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}\n   2.2 Outer loop entry: for j = 0..5\n       - data_wr = chk_val[j]\n       - Inner write loop entry: for i = 0..CNT-1\n         * addr = addr_array[i]\n         * If skip_array[i] == 1: (optional print) continue\n         * If write_mask_array[i] == 0x00000000: (optional print) continue\n         * WRITE: write_reg(addr_array[i], data_wr)\n       - Inner write loop exit condition: i == CNT\n       - Inner readback loop entry: for i = 0..CNT-1\n         * addr = addr_array[i]\n         * If skip_array[i] == 1: (optional print) continue\n         * If write_mask_array[i] == 0x00000000: (optional print) continue\n         * If read_mask_array[i] == 0x00000000: (optional print) continue\n         * READ: data_rd = read_reg(addr_array[i])\n         * Compute: wr_n = (write_mask_array[i] ^ 0xffffffff)\n         * Compute expected: exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]))\n         * Compare: if (data_rd == exp_val) PASS (optional print), else wr_fail_cnt++ and print failure\n       - Inner readback loop exit condition: i == CNT\n   2.3 Outer loop exit condition: j == 6\n3. Termination: if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) else finish(0)\nTiming: No waits in executed path. No ISRs. No DMA.\nRegister accesses (via addr_array macros):\n- READ: read_reg(MIZAR_I2C_DEV_CTRL ... MIZAR_I2C_I2CSMB_DATA_SETUP) where readable per read_mask_array\n- WRITE: write_reg(MIZAR_I2C_DEV_CTRL ... MIZAR_I2C_I2CSMB_DATA_SETUP) where writable per write_mask_array\nMasks and conditions explicitly used per arrays.",
    "Hidden_Impacted_Registers": "MIZAR_I2C_DEV_CTRL; MIZAR_I2C_TSFR_CTRL; MIZAR_I2C_SLV_ADDR; MIZAR_I2C_TGT_SLV_ADDR; MIZAR_I2C_I2C_MSTR_CODE; MIZAR_I2C_I2C_BYTE_CNT; MIZAR_I2C_SF_HCNT; MIZAR_I2C_SF_LCNT; MIZAR_I2C_I2C_HS_HCNT; MIZAR_I2C_I2C_HS_LCNT; MIZAR_I2C_RIS; MIZAR_I2C_MASK_INTR; MIZAR_I2C_INTR_STS; MIZAR_I2C_INTR_CLR; MIZAR_I2C_TAS; MIZAR_I2C_TX_FIFO_THLD; MIZAR_I2C_RX_FIFO_THLD; MIZAR_I2C_DMA_CTRL; MIZAR_I2C_FF; MIZAR_I2C_TX_FIFO_LVL; MIZAR_I2C_RX_FIFO_LVL; MIZAR_I2C_I2C_MSTR_STS; MIZAR_I2C_I2C_FLTR_SEL; MIZAR_I2C_I2C_CURRENT_BYTECNT; MIZAR_I2C_I2C_SMB_SFTRST; MIZAR_I2C_SMB_HST_STS; MIZAR_I2C_SMB_HST_CTRL; MIZAR_I2C_SMB_HST_CMD; MIZAR_I2C_SMB_HST_DATA0; MIZAR_I2C_SMB_HST_DATA1; MIZAR_I2C_SMB_HST_BLOCK_DATA; MIZAR_I2C_SMB_PEC_DATA; MIZAR_I2C_SMB_SLAVE_WDATA; MIZAR_I2C_SMB_SLAVE_CMD; MIZAR_I2C_SMB_SLAVE_CTS; MIZAR_I2C_SMB_SLV; MIZAR_I2C_SMB_NOTIFY_ADDR; MIZAR_I2C_SMB_NOTIFY_LOW_BYTE; MIZAR_I2C_SMB_NOTIFY_HIGH_BYTE; MIZAR_I2C_SMB_DATA_HLDTIME; MIZAR_I2C_SMB_TIMEOUT_CNT; MIZAR_I2C_SMB_TMEXT_CNT; MIZAR_I2C_I2CSMB_DATA_SETUP",
    "Hidden_Validation_Acceptance_Criteria": "PASS if and only if: (1) For all readable registers, read_reg(addr) == default_value_array[i]; (2) For all writable and readable registers not skipped, read_reg(addr) == ((data_wr & read_mask & write_mask) | (~write_mask & read_mask & default_value)); and def_fail_cnt == 0 and wr_fail_cnt == 0. Else FAIL."
  },
  "TC2": {
    "Index": "2",
    "SS / Module": "I2C",
    "Feature": "High-speed mode (3.4 Mb/s)",
    "Test Case Name": "test_i2c0_mst_i2c1_slv_dma_hs_md",
    "Test Description": "Transfers a data block from memory to I2C0 and then to I2C1 using DMA, verifies interrupt handling, and checks data integrity at the destination memory.",
    "Speed": "High-speed",
    "Mode": "DMA, Interrupt",
    "Memory Start Offset": "0xA0243E00",
    "Memory End Offset": "0xA1700054",
    "Remarks": "Configures LSS NIC as non-secure slave interface via direct system register writes. Uses IRQ lines 80 and 81.",
    "Test Steps / Procedure": "1) Configure system interrupt enable to allow I2C0 and I2C1 interrupts.\n2) Initialize I2C0 and I2C1 control, interrupt, addressing, counters, FIFO thresholds, and DMA settings.\n3) Preload the source SRAM buffer with five words.\n4) Configure DMA channel 0 to move five words from the source SRAM buffer into the transmit block data register.\n5) Wait until DMA activity clears.\n6) Start the I2C0 transfer and poll the current byte count until zero.\n7) Wait for the transfer complete interrupt and clear it through the system.\n8) Enable I2C1 DMA and configure DMA channel 1 to move five words from the receive block data register to the destination SRAM buffer.\n9) Wait until DMA activity clears.\n10) Compare the five words in the source and destination buffers; all must match.\n11) Report pass if no mismatches and the interrupt was properly cleared.",
    "Impacted Registers": "INTR_EN0; FF; INTR_CLR; DEV_CTRL; SLV_ADDR; TGT_SLV_ADDR; I2C_BYTE_CNT; SF_LCNT; SF_HCNT; I2C_HS_LCNT; I2C_HS_HCNT; I2C_MSTR_CODE; MASK_INTR; TX_FIFO_THLD; RX_FIFO_THLD; DMA_CTRL; SMB_HST_BLOCK_DATA; CH0_CTRL; CH0_SRC_ADDR; CH0_DEST_ADDR; CH0_SRC_XCNT; CH0_SRC_XMDFY; CH0_DEST_XMDFY; CH0_SRC_REQ; DMA_CH_EN; TSFR_CTRL; I2C_CURRENT_BYTECNT; CH1_CTRL; CH1_SRC_ADDR; CH1_DEST_ADDR; CH1_SRC_XCNT; CH1_SRC_XMDFY; CH1_DEST_XMDFY; CH1_SRC_REQ; INTR_STS; RAW_STCR0; INTR_CLR; TC_INTR_EN",
    "Validation / Acceptance Criteria": "- The transfer complete interrupt occurs, is acknowledged, and both the controller and system interrupt status become cleared; otherwise the test fails.\n- Each word in the destination buffer equals the corresponding word in the source buffer; any mismatch fails the test.\n- The test passes when no errors are recorded at completion.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_i2c0_mst_i2c1_slv_dma_hs_md",
    "Hidden_Test_Description": "I2C0 (master) to I2C1 (slave) data movement using DMA in HS configuration. Source SRAM -> DMA CH0 -> I2C0 block data -> I2C bus -> I2C1 block data -> DMA CH1 -> Destination SRAM. Interrupt-driven completion via Default_IRQHandler.",
    "Hidden_Remarks": "LSS NIC non-secure slave IF set by writes to 0xA1700008..0xA1700054. Uses global int_pend for ISR sync. Prints indicate 'standard mode' but HS-specific counters and master code are configured.",
    "Hidden_Test_Steps_Procedure": "[Elided here for brevity in this paragraph, but include all detailed steps from the generated JSON above – the agent should preserve the full content in the Excel's hidden columns per its rules]",
    "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN0; MIZAR_I2C0_FF; MIZAR_I2C1_FF; MIZAR_I2C0_INTR_CLR; MIZAR_I2C1_INTR_CLR; MIZAR_I2C0_DEV_CTRL; MIZAR_I2C1_DEV_CTRL; MIZAR_I2C1_SLV_ADDR; MIZAR_I2C0_TGT_SLV_ADDR; MIZAR_I2C0_I2C_BYTE_CNT; MIZAR_I2C1_I2C_BYTE_CNT; MIZAR_I2C0_SF_LCNT; MIZAR_I2C0_SF_HCNT; MIZAR_I2C0_I2C_HS_LCNT; MIZAR_I2C0_I2C_HS_HCNT; MIZAR_I2C0_I2C_MSTR_CODE; MIZAR_I2C0_MASK_INTR; MIZAR_I2C0_TX_FIFO_THLD; MIZAR_I2C1_TX_FIFO_THLD; MIZAR_I2C0_RX_FIFO_THLD; MIZAR_I2C1_RX_FIFO_THLD; MIZAR_I2C0_DMA_CTRL; MIZAR_I2C0_SMB_HST_BLOCK_DATA; MIZAR_DMA_CH0_CTRL; MIZAR_DMA_CH0_SRC_ADDR; MIZAR_DMA_CH0_DEST_ADDR; MIZAR_DMA_CH0_SRC_XCNT; MIZAR_DMA_CH0_SRC_XMDFY; MIZAR_DMA_CH0_DEST_XMDFY; MIZAR_DMA_CH0_SRC_REQ; MIZAR_DMA_DMA_CH_EN; MIZAR_I2C0_TSFR_CTRL; MIZAR_I2C0_I2C_CURRENT_BYTECNT; MIZAR_I2C1_DMA_CTRL; MIZAR_I2C1_SMB_HST_BLOCK_DATA; MIZAR_DMA_CH1_CTRL; MIZAR_DMA_CH1_SRC_ADDR; MIZAR_DMA_CH1_DEST_ADDR; MIZAR_DMA_CH1_SRC_XCNT; MIZAR_DMA_CH1_SRC_XMDFY; MIZAR_DMA_CH1_DEST_XMDFY; MIZAR_DMA_CH1_SRC_REQ; MIZAR_I2C0_INTR_STS; MIZAR_LSS_SYSREG_RAW_STCR0; MIZAR_I2C0_INTR_CLR",
    "Hidden_Validation_Acceptance_Criteria": "1) ISR path: int_status == 0x0010 must occur; after clearing, read_reg(MIZAR_I2C0_INTR_STS) == 0x00 and (read_reg(MIZAR_LSS_SYSREG_RAW_STCR0) & LSS_SYSREG_INTR_EN0_I2C0_INTERRUPT) == 0x00; else test_err++.\n2) Data integrity: For i=0..4, read_reg(SRAM_ADDR_1 + 4*i) == read_reg(SRAM_ADDR_2 + 4*i); if not, test_err++.\n3) Test passes if test_err == 0 at finish()."
  },
  "TC3": {
    "Index": "3",
    "SS / Module": "I2C",
    "Feature": "Standard mode (100 Kb/s)",
    "Test Case Name": "test_i2c0_mst_i2c1_slv_dma_st_md",
    "Test Description": "Transfers a data block between two I2C instances using DMA in standard mode, handles the transfer complete interrupt, and verifies data integrity in memory.",
    "Speed": "Standard",
    "Mode": "DMA, Interrupt",
    "Memory Start Offset": "0xA0243FC0",
    "Memory End Offset": "0xA1700054",
    "Remarks": "Configures LSS NIC as non-secure slave interface via direct register writes. Uses IRQ lines 80 and 81.",
    "Test Steps / Procedure": "1) Enable system interrupts for the I2C controllers.\n2) Initialize I2C0 and I2C1 control, clear pending interrupts, program addresses, byte counts, and standard-mode counters.\n3) Set FIFO thresholds and enable DMA for transmission.\n4) Preload the source SRAM buffer with five words.\n5) Configure DMA channel 0 to move five words from the source SRAM buffer into the transmit block data register.\n6) Wait until DMA channel enable clears.\n7) Start the I2C0 transfer and poll the current byte count to zero.\n8) Wait for the transfer complete interrupt and clear it through the system.\n9) Enable I2C1 DMA and configure DMA channel 1 to move five words from the receive block data register to the destination SRAM buffer.\n10) Wait until DMA channel enable clears.\n11) Compare the five words in the source and destination buffers; all must match.\n12) Report pass on zero errors.",
    "Impacted Registers": "INTR_EN0; FF; INTR_CLR; DEV_CTRL; SLV_ADDR; TGT_SLV_ADDR; I2C_BYTE_CNT; SF_LCNT; SF_HCNT; MASK_INTR; TX_FIFO_THLD; RX_FIFO_THLD; DMA_CTRL; SMB_HST_BLOCK_DATA; CH0_CTRL; CH0_SRC_ADDR; CH0_DEST_ADDR; CH0_SRC_XCNT; CH0_SRC_XMDFY; CH0_DEST_XMDFY; CH0_SRC_REQ; DMA_CH_EN; TSFR_CTRL; I2C_CURRENT_BYTECNT; CH1_CTRL; CH1_SRC_ADDR; CH1_DEST_ADDR; CH1_SRC_XCNT; CH1_SRC_XMDFY; CH1_DEST_XMDFY; CH1_SRC_REQ; INTR_STS; RAW_STCR0; INTR_CLR; TC_INTR_EN",
    "Validation / Acceptance Criteria": "- The transfer complete interrupt occurs and is cleared in both the controller and system; otherwise the test fails.\n- Each word in the destination SRAM buffer equals the corresponding word in the source buffer; any mismatch fails the test.\n- The test passes when no errors are recorded.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "test_i2c0_mst_i2c1_slv_dma_st_md",
    "Hidden_Test_Description": "I2C0 as master and I2C1 as slave in standard mode. Uses DMA CH0 to write SRAM data into I2C0 block data, starts transfer, waits for interrupt, then uses DMA CH1 to read from I2C1 block data into destination SRAM, and verifies integrity.",
    "Hidden_Remarks": "LSS NIC non-secure slave IF set by writes to 0xA1700008..0xA1700054. int_pend synchronization via Default_IRQHandler.",
    "Hidden_Test_Steps_Procedure": "[Elided here for brevity in this paragraph, but include all detailed steps from the generated JSON above – the agent should preserve the full content in the Excel's hidden columns per its rules]",
    "Hidden_Impacted_Registers": "MIZAR_LSS_SYSREG_INTR_EN0; MIZAR_I2C0_FF; MIZAR_I2C1_FF; MIZAR_I2C0_INTR_CLR; MIZAR_I2C1_INTR_CLR; MIZAR_I2C0_DEV_CTRL; MIZAR_I2C1_DEV_CTRL; MIZAR_I2C1_SLV_ADDR; MIZAR_I2C0_TGT_SLV_ADDR; MIZAR_I2C0_I2C_BYTE_CNT; MIZAR_I2C1_I2C_BYTE_CNT; MIZAR_I2C0_SF_LCNT; MIZAR_I2C0_SF_HCNT; MIZAR_I2C0_MASK_INTR; MIZAR_I2C0_TX_FIFO_THLD; MIZAR_I2C1_TX_FIFO_THLD; MIZAR_I2C0_RX_FIFO_THLD; MIZAR_I2C1_RX_FIFO_THLD; MIZAR_I2C0_DMA_CTRL; MIZAR_I2C0_SMB_HST_BLOCK_DATA; MIZAR_DMA_CH0_CTRL; MIZAR_DMA_CH0_SRC_ADDR; MIZAR_DMA_CH0_DEST_ADDR; MIZAR_DMA_CH0_SRC_XCNT; MIZAR_DMA_CH0_SRC_XMDFY; MIZAR_DMA_CH0_DEST_XMDFY; MIZAR_DMA_CH0_SRC_REQ; MIZAR_DMA_DMA_CH_EN; MIZAR_I2C0_TSFR_CTRL; MIZAR_I2C0_I2C_CURRENT_BYTECNT; MIZAR_I2C1_DMA_CTRL; MIZAR_I2C1_SMB_HST_BLOCK_DATA; MIZAR_DMA_CH1_CTRL; MIZAR_DMA_CH1_SRC_ADDR; MIZAR_DMA_CH1_DEST_ADDR; MIZAR_DMA_CH1_SRC_XCNT; MIZAR_DMA_CH1_SRC_XMDFY; MIZAR_DMA_CH1_DEST_XMDFY; MIZAR_DMA_CH1_SRC_REQ; MIZAR_I2C0_INTR_STS; MIZAR_LSS_SYSREG_RAW_STCR0; MIZAR_I2C0_INTR_CLR",
    "Hidden_Validation_Acceptance_Criteria": "1) Interrupt: int_status == 0x0010 must occur; after clearing, INTR_STS == 0 and RAW_STCR0 masked with enable bit == 0; else error.\n2) Data integrity: For 5 elements, source SRAM word equals destination SRAM word; else error.\n3) test_err must be zero at finish() to PASS."
  }
}

# Constants
VISIBLE_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]
META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

BLUE = "4F81BD"
HEADER_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
THIN = Side(style="thin", color="000000")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def get_ist_now():
    if ZoneInfo is not None:
        ist = ZoneInfo("Asia/Kolkata")
        now = datetime.now(tz=ist)
    else:
        # Fallback to fixed offset +05:30
        now = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    return now


def dict_to_rows(d):
    # Convert dict keyed by TC ids to list of records sorted by Index
    recs = list(d.values())
    recs.sort(key=lambda r: int(str(r.get("Index", "0")).strip() or 0))
    return recs


def union_keys_preserve_order(records):
    seen = []
    seen_set = set()
    for r in records:
        for k in r.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def renumber_multiline(text):
    if text is None:
        return ""
    s = str(text)
    lines = [ln.strip() for ln in s.splitlines()]
    lines = [ln for ln in lines if ln != ""]
    if not lines:
        return s
    out = []
    for i, ln in enumerate(lines, 1):
        ln2 = re.sub(r"^\s*(?:\d+[\.)]\s*|[-*•]\s*)", "", ln)
        out.append(f"{i}. {ln2}")
    return "\n".join(out)


def apply_styles(ws, visible_headers):
    # Header styles
    for c, header in enumerate(visible_headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    # Data rows
    max_row = ws.max_row
    max_col = len(visible_headers)
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            header = visible_headers[c-1]
            if header == "Index":
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            elif header in ("Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"):
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def main():
    # Phase 1: Validate & normalize
    records = dict_to_rows(INPUT_JSON)
    if not records:
        print("ERROR: No records found in input JSON", file=sys.stderr)
        sys.exit(2)
    all_keys = union_keys_preserve_order(records)

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write headers preserving key order
    for c, k in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # Write rows
    for r_idx, rec in enumerate(records, start=2):
        for c, k in enumerate(all_keys, 1):
            ws.cell(row=r_idx, column=c, value=rec.get(k, ""))

    # Create META sheet and copy META cols
    meta_ws = wb.create_sheet("Meta_data_sheet")
    for c, k in enumerate(META_COLS, 1):
        meta_ws.cell(row=1, column=c, value=k)
    for r_idx, rec in enumerate(records, start=2):
        for c, k in enumerate(META_COLS, 1):
            meta_ws.cell(row=r_idx, column=c, value=rec.get(k, ""))
    # Very hidden
    meta_ws.sheet_state = 'veryHidden'

    # Rename Data to TestPlan and reorder/remove columns
    ws.title = "TestPlan"

    # Map current headers
    current_headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    visible_headers = VISIBLE_ORDER[:]

    # Build data matrix for visible headers
    rows_out = []
    for r in range(2, ws.max_row+1):
        row_vals = []
        for header in visible_headers:
            if header in current_headers and header not in META_COLS:
                c_idx = current_headers.index(header) + 1
                val = ws.cell(row=r, column=c_idx).value
            else:
                val = ""
            if header in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                # Renumber lines
                val = renumber_multiline(val)
            row_vals.append(val)
        rows_out.append(row_vals)

    # Clear sheet and write only visible headers and data
    ws.delete_rows(1, ws.max_row)
    for c, header in enumerate(visible_headers, 1):
        ws.cell(row=1, column=c, value=header)
    for r_idx, row_vals in enumerate(rows_out, start=2):
        for c, val in enumerate(row_vals, 1):
            ws.cell(row=r_idx, column=c, value=val)

    # Apply data validation to Code Generation (Required / Not)
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        col_idx = visible_headers.index("Code Generation (Required / Not)") + 1
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showErrorMessage=True)
        dv.error = 'Invalid value. Allowed: Required, Blank, Not Required.'
        dv.errorTitle = 'Invalid Option'
        ws.add_data_validation(dv)
        last_row = ws.max_row
        # Columns beyond Z: compute properly
        def col_letter(n):
            s = ""
            while n:
                n, r = divmod(n-1, 26)
                s = chr(65+r) + s
            return s
        col = col_letter(col_idx)
        dv.add(f"{col}2:{col}{last_row}")
    except Exception as e:
        print(f"WARN: Data validation not applied: {e}")

    # Formatting and styling
    apply_styles(ws, visible_headers)

    # Autofit columns and row heights
    from openpyxl.utils import get_column_letter
    max_width = {}
    for r in range(1, ws.max_row+1):
        for c in range(1, len(visible_headers)+1):
            val = ws.cell(row=r, column=c).value
            txt = "" if val is None else str(val)
            width = 0
            for part in txt.split("\n"):
                width = max(width, len(part))
            max_width[c] = max(max_width.get(c, 0), width)
    for c, width in max_width.items():
        adj = min(max(12, int(width * 1.1)), 120)
        ws.column_dimensions[get_column_letter(c)].width = adj
    for r in range(1, ws.max_row+1):
        max_lines = 1
        for c in range(1, len(visible_headers)+1):
            val = ws.cell(row=r, column=c).value
            txt = "" if val is None else str(val)
            lines = txt.count("\n") + 1 if txt else 1
            if lines > max_lines:
                max_lines = lines
        ws.row_dimensions[r].height = max(18, min(16 * max_lines + 4, 409))

    # Safety check on sheets
    names = [s.title for s in wb.worksheets]
    if "Data" in names:
        for s in wb.worksheets:
            if s.title == 'Data':
                wb.remove(s)
    names = [s.title for s in wb.worksheets]
    if set(names) != set(["TestPlan", "Meta_data_sheet"]):
        print(f"ERROR: Unexpected worksheets present: {names}", file=sys.stderr)
        sys.exit(3)

    # Compute IST time for naming
    now_ist = get_ist_now()
    date_tag = now_ist.strftime("%Y%m%d")
    time_tag = now_ist.strftime("%H%M%S")
    human_ts = now_ist.strftime("%Y-%m-%d %H:%M:%S")

    out_dir = os.path.join("Test_Output", "I2C", "TestPlan")
    os.makedirs(out_dir, exist_ok=True)
    out_name = f"I2C_TestPlan_{date_tag}_{time_tag}.xlsx"
    out_path = os.path.join(out_dir, out_name)

    wb.save(out_path)

    with zipfile.ZipFile(out_path, 'r') as zf:
        names = zf.namelist()
        if '[Content_Types].xml' not in names or 'xl/workbook.xml' not in names:
            print('ERROR: Not a valid XLSX structure', file=sys.stderr)
            sys.exit(4)
    load_workbook(out_path)

    meta = {
        "output_path": out_path.replace('\\', '/'),
        "commit_message": f"Add I2C TestPlan generated on {human_ts} (GMT+05:30)",
        "rows": len(records),
        "columns_visible": len(VISIBLE_ORDER)
    }
    with open("scripts/.gen_meta.json", "w", encoding="utf-8") as f:
        json.dump(meta, f)
    print(json.dumps(meta))

if __name__ == '__main__':
    main()
