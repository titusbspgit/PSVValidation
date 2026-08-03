#!/usr/bin/env python3
import json
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

# Configuration (bound from task inputs)
OWNER = "titusbspgit"
REPO = "PSVValidation"
BRANCH = "main"
OUTPUT_DIR = os.path.join("Test_Output", "USB", "TestPlan")
IP_NAME = "USB"

# Final aggregated Test Plan JSON (preserve order exactly as provided)
JSON_DATA = [
  {
    "Index": "1",
    "SS / Module": "USB",
    "Feature": "Register Access Sanity",
    "Test Case Name": "basic_test",
    "Test Description": "Perform a basic USB MMIO access check by reading two USB registers and executing a write-then-readback on a third location to confirm bus access integrity.",
    "Meta Test Description": "The test prints a banner, reads from address 0xA0000000, logs the value, reads from 0xA001706C, logs the value, writes 0xDEADBEEF to 0xA0240000, reads back from 0xA0240000, logs the value, and then calls finish(0). No explicit assertions or comparisons are present in the code.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "0xA0000000",
    "Memory End Offset": "0xA0240000",
    "Remarks": "NA",
    "Test Steps / Procedure": "1) Start the test and log the start banner. 2) Read and log the value from the first USB MMIO location. 3) Read and log the value from the second USB MMIO location. 4) Write 0xDEADBEEF to the target USB MMIO location and read back the same location. 5) Verify the read-back value equals the written value and record the result. 6) End the test.",
    "Meta Test Steps / Procedure": "1) printf(\"[C-Programme] Hello world\\n\"); 2) int rd_data = 0; 3) rd_data = read_reg(0xA0000000); 4) printf(\"THE READ DATA is %X\\n\", rd_data); 5) rd_data = read_reg(0xA001706C); 6) printf(\"THE READ DATA1 is %X\\n\", rd_data); 7) write_reg(0xA0240000, 0xDEADBEEF); 8) rd_data = read_reg(0xA0240000); 9) printf(\"THE READ DATA2 is %X\\n\", rd_data); 10) finish(0);",
    "Impacted Registers": "NA",
    "Meta Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "The value read back from the written MMIO location must equal 0xDEADBEEF; values should be readable from the other MMIO locations without access faults.",
    "Meta Validation / Acceptance Criteria": "No explicit assert/if in code. Implicit PASS if read_reg(0xA0240000) returns 0xDEADBEEF after write_reg(0xA0240000, 0xDEADBEEF). Console logs show: \"THE READ DATA is ...\", \"THE READ DATA1 is ...\", \"THE READ DATA2 is ...\". finish(0) ends the test.",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>\n#include <stdlib.h>\n#include <usb/usb_def.h>\n#include <usb/usb_offsets.h>",
    "Meta Macros": "NA",
    "Meta Arrays": "NA"
  },
  {
    "Index": "2",
    "SS / Module": "USB",
    "Feature": "Low-Speed (LS) Operation",
    "Test Case Name": "usb_host_enumeration_ls",
    "Test Description": "Exercise USB host stack to detect and enumerate a Low-Speed (LS) device on the root hub port.",
    "Meta Test Description": "NA",
    "Speed": "NA",
    "Mode": "NA",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Requires a Low-Speed USB device connected to the root hub port.",
    "Test Steps / Procedure": "1) Initialize the USB controller in Host mode. 2) Apply power to the root hub port and issue a port reset. 3) Detect device connect and confirm Low-Speed signaling. 4) Perform standard USB enumeration sequence (default address to assigned address and read descriptors). 5) Verify enumeration completes without errors.",
    "Meta Test Steps / Procedure": "NA",
    "Impacted Registers": "GUSB2PHYCFG; GAHBCFG; GUSBCFG; PCGCCTL; HCFG; HPRT; GINTSTS; GINTMSK; HPTXSTS; GRXSTSP; GRXFSIZ; GNPTXFSIZ; HCCHAR; HCTSIZ",
    "Meta Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "The Low-Speed device is detected on the root hub port, assigned an address, and responds to descriptor requests without controller error indications.",
    "Meta Validation / Acceptance Criteria": "NA",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "NA",
    "Meta Macros": "NA",
    "Meta Arrays": "NA"
  }
]

# Column schemas (must match exactly)
TESTPLAN_COLUMNS = [
  "Index",
  "SS / Module",
  "Feature",
  "Test Case Name",
  "Test Description",
  "Speed",
  "Mode",
  "Memory Start Offset",
  "Memory End Offset",
  "Remarks",
  "Test Steps / Procedure",
  "Impacted Registers",
  "Validation / Acceptance Criteria",
  "Code Generation (Required / Not)",
]

METADATA_COLUMNS = [
  "Index",
  "Test Case Name",
  "Meta Test Description",
  "Meta Test Steps / Procedure",
  "Meta Impacted Registers",
  "Meta Validation / Acceptance Criteria",
  "Meta Headers",
  "Meta Macros",
  "Meta Arrays",
]

# Styles
HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CELL_ALIGN = Alignment(wrap_text=True, vertical="top")


def write_sheet(ws, rows, columns):
    # Header
    for c_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CELL_ALIGN
    # Rows
    for r_idx, item in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(columns, start=1):
            val = item.get(col_name, "")
            ws.cell(row=r_idx, column=c_idx, value=val).alignment = CELL_ALIGN
    # Freeze header row
    ws.freeze_panes = "A2"
    # Column widths based on content length
    for c_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for r_idx in range(2, len(rows) + 2):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        # heuristic width: cap at 80, min 12
        width = min(max(12, max_len + 2), 80)
        ws.column_dimensions[get_column_letter(c_idx)].width = width


def main():
    # STEP 1: Validate JSON
    if not isinstance(JSON_DATA, list) or len(JSON_DATA) == 0:
        raise SystemExit("json_data must be a non-empty list of objects")
    for i, obj in enumerate(JSON_DATA):
        if not isinstance(obj, dict):
            raise SystemExit(f"Item at index {i} is not an object/dict")

    # STEP 2: Split data
    testplan_rows = []
    metadata_rows = []
    for obj in JSON_DATA:
        testplan_rows.append({k: obj.get(k, "") for k in TESTPLAN_COLUMNS})
        metadata_rows.append({k: obj.get(k, "") for k in METADATA_COLUMNS})

    # Timestamp in IST
    ist = ZoneInfo("Asia/Kolkata")
    ts = datetime.now(ist).strftime("%Y%m%d_%H%M%S")
    filename = f"{IP_NAME}_TestPlan_{ts}.xlsx"

    # Ensure output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, filename)

    # STEP 3: Create workbook, sheets, formatting
    wb = Workbook()
    ws_testplan = wb.active
    ws_testplan.title = "TestPlan"
    write_sheet(ws_testplan, testplan_rows, TESTPLAN_COLUMNS)

    ws_meta = wb.create_sheet(title="MetaData")
    write_sheet(ws_meta, metadata_rows, METADATA_COLUMNS)
    # VERY HIDDEN for MetaData
    ws_meta.sheet_state = "veryHidden"

    # STEP 4: Save real .xlsx
    wb.save(out_path)

    print(out_path)

if __name__ == "__main__":
    main()
