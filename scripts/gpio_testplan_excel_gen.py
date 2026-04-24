import json
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

HIDDEN_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

def list_to_multiline(value):
    if value is None:
        return ""
    if isinstance(value, list):
        lines = []
        for idx, item in enumerate(value, 1):
            if isinstance(item, (list, tuple)):
                sub = "; ".join(str(x) for x in item)
                lines.append(f"{idx}. {sub}")
            else:
                lines.append(f"{idx}. {item}")
        return "\n".join(lines)
    return str(value)


def autosize(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column if hasattr(col[0], 'column') else col[0].column_letter
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            if v:
                for line in v.splitlines():
                    if len(line) > max_length:
                        max_length = len(line)
        adjusted_width = min(120, max(10, max_length + 2))
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


def build_workbook(data, output_path: Path):
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "MAIN"

    # Headers
    headers = MAIN_COLUMNS + HIDDEN_COLUMNS
    ws_main.append(headers)
    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws_main.cell(row=1, column=c).font = header_font

    # Rows per test case, preserving order
    for tc in data.get("TEST_CASES", []):
        row = []
        for key in MAIN_COLUMNS:
            val = tc.get(key)
            if key in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                val = list_to_multiline(val)
            row.append(val)
        for key in HIDDEN_COLUMNS:
            val = tc.get(key)
            if key in ("Hidden_Test_Steps_Procedure", "Hidden_Validation_Acceptance_Criteria"):
                val = list_to_multiline(val)
            row.append(val)
        ws_main.append(row)

    # Wrap text for multi-line columns
    wrap_cols = [headers.index("Test Steps / Procedure") + 1,
                 headers.index("Validation / Acceptance Criteria") + 1,
                 headers.index("Hidden_Test_Steps_Procedure") + 1,
                 headers.index("Hidden_Validation_Acceptance_Criteria") + 1,
                 headers.index("Test Description") + 1,
                 headers.index("Remarks") + 1]
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row, min_col=1, max_col=ws_main.max_column):
        for cell in row:
            if cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top')

    autosize(ws_main)

    # META_DATA sheet
    ws_meta = wb.create_sheet(title="META_DATA")
    meta = data.get("META_DATA", {})
    for k, v in meta.items():
        ws_meta.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])
    autosize(ws_meta)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--output', required=True)
    args = ap.parse_args()

    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)

    build_workbook(data, Path(args.output))

if __name__ == '__main__':
    main()
