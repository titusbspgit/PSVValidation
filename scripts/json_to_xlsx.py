#!/usr/bin/env python3
import json, os, sys
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook

SHEET_NAME = 'Data'
JSON_PATH = 'data/json/testing_excel_local.json'
OUT_PATH = 'Test_Output/GPIO/TestPlan/testing_excel_local.xlsx'

def log(msg):
    print(msg, flush=True)

def parse_json(path):
    log('Parsing started')
    with open(path, 'r', encoding='utf-8') as f:
        raw = json.load(f)
    if isinstance(raw, dict):
        data = [raw]
    elif isinstance(raw, list):
        data = raw
    else:
        raise ValueError('JSON root must be object or array')
    if len(data) == 0:
        raise ValueError('JSON array is empty')
    headers = sorted({k for row in data for k in row.keys()})
    log(f'Parsing completed; rows={len(data)}; headers={headers}')
    return data, headers

def write_excel(data, headers, out_path):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(headers)
    for row in data:
        ws.append([row.get(h, '') for h in headers])
    wb.save(out_path)
    return out_path

def validate(out_path, data, headers):
    log('Validation started')
    wb = load_workbook(out_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise AssertionError(f'Missing sheet: {SHEET_NAME}')
    ws = wb[SHEET_NAME]
    read_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if read_headers != headers:
        raise AssertionError(f'Header mismatch. expected={headers} actual={read_headers}')
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append([r[i] for i in range(len(headers))])
    if len(rows) != len(data):
        raise AssertionError(f'Row count mismatch. expected={len(data)} actual={len(rows)}')
    for i, orig in enumerate(data):
        expected = [orig.get(h, '') for h in headers]
        actual = rows[i]
        exp_norm = [str(x) if x is not None else '' for x in expected]
        act_norm = [str(x) if x is not None else '' for x in actual]
        if exp_norm != act_norm:
            raise AssertionError(f'Row {i+1} mismatch. expected={exp_norm} actual={act_norm}')
    log('Validation completed successfully')

def main():
    retries = 3
    last_err = None
    data, headers = parse_json(JSON_PATH)
    for attempt in range(1, retries+1):
        try:
            out = write_excel(data, headers, OUT_PATH)
            validate(out, data, headers)
            log(f'File saved: {os.path.abspath(out)}')
            print(f'ROWS_PROCESSED={len(data)}')
            return 0
        except Exception as e:
            last_err = e
            log(f'Attempt {attempt} failed: {e}')
            if attempt == retries:
                break
    log(f'Failed after {retries} attempts: {last_err}')
    return 1

if __name__ == '__main__':
    sys.exit(main())
