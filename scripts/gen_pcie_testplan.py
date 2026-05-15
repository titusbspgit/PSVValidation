#!/usr/bin/env python3
# Deterministic fallback Excel generator for PCIE TestPlan
# - Consumes embedded JSON (array of objects)
# - Produces a single OOXML .xlsx with strict formatting rules
# - Writes the final file to Test_Output/PCIE/TestPlan/<PCIE_TestPlan_YYYYMMDD_HHMMSS.xlsx>
# - Emits generated_path.txt and ist_timestamp.txt for the workflow to commit with the correct message

import json
import os
import re
import zipfile
from datetime import datetime

try:
    from zoneinfo import ZoneInfo
except Exception:
    # Fallback for very old Pythons (not expected on GH runners)
    from pytz import timezone as ZoneInfo  # type: ignore

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ------------------------ INPUT JSON (embedded) ------------------------
JSON_INPUT = r'''[
  {
    "Index": 1,
    "SS / Module": "PCIE",
    "Feature": "SLOT_IMPLEMENTED_n",
    "Test Case Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Test Description": "Verifies default values of PCIe Root Complex interface registers and checks masked write and read behavior across the register set.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "0xE68C1000",
    "Memory End Offset": "NA",
    "Remarks": "Non-readable addresses are skipped. Non-writable addresses are skipped. Registers listed in the skip list are skipped. The PHY reset control register is excluded from default checks.",
    "Test Steps / Procedure": "1) Read all readable registers and compare each value against its documented default.\n2) For each test data pattern, write to all writable registers except those explicitly skipped.\n3) Read back each affected register and compare to the expected value derived from the write and read masks combined with preserved default bits.",
    "Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "1) Each readable register equals its documented default value → Pass\n2) For every written register, the read value matches the masked expected value → Pass\n3) No mismatches accumulated at test end → Pass",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Hidden_Test_Description": "DEFAULT VALUE_CHECK and WRITE & READ CHECK on PCIe1 SII RC registers. The test iterates over addr_array using read_mask_array, write_mask_array, default_value_array, and skip_array. It checks default values for readable addresses (excluding mizar_PCIE1_SII_PHY_RST_CONTROL), performs masked writes with multiple patterns, reads back with masks applied, compares to expected values, updates def_fail_cnt and wr_fail_cnt on mismatches, and calls finish(1) if any failures else finish(0).",
    "Hidden_Remarks": "1) Addresses with read_mask_array[i] == 0x00000000 are skipped for reading.\n2) Addresses with write_mask_array[i] == 0x00000000 are skipped for writing (and skipped for reading in the write-read phase).\n3) Addresses with skip_array[i] == 1 are skipped for both writing and reading.\n4) Address equal to mizar_PCIE1_SII_PHY_RST_CONTROL is skipped in default value check.\n5) DEBUG_DISPLAY, when enabled, prints diagnostic messages.",
    "Hidden_Test_Steps_Procedure": "Top-level flow (test_case):\n1) Call chk_rst_val().\n2) Print \"********* Default value check end ************\" if DEBUG_DISPLAY.\n3) Call chk_rd_wr().\n4) Print \"********* Write & Read from registers end ************\" if DEBUG_DISPLAY.\n5) If (def_fail_cnt > 0 || wr_fail_cnt > 0) then finish(1) else finish(0).\n\nFunction chk_rst_val():\n1) For i from 0 to CNT-1:\n   a) addr = addr_array[i].\n   b) If (read_mask_array[i] == 0x00000000):\n      - If DEBUG_DISPLAY, print: \"RST : This address 0x%x is not readable, hence skipped for reading\" with addr.\n      - continue.\n   c) If (addr_array[i] == mizar_PCIE1_SII_PHY_RST_CONTROL):\n      - continue.\n   d) data_rd = read_reg(addr).\n   e) If (data_rd == default_value_array[i]):\n      - If DEBUG_DISPLAY, print: \"RST : PASS Reading Default value from Address :0x%x Expected : 0x%x\\tRead_data : 0x%x\" with addr, default_value_array[i], data_rd.\n     Else:\n      - def_fail_cnt++.\n      - Print: \"RST : Failed Default value mismatch Addr :0x%x Expected : 0x%x\\tRead_data : 0x%x\" with addr, default_value_array[i], data_rd.\n\nFunction chk_rd_wr():\n1) Define chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0x00000000, 0xA5A5A5A5, 0xffff0000}.\n2) For j from 0 to 5:\n   a) data_wr = chk_val[j].\n   b) Write phase: For i from 0 to CNT-1:\n      - addr = addr_array[i].\n      - If (skip_array[i] == 1):\n        * If DEBUG_DISPLAY, print: \"Read_write : Writing into this Address : 0x%x is skipped because address present in skip_array\" with addr.\n        * continue.\n      - If (write_mask_array[i] == 0x00000000):\n        * If DEBUG_DISPLAY, print: \"Read_write : This address 0x%x is not writable, hence skipped for writing\" with addr.\n        * continue.\n      - Else:\n        * write_reg(addr, data_wr).\n        * If DEBUG_DISPLAY, print: \"Read_write : Writing into register Address : 0x%x\\tdata :0x%x\" with addr, data_wr.\n   c) Read/verify phase: For i from 0 to CNT-1:\n      - addr = addr_array[i].\n      - If (skip_array[i] == 1):\n        * If DEBUG_DISPLAY, print: \"Read_write : Reading from this Address : 0x%x is skipped because address present in skip_array\" with addr.\n        * continue.\n      - If (write_mask_array[i] == 0x00000000):\n        * If DEBUG_DISPLAY, print: \"Read_write : This address 0x%x is not Writable , hence skipped for reading\" with addr.\n        * continue.\n      - If (read_mask_array[i] == 0x00000000):\n        * If DEBUG_DISPLAY, print: \"Read_write : This address 0x%x is not Readable , hence skipped for reading\" with addr.\n        * continue.\n      - Else:\n        * data_rd = read_reg(addr).\n        * wr_n = (write_mask_array[i] ^ 0xffffffff).\n        * exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])).\n        * If (data_rd == exp_val):\n          - If DEBUG_DISPLAY, print: \"Read_write : PASS : For Address %x, Expected value=0x%x\\tRead value=0x%x\" with addr, exp_val, data_rd.\n         Else:\n          - wr_fail_cnt++.\n          - Print: \"Read_write : Failed : Write Read mismatch For Address %x, Expected value=0x%x\\tRead value=0x%x\" with addr, exp_val, data_rd.\n\nNote: soft_reset_chk() present but not invoked in test_case. It reads SOFT_RST_REG_ADDRESS, writes SOFT_RST_REG_DATA, waits, restores default_value, waits.",
    "Hidden_Impacted_Registers": "mizar_PCIE1_SII_CFG_BAR0_START1\nmizar_PCIE1_SII_CFG_BAR0_START2\nmizar_PCIE1_SII_CFG_BAR0_LIMIT1\nmizar_PCIE1_SII_CFG_BAR0_LIMIT2\nmizar_PCIE1_SII_CFG_BAR1_START\nmizar_PCIE1_SII_CFG_BAR1_LIMIT1\nmizar_PCIE1_SII_CFG_BAR2_START1\nmizar_PCIE1_SII_CFG_BAR2_START2\nmizar_PCIE1_SII_CFG_BAR2_LIMIT1\nmizar_PCIE1_SII_CFG_BAR2_LIMIT2\nmizar_PCIE1_SII_CFG_BAR3_START\nmizar_PCIE1_SII_CFG_BAR3_LIMIT\nmizar_PCIE1_SII_CFG_BAR4_START1\nmizar_PCIE1_SII_CFG_BAR4_START2\nmizar_PCIE1_SII_CFG_BAR4_LIMIT1\nmizar_PCIE1_SII_CFG_BAR4_LIMIT2\nmizar_PCIE1_SII_CFG_BAR5_START\nmizar_PCIE1_SII_CFG_BAR5_LIMIT\nmizar_PCIE1_SII_PCIE1_CONFIG_INFO1\nmizar_PCIE1_SII_PCIE1_CONFIG_INFO2\nmizar_PCIE1_SII_PCIE1_GEN_CONTROL1\nmizar_PCIE1_SII_PCIE1_GEN_CONTROL2\nmizar_PCIE1_SII_PCIE1_GEN_CONTROL3\nmizar_PCIE1_SII_PCIE1_PM_CONTROL\nmizar_PCIE1_SII_PCIE1_CONTROL_PM_STS\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER1\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3\nmizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER4\nmizar_PCIE1_SII_PCIE1_TRANSMIT_REQ\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR1\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR2\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR3\nmizar_PCIE1_SII_PCIE1_RCV_MSG_HDR4\nmizar_PCIE1_SII_PCIE1_RCV_MSG_STS\nmizar_PCIE1_SII_RCV_INTERRPUT_CTRL\nmizar_PCIE1_SII_CFG_EXP_ROM_START\nmizar_PCIE1_SII_CFG_EXP_ROM_LIMIT\nmizar_PCIE1_SII_CFG_EXP_ROM_INFO\nmizar_PCIE1_SII_CXPL_DEBUG_INFO1\nmizar_PCIE1_SII_CXPL_DEBUG_INFO2\nmizar_PCIE1_SII_CXPL_DEBUG_INFO_EI\nmizar_PCIE1_SII_PCIE1_TARGET_INFO1\nmizar_PCIE1_SII_PCIE1_TARGET_INFO2\nmizar_PCIE1_SII_PCIE1_CONTOLLER_ERROR_STATUS\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL\nmizar_PCIE1_SII_PHY_RST_CONTROL\nmizar_PCIE1_SII_LINK_DEBUG_DATA\nmizar_PCIE1_SII_PCIE1_ERR_STS\nmizar_PCIE1_SII_PCIE1_ERR_INTERRUPT_CTRL\nmizar_PCIE1_SII_CFG_MSI_INT\nmizar_PCIE1_SII_LTR_MSG\nmizar_PCIE1_SII_LTR_MSG_LATENCY\nmizar_PCIE1_SII_APP_LTR_LATENCY\nmizar_PCIE1_SII_CFG_LTR_MAX_LATENCY\nmizar_PCIE1_SII_OBFF_CNTRL\nmizar_PCIE1_SII_SLV_AWMISC_INFO\nmizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_HI\nmizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_LO\nmizar_PCIE1_SII_SLV_MISC_INFO\nmizar_PCIE1_SII_SLV_MISC_RESP_INFO\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_CNTRL\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_1\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_0\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_HI\nmizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_LO\nmizar_PCIE1_SII_MSTR_ARMISC_INFO_CNTRL\nmizar_PCIE1_SII_MSTR_ARMISC_INFO_1\nmizar_PCIE1_SII_MSTR_ARMISC_INFO_0\nmizar_PCIE1_SII_MSTR_BMISC_RMISC_CPL_STAT_INFO\nmizar_PCIE1_SII_RADM_TIMEOUT_INFO\nmizar_PCIE1_SII_CFG_MSI_INFO\nmizar_PCIE1_SII_CFG_MSI_DATA\nmizar_PCIE1_SII_CFG_MSI_ADDR_HI\nmizar_PCIE1_SII_CFG_MSI_ADDR_LO\nmizar_PCIE1_SII_CFG_AER_INT_AND_PCIE1_CAP_INT_MSG\nmizar_PCIE1_SII_RTLH_RFC_DATA\nmizar_PCIE1_SII_APP_HDR_INFO\nmizar_PCIE1_SII_APP_HDR_LOG_3\nmizar_PCIE1_SII_APP_HDR_LOG_2\nmizar_PCIE1_SII_APP_HDR_LOG_1\nmizar_PCIE1_SII_APP_HDR_LOG_0\nmizar_PCIE1_SII_CFG_BUS_NUM\nmizar_PCIE1_SII_CFG_BR_CTRL_SERREN\nmizar_PCIE1_SII_APP_DEV_AND_BUS_NUM\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS_1\nmizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL_1\nmizar_PCIE1_SII_APP_AND_SLOT_CONTROL_REG\nmizar_PCIE1_SII_DIAG_CTRL_BUS\nmizar_PCIE1_SII_CFG_REG_RO\nmizar_PCIE1_SII_CFG_ARI_FWD_EN\nmizar_PCIE1_SII_RADM_SLOT_PWR_PAYLOAD\nmizar_PCIE1_SII_DIAG_STATUS_BUS_0\nmizar_PCIE1_SII_DIAG_STATUS_BUS_1\nmizar_PCIE1_SII_DIAG_STATUS_BUS_2\nmizar_PCIE1_SII_DIAG_STATUS_BUS_3\nmizar_PCIE1_SII_DIAG_STATUS_BUS_4\nmizar_PCIE1_SII_DIAG_STATUS_BUS_5\nmizar_PCIE1_SII_DIAG_STATUS_BUS_6\nmizar_PCIE1_SII_DIAG_STATUS_BUS_7\nmizar_PCIE1_SII_DIAG_STATUS_BUS_8\nmizar_PCIE1_SII_DIAG_STATUS_BUS_9\nmizar_PCIE1_SII_DIAG_STATUS_BUS_10\nmizar_PCIE1_SII_DIAG_STATUS_BUS_11\nmizar_PCIE1_SII_DIAG_STATUS_BUS_12\nmizar_PCIE1_SII_DIAG_STATUS_BUS_13\nmizar_PCIE1_SII_DIAG_STATUS_BUS_14\nmizar_PCIE1_SII_DIAG_STATUS_BUS_15\nmizar_PCIE1_SII_DIAG_STATUS_BUS_16\nmizar_PCIE1_SII_DIAG_STATUS_BUS_17\nmizar_PCIE1_SII_DIAG_STATUS_BUS_18\nmizar_PCIE1_SII_DIAG_STATUS_BUS_19\nmizar_PCIE1_SII_RAM_PWR_CNTRL_0\nmizar_PCIE1_SII_RAM_PWR_CNTRL_1\nmizar_PCIE1_SII_SOFT_RESET_CTRL\nmizar_PCIE1_SII_CFG_MSI_PENDING_B\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_1\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_2\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_3\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_4\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_5\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_6\nmizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_7\nmizar_PCIE1_SII_PHY_CONTROL_0\nmizar_PCIE1_SII_PHY_CONTROL_1\nmizar_PCIE1_SII_PHY_CONTROL_2\nmizar_PCIE1_SII_PHY_CONTROL_3\nmizar_PCIE1_SII_PHY_CONTROL_4\nmizar_PCIE1_SII_PHY_CONTROL_5\nmizar_PCIE1_SII_PHY_CONTROL_6\nmizar_PCIE1_SII_PHY_CONTROL_7\nmizar_PCIE1_SII_PHY_CONTROL_8\nmizar_PCIE1_SII_PHY_CONTROL_9\nmizar_PCIE1_SII_PHY_CONTROL_10\nmizar_PCIE1_SII_PHY_CONTROL_11\nmizar_PCIE1_SII_PHY_CONTROL_12\nmizar_PCIE1_SII_PHY_CONTROL_13\nmizar_PCIE1_SII_PHY_CONTROL_14\nmizar_PCIE1_SII_PHY_CONTROL_15\nmizar_PCIE1_SII_PHY_CONTROL_16\nmizar_PCIE1_SII_PHY_CONTROL_17\nmizar_PCIE1_SII_PHY_CONTROL_18\nmizar_PCIE1_SII_PHY_CONTROL_19\nmizar_PCIE1_SII_PHY_CONTROL_20\nmizar_PCIE1_SII_PHY_CONTROL_21\nmizar_PCIE1_SII_PHY_CONTROL_22\nmizar_PCIE1_SII_PHY_CONTROL_23\nmizar_PCIE1_SII_PHY_CONTROL_24\nmizar_PCIE1_SII_PHY_CONTROL_25\nmizar_PCIE1_SII_MSI_CTRL_IO\nmizar_PCIE1_SII_MSI_CTRL_INT_VEC",
    "Hidden_Validation_Acceptance_Criteria": "Default value check: For each i where read_mask_array[i] != 0x00000000 and addr_array[i] != mizar_PCIE1_SII_PHY_RST_CONTROL, data_rd = read_reg(addr_array[i]) must equal default_value_array[i]. On mismatch, def_fail_cnt++ and a failure message is printed. Write/read check: For each pattern data_wr in {0xffffffff, 0xaaaaaaaa, 0x55555555, 0x00000000, 0xA5A5A5A5, 0xffff0000}, for each i where skip_array[i] == 0 and write_mask_array[i] != 0x00000000 and read_mask_array[i] != 0x00000000, after write_reg(addr_array[i], data_wr), data_rd = read_reg(addr_array[i]) must equal exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | ((write_mask_array[i] ^ 0xffffffff) & read_mask_array[i] & default_value_array[i])). On mismatch, wr_fail_cnt++ and a failure message is printed. End criterion: if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) else finish(0).",
    "Hidden_Header_Includes": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<pcie1/pcie_sii_rc_def.h>\n#include<pcie1/pcie_sii_rc_offset.h>",
    "Hidden_Macro_Defines": "#define MIZAR_PCIE1_SII_BASE     0xE68C1000\n#define CNT 153\n#define SOFT_RST_REG_ADDRESS 0x00000000\n#define SOFT_RST_REG_DATA 0x00000000",
    "Hidden_Skip_Array_Definition": "const int skip_array[153]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,}"
  },
  {
    "Index": 2,
    "SS / Module": "PCIE",
    "Feature": "BAR5_ENABLED_n",
    "Test Case Name": "pcie_cfg_wr_rd_test",
    "Test Description": "Validates PCIe configuration space access and BAR programming after link bring-up. Ensures status readiness and completion handshake before finishing.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "0xE6004100",
    "Memory End Offset": "NA",
    "Remarks": "Compile-time flags select instance and role. Debug prints are optional. Status is polled until ready. External handshake register is used to signal test completion.",
    "Test Steps / Procedure": "1) Configure coherency control registers and apply settings.\n2) Poll the status register until required bits indicate ready.\n3) Read initial configuration header registers.\n4) Program all BAR registers with a probing pattern, read back values, then program target addresses and read back again.\n5) Enable the command register for memory, I/O, and bus mastering.\n6) Wait for the external completion handshake via the control register and finish.",
    "Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "1) Status condition indicates ready → Proceed with configuration access.\n2) External completion control equals the expected value → Test passes.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie_cfg_wr_rd_test",
    "Hidden_Test_Description": "Performs PCIe configuration-space BAR write/read sequencing after link training and coherency setup. Steps: write_reg(0xE6004100,0x0); optional link training via compile-time paths; CACHE PROGRAMMING: for each of PCIE0 and PCIE1 DBI coherency control 3 registers, set bits [11:14] and [3:6] to 0xF, write back; then set bits [27:30] and [19:22] to 0xF, write back; wait_on(20); for each, read current, set [11:14], [3:6], [27:30], [19:22] to 0xF, and write back. Read SII0 status register at offset 0xC0 and poll while ((data_rd & 0xD1) != 0xD1); if DM1_RC, do the same for SII1. Write_reg(0xE6004100,0x11111111); wait_on(15000). If DM0_RC path: mem_base_program_dm0_x4(); read first 10 configuration dwords (offsets 0x00..0x24) via read_pcie_slv0_reg. First set of BAR writes: write 0xFFFFFFFF to offsets 0x10, 0x14, 0x18, 0x1C, 0x20, 0x24; then read each back. Second set of BAR writes: write 0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000 respectively to offsets 0x10..0x24; then read each back. Enable memory, I/O, and bus mastering by write_pcie_slv0_reg(0x4, 0x7). If DM1_RC path: mem_base_program_dm1_x4(); read first 10 configuration dwords for instance 1; enable command (0x4←0x7); perform analogous BAR writes (0xFFFFFFFF, reads; then 0x0/0x4/0x20000000/0x40000000/0x60000000/0x80000000) and reads for instance 1. Finally poll read_reg(0xE6004100) until it equals 0x12345678; finish(0).",
    "Hidden_Remarks": "Conditional compilation controls RC/EP instance selection (DM0_RC, DM1_RC, DM0_EP, DM1_EP). DEBUG_DISPLAY enables prints. Status polling uses SII0/SII1 offset 0xC0 with mask 0xD1 until equal to 0xD1. Handshake at control address 0xE6004100 waits for 0x12345678 before finish(0).",
    "Hidden_Test_Steps_Procedure": "1) write_reg(0xE6004100, 0x0).\n2) Optionally call link_training_dm0_x4(4)/link_training_dm1_x4(4) depending on compile-time flags.\n3) For mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF: rd=read_reg(...); rd=set_data(rd,11,14,0xF); rd=set_data(rd,3,6,0xF); write_reg(..., rd); then rd=read_reg(...); rd=set_data(rd,27,30,0xF); rd=set_data(rd,19,22,0xF); write_reg(..., rd).\n4) Repeat step 3 for mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF.\n5) wait_on(20); For PCIE0 coherency register: rd=read_reg(...); set [11:14],[3:6],[27:30],[19:22] to 0xF; write back. Repeat for PCIE1.\n6) data_rd = read_sii0_reg(0xC0); while(((data_rd)&(0xD1)) != 0xD1) { data_rd = read_sii0_reg(0xC0); }.\n7) If DM1_RC: data_rd = read_sii1_reg(0xC0); while(((data_rd)&(0xD1)) != 0xD1) { data_rd = read_sii1_reg(0xC0); }.\n8) write_reg(0xE6004100, 0x11111111).\n9) wait_on(15000).\n10) If DM0_RC: mem_base_program_dm0_x4(); for (i=0;i<10;i++) read_pcie_slv0_reg(i*0x4). Write 0xFFFFFFFF to 0x10,0x14,0x18,0x1C,0x20,0x24 via write_pcie_slv0_reg(); read back each via read_pcie_slv0_reg(). Then write 0x0,0x4,0x20000000,0x40000000,0x60000000,0x80000000 to 0x10..0x24 and read back each. write_pcie_slv0_reg(0x4,0x7).\n11) If DM1_RC: mem_base_program_dm1_x4(); read first 10 dwords via read_pcie_slv1_reg(i*0x4); then perform analogous writes/reads at 0x10..0x24; write_pcie_slv1_reg(0x4,0x7).\n12) wait_on(10); data_rd = read_reg(0xE6004100); while(data_rd != 0x12345678) { wait_on(5); data_rd = read_reg(0xE6004100); } finish(0).",
    "Hidden_Impacted_Registers": "mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF\nmizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF",
    "Hidden_Validation_Acceptance_Criteria": "Link/status readiness: loop until (read_sii0_reg(0xC0) & 0xD1) == 0xD1; if DM1_RC, also ensure (read_sii1_reg(0xC0) & 0xD1) == 0xD1. Completion handshake: poll read_reg(0xE6004100) until it equals 0x12345678. On reaching expected handshake value, call finish(0) → pass.",
    "Hidden_Header_Includes": "#include <stdlib.h>\n#include <stdio.h>\n#include <test_common.h>\n#include \"pcie.h\"",
    "Hidden_Macro_Defines": "NA",
    "Hidden_Skip_Array_Definition": "NA"
  }
]'''
# ---------------------------------------------------------------------

# Configuration
IP_NAME = "PCIE"
OUTPUT_DIR = os.environ.get("EXPORT_DIR", os.path.join("Test_Output", IP_NAME, "TestPlan"))
TIMEZONE = ZoneInfo("Asia/Kolkata")

# Required column orders
META_COLS_CANONICAL = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
    "Hidden_Header_Includes",
    # Accept either singular or plural macro key; will pick what exists in data
    "Hidden_Macro_Define",
    "Hidden_Skip_Array_Definition",
]

MAIN_COLS_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

# Helper: schema union preserving first-seen order
def build_schema_union(rows):
    seen = []
    seen_set = set()
    for obj in rows:
        for k in obj.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen

# Helper: approximate autofit width from text length
# Use a cap to avoid excessive widths
COL_WIDTH_MIN = 10
COL_WIDTH_MAX = 80

def calc_width(text: str) -> int:
    if text is None:
        return COL_WIDTH_MIN
    s = str(text)
    # Tab/newlines increase perceived width slightly
    s = s.replace("\t", "    ")
    max_line = max((len(line) for line in s.splitlines()), default=0)
    return max(COL_WIDTH_MIN, min(COL_WIDTH_MAX, max_line + 2))

# Renumber content inside a single cell as "1. ...\n2. ..."
# Keep original order; strip any existing leading numbering/bullets
LEADING_NUM_RE = re.compile(r"^\s*([0-9]+[\).:-]?\s*|-\s*|•\s*|\*\s*)")

def renumber_cell(val: str) -> str:
    if val is None:
        return ""
    text = str(val).strip()
    if not text:
        return ""
    # Split on newlines; ignore empty lines
    raw_lines = [ln for ln in re.split(r"\r?\n", text) if ln.strip()]
    out_lines = []
    for idx, ln in enumerate(raw_lines, start=1):
        ln2 = LEADING_NUM_RE.sub("", ln.strip())
        out_lines.append(f"{idx}. {ln2}")
    return "\n".join(out_lines)

# Validate XLSX is a proper OOXML zip
REQUIRED_XLSX_ENTRIES = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}

def validate_xlsx(path: str) -> bool:
    try:
        with zipfile.ZipFile(path, "r") as zf:
            names = set(zf.namelist())
        return REQUIRED_XLSX_ENTRIES.issubset(names)
    except Exception:
        return False

# Load JSON
rows = json.loads(JSON_INPUT)
if not isinstance(rows, list) or len(rows) == 0:
    raise SystemExit("ERROR: JSON input is empty or not an array")

# Build full schema (for staging Data sheet)
schema = build_schema_union(rows)

# Create workbook and staging sheet "Data"
wb = Workbook()
ws = wb.active
ws.title = "Data"

# Write headers
for c, key in enumerate(schema, start=1):
    ws.cell(row=1, column=c, value=key)

# Write rows exactly preserving values
for r, obj in enumerate(rows, start=2):
    for c, key in enumerate(schema, start=1):
        val = obj.get(key, "")
        ws.cell(row=r, column=c, value=val)

# Base formatting on staging sheet
ws.freeze_panes = "A2"
header_font = Font(bold=True)
for c in range(1, len(schema) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = header_font

# Approximate auto-fit column widths
for c, key in enumerate(schema, start=1):
    maxw = calc_width(key)
    for r in range(2, len(rows) + 2):
        v = ws.cell(row=r, column=c).value
        maxw = max(maxw, calc_width(v))
    ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = maxw

# Create META sheet and copy existing META columns (AS-IS, raw)
meta_ws = wb.create_sheet("Meta_data_sheet")
# Determine which META keys exist in schema; handle singular/plural macro define
existing_meta_keys = []
for k in META_COLS_CANONICAL:
    if k in schema:
        existing_meta_keys.append(k)
# If singular not found but plural present, include plural
if "Hidden_Macro_Define" not in existing_meta_keys and "Hidden_Macro_Defines" in schema:
    existing_meta_keys.append("Hidden_Macro_Defines")

# Write META headers
for c, key in enumerate(existing_meta_keys, start=1):
    meta_ws.cell(row=1, column=c, value=key)

# Write META rows
for r, obj in enumerate(rows, start=2):
    for c, key in enumerate(existing_meta_keys, start=1):
        meta_ws.cell(row=r, column=c, value=obj.get(key, ""))

# Very hide META sheet
meta_ws.sheet_state = 'veryHidden'

# Normalize MAIN sheet in-place: rename Data -> TestPlan
ws.title = "TestPlan"

# Rebuild TestPlan content strictly with MAIN_COLS_ORDER (remove META columns)
# Construct table from input rows for visible columns only
visible_headers = MAIN_COLS_ORDER[:]
# Ensure any missing columns are still present (will be blanks)
# Clear current TestPlan sheet
for row in ws[1:ws.max_row]:
    for cell in row:
        cell.value = None

# Write visible headers
for c, key in enumerate(visible_headers, start=1):
    ws.cell(row=1, column=c, value=key)

# Write visible rows (with numbering for the two specific columns)
for r, obj in enumerate(rows, start=2):
    for c, key in enumerate(visible_headers, start=1):
        val = obj.get(key, "")
        if key in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
            val = renumber_cell(val)
        ws.cell(row=r, column=c, value=val)

# Strict formatting for TestPlan
blue_fill = PatternFill("solid", fgColor="B7DEE8")  # light blue for readability
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
cell_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Header formatting
for c in range(1, len(visible_headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True)
    cell.alignment = hdr_align
    cell.fill = blue_fill

# Column-specific wrapping and alignment
wrap_cols = {"Test Description", "Remarks", "Test Steps / Procedure", "Validation / Acceptance Criteria"}
for c, key in enumerate(visible_headers, start=1):
    # Auto-fit widths again for visible columns
    maxw = calc_width(key)
    for r in range(2, len(rows) + 2):
        v = ws.cell(row=r, column=c).value
        maxw = max(maxw, calc_width(v))
        # Borders
        ws.cell(row=r, column=c).border = thin_border
        # Alignment per column
        if key in wrap_cols:
            ws.cell(row=r, column=c).alignment = cell_left
        elif key == "Index":
            ws.cell(row=r, column=c).alignment = cell_center
        else:
            # Default text left/top
            ws.cell(row=r, column=c).alignment = cell_left
    ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = maxw

# Apply borders to header too
for c in range(1, len(visible_headers) + 1):
    ws.cell(row=1, column=c).border = thin_border

# Approximate row heights post-wrap: set a minimum height, scale with line breaks
BASE_ROW_HEIGHT = 15
for r in range(2, len(rows) + 2):
    # Estimate max lines across wrapped columns
    max_lines = 1
    for c, key in enumerate(visible_headers, start=1):
        if key in wrap_cols:
            val = ws.cell(row=r, column=c).value or ""
            lines = str(val).count("\n") + 1
            if lines > max_lines:
                max_lines = lines
    ws.row_dimensions[r].height = BASE_ROW_HEIGHT * max(1, min(max_lines, 10))

# Data validation for Code Generation (Required / Not)
if "Code Generation (Required / Not)" in visible_headers:
    col_idx = visible_headers.index("Code Generation (Required / Not)") + 1
    start_row = 2
    end_row = len(rows) + 1
    rng = f"{ws.cell(row=1, column=col_idx).column_letter}{start_row}:{ws.cell(row=1, column=col_idx).column_letter}{end_row}"
    dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showDropDown=True)
    dv.error = "Select only: Required, Blank, Not Required"
    dv.errorTitle = "Invalid choice"
    ws.add_data_validation(dv)
    dv.add(rng)

# Safety: ensure no sheet named 'Data' remains
if "Data" in [s.title for s in wb.worksheets]:
    # Attempt deletion
    try:
        del wb["Data"]
    except Exception as e:
        raise SystemExit(f"Validation error: Unable to delete 'Data' sheet: {e}")

# Ensure only TestPlan (visible) and Meta_data_sheet (veryHidden) exist
allowed = {"TestPlan", "Meta_data_sheet"}
existing = set(ws.title for ws in wb.worksheets)
existing.add(meta_ws.title)
for s in list(wb.sheetnames):
    if s not in allowed:
        # Should not occur, but be defensive
        try:
            del wb[s]
        except Exception:
            pass

# Compute IST timestamp and filename
now_ist = datetime.now(TIMEZONE)
filename = f"{IP_NAME}_TestPlan_{now_ist:%Y%m%d}_{now_ist:%H%M%S}.xlsx"
rel_path = os.path.join(OUTPUT_DIR, filename)
abs_dir = OUTPUT_DIR
os.makedirs(abs_dir, exist_ok=True)

# Save workbook
wb.save(rel_path)

# Validate OOXML zip structure
if not validate_xlsx(rel_path):
    raise SystemExit("ERROR: XLSX validation failed (not a proper OOXML workbook)")

# Emit helper files for the workflow
with open("generated_path.txt", "w", encoding="utf-8") as f:
    f.write(rel_path)
with open("ist_timestamp.txt", "w", encoding="utf-8") as f:
    f.write(now_ist.strftime("%Y-%m-%d %H:%M:%S"))

print(f"Generated: {rel_path}")
