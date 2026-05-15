#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Deterministic fallback automation: Convert embedded JSON array to a true binary XLSX
and commit only the finalized Excel (commit performed by the workflow job).

Strictly follows the required formatting and sheet rules.
"""
import argparse
import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from zipfile import ZipFile
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Embedded FULL_JSON_STRUCTURE (provided by agent). Do not modify.
EMBEDDED_JSON = r'''[
  {
    "Index": 1,
    "SS / Module": "PCIE1 SII RC",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Test Description": "Verify that the PCIe controller registers report their reset defaults and support masked write and read-back across the defined register set.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Registers not readable or not writable are skipped. Entries in the skip list are skipped. One reset control register is omitted from the default check. Debug prints are optional.",
    "Test Steps / Procedure": "1) For each register in the defined list, if it is readable, read the value and compare with the documented default.\n2) For each data pattern, write the pattern to every writable register that is not skipped.\n3) For each such register, read back the value and compare against the expected value derived from the read and write masks and the default.\n4) Record any mismatches and report the final pass or fail result.",
    "Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "1) Default value read for each readable register → Equals documented default.\n2) Read after write for each writable and readable register → Equals expected masked value.\n3) No accumulated failures at test end → Overall pass.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie1_sii_rc_reg_wr_rd_test",
    "Hidden_Test_Description": "Performs two checks: (a) default value check across the PCIE1 SII RC register list using read_mask_array and default_value_array; (b) write & read check across the same list using test patterns and write_mask_array/read_mask_array to compute expected values. In test_case(): chk_rst_val() is called, then chk_rd_wr(); based on def_fail_cnt and wr_fail_cnt, finish(1) is invoked on any failure else finish(0). In chk_rst_val(): for i = 0..CNT-1, addr = addr_array[i]; if read_mask_array[i] == 0x00000000 then print skip and continue; if addr == mizar_PCIE1_SII_PHY_RST_CONTROL then continue; data_rd = read_reg(addr); if (data_rd == default_value_array[i]) print PASS; else increment def_fail_cnt and print failure. In chk_rd_wr(): define chk_val[6] = {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}; for each j pattern: data_wr = chk_val[j]; write loop i = 0..CNT-1: addr = addr_array[i]; if skip_array[i] == 1 print skip and continue; if write_mask_array[i] == 0x00000000 print skip and continue; else write_reg(addr, data_wr) and optionally print. Read/verify loop i = 0..CNT-1: addr = addr_array[i]; if skip_array[i] == 1 continue; if write_mask_array[i] == 0x00000000 continue; if read_mask_array[i] == 0x00000000 continue; else data_rd = read_reg(addr); wr_n = (write_mask_array[i] ^ 0xffffffff); exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if (data_rd == exp_val) print PASS; else increment wr_fail_cnt and print failure. soft_reset_chk() exists (reads SOFT_RST_REG_ADDRESS, writes SOFT_RST_REG_DATA, waits, restores default and waits) but is commented out and not executed.",
    "Hidden_Remarks": "Addresses with read_mask_array[i] == 0x00000000 are skipped for default read. In default value check, the address equal to mizar_PCIE1_SII_PHY_RST_CONTROL is skipped. During write/read, entries with skip_array[i] == 1 are skipped; addresses with write_mask_array[i] == 0x00000000 or read_mask_array[i] == 0x00000000 are skipped. DEBUG_DISPLAY gates printf diagnostics. soft_reset_chk() is defined but not invoked in test_case().",
    "Hidden_Test_Steps_Procedure": "Entry: test_case()\n- Call chk_rst_val(). If DEBUG_DISPLAY, print \"********* Default value check end ************\" after completion.\n- Call chk_rd_wr(). If DEBUG_DISPLAY, print \"********* Write & Read from registers end ************\" after completion.\n- If (def_fail_cnt > 0 || wr_fail_cnt > 0) then finish(1); else finish(0).\n\nFunction: chk_rst_val()\n- For (i = 0; i < CNT; i++):\n  - addr = addr_array[i].\n  - If (read_mask_array[i] == 0x00000000): if DEBUG_DISPLAY print skip message; continue.\n  - If (addr_array[i] == mizar_PCIE1_SII_PHY_RST_CONTROL): continue.\n  - data_rd = read_reg(addr).\n  - If (data_rd == default_value_array[i]): if DEBUG_DISPLAY print PASS with address, expected, read_data.\n    Else: def_fail_cnt++; print failure with address, expected, read_data.\n\nFunction: chk_rd_wr()\n- Define patterns: int chk_val[6] = {0xffffffff, 0xaaaaaaaa, 0x55555555, 0x00000000, 0xA5A5A5A5, 0xffff0000}.\n- For (j = 0; j < 6; j++):\n  - data_wr = chk_val[j].\n  - Write phase: For (i = 0; i < CNT; i++):\n    - addr = addr_array[i].\n    - If (skip_array[i] == 1): if DEBUG_DISPLAY print skip message; continue.\n    - If (write_mask_array[i] == 0x00000000): if DEBUG_DISPLAY print not-writable skip; continue.\n    - Else: write_reg(addr, data_wr); if DEBUG_DISPLAY print write details.\n  - Read/verify phase: For (i = 0; i < CNT; i++):\n    - addr = addr_array[i].\n    - If (skip_array[i] == 1): if DEBUG_DISPLAY print skip message; continue.\n    - If (write_mask_array[i] == 0x00000000): if DEBUG_DISPLAY print not-writable skip; continue.\n    - If (read_mask_array[i] == 0x00000000): if DEBUG_DISPLAY print not-readable skip; continue.\n    - Else:\n      - data_rd = read_reg(addr).\n      - wr_n = (write_mask_array[i] ^ 0xffffffff).\n      - exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i]) ).\n      - If (data_rd == exp_val): if DEBUG_DISPLAY print PASS with address, expected, read.\n        Else: wr_fail_cnt++; print failure with address, expected, read.\n\nFunction: soft_reset_chk() [not invoked]\n- default_value = read_reg(SOFT_RST_REG_ADDRESS).\n- write_reg(SOFT_RST_REG_ADDRESS, SOFT_RST_REG_DATA).\n- wait_on(1000).\n- write_reg(SOFT_RST_REG_ADDRESS, default_value).\n- wait_on(1000).",
    "Hidden_Impacted_Registers": "mizar_PCIE1_SII_CFG_BAR0_START1,mizar_PCIE1_SII_CFG_BAR0_START2,mizar_PCIE1_SII_CFG_BAR0_LIMIT1,mizar_PCIE1_SII_CFG_BAR0_LIMIT2,mizar_PCIE1_SII_CFG_BAR1_START,mizar_PCIE1_SII_CFG_BAR1_LIMIT1,mizar_PCIE1_SII_CFG_BAR2_START1,mizar_PCIE1_SII_CFG_BAR2_START2,mizar_PCIE1_SII_CFG_BAR2_LIMIT1,mizar_PCIE1_SII_CFG_BAR2_LIMIT2,mizar_PCIE1_SII_CFG_BAR3_START,mizar_PCIE1_SII_CFG_BAR3_LIMIT,mizar_PCIE1_SII_CFG_BAR4_START1,mizar_PCIE1_SII_CFG_BAR4_START2,mizar_PCIE1_SII_CFG_BAR4_LIMIT1,mizar_PCIE1_SII_CFG_BAR4_LIMIT2,mizar_PCIE1_SII_CFG_BAR5_START,mizar_PCIE1_SII_CFG_BAR5_LIMIT,mizar_PCIE1_SII_PCIE1_CONFIG_INFO1,mizar_PCIE1_SII_PCIE1_CONFIG_INFO2,mizar_PCIE1_SII_PCIE1_GEN_CONTROL1,mizar_PCIE1_SII_PCIE1_GEN_CONTROL2,mizar_PCIE1_SII_PCIE1_GEN_CONTROL3,mizar_PCIE1_SII_PCIE1_PM_CONTROL,mizar_PCIE1_SII_PCIE1_CONTROL_PM_STS,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER1,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER2,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER3,mizar_PCIE1_SII_PCIE1_TRANSMIT_HEADER4,mizar_PCIE1_SII_PCIE1_TRANSMIT_REQ,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR1,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR2,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR3,mizar_PCIE1_SII_PCIE1_RCV_MSG_HDR4,mizar_PCIE1_SII_PCIE1_RCV_MSG_STS,mizar_PCIE1_SII_RCV_INTERRPUT_CTRL,mizar_PCIE1_SII_CFG_EXP_ROM_START,mizar_PCIE1_SII_CFG_EXP_ROM_LIMIT,mizar_PCIE1_SII_CFG_EXP_ROM_INFO,mizar_PCIE1_SII_CXPL_DEBUG_INFO1,mizar_PCIE1_SII_CXPL_DEBUG_INFO2,mizar_PCIE1_SII_CXPL_DEBUG_INFO_EI,mizar_PCIE1_SII_PCIE1_TARGET_INFO1,mizar_PCIE1_SII_PCIE1_TARGET_INFO2,mizar_PCIE1_SII_PCIE1_CONTOLLER_ERROR_STATUS,mizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS,mizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL,mizar_PCIE1_SII_PHY_RST_CONTROL,mizar_PCIE1_SII_LINK_DEBUG_DATA,mizar_PCIE1_SII_PCIE1_ERR_STS,mizar_PCIE1_SII_PCIE1_ERR_INTERRUPT_CTRL,mizar_PCIE1_SII_CFG_MSI_INT,mizar_PCIE1_SII_LTR_MSG,mizar_PCIE1_SII_LTR_MSG_LATENCY,mizar_PCIE1_SII_APP_LTR_LATENCY,mizar_PCIE1_SII_CFG_LTR_MAX_LATENCY,mizar_PCIE1_SII_OBFF_CNTRL,mizar_PCIE1_SII_SLV_AWMISC_INFO,mizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_HI,mizar_PCIE1_SII_SLV_AWMISC_INFO_HDR_34DW_LO,mizar_PCIE1_SII_SLV_MISC_INFO,mizar_PCIE1_SII_SLV_MISC_RESP_INFO,mizar_PCIE1_SII_MSTR_AWMISC_INFO_CNTRL,mizar_PCIE1_SII_MSTR_AWMISC_INFO_1,mizar_PCIE1_SII_MSTR_AWMISC_INFO_0,mizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_HI,mizar_PCIE1_SII_MSTR_AWMISC_INFO_HDR_34DW_LO,mizar_PCIE1_SII_MSTR_ARMISC_INFO_CNTRL,mizar_PCIE1_SII_MSTR_ARMISC_INFO_1,mizar_PCIE1_SII_MSTR_ARMISC_INFO_0,mizar_PCIE1_SII_MSTR_BMISC_RMISC_CPL_STAT_INFO,mizar_PCIE1_SII_RADM_TIMEOUT_INFO,mizar_PCIE1_SII_CFG_MSI_INFO,mizar_PCIE1_SII_CFG_MSI_DATA,mizar_PCIE1_SII_CFG_MSI_ADDR_HI,mizar_PCIE1_SII_CFG_MSI_ADDR_LO,mizar_PCIE1_SII_CFG_AER_INT_AND_PCIE1_CAP_INT_MSG,mizar_PCIE1_SII_RTLH_RFC_DATA,mizar_PCIE1_SII_APP_HDR_INFO,mizar_PCIE1_SII_APP_HDR_LOG_3,mizar_PCIE1_SII_APP_HDR_LOG_2,mizar_PCIE1_SII_APP_HDR_LOG_1,mizar_PCIE1_SII_APP_HDR_LOG_0,mizar_PCIE1_SII_CFG_BUS_NUM,mizar_PCIE1_SII_CFG_BR_CTRL_SERREN,mizar_PCIE1_SII_APP_DEV_AND_BUS_NUM,mizar_PCIE1_SII_PCIE1_CONTROLLER_INT_STS_1,mizar_PCIE1_SII_PCIE1_CONTROLLER_INTERRUPT_CONTROL_1,mizar_PCIE1_SII_APP_AND_SLOT_CONTROL_REG,mizar_PCIE1_SII_DIAG_CTRL_BUS,mizar_PCIE1_SII_CFG_REG_RO,mizar_PCIE1_SII_CFG_ARI_FWD_EN,mizar_PCIE1_SII_RADM_SLOT_PWR_PAYLOAD,mizar_PCIE1_SII_DIAG_STATUS_BUS_0,mizar_PCIE1_SII_DIAG_STATUS_BUS_1,mizar_PCIE1_SII_DIAG_STATUS_BUS_2,mizar_PCIE1_SII_DIAG_STATUS_BUS_3,mizar_PCIE1_SII_DIAG_STATUS_BUS_4,mizar_PCIE1_SII_DIAG_STATUS_BUS_5,mizar_PCIE1_SII_DIAG_STATUS_BUS_6,mizar_PCIE1_SII_DIAG_STATUS_BUS_7,mizar_PCIE1_SII_DIAG_STATUS_BUS_8,mizar_PCIE1_SII_DIAG_STATUS_BUS_9,mizar_PCIE1_SII_DIAG_STATUS_BUS_10,mizar_PCIE1_SII_DIAG_STATUS_BUS_11,mizar_PCIE1_SII_DIAG_STATUS_BUS_12,mizar_PCIE1_SII_DIAG_STATUS_BUS_13,mizar_PCIE1_SII_DIAG_STATUS_BUS_14,mizar_PCIE1_SII_DIAG_STATUS_BUS_15,mizar_PCIE1_SII_DIAG_STATUS_BUS_16,mizar_PCIE1_SII_DIAG_STATUS_BUS_17,mizar_PCIE1_SII_DIAG_STATUS_BUS_18,mizar_PCIE1_SII_DIAG_STATUS_BUS_19,mizar_PCIE1_SII_RAM_PWR_CNTRL_0,mizar_PCIE1_SII_RAM_PWR_CNTRL_1,mizar_PCIE1_SII_SOFT_RESET_CTRL,mizar_PCIE1_SII_CFG_MSI_PENDING_B,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_1,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_2,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_3,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_4,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_5,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_6,mizar_PCIE1_SII_SMLH_LTSSM_STATE_TRAN_7,mizar_PCIE1_SII_PHY_CONTROL_0,mizar_PCIE1_SII_PHY_CONTROL_1,mizar_PCIE1_SII_PHY_CONTROL_2,mizar_PCIE1_SII_PHY_CONTROL_3,mizar_PCIE1_SII_PHY_CONTROL_4,mizar_PCIE1_SII_PHY_CONTROL_5,mizar_PCIE1_SII_PHY_CONTROL_6,mizar_PCIE1_SII_PHY_CONTROL_7,mizar_PCIE1_SII_PHY_CONTROL_8,mizar_PCIE1_SII_PHY_CONTROL_9,mizar_PCIE1_SII_PHY_CONTROL_10,mizar_PCIE1_SII_PHY_CONTROL_11,mizar_PCIE1_SII_PHY_CONTROL_12,mizar_PCIE1_SII_PHY_CONTROL_13,mizar_PCIE1_SII_PHY_CONTROL_14,mizar_PCIE1_SII_PHY_CONTROL_15,mizar_PCIE1_SII_PHY_CONTROL_16,mizar_PCIE1_SII_PHY_CONTROL_17,mizar_PCIE1_SII_PHY_CONTROL_18,mizar_PCIE1_SII_PHY_CONTROL_19,mizar_PCIE1_SII_PHY_CONTROL_20,mizar_PCIE1_SII_PHY_CONTROL_21,mizar_PCIE1_SII_PHY_CONTROL_22,mizar_PCIE1_SII_MSI_CTRL_IO,mizar_PCIE1_SII_MSI_CTRL_INT_VEC",
    "Hidden_Validation_Acceptance_Criteria": "Default value check: For each i where read_mask_array[i] != 0x00000000 and addr_array[i] != mizar_PCIE1_SII_PHY_RST_CONTROL, data_rd = read_reg(addr_array[i]) must equal default_value_array[i]; otherwise def_fail_cnt++ and a failure message is printed. Write/read check: For each test pattern data_wr in {0xffffffff,0xaaaaaaaa,0x55555555,0x00000000,0xA5A5A5A5,0xffff0000}, and for each i where skip_array[i] == 0 and write_mask_array[i] != 0x00000000 and read_mask_array[i] != 0x00000000, after write_reg(addr_array[i], data_wr) the subsequent read_reg must return exp_val = ((data_wr & read_mask_array[i] & write_mask_array[i]) | ((write_mask_array[i] ^ 0xffffffff) & read_mask_array[i] & default_value_array[i])); if not, wr_fail_cnt++ and a failure message is printed. Final result: if (def_fail_cnt > 0 || wr_fail_cnt > 0) → finish(1) (FAIL); else → finish(0) (PASS).",
    "Hidden_Header_Includes": "#include <stdio.h>\n#include <stdlib.h>\n#include \"test_common.h\"\n#include \"test_define.c\"\n#include<pcie1/pcie_sii_rc_def.h>\n#include<pcie1/pcie_sii_rc_offset.h>",
    "Hidden_Macro_Defines": "#define SOFT_RST_REG_ADDRESS\t0x00000000\n#define SOFT_RST_REG_DATA\t0x00000000\n#define MIZAR_PCIE1_SII_BASE     0xE68C1000\n#define CNT 153",
    "Hidden_Skip_Array_Definition": "const int skip_array[153]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,1,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,1,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,}"
  },
  {
    "Index": 2,
    "SS / Module": "PCIE",
    "Feature": "Testable: writeAsRead",
    "Test Case Name": "pcie_cfg_wr_rd_test",
    "Test Description": "Verify PCIe configuration space write and readback after link training and coherency programming, enabling device features and confirming readiness before completion.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Requires link training to complete. Debug prints are optional. Instance selection is controlled by build-time flags.",
    "Test Steps / Procedure": "1) Clear the test control register and perform link training for the selected instance.\n2) Program the coherency control registers for enabled instances and reapply settings after a short delay.\n3) Poll the status register until the ready state is indicated; repeat for the second instance if enabled.\n4) Program memory base and perform configuration writes to BAR registers, then read back the values.\n5) Enable memory, I/O, and bus master in the command register for the enabled instance.\n6) Poll the test control register until it indicates completion, then end the test.",
    "Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "1) Status register indicates ready state → Proceed with configuration.\n2) Configuration space reads after writes return the programmed values → Pass.\n3) Test control register indicates completion → Pass.",
    "Code Generation (Required / Not)": "",
    "Hidden_Test_Case_Name": "pcie_cfg_wr_rd_test",
    "Hidden_Test_Description": "The test performs PCIe link training (variant selected by build-time flags DM0_RC, DM1_RC, DM0_EP, DM1_EP). It clears a control register at 0xE6004100, programs coherency settings via mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF and mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF using set_data for bit ranges [11:14], [3:6], [27:30], [19:22], waits, reapplies combined settings, and then polls the SII status register at offset 0xC0 via read_sii0_reg until (data_rd & 0xD1) == 0xD1 (and similarly for SII1 if DM1_RC is defined). It writes 0x11111111 to 0xE6004100, waits, and for DM0_RC executes mem_base_program_dm0_x4(), reads first 10 config dwords, writes 0xFFFFFFFF to BAR offsets 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24, reads them back, then writes 0x0, 0x4, 0x20000000, 0x40000000, 0x60000000, 0x80000000 respectively and reads back. It enables memory, I/O, and bus master by writing 0x7 at offset 0x4. For DM1_RC, analogous operations use read_pcie_slv1_reg/write_pcie_slv1_reg. Finally it polls read_reg(0xE6004100) until it equals 0x12345678, then finish(0).",
    "Hidden_Remarks": "Conditional compilation controls which instance is trained and programmed (DM0_RC, DM1_RC, DM0_EP, DM1_EP). The test uses polling loops to wait for SII status readiness ((read_siiX_reg(0xC0) & 0xD1) == 0xD1). The test control register at 0xE6004100 is first cleared, later written with 0x11111111, and finally polled until it becomes 0x12345678. DEBUG_DISPLAY gates diagnostic printf messages.",
    "Hidden_Test_Steps_Procedure": "Entry: test_case()\n- write_reg(0xE6004100, 0x0).\n- Optionally call a link training function depending on build-time flags:\n  - If DM0_RC: link_training_dm0_x4(4).\n  - If DM1_RC: link_training_dm1_x4(4).\n  - If DM0_EP: link_training_dm0_x4(4).\n  - If DM1_EP: link_training_dm1_x4(4).\n- CACHE PROGRAMMING for PCIE0:\n  - rd_wr_data1 = set_data(read_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF), 11, 14, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 3, 6, 0xf).\n  - write_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, rd_wr_data1).\n- CACHE PROGRAMMING for PCIE0 (upper fields):\n  - rd_wr_data1 = set_data(read_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF), 27, 30, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 19, 22, 0xf).\n  - write_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, rd_wr_data1).\n- CACHE PROGRAMMING for PCIE1:\n  - rd_wr_data1 = set_data(read_reg(mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF), 11, 14, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 3, 6, 0xf).\n  - write_reg(mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, rd_wr_data1).\n- CACHE PROGRAMMING for PCIE1 (upper fields):\n  - rd_wr_data1 = set_data(read_reg(mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF), 27, 30, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 19, 22, 0xf).\n  - write_reg(mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, rd_wr_data1).\n- wait_on(20).\n- Reapply combined settings for PCIE0:\n  - rd_wr_data1 = set_data(read_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF), 11, 14, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 3, 6, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 27, 30, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 19, 22, 0xf).\n  - write_reg(mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF, rd_wr_data1).\n- Reapply combined settings for PCIE1:\n  - rd_wr_data1 = set_data(read_reg(mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF), 11, 14, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 3, 6, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 27, 30, 0xf).\n  - rd_wr_data1 = set_data(rd_wr_data1, 19, 22, 0xf).\n  - write_reg(mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF, rd_wr_data1).\n- data_rd = read_sii0_reg(0xC0); while ((data_rd & 0xD1) != 0xD1) { data_rd = read_sii0_reg(0xC0); }.\n- If DM1_RC: data_rd = read_sii1_reg(0xC0); while ((data_rd & 0xD1) != 0xD1) { data_rd = read_sii1_reg(0xC0); }.\n- write_reg(0xE6004100, 0x11111111).\n- wait_on(15000).\n- If DM0_RC:\n  - mem_base_program_dm0_x4(); wait_on(10).\n  - For (i = 0; i < 10; i++): rd_wr_data1 = read_pcie_slv0_reg(i * 0x4).\n  - Write 0xFFFFFFFF to BAR registers: write_pcie_slv0_reg(0x10, 0xFFFFFFFF); 0x14 → 0xFFFFFFFF; 0x18 → 0xFFFFFFFF; 0x1c → 0xFFFFFFFF; 0x20 → 0xFFFFFFFF; 0x24 → 0xFFFFFFFF.\n  - Read back BAR registers: read_pcie_slv0_reg(0x10), 0x14, 0x18, 0x1c, 0x20, 0x24.\n  - Write specific values: 0x10 → 0x0; 0x14 → 0x4; 0x18 → 0x20000000; 0x1c → 0x40000000; 0x20 → 0x60000000; 0x24 → 0x80000000.\n  - Read back BAR registers: read_pcie_slv0_reg(0x10), 0x14, 0x18, 0x1c, 0x20, 0x24.\n  - Enable MSE/IO/BME: write_pcie_slv0_reg(0x4, 0x7).\n- If DM1_RC:\n  - mem_base_program_dm1_x4().\n  - For (i = 0; i < 10; i++): rd_wr_data1 = read_pcie_slv1_reg(i * 0x4).\n  - Enable MSE/IO/BME: write_pcie_slv1_reg(0x4, 0x7).\n  - BAR write 0xFFFFFFFF to 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24.\n  - Readback BARs at 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24.\n  - Program BARs: 0x10 → 0x0; 0x14 → 0x4; 0x18 → 0x20000000; 0x1c → 0x40000000; 0x20 → 0x60000000; 0x24 → 0x80000000.\n  - Readback BARs at 0x10, 0x14, 0x18, 0x1c, 0x20, 0x24.\n- wait_on(10); data_rd = read_reg(0xE6004100); while (data_rd != 0x12345678) { wait_on(5); data_rd = read_reg(0xE6004100); }.\n- finish(0).",
    "Hidden_Impacted_Registers": "mizar_PCIE0_DBI_DSP_COHERENCY_CONTROL_3_OFF,mizar_PCIE1_DBI_DSP_COHERENCY_CONTROL_3_OFF",
    "Hidden_Validation_Acceptance_Criteria": "SII0 readiness: Loop on data_rd = read_sii0_reg(0xC0) until ((data_rd & 0xD1) == 0xD1). If DM1_RC, also require SII1 readiness via read_sii1_reg(0xC0) with the same mask/compare. Final completion: Poll read_reg(0xE6004100) until it equals 0x12345678; then finish(0). Configuration BAR writes are performed and read back; the code prints values but does not compare them against explicit expected values, so no failure is raised based on those reads within this test.",
    "Hidden_Header_Includes": "#include <stdlib.h>\n#include <stdio.h>\n#include <test_common.h>\n#include \"pcie.h\"",
    "Hidden_Macro_Defines": "NA",
    "Hidden_Skip_Array_Definition": "NA"
  }
]'''

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

# Explicit meta names from spec
META_EXPLICIT = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
    "Hidden_Header_Includes",
    "Hidden_Macro_Define",
    "Hidden_Skip_Array_Definition",
]

# Any key beginning with Hidden_ should also be treated as META to preserve and remove from main

def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ip-name", required=True)
    ap.add_argument("--output-dir", required=True)
    return ap.parse_args()


def validate_and_load_json():
    try:
        data = json.loads(EMBEDDED_JSON)
    except Exception as e:
        raise SystemExit(f"Invalid JSON input: {e}")
    if not isinstance(data, list) or len(data) == 0:
        raise SystemExit("Invalid or empty JSON array")
    for i, rec in enumerate(data, 1):
        if not isinstance(rec, dict):
            raise SystemExit(f"Record {i} is not an object")
    return data


def union_keys_preserve_order(records):
    seen = []
    sset = set()
    for rec in records:
        for k in rec.keys():
            if k not in sset:
                sset.add(k)
                seen.append(k)
    return seen


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def numerically_number(cell_value):
    if not isinstance(cell_value, str) or not cell_value.strip():
        return cell_value
    # Split by newlines and rebuild with 1., 2., 3.
    lines = [ln.strip() for ln in cell_value.replace("\r", "").split("\n") if ln.strip()]
    if not lines:
        return ""
    renum = []
    for idx, ln in enumerate(lines, 1):
        # Strip any existing numeric/bullet prefix
        lns = ln
        # Remove common leading patterns like '1)', '1.', '-', '•'
        while True:
            stripped = False
            for pref in [")", ".", "-", "•", ":"]:
                if lns and lns[0].isdigit():
                    # remove digits and immediate ) or .
                    j = 0
                    while j < len(lns) and lns[j].isdigit():
                        j += 1
                    if j < len(lns) and lns[j] in (")", "."):
                        lns = lns[j+1:].lstrip()
                        stripped = True
                        break
                if lns.startswith(pref):
                    lns = lns[len(pref):].lstrip()
                    stripped = True
                    break
            if not stripped:
                break
        renum.append(f"{idx}. {lns}")
    return "\n".join(renum)


def build_workbook(records, ip_name, out_dir):
    # Determine schema
    all_keys = union_keys_preserve_order(records)

    # Create workbook with single sheet 'Data'
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header style
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="BDD7EE")  # light blue for readability

    # Data style defaults
    text_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    idx_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write headers
    ws.append(all_keys)
    for c, _ in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = border_thin

    # Write rows preserving exact values
    for rec in records:
        row = [rec.get(k, "") for k in all_keys]
        ws.append(row)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Apply base borders and alignment for all data cells
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.border = border_thin
            # Index column heuristic: match exact header name
            if ws.cell(row=1, column=cell.column).value == "Index":
                cell.alignment = idx_align
            else:
                cell.alignment = text_align

    # Create Meta_data_sheet and copy META columns (explicit list + all Hidden_*)
    meta_ws = wb.create_sheet("Meta_data_sheet")
    meta_keys = []
    # First, any keys that start with Hidden_
    for k in all_keys:
        if isinstance(k, str) and k.startswith("Hidden_") and k not in meta_keys:
            meta_keys.append(k)
    # Ensure all explicit names are included (even if not present in records)
    for k in META_EXPLICIT:
        if k not in meta_keys:
            meta_keys.append(k)

    # Write meta headers
    meta_ws.append(meta_keys)
    for c, _ in enumerate(meta_keys, 1):
        cell = meta_ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = border_thin

    # Write meta rows from records, preserving exact existing values where key exists
    for rec in records:
        row = [rec.get(k, "") for k in meta_keys]
        meta_ws.append(row)
        # Apply borders to meta rows
        for c in range(1, len(meta_keys) + 1):
            meta_ws.cell(row=meta_ws.max_row, column=c).border = border_thin

    # Very hidden meta sheet
    meta_ws.sheet_state = 'veryHidden'

    # Rename 'Data' to 'TestPlan' and normalize columns there
    ws.title = "TestPlan"

    # Build final visible column order based on MAIN_ORDER; include only those present,
    # but also add any remaining non-hidden, non-meta keys at the end to preserve schema if present in input
    non_hidden_keys = [k for k in all_keys if not (isinstance(k, str) and k.startswith("Hidden_"))]
    final_order = [k for k in MAIN_ORDER if k in non_hidden_keys]
    # Append any additional non-hidden keys not in MAIN_ORDER (schema preservation)
    for k in non_hidden_keys:
        if k not in final_order:
            final_order.append(k)

    # Extract data rows from current TestPlan
    rows = []
    headers = [cell.value for cell in ws[1]]
    key_index = {k: i for i, k in enumerate(headers)}
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rec = {headers[i]: ("" if r[i] is None else r[i]) for i in range(len(headers))}
        rows.append(rec)

    # Clear sheet and write normalized headers
    ws.delete_rows(1, ws.max_row)
    ws.append(final_order)
    for c, _ in enumerate(final_order, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = border_thin

    # Write normalized rows with numbering rules applied
    for rec in rows:
        out_row = []
        for k in final_order:
            v = rec.get(k, "")
            if k in ("Test Steps / Procedure", "Validation / Acceptance Criteria"):
                v = numerically_number(v)
            out_row.append(v)
        ws.append(out_row)

    # Freeze top row again
    ws.freeze_panes = "A2"

    # Apply borders and alignment
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.border = border_thin
            if ws.cell(row=1, column=cell.column).value == "Index":
                cell.alignment = idx_align
            else:
                cell.alignment = text_align

    # Data validation for 'Code Generation (Required / Not)'
    if "Code Generation (Required / Not)" in final_order and ws.max_row >= 2:
        col_idx = final_order.index("Code Generation (Required / Not)") + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True)
        dv.error = "Select a value from the list"
        dv.errorTitle = "Invalid Input"
        ws.add_data_validation(dv)
        dv_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        dv.add(dv_range)

    # Auto-fit columns roughly based on max text length
    def str_len(v):
        if v is None:
            return 0
        s = str(v)
        # account for newlines: take longest line
        return max((len(p) for p in s.split("\n")), default=0)

    for idx, k in enumerate(final_order, 1):
        max_len = str_len(k)
        for r in range(2, ws.max_row + 1):
            max_len = max(max_len, str_len(ws.cell(row=r, column=idx).value))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(120, max(10, int(max_len * 1.2) + 2))

    # Approximate row heights based on number of lines
    for r in range(2, ws.max_row + 1):
        max_lines = 1
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                max_lines = max(max_lines, len(v.split("\n")))
        ws.row_dimensions[r].height = 15 * max_lines

    # Mandatory safety check: ensure no sheet named 'Data' remains
    for sh in list(wb.worksheets):
        if sh.title == "Data":
            wb.remove(sh)

    # Save using IST timestamp and validate as true XLSX
    ist = ZoneInfo("Asia/Kolkata")
    now_ist = datetime.now(ist)
    fname = f"{ip_name}_TestPlan_{now_ist.strftime('%Y%m%d')}_{now_ist.strftime('%H%M%S')}.xlsx"
    ensure_dir(out_dir)
    out_path = os.path.join(out_dir, fname)
    wb.save(out_path)

    # Validate OOXML zip structure
    with ZipFile(out_path, 'r') as zf:
        namelist = set(zf.namelist())
        required = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
        if not required.issubset(namelist):
            raise SystemExit("XLSX validation failed: Missing OOXML parts")

    print(out_path)
    return out_path


def main():
    args = parse_args()
    records = validate_and_load_json()
    build_workbook(records, args.ip_name, args.output_dir)

if __name__ == "__main__":
    main()
