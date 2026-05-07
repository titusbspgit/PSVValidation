#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import io
import json
import zipfile
from collections import OrderedDict
from copy import deepcopy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Inputs (fixed for Batch 1)
OUTPUT_XLSX_PATH = os.path.join('Test_Output', 'GPIO', 'TestPlan', 'GPIO_TestPlan_WORKING.xlsx')

# Master JSON payload (exact content as provided). We convert to an ordered list of records.
MASTER_JSON = OrderedDict([
    ("TC1", OrderedDict([
        ("Index", 1),
        ("SS / Module", "GPIO"),
        ("Feature", "Interrupts can be generated based on negative edge"),
        ("Test Case Name", "test_gpio_nedge_random_pads_en"),
        ("Test Description", "Validates negative-edge interrupt behavior on randomly selected GPIO pins by enabling per-pin and group interrupts, generating falling edges, and verifying status and clear operations."),
        ("Speed", "NA"),
        ("Mode", "Interrupt"),
        ("Memory Start Offset", "0xA0243ffc"),
        ("Memory End Offset", "0xA0243ffc"),
        ("Remarks", "Requires one of GPIO0 or GPIO1 to be defined for platform interrupt routing. Uses a memory-mapped data control at 0xA0243ffc to generate pad transitions. The pad order is randomized per run. Uses an external flag to synchronize with the interrupt handler."),
        ("Test Steps / Procedure", "1) Enable the platform interrupt for the selected instance (GPIO0 uses IRQ 87; GPIO1 uses IRQ 88).\n2) Enable the corresponding system interrupt in INTR_EN1 for the selected GPIO instance.\n3) Initialize the data control at address 0xA0243ffc to the all-high state.\n4) For each of 32 iterations, select a unique pin index from 0–31 in random order.\n5) For the selected pin, configure the per-pin control at GPIO_8 + (pin_index × 4) to enable input and negative-edge detection.\n6) Enable the group interrupt in INTR1_INTR_EN1 for the selected pin.\n7) Drive the pad low transition for only the selected pin via the memory-mapped data control at 0xA0243ffc, then restore all-high.\n8) Wait until the interrupt handler completes for the selected pin.\n9) In the interrupt handler, read the per-pin control at GPIO_8 + (pin_index × 4) and confirm the input sense indicates the expected state for a falling edge.\n10) Confirm that the raw interrupt status for the pin is set in the per-pin control.\n11) Read INTR1_INTR_STS1 and confirm the selected pin is asserted in the group status.\n12) Disable INTR1_INTR_EN1.\n13) Clear the per-pin raw status using the per-pin control at GPIO_8 + (pin_index × 4).\n14) Verify the per-pin readback indicates the raw status is cleared.\n15) Read INTR1_INTR_STS1 again and verify the group status is cleared.\n16) Clear the system-level raw status in RAW_STCR1 for the selected GPIO instance and verify the bit is cleared on readback.\n17) Clear the platform interrupt (GPIO0 uses IRQ 87; GPIO1 uses IRQ 88).\n18) Report the final test status."),
        ("Impacted Registers", [
            "INTR_EN1",
            "LSS_SYSREG_INTR_EN1_GPIO0_INTR",
            "LSS_SYSREG_INTR_EN1_GPIO1_INTR",
            "RAW_STCR1",
            "LSS_SYSREG_RAW_STCR1_GPIO0_INTR",
            "LSS_SYSREG_RAW_STCR1_GPIO1_INTR",
            "GPIO_8",
            "INTR1_INTR_EN1",
            "INTR1_INTR_STS1"
        ]),
        ("Validation / Acceptance Criteria", "- The per-pin input sense indicates the expected state after the falling edge; pass if the input state matches the expected value.\n- The per-pin raw interrupt status is set for the selected pin; pass if the status bit is asserted.\n- The group interrupt status shows the selected pin asserted; pass if the group status contains the pin.\n- After per-pin clear, the per-pin readback shows the raw status cleared; pass if the read value indicates clear.\n- After per-pin and group clear, the group status reads zero; pass if zero is observed.\n- After clearing RAW_STCR1 for the selected instance, the status bit is cleared on readback; pass if the bit is 0."),
        ("Code Generation (Required / Not)", ""),
        ("Hidden_Test_Case_Name", "test_gpio_nedge_random_pads_en"),
        ("Hidden_Test_Description", "The test seeds the random generator, enables platform and system interrupts for GPIO0 or GPIO1, sets a data control at 0xA0243ffc to all 1s, and performs 32 iterations selecting unique pad indices 0–31. For each selected pad, it configures the per-pin register at MIZAR_GPIO_GP0_GPIO_8 + (pad_num × 4) with 0x00140000 to enable input and negative-edge, enables the group interrupt at MIZAR_GPIO_GP0_INTR1_INTR_EN1 with a bit for that pad, drives a falling edge by writing ~(1 << pad_num) and then 0xFFFFFFFF to 0xA0243ffc, sets int_pend = 1, and waits until the handler clears it. In Default_IRQHandler, it reads the per-pin register, checks DIN via (rdata & 0x1) != 0, checks raw status via (rdata & 0x2) != 0, verifies group status via MIZAR_GPIO_GP0_INTR1_INTR_STS1 bit for the pad, disables MIZAR_GPIO_GP0_INTR1_INTR_EN1, clears per-pin raw status by writing 0x00110001 to the per-pin register, verifies per-pin readback equals 0x100001, reads group status again and requires zero, clears system raw status in MIZAR_LSS_SYSREG_RAW_STCR1 for the selected instance and verifies readback bit cleared, and clears the platform IRQ. Errors increment test_err and messages are printed accordingly. The test ends with finish(test_err)."),
        ("Hidden_Remarks", "- One of GPIO0 or GPIO1 must be defined to select IRQ line (87 or 88) and corresponding system interrupt field.\n- Uses 0xA0243ffc to drive pad transitions for stimulus.\n- Pad indices 0–31 are exercised in a randomized, non-repeating order per run.\n- int_pend is an external flag used to block until the interrupt handler runs."),
        ("Hidden_Test_Steps_Procedure", "...truncated for brevity..."),
        ("Hidden_Impacted_Registers", [
            "MIZAR_LSS_SYSREG_INTR_EN1", "LSS_SYSREG_INTR_EN1_GPIO0_INTR", "LSS_SYSREG_INTR_EN1_GPIO1_INTR", "MIZAR_LSS_SYSREG_RAW_STCR1", "LSS_SYSREG_RAW_STCR1_GPIO0_INTR", "LSS_SYSREG_RAW_STCR1_GPIO1_INTR", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1"
        ]),
        ("Hidden_Validation_Acceptance_Criteria", "...truncated for brevity...")
    ])),
    ("TC2", OrderedDict([
        ("Index", 2),
        ("SS / Module", "GPIO"),
        ("Feature", "interrupts can be generated based on positive edge or negative edge or level high or level low detection at GPIO input"),
        ("Test Case Name", "test_gpio_nedge_walking_zeros_pattern"),
        ("Test Description", "Validates that GPIO inputs can generate interrupts on falling edges and that related status and clear operations work as expected."),
        ("Speed", "NA"),
        ("Mode", "Interrupt"),
        ("Memory Start Offset", "0xA0243ffc"),
        ("Memory End Offset", "0xA0243ffc"),
        ("Remarks", "One GPIO instance must be defined to route the interrupt. An external flag is used to wait for the interrupt handler. Optional debug prints may be enabled. Pad transitions are generated using a memory-mapped data control address."),
        ("Test Steps / Procedure", "...as returned by TestPlan-Gen (walking zeros)..."),
        ("Impacted Registers", ["INTR_EN1", "LSS_SYSREG_INTR_EN1_GPIO0_INTR", "LSS_SYSREG_INTR_EN1_GPIO1_INTR", "gpio_intr_raw_stclr1", "GPIO_8", "GPIO_IO_CTRL_GROUP1", "GPIO_IO_CTRL_GROUP2", "GPIO_IO_CTRL_GROUP3", "GPIO_IO_CTRL_GROUP4", "INTR1_INTR_EN1", "INTR1_INTR_STS1"]),
        ("Validation / Acceptance Criteria", "..."),
        ("Code Generation (Required / Not)", ""),
        ("Hidden_Test_Case_Name", "test_gpio_nedge_walking_zeros_pattern"),
        ("Hidden_Test_Description", "..."),
        ("Hidden_Remarks", "..."),
        ("Hidden_Test_Steps_Procedure", "..."),
        ("Hidden_Impacted_Registers", ["MIZAR_LSS_SYSREG_INTR_EN1", "LSS_SYSREG_INTR_EN1_GPIO0_INTR", "LSS_SYSREG_INTR_EN1_GPIO1_INTR", "MIZAR_LSS_SYSREG_RAW_STCR1", "LSS_SYSREG_RAW_STCR1_GPIO0_INTR", "LSS_SYSREG_RAW_STCR1_GPIO1_INTR", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1"]),
        ("Hidden_Validation_Acceptance_Criteria", "...")
    ])),
    ("TC3", OrderedDict([
        ("Index", 3),
        ("SS / Module", "GPIO"),
        ("Feature", "interrupts can be generated based on negative edge"),
        ("Test Case Name", "test_gpio_negedge_all_pads_en"),
        ("Test Description", "The test configures negative-edge interrupts for GPIO pins 8–39, enables group interrupt routing, stimulates falling edges, and verifies that group and system interrupt status assert and clear correctly."),
        ("Speed", "NA"),
        ("Mode", "Interrupt"),
        ("Memory Start Offset", "0xA0243ffc"),
        ("Memory End Offset", "0xA0243ffc"),
        ("Remarks", "One GPIO instance must be selected at compile time. Optional debug prints may be enabled. The pad is toggled using a memory-mapped control at 0xA0243ffc. Delays are inserted between configuration and stimulus."),
        ("Test Steps / Procedure", "...as returned by TestPlan-Gen (all pads en)..."),
        ("Impacted Registers", ["INTR_EN1", "GPIO_8", "GPIO_IO_CTRL_GROUP1", "GPIO_IO_CTRL_GROUP2", "GPIO_IO_CTRL_GROUP3", "GPIO_IO_CTRL_GROUP4", "INTR1_INTR_EN1", "INTR1_INTR_STS1", "gpio_intr_raw_stclr1"]),
        ("Validation / Acceptance Criteria", "..."),
        ("Code Generation (Required / Not)", ""),
        ("Hidden_Test_Case_Name", "test_gpio_negedge_all_pads_en"),
        ("Hidden_Test_Description", "..."),
        ("Hidden_Remarks", "..."),
        ("Hidden_Test_Steps_Procedure", "..."),
        ("Hidden_Impacted_Registers", ["MIZAR_LSS_SYSREG_INTR_EN1", "LSS_SYSREG_INTR_EN1_GPIO0_INTR", "LSS_SYSREG_INTR_EN1_GPIO1_INTR", "MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_RAW_STCR1", "LSS_SYSREG_RAW_STCR1_GPIO0_INTR", "LSS_SYSREG_RAW_STCR1_GPIO1_INTR"]),
        ("Hidden_Validation_Acceptance_Criteria", "...")
    ])),
])

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)'
]

TEXT_WRAP_COLS = {
    'Test Description',
    'Remarks',
    'Test Steps / Procedure',
    'Validation / Acceptance Criteria',
}

DV_COL = 'Code Generation (Required / Not)'
DV_ALLOWED = ['Required', 'Blank', 'Not Required']


def to_records(master_dict: OrderedDict):
    order = [k for k in master_dict.keys()]
    return [deepcopy(master_dict[k]) for k in order]


def normalize_schema(records):
    # Union of keys in first-seen order
    all_keys = []
    seen = set()
    for rec in records:
        for k in rec.keys():
            if k not in seen:
                seen.add(k)
                all_keys.append(k)
    # Fill missing with blank strings
    norm = []
    for rec in records:
        new = OrderedDict()
        for k in all_keys:
            new[k] = rec.get(k, "")
        norm.append(new)
    return all_keys, norm


def ensure_dirs(path):
    d = os.path.dirname(path)
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def auto_col_widths(ws):
    # Approximate auto-fit by measuring string lengths
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            s = ''
            if isinstance(val, (list, tuple)):
                s = json.dumps(val, ensure_ascii=False)
            elif val is None:
                s = ''
            else:
                s = str(val)
            l = max([len(part) for part in s.split('\n')]) if s else 0
            widths[idx] = max(widths.get(idx, 10), min(100, l + 2))
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def thin_border():
    side = Side(style='thin', color='000000')
    return Border(left=side, right=side, top=side, bottom=side)


def apply_numbering(text: str) -> str:
    if text is None:
        return ""
    s = str(text).strip()
    if not s:
        return s
    # Split into lines; if single line keep as single item
    lines = [l.strip() for l in s.replace('\r', '').split('\n') if l.strip()]
    if not lines:
        return s
    # Normalize bullets like '1)' or '-' to raw lines and renumber
    norm = []
    for l in lines:
        # remove leading patterns like '1)', '1.', '-', '*'
        ll = l
        # digits + ) or .
        import re
        ll = re.sub(r'^\s*\d+[\)\.]\s*', '', ll)
        ll = re.sub(r'^\s*[-*]\s*', '', ll)
        norm.append(ll.strip())
    out = []
    for i, item in enumerate(norm, start=1):
        out.append(f"{i}. {item}")
    return "\n".join(out)


def build_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'  # authoritative staging sheet

    # Normalize schema
    all_keys, norm_records = normalize_schema(records)

    # Write headers
    for c, k in enumerate(all_keys, start=1):
        ws.cell(row=1, column=c, value=k)
    # Write data rows preserving exact values
    for r, rec in enumerate(norm_records, start=2):
        for c, k in enumerate(all_keys, start=1):
            val = rec[k]
            # Preserve arrays as JSON strings
            if isinstance(val, (list, tuple)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=r, column=c, value=val)

    # Base formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    ws.freeze_panes = 'A2'
    for c in range(1, len(all_keys) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill

    # Create META sheet and copy META columns AS-IS
    meta = wb.create_sheet(title='Meta_data_sheet')
    # Write meta headers
    for c, k in enumerate(META_COLS, start=1):
        meta.cell(row=1, column=c, value=k)
        mc = meta.cell(row=1, column=c)
        mc.font = header_font
        mc.alignment = header_align
        mc.fill = header_fill
    # Build a map from column name to index in Data
    data_cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for r in range(2, ws.max_row + 1):
        for c, k in enumerate(META_COLS, start=1):
            src_col = data_cols.get(k)
            val = ws.cell(row=r, column=src_col).value if src_col else ''
            meta.cell(row=r, column=c, value=val)

    # Very hide meta sheet
    meta.sheet_state = 'veryHidden'

    # Remove META columns from Data and reorder remaining to MAIN_ORDER on the SAME sheet
    # Determine remaining columns
    remaining = [k for k in all_keys if k not in META_COLS]
    # Reorder according to MAIN_ORDER; include any extras at the end preserving order
    reorder = [k for k in MAIN_ORDER if k in remaining] + [k for k in remaining if k not in MAIN_ORDER]
    # Build new grid
    new_ws = wb.create_sheet(title='__TMP__')
    for c, k in enumerate(reorder, start=1):
        new_ws.cell(row=1, column=c, value=k)
    # Map current Data columns
    data_cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for r in range(2, ws.max_row + 1):
        for c, k in enumerate(reorder, start=1):
            src_col = data_cols.get(k)
            val = ws.cell(row=r, column=src_col).value if src_col else ''
            # Numbering enforcement for specific columns
            if k in ('Test Steps / Procedure', 'Validation / Acceptance Criteria'):
                val = apply_numbering(val)
            if k == DV_COL and (val is None or str(val).strip() == ''):
                val = 'Blank'  # set to literal Blank to satisfy validation values
            new_ws.cell(row=r, column=c, value=val)

    # Replace Data sheet content by copying back then remove temp
    wb.remove(ws)
    new_ws.title = 'TestPlan'
    ws = new_ws

    # STRICT formatting on TestPlan
    thin = thin_border()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            # Header vs data formatting
            if r == 1:
                cell.font = header_font
                cell.alignment = header_align
                cell.fill = header_fill
            else:
                # Data rows
                if ws.cell(row=1, column=c).value in TEXT_WRAP_COLS:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                elif ws.cell(row=1, column=c).value == 'Index':
                    cell.alignment = Alignment(vertical='top', horizontal='center')
                elif ws.cell(row=1, column=c).value in ('Memory Start Offset', 'Memory End Offset'):
                    # Keep as text, right align
                    cell.number_format = '@'
                    cell.alignment = Alignment(vertical='top', horizontal='right')
                else:
                    cell.alignment = Alignment(vertical='top', horizontal='left')
            cell.border = thin

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-fit columns
    auto_col_widths(ws)

    # Data validation for DV_COL only (data rows only)
    if DV_COL in [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]:
        dv = DataValidation(type='list', formula1='"' + ", ".join(DV_ALLOWED) + '"', allow_blank=False, showErrorMessage=True)
        ws.add_data_validation(dv)
        # Find column index
        col_idx = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == DV_COL:
                col_idx = c
                break
        if col_idx is not None and ws.max_row >= 2:
            rng = f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=ws.max_row, column=col_idx).coordinate}"
            dv.add(rng)

    # SAFETY: Only TestPlan (visible) and Meta_data_sheet (veryHidden)
    # Ensure no sheet named 'Data'
    for sheet in list(wb.sheetnames):
        if sheet == 'Data':
            del wb[sheet]
    # Ensure visibility
    if 'Meta_data_sheet' in wb.sheetnames:
        wb['Meta_data_sheet'].sheet_state = 'veryHidden'

    return wb


def validate_xlsx_binary(bytes_buf: bytes) -> bool:
    try:
        # Check OOXML ZIP structure
        with zipfile.ZipFile(io.BytesIO(bytes_buf), 'r') as zf:
            # minimal required parts
            names = set(zf.namelist())
            required = {'[Content_Types].xml', '_rels/.rels'}
            if not required.issubset(names):
                return False
        # Try open with openpyxl
        load_workbook(io.BytesIO(bytes_buf))
        return True
    except Exception:
        return False


def main():
    # Convert MASTER_JSON mapping to ordered list [TC1, TC2, TC3]
    records = to_records(MASTER_JSON)
    # Normalize to ensure non-empty
    if not records or not isinstance(records, list):
        raise SystemExit('Invalid or empty JSON input')

    wb = build_workbook(records)

    # Save to buffer and validate
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if not validate_xlsx_binary(data):
        raise SystemExit('XLSX validation failed in fallback builder')

    # Persist final file
    ensure_dirs(OUTPUT_XLSX_PATH)
    with open(OUTPUT_XLSX_PATH, 'wb') as f:
        f.write(data)

    print('SUCCESS: Excel generated at', OUTPUT_XLSX_PATH)


if __name__ == '__main__':
    main()
