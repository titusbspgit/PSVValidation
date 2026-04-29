#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stage1 JSON→Excel generator for I2C TestPlan
- Reads embedded JSON payload (from this script) and/or env overrides
- Generates base workbook (Data)
- Creates Meta_data_sheet with META columns and set Very Hidden
- Renames Data→TestPlan and removes META columns leaving only approved MAIN columns (in exact order)
- Applies strict formatting and data validation to TestPlan
- Saves the workbook to Test_Output/I2C/TestPlan/I2C_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx using IST time
Environment overrides (optional):
- IST_TS: timestamp for filename (YYYYMMDD_HHMMSS)
- IST_HUMAN: human-friendly timestamp (YYYY-MM-DD HH:MM:SS)
"""
import os
import json
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    TZ = ZoneInfo("Asia/Kolkata")
except Exception:
    TZ = None

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# -------------------- JSON payload (embedded) --------------------
JSON_PAYLOAD = r'''{
  "IP": "I2C",
  "TestCases": [
    {
      "Index": 1,
      "SS / Module": "I2C",
      "Feature": "AHB interface to access its register space",
      "Test Case Name": "i2c_reg_rd_wr_test",
      "Test Description": "Verifies reset defaults and masked read/write behavior across a defined list of I2C registers. For each readable register, compares the current value with its default. For each writable-and-readable register, writes test patterns and validates readback against expected values derived from read/write masks and defaults. Aggregates failures and reports overall pass/fail.",
      "Speed": "NA",
      "Mode": "NA",
      "Memory Start Offset": "NA",
      "Memory End Offset": "NA",
      "Remarks": "Addresses marked non-readable are skipped when checking defaults. Addresses marked non-writable are skipped for write/read tests. Entries flagged in the skip array are excluded. Soft reset routine exists but is not executed in this test.",
      "Test Steps / Procedure": [
        "Entry: test_case()",
        "Invoke default value check: call chk_rst_val()",
        "Loop i=0..(CNT-1): set addr = addr_array[i]",
        "If read_mask_array[i] == 0x00000000 then skip default read for this addr",
        "Else READ register at addr; store to data_rd",
        "Compare data_rd to default_value_array[i]; if mismatch, increment def_fail_cnt and log failure",
        "Invoke write/read verification: call chk_rd_wr()",
        "Define chk_val[6] = {0xFFFFFFFF, 0xAAAAAAAA, 0x55555555, 0x00000000, 0xA5A5A5A5, 0xFFFF0000}",
        "Outer loop j=0..5: set data_wr = chk_val[j]",
        "Inner loop (write phase) i=0..(CNT-1): set addr = addr_array[i]",
        "If skip_array[i] == 1 then continue to next i (write skipped)",
        "If write_mask_array[i] == 0x00000000 then continue to next i (not writable)",
        "Else WRITE data_wr to register at addr",
        "Inner loop (read/compare phase) i=0..(CNT-1): set addr = addr_array[i]",
        "If skip_array[i] == 1 then continue to next i (read skipped)",
        "If write_mask_array[i] == 0x00000000 then continue to next i (write was skipped → read skipped)",
        "If read_mask_array[i] == 0x00000000 then continue to next i (not readable)",
        "Else READ register at addr; store to data_rd",
        "Compute wr_n = bitwise NOT of write_mask_array[i]",
        "Compute exp_val = (data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])",
        "If data_rd != exp_val then increment wr_fail_cnt and log failure",
        "Return from chk_rd_wr()",
        "Evaluate overall status: if (def_fail_cnt > 0 || wr_fail_cnt > 0) then finish(1) else finish(0)"
      ],
      "Impacted Registers": "MIZAR_I2C_DEV_CTRL,MIZAR_I2C_TSFR_CTRL,MIZAR_I2C_SLV_ADDR,MIZAR_I2C_TGT_SLV_ADDR,MIZAR_I2C_I2C_MSTR_CODE,MIZAR_I2C_I2C_BYTE_CNT,MIZAR_I2C_SF_HCNT,MIZAR_I2C_SF_LCNT,MIZAR_I2C_I2C_HS_HCNT,MIZAR_I2C_I2C_HS_LCNT,MIZAR_I2C_RIS,MIZAR_I2C_MASK_INTR,MIZAR_I2C_INTR_STS,MIZAR_I2C_INTR_CLR,MIZAR_I2C_TAS,MIZAR_I2C_TX_FIFO_THLD,MIZAR_I2C_RX_FIFO_THLD,MIZAR_I2C_DMA_CTRL,MIZAR_I2C_FF,MIZAR_I2C_TX_FIFO_LVL,MIZAR_I2C_RX_FIFO_LVL,MIZAR_I2C_I2C_MSTR_STS,MIZAR_I2C_I2C_FLTR_SEL,MIZAR_I2C_I2C_CURRENT_BYTECNT,MIZAR_I2C_I2C_SMB_SFTRST,MIZAR_I2C_SMB_HST_STS,MIZAR_I2C_SMB_HST_CTRL,MIZAR_I2C_SMB_HST_CMD,MIZAR_I2C_SMB_HST_DATA0,MIZAR_I2C_SMB_HST_DATA1,MIZAR_I2C_SMB_HST_BLOCK_DATA,MIZAR_I2C_SMB_PEC_DATA,MIZAR_I2C_SMB_SLAVE_WDATA,MIZAR_I2C_SMB_SLAVE_CMD,MIZAR_I2C_SMB_SLAVE_CTS,MIZAR_I2C_SMB_SLV,MIZAR_I2C_SMB_NOTIFY_ADDR,MIZAR_I2C_SMB_NOTIFY_LOW_BYTE,MIZAR_I2C_SMB_NOTIFY_HIGH_BYTE,MIZAR_I2C_SMB_DATA_HLDTIME,MIZAR_I2C_SMB_TIMEOUT_CNT,MIZAR_I2C_SMB_TMEXT_CNT,MIZAR_I2C_I2CSMB_DATA_SETUP",
      "Hidden_Test_Case_Name": "i2c_reg_rd_wr_test",
      "ArtifactsPath": "TestRepo/i2c/i2c_reg_rd_wr_test"
    },
    {
      "Index": 2,
      "SS / Module": "I2C",
      "Feature": "Three speeds in I2C mode: High-speed (3.4 Mb/s)",
      "Test Case Name": "test_i2c0_mst_i2c1_slv_dma_hs_md",
      "Test Description": "Configures two I2C instances for a DMA-assisted transfer from controller 0 to controller 1, triggers a transfer, waits for completion via status and interrupt, then performs a DMA move from the receiver to memory and verifies data integrity. Uses high-speed timing configuration on the transmitter.",
      "Speed": "High-speed",
      "Mode": "DMA + Interrupt + Polling",
      "Remarks": "Initial writes configure a NIC interface at fixed addresses before I2C setup. Interrupts are enabled through the system register interface and serviced by a default handler that clears the event and validates status. Data buffers are preloaded in SRAM and moved using DMA channels.",
      "Test Steps / Procedure": ["... see description above ..."],
      "Impacted Registers": "MIZAR_LSS_SYSREG_INTR_EN0,MIZAR_I2C0_FF,MIZAR_I2C1_FF,MIZAR_I2C0_INTR_CLR,MIZAR_I2C1_INTR_CLR,MIZAR_I2C0_DEV_CTRL,MIZAR_I2C1_DEV_CTRL,MIZAR_I2C1_SLV_ADDR,MIZAR_I2C0_TGT_SLV_ADDR,MIZAR_I2C0_I2C_BYTE_CNT,MIZAR_I2C1_I2C_BYTE_CNT,MIZAR_I2C0_SF_LCNT,MIZAR_I2C0_SF_HCNT,MIZAR_I2C0_I2C_HS_LCNT,MIZAR_I2C0_I2C_HS_HCNT,MIZAR_I2C0_I2C_MSTR_CODE,MIZAR_I2C0_MASK_INTR,MIZAR_I2C0_TX_FIFO_THLD,MIZAR_I2C1_TX_FIFO_THLD,MIZAR_I2C0_RX_FIFO_THLD,MIZAR_I2C1_RX_FIFO_THLD,MIZAR_I2C0_DMA_CTRL,MIZAR_I2C0_SMB_HST_BLOCK_DATA,MIZAR_DMA_CH0_CTRL,MIZAR_DMA_CH0_SRC_ADDR,MIZAR_DMA_CH0_DEST_ADDR,MIZAR_DMA_CH0_SRC_XCNT,MIZAR_DMA_CH0_SRC_XMDFY,MIZAR_DMA_CH0_DEST_XMDFY,MIZAR_DMA_CH0_SRC_REQ,MIZAR_DMA_DMA_CH_EN,MIZAR_I2C0_TSFR_CTRL,MIZAR_I2C0_I2C_CURRENT_BYTECNT,MIZAR_I2C1_DMA_CTRL,MIZAR_I2C1_SMB_HST_BLOCK_DATA,MIZAR_DMA_CH1_CTRL,MIZAR_DMA_CH1_SRC_ADDR,MIZAR_DMA_CH1_DEST_ADDR,MIZAR_DMA_CH1_SRC_XCNT,MIZAR_DMA_CH1_SRC_XMDFY,MIZAR_DMA_CH1_DEST_XMDFY,MIZAR_DMA_CH1_SRC_REQ,MIZAR_I2C0_INTR_STS,MIZAR_LSS_SYSREG_RAW_STCR0",
      "Hidden_Test_Case_Name": "test_i2c0_mst_i2c1_slv_dma_hs_md",
      "ArtifactsPath": "TestRepo/i2c/test_i2c0_mst_i2c1_slv_dma_hs_md"
    },
    {
      "Index": 3,
      "SS / Module": "I2C",
      "Feature": "Three speeds in I2C mode: Standard (100 Kb/s)",
      "Test Case Name": "test_i2c0_mst_i2c1_slv_dma_st_md",
      "Test Description": "Performs a DMA-backed transfer from controller 0 to controller 1 in standard timing configuration. Waits for completion by polling a byte-count register and an interrupt, then DMAs received data to memory to validate data integrity.",
      "Speed": "Standard",
      "Mode": "DMA + Interrupt + Polling",
      "Remarks": "System interface configuration occurs via fixed address writes prior to I2C initialization. Interrupt handling validates the transfer-complete condition and clear sequence. Data buffers are set in SRAM and moved via DMA for comparison.",
      "Test Steps / Procedure": ["... see description above ..."],
      "Impacted Registers": "MIZAR_LSS_SYSREG_INTR_EN0,MIZAR_I2C0_FF,MIZAR_I2C1_FF,MIZAR_I2C0_INTR_CLR,MIZAR_I2C1_INTR_CLR,MIZAR_I2C0_DEV_CTRL,MIZAR_I2C1_DEV_CTRL,MIZAR_I2C1_SLV_ADDR,MIZAR_I2C0_TGT_SLV_ADDR,MIZAR_I2C0_I2C_BYTE_CNT,MIZAR_I2C1_I2C_BYTE_CNT,MIZAR_I2C0_SF_LCNT,MIZAR_I2C0_SF_HCNT,MIZAR_I2C0_MASK_INTR,MIZAR_I2C0_TX_FIFO_THLD,MIZAR_I2C1_TX_FIFO_THLD,MIZAR_I2C0_RX_FIFO_THLD,MIZAR_I2C1_RX_FIFO_THLD,MIZAR_I2C0_DMA_CTRL,MIZAR_I2C0_SMB_HST_BLOCK_DATA,MIZAR_DMA_CH0_CTRL,MIZAR_DMA_CH0_SRC_ADDR,MIZAR_DMA_CH0_DEST_ADDR,MIZAR_DMA_CH0_SRC_XCNT,MIZAR_DMA_CH0_SRC_XMDFY,MIZAR_DMA_CH0_DEST_XMDFY,MIZAR_DMA_CH0_SRC_REQ,MIZAR_DMA_DMA_CH_EN,MIZAR_I2C0_TSFR_CTRL,MIZAR_I2C0_I2C_CURRENT_BYTECNT,MIZAR_I2C1_DMA_CTRL,MIZAR_I2C1_SMB_HST_BLOCK_DATA,MIZAR_DMA_CH1_CTRL,MIZAR_DMA_CH1_SRC_ADDR,MIZAR_DMA_CH1_DEST_ADDR,MIZAR_DMA_CH1_SRC_XCNT,MIZAR_DMA_CH1_SRC_XMDFY,MIZAR_DMA_CH1_DEST_XMDFY,MIZAR_DMA_CH1_SRC_REQ,MIZAR_I2C0_INTR_STS,MIZAR_LSS_SYSREG_RAW_STCR0",
      "Hidden_Test_Case_Name": "test_i2c0_mst_i2c1_slv_dma_st_md",
      "ArtifactsPath": "TestRepo/i2c/test_i2c0_mst_i2c1_slv_dma_st_md"
    }
  ],
  "META_DATA": {
    "Repository": "https://github.com/titusbspgit/PSVValidation",
    "Branch": "main",
    "BaseDirectory": "TestRepo/i2c"
  }
}'''
# ---------------------------------------------------------------

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

OUTPUT_DIR = os.path.join("Test_Output", "I2C", "TestPlan")


def to_str(val):
    if isinstance(val, list):
        return "\n".join(str(x) for x in val)
    return "" if val is None else str(val)


def compute_ist_timestamps():
    ts_env = os.getenv("IST_TS")
    human_env = os.getenv("IST_HUMAN")
    if ts_env and human_env:
        return ts_env, human_env
    now = datetime.now(TZ) if TZ else datetime.utcnow()
    return now.strftime("%Y%m%d_%H%M%S"), now.strftime("%Y-%m-%d %H:%M:%S")


def autofit_columns(ws):
    # Approximate autofit by computing max string lengths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = cell.value
                if val is None:
                    length = 0
                else:
                    s = str(val)
                    # account for wrapped lines: take longest line
                    length = max(len(line) for line in s.split("\n")) if "\n" in s else len(s)
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        # set width with padding; cap to a reasonable max
        width = min(max_len + 2, 80)
        ws.column_dimensions[col_letter].width = width


def add_thin_borders(ws):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def main():
    data = json.loads(JSON_PAYLOAD)
    rows = data.get("TestCases", [])
    if not isinstance(rows, list) or not rows:
        raise SystemExit("ERROR: TestCases must be a non-empty array")

    # Build union of keys (first-seen order)
    header = []
    for row in rows:
        if not isinstance(row, dict):
            raise SystemExit("ERROR: each TestCases entry must be an object")
        for k in row.keys():
            if k not in header:
                header.append(k)

    # Create workbook and Data sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(header)
    # Write data preserving values; arrays become newline-joined strings here
    for row in rows:
        ws.append([to_str(row.get(k, "")) for k in header])

    # Base formatting on Data
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    autofit_columns(ws)

    # Create Meta_data_sheet with only META columns, preserve order and values
    meta_ws = wb.create_sheet(title="Meta_data_sheet")
    meta_ws.append(META_COLUMNS)
    # Build mapping from header to index for reading from Data rows
    idx = {name: i for i, name in enumerate(header)}
    for r in range(2, ws.max_row + 1):
        meta_row = []
        for name in META_COLUMNS:
            if name in idx:
                meta_row.append(ws.cell(row=r, column=idx[name] + 1).value)
            else:
                meta_row.append("")
        meta_ws.append(meta_row)
    # Very hidden meta sheet
    meta_ws.sheet_state = "veryHidden"

    # Rename Data to TestPlan
    ws.title = "TestPlan"

    # Remove META columns and ensure only approved MAIN columns remain in exact order
    # Build a map from existing header to its column index
    current_header = [c.value for c in ws[1]]
    existing_map = {name: i for i, name in enumerate(current_header)}

    # Build new table with MAIN_COLUMNS only
    new_ws = wb.create_sheet(title="__TMP_TestPlan__")
    new_ws.append(MAIN_COLUMNS)
    # Copy rows
    for r in range(2, ws.max_row + 1):
        new_row = []
        for name in MAIN_COLUMNS:
            if name in existing_map:
                val = ws.cell(row=r, column=existing_map[name] + 1).value
            else:
                val = ""
            new_row.append(val)
        new_ws.append(new_row)

    # Delete old TestPlan sheet and rename tmp
    wb.remove(ws)
    new_ws.title = "TestPlan"
    ws = new_ws

    # Strict formatting on TestPlan
    wrap_cols = set([
        "Test Description",
        "Remarks",
        "Test Steps / Procedure",
        "Validation / Acceptance Criteria",
    ])

    # Header formatting with solid blue fill
    header_fill = PatternFill(fill_type="solid", start_color="FF0070C0", end_color="FF0070C0")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    # Data rows formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            if col_name in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            elif col_name == "Index":
                cell.alignment = Alignment(vertical="top", horizontal="center")
            else:
                cell.alignment = Alignment(vertical="top", horizontal="left")

    ws.freeze_panes = "A2"
    autofit_columns(ws)

    # Adjust row heights after wrap (approximation by openpyxl default; explicit auto-height not exposed)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = None  # let Excel auto-fit on open

    # Apply thin borders to all populated cells
    add_thin_borders(ws)

    # Data validation dropdown for Code Generation (Required / Not)
    if "Code Generation (Required / Not)" in MAIN_COLUMNS:
        col_index = MAIN_COLUMNS.index("Code Generation (Required / Not)") + 1
        dv = DataValidation(type="list", formula1='"Required,Blank,Not Required"', allow_blank=True, showDropDown=True)
        dv.error = "Select one of: Required, Blank, Not Required"
        dv.prompt = "Choose code-generation requirement"
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col_index).coordinate}:{ws.cell(row=ws.max_row, column=col_index).coordinate}")

    # Compute IST timestamps for filename
    ts_compact, ts_human = compute_ist_timestamps()

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"I2C_TestPlan_{ts_compact}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(out_path)

    print(f"GENERATED: {out_path}")
    print(f"IST_TIMESTAMP: {ts_compact}")
    print(f"IST_HUMAN: {ts_human}")


if __name__ == "__main__":
    main()
