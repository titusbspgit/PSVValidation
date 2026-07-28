#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate a REAL .xlsx Test Plan with two sheets (TestPlan and MetaData) from embedded JSON
- Applies formatting (bold blue headers with white font, wrap text, freeze first row, column widths)
- Sets MetaData sheet to VERY HIDDEN
- Saves into Test_Output/USB/TestPlan/USB_TestPlan_<YYYYMMDD>_<HHMMSS>.xlsx where time is IST
"""
from __future__ import annotations
import os
from pathlib import Path
from datetime import datetime, timedelta, timezone
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    TZ = ZoneInfo("Asia/Kolkata")
except Exception:
    # Fallback fixed IST offset if zoneinfo is unavailable
    TZ = timezone(timedelta(hours=5, minutes=30))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --------------------------- Embedded JSON data (immutable source) --------------------------- #
JSON_DATA = [
  {
    "Index": "1",
    "SS / Module": "USB",
    "Feature": "USB Register Read/Write Sanity",
    "Test Case Name": "basic_test",
    "Test Description": "Sanity check of USB memory-mapped register access by reading two registers and performing a write-readback on a third, logging the observed values.",
    "Meta Test Description": "Program prints a greeting, reads from memory-mapped addresses 0xA0000000 and 0xA001706C using read_reg(), prints both values, writes 0xDEADBEEF to address 0xA0240000 using write_reg(), reads back from 0xA0240000 with read_reg(), prints the read-back value, and calls finish(0).",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "0xA0000000",
    "Memory End Offset": "0xA0240000",
    "Remarks": "NA",
    "Test Steps / Procedure": "1. Start the test application.\n2. Read the value from address 0xA0000000 and log it.\n3. Read the value from address 0xA001706C and log it.\n4. Write 0xDEADBEEF to address 0xA0240000.\n5. Read back from address 0xA0240000 and log the value.\n6. End the test.",
    "Meta Test Steps / Procedure": "- printf(\"[C-Programme] Hello world\\n\").\n- int rd_data = 0.\n- rd_data = read_reg(0xA0000000); printf(\"THE READ DATA is %X\\n\", rd_data).\n- rd_data = read_reg(0xA001706C); printf(\"THE READ DATA1 is %X\\n\", rd_data).\n- write_reg(0xA0240000, 0xDEADBEEF).\n- rd_data = read_reg(0xA0240000); printf(\"THE READ DATA2 is %X\\n\", rd_data).\n- finish(0).",
    "Impacted Registers": "NA",
    "Meta Impacted Registers": "NA",
    "Validation / Acceptance Criteria": "NA",
    "Meta Validation / Acceptance Criteria": "NA",
    "Code Generation (Required / Not)": "NA",
    "Meta Headers": "#include <stdio.h>\n#include <stdlib.h>\n#include <usb/usb_def.h>\n#include <usb/usb_offsets.h>",
    "Meta Macros": "NA",
    "Meta Arrays": "NA"
  },
  {
    "Index": "2",
    "SS / Module": "USB",
    "Feature": "USB Host Enumeration (Low-Speed)",
    "Test Case Name": "usb_host_enumeration_ls",
    "Test Description": "Initializes the USB host controller, configures port and event/command rings, enables interrupts, issues Enable Slot and Address Device commands, and performs EP0 control transfers (GET DESCRIPTOR, SET CONFIGURATION) to enumerate a low-speed device. Completion is verified via event ring updates and successful test termination.",
    "Meta Test Description": "The test programs NIC and enables all IRQs. It configures global control/tuning registers (GCTL, GFLADJ, GUCTL), programs PIPE/PHY control via MIZAR_USB_BASE+0xC2C0/0xC200, reads capability registers (HCSPARAMS1/2, PAGESIZE), enables port change wake events and resets the port using PORTSC_20. It sets up the Event Ring Segment Table and Scratchpad Buffers, loads the Device Context Base Address Array (DCBAA), points CRCR to the Default Command Ring, sets MaxSlots and CIE in CONFIG, programs DCBAAP, ERSTSZ, ERDP, ERSTBA, IMOD/IMAN, enables interrupts (USBCMD, sysreg INTR_EN0), and starts the host (USBCMD run). It waits for interrupt completion (int_pend cleared by ISR), acknowledges USBSTS, re-arms IMAN/ERDP, confirms port connect, increments ERDP, performs a port reset sequence, and waits for another interrupt. It builds and submits an Enable Slot command TRB on the command ring, services resulting events, then prepares an input context for EP0 and submits Address Device commands (twice to handle BSR). After each submission, it acknowledges USBSTS/IMAN, updates ERDP, rings the doorbell (DB), and waits for completion events from the Default Event Ring. The enumeration() routine then builds EP0 transfer TRBs to GET DEVICE DESCRIPTOR, GET CONFIGURATION DESCRIPTOR, SET CONFIGURATION, and additional GET CONFIG stages, rings the host (MIZAR_USB_BASE+0x484), and polls Default_Event_Ring_Array+0x110 until non-zero. The test completes with finish(0).",
    "Speed": "NA",
    "Mode": "ISR",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Requires a connected low-speed USB device on the root port. Ensure host interrupts are enabled and routed and that controller memory structures (command/event rings, scratchpads, DCBAA, input contexts) are allocated and accessible.",
    "Test Steps / Procedure": "1. Program the USB host global control and tuning registers and initialize PIPE/PHY controls.\n2. Read host capability registers and enable port change/warm wake events; apply a port reset sequence.\n3. Initialize Event Ring Segment Table, Scratchpad Buffers, and load the Device Context Base Address Array.\n4. Set up the Command Ring, configure MaxSlots and Context Enable in the configuration register, and program DCBAA/ERST/ERDP/ERSTBA/IMOD/IMAN.\n5. Enable host controller interrupts and start the controller; verify an interrupt occurs and service it.\n6. Acknowledge status, re-arm interrupt enables and event dequeue pointers; confirm the port is connected.\n7. Submit an Enable Slot command and service the resulting events.\n8. Build the EP0 input context for the default control endpoint and submit Address Device commands; service events after each submission.\n9. Perform EP0 control transfers: GET DEVICE DESCRIPTOR, GET CONFIGURATION DESCRIPTORS, and SET CONFIGURATION; ring the controller and poll the event ring for completion.\n10. Confirm successful completion events and terminate the test successfully.",
    "Meta Test Steps / Procedure": "- nic_programming(); enable interrupts with GIC_EnableAllIRQ().\n- write_reg(MIZAR_USB_GCTL, set_data(read_reg(MIZAR_USB_GCTL), 0xFFFFFFFF, 0x30c11234)).\n- write_reg(MIZAR_USB_GFLADJ, set_data(read_reg(MIZAR_USB_GFLADJ), 0xFFFFFFFF, 0x0A87F000)).\n- write_reg(MIZAR_USB_GUCTL, set_data(read_reg(MIZAR_USB_GUCTL), 0xFFFFFFFF, 0x02000010)).\n- rd_data = read_reg(MIZAR_USB_BASE+0xC2C0); write_reg(MIZAR_USB_BASE+0xC2C0, 0x010C0002) // PIPECTL.\n- rd_data = read_reg(MIZAR_USB_BASE+0xC200); write_reg(MIZAR_USB_BASE+0xC200, 0x00102407) // PHYCTL.\n- port_count = read_reg(MIZAR_USB_HCSPARAMS1); rd_data = read_reg(MIZAR_USB_SUPTPRT2_DW2); rd_data = read_reg(MIZAR_USB_SUPTPRT3_DW2); rd_data = read_reg(MIZAR_USB_PORTSC_20).\n- Enable port wake events: write_reg(MIZAR_USB_PORTSC_20, set_data(read_reg(MIZAR_USB_PORTSC_20), USB_PORTSC_20_WCE, 1)); similarly set WDE and WOE.\n- write_reg(MIZAR_USB_PORTSC_20, 0x0E0002A0) // port config.\n- db_offset = read_reg(MIZAR_USB_DBOFF).\n- Program Event Ring Segment Table: write_reg(Event_Ring_Segment_Table, Default_Event_Ring_Array); write_reg(Event_Ring_Segment_Table+DWORD, 0x0); write_reg(Event_Ring_Segment_Table+2*DWORD, 0x30).\n- rd_data = read_reg(MIZAR_USB_HCSPARAMS2); rd_data = read_reg(MIZAR_USB_PAGESIZE).\n- Program Scratchpad Buffer Array: write_reg(Scratchpad_Buffer_Array, SCRATCHPAD0); write_reg(Scratchpad_Buffer_Array+DWORD, 0x0); write_reg(Scratchpad_Buffer_Array+2*DWORD, SCRATCHPAD1); write_reg(Scratchpad_Buffer_Array+3*DWORD, 0x0).\n- Load DCBAA: write_reg(Device_Context_Base_Address_Array, Scratchpad_Buffer_Array); write_reg(Device_Context_Base_Address_Array+DWORD, 0x0); write_reg(Device_Context_Base_Address_Array+2*DWORD, Device_Context_Array+0x100); write_reg(Device_Context_Base_Address_Array+3*DWORD, 0x0); write_reg(Device_Context_Base_Address_Array+4*DWORD, Device_Context_Array+0x0D00); write_reg(Device_Context_Base_Address_Array+5*DWORD, 0x0).\n- Command ring base: write_reg(MIZAR_USB_CRCR_LO, Default_Command_Ring+0x1); write_reg(MIZAR_USB_CRCR_HI, 0x0).\n- Configure MaxSlots: write_reg(MIZAR_USB_CONFIG, 0x10); rd_data = read_reg(MIZAR_USB_CONFIG); enable CIE: write_reg(MIZAR_USB_CONFIG, 0x110).\n- Set DCBAAP/ERST: write_reg(MIZAR_USB_DCBAAP_LO, Device_Context_Base_Address_Array); write_reg(MIZAR_USB_DCBAAP_HI, 0x0); write_reg(MIZAR_USB_ERSTSZ, 0x1);\n  write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array); write_reg(MIZAR_USB_ERDP_HI, 0x0);\n  write_reg(MIZAR_USB_ERSTBA_LO, Event_Ring_Segment_Table); write_reg(MIZAR_USB_ERSTBA_HI, 0x0);\n  write_reg(MIZAR_USB_IMOD, 0x0); write_reg(MIZAR_USB_IMAN, 0x2).\n- Enable interrupts and run: write_reg(MIZAR_USB_USBCMD, 0x4); write_reg(MIZAR_LSS_SYSREG_INTR_EN0, 0x80000000); write_reg(MIZAR_USB_USBCMD, 0x5); int_pend=1; while(int_pend){wait_on(100);} .\n- Acknowledge and re-arm: usb_status=read_reg(MIZAR_USB_USBSTS); write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2); write_reg(MIZAR_USB_ERDP_HI, 0x0); port_status=read_reg(MIZAR_USB_PORTSC_20).\n- Advance ERDP and reset port: write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array+0x18); write_reg(MIZAR_USB_PORTSC_20, 0x0E0006F1); write_reg(MIZAR_USB_USBSTS, 0x8); write_reg(MIZAR_USB_IMAN, 0x2); write_reg(MIZAR_USB_PORTSC_20, 0x0E220200); port_status=read_reg(MIZAR_USB_PORTSC_20).\n- Wait for interrupt: int_pend=1; while(int_pend){wait_on(100);} .\n- Enable Slot TRB: write_reg(Default_Command_Ring+0x0,0x0); write_reg(Default_Command_Ring+0x4,0x0); write_reg(Default_Command_Ring+0x8,0x0); write_reg(Default_Command_Ring+0xC,0x00002401).\n- Service events: usb_status=read_reg(MIZAR_USB_USBSTS); write_reg(MIZAR_USB_USBSTS,0x8); write_reg(MIZAR_USB_IMAN,0x2); write_reg(MIZAR_USB_ERDP_HI,0x0); write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array+0x28);\n  write_reg(MIZAR_USB_USBSTS,0x8); write_reg(MIZAR_USB_IMAN,0x2); write_reg(MIZAR_USB_PORTSC_20,0x0E200E01); write_reg(MIZAR_USB_DB,0x0); int_pend=1; while(int_pend){wait_on(100);} .\n- Build Input Context: write_reg(Default_Input_Context,0x0); write_reg(Default_Input_Context+DWORD,0x3);\n  write_reg(Default_Input_Context+0x40,0x08200000); write_reg(Default_Input_Context+0x44,0x00010000);\n  write_reg(Default_Input_Context+0x80,0x00); write_reg(Default_Input_Context+0x84,0x00080020);\n  write_reg(Default_Input_Context+0x88, EP0_TR_Dequeue_Pointer | 0x1); write_reg(Default_Input_Context+0x90,0x08).\n- Address Device (BSR set): input_context_address = (Default_Input_Context)+0x0; write_reg(Default_Command_Ring+0x10, input_context_address); write_reg(Default_Command_Ring+0x1C,0x01002E01);\n  write_reg(MIZAR_USB_USBSTS,0x8); write_reg(MIZAR_USB_IMAN,0x2); write_reg(MIZAR_USB_ERDP_HI,0x0); write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array+0x38); write_reg(MIZAR_USB_DB,0x0);\n  int_pend=1; while(int_pend){wait_on(100);} .\n- Read completion and clear BSR path: event_completion=read_reg(Default_Event_Ring_Array+0x30); write_reg(Default_Input_Context+DWORD,0x3);\n  re-write same EP0 context fields and TR dequeue pointer.\n- Address Device (BSR clear): write_reg(Default_Command_Ring+0x20,input_context_address); write_reg(Default_Command_Ring+0x2C,0x01002C01);\n  write_reg(MIZAR_USB_USBSTS,0x8); write_reg(MIZAR_USB_IMAN,0x2); write_reg(MIZAR_USB_ERDP_HI,0x0); write_reg(MIZAR_USB_ERDP_LO, Default_Event_Ring_Array+0x48); write_reg(MIZAR_USB_DB,0x0);\n  int_pend=1; while(int_pend){wait_on(100);} .\n- event_completion = read_reg(Default_Event_Ring_Array+0x40); call enumeration().\n- enumeration(): Build EP0 TRBs sequence for control transfers:\n  • GET DEVICE DESCRIPTOR setup/data/status at EP0_TR_Dequeue_Pointer (+0x0..0x2C).\n  • GET CONFIGURATION (len=9) setup/data/status at (+0x30..0x5C).\n  • SET CONFIGURATION setup/status at (+0x60..0x7C).\n  • GET CONFIGURATION2 setup/data/status at (+0x80..0xAC).\n  • GET CONFIGURATION3 setup/data/status at (+0xB0..0xDC).\n  • write_reg(MIZAR_USB_BASE+0x484,0x1); printf of event_completion; poll event_completion=read_reg(Default_Event_Ring_Array+0x110) until non-zero using wait_on(4).\n- printf(\"after enumeration\\n\"); finish(0).",
    "Impacted Registers": "MIZAR_USB_GCTL, MIZAR_USB_GFLADJ, MIZAR_USB_GUCTL, MIZAR_USB_HCSPARAMS1, MIZAR_USB_SUPTPRT2_DW2, MIZAR_USB_SUPTPRT3_DW2, MIZAR_USB_PORTSC_20, MIZAR_USB_DBOFF, MIZAR_USB_HCSPARAMS2, MIZAR_USB_PAGESIZE, MIZAR_USB_CRCR_LO, MIZAR_USB_CRCR_HI, MIZAR_USB_CONFIG, MIZAR_USB_DCBAAP_LO, MIZAR_USB_DCBAAP_HI, MIZAR_USB_ERSTSZ, MIZAR_USB_ERDP_LO, MIZAR_USB_ERDP_HI, MIZAR_USB_ERSTBA_LO, MIZAR_USB_ERSTBA_HI, MIZAR_USB_IMOD, MIZAR_USB_IMAN, MIZAR_USB_USBCMD, MIZAR_USB_USBSTS, MIZAR_USB_DB, INTR_EN0, MSK_STS0, RAW_STCR0",
    "Meta Impacted Registers": "MIZAR_USB_GCTL, MIZAR_USB_GFLADJ, MIZAR_USB_GUCTL, MIZAR_USB_HCSPARAMS1, MIZAR_USB_SUPTPRT2_DW2, MIZAR_USB_SUPTPRT3_DW2, MIZAR_USB_PORTSC_20, MIZAR_USB_DBOFF, MIZAR_USB_HCSPARAMS2, MIZAR_USB_PAGESIZE, MIZAR_USB_CRCR_LO, MIZAR_USB_CRCR_HI, MIZAR_USB_CONFIG, MIZAR_USB_DCBAAP_LO, MIZAR_USB_DCBAAP_HI, MIZAR_USB_ERSTSZ, MIZAR_USB_ERDP_LO, MIZAR_USB_ERDP_HI, MIZAR_USB_ERSTBA_LO, MIZAR_USB_ERSTBA_HI, MIZAR_USB_IMOD, MIZAR_USB_IMAN, MIZAR_USB_USBCMD, MIZAR_USB_USBSTS, MIZAR_USB_DB, MIZAR_LSS_SYSREG_INTR_EN0, MIZAR_LSS_SYSREG_MSK_STS0, MIZAR_LSS_SYSREG_RAW_STCR0",
    "Validation / Acceptance Criteria": "- Host controller enters run state with interrupts enabled; at least one interrupt is serviced after start and after port reset, with status acknowledged via USBSTS and IMAN re-armed.\n- Port status indicates device connection following reset using the Port Status and Control register.\n- Enable Slot and Address Device commands complete successfully as indicated by event ring updates and no error conditions.\n- EP0 control transfers (GET DEVICE/CONFIGURATION and SET CONFIGURATION) complete; event ring entry at the monitored offset becomes non-zero.\n- Test terminates with a PASS (finish(0)).",
    "Meta Validation / Acceptance Criteria": "- int_pend transitions from 1 to 0 due to Default_IRQHandler() and while(int_pend){wait_on(...)} loops exit each time before proceeding.\n- After start: usb_status=read_reg(MIZAR_USB_USBSTS); status is acknowledged by write_reg(MIZAR_USB_USBSTS,0x8) and IMAN is re-armed (write_reg(MIZAR_USB_IMAN,0x2)); ERDP updated.\n- port_status=read_reg(MIZAR_USB_PORTSC_20) reflects connect high after reset sequence (writes 0xE0006F1 then 0xE220200).\n- Command ring submissions (0x00002401, 0x01002E01, 0x01002C01) are followed by write_reg(MIZAR_USB_DB,0x0) and corresponding event completions read from Default_Event_Ring_Array+0x30/0x40 (non-zero).\n- enumeration(): After programming EP0 TRBs, write_reg(MIZAR_USB_BASE+0x484,0x1) and polling read_reg(Default_Event_Ring_Array+0x110) eventually yields non-zero.\n- finish(0) is executed indicating PASS.",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>\n#include <stdlib.h>\n\"usb.h\"",
    "Meta Macros": "NA",
    "Meta Arrays": "NA"
  }
]

# Define columns for both sheets
TESTPLAN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

METADATA_COLUMNS = [
    "Index",
    "Test Case Name",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")  # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

TEXT_HEAVY_COLUMNS = {
    "Test Description",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Validation / Acceptance Criteria",
}


def auto_width(ws, headers):
    # Compute max length per column and set a reasonable width
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = "" if cell.value is None else str(cell.value)
                # account for line breaks by taking the longest line
                val_len = max(len(part) for part in val.splitlines()) if "\n" in val else len(val)
                if val_len > max_len:
                    max_len = val_len
        # Base width
        width = max_len + 2
        # Provide larger width for text-heavy columns
        if header in TEXT_HEAVY_COLUMNS:
            width = max(30, min(width, 80))
        else:
            width = max(12, min(width, 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def apply_header_style(ws, headers):
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP


def append_rows(ws, headers, rows):
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    # Wrap text for all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = WRAP


def build_workbook(data: list[dict]) -> Workbook:
    wb = Workbook()
    # First sheet: TestPlan
    ws1 = wb.active
    ws1.title = "TestPlan"
    apply_header_style(ws1, TESTPLAN_COLUMNS)
    # Prepare data rows in the given order
    tp_rows = []
    for item in data:
        tp_rows.append({k: item.get(k, "") for k in TESTPLAN_COLUMNS})
    append_rows(ws1, TESTPLAN_COLUMNS, tp_rows)
    ws1.freeze_panes = "A2"
    auto_width(ws1, TESTPLAN_COLUMNS)

    # Second sheet: MetaData (Very Hidden)
    ws2 = wb.create_sheet(title="MetaData")
    apply_header_style(ws2, METADATA_COLUMNS)
    md_rows = []
    for item in data:
        md_rows.append({k: item.get(k, "") for k in METADATA_COLUMNS})
    append_rows(ws2, METADATA_COLUMNS, md_rows)
    ws2.freeze_panes = "A2"
    auto_width(ws2, METADATA_COLUMNS)
    # Set MetaData to very hidden
    ws2.sheet_state = "veryHidden"

    return wb


def main() -> None:
    # Validate JSON
    if not isinstance(JSON_DATA, list) or not all(isinstance(x, dict) for x in JSON_DATA):
        raise SystemExit("json_data must be a list of dicts")

    ip_name = "USB"
    ts = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
    out_dir = Path("Test_Output/USB/TestPlan")
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{ip_name}_TestPlan_{ts}.xlsx"
    out_path = out_dir / filename

    wb = build_workbook(JSON_DATA)
    wb.save(str(out_path))  # REAL .xlsx
    print(str(out_path))


if __name__ == "__main__":
    main()
