#!/usr/bin/env python3
import argparse
import json
import os
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)


def load_json(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_value(v: Any) -> Any:
    # Preserve values; join lists with newlines to keep single-cell multi-line representation
    if isinstance(v, list):
        return "\n".join(str(x) for x in v)
    return v


def build_union_keys(rows: List[Dict[str, Any]]) -> List[str]:
    seen = []
    seen_set = set()
    for row in rows:
        for k in row.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def autofit_columns(ws: Worksheet):
    # Approximate column width based on max content length
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_len = 0
        for cell in col:
            val = cell.value
            if val is None:
                continue
            s = str(val)
            if "\n" in s:
                # For wrapped cells, consider average line length
                s = max((len(line) for line in s.splitlines()), default=0)
                length = s
            else:
                length = len(s)
            if length > max_len:
                max_len = length
        # add a little padding
        adjusted_width = min(max(10, max_len + 2), 80)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width


def apply_table_format(ws: Worksheet):
    # Header formatting
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align
        cell.border = THIN_BORDER
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # Data rows formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            header = ws.cell(row=1, column=cell.col_idx).value
            if header in WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            elif header == "Index":
                cell.alignment = Alignment(wrap_text=False, vertical="top", horizontal="center")
            else:
                cell.alignment = Alignment(wrap_text=False, vertical="top", horizontal="left")
            cell.border = THIN_BORDER

    # Autofit and row height adjust (best-effort)
    autofit_columns(ws)
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--ip-name", required=True)
    ap.add_argument("--ist-stamp", required=False, help="YYYYMMDD_HHMMSS in IST; if omitted, compute now")
    args = ap.parse_args()

    data = load_json(args.input)

    # Validate and extract rows
    tests = data.get("tests")
    if tests is None or not isinstance(tests, list) or len(tests) == 0:
        raise SystemExit("Invalid or empty JSON: 'tests' array is required and must be non-empty")

    # Normalize test dicts (preserve keys as-is; convert lists to newline strings)
    norm_rows: List[Dict[str, Any]] = []
    for t in tests:
        row = {}
        for k, v in t.items():
            row[k] = normalize_value(v)
        norm_rows.append(row)

    # Create workbook and Data sheet
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"

    # Build union of keys preserving first appearance across rows
    union_keys = build_union_keys(norm_rows)

    # Write header
    ws_data.append(union_keys)

    # Write rows
    for r in norm_rows:
        ws_data.append([r.get(k, "") for k in union_keys])

    # Prepare Meta_data_sheet from META_COLUMNS (copy values if present else blank)
    ws_meta = wb.create_sheet(title="Meta_data_sheet")
    ws_meta.append(META_COLUMNS)
    for r in norm_rows:
        ws_meta.append([r.get(k, "") for k in META_COLUMNS])
    # Very hide meta sheet
    ws_meta.sheet_state = "veryHidden"

    # Rename Data to TestPlan and filter/reorder columns to MAIN_COLUMNS
    ws = ws_data
    ws.title = "TestPlan"

    # Build mapping from header name to column index in current sheet
    header_to_idx = {ws.cell(row=1, column=i+1).value: i+1 for i in range(ws.max_column)}

    # Create a temporary 2D array for filtered data
    filtered_rows = []
    filtered_rows.append(MAIN_COLUMNS)
    for row_idx in range(2, ws.max_row + 1):
        filtered_rows.append([
            ws.cell(row=row_idx, column=header_to_idx.get(col)).value if header_to_idx.get(col) else ""
            for col in MAIN_COLUMNS
        ])

    # Clear sheet and write filtered content
    ws.delete_rows(1, ws.max_row)
    for r in filtered_rows:
        ws.append(r)

    # Apply formatting to TestPlan only
    apply_table_format(ws)

    # Construct output filename
    if args.ist_stamp:
        ist_stamp = args.ist_stamp
    else:
        ist = timezone(timedelta(hours=5, minutes=30))
        now = datetime.now(ist)
        ist_stamp = now.strftime("%Y%m%d_%H%M%S")

    out_dir = args.output_dir.rstrip("/\\")
    os.makedirs(out_dir, exist_ok=True)
    out_name = f"{args.ip_name}_TestPlan_{ist_stamp}.xlsx"
    out_path = os.path.join(out_dir, out_name)
    wb.save(out_path)

    # Write helper file with filename for the workflow to read
    with open(os.path.join(out_dir, ".last_generated_filename"), "w", encoding="utf-8") as fh:
        fh.write(out_name)

    print(out_path)

if __name__ == "__main__":
    main()
