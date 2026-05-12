#!/usr/bin/env python3
import argparse
import json
import os
import sys
import zipfile
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as e:
    print(f"ERROR: openpyxl not available: {e}")
    sys.exit(2)

META_COLS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_ORDER = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Imparted Registers",  # placeholder if typo occurs; corrected below if real key exists
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

WRAP_COLS = [
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
]

ALLOWED_DV = ["Required", "Blank", "Not Required"]


def parse_args():
    ap = argparse.ArgumentParser(description="Convert JSON to formatted TestPlan XLSX")
    ap.add_argument("--json", required=True, help="Path to JSON array file")
    ap.add_argument("--ipname", required=True, help="IP name for filename rule")
    ap.add_argument("--outdir", required=True, help="Output directory in repo")
    return ap.parse_args()


def load_json(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("JSON input must be a non-empty array of objects")
    # Ensure each item is a dict
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"JSON item at index {i} is not an object")
    return data


def schema_union_first_seen(rows: List[Dict[str, Any]]) -> List[str]:
    keys = []
    seen = set()
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    return keys


def normalize_rows(rows: List[Dict[str, Any]], keys: List[str]) -> List[Dict[str, Any]]:
    norm = []
    for row in rows:
        new = {}
        for k in keys:
            new[k] = row.get(k, "")
        norm.append(new)
    return norm


def numbering_from_text(text: str) -> str:
    if text is None:
        return ""
    # Split by newlines and filter empties
    parts = [p.strip() for p in str(text).replace("\r", "\n").split("\n")]
    parts = [p for p in parts if p]
    if not parts:
        return str(text) if text is not None else ""
    # Re-number with 1., 2., ...
    out_lines = []
    for i, p in enumerate(parts, start=1):
        # Remove any leading numeric bullets like '1) ' or '1. '
        q = p
        if len(q) > 2 and q[0].isdigit() and (q[1] in ")."):
            q = q[2:].lstrip()
        out_lines.append(f"{i}. {q}")
    return "\n".join(out_lines)


def autosize_columns(ws):
    # Approximate auto-fit by measuring max string length
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = cell.value
                if val is None:
                    length = 0
                else:
                    length = len(str(val))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(120, max(10, max_len + 2))


def apply_borders(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value) != "":
                cell.border = border


def set_meta_sheet_very_hidden(wb):
    meta = wb["Meta_data_sheet"]
    meta.sheet_state = "veryHidden"


def validate_xlsx(path: str) -> None:
    # Ensure it's a zip and has OOXML core files
    with zipfile.ZipFile(path, "r") as zf:
        names = set(zf.namelist())
        if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
            raise ValueError("Not a valid OOXML workbook (missing core files)")
    # Re-open with openpyxl for sheet checks
    wb = load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    if set(sheets) != {"TestPlan", "Meta_data_sheet"}:
        raise ValueError(f"Unexpected sheets: {sheets}")
    if wb["Meta_data_sheet"].sheet_state != "veryHidden":
        raise ValueError("Meta_data_sheet must be Very Hidden")
    # Basic check for data validation only on Code Generation column
    ws = wb["TestPlan"]
    # Gather DV ranges
    # We cannot fully assert ranges without heavy parsing; ensure at least one DV exists
    # and its formula list matches allowed values
    dvs = getattr(ws, 'data_validations', None)
    if not dvs or not dvs.dataValidation:
        raise ValueError("Data validation missing on TestPlan")
    ok_list = False
    for dv in dvs.dataValidation:
        if dv.type == "list" and dv.formula1:
            f = dv.formula1.replace('"', '').strip()
            if f == ",".join(ALLOWED_DV):
                ok_list = True
                break
    if not ok_list:
        raise ValueError("Data validation list does not match required values")


def main():
    args = parse_args()
    rows = load_json(args.json)
    # Normalize schema
    keys = schema_union_first_seen(rows)
    norm_rows = normalize_rows(rows, keys)

    # Prepare workbook with a single 'Data' sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")

    for j, k in enumerate(keys, start=1):
        c = ws.cell(row=1, column=j, value=k)
        c.font = header_font
        c.alignment = header_align
        c.fill = header_fill

    # Data rows (preserve values exactly)
    for i, row in enumerate(norm_rows, start=2):
        for j, k in enumerate(keys, start=1):
            ws.cell(row=i, column=j, value=row.get(k, ""))

    # Freeze top row
    ws.freeze_panes = "A2"

    # Base autosize before transformations
    autosize_columns(ws)

    # Create Meta_data_sheet and copy META cols as-is (unnumbered)
    meta_ws = wb.create_sheet("Meta_data_sheet")
    for j, k in enumerate(META_COLS, start=1):
        mc = meta_ws.cell(row=1, column=j, value=k)
        mc.font = header_font
        mc.alignment = header_align
        mc.fill = header_fill
    for i, row in enumerate(norm_rows, start=2):
        for j, k in enumerate(META_COLS, start=1):
            meta_ws.cell(row=i, column=j, value=row.get(k, ""))

    # Normalize MAIN sheet: operate on existing 'Data' sheet only
    # Remove META cols and reorder to MAIN_ORDER
    # Resolve potential typo in MAIN_ORDER (Imparted vs Impacted)
    effective_order = []
    for name in MAIN_ORDER:
        if name == "Imparted Registers":
            if "Impacted Registers" in keys:
                effective_order.append("Impacted Registers")
            elif name in keys:
                effective_order.append(name)
            continue
        if name in keys:
            effective_order.append(name)
    # Keep any non-meta and non-main extras in original order after main columns
    extras = [k for k in keys if k not in META_COLS and k not in effective_order]
    final_order = effective_order + extras

    # Snapshot current data
    data_matrix = [[r.get(k, "") for k in final_order] for r in norm_rows]

    # Clear and rewrite 'Data' as per final_order
    ws.delete_rows(1, ws.max_row)
    for j, k in enumerate(final_order, start=1):
        c = ws.cell(row=1, column=j, value=k)
        c.font = header_font
        c.alignment = header_align
        c.fill = header_fill
    for i, rowvals in enumerate(data_matrix, start=2):
        for j, val in enumerate(rowvals, start=1):
            ws.cell(row=i, column=j, value=val)

    # Rename Data -> TestPlan
    ws.title = "TestPlan"

    # Apply formatting to TestPlan
    wrap_cols_idx = [final_order.index(c) + 1 for c in WRAP_COLS if c in final_order]
    for i in range(2, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            cell = ws.cell(row=i, column=j)
            is_wrap = j in wrap_cols_idx
            halign = "left"
            if final_order[j - 1] == "Index":
                halign = "center"
            cell.alignment = Alignment(wrap_text=is_wrap, horizontal=halign, vertical="top")

    # Number inside cells for procedure and acceptance criteria
    for col_name in ["Test Steps / Procedure", "Validation / Acceptance Criteria"]:
        if col_name in final_order:
            ci = final_order.index(col_name) + 1
            for i in range(2, ws.max_row + 1):
                v = ws.cell(row=i, column=ci).value
                ws.cell(row=i, column=ci, value=numbering_from_text(v))

    # Autosize and approximate auto row height based on line breaks
    autosize_columns(ws)
    base_height = 15
    for i in range(2, ws.max_row + 1):
        max_lines = 1
        for j in wrap_cols_idx:
            v = ws.cell(row=i, column=j).value
            if v is None:
                continue
            lines = str(v).count("\n") + 1
            if lines > max_lines:
                max_lines = lines
        ws.row_dimensions[i].height = base_height * max_lines

    # Apply thin borders to all populated cells
    apply_borders(ws)

    # Data validation only on Code Generation (Required / Not)
    if "Code Generation (Required / Not)" in final_order:
        cg_col = final_order.index("Code Generation (Required / Not)") + 1
        dv = DataValidation(type="list", formula1='"' + ",".join(ALLOWED_DV) + '"', allow_blank=True)
        dv.errorTitle = "Invalid choice"
        dv.error = "Select one of: Required, Blank, Not Required"
        start_cell = ws.cell(row=2, column=cg_col).coordinate
        end_cell = ws.cell(row=max(2, ws.max_row), column=cg_col).coordinate
        dv.ranges.append(f"{start_cell}:{end_cell}")
        # Clear any existing validations then add ours
        if hasattr(ws, 'data_validations') and ws.data_validations is not None:
            ws.data_validations.dataValidation = []
        ws.add_data_validation(dv)

    # Meta sheet visibility
    set_meta_sheet_very_hidden(wb)

    # Safety: only TestPlan and Meta_data_sheet should exist
    if set(wb.sheetnames) != {"TestPlan", "Meta_data_sheet"}:
        raise RuntimeError(f"Unexpected sheets present: {wb.sheetnames}")

    # Compute IST timestamped filename
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    fname = f"{args.ipname}_TestPlan_{now_ist.strftime('%Y%m%d')}_{now_ist.strftime('%H%M%S')}.xlsx"
    outdir = args.outdir.rstrip("/")
    os.makedirs(outdir, exist_ok=True)
    outpath = os.path.join(outdir, fname)

    wb.save(outpath)

    # Validate saved workbook
    validate_xlsx(outpath)

    print(outpath)
    # Export for GitHub Actions
    ghe = os.getenv("GITHUB_ENV")
    if ghe:
        with open(ghe, "a", encoding="utf-8") as envf:
            envf.write(f"OUTPUT_FILE={outpath}\n")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"FATAL: {e}")
        sys.exit(1)
