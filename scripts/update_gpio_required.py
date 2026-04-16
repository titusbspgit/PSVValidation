import sys
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("TestRepo/gpio/GPIO_TestCase.xlsx")
DST = Path("TestRepo/gpio/GPIO_TestCase_Update.xlsx")
COLUMN_NAME = "Code Generation - Required/Not Required"
NEW_VALUE = "Required"

if not SRC.exists():
    print(f"Source file not found: {SRC}")
    sys.exit(1)

wb = load_workbook(SRC)

updated_any = False
sheets_with_column = []

for ws in wb.worksheets:
    # Build header map from the first row
    headers = {}
    try:
        header_row = ws[1]
    except Exception as e:
        print(f"Unable to access header row in sheet '{ws.title}': {e}")
        continue

    for cell in header_row:
        raw = cell.value
        if raw is None:
            continue
        val = str(raw).strip()
        if not val:
            continue
        headers[val] = cell.col_idx  # numeric 1-based index

    if COLUMN_NAME in headers:
        col = headers[COLUMN_NAME]
        max_row = ws.max_row or 1
        if max_row < 2:
            print(f"Sheet '{ws.title}' has no data rows; skipping updates but will still include in output.")
        else:
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=col, value=NEW_VALUE)
            updated_any = True
            sheets_with_column.append(ws.title)
            print(f"Updated sheet '{ws.title}' column '{COLUMN_NAME}' to '{NEW_VALUE}' for rows 2..{max_row}")
    else:
        print(f"Column '{COLUMN_NAME}' not found in sheet '{ws.title}', skipping.")

# Ensure destination directory exists and save
DST.parent.mkdir(parents=True, exist_ok=True)
wb.save(DST)

if updated_any:
    print(f"Saved updated workbook to {DST}; updated sheets: {sheets_with_column}")
else:
    print("No sheets contained the target column; saved copy without changes.")
