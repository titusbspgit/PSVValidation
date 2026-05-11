#!/usr/bin/env python3
import json, argparse, datetime, os, re
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

META_COLS = [
    'Hidden_Test_Case_Name',
    'Hidden_Test_Description',
    'Hidden_Remarks',
    'Hidden_Test_Steps_Procedure',
    'Hidden_Impacted_Registers',
    'Hidden_Validation_Acceptance_Criteria',
]

MAIN_ORDER = [
    'Index',
    'SS / Module',
    'Feature',
    'Test Case Name',
    'Test Description',
    'Speed',
    'Mode',
    'Memory Start Offset',
    'Memory End Offset',
    'Remarks',
    'Test Steps / Procedure',
    'Impacted Registers',
    'Validation / Acceptance Criteria',
    'Code Generation (Required / Not)',
]

TEXT_WRAP_COLS = [
    'Test Description', 'Remarks', 'Test Steps / Procedure', 'Validation / Acceptance Criteria'
]

ALLOWED_DV = ['Required', 'Blank', 'Not Required']

# Helpers

def normalize_schema(records):
    # ordered union of keys (first appearance)
    seen = []
    for rec in records:
        for k in rec.keys():
            if k not in seen:
                seen.append(k)
    # fill blanks
    norm = []
    for rec in records:
        row = {k: ('' if rec.get(k) is None else rec.get(k)) for k in seen}
        norm.append(row)
    return seen, norm

def auto_width(ws):
    col_widths = {}
    for r in ws.iter_rows(values_only=True):
        for idx, v in enumerate(r, start=1):
            val = '' if v is None else str(v)
            col_widths[idx] = max(col_widths.get(idx, 0), len(val))
    for idx, w in col_widths.items():
        ws.column_dimensions[chr(64+idx) if idx<=26 else _colname(idx)].width = min(max(w+2, 12), 80)

def _colname(n):
    s = ''
    while n:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

def set_borders(ws):
    thin = Side(style='thin', color='000000')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = bd

def renumber_block(text: str) -> str:
    if not text:
        return ''
    # split by lines, remove existing leading numbering/bullets, re-number 1., 2., ...
    lines = [l for l in str(text).splitlines() if l.strip()]
    out = []
    for i, l in enumerate(lines, start=1):
        cleaned = re.sub(r'^\s*(?:\d+[\.)]|[-•])\s*', '', l.strip())
        out.append(f"{i}. {cleaned}")
    return "\n".join(out) if out else ''

def make_very_hidden(ws):
    ws.sheet_state = 'veryHidden'


def build_workbook(data):
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError('json_data must be a non-empty array of objects')

    keys, rows = normalize_schema(data)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # header
    for c, k in enumerate(keys, start=1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = Font(bold=True)
    # rows
    for r, rec in enumerate(rows, start=2):
        for c, k in enumerate(keys, start=1):
            ws.cell(row=r, column=c, value=rec.get(k, ''))
    # freeze
    ws.freeze_panes = 'A2'

    # Meta_data_sheet
    meta = wb.create_sheet('Meta_data_sheet')
    # write meta headers
    for c, k in enumerate(META_COLS, start=1):
        meta.cell(row=1, column=c, value=k).font = Font(bold=True)
    # transfer meta values from Data (by column name lookup)
    header_map = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column+1)}
    for r in range(2, ws.max_row+1):
        for c, k in enumerate(META_COLS, start=1):
            src_col = header_map.get(k)
            val = ws.cell(row=r, column=src_col).value if src_col else ''
            meta.cell(row=r, column=c, value=val)

    # hide meta
    make_very_hidden(meta)

    # Transform Data -> TestPlan on same sheet
    # Remove META columns and reorder to MAIN_ORDER
    # Build current header order
    curr_header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
    # Determine columns to keep in final main
    final_cols = MAIN_ORDER
    # Build new table in a temp list
    table = []
    table.append(final_cols)
    col_index = {h: (curr_header.index(h)+1 if h in curr_header else None) for h in final_cols}
    for r in range(2, ws.max_row+1):
        row_vals = []
        for h in final_cols:
            ci = col_index.get(h)
            row_vals.append(ws.cell(row=r, column=ci).value if ci else '')
        table.append(row_vals)

    # Clear sheet and write back as TestPlan
    for row in ws[1:ws.max_row]:
        for cell in row:
            cell.value = None
    ws.delete_rows(1, ws.max_row)

    # write headers
    for c, h in enumerate(table[0], start=1):
        ws.cell(row=1, column=c, value=h)
    # write data
    for r, vals in enumerate(table[1:], start=2):
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)

    # Rename sheet to TestPlan
    ws.title = 'TestPlan'

    # Formatting
    # Header style
    header_fill = PatternFill('solid', fgColor='4472C4')  # blue
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align

    # Data alignment
    top_left = Alignment(vertical='top', horizontal='left', wrap_text=False)
    top_center = Alignment(vertical='top', horizontal='center', wrap_text=False)

    # Wrap text for specific columns
    wrap_cols_idx = {}
    for c in range(1, ws.max_column+1):
        h = ws.cell(row=1, column=c).value
        if h in TEXT_WRAP_COLS:
            wrap_cols_idx[c] = True
    for r in range(2, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            if c in wrap_cols_idx:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            else:
                # Index numeric center, others left
                if ws.cell(row=1, column=c).value == 'Index':
                    cell.alignment = top_center
                else:
                    cell.alignment = top_left

    # Numbering inside cells for two columns (without touching meta sheet)
    def col_by_name(name):
        for c in range(1, ws.max_column+1):
            if ws.cell(row=1, column=c).value == name:
                return c
        return None

    steps_c = col_by_name('Test Steps / Procedure')
    crit_c = col_by_name('Validation / Acceptance Criteria')
    if steps_c:
        for r in range(2, ws.max_row+1):
            val = ws.cell(row=r, column=steps_c).value
            ws.cell(row=r, column=steps_c, value=renumber_block(val))
    if crit_c:
        for r in range(2, ws.max_row+1):
            val = ws.cell(row=r, column=crit_c).value
            ws.cell(row=r, column=crit_c, value=renumber_block(val))

    # Data validation on Code Generation (Required / Not) for data rows only
    code_c = col_by_name('Code Generation (Required / Not)')
    if code_c:
        dv = DataValidation(type='list', formula1='"' + ','.join(ALLOWED_DV) + '"', allow_blank=True, showDropDown=True)
        ws.add_data_validation(dv)
        dv.add(f"{_colname(code_c)}2:{_colname(code_c)}{ws.max_row}")

    # Borders
    set_borders(ws)

    # Auto width and row height
    auto_width(ws)
    # approximate row height adjustment for wrapped columns
    for r in range(2, ws.max_row+1):
        max_lines = 1
        for c in wrap_cols_idx.keys():
            txt = ws.cell(row=r, column=c).value
            if txt:
                lines = str(txt).count('\n') + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[r].height = min(15 * max_lines, 300)

    # Safety check: ensure no sheet named 'Data' remains
    if 'Data' in [s.title for s in wb.worksheets]:
        # try to delete it
        for s in wb.worksheets:
            if s.title == 'Data':
                wb.remove(s)
    if 'Data' in [s.title for s in wb.worksheets]:
        raise RuntimeError("Data sheet still exists after normalization")

    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--json', required=True)
    ap.add_argument('--output-dir', required=True)
    ap.add_argument('--ip-name', required=True)
    args = ap.parse_args()

    with open(args.json, 'r', encoding='utf-8') as f:
        data = json.load(f)
    wb = build_workbook(data)

    # Naming with IST timezone
    tz = ZoneInfo('Asia/Kolkata')
    now = datetime.datetime.now(tz)
    stamp = now.strftime('%Y%m%d_%H%M%S')
    fname = f"{args.ip-name if False else ''}"
    # Avoid hyphen in variable; build string explicitly
    ipname = args.ip_name
    out_name = f"{ipname}_TestPlan_{stamp}.xlsx"

    out_dir = args.output_dir.rstrip('/')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, out_name)
    wb.save(out_path)
    print(out_path)

if __name__ == '__main__':
    main()
