#!/usr/bin/env python3
import json
import os
from datetime import datetime

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    def get_ist_now():
        return datetime.now(ZoneInfo("Asia/Kolkata"))
except Exception:  # fallback to pytz if zoneinfo unavailable
    import pytz
    def get_ist_now():
        tz = pytz.timezone("Asia/Kolkata")
        return datetime.now(tz)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---- Input bindings (do not modify) ----
IP_NAME = "GPIO"
OUTPUT_DIR = os.path.join("Test_Output", "GPIO", "TestPlan")

final_json_str = r'''[
  {
    "Index": "1",
    "SS / Module": "GPIO",
    "Feature": "NA",
    "Test Case Name": "gpio_reg_wr_rd_test",
    "Test Description": "Validates GP0 GPIO registers (gpio_8 to gpio_27) for default values and masked write/read behavior. The test reads each register’s default value (ignoring bit 0), then for multiple data patterns performs masked writes to writable registers and verifies masked readbacks match expected values derived from the write mask, read mask, and default values. Registers marked non-readable or non-writable by masks are skipped accordingly. Final result is PASS only if no default or write-read mismatches occur.",
    "Meta Test Description": "The testcase executes test_case() which calls chk_rst_val() followed by chk_rd_wr(), and then reports overall status via finish(1) on any failure or finish(0) if all checks pass. chk_rst_val() loops i=0..(CNT-1), fetching addr=addr_array[i]. If skip_rst_array[i]==1, the iteration is skipped. If read_mask_array[i]==0x00000000, the iteration is skipped. Otherwise, data_rd=read_reg(addr); data=(data_rd & 0xfffffffe); compare data against default_value_array[i]. On mismatch, def_fail_cnt++ and a failure message is printed. chk_rd_wr() defines unsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}. For each pattern j, data_wr=chk_val[j], it first writes: for i=0..(CNT-1), addr=addr_array[i]; if skip_array[i]==1 continue; if write_mask_array[i]==0x00000000 continue; else write_reg(addr,(data_wr & write_mask_array[i])). Then it reads/validates: for i=0..(CNT-1), addr=addr_array[i]; if skip_array[i]==1 continue; if write_mask_array[i]==0x00000000 continue; if read_mask_array[i]==0x00000000 continue; else data_rd=(read_reg(addr) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if(data_rd != exp_val) wr_fail_cnt++ and print failure else print pass under DEBUG_DISPLAY. After all, if(def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) else finish(0). A soft_reset_chk() exists under #ifdef 0 using SOFT_RST_REG_ADDRESS and SOFT_RST_REG_DATA with wait_on() calls but is compiled out.",
    "Speed": "NA",
    "Mode": "Polling",
    "Memory Start Offset": "NA",
    "Memory End Offset": "NA",
    "Remarks": "Default-value reads ignore bit 0 during comparison. Registers with zero read or write masks are skipped accordingly. Ensure stable input levels during default checks; if inputs are not driven, observed values may float and cause mismatches. The test iterates over indices defined by CNT and arrays; only entries present and marked by masks are effectively validated.",
    "Test Steps / Procedure": "1. For each of gp0_gpio_8 to gp0_gpio_27 that is marked readable, read the register and verify the default value matches the expected value while ignoring bit 0. 2. For each of the six data patterns, write the pattern masked by each register’s write mask to all writable registers in the gp0_gpio_8 to gp0_gpio_27 range. 3. After each write pass, read back each readable and writable register masked by its read mask and verify the value matches the expected combination of written bits and preserved default bits. 4. Skip any register that is not readable or not writable per its masks, or is listed in the skip lists. 5. Declare PASS only if no default-value or write-read mismatches are detected across all patterns and registers; otherwise declare FAIL.",
    "Meta Test Steps / Procedure": "Entry: test_case(). Step 1: Call chk_rst_val(). In chk_rst_val(): for (i=0; i<CNT; i++): addr=addr_array[i]; if (skip_rst_array[i]==1) continue; if (read_mask_array[i]==0x00000000) continue; data_rd=read_reg(addr); data=(data_rd & 0xfffffffe); if (data==default_value_array[i]) pass else {def_fail_cnt++; printf failure}. Return to test_case(). Step 2: Call chk_rd_wr(). In chk_rd_wr(): define unsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}; for (j=0; j<6; j++): data_wr=chk_val[j]; Write phase: for (i=0; i<CNT; i++): addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0x00000000) continue; write_reg(addr, (data_wr & write_mask_array[i])); Read/verify phase: for (i=0; i<CNT; i++): addr=addr_array[i]; if (skip_array[i]==1) continue; if (write_mask_array[i]==0x00000000) continue; if (read_mask_array[i]==0x00000000) continue; data_rd=(read_reg(addr) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if data_rd != exp_val, increment wr_fail_cnt and mark failure. Final decision: if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) [FAIL], else finish(0) [PASS].",
    "Validation / Acceptance Criteria": "PASS: For all gp0_gpio_8 to gp0_gpio_27 registers, default-value reads (with bit 0 ignored) equal the expected defaults, and for each of the six data patterns, masked write-read checks match expected values derived from read/write masks and default values. FAIL: Any mismatch in default-value comparison or in masked write-read verification for any tested register.",
    "Meta Validation / Acceptance Criteria": "Default check: For each i where read_mask_array[i] != 0x00000000 and skip_rst_array[i] == 0, compute data=(read_reg(addr_array[i]) & 0xfffffffe); compare against default_value_array[i]. Any inequality increments def_fail_cnt and results in FAIL at test end. Write-read check: For each pattern data_wr in {0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000}, for each i where skip_array[i]==0, write_mask_array[i] != 0x00000000, perform write_reg(addr_array[i], (data_wr & write_mask_array[i])); then if read_mask_array[i] != 0x00000000, compute data_rd=(read_reg(addr_array[i]) & read_mask_array[i]); wr_n=(write_mask_array[i] ^ 0xffffffff); exp_val=((data_wr & read_mask_array[i] & write_mask_array[i]) | (wr_n & read_mask_array[i] & default_value_array[i])); if data_rd != exp_val, increment wr_fail_cnt and mark failure. Final decision: if (def_fail_cnt > 0 || wr_fail_cnt > 0) finish(1) [FAIL], else finish(0) [PASS].",
    "Code Generation (Required / Not)": "Not",
    "Meta Headers": "#include <stdio.h>; #include <stdlib.h>; #include \"test_common.h\"; #include \"test_define.c\"; #include<gpio/gpio_def.h>; #include<gpio/gpio_offset.h>",
    "Meta Macros": "#define SOFT_RST_REG_ADDRESS\t0x00000000; #define SOFT_RST_REG_DATA\t0x00000000; #define CNT 49",
    "Meta Arrays": "const unsigned long int addr_array[20]={MIZAR_GPIO_GP0_GPIO_8,MIZAR_GPIO_GP0_GPIO_9,MIZAR_GPIO_GP0_GPIO_10,MIZAR_GPIO_GP0_GPIO_11,MIZAR_GPIO_GP0_GPIO_12,MIZAR_GPIO_GP0_GPIO_13,MIZAR_GPIO_GP0_GPIO_14,MIZAR_GPIO_GP0_GPIO_15,MIZAR_GPIO_GP0_GPIO_16,MIZAR_GPIO_GP0_GPIO_17,MIZAR_GPIO_GP0_GPIO_18,MIZAR_GPIO_GP0_GPIO_19,MIZAR_GPIO_GP0_GPIO_20,MIZAR_GPIO_GP0_GPIO_21,MIZAR_GPIO_GP0_GPIO_22,MIZAR_GPIO_GP0_GPIO_23,MIZAR_GPIO_GP0_GPIO_24,MIZAR_GPIO_GP0_GPIO_25,MIZAR_GPIO_GP0_GPIO_26,MIZAR_GPIO_GP0_GPIO_27,}; const unsigned int default_value_array[20]={GPIO_GP0_GPIO_8_DEFAULT_VAL,GPIO_GP0_GPIO_9_DEFAULT_VAL,GPIO_GP0_GPIO_10_DEFAULT_VAL,GPIO_GP0_GPIO_11_DEFAULT_VAL,GPIO_GP0_GPIO_12_DEFAULT_VAL,GPIO_GP0_GPIO_13_DEFAULT_VAL,GPIO_GP0_GPIO_14_DEFAULT_VAL,GPIO_GP0_GPIO_15_DEFAULT_VAL,GPIO_GP0_GPIO_16_DEFAULT_VAL,GPIO_GP0_GPIO_17_DEFAULT_VAL,GPIO_GP0_GPIO_18_DEFAULT_VAL,GPIO_GP0_GPIO_19_DEFAULT_VAL,GPIO_GP0_GPIO_20_DEFAULT_VAL,GPIO_GP0_GPIO_21_DEFAULT_VAL,GPIO_GP0_GPIO_22_DEFAULT_VAL,GPIO_GP0_GPIO_23_DEFAULT_VAL,GPIO_GP0_GPIO_24_DEFAULT_VAL,GPIO_GP0_GPIO_25_DEFAULT_VAL,GPIO_GP0_GPIO_26_DEFAULT_VAL,GPIO_GP0_GPIO_27_DEFAULT_VAL,}; const unsigned int read_mask_array[20]={GPIO_GP0_GPIO_8_READ_MASK,GPIO_GP0_GPIO_9_READ_MASK,GPIO_GP0_GPIO_10_READ_MASK,GPIO_GP0_GPIO_11_READ_MASK,GPIO_GP0_GPIO_12_READ_MASK,GPIO_GP0_GPIO_13_READ_MASK,GPIO_GP0_GPIO_14_READ_MASK,GPIO_GP0_GPIO_15_READ_MASK,GPIO_GP0_GPIO_16_READ_MASK,GPIO_GP0_GPIO_17_READ_MASK,GPIO_GP0_GPIO_18_READ_MASK,GPIO_GP0_GPIO_19_READ_MASK,GPIO_GP0_GPIO_20_READ_MASK,GPIO_GP0_GPIO_21_READ_MASK,GPIO_GP0_GPIO_22_READ_MASK,GPIO_GP0_GPIO_23_READ_MASK,GPIO_GP0_GPIO_24_READ_MASK,GPIO_GP0_GPIO_25_READ_MASK,GPIO_GP0_GPIO_26_READ_MASK,GPIO_GP0_GPIO_27_READ_MASK,}; const unsigned int write_mask_array[20]={GPIO_GP0_GPIO_8_WRITE_MASK,GPIO_GP0_GPIO_9_WRITE_MASK,GPIO_GP0_GPIO_10_WRITE_MASK,GPIO_GP0_GPIO_11_WRITE_MASK,GPIO_GP0_GPIO_12_WRITE_MASK,GPIO_GP0_GPIO_13_WRITE_MASK,GPIO_GP0_GPIO_14_WRITE_MASK,GPIO_GP0_GPIO_15_WRITE_MASK,GPIO_GP0_GPIO_16_WRITE_MASK,GPIO_GP0_GPIO_17_WRITE_MASK,GPIO_GP0_GPIO_18_WRITE_MASK,GPIO_GP0_GPIO_19_WRITE_MASK,GPIO_GP0_GPIO_20_WRITE_MASK,GPIO_GP0_GPIO_21_WRITE_MASK,GPIO_GP0_GPIO_22_WRITE_MASK,GPIO_GP0_GPIO_23_WRITE_MASK,GPIO_GP0_GPIO_24_WRITE_MASK,GPIO_GP0_GPIO_25_WRITE_MASK,GPIO_GP0_GPIO_26_WRITE_MASK,GPIO_GP0_GPIO_27_WRITE_MASK,}; const unsigned int skip_array[20]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,}; const unsigned int skip_rst_array[20]={0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,}; unsigned int chk_val[6]={0xffffffff,0xaaaaaaaa,0x55555555,0xf5f5f5f5,0xA5A5A5A5,0xffff0000};",
    "Meta Impacted Registers": "SOFT_RST_REG_ADDRESS; MIZAR_GPIO_GP0_GPIO_8; MIZAR_GPIO_GP0_GPIO_9; MIZAR_GPIO_GP0_GPIO_10; MIZAR_GPIO_GP0_GPIO_11; MIZAR_GPIO_GP0_GPIO_12; MIZAR_GPIO_GP0_GPIO_13; MIZAR_GPIO_GP0_GPIO_14; MIZAR_GPIO_GP0_GPIO_15; MIZAR_GPIO_GP0_GPIO_16; MIZAR_GPIO_GP0_GPIO_17; MIZAR_GPIO_GP0_GPIO_18; MIZAR_GPIO_GP0_GPIO_19; MIZAR_GPIO_GP0_GPIO_20; MIZAR_GPIO_GP0_GPIO_21; MIZAR_GPIO_GP0_GPIO_22; MIZAR_GPIO_GP0_GPIO_23; MIZAR_GPIO_GP0_GPIO_24; MIZAR_GPIO_GP0_GPIO_25; MIZAR_GPIO_GP0_GPIO_26; MIZAR_GPIO_GP0_GPIO_27",
    "Impacted Registers": "gp0_gpio_8; gp0_gpio_9; gp0_gpio_10; gp0_gpio_11; gp0_gpio_12; gp0_gpio_13; gp0_gpio_14; gp0_gpio_15; gp0_gpio_16; gp0_gpio_17; gp0_gpio_18; gp0_gpio_19; gp0_gpio_20; gp0_gpio_21; gp0_gpio_22; gp0_gpio_23; gp0_gpio_24; gp0_gpio_25; gp0_gpio_26; gp0_gpio_27"
  }
]'''

# Parse JSON exactly
records = json.loads(final_json_str)

# Column definitions and order
TESTPLAN_COLS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
    "Impacted Registers",
]

METADATA_COLS = [
    "Meta Test Description",
    "Meta Test Steps / Procedure",
    "Meta Headers",
    "Meta Macros",
    "Meta Arrays",
    "Meta Impacted Registers",
    "Meta Validation / Acceptance Criteria",
]

# Prepare workbook
wb = Workbook()
ws = wb.active
ws.title = "TestPlan"
wsm = wb.create_sheet("MetaData")
ws.freeze_panes = "A2"
wsm.freeze_panes = "A2"
# Very hidden metadata sheet
wsm.sheet_state = "veryHidden"

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(fill_type="solid", start_color="FFDDEBF7", end_color="FFDDEBF7")
wrap = Alignment(wrap_text=True, vertical="top")

# Write headers
for col_idx, key in enumerate(TESTPLAN_COLS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=key)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

for col_idx, key in enumerate(METADATA_COLS, start=1):
    cell = wsm.cell(row=1, column=col_idx, value=key)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap

# Write data rows (preserve order)
for r_idx, rec in enumerate(records, start=2):
    for c_idx, key in enumerate(TESTPLAN_COLS, start=1):
        val = rec.get(key, "")
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.alignment = wrap
    for c_idx, key in enumerate(METADATA_COLS, start=1):
        val = rec.get(key, "")
        cell = wsm.cell(row=r_idx, column=c_idx, value=val)
        cell.alignment = wrap

# Column widths (reasonable defaults)
col_widths_testplan = {
    "Index": 8,
    "SS / Module": 16,
    "Feature": 10,
    "Test Case Name": 28,
    "Test Description": 80,
    "Speed": 10,
    "Mode": 12,
    "Memory Start Offset": 18,
    "Memory End Offset": 18,
    "Remarks": 60,
    "Test Steps / Procedure": 90,
    "Validation / Acceptance Criteria": 80,
    "Code Generation (Required / Not)": 24,
    "Impacted Registers": 40,
}

for idx, key in enumerate(TESTPLAN_COLS, start=1):
    ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = col_widths_testplan.get(key, 20)

col_widths_meta = {
    "Meta Test Description": 90,
    "Meta Test Steps / Procedure": 100,
    "Meta Headers": 60,
    "Meta Macros": 60,
    "Meta Arrays": 110,
    "Meta Impacted Registers": 60,
    "Meta Validation / Acceptance Criteria": 100,
}

for idx, key in enumerate(METADATA_COLS, start=1):
    wsm.column_dimensions[wsm.cell(row=1, column=idx).column_letter].width = col_widths_meta.get(key, 40)

# Build IST timestamp and filename
now_ist = get_ist_now()
ts = now_ist.strftime('%Y%m%d_%H%M%S')
filename = f"{IP_NAME}_TestPlan_{ts}.xlsx"
output_path = os.path.join(OUTPUT_DIR, filename)

# Ensure directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Save workbook
wb.save(output_path)

# Emit sidecar files for workflow
with open('generated_path.txt', 'w', encoding='utf-8') as f:
    f.write(output_path)
with open('generated_timestamp.txt', 'w', encoding='utf-8') as f:
    f.write(now_ist.strftime('%Y-%m-%d %H:%M:%S %Z'))

print(f"Generated: {output_path}")

# no-op change to trigger workflow
