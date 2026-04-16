#!/usr/bin/env python3
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Config
INPUT = Path('data/input.json')
OUTPUT = Path('TestRepo/gpio/json_testing2.xlsx')
SHEET_NAME = 'Data'


def fail(msg):
    print(msg, file=sys.stderr)
    sys.exit(1)


def load_json():
    if not INPUT.exists():
        fail(f"Input JSON not found at {INPUT}")
    try:
        with INPUT.open('r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        fail(f"Invalid JSON: {e}")
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list) or not data:
        fail("JSON must be a non-empty array of objects or a single object")
    for i, row in enumerate(data):
        if not isinstance(row, dict):
            fail(f"Row {i} is not an object")
    return data


def build_headers(rows):
    seen = []
    seen_set = set()
    for row in rows:
        for k in row.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return seen


def write_xlsx(rows, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Rows
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-fit approx by measuring max length in characters; set width accordingly
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in range(2, len(rows) + 2):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                l = 0
            else:
                l = len(str(v))
            if l > max_len:
                max_len = l
        adj = min(120, max_len + 2)  # clamp to avoid excessively wide columns
        ws.column_dimensions[get_column_letter(col_idx)].width = adj

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


if __name__ == '__main__':
    rows = load_json()
    headers = build_headers(rows)
    write_xlsx(rows, headers)
    print(f"Wrote {OUTPUT} with {len(rows)} rows and {len(headers)} columns")

# no-op to retrigger workflow
