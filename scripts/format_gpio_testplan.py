#!/usr/bin/env python3

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import sys

# Parameters from the request
INPUT_PATH = "Test_Output/GPIO/TestPlan/GPIO_TestPlan_2.xlsx"
OUTPUT_PATH = "Test_Output/GPIO/TestPlan/GPIO_TestPlan_2_hidden.xlsx"

# Column definitions (exact names required)
META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

# Fail fast helper
def fail(msg: str):
    print(f"ERROR: {msg}")
    sys.exit(1)

# Load workbook
try:
    wb = load_workbook(INPUT_PATH)
except FileNotFoundError:
    fail(f"Input file not found at {INPUT_PATH}")
except Exception as e:
    fail(f"Failed to open workbook: {e}")

# Identify primary visible sheet
visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
if not visible_sheets:
    fail("No visible worksheet found in workbook")
ws = visible_sheets[0]

# Rename main sheet to TestPlan
ws.title = "TestPlan"

# Build header index map for exact header names (row 1)
header_row = 1
header_to_idx = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=header_row, column=c).value
    if v is not None:
        header_to_idx[str(v)] = c

# Validate META columns exist
missing_meta = [h for h in META_COLUMNS if h not in header_to_idx]
if missing_meta:
    fail("Missing required META headers: " + ", ".join(missing_meta))

# Create META sheet and copy META columns preserving values exactly
if "Meta_data_sheet" in [s.title for s in wb.worksheets]:
    # Remove existing to ensure a clean state
    meta_ws = wb["Meta_data_sheet"]
    wb.remove(meta_ws)
meta_ws = wb.create_sheet("Meta_data_sheet")

for j, h in enumerate(META_COLUMNS, start=1):
    src_col = header_to_idx[h]
    for r in range(1, ws.max_row + 1):
        meta_ws.cell(row=r, column=j).value = ws.cell(row=r, column=src_col).value

# Very-hide META sheet
meta_ws.sheet_state = "veryHidden"

# Remove META columns from TestPlan (delete from rightmost to leftmost)
meta_indices = sorted([header_to_idx[h] for h in META_COLUMNS], reverse=True)
for idx in meta_indices:
    ws.delete_cols(idx, 1)

# Rebuild header map after deletions
header_to_idx = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=header_row, column=c).value
    if v is not None:
        header_to_idx[str(v)] = c

# Validate MAIN columns presence
missing_main = [h for h in MAIN_COLUMNS if h not in header_to_idx]
if missing_main:
    fail("Missing required MAIN headers: " + ", ".join(missing_main))

# Remove extra columns not in MAIN
current_headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
extra_headers = [h for h in current_headers if h not in MAIN_COLUMNS and h is not None]
extra_indices = sorted([header_to_idx[h] for h in extra_headers if h in header_to_idx], reverse=True)
for idx in extra_indices:
    ws.delete_cols(idx, 1)

# Rebuild header map again
header_to_idx = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=header_row, column=c).value
    if v is not None:
        header_to_idx[str(v)] = c

# Reorder columns to match MAIN_COLUMNS exactly by creating a temp sheet
tmp = wb.create_sheet("TMP_TestPlan")
for j, h in enumerate(MAIN_COLUMNS, start=1):
    col_idx = header_to_idx[h]
    for r in range(1, ws.max_row + 1):
        tmp.cell(row=r, column=j).value = ws.cell(row=r, column=col_idx).value

# Delete old TestPlan and rename temp
wb.remove(ws)
tmp.title = "TestPlan"
ws = tmp

# -------- Formatting (TestPlan only) --------
# Header styles
header_font = Font(bold=True)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
header_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")

# Data alignments
align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=False)
align_center_top = Alignment(horizontal="center", vertical="top", wrap_text=False)
align_right_top = Alignment(horizontal="right", vertical="top", wrap_text=False)

# Wrap text for specified columns
WRAP_COLUMNS = {
    "Test Description",
    "Remarks",
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
}

# Thin border for all populated cells
thin = Side(style="thin", color="000000")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

max_row = ws.max_row
max_col = ws.max_column

# Apply header formatting and borders
for c in range(1, max_col + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = header_font
    cell.alignment = header_align
    cell.fill = header_fill
    cell.border = border_all

# Determine index of each main column quickly
col_name_to_idx = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

# Apply data row alignment and borders; enable wrap text for WRAP_COLUMNS
for r in range(2, max_row + 1):
    for h in MAIN_COLUMNS:
        c = col_name_to_idx[h]
        cell = ws.cell(row=r, column=c)
        # Set borders
        cell.border = border_all
        # Alignment per column type
        if h in {"Index"}:
            cell.alignment = align_center_top
        elif h in {"Speed", "Mode", "Code Generation (Required / Not)"}:
            cell.alignment = align_center_top
        elif h in {"Memory Start Offset", "Memory End Offset"}:
            cell.alignment = align_right_top
        else:
            cell.alignment = align_left_top
        # Wrap for specific columns
        if h in WRAP_COLUMNS:
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=True)

# Approximate autofit: set column widths based on max text length in each column
for j in range(1, max_col + 1):
    max_len = 0
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=j).value
        if v is None:
            continue
        s = str(v)
        if "\n" in s:
            parts = s.split("\n")
            max_len = max(max_len, max(len(p) for p in parts))
        else:
            max_len = max(max_len, len(s))
    # Add padding and cap width
    width = min(100, max(10, int(max_len * 1.2) + 2))
    ws.column_dimensions[get_column_letter(j)].width = width

# Approximate row height after wrapping: scale by longest wrapped column
base_height = 15  # typical Excel default
for r in range(2, max_row + 1):
    max_lines = 1
    for name in WRAP_COLUMNS:
        c = col_name_to_idx.get(name)
        if not c:
            continue
        v = ws.cell(row=r, column=c).value
        if v is None:
            continue
        s = str(v)
        width = ws.column_dimensions[get_column_letter(c)].width or 10
        chars_per_line = max(1, int(width))
        est_lines = 1
        if s:
            est_lines = max(1, (len(s) // chars_per_line) + (1 if len(s) % chars_per_line else 0))
            if "\n" in s:
                est_lines = max(est_lines, s.count("\n") + 1)
        if est_lines > max_lines:
            max_lines = est_lines
    ws.row_dimensions[r].height = base_height * max_lines

# Save to output path
wb.save(OUTPUT_PATH)
print(f"Saved formatted workbook to {OUTPUT_PATH}")
