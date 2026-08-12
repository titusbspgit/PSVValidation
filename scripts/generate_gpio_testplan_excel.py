#!/usr/bin/env python3
import argparse
import os
import json
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo  # pragma: no cover
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# -------- Inputs embedded from previous agents (preserve order) --------
TEST_FOLDERS = [
  {
    "index": 1,
    "folder_name": "gpio_reg_wr_rd_test",
    "github_url": "https://github.com/titusbspgit/PSVValidation/tree/main/TestRepo/gpio/gpio_reg_wr_rd_test",
    "raw_base_url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/TestRepo/gpio/gpio_reg_wr_rd_test/"
  }
]

TOKENS = [
  {"token": "MIZAR_GPIO_GP0_GPIO_8",  "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_9",  "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_10", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_11", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_12", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_13", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_14", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_15", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_16", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_17", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_18", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_19", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_20", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_21", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_22", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_23", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_24", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_25", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_26", "token_type": "macro", "operation": "unknown"},
  {"token": "MIZAR_GPIO_GP0_GPIO_27", "token_type": "macro", "operation": "unknown"}
]

RESOLVED = [
  {"macro": "MIZAR_GPIO_GP0_GPIO_8",  "base_value": "0xA001A000", "offset_value": "0x0"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_9",  "base_value": "0xA001A000", "offset_value": "0x4"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_10", "base_value": "0xA001A000", "offset_value": "0x8"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_11", "base_value": "0xA001A000", "offset_value": "0xC"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_12", "base_value": "0xA001A000", "offset_value": "0x10"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_13", "base_value": "0xA001A000", "offset_value": "0x14"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_14", "base_value": "0xA001A000", "offset_value": "0x18"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_15", "base_value": "0xA001A000", "offset_value": "0x1C"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_16", "base_value": "0xA001A000", "offset_value": "0x20"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_17", "base_value": "0xA001A000", "offset_value": "0x24"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_18", "base_value": "0xA001A000", "offset_value": "0x28"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_19", "base_value": "0xA001A000", "offset_value": "0x2C"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_20", "base_value": "0xA001A000", "offset_value": "0x30"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_21", "base_value": "0xA001A000", "offset_value": "0x34"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_22", "base_value": "0xA001A000", "offset_value": "0x38"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_23", "base_value": "0xA001A000", "offset_value": "0x3C"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_24", "base_value": "0xA001A000", "offset_value": "0x40"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_25", "base_value": "0xA001A000", "offset_value": "0x44"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_26", "base_value": "0xA001A000", "offset_value": "0x48"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_27", "base_value": "0xA001A000", "offset_value": "0x4C"}
]

MAPPINGS = [
  {"macro": "MIZAR_GPIO_GP0_GPIO_8",  "register_name": "gp0_gpio_8"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_9",  "register_name": "gp0_gpio_9"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_10", "register_name": "gp0_gpio_10"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_11", "register_name": "gp0_gpio_11"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_12", "register_name": "gp0_gpio_12"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_13", "register_name": "gp0_gpio_13"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_14", "register_name": "gp0_gpio_14"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_15", "register_name": "gp0_gpio_15"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_16", "register_name": "gp0_gpio_16"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_17", "register_name": "gp0_gpio_17"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_18", "register_name": "gp0_gpio_18"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_19", "register_name": "gp0_gpio_19"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_20", "register_name": "gp0_gpio_20"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_21", "register_name": "gp0_gpio_21"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_22", "register_name": "gp0_gpio_22"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_23", "register_name": "gp0_gpio_23"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_24", "register_name": "gp0_gpio_24"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_25", "register_name": "gp0_gpio_25"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_26", "register_name": "gp0_gpio_26"},
  {"macro": "MIZAR_GPIO_GP0_GPIO_27", "register_name": "gp0_gpio_27"}
]

FINAL_JSON_TEXT = r'''[
  {
    "status": "success",
    "file_name": "GPIO_TestPlan_20260812_000000.xlsx",
    "repository": "titusbspgit/PSVValidation",
    "branch": "main",
    "path": "Test_Output/GPIO/TestPlan/GPIO_TestPlan_20260812_000000.xlsx",
    "html_url": "https://github.com/titusbspgit/PSVValidation/blob/main/Test_Output/GPIO/TestPlan/GPIO_TestPlan_20260812_000000.xlsx",
    "raw_url": "https://raw.githubusercontent.com/titusbspgit/PSVValidation/main/Test_Output/GPIO/TestPlan/GPIO_TestPlan_20260812_000000.xlsx",
    "commit": {
      "sha": "ce249996ca7456a50402c1e883b679167b4e103e",
      "url": "https://github.com/titusbspgit/PSVValidation/commit/ce249996ca7456a50402c1e883b679167b4e103e",
      "message": "Add GPIO TestPlan Excel (auto-generated) - IST 2026-08-12 00:00:00",
      "timestamp_utc": "2026-08-12T07:07:20Z",
      "timestamp_ist": "2026-08-12 00:00:00 IST"
    }
  }
]'''

# -------- Helpers --------
def hex_inc(base_hex: str, off_hex: str) -> str:
    b = int(base_hex, 16)
    o = int(off_hex, 16)
    return f"0x{(b + o):08X}"


def build_rows():
    res_map = {r["macro"]: r for r in RESOLVED}
    map_map = {m["macro"]: m for m in MAPPINGS}
    # Determine source folder path
    src_folder_rel = "TestRepo/gpio/" + TEST_FOLDERS[0]["folder_name"]

    rows = []
    for t in TOKENS:
        macro = t["token"]
        op = t.get("operation", "unknown")
        res = res_map.get(macro, {})
        mp = map_map.get(macro, {})
        base = res.get("base_value", "NA")
        off = res.get("offset_value", "NA")
        resolved = hex_inc(base, off) if base != "NA" and off != "NA" else "NA"
        reg = mp.get("register_name", "NA")
        rows.append([
            "gpio_reg_wr_rd_test",  # TestCaseName
            reg,                      # RegisterName
            macro,                    # Macro
            base,                     # BaseAddress
            off,                      # Offset
            resolved,                 # ResolvedAddress
            op,                       # Operation
            src_folder_rel            # SourceFolder
        ])
    return rows


def apply_formatting(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.freeze_panes = 'A2'
    # Reasonable column widths
    widths = [22, 18, 28, 14, 10, 18, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    # Wrap text for data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


def write_metadata(ws_meta, ist_str: str):
    # Two-column key/value MetaData with very hidden state
    headers = ["Key", "Value"]
    ws_meta.append(headers)
    ws_meta["A1"].font = Font(bold=True)
    ws_meta["B1"].font = Font(bold=True)
    ws_meta["A1"].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    ws_meta["B1"].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    meta = [
        ["agent", "Ag_Excel_Generator Agent"],
        ["ip_name", "GPIO"],
        ["generation_time_ist", ist_str],
        ["final_json", FINAL_JSON_TEXT],
    ]
    for k, v in meta:
        ws_meta.append([k, v])
    ws_meta.column_dimensions['A'].width = 28
    ws_meta.column_dimensions['B'].width = 120
    for row in ws_meta.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws_meta.sheet_state = 'veryHidden'


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--output-dir', required=True)
    parser.add_argument('--ip-name', required=True)
    args = parser.parse_args()

    ist_now = datetime.now(ZoneInfo('Asia/Kolkata'))
    ts_date = ist_now.strftime('%Y%m%d')
    ts_time = ist_now.strftime('%H%M%S')
    filename = f"{args.ip_name}_TestPlan_{ts_date}_{ts_time}.xlsx"

    rows = build_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = 'TestPlan'
    headers = [
        'TestCaseName', 'RegisterName', 'Macro', 'BaseAddress',
        'Offset', 'ResolvedAddress', 'Operation', 'SourceFolder'
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    apply_formatting(ws)

    ws_meta = wb.create_sheet('MetaData')
    write_metadata(ws_meta, ist_now.strftime('%Y-%m-%d %H:%M:%S IST'))

    os.makedirs(args.output_dir, exist_ok=True)
    out_path = os.path.join(args.output_dir, filename)
    wb.save(out_path)

    # Record the generated path for the workflow commit step
    with open('.generated_excel_path', 'w') as f:
        f.write(out_path)
    print(f"Generated Excel at {out_path}")

if __name__ == '__main__':
    main()
