#!/usr/bin/env python3
# Append Batch 2 (Indices 6–9) to existing GPIO TestPlan WORKING.xlsx
# - Requires that Test_Output/GPIO/TestPlan/GPIO_TestPlan_WORKING.xlsx already exists (from Batch 1)
# - Reads existing TestPlan and Meta_data_sheet rows and appends Batch 2 JSON
# - Preserves formatting rules, meta sheet, in-cell numbering, and validation per Stage1

import os, sys, json, re, zipfile
from copy import deepcopy
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE_PATH = os.environ.get('OUTPUT_FILE_PATH', 'Test_Output/GPIO/TestPlan/GPIO_TestPlan_WORKING.xlsx')
OUTPUT_DIR = os.path.dirname(OUTPUT_FILE_PATH)

# Batch 2 JSON derived strictly from source code in TestRepo/gpio
JSON_DATA: List[Dict] = [
    {
        "Index": 6,
        "SS / Module": "GPIO",
        "Feature": "Negative edge interrupt (walking zeros)",
        "Test Case Name": "test_gpio_nedge_walking_zeros_pattern",
        "Test Description": "Configures all GPIO[0..31] for falling-edge interrupts and applies a walking-zeros pattern via 0xA0243ffc to trigger one pin at a time. ISR validates DIN indication, raw status, masked group status, clears raw, and verifies clears including system sticky.",
        "Speed": "NA",
        "Mode": "Interrupt",
        "Memory Start Offset": "0xA0243ffc",
        "Memory End Offset": "0xA0243ffc",
        "Remarks": "Per-pin control set to 0x00040000; IO_CTRL_GROUP1..4 set to 0x000000FF; group mask 0xFFFFFFFF. For each i: write 0xFFFFFFFF then ~(1<<i); wait for ISR.",
        "Test Steps / Procedure": "Enable GIC (GPIO0=87/GPIO1=88); enable LSS_SYSREG_INTR_EN1 for selected GPIO; program MIZAR_GPIO_GP0_GPIO_8+(i*4)=0x00040000 for i=0..31; set GPIO_IO_CTRL_GROUP1..4=0x000000FF; enable MIZAR_GPIO_GP0_INTR1_INTR_EN1=0xFFFFFFFF; for i=0..31 write 0xFFFFFFFF then ~(1<<i) to 0xA0243ffc and wait for ISR; in ISR read per-pin reg, check DIN bit set and raw bit set; read group status bit set; clear raw with 0x00110001; verify readback 0x100001; verify group status=0; clear sysreg RAW_STCR1 for GPIO0/GPIO1; clear IRQ.",
        "Impacted Registers": ["MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_LSS_SYSREG_RAW_STCR1"],
        "Validation / Acceptance Criteria": "Per-pin: DIN bit indicates active after falling edge; raw bit set; group status bit set; after write 0x00110001 readback equals 0x100001; group status clears to 0; system sticky cleared in RAW_STCR1.",
        "Code Generation (Required / Not)": "",
        "Hidden_Test_Case_Name": "test_gpio_nedge_walking_zeros_pattern",
        "Hidden_Test_Description": "Walking zeros falling-edge test using 0xA0243ffc with per-pin 0x00040000 config; ISR checks DIN, raw, group, clears raw with 0x00110001 and sysreg sticky.",
        "Hidden_Remarks": "int_pend gating loop; GIC_ClearIRQ for IRQ 87/88; IO_CTRL groups set to 0x000000FF per code.",
        "Hidden_Test_Steps_Procedure": "Source: program.c. See inline sequence for configuration and ISR checks.",
        "Hidden_Impacted_Registers": ["MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_LSS_SYSREG_RAW_STCR1"],
        "Hidden_Validation_Acceptance_Criteria": "DIN active, RAW set, group bit set; after 0x00110001 readback 0x100001; group clears; RAW_STCR1 cleared."
    },
    {
        "Index": 7,
        "SS / Module": "GPIO",
        "Feature": "Negative edge interrupt (all pads)",
        "Test Case Name": "test_gpio_negedge_all_pads_en",
        "Test Description": "Configures all GPIO[0..31] for falling-edge interrupts, enables group mask, and drives all pads low to trigger group interrupts. ISR validates group status nonzero, clears all per-pin raw status, verifies group clear, and clears system sticky.",
        "Speed": "NA",
        "Mode": "Interrupt",
        "Memory Start Offset": "0xA0243ffc",
        "Memory End Offset": "0xA0243ffc",
        "Remarks": "Per-pin 0x00040000; IO_CTRL_GROUP1..4 0x000000FF; group mask 0xFFFFFFFF. For each i: write 0xFFFFFFFF then 0x00000000 and wait for ISR.",
        "Test Steps / Procedure": "Enable GIC; enable LSS_SYSREG_INTR_EN1; for i=0..31 write MIZAR_GPIO_GP0_GPIO_8+(i*4)=0x00040000; set GPIO_IO_CTRL_GROUP1..4=0x000000FF; enable MIZAR_GPIO_GP0_INTR1_INTR_EN1=0xFFFFFFFF; for i=0..31 write 0xFFFFFFFF then 0x00000000 to 0xA0243ffc and wait for ISR; in ISR read MIZAR_GPIO_GP0_INTR1_INTR_STS1 (expect nonzero); disable group EN1=0; for i=0..31 write per-pin 0x00110001 to clear raw; verify group status=0; clear RAW_STCR1 for selected GPIO; clear IRQ.",
        "Impacted Registers": ["MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_LSS_SYSREG_RAW_STCR1"],
        "Validation / Acceptance Criteria": "Group status nonzero on interrupt; after clearing all per-pin raw, group status reads 0; system sticky cleared.",
        "Code Generation (Required / Not)": "",
        "Hidden_Test_Case_Name": "test_gpio_negedge_all_pads_en",
        "Hidden_Test_Description": "All pads falling-edge test; ISR checks only group status and performs bulk raw clear with 0x00110001.",
        "Hidden_Remarks": "Disables group EN1 before clearing; uses IO_CTRL groups 0x000000FF.",
        "Hidden_Test_Steps_Procedure": "Source: program.c. See ISR sequence and loop over all pads.",
        "Hidden_Impacted_Registers": ["MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_RAW_STCR1"],
        "Hidden_Validation_Acceptance_Criteria": "Group status set then cleared after raw clears; RAW_STCR1 cleared."
    },
    {
        "Index": 8,
        "SS / Module": "GPIO",
        "Feature": "Negative edge interrupt enable (per-pin)",
        "Test Case Name": "test_gpio_negedge_intr_en",
        "Test Description": "Programs each pin for input + falling-edge (0x00140000), enables its mask bit, applies falling transition via ~(1<<i) at 0xA0243ffc, and in ISR validates DIN, raw, masked group status, clears raw (0x00110001), verifies clear, clears system sticky, and clears IRQ.",
        "Speed": "NA",
        "Mode": "Interrupt",
        "Memory Start Offset": "0xA0243ffc",
        "Memory End Offset": "0xA0243ffc",
        "Remarks": "Loops i=0..31: write per-pin 0x00140000, enable corresponding bit in INTR1_INTR_EN1, drive SRAM pattern, wait for ISR.",
        "Test Steps / Procedure": "Enable GIC; enable LSS_SYSREG_INTR_EN1; write 0xFFFFFFFF to 0xA0243ffc; for i=0..31 program addr1=MIZAR_GPIO_GP0_GPIO_8+(i*4)=0x00140000; enable MIZAR_GPIO_GP0_INTR1_INTR_EN1=(1<<i); write 0xFFFFFFFF then ~(1<<i); wait ISR; in ISR read raddr per-pin and verify DIN bit set and raw bit set; read group status has (1<<i); write 0x00110001 to raddr and verify 0x100001; verify group status=0; clear RAW_STCR1 for selected GPIO and clear IRQ.",
        "Impacted Registers": ["MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_INTR_EN1", "MIZAR_LSS_SYSREG_RAW_STCR1"],
        "Validation / Acceptance Criteria": "For each i: DIN indicates active; raw bit set; group status (1<<i) set; after 0x00110001 readback 0x100001; group status cleared; sysreg sticky cleared.",
        "Code Generation (Required / Not)": "",
        "Hidden_Test_Case_Name": "test_gpio_negedge_intr_en",
        "Hidden_Test_Description": "Per-pin negedge with mask enable flow; ISR validates and clears per-pin and group state.",
        "Hidden_Remarks": "Uses addr1 and raddr2 per code; uses int_pend loop and clears via GIC_ClearIRQ.",
        "Hidden_Test_Steps_Procedure": "Source: program.c. See loop and ISR details.",
        "Hidden_Impacted_Registers": ["MIZAR_GPIO_GP0_GPIO_8", "MIZAR_GPIO_GP0_INTR1_INTR_EN1", "MIZAR_GPIO_GP0_INTR1_INTR_STS1", "MIZAR_LSS_SYSREG_RAW_STCR1"],
        "Hidden_Validation_Acceptance_Criteria": "DIN active, RAW set, group set; 0x00110001 readback 0x100001; group cleared; RAW_STCR1 cleared."
    },
    {
        "Index": 9,
        "SS / Module": "GPIO",
        "Feature": "Output mode – all pads",
        "Test Case Name": "test_gpio_op_mode_all_pad_en",
        "Test Description": "Enables output mode for GPIO[8..39] via IO control groups, toggles each pad i in 0..31 high then low by writing per-pin control (0x00200000 then 0x00000000), and verifies pad state via readback at 0xA0243ffc. Any unexpected interrupt is an error.",
        "Speed": "NA",
        "Mode": "Polling",
        "Memory Start Offset": "0xA0243ffc",
        "Memory End Offset": "0xA0243ffc",
        "Remarks": "IO_CTRL groups set to 0x00FF00FF; per-pin writes at MIZAR_GPIO_GP0_GPIO_8+(i*4). Uses helper check_for_pad_value to verify DOUT effect.",
        "Test Steps / Procedure": "Enable GIC; set IO_CTRL_GROUP1..4=0x00FF00FF; enable MIZAR_GPIO_GP0_INTR1_INTR_EN1=0xFFFFFFFF; for i=0..31 write per-pin 0x00200000 then verify pad bit at 0xA0243ffc is 1; write 0x00000000 then verify pad bit is 0; clear IRQ if defined; any Default_IRQHandler entry increments error.",
        "Impacted Registers": ["MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_GPIO_8"],
        "Validation / Acceptance Criteria": "For each i: pad bit at 0xA0243ffc reflects written DOUT high/low; no unexpected interrupts.",
        "Code Generation (Required / Not)": "",
        "Hidden_Test_Case_Name": "test_gpio_op_mode_all_pad_en",
        "Hidden_Test_Description": "Output mode across IO groups with per-pin DOUT toggling; verification via 0xA0243ffc; Default_IRQHandler increments test_err.",
        "Hidden_Remarks": "Uses gp0_flag_dout_one/zero logic in helper; GIC_ClearIRQ after checks.",
        "Hidden_Test_Steps_Procedure": "Source: program.c. See loop and helper function.",
        "Hidden_Impacted_Registers": ["MIZAR_GPIO_GPIO_IO_CTRL_GROUP1", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP2", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP3", "MIZAR_GPIO_GPIO_IO_CTRL_GROUP4", "MIZAR_GPIO_GP0_GPIO_8"],
        "Hidden_Validation_Acceptance_Criteria": "Pad bit reflects DOUT high/low; no unexpected interrupts."    }
]

META_COLUMNS = [
    "Hidden_Test_Case_Name",
    "Hidden_Test_Description",
    "Hidden_Remarks",
    "Hidden_Test_Steps_Procedure",
    "Hidden_Impacted_Registers",
    "Hidden_Validation_Acceptance_Criteria",
]

MAIN_COLUMNS = [
    "Index",
    "SS / Module",
    "Feature",
    "Test Case Name",
    "Test Description",
    "Speed",
    "Mode",
    "Memory Start Offset",
    "Memory End Offset",
    "Remarks",
    "Test Steps / Procedure",
    "Impacted Registers",
    "Validation / Acceptance Criteria",
    "Code Generation (Required / Not)",
]

NUMBER_WRAP_COLS = [
    "Test Steps / Procedure",
    "Validation / Acceptance Criteria",
]

ALLOWED_DV = ["Required", "Blank", "Not Required"]


def read_existing(path: str):
    rows = []
    meta_rows = []
    if not os.path.exists(path):
        return rows, meta_rows
    try:
        wb = load_workbook(path, data_only=True)
        if 'TestPlan' in wb.sheetnames:
            ws = wb['TestPlan']
            vals = list(ws.iter_rows(values_only=True))
            if vals:
                headers = [str(h) if h is not None else '' for h in vals[0]]
                for r in vals[1:]:
                    if all(v is None for v in r):
                        continue
                    rec = {}
                    for k, v in zip(headers, r):
                        if k == '':
                            continue
                        rec[k] = v if v is not None else ''
                    rows.append(rec)
        if 'Meta_data_sheet' in wb.sheetnames:
            ms = wb['Meta_data_sheet']
            mvals = list(ms.iter_rows(values_only=True))
            if mvals:
                mheaders = [str(h) if h is not None else '' for h in mvals[0]]
                for r in mvals[1:]:
                    meta = {}
                    for k, v in zip(mheaders, r):
                        if k == '':
                            continue
                        meta[k] = v if v is not None else ''
                    meta_rows.append(meta)
    except Exception:
        pass
    return rows, meta_rows


def first_seen_union_keys(records: List[Dict]) -> List[str]:
    keys: List[str] = []
    seen = set()
    for rec in records:
        for k in rec.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    return keys


def enforce_numbering(text: str) -> str:
    if text is None:
        return ''
    lines = [ln.strip() for ln in str(text).splitlines()]
    lines = [ln for ln in lines if ln]
    out = []
    for i, ln in enumerate(lines, start=1):
        ln = re.sub(r"^\s*(?:[0-9]+[\.)-]|[-*•])\s*", "", ln)
        out.append(f"{i}. {ln}")
    return "\n".join(out)


def build_meta_sheet(wb, records: List[Dict]):
    ws_meta = wb.create_sheet(title='Meta_data_sheet')
    for ci, k in enumerate(META_COLUMNS, start=1):
        ws_meta.cell(row=1, column=ci, value=k)
    for ri, rec in enumerate(records, start=2):
        for ci, k in enumerate(META_COLUMNS, start=1):
            ws_meta.cell(row=ri, column=ci, value=rec.get(k, ""))
    for cell in ws_meta[1]:
        cell.font = Font(bold=True)
    ws_meta.sheet_state = 'veryHidden'


def write_workbook(records: List[Dict]):
    headers = first_seen_union_keys(records)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
    for ri, rec in enumerate(records, start=2):
        for ci, h in enumerate(headers, start=1):
            v = rec.get(h, "")
            if isinstance(v, (list, dict)):
                ws.cell(row=ri, column=ci, value=json.dumps(v, ensure_ascii=False))
            else:
                ws.cell(row=ri, column=ci, value=v)

    # Base formatting
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = 'A2'

    # Build META
    build_meta_sheet(wb, records)

    # Rename Data -> TestPlan and reorder/hide meta
    ws.title = 'TestPlan'
    current_headers = [ws.cell(row=1, column=i+1).value for i in range(ws.max_column)]
    remaining = [h for h in current_headers if h not in set(MAIN_COLUMNS) and h not in set(META_COLUMNS)]
    final_order = [h for h in MAIN_COLUMNS if h in current_headers] + remaining

    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[0]
    idx_map = {h: i for i, h in enumerate(header_row)}
    table = []
    for r in rows[1:]:
        table.append({h: (r[idx_map[h]] if h in idx_map else '') for h in current_headers})

    ws.delete_rows(1, ws.max_row)
    for ci, h in enumerate(final_order, start=1):
        ws.cell(row=1, column=ci, value=h)
    for ri, rec in enumerate(table, start=2):
        for ci, h in enumerate(final_order, start=1):
            ws.cell(row=ri, column=ci, value=rec.get(h, ""))

    # Enforce numbering
    header_to_col = {ws.cell(row=1, column=i+1).value: i+1 for i in range(ws.max_column)}
    for col_name in NUMBER_WRAP_COLS:
        if col_name in header_to_col:
            cidx = header_to_col[col_name]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=cidx, value=enforce_numbering(ws.cell(row=r, column=cidx).value))

    # Styles and borders
    blue = PatternFill(fill_type='solid', fgColor='1F4E78')
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_top = Alignment(horizontal='left', vertical='top', wrap_text=False)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in ws[1]:
        c.fill = blue
        c.alignment = align_center

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if ws.cell(row=1, column=c).value == 'Index':
                cell.alignment = Alignment(horizontal='center', vertical='top')
            else:
                cell.alignment = align_left_top
            cell.border = border

    # Wrap for specific columns
    for name in NUMBER_WRAP_COLS + ['Test Description', 'Remarks']:
        if name in header_to_col:
            col = header_to_col[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Data validation for Code Generation column
    headers_now = [ws.cell(row=1, column=i+1).value for i in range(ws.max_column)]
    if 'Code Generation (Required / Not)' in headers_now and ws.max_row >= 2:
        col_idx = headers_now.index('Code Generation (Required / Not)') + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        dv = DataValidation(type='list', formula1='"Required,Blank,Not Required"', allow_blank=True, showDropDown=True)
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{ws.max_row}')

    # Remove any stray sheets except TestPlan and Meta_data_sheet
    for name in list(wb.sheetnames):
        if name not in {'TestPlan', 'Meta_data_sheet'}:
            wb.remove(wb[name])

    wb.save(OUTPUT_FILE_PATH)

    with zipfile.ZipFile(OUTPUT_FILE_PATH, 'r') as z:
        names = set(z.namelist())
        required = {'[Content_Types].xml', 'xl/workbook.xml', 'xl/worksheets/sheet1.xml'}
        missing = required - names
        if missing:
            raise RuntimeError(f'Missing OOXML parts: {missing}')


def main():
    if not os.path.exists(OUTPUT_FILE_PATH):
        print(json.dumps({"Status":"FAILURE","Error":"Batch 1 WORKING.xlsx not found; aborting Batch 2 append"}))
        sys.exit(1)

    existing_rows, existing_meta = read_existing(OUTPUT_FILE_PATH)

    # Merge existing TestPlan rows with their meta fields (by index order)
    merged_existing: List[Dict] = []
    for idx, rec in enumerate(existing_rows):
        rec2 = dict(rec)
        if idx < len(existing_meta):
            for k in META_COLUMNS:
                if k in existing_meta[idx]:
                    rec2[k] = existing_meta[idx][k]
        merged_existing.append(rec2)

    merged = []
    merged.extend(merged_existing)
    merged.extend(deepcopy(JSON_DATA))

    try:
        write_workbook(merged)
        cols = len(first_seen_union_keys(merged))
        print(json.dumps({
            "Status": "SUCCESS",
            "Rows": len(merged),
            "Columns": cols,
            "Final Excel file path": OUTPUT_FILE_PATH
        }))
    except Exception as e:
        print(json.dumps({"Status":"FAILURE","Error":str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()
